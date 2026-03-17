from __future__ import annotations

import base64
import io
import json
import math
import mimetypes
import hashlib
import os
import re
import secrets
import time
import urllib.error
import urllib.parse
import urllib.request
import xml.etree.ElementTree as ET
from pathlib import Path
from datetime import date, datetime, timedelta, timezone
from typing import Any
from zoneinfo import ZoneInfo

from fastapi import APIRouter, Depends, File, Form, HTTPException, UploadFile, Request, Response
from fastapi.responses import HTMLResponse, RedirectResponse, StreamingResponse
from sqlalchemy import String, cast, delete, func, or_, select, text
from sqlalchemy.orm import Session

try:
    from openpyxl import Workbook, load_workbook
except Exception:  # pragma: no cover
    Workbook = None
    load_workbook = None

from app.config import settings
from app.auth import create_access_token, decode_access_token, get_password_hash, verify_password
from app.db import get_db
from app.deps import get_admin_user, get_current_user, oauth2_scheme
from app.models import (
    AiServiceAccount,
    ApiCredential,
    AuditLog,
    BillingAccount,
    BillingEvent,
    AccountingExpense,
    AccountingSettings,
    ModuleAccess,
    MarketplaceApiCache,
    PositionSnapshot,
    Product,
    SeoJob,
    User,
    UserAiSettings,
    UserAiPreference,
    UserKnowledgeDoc,
    UserProfile,
    TeamMember,
    SocialCalendarEvent,
    SocialChatMessage,
    SocialChatThread,
    SocialChatThreadMember,
    SocialGameScore,
    SocialCheckersProfile,
    SocialCheckersRoom,
    SocialAnnouncement,
    SocialAnnouncementAck,
    SocialNote,
    SocialNoteFile,
    SocialNotification,
    SocialTask,
    SocialTaskComment,
    SocialTaskProject,
    SocialTaskProjectMember,
    UserQuestionAiSettings,
    UserKeyword,
    SystemSetting,
    WbAdsCampaignSnapshot,
    WbAdsBidderRule,
    WbAdsBidderRun,
    WorkItemClaim,
)
from app.schemas import (
    ActivityTrackIn,
    AiAssistantIn,
    AiAssistantOut,
    AiEffectiveOut,
    AiProfileOut,
    AiSelectionIn,
    AiSelectionOut,
    AiServiceIn,
    AiServiceReorderIn,
    AiServiceOut,
    AdminUserProfileOut,
    AdminCredentialRowOut,
    AdminCredentialIn,
    BillingOut,
    BillingPlanChangeIn,
    AdminPasswordResetIn,
    AdminRoleUpdateIn,
    AdminStatsOut,
    AvatarUploadOut,
    AuditLogOut,
    AuditLogPageOut,
    ApiCredentialIn,
    ApiCredentialOut,
    AccountingDataOut,
    AccountingMonthlyKpiOut,
    AccountingMonthlyMetaOut,
    AccountingMonthlyRowOut,
    AccountingMonthlySummaryOut,
    AccountingExpenseIn,
    AccountingExpenseListOut,
    AccountingExpenseOut,
    AccountingPurchasePriceImportOut,
    AccountingSettingsIn,
    AccountingSettingsOut,
    CurrentModuleOut,
    CredentialTestOut,
    DashboardOut,
    GenerateReviewReplyIn,
    GenerateReviewReplyOut,
    ImportProductsRequest,
    KeywordIn,
    KeywordOut,
    LoginRequest,
    MessageOut,
    ModuleAccessIn,
    ModuleAccessOut,
    PositionCheckOut,
    PositionCheckRequest,
    ProductDetailOut,
    ProductBulkDeleteIn,
    ProductPageOut,
    ProductOut,
    ProductReloadRequest,
    ProductUpdateIn,
    SalesStatsOut,
    HelpDocOut,
    HelpReleaseOut,
    MobileApkLatestOut,
    KnowledgeDocOut,
    ReviewAiSettingsIn,
    ReviewAiSettingsOut,
    CampaignIdsIn,
    RegisterRequest,
    SeoApplyRequest,
    SeoDeleteRequest,
    SeoGenerateRequest,
    SeoJobOut,
    SeoRecheckRequest,
    TokenResponse,
    TrendOut,
    TrendPointOut,
    TeamMemberIn,
    TeamMemberOut,
    UserProfileOut,
    UserProfilePasswordIn,
    UserProfileUpdateIn,
    UserOut,
    SocialCalendarEventIn,
    SocialCalendarEventOut,
    SocialCalendarGoogleSyncIn,
    SocialCalendarGoogleSyncOut,
    SocialCalendarReminderSettingsIn,
    SocialCalendarReminderSettingsOut,
    SocialChatDirectStartIn,
    SocialChatGroupIn,
    SocialChatGroupUpdateIn,
    SocialChatCompanyUpdateIn,
    SocialChatMessageIn,
    SocialChatMessageOut,
    SocialChatReactionIn,
    SocialChatThreadAvatarIn,
    SocialChatThreadOut,
    SocialCurrencyRatesOut,
    SocialGameScoreIn,
    SocialGameScoreOut,
    SocialLeaderboardOut,
    SocialNoteIn,
    SocialNoteFileOut,
    SocialNoteOut,
    SocialNotificationOut,
    SocialAnnouncementIn,
    SocialAnnouncementOut,
    SocialAnnouncementPublicOut,
    SocialTaskCommentIn,
    SocialTaskIn,
    SocialTaskOut,
    SocialTaskProjectIn,
    SocialTaskProjectOut,
    SocialTaskProjectMemberOut,
    SocialTaskProjectMembersUpdateIn,
    SocialTaskUpdateIn,
    NotificationSoundSettingsIn,
    NotificationSoundSettingsOut,
    WbCampaignRatesIn,
    WbCampaignDetailOut,
    WbAdsActionIn,
    WbAdsActionOut,
    WbAdsAnalyticsOut,
    WbAdsRecommendationsOut,
    WbAdsBalanceOut,
    WbBidderRuleIn,
    WbBidderRuleOut,
    WbBidderRuleUpdateIn,
    WbBidderRulesOut,
    WbBidderRunIn,
    WbBidderRunOut,
    WbBidderRunRowOut,
    WbBidderRunsOut,
    WbCampaignEnrichOut,
    WbCampaignRatesOut,
    WbCampaignsOut,
    WbReviewReplyIn,
    WbReviewReplyOut,
    WbReviewsOut,
    ReturnActionIn,
    ReturnActionOut,
    ReturnsOut,
    UiSettingsIn,
    UiSettingsOut,
)
from app.services.accounting import (
    build_accounting_payload,
    build_accounting_monthly_summary,
)
from app.services.sales import build_sales_report
from app.services.market_cache import (
    build_market_cache_key,
    get_market_cache_stats,
    get_or_refresh_market_cache,
)
from app.services.task_queue import enqueue_task, queue_available, queue_depth, queue_enabled
from app.telemetry import collect_perf_metrics
from app.services.ads_cache import (
    get_wb_snapshot_rows,
    is_wb_snapshot_stale,
    sync_wb_campaign_snapshots,
)
from app.services.social_checkers import (
    CHECKERS_GAME_CODE,
    apply_checkers_elo,
    apply_checkers_move,
    build_checkers_bot_identity,
    create_checkers_room_code,
    create_checkers_state,
    get_checkers_difficulties,
    get_checkers_legal_moves,
    load_checkers_state,
    pick_checkers_bot_move,
)
from app.services.social_chess import (
    CHESS_GAME_CODE,
    apply_chess_move,
    build_chess_bot_identity,
    create_chess_state,
    get_chess_difficulties,
    get_chess_legal_moves,
    load_chess_state,
    pick_chess_bot_move,
)
from app.services.social_battleship import (
    BATTLESHIP_GAME_CODE,
    apply_battleship_shot,
    assign_battleship_side,
    build_battleship_bot_identity,
    create_battleship_state,
    get_battleship_available_shots,
    get_battleship_difficulties,
    load_battleship_state,
    mask_enemy_board,
    pick_battleship_bot_move,
)
from app.services.marketplace import (
    enrich_ozon_category_names,
    fetch_marketplace_product_details,
    fetch_products_from_marketplace,
    find_competitors,
    resolve_wb_external_id,
    test_marketplace_credentials,
    update_product_description,
    update_product_photos_order,
)
from app.services.modules import DEFAULT_MODULES
from app.services.seo import (
    build_seo_description,
    discover_keywords,
    evaluate_position,
    evaluate_positions_for_keywords,
    schedule_next_check,
)
from app.services.wb_modules import (
    generate_help_assistant_reply,
    probe_ozon_feedback_access,
    probe_wb_feedback_access,
    fetch_ozon_reviews,
    fetch_ozon_questions,
    fetch_wb_ads_balance,
    fetch_wb_campaign_summaries,
    fetch_wb_campaign_details,
    fetch_wb_campaign_rates,
    fetch_wb_campaign_stats_bulk,
    fetch_wb_campaigns,
    fetch_wb_questions,
    fetch_wb_questions_fast,
    fetch_wb_reviews,
    fetch_wb_reviews_fast,
    generate_review_reply,
    post_ozon_question_reply,
    post_ozon_review_reply,
    post_wb_question_reply,
    post_wb_review_reply,
    update_wb_campaign_state,
    fetch_wb_returns,
    fetch_wb_return_details,
    action_wb_return,
    fetch_ozon_ads_campaigns,
    fetch_ozon_ads_analytics,
    fetch_ozon_returns,
    fetch_ozon_return_details,
)
from app.services.wb_bidder import (
    apply_rule_payload,
    list_bidder_rules,
    list_bidder_runs,
    normalize_rule_payload,
    run_bidder_rules,
    serialize_bidder_rule,
    serialize_bidder_run,
)

router = APIRouter(prefix="/api")
DISABLED_BY_DEFAULT_MODULES = {"billing", "wb_reviews_ai", "wb_questions_ai", "wb_ads", "wb_ads_analytics", "wb_ads_recommendations", "help_center", "social_hub"}
AVAILABLE_THEMES = ("classic", "dark", "light", "newyear", "summer", "autumn", "winter", "spring", "japan", "greenland", "moon")
DEFAULT_UI_SETTINGS = {
    "theme_choice_enabled": True,
    "force_theme": False,
    "default_theme": "classic",
    "allowed_themes": list(AVAILABLE_THEMES),
}
MOBILE_MODULE_OVERRIDES_SETTING_KEY = "mobile_module_overrides"
AUDIT_STORAGE_MAX_BYTES = 3 * 1024 * 1024 * 1024
AUDIT_STORAGE_TARGET_BYTES = int(AUDIT_STORAGE_MAX_BYTES * 0.9)
AUDIT_PRUNE_BATCH = 5000
_AUDIT_PRUNE_LAST_CHECK_AT: float = 0.0
PRODUCT_PAGE_SIZE_OPTIONS = (30, 50, 100, 200, 500, 1000)
_SERVER_CPU_SNAPSHOT: tuple[int, int] | None = None
_SERVER_NET_SNAPSHOT: tuple[int, int] | None = None
_SERVER_NET_TS: float = 0.0
_SOCIAL_REMINDER_THROTTLE_SEC = 60
_SOCIAL_REMINDER_LAST_RUN: dict[str, float] = {}
_MARKET_CACHE_TTL_SEC = {
    "dashboard": 60,
    "sales_stats": 180,
    "accounting": 240,
    "accounting_monthly": 300,
    "wb_reviews_ai": 120,
    "wb_questions_ai": 120,
    "returns": 180,
    "wb_ads": 180,
    "wb_ads_analytics": 180,
    "wb_ads_recommendations": 180,
}
SOCIAL_GOOGLE_OAUTH_TOKENS_KEY = "social_google_oauth_tokens"
SOCIAL_CALENDAR_SYNC_STATUS_KEY = "social_calendar_sync_status"
SOCIAL_CALENDAR_REMINDER_SETTINGS_PREFIX = "social_calendar_reminders_u_"
SOCIAL_CALENDAR_REMINDER_DEFAULT_OFFSETS_MIN = [-10080, -4320, -1440, -180, -60, 0]
SOCIAL_CALENDAR_REMINDER_ALLOWED_OFFSETS_MIN = [-10080, -4320, -1440, -720, -360, -180, -120, -60, -30, -15, -5, 0]
SOCIAL_GOOGLE_OAUTH_SCOPE = "https://www.googleapis.com/auth/calendar.readonly"
SOCIAL_CALENDAR_DEFAULT_TZ = "Europe/Moscow"


def _secret_revision(*values: str) -> str:
    raw = "|".join(str(v or "").strip() for v in values)
    return hashlib.sha1(raw.encode("utf-8")).hexdigest()[:12]


def _market_cache_ttl(module_code: str, *, fast_mode: bool = False) -> int:
    base = int(_MARKET_CACHE_TTL_SEC.get(str(module_code or "").strip(), 120))
    if fast_mode:
        return max(45, int(base * 0.7))
    return base


def _market_cache_latest_payload(
    db: Session,
    *,
    user_id: int,
    module_code: str,
    marketplace: str,
    max_age_sec: int = 24 * 60 * 60,
    exclude_cache_keys: set[str] | None = None,
    scan_limit: int = 80,
) -> tuple[dict[str, Any] | None, dict[str, Any]]:
    now = datetime.utcnow()
    excluded = {str(x or "").strip() for x in (exclude_cache_keys or set()) if str(x or "").strip()}
    rows = db.scalars(
        select(MarketplaceApiCache)
        .where(
            MarketplaceApiCache.user_id == int(user_id),
            MarketplaceApiCache.module_code == str(module_code or "").strip()[:80],
            MarketplaceApiCache.marketplace == str(marketplace or "").strip()[:30],
        )
        .order_by(MarketplaceApiCache.fetched_at.desc(), MarketplaceApiCache.id.desc())
        .limit(max(10, min(int(scan_limit or 80), 500)))
    ).all()
    for row in rows:
        if excluded and str(row.cache_key or "").strip() in excluded:
            continue
        fetched_at = row.fetched_at or row.last_hit_at
        if not fetched_at:
            continue
        age_sec = max(0, int((now - fetched_at).total_seconds()))
        if age_sec > max(60, int(max_age_sec or 0)):
            continue
        try:
            payload = json.loads(str(row.payload_json or ""))
        except Exception:
            continue
        if isinstance(payload, dict) and payload:
            return payload, {
                "source": "db-latest-module-fallback",
                "age_sec": age_sec,
                "cache_key": str(row.cache_key or ""),
            }
    return None, {}


def _warnings_indicate_upstream_failure(warnings: list[Any]) -> bool:
    raw = [str(x or "").strip().lower() for x in (warnings or []) if str(x or "").strip()]
    if not raw:
        return False
    markers = (
        "429",
        "rate_limited",
        "timed out",
        "timeout",
        "РЅРµРґРѕСЃС‚СѓРї",
        "РѕС€РёР±РєР°",
        "error",
        "api РІРµСЂРЅСѓР»",
        "api РЅРµРґРѕСЃС‚СѓРїРµРЅ",
    )
    return any(any(mark in item for mark in markers) for item in raw)


def _sales_payload_has_data(payload: dict[str, Any] | None) -> bool:
    if not isinstance(payload, dict):
        return False
    rows = payload.get("rows")
    if isinstance(rows, list) and rows:
        return True
    totals = payload.get("totals") if isinstance(payload.get("totals"), dict) else {}
    numeric_keys = (
        "orders",
        "units",
        "buyouts",
        "revenue",
        "wb_orders",
        "ozon_orders",
        "wb_revenue",
        "ozon_revenue",
    )
    for key in numeric_keys:
        try:
            if abs(float(totals.get(key) or 0.0)) > 1e-9:
                return True
        except Exception:
            continue
    return False


def _accounting_payload_has_data(payload: dict[str, Any] | None) -> bool:
    if not isinstance(payload, dict):
        return False
    overview = payload.get("overview") if isinstance(payload.get("overview"), dict) else {}
    numeric_keys = ("orders", "units", "buyouts", "revenue", "gross_profit", "net_profit")
    for key in numeric_keys:
        try:
            if abs(float(overview.get(key) or 0.0)) > 1e-9:
                return True
        except Exception:
            continue
    chart = payload.get("chart")
    if isinstance(chart, list):
        for row in chart:
            if not isinstance(row, dict):
                continue
            for key in ("orders", "units", "revenue", "gross_profit", "net_profit"):
                try:
                    if abs(float(row.get(key) or 0.0)) > 1e-9:
                        return True
                except Exception:
                    continue
    analysis_rows = payload.get("analysis_rows")
    if not isinstance(analysis_rows, list) or not analysis_rows:
        return False
    for row in analysis_rows:
        if not isinstance(row, dict):
            continue
        for key in (
            "orders",
            "units",
            "buyouts",
            "sold_units",
            "returns",
            "revenue",
            "gross_profit",
            "operating_profit",
            "net_profit",
            "marketplace_expenses",
            "cogs",
        ):
            try:
                if abs(float(row.get(key) or 0.0)) > 1e-9:
                    return True
            except Exception:
                continue
    return False


def _accounting_monthly_payload_has_data(payload: dict[str, Any] | None) -> bool:
    if not isinstance(payload, dict):
        return False
    months = payload.get("months")
    if not isinstance(months, list) or not months:
        return False
    for row in months:
        if not isinstance(row, dict):
            continue
        for section in ("wb", "ozon", "total"):
            part = row.get(section)
            if not isinstance(part, dict):
                continue
            for key in ("turnover", "orders", "units", "buyouts", "cogs", "net_profit"):
                try:
                    if abs(float(part.get(key) or 0.0)) > 1e-9:
                        return True
                except Exception:
                    continue
    return False

def _accounting_payload_needs_finance_fallback(payload: dict[str, Any] | None, warnings: list[str] | None = None) -> bool:
    if not isinstance(payload, dict):
        return False
    warning_text = " | ".join(str(x or "").strip().lower() for x in (warnings or []))
    if not warning_text:
        return False
    markers = (
        "finance api РЅРµРґРѕСЃС‚СѓРїРµРЅ",
        "finance api unavailable",
        "СЂРµРєР»Р°РјРЅС‹Рµ СЂР°СЃС…РѕРґС‹ РІСЂРµРјРµРЅРЅРѕ РЅРµРґРѕСЃС‚СѓРїРЅС‹",
        "rate_limited",
        "РѕРіСЂР°РЅРёС‡РёР» Р·Р°РїСЂРѕСЃС‹",
    )
    if not any(marker in warning_text for marker in markers):
        return False

    overview = payload.get("overview") if isinstance(payload.get("overview"), dict) else {}
    revenue = abs(float(_to_float_safe(overview.get("revenue"), 0.0)))
    if revenue <= 1e-9:
        return False

    for key in (
        "cogs",
        "marketplace_expense",
        "marketplace_expenses",
        "gross_profit",
        "operating_profit",
        "net_profit",
        "ads",
        "ad_spent",
        "commissions",
        "logistics",
    ):
        if abs(float(_to_float_safe(overview.get(key), 0.0))) > 1e-9:
            return False

    analysis_rows = payload.get("analysis_rows")
    if isinstance(analysis_rows, list):
        for row in analysis_rows:
            if not isinstance(row, dict):
                continue
            for key in ("cogs", "marketplace_expense", "marketplace_expenses", "gross_profit", "operating_profit", "net_profit"):
                if abs(float(_to_float_safe(row.get(key), 0.0))) > 1e-9:
                    return False

    return True


def _extract_returns_payload_rows(payload: Any) -> list[dict[str, Any]]:
    def _path_value(root: Any, path: str) -> Any:
        current = root
        for part in str(path or "").split("."):
            key = str(part or "").strip()
            if not key or not isinstance(current, dict):
                return None
            current = current.get(key)
        return current

    def _append_rows(out: list[dict[str, Any]], node: Any) -> None:
        if isinstance(node, list):
            for item in node:
                if isinstance(item, dict):
                    out.append(item)

    def _looks_like_return_row(node: Any) -> bool:
        if not isinstance(node, dict):
            return False
        keys = {str(key or "").strip().lower() for key in node.keys()}
        if {"claim", "return", "item", "product"} & keys:
            return True
        id_keys = {
            "id",
            "return_id",
            "returnid",
            "claimid",
            "claim_id",
            "posting_number",
            "srid",
            "order_id",
        }
        info_keys = {
            "status",
            "state",
            "created_at",
            "createdat",
            "updated_at",
            "updatedat",
            "name",
            "product",
            "reason",
            "comment",
            "description",
        }
        return bool(keys & id_keys) and bool(keys & info_keys)

    rows: list[dict[str, Any]] = []
    if isinstance(payload, list):
        _append_rows(rows, payload)
        return rows
    if not isinstance(payload, dict):
        return rows

    for path in (
        "rows",
        "claims",
        "returns",
        "items",
        "data.rows",
        "data.claims",
        "data.returns",
        "data.items",
        "result.rows",
        "result.claims",
        "result.returns",
        "result.items",
        "payload.rows",
        "payload.claims",
        "payload.returns",
        "payload.items",
        "response.rows",
        "response.claims",
        "response.returns",
        "response.items",
        "data.result.rows",
        "data.result.claims",
        "data.result.returns",
        "result.data.rows",
        "result.data.claims",
        "result.data.returns",
    ):
        _append_rows(rows, _path_value(payload, path))

    if rows:
        return rows

    for node in (payload.get("data"), payload.get("result"), payload.get("payload"), payload.get("response")):
        if isinstance(node, list):
            _append_rows(rows, node)
            if rows:
                return rows

    if _looks_like_return_row(payload):
        return [payload]
    return rows


def _normalize_returns_rows(rows: list[dict[str, Any]], marketplace: str) -> list[dict[str, Any]]:
    safe_market = str(marketplace or "").strip().lower()
    out: list[dict[str, Any]] = []
    seen: set[str] = set()

    def _path_value(root: Any, path: str) -> Any:
        current = root
        for part in str(path or "").split("."):
            key = str(part or "").strip()
            if not key or not isinstance(current, dict):
                return None
            current = current.get(key)
        return current

    def _clean_text(value: Any) -> str:
        if isinstance(value, (dict, list, tuple, set)):
            return ""
        text = " ".join(str(value or "").split()).strip()
        low = text.lower()
        if not text or low in {"-", "вЂ”", "null", "none", "undefined", "n/a"}:
            return ""
        return text

    def _pick_text(node: dict[str, Any], *paths: str) -> str:
        for path in paths:
            text = _clean_text(_path_value(node, path))
            if text:
                return text
        return ""

    def _number_from(value: Any) -> float | None:
        if isinstance(value, bool):
            return None
        if isinstance(value, (int, float)):
            num = float(value)
            return num if math.isfinite(num) else None
        if isinstance(value, str):
            text = str(value or "").strip()
            if not text:
                return None
            normalized = text.replace("\xa0", " ").replace(" ", "")
            normalized = normalized.replace("?", "").replace("СЂСѓР±.", "").replace("СЂСѓР±", "")
            normalized = normalized.replace(",", ".")
            match = re.search(r"-?\d+(?:\.\d+)?", normalized)
            if not match:
                return None
            try:
                num = float(match.group(0))
                return num if math.isfinite(num) else None
            except Exception:
                return None
        if isinstance(value, dict):
            for key in ("value", "amount", "price", "sum", "total", "qty", "count", "quantity"):
                if key in value:
                    parsed = _number_from(value.get(key))
                    if parsed is not None:
                        return parsed
            for nested in value.values():
                parsed = _number_from(nested)
                if parsed is not None:
                    return parsed
            return None
        if isinstance(value, (list, tuple)):
            for item in value:
                parsed = _number_from(item)
                if parsed is not None:
                    return parsed
            return None
        return None

    def _format_number(value: float | None) -> str:
        if value is None or not math.isfinite(value):
            return ""
        rounded = round(float(value), 3)
        if abs(rounded - round(rounded)) < 0.000001:
            return str(int(round(rounded)))
        text = f"{rounded:.3f}".rstrip("0").rstrip(".")
        return text

    def _pick_number_text(node: dict[str, Any], *paths: str) -> str:
        for path in paths:
            number = _number_from(_path_value(node, path))
            text = _format_number(number)
            if text:
                return text
        return ""

    def _clean_photo(*values: Any) -> list[str]:
        photos: list[str] = []
        seen_photo: set[str] = set()

        def _walk(node: Any, depth: int = 0) -> None:
            if node is None or depth > 4:
                return
            if isinstance(node, str):
                url = str(node or "").strip()
                if not url or url in seen_photo:
                    return
                if not (url.startswith("http://") or url.startswith("https://")):
                    return
                seen_photo.add(url)
                photos.append(url)
                return
            if isinstance(node, list):
                for item in node:
                    _walk(item, depth + 1)
                return
            if isinstance(node, dict):
                for key in ("url", "photo", "image", "src", "link", "href", "path", "big", "x1", "x2", "orig", "tm"):
                    _walk(node.get(key), depth + 1)
                for key, value in node.items():
                    if re.search(r"(photo|image|picture|attachment|file|evidence)", str(key or ""), flags=re.IGNORECASE):
                        _walk(value, depth + 1)
                return

        for value in values:
            _walk(value, 0)
        return photos

    def _join_lines(*values: Any) -> str:
        lines: list[str] = []
        seen_lines: set[str] = set()
        for value in values:
            text = _clean_text(value)
            if not text:
                continue
            low = text.lower()
            if low in seen_lines:
                continue
            seen_lines.add(low)
            lines.append(text)
        return "\n".join(lines)

    for row in rows:
        if not isinstance(row, dict):
            continue
        raw_node = row.get("raw") if isinstance(row.get("raw"), dict) else {}
        merged = dict(raw_node)
        for key, value in row.items():
            if key not in merged or merged.get(key) in (None, "", [], {}):
                merged[key] = value

        rid = _pick_text(
            merged,
            "id",
            "return_id",
            "returnId",
            "claimId",
            "claim_id",
            "claim.id",
            "claim.claimId",
            "return.id",
            "return.return_id",
            "posting_number",
            "posting.number",
            "srid",
            "order_id",
            "returnNumber",
        )
        status = _pick_text(
            merged,
            "status",
            "state",
            "claimStatus",
            "claim.status",
            "claim.state",
            "return.status",
            "return.state",
            "status.name",
            "state.name",
            "claim.status.name",
            "return.status.name",
        )
        status_code = _pick_text(
            merged,
            "status_code",
            "statusCode",
            "status.id",
            "status.code",
            "state.id",
            "state.code",
            "claim.status.id",
            "claim.status.code",
            "return.status.id",
            "return.status.code",
        )
        status_note = _pick_text(
            merged,
            "status_ex",
            "statusEx",
            "status_detail",
            "statusDetail",
            "status.note",
            "status.description",
            "decision_comment",
            "rejectReason",
            "reject_reason",
            "claim.status.note",
            "claim.status.description",
            "return.status.note",
            "return.status.description",
        )
        created_at = _pick_text(
            merged,
            "created_at",
            "createdAt",
            "date",
            "createdDate",
            "claim.createdAt",
            "return.createdAt",
        )
        updated_at = _pick_text(
            merged,
            "updated_at",
            "updatedAt",
            "status_updated_at",
            "statusUpdatedAt",
            "claim.updatedAt",
            "return.updatedAt",
        )
        vendor_code = _pick_text(
            merged,
            "vendor_code",
            "vendorCode",
            "supplierVendorCode",
            "origin_id_info.vendor_code",
            "item.vendorCode",
            "claim.item.vendorCode",
            "return.item.vendorCode",
        )
        nm_id = _pick_text(
            merged,
            "nm_id",
            "nmId",
            "imt_id",
            "imtId",
            "item.nm_id",
            "claim.item.nm_id",
            "return.item.nm_id",
        )
        article = _pick_text(
            merged,
            "article",
            "offer_id",
            "offerId",
            "sku",
            "vendorCode",
            "supplierVendorCode",
            "item.article",
            "item.offer_id",
            "claim.item.article",
            "return.item.article",
            "product.article",
            "product.offer_id",
        )
        product = _pick_text(
            merged,
            "product",
            "product_name",
            "productName",
            "name",
            "subjectName",
            "imtName",
            "item.name",
            "item.title",
            "claim.item.name",
            "return.item.name",
            "product.name",
        )
        quantity = _pick_number_text(
            merged,
            "quantity",
            "count",
            "itemsCount",
            "qty",
            "item.quantity",
            "claim.item.quantity",
            "return.item.quantity",
        )
        amount = _pick_number_text(
            merged,
            "amount",
            "sum",
            "total",
            "refundAmount",
            "returnAmount",
            "price",
            "claim.amount",
            "claim.price",
            "return.amount",
            "return.price",
            "item.price",
        )
        customer_comment = _pick_text(
            merged,
            "user_comment",
            "customer_comment",
            "comment",
            "claim.comment",
            "return.comment",
        )
        seller_comment = _pick_text(
            merged,
            "seller_comment",
            "wb_comment",
            "decision_comment",
            "rejectReason",
            "reject_reason",
            "claim.decision_comment",
            "return.decision_comment",
        )
        base_reason = _pick_text(
            merged,
            "description",
            "reason",
            "text",
            "comment",
            "rejectReason",
            "reject_reason",
            "claim.reason",
            "claim.description",
            "return.reason",
            "return.description",
        )
        reason = _join_lines(base_reason, status_note, customer_comment, seller_comment)
        photos = _clean_photo(
            merged.get("photos"),
            merged.get("images"),
            merged.get("pictures"),
            merged.get("attachments"),
            merged.get("files"),
            merged.get("evidences"),
            merged.get("claim"),
            merged.get("return"),
            merged.get("item"),
            merged.get("product"),
        )
        if not rid:
            fingerprint = "|".join(
                [
                    created_at,
                    updated_at,
                    vendor_code,
                    nm_id,
                    article,
                    product,
                    status,
                    reason,
                ]
            ).strip("|")
            if fingerprint:
                rid = hashlib.sha1(fingerprint.encode("utf-8", errors="ignore")).hexdigest()[:16]
        if not any([rid, status, status_code, created_at, updated_at, article, product, reason, photos, amount, quantity]):
            continue
        if not rid or rid in seen:
            continue
        seen.add(rid)
        out.append(
            {
                "id": rid,
                "status": status or status_code,
                "status_code": status_code or status,
                "status_note": status_note,
                "created_at": created_at,
                "updated_at": updated_at,
                "date": created_at,
                "article": article,
                "vendor_code": vendor_code,
                "nm_id": nm_id,
                "product": product,
                "quantity": quantity,
                "amount": amount,
                "customer_comment": customer_comment,
                "seller_comment": seller_comment,
                "reason": reason,
                "description": reason,
                "photos": photos,
                "marketplace": safe_market or str(merged.get("marketplace") or "").strip().lower(),
                "raw": raw_node if raw_node else merged,
            }
        )
    return out


def _enqueue_sales_cache_warmup(
    user_id: int,
    *,
    marketplace: str,
    date_from: date,
    date_to: date,
    granularity: str,
    tz_name: str,
) -> dict[str, Any]:
    return enqueue_task(
        "warm_sales_cache",
        {
            "user_id": int(user_id),
            "marketplace": str(marketplace or "all").strip().lower(),
            "date_from": date_from.isoformat(),
            "date_to": date_to.isoformat(),
            "granularity": str(granularity or "auto").strip().lower(),
            "tz": str(tz_name or "UTC").strip() or "UTC",
        },
        dedupe_key=(
            f"warm_sales:{int(user_id)}:{str(marketplace or 'all').strip().lower()}:"
            f"{date_from.isoformat()}:{date_to.isoformat()}:{str(granularity or 'auto').strip().lower()}:{str(tz_name or 'UTC').strip()}"
        ),
        dedupe_ttl_sec=90,
    )

BILLING_PLANS: dict[str, dict[str, Any]] = {
    "starter": {"title": "Starter", "price": 990, "limits": {"products": 500, "seo_jobs_month": 1500, "ai_replies_month": 800}},
    "pro": {"title": "Pro", "price": 2990, "limits": {"products": 5000, "seo_jobs_month": 10000, "ai_replies_month": 5000}},
    "business": {"title": "Business", "price": 8990, "limits": {"products": 50000, "seo_jobs_month": 100000, "ai_replies_month": 50000}},
}
AI_PROVIDER_CODES = ("openai", "openrouter", "deepseek", "groq", "together", "mistral", "xai", "custom")

HELP_DOCS_RU: dict[str, dict[str, str]] = {
    "dashboard": {
        "title": "Р”Р°СЂР±РѕСЂРґ",
        "content": (
            "РќР°Р·РЅР°С‡РµРЅРёРµ: РµР¶РµРґРЅРµРІРЅС‹Р№ РєРѕРЅС‚СЂРѕР»СЊ СЃРѕСЃС‚РѕСЏРЅРёСЏ РїСЂРѕРµРєС‚Р°.\n\n"
            "Р‘Р»РѕРєРё Рё С„СѓРЅРєС†РёРё:\n"
            "1) KPI-РєР°СЂС‚РѕС‡РєРё: РїРѕРєР°Р·С‹РІР°СЋС‚ РѕР±С‰РµРµ С‡РёСЃР»Рѕ С‚РѕРІР°СЂРѕРІ, SEO-Р·Р°РґР°С‡, Р·Р°РґР°С‡ РІ СЂР°Р±РѕС‚Рµ Рё РІ С‚РѕРї-5.\n"
            "2) РўСЂРµРЅРґ РїРѕР·РёС†РёР№ (21 РґРµРЅСЊ): РѕС‚РѕР±СЂР°Р¶Р°РµС‚ РґРёРЅР°РјРёРєСѓ СЃСЂРµРґРЅРµР№ РїРѕР·РёС†РёРё Рё РїР»РѕС‚РЅРѕСЃС‚СЊ РїСЂРѕРІРµСЂРѕРє.\n"
            "3) Р‘С‹СЃС‚СЂС‹Рµ РґРµР№СЃС‚РІРёСЏ:\n"
            "- РРјРїРѕСЂС‚РёСЂРѕРІР°С‚СЊ С‚РѕРІР°СЂС‹: РїРµСЂРµС…РѕРґ РІ РјРѕРґСѓР»СЊ В«РўРѕРІР°СЂС‹В».\n"
            "- Р—Р°РїСѓСЃС‚РёС‚СЊ SEO-РіРµРЅРµСЂР°С†РёСЋ: РїРµСЂРµС…РѕРґ РІ В«SEO Р·Р°РґР°С‡РёВ».\n"
            "- РџСЂРѕРІРµСЂРёС‚СЊ РїРѕР·РёС†РёРё РІСЃРµС…: РјР°СЃСЃРѕРІР°СЏ РїСЂРѕРІРµСЂРєР° РїРѕР·РёС†РёР№.\n\n"
            "РџСЂРёРјРµСЂ: СѓС‚СЂРѕРј РѕС‚РєСЂРѕР№С‚Рµ РґР°С€Р±РѕСЂРґ, РїСЂРѕРІРµСЂСЊС‚Рµ СЂРѕСЃС‚ KPI Рё Р·Р°РїСѓСЃС‚РёС‚Рµ РјР°СЃСЃРѕРІСѓСЋ РїСЂРѕРІРµСЂРєСѓ, РµСЃР»Рё СЃРЅРёР·РёР»РѕСЃСЊ С‡РёСЃР»Рѕ С‚РѕРІР°СЂРѕРІ РІ С‚РѕРї-5."
        ),
    },
    "products": {
        "title": "РўРѕРІР°СЂС‹",
        "content": (
            "РќР°Р·РЅР°С‡РµРЅРёРµ: РёРјРїРѕСЂС‚ РєР°С‚Р°Р»РѕРіР°, С„РёР»СЊС‚СЂР°С†РёСЏ, РїРѕСЃС‚СЂР°РЅРёС‡РЅР°СЏ СЂР°Р±РѕС‚Р° Рё СѓРїСЂР°РІР»РµРЅРёРµ РєР°СЂС‚РѕС‡РєР°РјРё.\n\n"
            "РљРЅРѕРїРєРё Рё РїРѕР»СЏ:\n"
            "- РРјРїРѕСЂС‚: Р·Р°РіСЂСѓР¶Р°РµС‚ С‚РѕРІР°СЂС‹ РёР· РІС‹Р±СЂР°РЅРЅРѕРіРѕ РјР°СЂРєРµС‚РїР»РµР№СЃР°.\n"
            "- РџРµСЂРµР·Р°РіСЂСѓР·РёС‚СЊ Р±Р°Р·Сѓ: РїРѕР»РЅРѕСЃС‚СЊСЋ РїРµСЂРµСЃРѕР±РёСЂР°РµС‚ Р»РѕРєР°Р»СЊРЅС‹Р№ РєР°С‚Р°Р»РѕРі.\n"
            "- Р’С‹Р±СЂР°С‚СЊ РІСЃРµ: РІС‹РґРµР»СЏРµС‚ РІСЃРµ СЃС‚СЂРѕРєРё РІ С‚РµРєСѓС‰РµР№ РІС‹РґР°С‡Рµ.\n"
            "- РџРµСЂРµРєР»СЋС‡РµРЅРёРµ СЃС‚СЂР°РЅРёС†: РЅР°РІРёРіР°С†РёСЏ РїРѕ РєР°С‚Р°Р»РѕРіСѓ СЃРІРµСЂС…Сѓ Рё СЃРЅРёР·Сѓ С‚Р°Р±Р»РёС†С‹.\n"
            "- Р Р°Р·РјРµСЂ СЃС‚СЂР°РЅРёС†С‹: 30/50/100/200/500/1000 СЃС‚СЂРѕРє.\n"
            "- РџРѕР»Рµ В«Р¤РёР»СЊС‚СЂВ»: Р±С‹СЃС‚СЂС‹Р№ РїРѕРёСЃРє РїРѕ Р°СЂС‚РёРєСѓР»Сѓ/РЅР°Р·РІР°РЅРёСЋ.\n"
            "- Р”РµС‚Р°Р»Рё СЃРїСЂР°РІР°: РїСЂРѕСЃРјРѕС‚СЂ С„РѕС‚Рѕ, Р°С‚СЂРёР±СѓС‚РѕРІ Рё СЂРµРґР°РєС‚РёСЂРѕРІР°РЅРёРµ РєР°СЂС‚РѕС‡РєРё.\n\n"
            "РџРѕР»СЏ С†РµРЅ РІ РєР°СЂС‚РѕС‡РєРµ:\n"
            "- Р—Р°РєСѓРїРѕС‡РЅР°СЏ С†РµРЅР°: СЃРµР±РµСЃС‚РѕРёРјРѕСЃС‚СЊ РµРґРёРЅРёС†С‹ (РёСЃРїРѕР»СЊР·СѓРµС‚СЃСЏ РІ Р±СѓС…РіР°Р»С‚РµСЂРёРё РґР»СЏ COGS/РјР°СЂР¶Рё).\n"
            "- Р¦РµРЅР° Р±РµР· СЃРєРёРґРєРё: Р±Р°Р·РѕРІР°СЏ (СЃС‚Р°СЂР°СЏ) С†РµРЅР° РґРѕ РїСЂРѕРјРѕ Рё СЃРєРёРґРѕРє.\n"
            "- Р¦РµРЅР° СЃРѕ СЃРєРёРґРєРѕР№: С‚РµРєСѓС‰Р°СЏ С„Р°РєС‚РёС‡РµСЃРєР°СЏ С†РµРЅР° РїСЂРѕРґР°Р¶Рё.\n"
            "- РњРёРЅ. С†РµРЅР°: РЅРёР¶РЅСЏСЏ РіСЂР°РЅРёС†Р° С†РµРЅС‹ РїРѕ РІР°СЂРµР№ СЃС‚СЂР°С‚РµРіРёРё.\n"
            "- РњР°СЂРєРµС‚РёРЅРі С†РµРЅР°: РїСЂРѕРјРѕ-С†РµРЅР° РґР»СЏ СЂРµРєР»Р°РјРЅС‹С…/РјР°СЂРєРµС‚РёРЅРіРѕРІС‹С… СЃС†РµРЅР°СЂРёРµРІ.\n\n"
            "Р¤РѕС‚Рѕ РІ СЂРµРґР°РєС‚РѕСЂРµ:\n"
            "- В«Р”РѕР±Р°РІРёС‚СЊ URLВ»: РґРѕР±Р°РІР»СЏРµС‚ С„РѕС‚Рѕ РїРѕ РїСЂСЏРјРѕР№ СЃСЃС‹Р»РєРµ.\n"
            "- В«Р”РѕР±Р°РІРёС‚СЊ С„РѕС‚РѕВ»: РѕС‚РєСЂС‹РІР°РµС‚ Р»РѕРєР°Р»СЊРЅС‹Рµ РїР°РїРєРё Рё РїРѕРґРґРµСЂР¶РёРІР°РµС‚ РІС‹Р±РѕСЂ РЅРµСЃРєРѕР»СЊРєРёС… С„Р°Р№Р»РѕРІ.\n"
            "- РџРµСЂРµС‚Р°СЃРєРёРІР°РЅРёРµ С„РѕС‚Рѕ РјРµРЅСЏРµС‚ РїРѕСЂСЏРґРѕРє; РїРµСЂРІРѕРµ С„РѕС‚Рѕ СЃС‚Р°РЅРѕРІРёС‚СЃСЏ РіР»Р°РІРЅС‹Рј.\n\n"
            "Р РµРєРѕРјРµРЅРґСѓРµРјС‹Р№ РїРѕСЂСЏРґРѕРє СЂР°Р±РѕС‚С‹:\n"
            "1) РРјРїРѕСЂС‚РёСЂСѓР№С‚Рµ РєР°С‚Р°Р»РѕРі.\n"
            "2) РћС‚С„РёР»СЊС‚СЂСѓР№С‚Рµ РЅСѓР¶РЅСѓСЋ РіСЂСѓРїРїСѓ SKU.\n"
            "3) РќР°СЃС‚СЂРѕР№С‚Рµ СЂР°Р·РјРµСЂ СЃС‚СЂР°РЅРёС†С‹ Рё РїСЂРѕР»РёСЃС‚Р°Р№С‚Рµ РЅСѓР¶РЅС‹Р№ СЃРµРіРјРµРЅС‚ РєР°С‚Р°Р»РѕРіР°.\n"
            "4) РћС‚РєСЂРѕР№С‚Рµ РєР°СЂС‚РѕС‡РєСѓ СЃРїСЂР°РІР° Рё РїСЂРё РЅРµРѕР±С…РѕРґРёРјРѕСЃС‚Рё РѕР±РЅРѕРІРёС‚Рµ РґР°РЅРЅС‹Рµ.\n"
            "5) Р”Р»СЏ РїСЂРѕРІРµСЂРєРё РїРѕР·РёС†РёР№ РїРµСЂРµР№РґРёС‚Рµ РІ РїРѕРґРјРѕРґСѓР»СЊ В«SEO Р·Р°РґР°С‡РёВ».\n\n"
            "Р”РёР°РіРЅРѕСЃС‚РёРєР°:\n"
            "- Р•СЃР»Рё С„РёР»СЊС‚СЂ WB/Ozon РЅРµ СЃСЂР°Р±Р°С‚С‹РІР°РµС‚, РїСЂРѕРІРµСЂСЊС‚Рµ РІС‹Р±СЂР°РЅРЅС‹Р№ РјР°СЂРєРµС‚РїР»РµР№СЃ РІ СЃРµР»РµРєС‚Рµ РЅР°Рґ С‚Р°Р±Р»РёС†РµР№.\n"
            "- Р•СЃР»Рё С‚Р°Р±Р»РёС†Р° РїСѓСЃС‚Р°СЏ РїРѕСЃР»Рµ РёРјРїРѕСЂС‚Р°, РѕС‡РёСЃС‚РёС‚Рµ С„РёР»СЊС‚СЂ Рё РїРѕРІС‚РѕСЂРёС‚Рµ Р·Р°РіСЂСѓР·РєСѓ.\n"
            "- РџРѕСЃР»Рµ РґРѕР±Р°РІР»РµРЅРёСЏ/РїРµСЂРµС‚Р°СЃРєРёРІР°РЅРёСЏ С„РѕС‚Рѕ РЅР°Р¶РјРёС‚Рµ В«РЎРѕС…СЂР°РЅРёС‚СЊВ», С‡С‚РѕР±С‹ Р·Р°С„РёРєСЃРёСЂРѕРІР°С‚СЊ РїРѕСЂСЏРґРѕРє.\n\n"
            "РџСЂРёРјРµСЂ: РёРјРїРѕСЂС‚РёСЂСѓР№С‚Рµ С‚РѕРІР°СЂС‹ Ozon, РІРєР»СЋС‡РёС‚Рµ 100 СЃС‚СЂРѕРє РЅР° СЃС‚СЂР°РЅРёС†Сѓ, РѕС‚С„РёР»СЊС‚СЂСѓР№С‚Рµ В«РґС‹РјРѕС…РѕРґВ», РѕС‚РєСЂРѕР№С‚Рµ РєР°СЂС‚РѕС‡РєСѓ Рё РѕР±РЅРѕРІРёС‚Рµ РѕРїРёСЃР°РЅРёРµ."
        ),
    },
    "seo_generation": {
        "title": "SEO Р·Р°РґР°С‡Рё",
        "content": (
            "РќР°Р·РЅР°С‡РµРЅРёРµ: РіРµРЅРµСЂР°С†РёСЏ SEO-РѕРїРёСЃР°РЅРёР№ Рё РёС‚РµСЂР°С‚РёРІРЅРѕРµ СѓР»СѓС‡СЂРµРЅРёРµ.\n\n"
            "РљРЅРѕРїРєРё Рё РґРµР№СЃС‚РІРёСЏ:\n"
            "- РЎРіРµРЅРµСЂРёСЂРѕРІР°С‚СЊ РґР»СЏ РІС‹Р±СЂР°РЅРЅС‹С… / РґР»СЏ РІСЃРµС…: СЃРѕР·РґР°РµС‚ SEO-Р·Р°РґР°С‡Рё.\n"
            "- РџСЂРёРјРµРЅРёС‚СЊ: РѕС‚РїСЂР°РІР»СЏРµС‚ СѓС‚РІРµСЂР¶РґРµРЅРЅС‹Рµ РѕРїРёСЃР°РЅРёСЏ РІ РјР°СЂРєРµС‚РїР»РµР№СЃ.\n"
            "- Recheck РІС‹Р±СЂР°РЅРЅС‹С…: РїРѕРІС‚РѕСЂРЅР°СЏ РїСЂРѕРІРµСЂРєР° РјРµС‚СЂРёРє/РїРѕР·РёС†РёР№ РїРѕ Р·Р°РґР°С‡Р°Рј.\n"
            "- Recheck РїСЂРѕСЃСЂРѕС‡РµРЅРЅС‹С…: РјР°СЃСЃРѕРІР°СЏ Р°РєС‚СѓР°Р»РёР·Р°С†РёСЏ Р·Р°РґР°С‡ СЃ РёСЃС‚РµРєСЂРµР№ РґР°С‚РѕР№.\n"
            "- Р’С‹Р±СЂР°С‚СЊ РІСЃРµ: РІС‹РґРµР»РµРЅРёРµ РІСЃРµС… Р·Р°РґР°С‡ РІ С‚РµРєСѓС‰РµРј СЃРїРёСЃРєРµ.\n"
            "- РЈРґР°Р»РёС‚СЊ РІС‹Р±СЂР°РЅРЅС‹Рµ / РћС‡РёСЃС‚РёС‚СЊ РІСЃРµ SEO Р·Р°РґР°С‡Рё: СѓРґР°Р»РµРЅРёРµ Р·Р°РґР°С‡.\n\n"
            "Р–РёР·РЅРµРЅРЅС‹Р№ С†РёРєР» Р·Р°РґР°С‡Рё:\n"
            "1) generated: С‚РµРєСЃС‚ СЃРіРµРЅРµСЂРёСЂРѕРІР°РЅ Рё РґРѕСЃС‚СѓРїРµРЅ РґР»СЏ РїСЂРѕСЃРјРѕС‚СЂР°.\n"
            "2) in_progress: Р·Р°РґР°С‡Р° РїСЂРѕРІРµСЂСЏРµС‚СЃСЏ РёР»Рё РґРѕСЂР°Р±Р°С‚С‹РІР°РµС‚СЃСЏ.\n"
            "3) applied: РѕРїРёСЃР°РЅРёРµ РїСЂРёРјРµРЅРµРЅРѕ РІ РјР°СЂРєРµС‚РїР»РµР№СЃРµ.\n"
            "4) top_5: РґРѕСЃС‚РёРіРЅСѓС‚Р° С†РµР»РµРІР°СЏ РїРѕР·РёС†РёСЏ (РµСЃР»Рё Р±С‹Р»Р° Р·Р°РґР°РЅР°).\n\n"
            "Р РµРєРѕРјРµРЅРґР°С†РёРё:\n"
            "- РСЃРїРѕР»СЊР·СѓР№С‚Рµ В«Recheck РІС‹Р±СЂР°РЅРЅС‹С…В» РїРѕСЃР»Рµ СЂСѓС‡РЅС‹С… РїСЂР°РІРѕРє С‚РµРєСЃС‚Р°.\n"
            "- РСЃРїРѕР»СЊР·СѓР№С‚Рµ В«Recheck РїСЂРѕСЃСЂРѕС‡РµРЅРЅС‹С…В» РєР°Рє СЂРµРіСѓР»СЏСЂРЅСѓСЋ РѕРїРµСЂР°С†РёСЋ, РЅР°РїСЂРёРјРµСЂ СЂР°Р· РІ РґРµРЅСЊ.\n"
            "- РџРµСЂРµРґ В«РџСЂРёРјРµРЅРёС‚СЊВ» РїСЂРѕРІРµСЂСЏР№С‚Рµ Р±Р»РѕРє В«РџСЂРµРґРїСЂРѕСЃРјРѕС‚СЂВ» Рё СЃРїРёСЃРѕРє РєРѕРЅРєСѓСЂРµРЅС‚РѕРІ.\n\n"
            "Р”РёР°РіРЅРѕСЃС‚РёРєР°:\n"
            "- Р•СЃР»Рё РїСЂРёРјРµРЅРµРЅРёРµ РЅРµ РїСЂРѕС…РѕРґРёС‚, РїСЂРѕРІРµСЂСЊС‚Рµ РґРѕСЃС‚СѓРїРЅРѕСЃС‚СЊ API РјР°СЂРєРµС‚РїР»РµР№СЃР° Рё РїСЂР°РІР° РєР»СЋС‡Р°.\n"
            "- Р•СЃР»Рё РїСЂРѕРіСЂРµСЃСЃ Р·Р°РґР°С‡ РЅРµ РјРµРЅСЏРµС‚СЃСЏ, Р·Р°РїСѓСЃС‚РёС‚Рµ recheck Рё СЃРІРµСЂСЏР№С‚Рµ РґР°С‚Сѓ СЃР»РµРґСѓСЋС‰РµР№ РїСЂРѕРІРµСЂРєРё.\n\n"
            "РџСЂРёРјРµСЂ: СЃРіРµРЅРµСЂРёСЂСѓР№С‚Рµ РѕРїРёСЃР°РЅРёСЏ РґР»СЏ 20 SKU, РїСЂРѕРІРµСЂСЊС‚Рµ РїСЂРµРІСЊСЋ, РѕС‚РїСЂР°РІСЊС‚Рµ В«РџСЂРёРјРµРЅРёС‚СЊВ», Р·Р°С‚РµРј С‡РµСЂРµР· recheck РѕС†РµРЅРёС‚Рµ РёР·РјРµРЅРµРЅРёРµ РїРѕР·РёС†РёРё."
        ),
    },
    "wb_reviews_ai": {
        "title": "РћС‚Р·С‹РІС‹ WB/Ozon",
        "content": (
            "РќР°Р·РЅР°С‡РµРЅРёРµ: РѕР±СЂР°Р±РѕС‚РєР° РѕС‚Р·С‹РІРѕРІ Рё РѕС‚РїСЂР°РІРєР° РѕС‚РІРµС‚РѕРІ.\n\n"
            "Р§С‚Рѕ РґРµР»Р°РµС‚ РєР°Р¶РґР°СЏ РєРЅРѕРїРєР°:\n"
            "- РћР±РЅРѕРІРёС‚СЊ РѕС‚Р·С‹РІС‹: Р·Р°РіСЂСѓР¶Р°РµС‚ РЅРѕРІС‹Рµ РґР°РЅРЅС‹Рµ РёР· WB/Ozon.\n"
            "- РЎРѕС…СЂР°РЅРёС‚СЊ AI-РЅР°СЃС‚СЂРѕР№РєРё: СЃРѕС…СЂР°РЅСЏРµС‚ СЂРµР¶РёРј Рё РїСЂРѕРјРїС‚ РіРµРЅРµСЂР°С‚РѕСЂР°.\n"
            "- РРєРѕРЅРєР° AI РІ СЃС‚СЂРѕРєРµ: РіРµРЅРµСЂРёСЂСѓРµС‚ С‡РµСЂРЅРѕРІРёРє РѕС‚РІРµС‚Р°.\n"
            "- РРєРѕРЅРєР° РѕС‚РїСЂР°РІРєРё: РїСѓР±Р»РёРєСѓРµС‚ РѕС‚РІРµС‚ (РёР»Рё РѕР±РЅРѕРІР»СЏРµС‚ СЂР°РЅРµРµ РѕС‚РїСЂР°РІР»РµРЅРЅС‹Р№).\n"
            "- Р¤РёР»СЊС‚СЂС‹: РјР°СЂРєРµС‚РїР»РµР№СЃ, РѕС†РµРЅРєР°, СЃС‚Р°С‚СѓСЃ, СЃРѕСЂС‚РёСЂРѕРІРєР°, РїРµСЂРёРѕРґ.\n\n"
            "Р’Р°Р¶РЅРѕ:\n"
            "- Р”Р°С‚Р° В«СЃ/РїРѕВ» С„РёР»СЊС‚СЂСѓРµС‚ СѓР¶Рµ Р·Р°РіСЂСѓР¶РµРЅРЅС‹Рµ Р·Р°РїРёСЃРё. Р•СЃР»Рё СЃС‚СЂРѕРє РЅРµС‚, СЃРЅР°С‡Р°Р»Р° РѕС‡РёСЃС‚РёС‚Рµ С„РёР»СЊС‚СЂ РґР°С‚.\n"
            "- РЎС‚Р°С‚СѓСЃ-Р±Р°СЂ РїРѕРєР°Р·С‹РІР°РµС‚ РїСЂРѕРіСЂРµСЃСЃ РґРѕРіСЂСѓР·РєРё: СЃРЅР°С‡Р°Р»Р° Р±С‹СЃС‚СЂС‹Р№ СЃР»РѕР№, Р·Р°С‚РµРј РїРѕР»РЅС‹Р№.\n"
            "- Р’ РєРѕР»РѕРЅРєРµ В«РћС‚РІРµС‚В» СЃРѕС…СЂР°РЅСЏРµС‚СЃСЏ С‡РµСЂРЅРѕРІРёРє; РѕС‚РїСЂР°РІРєР° РІ РјР°СЂРєРµС‚РїР»РµР№СЃ РІС‹РїРѕР»РЅСЏРµС‚СЃСЏ С‚РѕР»СЊРєРѕ РїРѕ РєРЅРѕРїРєРµ РґРµР№СЃС‚РІРёСЏ.\n\n"
            "РўРёРїРѕРІРѕР№ СЃС†РµРЅР°СЂРёР№:\n"
            "1) Р’С‹Р±РµСЂРёС‚Рµ WB РёР»Рё Ozon.\n"
            "2) РџРѕСЃС‚Р°РІСЊС‚Рµ С„РёР»СЊС‚СЂ В«РќРµРѕС‚РІРµС‡РµРЅРЅС‹РµВ».\n"
            "3) РќР°Р¶РјРёС‚Рµ РёРєРѕРЅРєСѓ AI, РѕС‚СЂРµРґР°РєС‚РёСЂСѓР№С‚Рµ С‚РµРєСЃС‚ РїСЂРё РЅРµРѕР±С…РѕРґРёРјРѕСЃС‚Рё.\n"
            "4) РќР°Р¶РјРёС‚Рµ РѕС‚РїСЂР°РІРєСѓ Рё РґРѕР¶РґРёС‚РµСЃСЊ СЃС‚Р°С‚СѓСЃР° В«РѕС‚РІРµС‡РµРЅВ».\n\n"
            "РџСЂРёРјРµСЂ: РІС‹Р±РµСЂРёС‚Рµ В«РќРµРѕС‚РІРµС‡РµРЅРЅС‹РµВ», РЅР°Р¶РјРёС‚Рµ AI-РёРєРѕРЅРєСѓ, РїРѕРїСЂР°РІСЊС‚Рµ С‚РѕРЅ РѕС‚РІРµС‚Р° Рё РѕС‚РїСЂР°РІСЊС‚Рµ РІ РєР°СЂС‚РѕС‡РєСѓ РѕС‚Р·С‹РІР°."
        ),
    },
    "wb_questions_ai": {
        "title": "Р’РѕРїСЂРѕСЃС‹ WB/Ozon",
        "content": (
            "РќР°Р·РЅР°С‡РµРЅРёРµ: РѕС‚РІРµС‚С‹ РЅР° РІРѕРїСЂРѕСЃС‹ РїРѕРєСѓРїР°С‚РµР»РµР№ РїРѕ С‚РѕРІР°СЂР°Рј.\n\n"
            "РљРЅРѕРїРєРё Рё Р±Р»РѕРєРё:\n"
            "- РћР±РЅРѕРІРёС‚СЊ РІРѕРїСЂРѕСЃС‹: Р·Р°РіСЂСѓР¶Р°РµС‚ РІРѕРїСЂРѕСЃС‹ РёР· WB/Ozon.\n"
            "- РЎРѕС…СЂР°РЅРёС‚СЊ AI-РЅР°СЃС‚СЂРѕР№РєРё: СЃРѕС…СЂР°РЅСЏРµС‚ СЂРµР¶РёРј Рё РїСЂРѕРјРїС‚.\n"
            "- Р—Р°РіСЂСѓР·РёС‚СЊ РІ Р±Р°Р·Сѓ Р·РЅР°РЅРёР№: РґРѕР±Р°РІР»СЏРµС‚ РґРѕРєСѓРјРµРЅС‚ РґР»СЏ РєРѕРЅС‚РµРєСЃС‚Р° AI.\n"
            "- РЈРґР°Р»РёС‚СЊ РІС‹Р±СЂР°РЅРЅС‹Р№ РґРѕРєСѓРјРµРЅС‚: СѓРґР°Р»СЏРµС‚ РґРѕРєСѓРјРµРЅС‚ РёР· Р±Р°Р·С‹ Р·РЅР°РЅРёР№.\n"
            "- РРєРѕРЅРєР° AI РІ СЃС‚СЂРѕРєРµ: РіРµРЅРµСЂРёСЂСѓРµС‚ РѕС‚РІРµС‚ РЅР° РІРѕРїСЂРѕСЃ.\n"
            "- РРєРѕРЅРєР° РѕС‚РїСЂР°РІРєРё: РїСѓР±Р»РёРєСѓРµС‚/РѕР±РЅРѕРІР»СЏРµС‚ РѕС‚РІРµС‚.\n\n"
            "Р’Р°Р¶РЅРѕ:\n"
            "- Р”Р»СЏ Р±С‹СЃС‚СЂС‹С… РѕС‚РІРµС‚РѕРІ Р·Р°РіСЂСѓР·РёС‚Рµ FAQ/СЂРµРіР»Р°РјРµРЅС‚ РІ Р±Р°Р·Сѓ Р·РЅР°РЅРёР№ РїРµСЂРµРґ РіРµРЅРµСЂР°С†РёРµР№.\n"
            "- РџСЂРё РїСѓСЃС‚РѕР№ С‚Р°Р±Р»РёС†Рµ СЃРЅР°С‡Р°Р»Р° РїСЂРѕРІРµСЂСЊС‚Рµ С„РёР»СЊС‚СЂ РґР°С‚ Рё РјР°СЂРєРµС‚РїР»РµР№СЃ.\n"
            "- РЎС‚Р°С‚СѓСЃ-Р±Р°СЂ Рё RAW-Р±Р»РѕРє РїРѕРјРѕРіР°СЋС‚ РїРѕРЅСЏС‚СЊ, РёРґРµС‚ Р»Рё Р·Р°РіСЂСѓР·РєР° РёР»Рё РµСЃС‚СЊ РѕС€РёР±РєР° API.\n\n"
            "РўРёРїРѕРІРѕР№ СЃС†РµРЅР°СЂРёР№:\n"
            "1) Р’С‹Р±РµСЂРёС‚Рµ РјР°СЂРєРµС‚РїР»РµР№СЃ Рё СЃС‚Р°С‚СѓСЃ В«РќРѕРІС‹РµВ».\n"
            "2) РџСЂРё РЅРµРѕР±С…РѕРґРёРјРѕСЃС‚Рё Р·Р°РіСЂСѓР·РёС‚Рµ РґРѕРєСѓРјРµРЅС‚ РІ Р±Р°Р·Сѓ Р·РЅР°РЅРёР№.\n"
            "3) РЎРіРµРЅРµСЂРёСЂСѓР№С‚Рµ РѕС‚РІРµС‚, РѕС‚СЂРµРґР°РєС‚РёСЂСѓР№С‚Рµ С„РѕСЂРјСѓР»РёСЂРѕРІРєРё.\n"
            "4) РћС‚РїСЂР°РІСЊС‚Рµ РѕС‚РІРµС‚ Рё РѕР±РЅРѕРІРёС‚Рµ СЃРїРёСЃРѕРє.\n\n"
            "РџСЂРёРјРµСЂ: Р·Р°РіСЂСѓР·РёС‚Рµ FAQ РїРѕСЃС‚Р°РІС‰РёРєР° РІ Р±Р°Р·Сѓ Р·РЅР°РЅРёР№, Р·Р°С‚РµРј СЃРіРµРЅРµСЂРёСЂСѓР№С‚Рµ Рё РѕС‚РїСЂР°РІСЊС‚Рµ РѕС‚РІРµС‚С‹ РЅР° РЅРѕРІС‹Рµ РІРѕРїСЂРѕСЃС‹."
        ),
    },
    "wb_ads": {
        "title": "Р РµРєР»Р°РјР° WB",
        "content": (
            "РќР°Р·РЅР°С‡РµРЅРёРµ: РјРѕРЅРёС‚РѕСЂРёРЅРі Рё РѕРїРµСЂР°С‚РёРІРЅРѕРµ СѓРїСЂР°РІР»РµРЅРёРµ СЂРµРєР»Р°РјРЅС‹РјРё РєР°РјРїР°РЅРёСЏРјРё WB.\n\n"
            "РљРЅРѕРїРєРё Рё С„СѓРЅРєС†РёРё:\n"
            "- Р—Р°РіСЂСѓР·РёС‚СЊ РєР°РјРїР°РЅРёРё: Р±С‹СЃС‚СЂР°СЏ Р·Р°РіСЂСѓР·РєР° СЃРїРёСЃРєР° РєР°РјРїР°РЅРёР№.\n"
            "- РџРѕР»СѓС‡РёС‚СЊ СЃС‚Р°РІРєРё: Р·Р°РїСЂРѕСЃ СЃС‚Р°РІРѕРє РїРѕ campaign_id.\n"
            "- РЎР±СЂРѕСЃРёС‚СЊ С„РёР»СЊС‚СЂС‹: РІРѕР·РІСЂР°С‚ С„РёР»СЊС‚СЂР°С†РёРё Рё СЃРѕСЂС‚РёСЂРѕРІРєРё Рє СѓРјРѕР»С‡Р°РЅРёСЋ.\n"
            "- РџРѕРёСЃРє/С„РёР»СЊС‚СЂС‹: ID, РЅР°Р·РІР°РЅРёРµ, СЃС‚Р°С‚СѓСЃ, С‚РёРї, СЂР°Р±РѕС‚Р°РµС‚, Р±СЋРґР¶РµС‚.\n"
            "- Р”РІРѕР№РЅРѕР№ РєР»РёРє РїРѕ СЃС‚СЂРѕРєРµ: РѕС‚РєСЂС‹РІР°РµС‚ РјРѕРґР°Р»СЊРЅРѕРµ РѕРєРЅРѕ СЃ РґРµС‚Р°Р»СЏРјРё.\n"
            "- Р’ РґРµС‚Р°Р»СЏС… РєР°РјРїР°РЅРёРё: Р—Р°РїСѓСЃС‚РёС‚СЊ / РџР°СѓР·Р° / РћСЃС‚Р°РЅРѕРІРёС‚СЊ / РћР±РЅРѕРІРёС‚СЊ РґРµС‚Р°Р»Рё.\n\n"
            "РџРѕРґРІРєР»Р°РґРєР° В«Р‘РёРґРµСЂ WB AdsВ»:\n"
            "- РћС‚РєСЂРѕР№С‚Рµ В«Р РµРєР»Р°РјР°В» -> В«Р‘РёРґРµСЂВ».\n"
            "- Р—Р°РїРѕР»РЅРёС‚Рµ campaign_id Рё nm_id.\n"
            "- Р’С‹Р±РµСЂРёС‚Рµ С†РµР»СЊ:\n"
            "  normquery: С„СЂР°Р·Р° (СѓРєР°Р¶РёС‚Рµ target_value),\n"
            "  nm: РєР°СЂС‚РѕС‡РєР° (target_value РјРѕР¶РЅРѕ РѕСЃС‚Р°РІРёС‚СЊ РїСѓСЃС‚С‹Рј).\n"
            "- Р’С‹Р±РµСЂРёС‚Рµ СЃС‚СЂР°С‚РµРіРёСЋ:\n"
            "  optimal: РґРµСЂР¶Р°С‚СЊ СЃС‚Р°РІРєСѓ РІ РѕРїС‚РёРјСѓРјРµ WB,\n"
            "  position: СѓРґРµСЂР¶РёРІР°С‚СЊ РїРѕР·РёС†РёСЋ РІ РґРёР°РїР°Р·РѕРЅРµ,\n"
            "  range: СѓРґРµСЂР¶РёРІР°С‚СЊ СЃС‚Р°РІРєСѓ РІ РєРѕСЂРёРґРѕСЂРµ min/max,\n"
            "  hold: С„РёРєСЃРёСЂРѕРІР°С‚СЊ С‚РµРєСѓС‰СѓСЋ Р»РѕРіРёРєСѓ Р±РµР· Р°РіСЂРµСЃСЃРёРІРЅС‹С… СЂР°РіРѕРІ.\n"
            "- РЈРєР°Р¶РёС‚Рµ min/max/step Рё cooldown (СЃРµРєСѓРЅРґС‹ РјРµР¶РґСѓ Р·Р°РїСѓСЃРєР°РјРё РїСЂР°РІРёР»Р°).\n"
            "- РќР°Р¶РјРёС‚Рµ В«РЎРѕС…СЂР°РЅРёС‚СЊ РїСЂР°РІРёР»РѕВ».\n"
            "- Р”Р»СЏ СЂСѓС‡РЅРѕРіРѕ Р·Р°РїСѓСЃРєР°: В«Р—Р°РїСѓСЃС‚РёС‚СЊ СЃРµР№С‡Р°СЃВ».\n"
            "- Р”Р»СЏ С„РѕРЅРѕРІРѕРіРѕ СЂРµР¶РёРјР°: РѕСЃС‚Р°РІСЊС‚Рµ В«РђРєС‚РёРІРЅРѕВ», РІРѕСЂРєРµСЂ Р±СѓРґРµС‚ Р·Р°РїСѓСЃРєР°С‚СЊ РїСЂР°РІРёР»Рѕ РїРѕ cooldown.\n\n"
            "Р‘С‹СЃС‚СЂС‹Р№ СЃС‚Р°СЂС‚ (3 РјРёРЅСѓС‚С‹):\n"
            "1) Р’ В«РљР°РјРїР°РЅРёРёВ» РЅР°Р№РґРёС‚Рµ Р°РєС‚РёРІРЅСѓСЋ РєР°РјРїР°РЅРёСЋ Рё СЃРєРѕРїРёСЂСѓР№С‚Рµ campaign_id.\n"
            "2) Р’ В«Р‘РёРґРµСЂВ» РІСЃС‚Р°РІСЊС‚Рµ campaign_id Рё nm_id, РІС‹Р±РµСЂРёС‚Рµ normquery РёР»Рё nm.\n"
            "3) Р”Р»СЏ normquery РѕР±СЏР·Р°С‚РµР»СЊРЅРѕ Р·Р°РїРѕР»РЅРёС‚Рµ С„СЂР°Р·Сѓ target_value.\n"
            "4) РЎС‚Р°СЂС‚РѕРІС‹Рµ РЅР°СЃС‚СЂРѕР№РєРё: step=50..100, cooldown=300, min_bid/max_bid РІ Р±РµР·РѕРїР°СЃРЅРѕРј РґРёР°РїР°Р·РѕРЅРµ.\n"
            "5) РќР°Р¶РјРёС‚Рµ В«РЎРѕС…СЂР°РЅРёС‚СЊ РїСЂР°РІРёР»РѕВ», Р·Р°С‚РµРј В«Р—Р°РїСѓСЃС‚РёС‚СЊ СЃРµР№С‡Р°СЃВ» Рё РїСЂРѕРІРµСЂСЊС‚Рµ СЃС‚СЂРѕРєСѓ РІ Р»РѕРіР°С….\n"
            "6) Р•СЃР»Рё Р»РѕРі ok/skipped Р±РµР· РѕС€РёР±РѕРє вЂ” РІРєР»СЋС‡Р°Р№С‚Рµ С„РѕРЅРѕРІС‹Р№ СЂРµР¶РёРј.\n\n"
            "РљР°Рє С‡РёС‚Р°С‚СЊ Р»РѕРіРё Р±РёРґРґРµСЂР°:\n"
            "- ok: СЃС‚Р°РІРєР° РїСЂРёРјРµРЅРµРЅР° РёР»Рё РїРѕРґС‚РІРµСЂР¶РґРµРЅР°.\n"
            "- skipped: Р·Р°РїСѓСЃРє РїСЂРѕРїСѓС‰РµРЅ РїРѕ СѓСЃР»РѕРІРёСЏРј (cooldown/РЅРµС‚ РґР°РЅРЅС‹С…/Р»РёРјРёС‚С‹).\n"
            "- error: РѕС€РёР±РєР° API РёР»Рё РЅРµРєРѕСЂСЂРµРєС‚РЅР°СЏ РєРѕРЅС„РёРіСѓСЂР°С†РёСЏ РїСЂР°РІРёР»Р°.\n"
            "- РџСЂРёС‡РёРЅР° РІСЃРµРіРґР° РїРёСЂРµС‚СЃСЏ РІ РєРѕР»РѕРЅРєРµ В«РџСЂРёС‡РёРЅР°В» Рё РІ СЃС‚Р°С‚СѓСЃРµ РїСЂР°РІРёР»Р°.\n\n"
            "Р•СЃР»Рё РїСЂР°РІРёР»Рѕ РЅРµ СЂР°Р±РѕС‚Р°РµС‚:\n"
            "- РџСЂРѕРІРµСЂСЊС‚Рµ, С‡С‚Рѕ РєР»СЋС‡ WB Ads Р°РєС‚РёРІРµРЅ Рё Сѓ РїРѕР»СЊР·РѕРІР°С‚РµР»СЏ РµСЃС‚СЊ РґРѕСЃС‚СѓРї Рє РјРѕРґСѓР»СЋ В«Р РµРєР»Р°РјР° WB/OzonВ».\n"
            "- РЈР±РµРґРёС‚РµСЃСЊ, С‡С‚Рѕ campaign_id/nm_id РёР· РѕРґРЅРѕР№ РєР°РјРїР°РЅРёРё, Р° РЅРµ РёР· СЂР°Р·РЅС‹С… РєР°Р±РёРЅРµС‚РѕРІ.\n"
            "- Р”Р»СЏ normquery РїСЂРѕРІРµСЂСЊС‚Рµ С‚РѕС‡РЅРѕРµ СЃРѕРІРїР°РґРµРЅРёРµ С„СЂР°Р·С‹ (Р±РµР· Р»РёСЂРЅРёС… РїСЂРѕР±РµР»РѕРІ).\n"
            "- Р•СЃР»Рё С‡Р°СЃС‚Рѕ skipped(cooldown) вЂ” СѓРјРµРЅСЊС€РёС‚Рµ С‡Р°СЃС‚РѕС‚Сѓ Р·Р°РїСѓСЃРєР° РІСЂСѓС‡РЅСѓСЋ Рё СѓРІРµР»РёС‡СЊС‚Рµ cooldown.\n"
            "- Р•СЃР»Рё error 401/403 вЂ” РїРµСЂРµСѓСЃС‚Р°РЅРѕРІРёС‚Рµ WB Ads РєР»СЋС‡ РІ РїСЂРѕС„РёР»Рµ.\n\n"
            "РџСЂР°РєС‚РёРєР° РЅР°СЃС‚СЂРѕР№РєРё:\n"
            "- РќР°С‡РёРЅР°Р№С‚Рµ СЃ РєРѕРЅСЃРµСЂРІР°С‚РёРІРЅРѕРіРѕ СЂР°РіР° (50-100), С‡С‚РѕР±С‹ РЅРµ СЂР°СЃРєР°С‡Р°С‚СЊ СЂР°СЃС…РѕРґ.\n"
            "- РЎС‚Р°РІСЊС‚Рµ cooldown РЅРµ РјРµРЅСЊС€Рµ 180-300 СЃРµРєСѓРЅРґ РґР»СЏ СЃС‚Р°Р±РёР»СЊРЅС‹С… РєР°РјРїР°РЅРёР№.\n"
            "- Р”Р»СЏ РЅРѕРІС‹С… РєР°РјРїР°РЅРёР№ СЃРЅР°С‡Р°Р»Р° РїСЂРѕРІРµСЂСЊС‚Рµ СЂСѓС‡РЅРѕР№ Р·Р°РїСѓСЃРє, Р·Р°С‚РµРј РІРєР»СЋС‡Р°Р№С‚Рµ С„РѕРЅ.\n"
            "- Р•СЃР»Рё РїСЂР°РІРёР»Рѕ С‡Р°СЃС‚Рѕ РІ skipped/error, РїСЂРѕРІРµСЂСЊС‚Рµ С†РµР»СЊ, РєР»СЋС‡ WB Ads Рё С‚РёРї РєР°РјРїР°РЅРёРё.\n\n"
            "РџСЂРёРјРµСЂ: РѕС‚С„РёР»СЊС‚СЂСѓР№С‚Рµ В«С‚РѕР»СЊРєРѕ СЂР°Р±РѕС‚Р°РµС‚В», РЅР°Р№РґРёС‚Рµ РєР°РјРїР°РЅРёРё СЃ РІС‹СЃРѕРєРёРј СЂР°СЃС…РѕРґРѕРј Рё РїСЂРѕРІРµСЂСЊС‚Рµ РґРµС‚Р°Р»Рё РґРІРѕР№РЅС‹Рј РєР»РёРєРѕРј."
        ),
    },
    "wb_ads_analytics": {
        "title": "РђРЅР°Р»РёС‚РёРєР° WB Ads",
        "content": (
            "РќР°Р·РЅР°С‡РµРЅРёРµ: РѕС‚С‡РµС‚ РїРѕ СЌС„С„РµРєС‚РёРІРЅРѕСЃС‚Рё СЂРµРєР»Р°РјС‹ Р·Р° РїРµСЂРёРѕРґ.\n\n"
            "РџРѕР»СЏ Рё РґРµР№СЃС‚РІРёСЏ:\n"
            "- Р”Р°С‚Р° СЃ / РїРѕ: РІСЂРµРјРµРЅРЅРѕР№ РґРёР°РїР°Р·РѕРЅ РѕС‚С‡РµС‚Р°.\n"
            "- campaign_id (РѕРїС†РёРѕРЅР°Р»СЊРЅРѕ): РІС‹Р±РѕСЂ РѕРґРЅРѕР№ РєР°РјРїР°РЅРёРё.\n"
            "- РџРѕСЃС‚СЂРѕРёС‚СЊ РѕС‚С‡РµС‚: С„РѕСЂРјРёСЂСѓРµС‚ С‚Р°Р±Р»РёС†Сѓ Рё Р°РіСЂРµРіРёСЂРѕРІР°РЅРЅС‹Рµ totals.\n"
            "- РњРµС‚СЂРёРєРё: РїРѕРєР°Р·С‹, РєР»РёРєРё, CTR, Р·Р°РєР°Р·С‹, СЂР°СЃС…РѕРґ, CPC, CPO.\n\n"
            "РџСЂРёРјРµСЂ: СѓРєР°Р¶РёС‚Рµ РїРѕСЃР»РµРґРЅРёРµ 7 РґРЅРµР№ Рё СЃСЂР°РІРЅРёС‚Рµ CPO РјРµР¶РґСѓ РєР°РјРїР°РЅРёСЏРјРё, С‡С‚РѕР±С‹ РЅР°Р№С‚Рё РЅРµСЌС„С„РµРєС‚РёРІРЅС‹Рµ."
        ),
    },
    "wb_ads_recommendations": {
        "title": "Р РµРєРѕРјРµРЅРґР°С†РёРё WB Ads",
        "content": (
            "РќР°Р·РЅР°С‡РµРЅРёРµ: Р°РІС‚РѕРјР°С‚РёС‡РµСЃРєРёРµ СЂРµРєРѕРјРµРЅРґР°С†РёРё РїРѕ РѕРїС‚РёРјРёР·Р°С†РёРё WB Ads.\n\n"
            "РџР°СЂР°РјРµС‚СЂС‹:\n"
            "- Р”Р°С‚Р° СЃ / РїРѕ: РїРµСЂРёРѕРґ Р°РЅР°Р»РёР·Р°.\n"
            "- РњРёРЅ. СЂР°СЃС…РѕРґ: РїРѕСЂРѕРі РІРєР»СЋС‡РµРЅРёСЏ РєР°РјРїР°РЅРёРё РІ Р°РЅР°Р»РёР·.\n"
            "- РџРѕСЃС‚СЂРѕРёС‚СЊ СЂРµРєРѕРјРµРЅРґР°С†РёРё: СЂР°СЃСЃС‡РёС‚С‹РІР°РµС‚ РїСЂРёРѕСЂРёС‚РµС‚ Рё РґРµР№СЃС‚РІРёРµ.\n"
            "- РљРѕР»РѕРЅРєРё СЂРµР·СѓР»СЊС‚Р°С‚Р°: priority, recommendation, reason, action.\n\n"
            "РџСЂРёРјРµСЂ: Р·Р°РґР°Р№С‚Рµ В«РњРёРЅ. СЂР°СЃС…РѕРґ = 500В», РЅР°Р№РґРёС‚Рµ high-priority РєР°РјРїР°РЅРёРё Рё РЅР°С‡РЅРёС‚Рµ СЃ РґРµР№СЃС‚РІРёР№ pause/refresh."
        ),
    },
    "ads_bidder": {
        "title": "Р‘РёРґРµСЂ WB Ads",
        "content": (
            "РќР°Р·РЅР°С‡РµРЅРёРµ: Р°РІС‚РѕРјР°С‚РёС‡РµСЃРєРѕРµ СѓРїСЂР°РІР»РµРЅРёРµ СЃС‚Р°РІРєР°РјРё РїРѕ РїСЂР°РІРёР»Р°Рј РІ WB Ads.\n\n"
            "Р“РґРµ РѕС‚РєСЂС‹С‚СЊ:\n"
            "- РњРѕРґСѓР»СЊ В«Р РµРєР»Р°РјР° WB/OzonВ» в†’ РІРєР»Р°РґРєР° В«Р‘РёРґРµСЂВ».\n\n"
            "РљР°Рє СЃРѕР·РґР°С‚СЊ РїСЂР°РІРёР»Рѕ:\n"
            "1) РЈРєР°Р¶РёС‚Рµ campaign_id Рё nm_id.\n"
            "2) Р’С‹Р±РµСЂРёС‚Рµ target_type: normquery РёР»Рё nm.\n"
            "3) Р”Р»СЏ normquery Р·Р°РїРѕР»РЅРёС‚Рµ target_value (С„СЂР°Р·Сѓ).\n"
            "4) Р’С‹Р±РµСЂРёС‚Рµ СЃС‚СЂР°С‚РµРіРёСЋ (optimal/position/range/hold).\n"
            "5) Р—Р°РґР°Р№С‚Рµ min_bid, max_bid, step Рё cooldown.\n"
            "6) РќР°Р¶РјРёС‚Рµ В«РЎРѕС…СЂР°РЅРёС‚СЊ РїСЂР°РІРёР»РѕВ» Рё Р·Р°С‚РµРј В«Р—Р°РїСѓСЃС‚РёС‚СЊ СЃРµР№С‡Р°СЃВ».\n\n"
            "РџСЂРѕРІРµСЂРєР° СЂРµР·СѓР»СЊС‚Р°С‚Р°:\n"
            "- Р’ С‚Р°Р±Р»РёС†Рµ Р·Р°РїСѓСЃРєРѕРІ СЃРјРѕС‚СЂРёС‚Рµ СЃС‚Р°С‚СѓСЃ ok/skipped/error Рё РїСЂРёС‡РёРЅСѓ.\n"
            "- Р•СЃР»Рё РјРЅРѕРіРѕ skipped(cooldown), СѓРІРµР»РёС‡СЊС‚Рµ cooldown Рё СѓРјРµРЅСЊСЂРёС‚Рµ С‡Р°СЃС‚РѕС‚Сѓ СЂСѓС‡РЅС‹С… Р·Р°РїСѓСЃРєРѕРІ.\n"
            "- Р•СЃР»Рё 401/403 РёР»Рё api_failed, РїРµСЂРµРїСЂРѕРІРµСЂСЊС‚Рµ WB Ads РєР»СЋС‡ РІ РїСЂРѕС„РёР»Рµ.\n\n"
            "РџСЂРёРјРµСЂ: СЃРѕР·РґР°Р№С‚Рµ РїСЂР°РІРёР»Рѕ СЃ step=50 Рё cooldown=300, Р·Р°РїСѓСЃС‚РёС‚Рµ РІСЂСѓС‡РЅСѓСЋ Рё РїСЂРѕРІРµСЂСЊС‚Рµ РїРѕСЃР»РµРґРЅСЋСЋ СЃС‚СЂРѕРєСѓ Р»РѕРіРѕРІ."
        ),
    },
    "wb_ads_bidder": {
        "title": "Р‘РёРґРµСЂ WB Ads",
        "content": (
            "РќР°Р·РЅР°С‡РµРЅРёРµ: Р°РІС‚РѕРјР°С‚РёС‡РµСЃРєРѕРµ СѓРїСЂР°РІР»РµРЅРёРµ СЃС‚Р°РІРєР°РјРё РїРѕ РїСЂР°РІРёР»Р°Рј РІ WB Ads.\n\n"
            "Р“РґРµ РѕС‚РєСЂС‹С‚СЊ:\n"
            "- РњРѕРґСѓР»СЊ В«Р РµРєР»Р°РјР° WB/OzonВ» в†’ РІРєР»Р°РґРєР° В«Р‘РёРґРµСЂВ».\n\n"
            "РљР°Рє СЃРѕР·РґР°С‚СЊ РїСЂР°РІРёР»Рѕ:\n"
            "1) РЈРєР°Р¶РёС‚Рµ campaign_id Рё nm_id.\n"
            "2) Р’С‹Р±РµСЂРёС‚Рµ target_type: normquery РёР»Рё nm.\n"
            "3) Р”Р»СЏ normquery Р·Р°РїРѕР»РЅРёС‚Рµ target_value (С„СЂР°Р·Сѓ).\n"
            "4) Р’С‹Р±РµСЂРёС‚Рµ СЃС‚СЂР°С‚РµРіРёСЋ (optimal/position/range/hold).\n"
            "5) Р—Р°РґР°Р№С‚Рµ min_bid, max_bid, step Рё cooldown.\n"
            "6) РќР°Р¶РјРёС‚Рµ В«РЎРѕС…СЂР°РЅРёС‚СЊ РїСЂР°РІРёР»РѕВ» Рё Р·Р°С‚РµРј В«Р—Р°РїСѓСЃС‚РёС‚СЊ СЃРµР№С‡Р°СЃВ».\n\n"
            "РџСЂРѕРІРµСЂРєР° СЂРµР·СѓР»СЊС‚Р°С‚Р°:\n"
            "- Р’ С‚Р°Р±Р»РёС†Рµ Р·Р°РїСѓСЃРєРѕРІ СЃРјРѕС‚СЂРёС‚Рµ СЃС‚Р°С‚СѓСЃ ok/skipped/error Рё РїСЂРёС‡РёРЅСѓ.\n"
            "- Р•СЃР»Рё РјРЅРѕРіРѕ skipped(cooldown), СѓРІРµР»РёС‡СЊС‚Рµ cooldown Рё СѓРјРµРЅСЊСЂРёС‚Рµ С‡Р°СЃС‚РѕС‚Сѓ СЂСѓС‡РЅС‹С… Р·Р°РїСѓСЃРєРѕРІ.\n"
            "- Р•СЃР»Рё 401/403 РёР»Рё api_failed, РїРµСЂРµРїСЂРѕРІРµСЂСЊС‚Рµ WB Ads РєР»СЋС‡ РІ РїСЂРѕС„РёР»Рµ.\n\n"
            "РџСЂРёРјРµСЂ: СЃРѕР·РґР°Р№С‚Рµ РїСЂР°РІРёР»Рѕ СЃ step=50 Рё cooldown=300, Р·Р°РїСѓСЃС‚РёС‚Рµ РІСЂСѓС‡РЅСѓСЋ Рё РїСЂРѕРІРµСЂСЊС‚Рµ РїРѕСЃР»РµРґРЅСЋСЋ СЃС‚СЂРѕРєСѓ Р»РѕРіРѕРІ."
        ),
    },
    "billing": {
        "title": "Р‘РёР»Р»РёРЅРі",
        "content": (
            "РќР°Р·РЅР°С‡РµРЅРёРµ: СѓРїСЂР°РІР»РµРЅРёРµ С‚Р°СЂРёС„РѕРј Рё Р»РёРјРёС‚Р°РјРё.\n\n"
            "РљРЅРѕРїРєРё:\n"
            "- РЎРјРµРЅРёС‚СЊ С‚Р°СЂРёС„: РїРµСЂРµРІРѕРґРёС‚ Р°РєРєР°СѓРЅС‚ РЅР° РІС‹Р±СЂР°РЅРЅС‹Р№ РїР»Р°РЅ.\n"
            "- РџСЂРѕРґР»РёС‚СЊ РЅР° 30 РґРЅРµР№: РїСЂРѕРґР»РµРЅРёРµ С‚РµРєСѓС‰РµРіРѕ РїР»Р°РЅР°.\n"
            "- РћР±РЅРѕРІРёС‚СЊ: РїРѕРІС‚РѕСЂРЅР°СЏ Р·Р°РіСЂСѓР·РєР° СЃС‚Р°С‚СѓСЃР° Р±РёР»Р»РёРЅРіР°.\n"
            "- Р‘Р»РѕРєРё В«РЎС‚Р°С‚СѓСЃ Рё Р»РёРјРёС‚С‹В» Рё В«РСЃС‚РѕСЂРёСЏВ»: С‚РµРєСѓС‰РёРµ РєРІРѕС‚С‹ Рё РѕРїРµСЂР°С†РёРё.\n\n"
            "РџСЂРёРјРµСЂ: РµСЃР»Рё СѓРїРёСЂР°РµС‚РµСЃСЊ РІ Р»РёРјРёС‚ AI-РѕС‚РІРµС‚РѕРІ, РІС‹Р±РµСЂРёС‚Рµ plan pro Рё РЅР°Р¶РјРёС‚Рµ В«РЎРјРµРЅРёС‚СЊ С‚Р°СЂРёС„В»."
        ),
    },
    "accounting": {
        "title": "Р‘СѓС…РіР°Р»С‚РµСЂРёСЏ",
        "content": (
            "РќР°Р·РЅР°С‡РµРЅРёРµ: СѓС‡РµС‚ РїСЂРёР±С‹Р»Рё Рё СЂР°СЃС…РѕРґРѕРІ РїРѕ WB/Ozon РІ РѕРґРЅРѕРј РјРѕРґСѓР»Рµ.\n\n"
            "РџРѕРґРІРєР»Р°РґРєРё:\n"
            "- РћР±Р·РѕСЂ: РІС‹СЂСѓС‡РєР°, СЃРµР±РµСЃС‚РѕРёРјРѕСЃС‚СЊ, РІР°Р»РѕРІР°СЏ/РѕРїРµСЂР°С†РёРѕРЅРЅР°СЏ/С‡РёСЃС‚Р°СЏ РїСЂРёР±С‹Р»СЊ, РјР°СЂР¶Р°, РіСЂР°С„РёРє РґРёРЅР°РјРёРєРё.\n"
            "- РђРЅР°Р»РёР·: С‚РѕРІР°СЂРЅР°СЏ С‚Р°Р±Р»РёС†Р° РїРѕ SKU/Р°СЂС‚РёРєСѓР»Сѓ (СѓР±С‹С‚РѕС‡РЅС‹Рµ, Р»РёРґРµСЂС‹ РїСЂРёР±С‹Р»Рё, РІРѕР·РІСЂР°С‚С‹, СЂР°СЃС…РѕРґС‹).\n"
            "- Р Р°СЃС…РѕРґС‹: РґРѕР±Р°РІР»РµРЅРёРµ/СЂРµРґР°РєС‚РёСЂРѕРІР°РЅРёРµ/СѓРґР°Р»РµРЅРёРµ СЂРµРіСѓР»СЏСЂРЅС‹С… Рё СЂР°Р·РѕРІС‹С… СЂР°СЃС…РѕРґРѕРІ.\n"
            "- РќР°СЃС‚СЂРѕР№РєРё СЂР°СЃС‡РµС‚Р°: РќР”РЎ, РЅР°Р»РѕРі, РґРѕРїРѕР»РЅРёС‚РµР»СЊРЅС‹Рµ РєРѕСЌС„С„РёС†РёРµРЅС‚С‹ Рё С„РёРєСЃРёСЂРѕРІР°РЅРЅС‹Рµ СЂР°СЃС…РѕРґС‹.\n\n"
            "Excel РїРѕ Р·Р°РєСѓРїРѕС‡РЅС‹Рј С†РµРЅР°Рј:\n"
            "- РЎРєР°С‡Р°Р№С‚Рµ СЂР°Р±Р»РѕРЅ РІ В«РќР°СЃС‚СЂРѕР№РєР°С… СЂР°СЃС‡РµС‚Р°В».\n"
            "- Р—Р°РїРѕР»РЅРёС‚Рµ Р·Р°РєСѓРїРѕС‡РЅС‹Рµ С†РµРЅС‹ Рё Р·Р°РіСЂСѓР·РёС‚Рµ С„Р°Р№Р» РѕР±СЂР°С‚РЅРѕ.\n"
            "- РЎРµСЂРІРёСЃ РїСЂРѕРІР°Р»РёРґРёСЂСѓРµС‚ СЃС‚СЂРѕРєРё Рё РїРѕРєР°Р¶РµС‚, С‡С‚Рѕ РѕР±РЅРѕРІР»РµРЅРѕ/С‡С‚Рѕ РЅРµ СЃРѕРїРѕСЃС‚Р°РІРёР»РѕСЃСЊ.\n\n"
            "РљР°Рє СЃС‡РёС‚Р°РµС‚СЃСЏ РїСЂРёР±С‹Р»СЊ:\n"
            "- Р‘РµСЂРµРј РґР°РЅРЅС‹Рµ РїСЂРѕРґР°Р¶ Рё С„РёРЅР°РЅСЃРѕРІС‹С… РѕРїРµСЂР°С†РёР№ WB/Ozon.\n"
            "- Р—Р°РєСѓРїРѕС‡РЅС‹Рµ С†РµРЅС‹ Р±РµСЂСѓС‚СЃСЏ РёР· РѕР±С‰РµР№ С‚Р°Р±Р»РёС†С‹ С‚РѕРІР°СЂРѕРІ (Product), Р±РµР· РѕС‚РґРµР»СЊРЅРѕРіРѕ РґСѓР±Р»РёСЂСѓСЋС‰РµРіРѕ РєР°С‚Р°Р»РѕРіР°.\n"
            "- РЈС‡РёС‚С‹РІР°РµРј РєРѕРјРёСЃСЃРёРё, Р»РѕРіРёСЃС‚РёРєСѓ, С…СЂР°РЅРµРЅРёРµ, СѓРґРµСЂР¶Р°РЅРёСЏ, С€С‚СЂР°С„С‹, СЂРµРєР»Р°РјСѓ.\n"
            "- Р’С‹С‡РёС‚Р°РµРј СЃРµР±РµСЃС‚РѕРёРјРѕСЃС‚СЊ (Р·Р°РєСѓРїРѕС‡РЅР°СЏ С†РµРЅР° * РІС‹РєСѓРїР»РµРЅРЅС‹Рµ СЂС‚СѓРєРё).\n"
            "- РџСЂРёРјРµРЅСЏРµРј РїРѕР»СЊР·РѕРІР°С‚РµР»СЊСЃРєРёРµ СЂР°СЃС…РѕРґС‹, РќР”РЎ/РЅР°Р»РѕРіРѕРІС‹Рµ РїР°СЂР°РјРµС‚СЂС‹.\n"
            "- РџРѕР»СѓС‡Р°РµРј С‡РёСЃС‚СѓСЋ РїСЂРёР±С‹Р»СЊ Рё РјР°СЂР¶Сѓ.\n\n"
            "РћРіСЂР°РЅРёС‡РµРЅРёСЏ:\n"
            "- Р•СЃР»Рё РјР°СЂРєРµС‚РїР»РµР№СЃ РЅРµ РѕС‚РґР°Р» С‚РѕРІР°СЂРЅС‹Рµ РїРѕР»СЏ РїРѕ С‡Р°СЃС‚Рё РѕРїРµСЂР°С†РёР№, СЃС‚СЂРѕРєР° РїРѕРїР°РґРµС‚ РІ Р°РіСЂРµРіРёСЂРѕРІР°РЅРЅС‹Р№ РІРёРґ.\n"
            "- РўРѕС‡РЅРѕСЃС‚СЊ СЃРµР±РµСЃС‚РѕРёРјРѕСЃС‚Рё Р·Р°РІРёСЃРёС‚ РѕС‚ Р·Р°РїРѕР»РЅРµРЅРЅРѕСЃС‚Рё Р·Р°РєСѓРїРѕС‡РЅС‹С… С†РµРЅ.\n\n"
            "РџСЂРёРјРµСЂ: Р·Р°РіСЂСѓР·РёС‚Рµ Р·Р°РєСѓРїРѕС‡РЅС‹Рµ С†РµРЅС‹ С‡РµСЂРµР· Excel, РґРѕР±Р°РІСЊС‚Рµ РµР¶РµРјРµСЃСЏС‡РЅС‹Рµ СЂР°СЃС…РѕРґС‹ (Р·Р°СЂРїР»Р°С‚Р°/СѓРїР°РєРѕРІРєР°), РІС‹Р±РµСЂРёС‚Рµ В«РњРµСЃСЏС†В» Рё РѕС†РµРЅРёС‚Рµ С‡РёСЃС‚СѓСЋ РїСЂРёР±С‹Р»СЊ РїРѕ WB Рё Ozon."
        ),
    },
    "sales_stats": {
        "title": "РЎС‚Р°С‚РёСЃС‚РёРєР° РїСЂРѕРґР°Р¶",
        "content": (
            "РќР°Р·РЅР°С‡РµРЅРёРµ: Р°РЅР°Р»РёР· РїСЂРѕРґР°Р¶ РїРѕ РґР°С‚Р°Рј Рё РјР°СЂРєРµС‚РїР»РµР№СЃР°Рј.\n\n"
            "РљРЅРѕРїРєРё Рё РїРѕР»СЏ:\n"
            "- РњР°СЂРєРµС‚РїР»РµР№СЃ: all / wb / ozon.\n"
            "- Р”Р°С‚Р° СЃ / РїРѕ: РїРµСЂРёРѕРґ РѕС‚С‡РµС‚Р°.\n"
            "- Р‘С‹СЃС‚СЂС‹Рµ РїРµСЂРёРѕРґС‹: РґРµРЅСЊ, РЅРµРґРµР»СЏ, РјРµСЃСЏС†, РєРІР°СЂС‚Р°Р», 6 РјРµСЃСЏС†РµРІ, РіРѕРґ.\n"
            "- РџРµСЂРµРєР»СЋС‡Р°С‚РµР»СЊ РјРµС‚СЂРёРєРё РіСЂР°С„РёРєР°: СЂС‚СѓРєРё / РІС‹СЂСѓС‡РєР° / Р·Р°РєР°Р·С‹ / РѕС‚РєР°Р·С‹ / СЂРµРєР»Р°РјР° / С€С‚СЂР°С„С‹.\n"
            "- Р§РµРєР±РѕРєСЃС‹ РіСЂР°С„РёРєР°: РІСЃРµРіРѕ / WB / Ozon РґР»СЏ СЃСЂР°РІРЅРµРЅРёСЏ Р»РёРЅРёР№.\n"
            "- РџРѕРєР°Р·Р°С‚РµР»СЊ В«РЁС‚СЂР°С„С‹В»: СѓРґРµСЂР¶Р°РЅРёСЏ/С€С‚СЂР°С„РЅС‹Рµ СЃРїРёСЃР°РЅРёСЏ РёР· РґРѕСЃС‚СѓРїРЅС‹С… РѕС‚С‡РµС‚РѕРІ РјР°СЂРєРµС‚РїР»РµР№СЃРѕРІ.\n"
            "- Р—Р°РіСЂСѓР·РёС‚СЊ СЃС‚Р°С‚РёСЃС‚РёРєСѓ: СЂСѓС‡РЅРѕР№ РїСЂРёРЅСѓРґРёС‚РµР»СЊРЅС‹Р№ СЂРµС„СЂРµСЂ.\n\n"
            "Р’Р°Р¶РЅРѕ:\n"
            "- KPI-РєР°СЂС‚РѕС‡РєРё СЃРІРµСЂС…Сѓ РїРѕРєР°Р·С‹РІР°СЋС‚ Р°РіСЂРµРіР°С‚С‹ Р·Р° РІС‹Р±СЂР°РЅРЅС‹Р№ РїРµСЂРёРѕРґ.\n"
            "- Р”Р°РЅРЅС‹Рµ Р·Р°РіСЂСѓР¶Р°СЋС‚СЃСЏ Р°РІС‚РѕРјР°С‚РёС‡РµСЃРєРё РїСЂРё СЃРјРµРЅРµ РјР°СЂРєРµС‚РїР»РµР№СЃР° Рё РґР°С‚.\n"
            "- Р¤РѕСЂРјР°С‚ С‡РёСЃРµР» Р°РґР°РїС‚РёСЂСѓРµС‚СЃСЏ РїРѕРґ СЏР·С‹Рє РёРЅС‚РµСЂС„РµР№СЃР° (ru/en).\n"
            "- Р”Р»СЏ СЃСЂР°РІРЅРµРЅРёСЏ РєР°РЅР°Р»РѕРІ РѕСЃС‚Р°РІСЊС‚Рµ РІРєР»СЋС‡РµРЅРЅС‹РјРё СЃСЂР°Р·Сѓ WB Рё Ozon.\n\n"
            "Р”РёР°РіРЅРѕСЃС‚РёРєР°:\n"
            "- Р•СЃР»Рё РїРѕ WB РІРёРґРёС‚Рµ РїСЂРµРґСѓРїСЂРµР¶РґРµРЅРёРµ API 429, СЌС‚Рѕ Р»РёРјРёС‚ WB; РїРѕРІС‚РѕСЂРёС‚Рµ С‡СѓС‚СЊ РїРѕР·Р¶Рµ.\n"
            "- Р•СЃР»Рё С‚Р°Р±Р»РёС†Р° РїСѓСЃС‚Р°СЏ, СЃРЅР°С‡Р°Р»Р° РІС‹Р±РµСЂРёС‚Рµ В«Р”РµРЅСЊ/РќРµРґРµР»СЏВ» Рё РїСЂРѕРІРµСЂСЊС‚Рµ РєР»СЋС‡Рё API РІ В«РџСЂРѕС„РёР»РµВ».\n"
            "- Р•СЃР»Рё РіСЂР°С„РёРє РєР°Р¶РµС‚СЃСЏ РїСѓСЃС‚С‹Рј, РїРµСЂРµРєР»СЋС‡РёС‚Рµ РјРµС‚СЂРёРєСѓ Рё СѓР±РµРґРёС‚РµСЃСЊ, С‡С‚Рѕ РІРєР»СЋС‡РµРЅР° С…РѕС‚СЏ Р±С‹ РѕРґРЅР° Р»РёРЅРёСЏ (Р’СЃРµРіРѕ/WB/Ozon).\n\n"
            "РџСЂРёРјРµСЂ: РІС‹Р±РµСЂРёС‚Рµ В«РљРІР°СЂС‚Р°Р»В», РјРµС‚СЂРёРєСѓ В«Р’С‹СЂСѓС‡РєР°В», РІРєР»СЋС‡РёС‚Рµ WB+Ozon Рё СЃСЂР°РІРЅРёС‚Рµ РґРёРЅР°РјРёРєСѓ РєР°РЅР°Р»РѕРІ. Р—Р°С‚РµРј РїРµСЂРµРєР»СЋС‡РёС‚РµСЃСЊ РЅР° В«Р РµРєР»Р°РјР°В», С‡С‚РѕР±С‹ СЃРѕРїРѕСЃС‚Р°РІРёС‚СЊ СЂРѕСЃС‚ РІС‹СЂСѓС‡РєРё Рё Р·Р°С‚СЂР°С‚."
        ),
    },
    "user_profile": {
        "title": "РџСЂРѕС„РёР»СЊ",
        "content": (
            "РќР°Р·РЅР°С‡РµРЅРёРµ: СѓРїСЂР°РІР»РµРЅРёРµ Р»РёС‡РЅС‹РјРё Рё СЋСЂРёРґРёС‡РµСЃРєРёРјРё РґР°РЅРЅС‹РјРё, С‚Р°СЂРёС„РѕРј Рё РєР»СЋС‡Р°РјРё API.\n\n"
            "Р§С‚Рѕ РґРѕСЃС‚СѓРїРЅРѕ:\n"
            "- РџРѕР»СЏ РїСЂРѕС„РёР»СЏ: Р¤РРћ, РєРѕРјРїР°РЅРёСЏ, РіРѕСЂРѕРґ, РРќРќ, РЅР°Р»РѕРіРѕРІР°СЏ СЃС‚Р°РІРєР°, С‚РµР»РµС„РѕРЅ, СЃС‚СЂСѓРєС‚СѓСЂР° РєРѕРјР°РЅРґС‹.\n"
            "- РўР°СЂРёС„: РІС‹Р±РѕСЂ Рё РїСЂРѕРґР»РµРЅРёРµ РІ СЌС‚РѕРј Р¶Рµ РјРѕРґСѓР»Рµ.\n"
            "- API РєР»СЋС‡Рё WB/Ozon: РґРѕР±Р°РІР»РµРЅРёРµ, РїСЂРѕРІРµСЂРєР°, СѓРґР°Р»РµРЅРёРµ.\n"
            "- Р‘РµР·РѕРїР°СЃРЅРѕСЃС‚СЊ: СЃРјРµРЅР° РїР°СЂРѕР»СЏ.\n\n"
            "РџСЂР°РєС‚РёРєР°:\n"
            "1) РЎРЅР°С‡Р°Р»Р° Р·Р°РїРѕР»РЅРёС‚Рµ СЂРµРєРІРёР·РёС‚С‹ Рё РєРѕРЅС‚Р°РєС‚РЅС‹Рµ РїРѕР»СЏ.\n"
            "2) Р—Р°С‚РµРј РїРѕРґРєР»СЋС‡РёС‚Рµ WB/Ozon РєР»СЋС‡Рё Рё РЅР°Р¶РјРёС‚Рµ В«РџСЂРѕРІРµСЂРёС‚СЊВ».\n"
            "3) РџРѕСЃР»Рµ СѓСЃРїРµСЂРЅРѕР№ РїСЂРѕРІРµСЂРєРё РѕР±РЅРѕРІРёС‚Рµ РјРѕРґСѓР»СЊ В«РўРѕРІР°СЂС‹В» Рё Р·Р°РїСѓСЃС‚РёС‚Рµ РёРјРїРѕСЂС‚.\n"
            "4) Р•СЃР»Рё Р»РёРјРёС‚РѕРІ РјР°Р»Рѕ, СЃРјРµРЅРёС‚Рµ С‚Р°СЂРёС„ Рё РѕР±РЅРѕРІРёС‚Рµ СЌРєСЂР°РЅ.\n\n"
            "Р”РёР°РіРЅРѕСЃС‚РёРєР°:\n"
            "- РћС€РёР±РєР° РєР»СЋС‡Р° Ozon РѕР±С‹С‡РЅРѕ СЃРІСЏР·Р°РЅР° СЃ С„РѕСЂРјР°С‚РѕРј: РЅСѓР¶РµРЅ `client_id:api_key`.\n"
            "- Р•СЃР»Рё РјРѕРґСѓР»СЊ РЅРµ РѕС‚РєСЂС‹РІР°РµС‚СЃСЏ, РїСЂРѕРІРµСЂСЊС‚Рµ РґРѕСЃС‚СѓРїС‹ РІ Р°РґРјРёРЅРєРµ (СЂР°Р·РґРµР» В«РњРѕРґСѓР»РёВ»).\n"
            "- Р•СЃР»Рё РїР°СЂРѕР»СЊ РЅРµ РјРµРЅСЏРµС‚СЃСЏ, СѓР±РµРґРёС‚РµСЃСЊ, С‡С‚Рѕ РЅРѕРІС‹Р№ РїР°СЂРѕР»СЊ РЅРµ РєРѕСЂРѕС‡Рµ 8 СЃРёРјРІРѕР»РѕРІ.\n\n"
            "РџСЂРёРјРµСЂ: РѕР±РЅРѕРІРёС‚Рµ СЂРµРєРІРёР·РёС‚С‹ РєРѕРјРїР°РЅРёРё, СЃРјРµРЅРёС‚Рµ С‚Р°СЂРёС„ РЅР° pro Рё РїСЂРѕРІРµСЂСЊС‚Рµ РєР»СЋС‡Рё РёРЅС‚РµРіСЂР°С†РёРё РІ РѕРґРЅРѕРј РјРµСЃС‚Рµ."
        ),
    },
    "help_center": {
        "title": "РЎРїСЂР°РІРєР°",
        "content": (
            "РќР°Р·РЅР°С‡РµРЅРёРµ: С†РµРЅС‚СЂР°Р»РёР·РѕРІР°РЅРЅР°СЏ РёРЅСЃС‚СЂСѓРєС†РёСЏ РїРѕ РІСЃРµРј РјРѕРґСѓР»СЏРј.\n\n"
            "РљР°Рє РїРѕР»СЊР·РѕРІР°С‚СЊСЃСЏ:\n"
            "- Р’С‹Р±РµСЂРёС‚Рµ РјРѕРґСѓР»СЊ РІ РїРµСЂРІРѕРј СЃРїРёСЃРєРµ.\n"
            "- РќР°Р¶РјРёС‚Рµ В«РћР±РЅРѕРІРёС‚СЊ СЃРїСЂР°РІРєСѓВ» РґР»СЏ РїРµСЂРµР·Р°РіСЂСѓР·РєРё СЃРѕРґРµСЂР¶РёРјРѕРіРѕ.\n\n"
            "- РџРѕРґРІРєР»Р°РґРєР° В«Р—Р°РіСЂСѓР·РєРёВ»: С‚РµРєСѓС‰Р°СЏ РІРµСЂСЃРёСЏ, РѕС‚Р»РёС‡РёСЏ РѕС‚ РїСЂРµРґС‹РґСѓС‰РµР№, РёСЃС‚РѕСЂРёСЏ СЂРµР»РёР·РѕРІ Рё СЃСЃС‹Р»РєР° РЅР° РјРѕР±РёР»СЊРЅС‹Р№ РєР»РёРµРЅС‚.\n\n"
            "Р§С‚Рѕ РµСЃС‚СЊ РІ РєР°Р¶РґРѕРј СЂР°Р·РґРµР»Рµ:\n"
            "- РќР°Р·РЅР°С‡РµРЅРёРµ РјРѕРґСѓР»СЏ Рё РѕР¶РёРґР°РµРјС‹Р№ СЂРµР·СѓР»СЊС‚Р°С‚.\n"
            "- Р Р°СЃСЂРёС„СЂРѕРІРєР° РєРЅРѕРїРѕРє, РїРѕР»РµР№ Рё РїРµСЂРµРєР»СЋС‡Р°С‚РµР»РµР№.\n"
            "- РўРёРїРѕРІРѕР№ СЂР°Р±РѕС‡РёР№ СЃС†РµРЅР°СЂРёР№ РїРѕ СЂР°РіР°Рј.\n"
            "- РџСЂРёРјРµСЂ РїСЂР°РєС‚РёС‡РµСЃРєРѕРіРѕ РёСЃРїРѕР»СЊР·РѕРІР°РЅРёСЏ.\n\n"
            "Р”РѕРїРѕР»РЅРёС‚РµР»СЊРЅРѕ:\n"
            "- РљРЅРѕРїРєРё РјРѕРґСѓР»СЏ РїРѕРґСЃРІРµС‡РёРІР°СЋС‚ РІС‹Р±СЂР°РЅРЅС‹Р№ СЂР°Р·РґРµР», РЅРµ СЃРєСЂС‹РІР°СЏ РѕСЃС‚Р°Р»СЊРЅС‹Рµ РєР°СЂС‚РѕС‡РєРё.\n"
            "- Р§РµРє-Р»РёСЃС‚ РІ РєР°СЂС‚РѕС‡РєРµ РїРѕРјРѕРіР°РµС‚ Р±С‹СЃС‚СЂРѕ РїСЂРѕРІРµСЂРёС‚СЊ, С‡С‚Рѕ РІС‹ РЅРµ РїСЂРѕРїСѓСЃС‚РёР»Рё РѕР±СЏР·Р°С‚РµР»СЊРЅС‹Рµ СЂР°РіРё.\n\n"
            "РљР°Рє РёСЃРїРѕР»СЊР·РѕРІР°С‚СЊ СЃРїСЂР°РІРєСѓ РІ СЂР°Р±РѕС‚Рµ:\n"
            "1) РћС‚С„РёР»СЊС‚СЂСѓР№С‚Рµ РЅСѓР¶РЅС‹Р№ РјРѕРґСѓР»СЊ РІ РєР°СЂС‚РѕС‡РєРµ.\n"
            "2) Р’С‹РїРѕР»РЅРёС‚Рµ РґРµР№СЃС‚РІРёСЏ РёР· Р±Р»РѕРєР° В«РўРёРїРѕРІРѕР№ СЃС†РµРЅР°СЂРёР№В».\n"
            "3) РџСЂРё РѕСЂРёР±РєРµ СЃРІРµСЂСЏР№С‚РµСЃСЊ СЃ Р±Р»РѕРєРѕРј В«Р”РёР°РіРЅРѕСЃС‚РёРєР°В» РІ СЃРѕРѕС‚РІРµС‚СЃС‚РІСѓСЋС‰РµРј СЂР°Р·РґРµР»Рµ.\n"
            "4) Р’РѕР·РІСЂР°С‰Р°Р№С‚РµСЃСЊ РІ СЃРїСЂР°РІРєСѓ Рё С„РёРєСЃРёСЂСѓР№С‚Рµ СЂР°Р±РѕС‡РёРµ СЃС†РµРЅР°СЂРёРё РєРѕРјР°РЅРґС‹.\n\n"
            "РЎРѕРІРµС‚: С„РёР»СЊС‚СЂСѓР№С‚Рµ СЃРїСЂР°РІРєСѓ РїРѕ РѕРґРЅРѕРјСѓ РјРѕРґСѓР»СЋ, С‡С‚РѕР±С‹ РєРѕРјР°РЅРґР° СЂР°Р±РѕС‚Р°Р»Р° РїРѕ РµРґРёРЅРѕРјСѓ С‡РµРє-Р»РёСЃС‚Сѓ."
        ),
    },
    "ai_assistant": {
        "title": "AI РїРѕРјРѕС‰РЅРёРє",
        "content": (
            "РќР°Р·РЅР°С‡РµРЅРёРµ: Р±С‹СЃС‚СЂС‹Рµ РѕС‚РІРµС‚С‹ РїРѕ WB/Ozon, РјР°СЂРєРµС‚РїР»РµР№СЃР°Рј Рё СЂР°Р±РѕС‚Рµ РІРЅСѓС‚СЂРё SEO WIBE.\n\n"
            "РљР°Рє РёСЃРїРѕР»СЊР·РѕРІР°С‚СЊ:\n"
            "- РћС‚РєСЂРѕР№С‚Рµ РІРєР»Р°РґРєСѓ В«РЎРїСЂР°РІРєР°В» -> В«AI РїРѕРјРѕС‰РЅРёРєВ».\n"
            "- Р—Р°РґР°Р№С‚Рµ РІРѕРїСЂРѕСЃ РїСЂРѕСЃС‚С‹Рј СЏР·С‹РєРѕРј.\n"
            "- РЈС‚РѕС‡РЅРёС‚Рµ РјР°СЂРєРµС‚РїР»РµР№СЃ, РїРµСЂРёРѕРґ Рё РјРѕРґСѓР»СЊ, РµСЃР»Рё РЅСѓР¶РµРЅ С‚РѕС‡РЅС‹Р№ РѕС‚РІРµС‚.\n\n"
            "Р’Р°Р¶РЅРѕ:\n"
            "- РћС‚РІРµС‚С‹ СЃС‚СЂРѕСЏС‚СЃСЏ РЅР° РІС‹Р±СЂР°РЅРЅРѕРј AI-РїСЂРѕРІР°Р№РґРµСЂРµ (РїСЂРѕС„РёР»СЊ/Р°РґРјРёРЅРєР°).\n"
            "- РџСЂРё СЃРјРµРЅРµ РїСЂРѕРІР°Р№РґРµСЂР° РѕС‚РІРµС‚ РјРѕР¶РµС‚ РѕС‚Р»РёС‡Р°С‚СЊСЃСЏ РїРѕ СЃС‚РёР»СЋ Рё РґРµС‚Р°Р»РёР·Р°С†РёРё.\n"
            "- Р”Р»СЏ Р·Р°РґР°С‡ СЃ API/РјРµС‚СЂРёРєР°РјРё РїСЂРѕРІРµСЂСЏР№С‚Рµ РєР»СЋС‡Рё Рё С„РёР»СЊС‚СЂС‹ РїРµСЂРёРѕРґР°.\n\n"
            "РџСЂРёРјРµСЂ: В«РџРѕС‡РµРјСѓ WB СЃС‚Р°С‚РёСЃС‚РёРєР° Р·Р° РјРµСЃСЏС† РіСЂСѓР·РёС‚СЃСЏ С‡Р°СЃС‚РёС‡РЅРѕ Рё С‡С‚Рѕ РїСЂРѕРІРµСЂРёС‚СЊ РІ РїРµСЂРІСѓСЋ РѕС‡РµСЂРµРґСЊ?В»"
        ),
    },
}

HELP_DOCS_EN: dict[str, dict[str, str]] = {
    "dashboard": {
        "title": "Dashboard",
        "content": (
            "Purpose: daily control center for your workspace.\n\n"
            "Sections and actions:\n"
            "1) KPI cards: products, SEO jobs, in-progress jobs, top-5 products.\n"
            "2) 21-day trend: average ranking trend and checks density.\n"
            "3) Quick actions:\n"
            "- Import products\n"
            "- Open SEO jobs\n"
            "- Check all rankings\n\n"
            "Example: open dashboard every morning, verify top-5 trend, then run a full ranking check if metrics dropped."
        ),
    },
    "products": {
        "title": "Products",
        "content": (
            "Purpose: catalog import, filtering, pagination, and product card management.\n\n"
            "Buttons and fields:\n"
            "- Import: imports products from selected marketplace.\n"
            "- Reload Catalog: fully rebuilds local catalog.\n"
            "- Select All: selects all rows in current table view.\n"
            "- Page navigation: top and bottom paging controls.\n"
            "- Page size: 30/50/100/200/500/1000 rows.\n"
            "- Filter input: quick search by article/title.\n"
            "- Right-side details panel: photos, attributes, card editing.\n\n"
            "Price fields in product card:\n"
            "- Purchase price: unit cost (used by accounting for COGS/margin).\n"
            "- Base price: list/old price before discounts.\n"
            "- Discounted price: current sell price.\n"
            "- Min price: lower bound allowed by your pricing strategy.\n"
            "- Marketing price: promo-focused price for campaign workflows.\n\n"
            "Photos in editor:\n"
            "- Add URL: attach image by direct link.\n"
            "- Add photo: opens local folders and supports multi-file selection.\n"
            "- Drag and drop changes order; first image is main photo.\n\n"
            "Recommended workflow:\n"
            "1) Import catalog.\n"
            "2) Filter priority SKU group.\n"
            "3) Set page size and browse catalog pages.\n"
            "4) Open and update card details when needed.\n"
            "5) Run ranking checks in the SEO Jobs submodule.\n\n"
            "Troubleshooting:\n"
            "- If WB/Ozon filter looks wrong, verify selected marketplace in top selector.\n"
            "- If table looks empty after import, clear filters and reload data.\n"
            "- After adding/reordering photos, press Save to persist the new order.\n\n"
            "Example: import Ozon catalog, set 100 rows per page, filter \"chimney\", open card details, and update description."
        ),
    },
    "seo_generation": {
        "title": "SEO Jobs",
        "content": (
            "Purpose: generate, validate, and apply SEO descriptions.\n\n"
            "Actions:\n"
            "- Generate Selected / Generate All.\n"
            "- Apply: publish approved descriptions.\n"
            "- Recheck Selected / Recheck Due.\n"
            "- Select All.\n"
            "- Delete Selected / Delete All SEO Jobs.\n\n"
            "Job lifecycle:\n"
            "1) generated: draft text is ready.\n"
            "2) in_progress: job is being re-evaluated.\n"
            "3) applied: marketplace update is completed.\n"
            "4) top_5: target rank reached.\n\n"
            "Best practices:\n"
            "- Run Recheck Selected after manual edits.\n"
            "- Use Recheck Due daily for scheduled maintenance.\n"
            "- Review preview and competitor context before Apply.\n\n"
            "Troubleshooting:\n"
            "- If Apply fails, verify API permissions and marketplace availability.\n"
            "- If status does not change, trigger recheck and inspect next-check timestamp.\n\n"
            "Example: generate tasks for 30 products, review preview text, apply approved jobs, then recheck to measure ranking delta."
        ),
    },
    "wb_reviews_ai": {
        "title": "WB/Ozon Reviews",
        "content": (
            "Purpose: process marketplace reviews and publish replies.\n\n"
            "Controls:\n"
            "- Refresh Reviews: reload reviews from WB/Ozon.\n"
            "- Save AI settings: stores mode and prompt.\n"
            "- AI icon in row: generate draft reply.\n"
            "- Send icon in row: publish or update reply.\n"
            "- Filters: marketplace, rating, status, sort, date range.\n\n"
            "Important:\n"
            "- Date range filters already loaded rows. If the table is empty, clear date filters first.\n"
            "- Status bar shows progressive loading (fast layer first, full layer next).\n"
            "- Reply textarea is a draft until you explicitly press Send/Update.\n\n"
            "Typical workflow:\n"
            "1) Select marketplace.\n"
            "2) Filter by Unanswered.\n"
            "3) Generate draft with AI icon and adjust tone.\n"
            "4) Publish and verify status.\n\n"
            "Example: filter Unanswered reviews, generate draft, edit tone, then publish."
        ),
    },
    "wb_questions_ai": {
        "title": "WB/Ozon Questions",
        "content": (
            "Purpose: answer customer questions for product cards.\n\n"
            "Controls:\n"
            "- Refresh Questions.\n"
            "- Save AI settings.\n"
            "- Upload to Knowledge Base / Delete Selected Document.\n"
            "- AI icon in row: generate answer.\n"
            "- Send icon in row: publish/update answer.\n\n"
            "Important:\n"
            "- Upload FAQ/guidelines to knowledge base before bulk answering.\n"
            "- If rows are missing, verify marketplace and date filters.\n"
            "- Use status bar and RAW block to diagnose API-side issues.\n\n"
            "Typical workflow:\n"
            "1) Choose marketplace and New status.\n"
            "2) Upload a reference document if needed.\n"
            "3) Generate and edit answer.\n"
            "4) Publish and refresh list.\n\n"
            "Example: upload supplier FAQ into knowledge base, then generate consistent answers for new questions."
        ),
    },
    "wb_ads": {
        "title": "WB Ads",
        "content": (
            "Purpose: monitor and manage WB ad campaigns.\n\n"
            "Main controls:\n"
            "- Load Campaigns: fast campaign list load.\n"
            "- Get Rates: fetch bid rates by campaign_id.\n"
            "- Reset Filters.\n"
            "- Filters: search, status, type, running flag, budget range.\n"
            "- Double-click row: open campaign details modal.\n"
            "- In modal: Start / Pause / Stop / Refresh details.\n\n"
            "Bidder subtab:\n"
            "- Open Ads -> Bidder.\n"
            "- Set campaign_id and nm_id.\n"
            "- Target type:\n"
            "  normquery: phrase target (fill target_value),\n"
            "  nm: card target (target_value can be empty).\n"
            "- Strategy:\n"
            "  optimal, position, range, hold.\n"
            "- Set min/max/step and cooldown (seconds between rule runs).\n"
            "- Save rule.\n"
            "- Manual run: Run now.\n"
            "- Background mode: keep rule active and worker will run it by cooldown.\n\n"
            "Quick start (3 minutes):\n"
            "1) In Campaigns, choose a running campaign and copy campaign_id.\n"
            "2) In Bidder, fill campaign_id + nm_id and pick normquery or nm target.\n"
            "3) For normquery, target_value phrase is required.\n"
            "4) Safe starting profile: step=50..100, cooldown=300, conservative min/max bid range.\n"
            "5) Save rule, run once manually, then check one fresh log row.\n"
            "6) If no errors, keep rule active for background runs.\n\n"
            "Bidder logs:\n"
            "- ok: bid applied/confirmed.\n"
            "- skipped: run was skipped by conditions (cooldown/no data/limits).\n"
            "- error: API/configuration issue.\n"
            "- Reason column explains each run result.\n\n"
            "If a rule does not work:\n"
            "- Verify WB Ads key is valid and module access is enabled for the user.\n"
            "- Ensure campaign_id and nm_id belong to the same campaign/account.\n"
            "- For normquery, use exact phrase without trailing spaces.\n"
            "- If you often get skipped(cooldown), increase cooldown and avoid frequent manual reruns.\n"
            "- For 401/403 errors, reconnect WB Ads key in Profile.\n\n"
            "Tuning tips:\n"
            "- Start with conservative step (50-100).\n"
            "- Use cooldown >= 180-300 sec for stable campaigns.\n"
            "- Validate with manual run first, then enable background mode.\n\n"
            "Example: filter running campaigns, sort by spend, inspect a campaign in detail modal."
        ),
    },
    "wb_ads_analytics": {
        "title": "WB Ads Analytics",
        "content": (
            "Purpose: period analytics report for WB Ads.\n\n"
            "Inputs and output:\n"
            "- Date from / date to.\n"
            "- Optional campaign_id.\n"
            "- Build Report: produces table + totals.\n"
            "- Metrics: views, clicks, CTR, orders, spend, CPC, CPO.\n\n"
            "Example: run a 7-day report and compare CPO between campaigns before scaling budget."
        ),
    },
    "wb_ads_recommendations": {
        "title": "WB Ads Recommendations",
        "content": (
            "Purpose: prioritize optimization actions automatically.\n\n"
            "Parameters:\n"
            "- Date range.\n"
            "- Minimum spend threshold.\n"
            "- Build Recommendations: calculates priority/action/reason.\n"
            "- Output columns: priority, recommendation, reason, action.\n\n"
            "Example: set min spend to 500 and start from high-priority campaigns with pause/refresh actions."
        ),
    },
    "ads_bidder": {
        "title": "WB Ads Bidder",
        "content": (
            "Purpose: automate bid control for WB Ads campaigns using rules.\n\n"
            "Where to open:\n"
            "- Ads module -> Bidder subtab.\n\n"
            "Create a rule:\n"
            "1) Fill campaign_id and nm_id.\n"
            "2) Choose target_type: normquery or nm.\n"
            "3) For normquery, provide target_value phrase.\n"
            "4) Choose strategy (optimal/position/range/hold).\n"
            "5) Set min_bid, max_bid, step, cooldown.\n"
            "6) Save rule and run once manually.\n\n"
            "Validate runs:\n"
            "- Check run log status: ok / skipped / error and reason text.\n"
            "- If skipped(cooldown) appears too often, increase cooldown.\n"
            "- For 401/403 or api_failed, reconnect WB Ads API key in Profile.\n\n"
            "Example: create a conservative rule (step=50, cooldown=300), run now, verify the latest log row."
        ),
    },
    "wb_ads_bidder": {
        "title": "WB Ads Bidder",
        "content": (
            "Purpose: automate bid control for WB Ads campaigns using rules.\n\n"
            "Where to open:\n"
            "- Ads module -> Bidder subtab.\n\n"
            "Create a rule:\n"
            "1) Fill campaign_id and nm_id.\n"
            "2) Choose target_type: normquery or nm.\n"
            "3) For normquery, provide target_value phrase.\n"
            "4) Choose strategy (optimal/position/range/hold).\n"
            "5) Set min_bid, max_bid, step, cooldown.\n"
            "6) Save rule and run once manually.\n\n"
            "Validate runs:\n"
            "- Check run log status: ok / skipped / error and reason text.\n"
            "- If skipped(cooldown) appears too often, increase cooldown.\n"
            "- For 401/403 or api_failed, reconnect WB Ads API key in Profile.\n\n"
            "Example: create a conservative rule (step=50, cooldown=300), run now, verify the latest log row."
        ),
    },
    "billing": {
        "title": "Billing",
        "content": (
            "Purpose: plan and limit management.\n\n"
            "Buttons:\n"
            "- Change Plan.\n"
            "- Renew for 30 days.\n"
            "- Refresh.\n"
            "- Status/Limits and History blocks show current quota usage and billing events.\n\n"
            "Example: if AI limits are reached, switch to a higher plan, then refresh to verify new limits."
        ),
    },
    "accounting": {
        "title": "Accounting",
        "content": (
            "Purpose: marketplace profit accounting for WB/Ozon in one module.\n\n"
            "Subtabs:\n"
            "- Overview: revenue, COGS, gross/operating/net profit, margin, trend chart.\n"
            "- Analysis: SKU/article table (loss makers, top profit rows, return-heavy rows, expense-heavy rows).\n"
            "- Expenses: create/edit/delete recurring and one-time expenses.\n"
            "- Calculation settings: VAT/tax/additional coefficients and fixed monthly cost.\n\n"
            "Purchase price Excel flow:\n"
            "- Download template in settings.\n"
            "- Fill purchase prices and upload back.\n"
            "- Validation reports updated/skipped/unmatched rows.\n\n"
            "Profit model:\n"
            "- Uses WB/Ozon sales + finance operations.\n"
            "- Purchase prices are mapped from the shared Products table (no duplicate catalog for accounting).\n"
            "- Includes commissions, logistics, storage, deductions, penalties, ad spend.\n"
            "- Subtracts COGS (purchase price * sold units).\n"
            "- Applies custom expenses and tax/VAT parameters.\n"
            "- Produces net profit and margin.\n\n"
            "Limitations:\n"
            "- If marketplace item dimensions are missing for some operations, those records are shown in aggregated form.\n"
            "- COGS quality depends on purchase-price coverage.\n\n"
            "Example: import purchase prices from Excel, add monthly salary/packaging expenses, switch to Month range, and compare net profit for WB vs Ozon."
        ),
    },
    "sales_stats": {
        "title": "Sales Statistics",
        "content": (
            "Purpose: track sales by date range and marketplace.\n\n"
            "Controls:\n"
            "- Marketplace: all / wb / ozon.\n"
            "- Date from / date to.\n"
            "- Quick ranges: day, week, month, quarter, 6 months, year.\n"
            "- Chart metric switch: units / revenue / orders / returns / ad spend / penalties.\n"
            "- Chart series toggles: total / WB / Ozon for side-by-side comparison.\n"
            "- Penalties metric: marketplace deductions/penalty charges when available.\n"
            "- Load stats: manual forced refresh.\n\n"
            "Important:\n"
            "- KPI cards show total orders, units, and revenue for the selected range.\n"
            "- Data also auto-refreshes when marketplace or dates are changed.\n"
            "- Number formatting follows selected UI language.\n"
            "- Keep WB and Ozon enabled to compare trends visually.\n\n"
            "Troubleshooting:\n"
            "- WB warning 429 means API rate limit, retry later.\n"
            "- Empty table usually means no sales in range or invalid API keys.\n"
            "- If chart looks empty, switch metric and verify at least one line is enabled.\n\n"
            "Example: choose Quarter, switch to Revenue, enable WB+Ozon, compare line trends, then switch to Ads Spend to compare revenue growth against ad costs."
        ),
    },
    "user_profile": {
        "title": "Profile",
        "content": (
            "Purpose: manage personal/company data, plan, API keys, and password.\n\n"
            "Includes:\n"
            "- Profile fields: full name, company, city, legal details, tax rate, phone, team structure.\n"
            "- Plan management in the same module.\n"
            "- WB/Ozon API keys management.\n"
            "- Password change.\n\n"
            "Recommended flow:\n"
            "1) Fill profile and legal fields.\n"
            "2) Connect and validate WB/Ozon keys.\n"
            "3) Re-open Products and run import.\n"
            "4) Upgrade plan if limits are insufficient.\n\n"
            "Troubleshooting:\n"
            "- Ozon key must be in `client_id:api_key` format.\n"
            "- If a module is unavailable, check admin module access toggles.\n"
            "- New password must contain at least 8 characters.\n\n"
            "Example: update legal details, switch plan, and validate marketplace API keys from a single screen."
        ),
    },
    "help_center": {
        "title": "Help Center",
        "content": (
            "Purpose: centralized documentation for every module.\n\n"
            "Usage:\n"
            "- Select module in the first dropdown.\n"
            "- Click Refresh Help.\n\n"
            "- Downloads subtab: Android APK current version, differences from previous APK release, and APK history.\n\n"
            "Each section includes:\n"
            "- Module purpose and expected result.\n"
            "- Button/field reference.\n"
            "- Typical workflow.\n"
            "- Practical example.\n\n"
            "Additional tools:\n"
            "- Module buttons highlight the selected card while keeping all modules visible.\n"
            "- Built-in checklist helps verify required steps.\n\n"
            "How teams use it:\n"
            "1) Filter to the needed module in help card.\n"
            "2) Execute workflow section step by step.\n"
            "3) Use troubleshooting notes when API/UI behavior is unexpected.\n"
            "4) Keep internal team SOPs aligned with these cards.\n\n"
            "Tip: keep documentation filtered to one module during onboarding."
        ),
    },
    "ai_assistant": {
        "title": "AI Assistant",
        "content": (
            "Purpose: quick assistant for WB/Ozon, marketplace operations, and SEO WIBE usage.\n\n"
            "How to use:\n"
            "- Open Help -> AI Assistant submodule.\n"
            "- Ask your question in plain language.\n"
            "- Mention marketplace, period, and module for better precision.\n\n"
            "Important:\n"
            "- Replies depend on currently selected AI provider (profile/admin).\n"
            "- Different providers may produce different style/detail.\n"
            "- For API/metrics topics, always verify keys and date filters.\n\n"
            "Example: \"Why does monthly WB stats load partially and what should I check first?\""
        ),
    },
}

HELP_RELEASES: list[dict[str, Any]] = [
    {
        "version": "0.4.9",
        "android_version_code": 21,
        "released_at": "2026-03-12",
        "current": True,
        "summary": "APK 1.5.15: С‡Р°С‚ РІ APK СЃС‚Р°Р» Р±Р»РёР¶Рµ Рє Telegram РїРѕ С€Р°РїРєРµ Рё РїРѕР»СЋ РІРІРѕРґР°, СѓСЃС‚СЂР°РЅРµРЅС‹ РєСЂРѕРєРѕР·СЏР±СЂС‹ Рё Р·Р°С„РёРєСЃРёСЂРѕРІР°РЅ РєРѕСЂСЂРµРєС‚РЅС‹Р№ Р·Р°РіРѕР»РѕРІРѕРє Р°РєС‚РёРІРЅРѕРіРѕ РјРѕРґСѓР»СЏ.",
        "diff_from_previous": [
            "Android APK РѕР±РЅРѕРІР»РµРЅ РґРѕ versionCode=21 / versionName=1.5.15.",
            "РљР°Р»РµРЅРґР°СЂСЊ: СЃРѕС…СЂР°РЅРµРЅ РїРµСЂСЃРѕРЅР°Р»СЊРЅС‹Р№ Google OAuth РїРѕ actor_key РґР»СЏ РєР°Р¶РґРѕРіРѕ СЃРѕС‚СЂСѓРґРЅРёРєР°, Р° APK-РїСЂРёР»РѕР¶РµРЅРёРµ РІРѕР·РІСЂР°С‰Р°РµС‚СЃСЏ РІ РєР°Р»РµРЅРґР°СЂСЊ С‡РµСЂРµР· deep link РїРѕСЃР»Рµ Р°РІС‚РѕСЂРёР·Р°С†РёРё.",
            "РљР°Р»РµРЅРґР°СЂСЊ: РїСѓР±Р»РёС‡РЅС‹Рµ URL РїСЂРѕРґРѕР»Р¶Р°СЋС‚ СЃС‚СЂРѕРёС‚СЊСЃСЏ РѕС‚ canonical public base, РїРѕСЌС‚РѕРјСѓ callback Рё APK download РѕСЃС‚Р°СЋС‚СЃСЏ HTTPS-СЃРѕРІРјРµСЃС‚РёРјС‹РјРё.",
            "WB Ads: snapshot sync С‚РµРїРµСЂСЊ РёСЃРїРѕР»СЊР·СѓРµС‚ Р±РѕР»РµРµ СЂРёСЂРѕРєРёР№ РЅР°Р±РѕСЂ list/endpoints Рё РїРѕРІС‚РѕСЂРЅРѕ РґРѕР±РёСЂР°РµС‚ СЃС‚Р°С‚РёСЃС‚РёРєСѓ РїРѕ РєР°РјРїР°РЅРёСЏРј, РµСЃР»Рё bulk fullstats РІРµСЂРЅСѓР» С‚РѕР»СЊРєРѕ РїСѓСЃС‚РѕР№ РёР»Рё partial РєРѕРЅС‚РµРєСЃС‚.",
            "Р’РѕР·РІСЂР°С‚С‹: РєР°СЂС‚РѕС‡РєР° Рё С‚Р°Р±Р»РёС†Р° РїРѕРєР°Р·С‹РІР°СЋС‚ Р±РѕР»СЊС€Рµ РїРѕРЅСЏС‚РЅС‹С… РїРѕР»РµР№, Р° РґРµР№СЃС‚РІРёСЏ РїРµСЂРµРёРјРµРЅРѕРІР°РЅС‹ РІ С‡РµР»РѕРІРµРєРѕС‡РёС‚Р°РµРјС‹Рµ РєРЅРѕРїРєРё Р±РµР· РґРІСѓСЃРјС‹СЃР»РµРЅРЅРѕСЃС‚Рё.",
            "APK UI: bidder СЃРЅРѕРІР° РґРµСЂР¶РёС‚СЃСЏ РІ РјРѕР±РёР»СЊРЅРѕР№ РЅР°РІРёРіР°С†РёРё, Р° Р·Р°РіРѕР»РѕРІРѕРє РІРµСЂС…РЅРµР№ СЃС‚СЂРѕРєРё РѕС‚СЂР°Р¶Р°РµС‚ С‚РµРєСѓС‰РёР№ РјРѕРґСѓР»СЊ РІРјРµСЃС‚Рѕ РґСѓР±Р»РµР№ РІРЅСѓС‚СЂРё РїРѕР»РµР·РЅРѕР№ Р·РѕРЅС‹.",
        ],
        "changes": [
            "РЎРѕС†СЃРµС‚СЊ -> РљР°Р»РµРЅРґР°СЂСЊ: РѕРґРЅР° РєРЅРѕРїРєР° В«РџРѕРґРєР»СЋС‡РёС‚СЊ Google РєР°Р»РµРЅРґР°СЂСЊВ» РґР»СЏ РєР°Р¶РґРѕРіРѕ РїРѕР»СЊР·РѕРІР°С‚РµР»СЏ, РїРµСЂСЃРѕРЅР°Р»СЊРЅРѕРµ С…СЂР°РЅРµРЅРёРµ С‚РѕРєРµРЅР° Рё СЃРѕС…СЂР°РЅРµРЅРёРµ account_email / last sync state.",
            "Р РµРєР»Р°РјР° WB: СѓР»СѓС‡СЂРµРЅРѕ РґРѕР±РёСЂР°РЅРёРµ РЅР°Р·РІР°РЅРёР№ РєР°РјРїР°РЅРёР№, СЃС‚Р°С‚СѓСЃРѕРІ, Р·Р°РєР°Р·РѕРІ, СЂР°СЃС…РѕРґР° Рё РґСЂСѓРіРёС… РјРµС‚СЂРёРє РёР· РЅРµСЃРєРѕР»СЊРєРёС… WB Ads API-РёСЃС‚РѕС‡РЅРёРєРѕРІ.",
            "РћС‚Р·С‹РІС‹/Р’РѕРїСЂРѕСЃС‹ -> Р’РѕР·РІСЂР°С‚С‹: С‚Р°Р±Р»РёС†Р° Рё РјРѕРґР°Р»РєР° РґРµС‚Р°Р»РµР№ СЃС‚Р°Р»Рё РїРѕРЅСЏС‚РЅРµРµ РґР»СЏ СЃРѕС‚СЂСѓРґРЅРёРєРѕРІ, СЃ СЏРІРЅС‹РјРё РґРµР№СЃС‚РІРёСЏРјРё Рё СЂР°СЃСЂРёСЂРµРЅРЅС‹Рј РєРѕРЅС‚РµРєСЃС‚РѕРј РїРѕ Р·Р°СЏРІРєРµ.",
            "APK: РІРµСЂС…РЅСЏСЏ СЃС‚СЂРѕРєР° РїСЂРѕРґРѕР»Р¶Р°РµС‚ СЂР°Р±РѕС‚Р°С‚СЊ РєР°Рє Р·Р°РіРѕР»РѕРІРѕРє РјРѕРґСѓР»СЏ, Р° С‡Р°С‚ СЃРѕС…СЂР°РЅСЏРµС‚ РїРµСЂСЃРѕРЅР°Р»СЊРЅС‹Р№ header С‚РѕР»СЊРєРѕ РІРЅСѓС‚СЂРё РєРѕРЅРєСЂРµС‚РЅРѕРіРѕ С‚СЂРµРґР°.",
        ],
        "android_download_url": "/static/downloads/seo-wibe-mobile-latest.apk",
        "android_download_name": "SEO WIBE Android (.apk)",
        "app_entry_url": "/mobile",
        "notes": "APK 1.5.15 СѓСЃС‚Р°РЅР°РІР»РёРІР°РµС‚СЃСЏ РїРѕРІРµСЂС… APK 1.5.12+ (РЅРѕРІС‹Р№ release-РєР»СЋС‡). Р•СЃР»Рё РЅР° СѓСЃС‚СЂРѕР№СЃС‚РІРµ РѕС‡РµРЅСЊ СЃС‚Р°СЂР°СЏ СЃР±РѕСЂРєР° РґРѕ 1.5.12, СѓРґР°Р»РёС‚Рµ РїСЂРёР»РѕР¶РµРЅРёРµ РѕРґРёРЅ СЂР°Р·, Р·Р°С‚РµРј СѓСЃС‚Р°РЅРѕРІРёС‚Рµ APK 1.5.15 Рё РІС‹РїРѕР»РЅРёС‚Рµ РІС…РѕРґ Р·Р°РЅРѕРІРѕ.",
    },
    {
        "version": "0.4.8",
        "android_version_code": 19,
        "released_at": "2026-03-11",
        "current": False,
        "summary": "APK 1.5.13: Google Calendar С‚РµРїРµСЂСЊ РїРѕРґРєР»СЋС‡Р°РµС‚СЃСЏ РѕРґРЅРѕР№ РєРЅРѕРїРєРѕР№ РїРѕ Р»РёС‡РЅРѕРјСѓ Google-Р°РєРєР°СѓРЅС‚Сѓ РїРѕР»СЊР·РѕРІР°С‚РµР»СЏ, WB Ads РґРѕР±РёСЂР°РµС‚ Р±РѕР»СЊС€Рµ СЂРµР°Р»СЊРЅС‹С… РґР°РЅРЅС‹С…, Р° РІ APK РІРѕР·РІСЂР°С‰РµРЅС‹ Р±РёРґРµСЂ Рё РїРѕРЅСЏС‚РЅС‹Рµ Р·Р°РіРѕР»РѕРІРєРё РјРѕРґСѓР»РµР№.",
        "diff_from_previous": [
            "Android APK РѕР±РЅРѕРІР»РµРЅ РґРѕ versionCode=19 / versionName=1.5.13.",
            "РљР°Р»РµРЅРґР°СЂСЊ: Google OAuth РїРµСЂРµРІРµРґРµРЅ РЅР° РїРµСЂСЃРѕРЅР°Р»СЊРЅРѕРµ С…СЂР°РЅРµРЅРёРµ РїРѕ actor_key, РїРѕСЌС‚РѕРјСѓ РєР°Р¶РґС‹Р№ СЃРѕС‚СЂСѓРґРЅРёРє РїРѕРґРєР»СЋС‡Р°РµС‚ СЃРІРѕР№ Google-РєР°Р»РµРЅРґР°СЂСЊ РїРѕ СЃРІРѕРµР№ РїРѕС‡С‚Рµ, Р° РЅРµ РѕР±С‰РёР№ С‚РѕРєРµРЅ РІР»Р°РґРµР»СЊС†Р° РєР°Р±РёРЅРµС‚Р°.",
            "РљР°Р»РµРЅРґР°СЂСЊ: РґР»СЏ APK РґРѕР±Р°РІР»РµРЅ deep link seowibe://open Рё callback-С‚СЂР°РјРїР»РёРЅ, РїРѕСЌС‚РѕРјСѓ РїРѕСЃР»Рµ РІС…РѕРґР° РІ Google РїРѕР»СЊР·РѕРІР°С‚РµР»СЊ РІРѕР·РІСЂР°С‰Р°РµС‚СЃСЏ РїСЂСЏРјРѕ РІ РїСЂРёР»РѕР¶РµРЅРёРµ РЅР° СЌРєСЂР°РЅ РєР°Р»РµРЅРґР°СЂСЏ.",
            "Р’РѕР·РІСЂР°С‚С‹: backend С‚РµРїРµСЂСЊ РІС‹С‚Р°СЃРєРёРІР°РµС‚ СЃС‚СЂРѕРєРё РЅРµ С‚РѕР»СЊРєРѕ РёР· payload.rows, РЅРѕ Рё РёР· claims/returns/items/data/result, Р° РєР°СЂС‚РѕС‡РєР° РїРѕР»СѓС‡Р°РµС‚ СЃС‚Р°С‚СѓСЃ, РєРѕРјРјРµРЅС‚Р°СЂРёРё, РєРѕР»РёС‡РµСЃС‚РІРѕ, СЃСѓРјРјСѓ, vendor code Рё nm_id.",
            "WB Ads: fullstats Р±РѕР»СЊС€Рµ РЅРµ СЃСѓРјРјРёСЂСѓРµС‚ РѕРґРЅРѕРІСЂРµРјРµРЅРЅРѕ РІРµСЂС…РЅРёРµ РёС‚РѕРіРё Рё РІР»РѕР¶РµРЅРЅС‹Рµ РґРЅРµРІРЅС‹Рµ СЂСЏРґС‹; РЅСѓР»РµРІС‹Рµ СЂРµР°Р»СЊРЅС‹Рµ РјРµС‚СЂРёРєРё СЃС‡РёС‚Р°СЋС‚СЃСЏ РІР°Р»РёРґРЅС‹Рј РєРѕРЅС‚РµРєСЃС‚РѕРј, Р° РїСѓСЃС‚С‹Рµ placeholder-РѕС‚РІРµС‚С‹ Р±РѕР»СЊС€Рµ РЅРµ РјР°СЃРєРёСЂСѓСЋС‚ РЅРµРґРѕРіСЂСѓР·РєСѓ.",
            "APK UI: РІ СЂР°РїРєРµ РѕС‚РѕР±СЂР°Р¶Р°РµС‚СЃСЏ РЅР°Р·РІР°РЅРёРµ Р°РєС‚РёРІРЅРѕРіРѕ РјРѕРґСѓР»СЏ, bidder СЃРЅРѕРІР° РґРѕСЃС‚СѓРїРµРЅ РІ РјРѕР±РёР»СЊРЅРѕР№ РЅР°РІРёРіР°С†РёРё Рё СЃР°Р±С‚Р°Р±Р°С… СЂРµРєР»Р°РјС‹, Р° РЅРёР¶РЅРµРµ РїРѕР»РµР·РЅРѕРµ РїСЂРѕСЃС‚СЂР°РЅСЃС‚РІРѕ СЃС‚Р°Р»Рѕ С‡РёС‰Рµ.",
        ],
        "changes": [
            "РЎРѕС†СЃРµС‚СЊ -> РљР°Р»РµРЅРґР°СЂСЊ: РѕРґРЅР° РєРЅРѕРїРєР° В«РџРѕРґРєР»СЋС‡РёС‚СЊ Google РєР°Р»РµРЅРґР°СЂСЊВ» РґР»СЏ РєР°Р¶РґРѕРіРѕ РїРѕР»СЊР·РѕРІР°С‚РµР»СЏ, РїР»СЋСЃ СЃРѕС…СЂР°РЅРµРЅРёРµ account_email Рё СЃС‚Р°С‚СѓСЃР° РїРѕСЃР»РµРґРЅРµР№ СЃРёРЅС…СЂРѕРЅРёР·Р°С†РёРё.",
            "Р РµРєР»Р°РјР° WB: СѓР»СѓС‡СЂРµРЅРѕ РґРѕР±РёСЂР°РЅРёРµ РЅР°Р·РІР°РЅРёР№ РєР°РјРїР°РЅРёР№, РїРѕРєР°Р·РѕРІ, РєР»РёРєРѕРІ, Р·Р°РєР°Р·РѕРІ Рё СЂР°СЃС…РѕРґР° РёР· РЅРµСЃРєРѕР»СЊРєРёС… WB Ads API-РёСЃС‚РѕС‡РЅРёРєРѕРІ СЃ Р±РѕР»РµРµ РєРѕСЂСЂРµРєС‚РЅС‹Рј partial-state.",
            "РћС‚Р·С‹РІС‹/Р’РѕРїСЂРѕСЃС‹ -> Р’РѕР·РІСЂР°С‚С‹: С‚Р°Р±Р»РёС†Р° Рё РєР°СЂС‚РѕС‡РєР° РґРµС‚Р°Р»РµР№ РїРѕРєР°Р·С‹РІР°СЋС‚ Р±РѕР»СЊС€Рµ РїРѕРЅСЏС‚РЅС‹С… РїРѕР»РµР№ РІРјРµСЃС‚Рѕ РїСѓСЃС‚С‹С… РєРЅРѕРїРѕРє Рё РѕР±СЂС‹РІРєРѕРІ raw-СЃС‚СЂСѓРєС‚СѓСЂС‹.",
            "APK: РІРµСЂС…РЅСЏСЏ СЃС‚СЂРѕРєР° С‚РµРїРµСЂСЊ РёСЃРїРѕР»СЊР·СѓРµС‚СЃСЏ РєР°Рє Р·Р°РіРѕР»РѕРІРѕРє С‚РµРєСѓС‰РµРіРѕ РјРѕРґСѓР»СЏ, Р° С‡Р°С‚ СЃРѕС…СЂР°РЅСЏРµС‚ РїРµСЂСЃРѕРЅР°Р»СЊРЅС‹Р№ header С‚РѕР»СЊРєРѕ РІРЅСѓС‚СЂРё РѕС‚РєСЂС‹С‚РѕРіРѕ С‚СЂРµРґР°.",
        ],
        "android_download_url": "/static/downloads/seo-wibe-mobile-latest.apk",
        "android_download_name": "SEO WIBE Android (.apk)",
        "app_entry_url": "/mobile",
        "notes": "Р•СЃР»Рё Сѓ РїРѕР»СЊР·РѕРІР°С‚РµР»СЏ СѓР¶Рµ СѓСЃС‚Р°РЅРѕРІР»РµРЅ APK 1.5.12 СЃ РЅРѕРІС‹Рј release-РєР»СЋС‡РѕРј, APK 1.5.13 СЃС‚Р°РІРёС‚СЃСЏ РїРѕРІРµСЂС… РЅРµРіРѕ РєР°Рє РѕР±С‹С‡РЅРѕРµ РѕР±РЅРѕРІР»РµРЅРёРµ. Р•СЃР»Рё РЅР° СѓСЃС‚СЂРѕР№СЃС‚РІРµ РІСЃРµ РµС‰Рµ СЃС‚Р°СЂР°СЏ РІРµСЂСЃРёСЏ РґРѕ 1.5.12, СЃРЅР°С‡Р°Р»Р° СѓРґР°Р»РёС‚Рµ СЃС‚Р°СЂРѕРµ РїСЂРёР»РѕР¶РµРЅРёРµ РѕРґРёРЅ СЂР°Р·, Р·Р°С‚РµРј СѓСЃС‚Р°РЅРѕРІРёС‚Рµ РЅРѕРІС‹Р№ APK 1.5.13 Рё РІРѕР№РґРёС‚Рµ РІ Р°РєРєР°СѓРЅС‚ Р·Р°РЅРѕРІРѕ.",
    },
    {
        "version": "0.4.7",
        "android_version_code": 18,
        "released_at": "2026-03-10",
        "current": False,
        "summary": "APK 1.5.12: РІС‹РїСѓС‰РµРЅ РЅР° РЅРѕРІРѕРј РїРѕСЃС‚РѕСЏРЅРЅРѕРј release-РєР»СЋС‡Рµ, РїРѕСЌС‚РѕРјСѓ СѓСЃС‚Р°РЅР°РІР»РёРІР°РµС‚СЃСЏ С‚РѕР»СЊРєРѕ РїРѕСЃР»Рµ СѓРґР°Р»РµРЅРёСЏ СЃС‚Р°СЂРѕРіРѕ РїСЂРёР»РѕР¶РµРЅРёСЏ.",
        "diff_from_previous": [
            "Android APK РѕР±РЅРѕРІР»РµРЅ РґРѕ versionCode=18 / versionName=1.5.12.",
            "РџРѕРґРїРёСЃСЊ APK РїРµСЂРµРІРµРґРµРЅР° РЅР° РЅРѕРІС‹Р№ РїРѕСЃС‚РѕСЏРЅРЅС‹Р№ release-РєР»СЋС‡: future-РѕР±РЅРѕРІР»РµРЅРёСЏ С‚РµРїРµСЂСЊ РґРѕР»Р¶РЅС‹ СЃРѕР±РёСЂР°С‚СЊСЃСЏ СЌС‚РёРј Р¶Рµ keystore, Р° РЅРµ debug-РїРѕРґРїРёСЃСЊСЋ.",
            "РњРёРіСЂР°С†РёСЏ РЅР° 1.5.12 СЂР°Р·РѕРІР°СЏ: Android РЅРµ РјРѕР¶РµС‚ РѕР±РЅРѕРІРёС‚СЊ РїРѕРІРµСЂС… СЃС‚Р°СЂРѕР№ СЃР±РѕСЂРєРё СЃ РґСЂСѓРіРёРј СЃРµСЂС‚РёС„РёРєР°С‚РѕРј, РїРѕСЌС‚РѕРјСѓ РїРµСЂРµРґ СѓСЃС‚Р°РЅРѕРІРєРѕР№ РЅСѓР¶РЅРѕ СѓРґР°Р»РёС‚СЊ РїСЂРµР¶РЅРµРµ РїСЂРёР»РѕР¶РµРЅРёРµ.",
            "РџСЂРѕРІРµСЂРєР° РѕР±РЅРѕРІР»РµРЅРёР№ /api/mobile/apk/latest Рё РєР°СЂС‚РѕС‡РєР° СЂРµР»РёР·Р° С‚РµРїРµСЂСЊ СЏРІРЅРѕ РїСЂРµРґСѓРїСЂРµР¶РґР°СЋС‚ Рѕ СЂСѓС‡РЅРѕР№ РїРµСЂРµСѓСЃС‚Р°РЅРѕРІРєРµ, С‡С‚РѕР±С‹ РїРѕР»СЊР·РѕРІР°С‚РµР»СЊ РЅРµ СѓРїРёСЂР°Р»СЃСЏ РІ РєРѕРЅС„Р»РёРєС‚ РїР°РєРµС‚РѕРІ.",
            "РљР°Р»РµРЅРґР°СЂСЊ, WB Ads Рё РјРѕР±РёР»СЊРЅС‹Р№ С‡Р°С‚ РѕСЃС‚Р°СЋС‚СЃСЏ РІ СЃРѕСЃС‚Р°РІРµ СЂРµР»РёР·Р° 1.5.12 РІРјРµСЃС‚Рµ СЃ СЂР°РЅРµРµ РІРЅРµСЃРµРЅРЅС‹РјРё С„СѓРЅРєС†РёРѕРЅР°Р»СЊРЅС‹РјРё РёСЃРїСЂР°РІР»РµРЅРёСЏРјРё.",
        ],
        "changes": [
            "Android APK: СЂРµР»РёР·РЅР°СЏ СЃР±РѕСЂРєР° Р±РѕР»СЊС€Рµ РЅРµ РґРѕР»Р¶РЅР° РїСѓР±Р»РёРєРѕРІР°С‚СЊСЃСЏ СЃ debug-keystore; РґР»СЏ РЅРµРµ РІРІРµРґРµРЅ РїРѕСЃС‚РѕСЏРЅРЅС‹Р№ release-РєР»СЋС‡ Рё РѕС‚РґРµР»СЊРЅР°СЏ РёРЅСЃС‚СЂСѓРєС†РёСЏ РїРѕ СЃР±РѕСЂРєРµ.",
            "Р—Р°РіСЂСѓР·РєРё/РЎРїСЂР°РІРєР°: С‚РµРєСѓС‰РёР№ СЂРµР»РёР· РїРѕРјРµС‡РµРЅ РєР°Рє migration build СЃ СЏРІРЅС‹Рј СЃС†РµРЅР°СЂРёРµРј В«СѓРґР°Р»РёС‚СЊ СЃС‚Р°СЂРѕРµ РїСЂРёР»РѕР¶РµРЅРёРµ -> СѓСЃС‚Р°РЅРѕРІРёС‚СЊ APK 1.5.12В».",
            "РћР±РЅРѕРІР»РµРЅС‹ РјРµС‚Р°РґР°РЅРЅС‹Рµ help/releases Рё mobile/apk/latest, С‡С‚РѕР±С‹ СЃС‚Р°СЂС‹Р№ РєР»РёРµРЅС‚ РїРѕРєР°Р·С‹РІР°Р» РєРѕСЂСЂРµРєС‚РЅС‹Рµ РёРЅСЃС‚СЂСѓРєС†РёРё РІРјРµСЃС‚Рѕ РїРѕРїС‹С‚РєРё Р±РµСЃСЂРѕРІРЅРѕРіРѕ РѕР±РЅРѕРІР»РµРЅРёСЏ.",
        ],
        "android_download_url": "/static/downloads/seo-wibe-mobile-latest.apk",
        "android_download_name": "SEO WIBE Android (.apk)",
        "app_entry_url": "/mobile",
        "notes": "Р’Р°Р¶РЅРѕ: APK 1.5.12 РЅРµ СѓСЃС‚Р°РЅР°РІР»РёРІР°РµС‚СЃСЏ РїРѕРІРµСЂС… СЃС‚Р°СЂС‹С… РІРµСЂСЃРёР№, РїРѕРґРїРёСЃР°РЅРЅС‹С… РґСЂСѓРіРёРј РєР»СЋС‡РѕРј. РЈРґР°Р»РёС‚Рµ С‚РµРєСѓС‰РµРµ РїСЂРёР»РѕР¶РµРЅРёРµ SEO WIBE, Р·Р°С‚РµРј СЃРєР°С‡Р°Р№С‚Рµ Рё СѓСЃС‚Р°РЅРѕРІРёС‚Рµ РЅРѕРІС‹Р№ APK 1.5.12 РёР· В«РЎРїСЂР°РІРєР° -> Р—Р°РіСЂСѓР·РєРёВ». РџРѕСЃР»Рµ СѓСЃС‚Р°РЅРѕРІРєРё Р·Р°РЅРѕРІРѕ РІРѕР№РґРёС‚Рµ РІ Р°РєРєР°СѓРЅС‚ Рё РїСЂРѕРІРµСЂСЊС‚Рµ В«РЎРѕС†СЃРµС‚СЊ -> РљР°Р»РµРЅРґР°СЂСЊВ», В«Р РµРєР»Р°РјР° -> РђРЅР°Р»РёС‚РёРєР° WB AdsВ» Рё Р»СЋР±РѕР№ С‡Р°С‚ РЅР° РјРѕР±РёР»СЊРЅРѕРј СЌРєСЂР°РЅРµ.",
    },
    {
        "version": "0.4.6",
        "android_version_code": 17,
        "released_at": "2026-03-10",
        "current": False,
        "summary": "APK 1.5.11: РєР°Р»РµРЅРґР°СЂСЊ СЃРёРЅС…СЂРѕРЅРёР·РёСЂСѓРµС‚СЃСЏ СЃС‚Р°Р±РёР»СЊРЅРѕ, WB Ads РїРѕРєР°Р·С‹РІР°РµС‚ СЂРµР°Р»СЊРЅС‹Рµ РЅР°Р·РІР°РЅРёСЏ Рё РјРµС‚СЂРёРєРё, Р° РјРѕР±РёР»СЊРЅС‹Р№ С‡Р°С‚ РёСЃРїСЂР°РІР»РµРЅ РґР»СЏ APK.",
        "diff_from_previous": [
            "Android APK РѕР±РЅРѕРІР»РµРЅ РґРѕ versionCode=17 / versionName=1.5.11.",
            "РљР°Р»РµРЅРґР°СЂСЊ: Google OAuth Рё /mobile/apk/latest С‚РµРїРµСЂСЊ СЃС‚СЂРѕСЏС‚СЃСЏ РѕС‚ РєР°РЅРѕРЅРёС‡РµСЃРєРѕРіРѕ public base URL, РєРѕСЂСЂРµРєС‚РЅРѕ СЂР°Р±РѕС‚Р°СЋС‚ Р·Р° proxy/HTTPS.",
            "РљР°Р»РµРЅРґР°СЂСЊ: СЃРёРЅС…СЂРѕРЅРёР·Р°С†РёСЏ РїРѕ Google OAuth Рё ICS URL РїРѕР»СѓС‡РёР»Р° СЃРѕС…СЂР°РЅРµРЅРёРµ last-sync СЃС‚Р°С‚СѓСЃР°, РїРѕРЅСЏС‚РЅС‹Рµ РѕСЂРёР±РєРё Рё Р»РѕРєР°Р»СЊРЅСѓСЋ РѕР±СЂР°Р±РѕС‚РєСѓ РіСЂР°РЅРёС† РјРµСЃСЏС†Р°/РґРЅСЏ Р±РµР· UTC-СЃРґРІРёРіРѕРІ.",
            "РљР°Р»РµРЅРґР°СЂСЊ: СЌРєСЂР°РЅ РїРµСЂРµСЂР°Р±РѕС‚Р°РЅ РІ РЅРѕРІС‹Р№ productivity-layout СЃ РЅР°РІРёРіР°С†РёРѕРЅРЅС‹Рј РєР»Р°СЃС‚РµСЂРѕРј, sync-card Рё Р±РѕР»РµРµ СѓРґРѕР±РЅРѕР№ РјРѕР±РёР»СЊРЅРѕР№ РІРµСЂСЃС‚РєРѕР№.",
            "WB Ads: fullstats Р±РѕР»СЊС€Рµ РЅРµ РїСЂРёРЅРёРјР°РµС‚ РїСѓСЃС‚С‹Рµ zero-like РѕС‚РІРµС‚С‹ Р·Р° РІР°Р»РёРґРЅСѓСЋ СЃС‚Р°С‚РёСЃС‚РёРєСѓ, РґРѕР±Р°РІР»РµРЅС‹ РїРѕРІС‚РѕСЂС‹ Рё richer meta РїРѕ С‡Р°СЃС‚РёС‡РЅРѕР№ Р·Р°РіСЂСѓР·РєРµ.",
            "WB Ads: Р°РЅР°Р»РёС‚РёРєР° Рё С‚Р°Р±Р»РёС†С‹ РёСЃРїРѕР»СЊР·СѓСЋС‚ summary enrichment, РїРѕСЌС‚РѕРјСѓ СЂРµР°Р»СЊРЅС‹Рµ name/status/type/budget РїСЂРёС…РѕРґСЏС‚ РІРјРµСЃС‚Рµ СЃ РїРѕРєР°Р·Р°РјРё, РєР»РёРєР°РјРё Рё СЂР°СЃС…РѕРґРѕРј, РєРѕРіРґР° WB API РёС… РѕС‚РґР°РµС‚.",
            "Р§Р°С‚/APK: Android back СЃРЅРѕРІР° РёРґРµС‚ РїРѕ С†РµРїРѕС‡РєРµ thread в†’ chat list в†’ modules menu, Р° РјРµРЅСЋ СЂРµР°РєС†РёР№ СѓРґРµСЂР¶РёРІР°РµС‚СЃСЏ РІ РІРёРґРёРјРѕР№ Р·РѕРЅРµ СЌРєСЂР°РЅР°.",
        ],
        "changes": [
            "РЎРѕС†РёР°Р»СЊРЅС‹Р№ РјРѕРґСѓР»СЊ: РєР°Р»РµРЅРґР°СЂСЊ РїРѕР»СѓС‡РёР» РїРѕСЃС‚РѕСЏРЅРЅС‹Р№ СЃС‚Р°С‚СѓСЃ СЃРёРЅС…СЂРѕРЅРёР·Р°С†РёРё, РєР°СЂС‚РѕС‡РєСѓ РїРѕРґРєР»СЋС‡РµРЅРёСЏ Google/ICS Рё С‡РёС‚Р°РµРјС‹Р№ СЃРїРёСЃРѕРє СЃРѕР±С‹С‚РёР№ РІС‹Р±СЂР°РЅРЅРѕРіРѕ РґРЅСЏ.",
            "РЎРѕС†РёР°Р»СЊРЅС‹Р№ РјРѕРґСѓР»СЊ: РёСЃРїСЂР°РІР»РµРЅР° Р»РѕРіРёРєР° РјРѕР±РёР»СЊРЅРѕРіРѕ back Рё РїРѕР·РёС†РёРѕРЅРёСЂРѕРІР°РЅРёРµ reaction menu Сѓ РЅРёР¶РЅРµРіРѕ РєСЂР°СЏ СЌРєСЂР°РЅР°.",
            "Р РµРєР»Р°РјР°: РІ WB Ads Р°РЅР°Р»РёС‚РёРєРµ РґРѕР±Р°РІР»РµРЅС‹ РїСЂРёР·РЅР°РєРё partial/temporary_unavailable Рё СЃРѕС…СЂР°РЅРµРЅС‹ СЂРµР°Р»СЊРЅС‹Рµ РЅР°Р·РІР°РЅРёСЏ РєР°РјРїР°РЅРёР№ РїСЂРё РЅР°Р»РёС‡РёРё summary API.",
            "РћР±РЅРѕРІР»РµРЅС‹ РІРµСЂСЃРёРё СЃС‚Р°С‚РёС‡РµСЃРєРёС… СЂРµСЃСѓСЂСЃРѕРІ (styles/app/social) Рё РјРµС‚Р°РґР°РЅРЅС‹Рµ help/releases РґР»СЏ РїСЂРёРЅСѓРґРёС‚РµР»СЊРЅРѕРіРѕ РѕР±РЅРѕРІР»РµРЅРёСЏ РєР»РёРµРЅС‚Р°.",
        ],
        "android_download_url": "/static/downloads/seo-wibe-mobile-latest.apk",
        "android_download_name": "SEO WIBE Android (.apk)",
        "app_entry_url": "/mobile",
        "notes": "APK 1.5.11 СЃС‚Р°РІРёС‚СЃСЏ РїРѕРІРµСЂС… РїСЂРµРґС‹РґСѓС‰РёС… РІРµСЂСЃРёР№. РџРѕСЃР»Рµ РѕР±РЅРѕРІР»РµРЅРёСЏ РѕС‚РєСЂРѕР№С‚Рµ В«РЎРѕС†СЃРµС‚СЊ в†’ РљР°Р»РµРЅРґР°СЂСЊВ», В«Р РµРєР»Р°РјР° в†’ РђРЅР°Р»РёС‚РёРєР° WB AdsВ» Рё Р»СЋР±РѕР№ С‡Р°С‚ РЅР° РјРѕР±РёР»СЊРЅРѕРј СЌРєСЂР°РЅРµ, С‡С‚РѕР±С‹ РїСЂРѕРІРµСЂРёС‚СЊ sync, СЂРµР°Р»СЊРЅС‹Рµ РјРµС‚СЂРёРєРё Рё РёСЃРїСЂР°РІР»РµРЅРЅС‹Р№ back/reactions.",
    },
    {
        "version": "0.4.5",
        "android_version_code": 16,
        "released_at": "2026-03-09",
        "current": False,
        "summary": "APK 1.5.10: РІРѕР·РІСЂР°С‚С‹ Р±РµР· RAW, РїСЂР°РІР° СЃРѕС‚СЂСѓРґРЅРёРєРѕРІ СѓР¶РµСЃС‚РѕС‡РµРЅС‹, Р±РёРґРґРµСЂ/СЃРїСЂР°РІРєР° Р·Р°РєСЂРµРїР»РµРЅС‹ РІ РјРѕР±РёР»СЊРЅРѕР№ РЅР°РІРёРіР°С†РёРё.",
        "diff_from_previous": [
            "Android APK РѕР±РЅРѕРІР»РµРЅ РґРѕ versionCode=16 / versionName=1.5.10.",
            "Р’РѕР·РІСЂР°С‚С‹: СѓР±СЂР°РЅ РїРѕР»СЊР·РѕРІР°С‚РµР»СЊСЃРєРёР№ RAW-Р±Р»РѕРє, РєР°СЂС‚РѕС‡РєР° РґРµС‚Р°Р»РµР№ СЃС‚Р°Р»Р° С‚РѕР»СЊРєРѕ РІ РїРѕРЅСЏС‚РЅРѕРј С„РѕСЂРјР°С‚Рµ (РїРѕР»СЏ, С„РѕС‚Рѕ, РѕРїРёСЃР°РЅРёРµ).",
            "Р’РѕР·РІСЂР°С‚С‹: СѓР»СѓС‡СЂРµРЅ СЂР°Р·Р±РѕСЂ ID/СЃС‚Р°С‚СѓСЃР°/С‚РѕРІР°СЂР° РёР· РІР»РѕР¶РµРЅРЅС‹С… WB/Ozon СЃС‚СЂСѓРєС‚СѓСЂ, СѓРјРµРЅСЊС€РµРЅС‹ В«РїСѓСЃС‚С‹РµВ» СЃС‚СЂРѕРєРё РІ С‚Р°Р±Р»РёС†Рµ.",
            "РџСЂРѕС„РёР»СЊ: Сѓ СЃРѕС‚СЂСѓРґРЅРёРєРѕРІ СЃРєСЂС‹С‚Р° РєРЅРѕРїРєР° В«Р”РѕР±Р°РІРёС‚СЊ СЃРѕС‚СЂСѓРґРЅРёРєР°В», РґРѕР±Р°РІР»РµРЅР° Р·Р°С‰РёС‚Р° РЅР° РєР»РёРµРЅС‚Рµ РѕС‚ СЃРѕР·РґР°РЅРёСЏ СЃРѕС‚СЂСѓРґРЅРёРєРѕРІ.",
            "Р—Р°РґР°С‡Рё: СЃРѕС‚СЂСѓРґРЅРёРє РјРѕР¶РµС‚ Р·Р°РєСЂС‹РІР°С‚СЊ С‚РѕР»СЊРєРѕ СЃРІРѕРё Р·Р°РґР°С‡Рё (Рё РІ UI, Рё РЅР° API).",
            "РџСЂРѕС„РёР»СЊ РІ СЂР°РїРєРµ: С‚РµРїРµСЂСЊ РїРѕРєР°Р·С‹РІР°РµС‚СЃСЏ email С‚РµРєСѓС‰РµРіРѕ СЃРѕС‚СЂСѓРґРЅРёРєР° (actor_email), Р° РЅРµ email РІР»Р°РґРµР»СЊС†Р°.",
            "РЎРїСЂР°РІРєР°: РѕР±С‰РёР№ С‡РµРє-Р»РёСЃС‚ РѕР±РЅРѕРІР»РµРЅ Р±РµР· СѓРїРѕСЂР° РЅР° RAW-РґРёР°РіРЅРѕСЃС‚РёРєСѓ РґР»СЏ РїРѕР»СЊР·РѕРІР°С‚РµР»СЊСЃРєРёС… СЃС†РµРЅР°СЂРёРµРІ.",
        ],
        "changes": [
            "РЎРѕС†РёР°Р»СЊРЅС‹Р№ РјРѕРґСѓР»СЊ: РІ СЃРїРёСЃРєРµ Р·Р°РґР°С‡ РєРЅРѕРїРєР° В«Р—Р°РєСЂС‹С‚СЊВ» РІРёРґРЅР° С‚РѕР»СЊРєРѕ РІР»Р°РґРµР»СЊС†Сѓ РёР»Рё РЅР°Р·РЅР°С‡РµРЅРЅРѕРјСѓ РёСЃРїРѕР»РЅРёС‚РµР»СЋ.",
            "РџСЂРѕС„РёР»СЊ РІ APK: РІ СЂР°РїРєРµ Рё РїСЂРѕС„РёР»Рµ С‚РµРїРµСЂСЊ РїРѕРєР°Р·С‹РІР°РµС‚СЃСЏ email Р°РєС‚РёРІРЅРѕРіРѕ СЃРѕС‚СЂСѓРґРЅРёРєР°, Р° РЅРµ РІР»Р°РґРµР»СЊС†Р° РєРѕРјРїР°РЅРёРё.",
            "РћР±РЅРѕРІР»РµРЅС‹ РІРµСЂСЃРёРё СЃС‚Р°С‚РёС‡РµСЃРєРёС… СЂРµСЃСѓСЂСЃРѕРІ (styles/app/social) РґР»СЏ РїСЂРёРЅСѓРґРёС‚РµР»СЊРЅРѕРіРѕ РѕР±РЅРѕРІР»РµРЅРёСЏ РєСЌСЂР° Сѓ РєР»РёРµРЅС‚РѕРІ.",
            "РЎРѕС…СЂР°РЅРµРЅС‹ РґРѕСЂР°Р±РѕС‚РєРё 1.5.9 РїРѕ Р±С‹СЃС‚СЂС‹Рј РїРµСЂРµС…РѕРґР°Рј Рє Р‘РёРґРµСЂСѓ/РЎРїСЂР°РІРєРµ Рё С„РѕС‚Рѕ-СЂРµРґР°РєС‚РѕСЂСѓ С‚РѕРІР°СЂР°.",
        ],
        "android_download_url": "/static/downloads/seo-wibe-mobile-latest.apk",
        "android_download_name": "SEO WIBE Android (.apk)",
        "app_entry_url": "/mobile",
        "notes": "APK 1.5.10 СЃС‚Р°РІРёС‚СЃСЏ РїРѕРІРµСЂС… РїСЂРµРґС‹РґСѓС‰РµР№ РІРµСЂСЃРёРё. РџРѕСЃР»Рµ РѕР±РЅРѕРІР»РµРЅРёСЏ РѕС‚РєСЂРѕР№С‚Рµ В«Р РµРєР»Р°РјР° в†’ Р‘РёРґРµСЂВ», В«РЎРїСЂР°РІРєР°В» Рё В«РћС‚Р·С‹РІС‹/Р’РѕР·РІСЂР°С‚С‹В» РґР»СЏ РїСЂРѕРІРµСЂРєРё РЅРѕРІРѕРіРѕ РёРЅС‚РµСЂС„РµР№СЃР°.",
    },
    {
        "version": "0.4.4",
        "android_version_code": 15,
        "released_at": "2026-03-09",
        "current": False,
        "summary": "APK 1.5.9: РґРѕР±Р°РІР»РµРЅС‹ Р±С‹СЃС‚СЂС‹Рµ РїРµСЂРµС…РѕРґС‹ Рє Р‘РёРґРµСЂСѓ/РЎРїСЂР°РІРєРµ, СѓР»СѓС‡СЂРµРЅ СЂРµРґР°РєС‚РѕСЂ С„РѕС‚Рѕ С‚РѕРІР°СЂРѕРІ Рё СѓСЃРёР»РµРЅС‹ СѓРІРµРґРѕРјР»РµРЅРёСЏ РїРѕ Р·Р°РґР°С‡Р°Рј.",
        "diff_from_previous": [
            "Android APK РѕР±РЅРѕРІР»РµРЅ РґРѕ versionCode=15 / versionName=1.5.9.",
            "Р’ РјРѕР±РёР»СЊРЅРѕРј Р±С‹СЃС‚СЂРѕРј РїРµСЂРµРєР»СЋС‡Р°С‚РµР»Рµ РґРѕР±Р°РІР»РµРЅС‹ РїСѓРЅРєС‚С‹ В«Р РµРєР»Р°РјР°: РєР°РјРїР°РЅРёРёВ», В«Р РµРєР»Р°РјР°: Р±РёРґРµСЂВ», В«РЎРїСЂР°РІРєР°В».",
            "Р РµРґР°РєС‚РѕСЂ С‚РѕРІР°СЂР°: РєРЅРѕРїРєР° В«Р”РѕР±Р°РІРёС‚СЊ С„РѕС‚РѕВ» РѕС‚РєСЂС‹РІР°РµС‚ Р»РѕРєР°Р»СЊРЅС‹Рµ РїР°РїРєРё Рё РїРѕРґРґРµСЂР¶РёРІР°РµС‚ РјРЅРѕР¶РµСЃС‚РІРµРЅРЅС‹Р№ РІС‹Р±РѕСЂ С„Р°Р№Р»РѕРІ.",
            "Р’РѕР·РІСЂР°С‚С‹: РІРјРµСЃС‚Рѕ RAW-alert РѕС‚РєСЂС‹РІР°РµС‚СЃСЏ РґРµС‚Р°Р»СЊРЅРѕРµ РјРѕРґР°Р»СЊРЅРѕРµ РѕРєРЅРѕ СЃ РїРѕРЅСЏС‚РЅС‹РјРё РїРѕР»СЏРјРё Рё С„РѕС‚Рѕ.",
            "Р—Р°РґР°С‡Рё: СѓР»СѓС‡СЂРµРЅР° РґРѕСЃС‚Р°РІРєР° СѓРІРµРґРѕРјР»РµРЅРёР№ РїСЂРё alias actor_key (u:/m:), РґРѕР±Р°РІР»РµРЅР° РІРёР·СѓР°Р»СЊРЅР°СЏ РїРѕРґСЃРІРµС‚РєР° done/overdue.",
            "РљР°Р»РµРЅРґР°СЂСЊ: РґРѕР±Р°РІР»РµРЅР° СЃРёРЅС…СЂРѕРЅРёР·Р°С†РёСЏ РїРѕ Google Calendar ICS URL (РёРјРїРѕСЂС‚/РѕР±РЅРѕРІР»РµРЅРёРµ СЃРѕР±С‹С‚РёР№).",
        ],
        "changes": [
            "WB Ads РєР°РјРїР°РЅРёРё: СѓСЃРёР»РµРЅР° СЃРµСЂРІРµСЂРЅР°СЏ РґРѕРіСЂСѓР·РєР° РЅР°Р·РІР°РЅРёР№/СЃС‚Р°С‚СѓСЃРѕРІ/СЃС‚Р°С‚РёСЃС‚РёРєРё РґР»СЏ placeholder-СЃС‚СЂРѕРє.",
            "Р‘СѓС…РіР°Р»С‚РµСЂРёСЏ: РЅРѕСЂРјР°Р»РёР·РѕРІР°РЅС‹ СЂСѓРјРЅС‹Рµ СЃРѕРѕР±С‰РµРЅРёСЏ WB finance/sales API РІ Р±РѕР»РµРµ РїРѕРЅСЏС‚РЅС‹Рµ РїСЂРµРґСѓРїСЂРµР¶РґРµРЅРёСЏ.",
            "РўРѕРІР°СЂС‹: РґРѕР±Р°РІР»РµРЅС‹ РїРѕР»СЏ С†РµРЅ (base/discount/min/marketing), СЃРёРЅС…СЂРѕРЅРёР·РёСЂРѕРІР°РЅРЅС‹Рµ СЃ РїСЂРѕСЃРјРѕС‚СЂРѕРј/СЂРµРґР°РєС‚РёСЂРѕРІР°РЅРёРµРј.",
            "РЎРїСЂР°РІРєР°: СЂР°СЃСЂРёСЂРµРЅРѕ РѕРїРёСЃР°РЅРёРµ РјРѕРґСѓР»СЏ Р‘РёРґРµСЂ Рё РєР°СЂС‚РѕС‡РєРё С‚РѕРІР°СЂРѕРІ (РїРѕР»СЏ С†РµРЅ, СЂР°Р±РѕС‚Р° СЃ С„РѕС‚Рѕ).",
        ],
        "android_download_url": "/static/downloads/seo-wibe-mobile-latest.apk",
        "android_download_name": "SEO WIBE Android (.apk)",
        "app_entry_url": "/mobile",
        "notes": "APK 1.5.9 СѓСЃС‚Р°РЅР°РІР»РёРІР°РµС‚СЃСЏ РїРѕРІРµСЂС… РїСЂРµРґС‹РґСѓС‰РёС… РІРµСЂСЃРёР№. РџРѕСЃР»Рµ РѕР±РЅРѕРІР»РµРЅРёСЏ РѕС‚РєСЂРѕР№С‚Рµ В«Р РµРєР»Р°РјР° в†’ Р‘РёРґРµСЂВ» Рё В«РўРѕРІР°СЂС‹ в†’ Р РµРґР°РєС‚РёСЂРѕРІР°С‚СЊВ», С‡С‚РѕР±С‹ РїСЂРѕРІРµСЂРёС‚СЊ РЅРѕРІС‹Рµ СЌР»РµРјРµРЅС‚С‹.",
    },
    {
        "version": "0.4.3",
        "android_version_code": 14,
        "released_at": "2026-03-09",
        "current": False,
        "summary": "APK 1.5.8: СѓСЃРєРѕСЂРµРЅР° СЃС‚Р°С‚РёСЃС‚РёРєР° (С„РѕСЂСЃ-СЂРµС„СЂРµСЂ), СѓСЃРёР»РµРЅР° РґРѕРіСЂСѓР·РєР° WB Ads РєР°РјРїР°РЅРёР№ Рё РёСЃРїСЂР°РІР»РµРЅ back РёР· С‡Р°С‚Р° РІ Android.",
        "diff_from_previous": [
            "Android APK РѕР±РЅРѕРІР»РµРЅ РґРѕ versionCode=14 / versionName=1.5.8.",
            "РЎС‚Р°С‚РёСЃС‚РёРєР° РїСЂРѕРґР°Р¶: РєРЅРѕРїРєР° В«Р—Р°РіСЂСѓР·РёС‚СЊ СЃС‚Р°С‚РёСЃС‚РёРєСѓВ» РІС‹РїРѕР»РЅСЏРµС‚ РїСЂРёРЅСѓРґРёС‚РµР»СЊРЅС‹Р№ refresh Р±РµР· РѕР¶РёРґР°РЅРёСЏ cache TTL.",
            "РЎС‚Р°С‚РёСЃС‚РёРєР°: РґР»СЏ РґР»РёРЅРЅС‹С… РїРµСЂРёРѕРґРѕРІ СѓСЃРєРѕСЂРµРЅ СЂРµР¶РёРј СЃСЂР°РІРЅРµРЅРёСЏ (РїСЂРµРґС‹РґСѓС‰РёР№ РїРµСЂРёРѕРґ РѕРіСЂР°РЅРёС‡РµРЅ РјРµРЅСЊСЂРёРј РґРёР°РїР°Р·РѕРЅРѕРј).",
            "WB Ads: СѓРІРµР»РёС‡РµРЅС‹ Р»РёРјРёС‚С‹ РґРѕР±РѕСЂР° summary/stat, С‡С‚РѕР±С‹ С‡Р°С‰Рµ РїРѕРґС‚СЏРіРёРІР°Р»РёСЃСЊ СЂРµР°Р»СЊРЅС‹Рµ РЅР°Р·РІР°РЅРёСЏ Рё РјРµС‚СЂРёРєРё РєР°РјРїР°РЅРёР№.",
            "РњРѕР±РёР»СЊРЅС‹Р№ back РІ СЃРѕС†РёР°Р»СЊРЅРѕРј РјРѕРґСѓР»Рµ: РёР· РѕС‚РєСЂС‹С‚РѕРіРѕ РґРёР°Р»РѕРіР° РІСЃРµРіРґР° РІРѕР·РІСЂР°С‰Р°РµС‚ РІ СЃРїРёСЃРѕРє С‡Р°С‚РѕРІ.",
        ],
        "changes": [
            "WB campaigns enrich СѓСЃРєРѕСЂРµРЅ РЅР° Р±РѕР»СЊСЂРёС… РєР°Р±РёРЅРµС‚Р°С… Р·Р° СЃС‡РµС‚ Р±РѕР»РµРµ РєСЂСѓРїРЅРѕРіРѕ batching РЅР° РєР»РёРµРЅС‚Рµ.",
            "Р’ audit РґРѕР±Р°РІР»РµРЅ С„Р»Р°Рі force-refresh РґР»СЏ РґРёР°РіРЅРѕСЃС‚РёРєРё Р·Р°РїСЂРѕСЃРѕРІ СЃС‚Р°С‚РёСЃС‚РёРєРё.",
            "РЎРѕС…СЂР°РЅРµРЅС‹ РґРѕСЂР°Р±РѕС‚РєРё 1.5.7 РїРѕ РІРѕР·РІСЂР°С‚Р°Рј, СЃРµСЃСЃРёРё, С„РѕС‚Рѕ-СЂРµРґР°РєС‚РѕСЂСѓ Рё bidder.",
        ],
        "android_download_url": "/static/downloads/seo-wibe-mobile-latest.apk",
        "android_download_name": "SEO WIBE Android (.apk)",
        "app_entry_url": "/mobile",
        "notes": "APK 1.5.8 СѓСЃС‚Р°РЅР°РІР»РёРІР°РµС‚СЃСЏ РїРѕРІРµСЂС… РїСЂРµРґС‹РґСѓС‰РµР№ РІРµСЂСЃРёРё. РџРѕСЃР»Рµ РѕР±РЅРѕРІР»РµРЅРёСЏ РѕС‚РєСЂРѕР№С‚Рµ РјРѕРґСѓР»СЊ В«РЎС‚Р°С‚РёСЃС‚РёРєР°В» Рё РІС‹РїРѕР»РЅРёС‚Рµ СЂСѓС‡РЅРѕР№ refresh РѕРґРёРЅ СЂР°Р·.",
    },
    {
        "version": "0.4.2",
        "android_version_code": 13,
        "released_at": "2026-03-09",
        "current": False,
        "summary": "APK 1.5.7: СЃС‚Р°Р±РёР»СЊРЅРѕСЃС‚СЊ СЃРµСЃСЃРёРё Рё Р·Р°РіСЂСѓР·РѕРє, WB Ads/Returns РґРѕСЂР°Р±РѕС‚Р°РЅС‹, РІ СЂРµРґР°РєС‚РѕСЂ С‚РѕРІР°СЂР° РґРѕР±Р°РІР»РµРЅР° Р·Р°РіСЂСѓР·РєР° С„РѕС‚Рѕ С„Р°Р№Р»Р°РјРё.",
        "diff_from_previous": [
            "Android APK РѕР±РЅРѕРІР»РµРЅ РґРѕ versionCode=13 / versionName=1.5.7.",
            "WB Ads: СЂР°СЃС€РёСЂРµРЅР° РґРѕРіСЂСѓР·РєР° РєР°РјРїР°РЅРёР№ (Р±РѕР»СЊС€Рµ Р»РёРјРёС‚С‹ enrichment + РґРѕРїРѕР»РЅРёС‚РµР»СЊРЅС‹Р№ fallback РїРѕ summary/stat).",
            "WB Returns: РґРѕР±Р°РІР»РµРЅС‹ fallback-РІР°СЂРёР°РЅС‚С‹ РїР°СЂР°РјРµС‚СЂРѕРІ is_archive Рё РїРѕСЃС‚СЂР°РЅРёС‡РЅС‹Р№ СЃР±РѕСЂ СЃРїРёСЃРєРѕРІ Р·Р°СЏРІРѕРє.",
            "Р’ СЂРµРґР°РєС‚РѕСЂ С‚РѕРІР°СЂР° РґРѕР±Р°РІР»РµРЅР° Р·Р°РіСЂСѓР·РєР° С„РѕС‚Рѕ С„Р°Р№Р»Р°РјРё (multiple upload) + Р·Р°РїРёСЃСЊ РІ Р»РѕРєР°Р»СЊРЅСѓСЋ РєР°СЂС‚РѕС‡РєСѓ.",
            "РЈР»СѓС‡С€РµРЅР° СѓСЃС‚РѕР№С‡РёРІРѕСЃС‚СЊ СЃРµСЃСЃРёРё: РїСЂРё РєСЂР°С‚РєРѕРІСЂРµРјРµРЅРЅС‹С… 401/403 РїСЂРёР»РѕР¶РµРЅРёРµ РЅРµ РІС‹Р±СЂР°СЃС‹РІР°РµС‚ РІ Р°РІС‚РѕСЂРёР·Р°С†РёСЋ РїСЂРµР¶РґРµРІСЂРµРјРµРЅРЅРѕ.",
            "РЎРЅРёР¶РµРЅР° РЅР°РіСЂСѓР·РєР° РїСЂРё РѕС‡РµСЂРµРґРё: worker Р°РіСЂРµСЃСЃРёРІРЅРµРµ РѕС‚Р±СЂР°СЃС‹РІР°РµС‚ СѓСЃС‚Р°СЂРµРІСЂРёРµ warmup/snapshot/bidder Р·Р°РґР°С‡Рё.",
        ],
        "changes": [
            "Р‘СѓС…РіР°Р»С‚РµСЂРёСЏ: РїР°СЂР°Р»Р»РµР»СЊРЅР°СЏ Р·Р°РіСЂСѓР·РєР° РїРѕРґРјРѕРґСѓР»РµР№ Рё Р·Р°С‰РёС‚Р° РѕС‚ РЅРµРєРѕСЂСЂРµРєС‚РЅРѕРіРѕ payload РЅР° РєР»РёРµРЅС‚Рµ.",
            "РЎРµСЂРІРµСЂРЅР°СЏ СЃС‚Р°С‚РёСЃС‚РёРєР° cache-СЃР»РѕСЏ РєСЌСЂРёСЂСѓРµС‚СЃСЏ РєСЂР°С‚РєРѕ РІ РїР°РјСЏС‚Рё, С‡С‚РѕР±С‹ СЃРѕРєСЂР°С‚РёС‚СЊ РїРѕРІС‚РѕСЂРЅС‹Рµ С‚СЏР¶РµР»С‹Рµ Р°РіСЂРµРіР°С‚С‹.",
            "Р’ help/Р·Р°РіСЂСѓР·РєР°С… РѕР±РЅРѕРІР»РµРЅР° РєР°СЂС‚РѕС‡РєР° СЂРµР»РёР·Р° Рё СЃСЃС‹Р»РєР° РЅР° Р°РєС‚СѓР°Р»СЊРЅС‹Р№ APK.",
        ],
        "android_download_url": "/static/downloads/seo-wibe-mobile-latest.apk",
        "android_download_name": "SEO WIBE Android (.apk)",
        "app_entry_url": "/mobile",
        "notes": "APK 1.5.7 СѓСЃС‚Р°РЅР°РІР»РёРІР°РµС‚СЃСЏ РїРѕРІРµСЂС… РїСЂРµРґС‹РґСѓС‰РµР№ РІРµСЂСЃРёРё. РџРѕСЃР»Рµ РѕР±РЅРѕРІР»РµРЅРёСЏ РїРµСЂРµР·Р°РїСѓСЃС‚РёС‚Рµ РїСЂРёР»РѕР¶РµРЅРёРµ РґР»СЏ РїСЂРёРјРµРЅРµРЅРёСЏ РЅРѕРІРѕРіРѕ РєР»РёРµРЅС‚Р°.",
    },
    {
        "version": "0.4.0",
        "android_version_code": 11,
        "released_at": "2026-03-09",
        "current": False,
        "summary": "APK 1.5.5: РёСЃРїСЂР°РІР»РµРЅС‹ back-РЅР°РІРёРіР°С†РёСЏ С‡Р°С‚РѕРІ, РѕРЅР»Р°Р№РЅ-СЃС‚Р°С‚СѓСЃ/РїСЂРѕС„РёР»СЊ РІ Р»РёС‡РЅС‹С… С‡Р°С‚Р°С… Рё РґРѕР±Р°РІР»РµРЅС‹ РіР»РѕР±Р°Р»СЊРЅС‹Рµ РѕР±СЉСЏРІР»РµРЅРёСЏ.",
        "diff_from_previous": [
            "Android APK РѕР±РЅРѕРІР»РµРЅ РґРѕ versionCode=11 / versionName=1.5.5.",
            "Р’ С‡Р°С‚Рµ APK РёСЃРїСЂР°РІР»РµРЅР° С†РµРїРѕС‡РєР° РЅР°Р·Р°Рґ: РґРёР°Р»РѕРі в†’ С‡Р°С‚С‹ в†’ РєР°Р»РµРЅРґР°СЂСЊ в†’ СЃРїРёСЃРѕРє РјРѕРґСѓР»РµР№; РІС‹С…РѕРґ РїРѕ РґРІРѕР№РЅРѕРјСѓ back СЃРѕС…СЂР°РЅРµРЅ.",
            "Р’ Р»РёС‡РЅС‹С… С‡Р°С‚Р°С… РІРѕР·РІСЂР°С‰РµРЅ В«СЃРµР№С‡Р°СЃ РѕРЅР»Р°Р№РЅ / РїРѕСЃР»РµРґРЅРёР№ СЂР°Р· РІ СЃРµС‚РёВ» Рё РєР»РёРєР°Р±РµР»СЊРЅС‹Р№ РїСЂРѕС„РёР»СЊ СѓС‡Р°СЃС‚РЅРёРєР° РІ Р·Р°РіРѕР»РѕРІРєРµ.",
            "Р’ РѕС‚Р·С‹РІР°С…/РІРѕРїСЂРѕСЃР°С… РґР°С‚Р° РєР°СЂС‚РѕС‡РєРё С‚РµРїРµСЂСЊ РїРѕРґСЃРІРµС‡РёРІР°РµС‚СЃСЏ РїРѕ РјР°СЂРєРµС‚РїР»РµР№СЃСѓ (WB/Ozon) РІ РІРµР±Рµ Рё APK.",
            "Р”РѕР±Р°РІР»РµРЅС‹ РіР»РѕР±Р°Р»СЊРЅС‹Рµ РѕР±СЉСЏРІР»РµРЅРёСЏ (Р°РґРјРёРЅРєР° + РїРѕРєР°Р· РїРѕР»СЊР·РѕРІР°С‚РµР»СЏРј СЃ РїРѕРґС‚РІРµСЂР¶РґРµРЅРёРµРј РћРљ).",
        ],
        "changes": [
            "Р’ РєРѕР»РѕРєРѕР»СЊС‡РёРєРµ РґРѕР±Р°РІР»РµРЅ РјРѕРјРµРЅС‚Р°Р»СЊРЅС‹Р№ local-sync read-all, С‡С‚РѕР±С‹ Р±РµР№РґР¶ Р±С‹СЃС‚СЂРµРµ СЃР±СЂР°СЃС‹РІР°Р»СЃСЏ РїРѕСЃР»Рµ РїСЂРѕС‡С‚РµРЅРёСЏ.",
            "РљР°Р»РµРЅРґР°СЂСЊ РІ СЃРѕС†РёР°Р»СЊРЅРѕРј РјРѕРґСѓР»Рµ РїРѕ СѓРјРѕР»С‡Р°РЅРёСЋ РѕС‚РєСЂС‹РІР°РµС‚СЃСЏ РЅР° С‚РµРєСѓС‰РµРј РґРЅРµ С‚РµРєСѓС‰РµРіРѕ РјРµСЃСЏС†Р°.",
            "Р’ help/Р·Р°РіСЂСѓР·РєР°С… РѕР±РЅРѕРІР»РµРЅР° РєР°СЂС‚РѕС‡РєР° СЂРµР»РёР·Р° Рё СЃСЃС‹Р»РєР° РЅР° Р°РєС‚СѓР°Р»СЊРЅС‹Р№ APK-С„Р°Р№Р».",
        ],
        "android_download_url": "/static/downloads/seo-wibe-mobile-latest.apk",
        "android_download_name": "SEO WIBE Android (.apk)",
        "app_entry_url": "/mobile",
        "notes": "РџРѕСЃР»Рµ СЃРєР°С‡РёРІР°РЅРёСЏ APK 1.5.5 РёР· СЂР°Р·РґРµР»Р° В«РЎРїСЂР°РІРєР° в†’ Р—Р°РіСЂСѓР·РєРёВ» РѕС‚РєСЂРѕРµС‚СЃСЏ СЃРёСЃС‚РµРјРЅС‹Р№ СѓСЃС‚Р°РЅРѕРІС‰РёРє Android РґР»СЏ РѕР±РЅРѕРІР»РµРЅРёСЏ РїРѕРІРµСЂС… РїСЂРµРґС‹РґСѓС‰РµР№ РІРµСЂСЃРёРё.",
    },
    {
        "version": "0.3.9",
        "android_version_code": 10,
        "released_at": "2026-03-09",
        "current": False,
        "summary": "РСЃРїСЂР°РІР»РµРЅР° СѓСЃС‚Р°РЅРѕРІРєР° Android-РѕР±РЅРѕРІР»РµРЅРёСЏ, СѓСЃРєРѕСЂРµРЅС‹ РґРµС‚Р°Р»Рё С‚РѕРІР°СЂРѕРІ Рё РєСЌСЂ WB Ads enrich.",
        "diff_from_previous": [
            "APK РѕР±РЅРѕРІР»РµРЅ РґРѕ versionCode=10 / versionName=1.5.4 Рё РїРѕРґРїРёСЃР°РЅ РєРѕСЂСЂРµРєС‚РЅРѕР№ СЃС…РµРјРѕР№ v2/v3 РґР»СЏ СѓСЃС‚Р°РЅРѕРІРєРё РїРѕРІРµСЂС… РїСЂРµРґС‹РґСѓС‰РµР№ РІРµСЂСЃРёРё.",
            "Р’ С‚РѕРІР°СЂР°С… СѓР±СЂР°РЅР° РѕС‚РґРµР»СЊРЅР°СЏ РєРЅРѕРїРєР° СЂРµРґР°РєС‚РёСЂРѕРІР°РЅРёСЏ РёР· СЃС‚СЂРѕРєРё: С‚РµРїРµСЂСЊ РІС…РѕРґ РІ СЂРµРґР°РєС‚РёСЂРѕРІР°РЅРёРµ РёРґРµС‚ РёР· РѕРєРЅР° РїСЂРѕСЃРјРѕС‚СЂР°.",
            "Р’ СЂРµРґР°РєС‚РѕСЂРµ С‚РѕРІР°СЂР° РґРѕР±Р°РІР»РµРЅ drag&drop РїРѕСЂСЏРґРєР° С„РѕС‚Рѕ; РїРµСЂРІРѕРµ С„РѕС‚Рѕ СЃС‚Р°РЅРѕРІРёС‚СЃСЏ РіР»Р°РІРЅС‹Рј.",
            "Р”РµС‚Р°Р»Рё РєР°СЂС‚РѕС‡РєРё С‚РѕРІР°СЂР° Рё WB Ads enrich РїРµСЂРµРІРµРґРµРЅС‹ РЅР° Р±С‹СЃС‚СЂС‹Р№ cache-first СЂРµР¶РёРј, СЃРЅРёР¶РµРЅС‹ Р·Р°РґРµСЂР¶РєРё РѕС‚РєСЂС‹С‚РёСЏ Рё p95 API.",
        ],
        "changes": [
            "Р”РѕР±Р°РІР»РµРЅ РїСЂРёРЅСѓРґРёС‚РµР»СЊРЅС‹Р№ refresh РґРµС‚Р°Р»РµР№ С‚РѕРІР°СЂР° (В«РћР±РЅРѕРІРёС‚СЊ РёР· APIВ») Р±РµР· С‚РѕСЂРјРѕР·РѕРІ РѕР±С‹С‡РЅРѕРіРѕ РїСЂРѕСЃРјРѕС‚СЂР°.",
            "Р’ help/Р·Р°РіСЂСѓР·РєР°С… РѕР±РЅРѕРІР»РµРЅР° РєР°СЂС‚РѕС‡РєР° СЂРµР»РёР·Р° РґР»СЏ Android APK 1.5.4.",
            "Р”РѕР±Р°РІР»РµРЅС‹ РёРЅРґРµРєСЃС‹ SQLite РґР»СЏ СѓСЃРєРѕСЂРµРЅРёСЏ count/С„РёР»СЊС‚СЂРѕРІ С‚РѕРІР°СЂРѕРІ РїРѕ user/owner/last_position.",
            "РЎРѕС…СЂР°РЅРµРЅС‹ РїСѓСЂ-СѓРІРµРґРѕРјР»РµРЅРёСЏ, Р±РµР№РґР¶Рё Рё Р±С‹СЃС‚СЂС‹Р№ РѕС‚РІРµС‚ РІ Android.",
        ],
        "android_download_url": "/static/downloads/seo-wibe-mobile-latest.apk",
        "android_download_name": "SEO WIBE Android (.apk)",
        "app_entry_url": "/mobile",
        "notes": "APK 1.5.4 РёСЃРїСЂР°РІР»СЏРµС‚ СЃС†РµРЅР°СЂРёР№ В«РѕР±РЅРѕРІР»РµРЅРёРµ СЃРєР°С‡Р°Р»РѕСЃСЊ, РЅРѕ РЅРµ СѓСЃС‚Р°РЅРѕРІРёР»РѕСЃСЊВ» Рё РѕС‚РєСЂС‹РІР°РµС‚ СЃРёСЃС‚РµРјРЅС‹Р№ СѓСЃС‚Р°РЅРѕРІС‰РёРє Android РїРѕСЃР»Рµ Р·Р°РіСЂСѓР·РєРё.",
    },
    {
        "version": "0.3.8",
        "android_version_code": 9,
        "released_at": "2026-03-08",
        "current": False,
        "summary": "РЎС‚Р°Р±РёР»РёР·Р°С†РёСЏ СЃРµСЂРІРµСЂР°, СѓСЃРєРѕСЂРµРЅРёРµ СЂРµРєР»Р°РјС‹/РєСЌСЂРµР№, online-СЃС‚Р°С‚СѓСЃС‹ Рё РѕР±РЅРѕРІР»РµРЅРЅС‹Р№ Android APK 1.5.3.",
        "diff_from_previous": [
            "РСЃРїСЂР°РІР»РµРЅР° РЅРµСЃС‚Р°Р±РёР»СЊРЅРѕСЃС‚СЊ Р°РІС‚РѕРґРµРїР»РѕСЏ РЅР° СЃРµСЂРІРµСЂРµ (safe.directory + HOME РґР»СЏ systemd unit).",
            "WB Ads: РґРѕРіСЂСѓР·РєР° summary/СЃС‚Р°С‚РѕРІ СЂР°СЃС€РёСЂРµРЅР° Рё РІС‹РЅРµСЃРµРЅР° РІ РїСЂРѕРіСЂРµРІ РєСЌСЂР°/СЃРЅР°РїСЂРѕС‚РѕРІ.",
            "Р§Р°С‚С‹: В«РїРѕСЃР»РµРґРЅРёР№ СЂР°Р· РІ СЃРµС‚РёВ» Рё В«СЃРµР№С‡Р°СЃ РѕРЅР»Р°Р№РЅВ» СЃС‡РёС‚Р°СЋС‚ Р°РєС‚РёРІРЅРѕСЃС‚СЊ РїРѕ actor/member РєРѕСЂСЂРµРєС‚РЅРѕ.",
            "Android: РґРѕСЂР°Р±РѕС‚Р°РЅР° back-РЅР°РІРёРіР°С†РёСЏ (РІРЅСѓС‚СЂРµРЅРЅСЏСЏ РёСЃС‚РѕСЂРёСЏ + РґРІРѕР№РЅРѕР№ back РґР»СЏ РІС‹С…РѕРґР°).",
        ],
        "changes": [
            "РЎР±РѕСЂРєР° Android РѕР±РЅРѕРІР»РµРЅР° РґРѕ versionCode=9 / versionName=1.5.3.",
            "РЎРЅРёР¶РµРЅ РІРёР·СѓР°Р»СЊРЅС‹Р№ flicker Р°РІР°С‚Р°СЂРѕРІ/РёРєРѕРЅРѕРє Р·Р° СЃС‡РµС‚ СЃС‚Р°Р±РёР»СЊРЅРѕРіРѕ СЂРµРЅРґРµСЂР° Рё cache-control РґР»СЏ СЃС‚Р°С‚РёРєРё.",
            "Р РµР»РёР· РІРєР»СЋС‡Р°РµС‚ С†РІРµС‚РѕРІС‹Рµ СѓР»СѓС‡СЂРµРЅРёСЏ Р±СѓС…РіР°Р»С‚РµСЂРёРё/РѕС‚Р·С‹РІРѕРІ Рё РјРѕР±РёР»СЊРЅСѓСЋ РєРѕРјРїРѕРЅРѕРІРєСѓ РіСЂР°С„РёРєР° СЃС‚Р°С‚РёСЃС‚РёРєРё.",
            "РџСЂРѕРІРµСЂРєР° РѕР±РЅРѕРІР»РµРЅРёР№ APK РѕСЃС‚Р°РµС‚СЃСЏ С‡РµСЂРµР· /api/mobile/apk/latest РЅР° СЃРµСЂРІРµСЂРµ 5.129.207.106:8016.",
            "РЎРѕС…СЂР°РЅРµРЅС‹ push-СѓРІРµРґРѕРјР»РµРЅРёСЏ, Р±РµР№РґР¶Рё Рё Р±С‹СЃС‚СЂС‹Р№ РѕС‚РІРµС‚ РёР· СЂС‚РѕСЂРєРё Android.",
        ],
        "android_download_url": "/static/downloads/seo-wibe-mobile-latest.apk",
        "android_download_name": "SEO WIBE Android (.apk)",
        "app_entry_url": "/mobile",
        "notes": "APK 1.5.3 РїСЂРѕРІРµСЂСЏРµС‚ РѕР±РЅРѕРІР»РµРЅРёСЏ РїСЂРё Р·Р°РїСѓСЃРєРµ Рё РїСЂРµРґР»Р°РіР°РµС‚ В«РЈСЃС‚Р°РЅРѕРІРёС‚СЊ / РџРѕР·Р¶РµВ». РџРѕСЃР»Рµ Р·Р°РіСЂСѓР·РєРё РѕС‚РєСЂС‹РІР°РµС‚СЃСЏ СЃРёСЃС‚РµРјРЅС‹Р№ СѓСЃС‚Р°РЅРѕРІС‰РёРє Android.",
    },
    {
        "version": "0.3.7",
        "android_version_code": 8,
        "released_at": "2026-03-08",
        "current": False,
        "summary": "Android APK: С‡Р°С‚ РїРµСЂРµСЂР°Р±РѕС‚Р°РЅ РІ Telegram-РїРѕРґРѕР±РЅС‹Р№ mobile layout, РёСЃРїСЂР°РІР»РµРЅ РІРѕР·РІСЂР°С‚ РёР· РґРёР°Р»РѕРіР°, РєР°Р»РµРЅРґР°СЂСЊ СЃС‚Р°СЂС‚СѓРµС‚ РїРѕ СѓРјРѕР»С‡Р°РЅРёСЋ.",
        "diff_from_previous": [
            "APK Chat UI: РґРѕР±Р°РІР»РµРЅ РєРѕРјРїР°РєС‚РЅС‹Р№ РІРµСЂС…РЅРёР№ Р±Р»РѕРє (back + РёРјСЏ/С‚РёРї РґРёР°Р»РѕРіР°), СѓР±СЂР°РЅР° Р»РёС€РЅСЏСЏ РєРЅРѕРїРєР° refresh, С‡Р°С‚ Р·Р°РЅРёРјР°РµС‚ Р±РѕР»СЊС€Рµ РїРѕР»РµР·РЅРѕР№ РІС‹СЃРѕС‚С‹.",
            "РСЃРїСЂР°РІР»РµРЅ Р±Р°Рі Р°РІС‚РѕРїРµСЂРµРѕС‚РєСЂС‹С‚РёСЏ Р»РёС‡РЅРѕРіРѕ С‡Р°С‚Р° РїРѕСЃР»Рµ РІРѕР·РІСЂР°С‚Р° Рє СЃРїРёСЃРєСѓ РґРёР°Р»РѕРіРѕРІ.",
            "РЎРѕРѕР±С‰РµРЅРёСЏ Р·Р°РіСЂСѓР¶Р°СЋС‚СЃСЏ СѓСЃС‚РѕР№С‡РёРІРµРµ: СѓРІРµР»РёС‡РµРЅС‹ СЂРµС‚СЂР°Рё РїСЂРё РІСЂРµРјРµРЅРЅРѕ РїСѓСЃС‚РѕРј РѕС‚РІРµС‚Рµ API.",
            "Р”Р»СЏ APK СЃРЅРёР¶РµРЅР° РЅР°РІСЏР·С‡РёРІРѕСЃС‚СЊ onboarding-РїРѕРґСЃРєР°Р·РєРё РїСЂРѕ API-РєР»СЋС‡Рё, СЃС‚Р°СЂС‚РѕРІС‹Р№ РїРѕРґРјРѕРґСѓР»СЊ Social РёР·РјРµРЅРµРЅ РЅР° В«РљР°Р»РµРЅРґР°СЂСЊВ».",
        ],
        "changes": [
            "РЎР±РѕСЂРєР° Android РѕР±РЅРѕРІР»РµРЅР° РґРѕ versionCode=8 / versionName=1.5.2.",
            "РџСЂРѕРІРµСЂРєР° РѕР±РЅРѕРІР»РµРЅРёР№ APK РїСЂРѕРґРѕР»Р¶Р°РµС‚ СЂР°Р±РѕС‚Р°С‚СЊ С‡РµСЂРµР· /api/mobile/apk/latest РЅР° РЅРѕРІРѕРј СЃРµСЂРІРµСЂРµ 5.129.207.106:8016.",
            "Mobile chat chrome СЃРёРЅС…СЂРѕРЅРёР·РёСЂРѕРІР°РЅ СЃ С‚РµРєСѓС‰РёРј РґРёР°Р»РѕРіРѕРј Рё СЃРєСЂС‹РІР°РµС‚СЃСЏ РІРЅРµ РѕС‚РєСЂС‹С‚РѕРіРѕ С‡Р°С‚Р°.",
            "Р’ СЂРµР¶РёРјРµ APK РѕС‚РєР»СЋС‡РµРЅ Р°РІС‚Рѕ-РІС‹Р±РѕСЂ С‚СЂРµРґР° РїРѕСЃР»Рµ СЂСѓС‡РЅРѕРіРѕ РІРѕР·РІСЂР°С‚Р°, С‡С‚РѕР±С‹ СЃРїРёСЃРѕРє РґРёР°Р»РѕРіРѕРІ РѕСЃС‚Р°РІР°Р»СЃСЏ СЃС‚Р°Р±РёР»СЊРЅС‹Рј.",
            "РЎРѕС…СЂР°РЅРµРЅ Р°РІС‚РѕР°РїРґРµР№С‚ APK С‡РµСЂРµР· DownloadManager Рё СѓРІРµРґРѕРјР»РµРЅРёСЏ СЃ Р±С‹СЃС‚СЂС‹Рј РѕС‚РІРµС‚РѕРј.",
        ],
        "android_download_url": "/static/downloads/seo-wibe-mobile-latest.apk",
        "android_download_name": "SEO WIBE Android (.apk)",
        "app_entry_url": "/mobile",
        "notes": "APK РїСЂРѕРІРµСЂСЏРµС‚ РѕР±РЅРѕРІР»РµРЅРёСЏ РїСЂРё Р·Р°РїСѓСЃРєРµ Рё РїСЂРµРґР»Р°РіР°РµС‚ В«РЈСЃС‚Р°РЅРѕРІРёС‚СЊ / РџРѕР·Р¶РµВ». РџРѕСЃР»Рµ Р·Р°РіСЂСѓР·РєРё РѕС‚РєСЂС‹РІР°РµС‚СЃСЏ СЃРёСЃС‚РµРјРЅС‹Р№ СѓСЃС‚Р°РЅРѕРІС‰РёРє Android. РЈРІРµРґРѕРјР»РµРЅРёСЏ РІ С€С‚РѕСЂРєРµ РїРѕРґРґРµСЂР¶РёРІР°СЋС‚ Р±С‹СЃС‚СЂС‹Р№ РѕС‚РІРµС‚.",
    },
    {
        "version": "0.3.0",
        "android_version_code": 1,
        "released_at": "2026-03-07",
        "current": False,
        "summary": "РЎС‚Р°Р±РёР»РёР·Р°С†РёСЏ СЂРµРєР»Р°РјС‹/СЃС‚Р°С‚РёСЃС‚РёРєРё, Р±СѓС…РіР°Р»С‚РµСЂРёСЏ РЅР° РѕР±С‰РµР№ Р±Р°Р·Рµ С‚РѕРІР°СЂРѕРІ Рё Android-РїСЂРёР»РѕР¶РµРЅРёРµ.",
        "diff_from_previous": [
            "РСЃРїСЂР°РІР»РµРЅР° РґРѕРіСЂСѓР·РєР° WB Ads: С‡Р°СЃС‚РёС‡РЅС‹Рµ РѕС‚РІРµС‚С‹ API Р±РѕР»СЊС€Рµ РЅРµ СЃС‡РёС‚Р°СЋС‚СЃСЏ С„Р°С‚Р°Р»СЊРЅРѕР№ РѕСЂРёР±РєРѕР№.",
            "РЎС‚Р°С‚РёСЃС‚РёРєР° РїСЂРѕРґР°Р¶ РїРµСЂРµСЂР°Р±РѕС‚Р°РЅР°: РєРѕРјРїР°РєС‚РЅС‹Р№ РІРµСЂС… KPI Рё РІС‹РЅРµСЃРµРЅРЅС‹Рµ РІРЅРёР· РґРѕРїРѕР»РЅРёС‚РµР»СЊРЅС‹Рµ РјРµС‚СЂРёРєРё.",
            "РЎРЅРёР¶РµРЅ СЂРёСЃРє РґРІРѕР№РЅРѕРіРѕ СѓС‡РµС‚Р° WB/Ozon РґРѕС…РѕРґРѕРІ РІ РїСЂРѕРґР°Р¶Р°С… Рё С„РёРЅР°РЅСЃР°С….",
            "Р”РѕР±Р°РІР»РµРЅРѕ Android-РїСЂРёР»РѕР¶РµРЅРёРµ: С‡Р°С‚ РїРѕ СѓРјРѕР»С‡Р°РЅРёСЋ, Р±С‹СЃС‚СЂС‹Р№ РїРµСЂРµРєР»СЋС‡Р°С‚РµР»СЊ СЂР°Р·РґРµР»РѕРІ, РїСЂРѕС„РёР»СЊ, СЏР·С‹Рє Рё СѓРІРµРґРѕРјР»РµРЅРёСЏ.",
        ],
        "changes": [
            "WB/Ozon sales: СѓР»СѓС‡СЂРµРЅР° РґРµРґСѓРїР»РёРєР°С†РёСЏ, РїРµСЂРµСЃС‡РёС‚Р°РЅС‹ РѕС‚РјРµРЅС‹/РІРѕР·РІСЂР°С‚С‹ Рё С„РёРЅРїРѕС‚РѕРєРё.",
            "Help Center: РґРѕР±Р°РІР»РµРЅ СЂР°Р·РґРµР» В«Р—Р°РіСЂСѓР·РєРёВ» СЃ РёСЃС‚РѕСЂРёРµР№ РІРµСЂСЃРёР№ Рё СЃСЃС‹Р»РєРѕР№ РЅР° Android APK.",
            "Р‘СѓС…РіР°Р»С‚РµСЂРёСЏ РёСЃРїРѕР»СЊР·СѓРµС‚ РµРґРёРЅСѓСЋ Р±Р°Р·Сѓ С‚РѕРІР°СЂРѕРІ (Product) Рё Р·Р°РєСѓРїРѕС‡РЅС‹С… С†РµРЅ Р±РµР· РґСѓР±Р»РёСЂСѓСЋС‰РµРіРѕ С…СЂР°РЅРёР»РёС‰Р°.",
        ],
        "android_download_url": "/static/downloads/seo-wibe-mobile-latest.apk",
        "android_download_name": "SEO WIBE Mobile Android (.apk)",
        "app_entry_url": "/mobile",
        "notes": "РЈСЃС‚Р°РЅРѕРІРєР° РЅР° Android: СЃРєР°С‡Р°Р№С‚Рµ .apk, РѕС‚РєСЂРѕР№С‚Рµ С„Р°Р№Р» Рё РїРѕРґС‚РІРµСЂРґРёС‚Рµ СѓСЃС‚Р°РЅРѕРІРєСѓ РёР· РЅРµРёР·РІРµСЃС‚РЅРѕРіРѕ РёСЃС‚РѕС‡РЅРёРєР°.",
    },
    {
        "version": "0.2.1",
        "android_version_code": 1,
        "released_at": "2026-03-06",
        "current": False,
        "summary": "РњРѕРґСѓР»СЊ Р±СѓС…РіР°Р»С‚РµСЂРёРё (РѕР±Р·РѕСЂ, Р°РЅР°Р»РёР·, СЂР°СЃС…РѕРґС‹, РЅР°СЃС‚СЂРѕР№РєРё, Excel РёРјРїРѕСЂС‚/СЌРєСЃРїРѕСЂС‚ Р·Р°РєСѓРїРѕС‡РЅС‹С… С†РµРЅ).",
        "diff_from_previous": [
            "Р”РѕР±Р°РІР»РµРЅ РїРѕР»РЅРѕС†РµРЅРЅС‹Р№ РјРѕРґСѓР»СЊ Р±СѓС…РіР°Р»С‚РµСЂРёРё РІ РјРѕРґСѓР»СЊРЅСѓСЋ СЃРёСЃС‚РµРјСѓ Рё РїСЂР°РІР° РґРѕСЃС‚СѓРїР°.",
            "Р РµР°Р»РёР·РѕРІР°РЅС‹ СЂР°СЃС‡РµС‚РЅС‹Рµ РїР°СЂР°РјРµС‚СЂС‹ РќР”РЎ/РЅР°Р»РѕРіРѕРІ Рё РїРѕР»СЊР·РѕРІР°С‚РµР»СЊСЃРєРёС… СЂР°СЃС…РѕРґРѕРІ.",
        ],
        "changes": [
            "РўР°Р±Р»РёС‡РЅС‹Р№ Р°РЅР°Р»РёР· РїСЂРёР±С‹Р»СЊРЅРѕСЃС‚Рё РїРѕ SKU/Р°СЂС‚РёРєСѓР»Сѓ.",
            "Р­РєСЃРїРѕСЂС‚ Рё РёРјРїРѕСЂС‚ Р·Р°РєСѓРїРѕС‡РЅС‹С… С†РµРЅ С‡РµСЂРµР· Excel/CSV.",
        ],
        "android_download_url": "/static/downloads/seo-wibe-mobile-latest.apk",
        "android_download_name": "SEO WIBE Mobile Android (.apk)",
        "app_entry_url": "/mobile",
        "notes": "",
    },
    {
        "version": "0.2.0",
        "android_version_code": 1,
        "released_at": "2026-03-05",
        "current": False,
        "summary": "РЎС‚Р°Р±РёР»РёР·Р°С†РёСЏ РјРѕРґСѓР»РµР№ С‚РѕРІР°СЂРѕРІ, РѕС‚Р·С‹РІРѕРІ/РІРѕРїСЂРѕСЃРѕРІ, СЂРµРєР»Р°РјС‹ Рё С‡Р°С‚Р°.",
        "diff_from_previous": [
            "Р”РѕР±Р°РІР»РµРЅС‹ bulk-РѕРїРµСЂР°С†РёРё РІ С‚РѕРІР°СЂР°С… Рё СѓР»СѓС‡СЂРµРЅРёСЏ РІ СЂРµРєР»Р°РјРЅРѕРј enrich.",
            "РЈСЃРєРѕСЂРµРЅС‹ СЃС†РµРЅР°СЂРёРё РіРµРЅРµСЂР°С†РёРё Рё РѕС‚РїСЂР°РІРєРё РѕС‚РІРµС‚РѕРІ.",
        ],
        "changes": [
            "РСЃРїСЂР°РІР»РµРЅС‹ РјРѕР±РёР»СЊРЅС‹Рµ UI-СЂРµРіСЂРµСЃСЃРёРё Рё РѕР±СЂР°Р±РѕС‚РєР° direct-С‡Р°С‚РѕРІ.",
            "РЈР»СѓС‡С€РµРЅС‹ СЃС‚Р°С‚СѓСЃС‹ Р·Р°РіСЂСѓР·РєРё СЂРµРєР»Р°РјРЅС‹С… РєР°РјРїР°РЅРёР№.",
        ],
        "android_download_url": "/static/downloads/seo-wibe-mobile-latest.apk",
        "android_download_name": "SEO WIBE Mobile Android (.apk)",
        "app_entry_url": "/mobile",
        "notes": "",
    },
]


@router.post("/auth/register", response_model=TokenResponse)
def register(payload: RegisterRequest, request: Request, response: Response, db: Session = Depends(get_db)):
    email = payload.email.strip().lower()
    exists = db.scalar(select(User).where(User.email == email))
    if exists:
        raise HTTPException(status_code=400, detail="РџРѕР»СЊР·РѕРІР°С‚РµР»СЊ СѓР¶Рµ СЃСѓС‰РµСЃС‚РІСѓРµС‚")
    member_exists = db.scalar(select(TeamMember).where(TeamMember.email == email))
    if member_exists:
        raise HTTPException(status_code=400, detail="Email СѓР¶Рµ РёСЃРїРѕР»СЊР·СѓРµС‚СЃСЏ СЃРѕС‚СЂСѓРґРЅРёРєРѕРј")

    users_count = db.scalar(select(func.count()).select_from(User)) or 0
    role = "admin" if users_count == 0 else "client"
    user = User(email=email, hashed_password=get_password_hash(payload.password), role=role)
    db.add(user)
    db.flush()
    db.add(
        TeamMember(
            user_id=user.id,
            email=user.email,
            full_name="",
            nickname="owner",
            hashed_password=get_password_hash(payload.password),
            access_scope=json.dumps(["*"], ensure_ascii=False),
            is_owner=True,
            is_active=True,
        )
    )

    for module_code in DEFAULT_MODULES:
        db.add(
            ModuleAccess(
                user_id=user.id,
                module_code=module_code,
                enabled=(module_code not in DISABLED_BY_DEFAULT_MODULES),
            )
        )

    _audit(
        db,
        user,
        action="user_registered",
        details=f"role={role}",
        module_code="auth",
        entity_type="user",
        entity_id=str(user.id),
        status="ok",
        request=request,
    )
    db.commit()

    token = create_access_token(f"u:{user.id}")
    response.set_cookie("seo_wibe_token", token, httponly=True, samesite="lax", path="/")
    return TokenResponse(access_token=token)


@router.post("/auth/login", response_model=TokenResponse)
def login(payload: LoginRequest, request: Request, response: Response, db: Session = Depends(get_db)):
    email = payload.email.strip().lower()
    user = db.scalar(select(User).where(User.email == email))
    if user and verify_password(payload.password, user.hashed_password):
        _audit(
            db,
            user,
            action="auth_login_success",
            details="kind=owner",
            module_code="auth",
            entity_type="user",
            entity_id=str(user.id),
            status="ok",
            request=request,
        )
        db.commit()
        token = create_access_token(f"u:{user.id}")
        response.set_cookie("seo_wibe_token", token, httponly=True, samesite="lax", path="/")
        return TokenResponse(access_token=token)

    member = db.scalar(
        select(TeamMember)
        .where(
            TeamMember.email == email,
            TeamMember.is_active.is_(True),
        )
        .order_by(TeamMember.id.desc())
    )
    if member and member.hashed_password and verify_password(payload.password, member.hashed_password):
        owner = db.get(User, member.user_id)
        if owner:
            owner._actor_email = member.email
            owner._actor_member_id = member.id
            owner._actor_is_owner = bool(member.is_owner)
            _audit(
                db,
                owner,
                action="auth_login_success",
                details=f"kind=team_member;member_id={member.id}",
                module_code="auth",
                entity_type="team_member",
                entity_id=str(member.id),
                status="ok",
                request=request,
            )
            db.commit()
            token = create_access_token(f"m:{member.id}")
            response.set_cookie("seo_wibe_token", token, httponly=True, samesite="lax", path="/")
            return TokenResponse(access_token=token)
    _audit(
        db,
        None,
        action="auth_login_failed",
        details=f"email={email}",
        module_code="auth",
        entity_type="user",
        entity_id=email,
        status="error",
        request=request,
    )
    db.commit()
    raise HTTPException(status_code=401, detail="РќРµРІРµСЂРЅС‹Р№ email РёР»Рё РїР°СЂРѕР»СЊ")


@router.post("/auth/logout", response_model=MessageOut)
def logout(request: Request, response: Response, user: User = Depends(get_current_user), db: Session = Depends(get_db)):
    _audit(
        db,
        user,
        action="auth_logout",
        details="kind=owner" if _actor_is_owner(user) else f"kind=team_member;member_id={_actor_member_id(user)}",
        module_code="auth",
        entity_type="user",
        entity_id=str(user.id),
        status="ok",
        request=request,
    )
    db.commit()
    response.delete_cookie("seo_wibe_token", path="/")
    response.delete_cookie("seo_wibe_token", path="/api/auth")
    return MessageOut(message="Р’С‹С…РѕРґ РІС‹РїРѕР»РЅРµРЅ")


@router.get("/auth/me", response_model=UserOut)
def me(request: Request, user: User = Depends(get_current_user), db: Session = Depends(get_db)):
    actor_key, actor_nick, actor_member_id = _social_actor_identity(db, user)
    avatar_url = ""
    if actor_key.startswith("m:") and actor_member_id:
        member = db.get(TeamMember, int(actor_member_id))
        avatar_url = str(member.avatar_url or "") if member else ""
    else:
        owner_member = db.scalar(
            select(TeamMember).where(
                TeamMember.user_id == user.id,
                TeamMember.is_owner.is_(True),
            ).order_by(TeamMember.id.asc())
        )
        avatar_url = str(owner_member.avatar_url or "") if owner_member else ""
    _audit(
        db,
        user,
        action="auth_session_check",
        details="session=active",
        module_code="auth",
        entity_type="user",
        entity_id=str(user.id),
        status="ok",
        request=request,
    )
    db.commit()
    payload = UserOut.model_validate(user)
    payload.actor_email = (_actor_email(user) or str(payload.email or "").strip().lower()) or None
    payload.actor_key = actor_key
    payload.actor_nick = actor_nick
    payload.actor_is_owner = bool(_actor_is_owner(user))
    payload.actor_member_id = int(actor_member_id or 0) if actor_member_id else None
    payload.avatar_url = avatar_url or None
    return payload


@router.get("/modules/current", response_model=list[CurrentModuleOut])
def current_modules(request: Request, user: User = Depends(get_current_user), db: Session = Depends(get_db)):
    rows = db.scalars(select(ModuleAccess).where(ModuleAccess.user_id == user.id)).all()
    out: list[CurrentModuleOut] = []
    enabled_count = 0
    for row in rows:
        effective_enabled = _module_enabled_for_context(db, user, row.module_code, bool(row.enabled))
        allowed = effective_enabled and _actor_can_use_module(user, row.module_code)
        if allowed:
            enabled_count += 1
        out.append(CurrentModuleOut(module_code=row.module_code, enabled=allowed))
    _audit(
        db,
        user,
        action="ui_modules_loaded",
        details=f"enabled={enabled_count};total={len(rows)}",
        module_code="auth",
        entity_type="module_access",
        status="ok",
        request=request,
    )
    db.commit()
    return out


@router.get("/ui/settings", response_model=UiSettingsOut)
def ui_settings(user: User = Depends(get_current_user), db: Session = Depends(get_db)):
    _ = user
    return _get_ui_settings(db)


@router.post("/activity/track", response_model=MessageOut)
def track_activity(
    payload: ActivityTrackIn,
    request: Request,
    user: User = Depends(get_current_user),
    db: Session = Depends(get_db),
):
    action = re.sub(r"[^a-z0-9_-]+", "", str(payload.action or "").strip().lower())[:120]
    if not action:
        raise HTTPException(status_code=400, detail="action is required")
    module_code = re.sub(r"[^a-z0-9_-]+", "", str(payload.module_code or "").strip().lower())[:80]
    entity_type = re.sub(r"[^a-z0-9_-]+", "", str(payload.entity_type or "").strip().lower())[:80]
    status = str(payload.status or "ok").strip().lower()
    if status not in {"ok", "error", "warn"}:
        status = "ok"
    _audit(
        db,
        user,
        action=action,
        details=str(payload.details or "")[:5000],
        module_code=module_code,
        entity_type=entity_type,
        entity_id=str(payload.entity_id or "")[:120],
        status=status,
        request=request,
    )
    db.commit()
    return MessageOut(message="tracked")


@router.post("/credentials", response_model=ApiCredentialOut)
def save_credential(payload: ApiCredentialIn, user: User = Depends(get_current_user), db: Session = Depends(get_db)):
    _require_owner_actor(user)
    marketplace = validate_marketplace(payload.marketplace)

    creds = db.scalars(
        select(ApiCredential)
        .where(ApiCredential.user_id == user.id, ApiCredential.marketplace == marketplace)
        .order_by(ApiCredential.id.desc())
    ).all()
    if creds:
        cred = creds[0]
        cred.api_key = payload.api_key
        cred.active = True
        for stale in creds[1:]:
            stale.active = False
    else:
        cred = ApiCredential(user_id=user.id, marketplace=marketplace, api_key=payload.api_key, active=True)
        db.add(cred)

    _audit(
        db,
        user,
        action="credential_saved",
        details=f"marketplace={marketplace}",
        module_code="user_profile",
        entity_type="api_credential",
        entity_id=marketplace,
    )
    db.commit()
    return ApiCredentialOut(id=cred.id, marketplace=cred.marketplace, api_key_masked=mask_key(cred.api_key), active=cred.active)


@router.get("/credentials", response_model=list[ApiCredentialOut])
def list_credentials(user: User = Depends(get_current_user), db: Session = Depends(get_db)):
    creds = db.scalars(
        select(ApiCredential).where(ApiCredential.user_id == user.id).order_by(ApiCredential.id.desc())
    ).all()
    return [ApiCredentialOut(id=c.id, marketplace=c.marketplace, api_key_masked=mask_key(c.api_key), active=c.active) for c in creds]


@router.post("/credentials/test", response_model=CredentialTestOut)
def test_credential(payload: ApiCredentialIn, user: User = Depends(get_current_user), db: Session = Depends(get_db)):
    _require_owner_actor(user)
    marketplace = validate_marketplace(payload.marketplace)
    ok, message = test_marketplace_credentials(marketplace, payload.api_key)
    _audit(
        db,
        user,
        action="credential_tested",
        details=f"marketplace={marketplace};ok={ok}",
        module_code="user_profile",
        entity_type="api_credential",
        entity_id=marketplace,
        status="ok" if ok else "error",
    )
    db.commit()
    return CredentialTestOut(ok=ok, message=message)


@router.delete("/credentials/{marketplace}", response_model=MessageOut)
def delete_credential(marketplace: str, user: User = Depends(get_current_user), db: Session = Depends(get_db)):
    _require_owner_actor(user)
    market = validate_marketplace(marketplace)
    creds = db.scalars(
        select(ApiCredential).where(ApiCredential.user_id == user.id, ApiCredential.marketplace == market)
    ).all()
    if not creds:
        raise HTTPException(status_code=404, detail="РљР»СЋС‡ РЅРµ РЅР°Р№РґРµРЅ")

    for cred in creds:
        db.delete(cred)
    _audit(
        db,
        user,
        action="credential_deleted",
        details=f"marketplace={market}",
        module_code="user_profile",
        entity_type="api_credential",
        entity_id=market,
    )
    db.commit()
    return MessageOut(message="РљР»СЋС‡ СѓРґР°Р»РµРЅ")


@router.get("/keywords", response_model=list[KeywordOut])
def list_keywords(user: User = Depends(get_current_user), db: Session = Depends(get_db)):
    return db.scalars(
        select(UserKeyword)
        .where(
            UserKeyword.user_id == user.id,
            _owned_by_actor_or_owner_filter(UserKeyword, user),
        )
        .order_by(UserKeyword.id.desc())
    ).all()


@router.post("/keywords", response_model=KeywordOut)
def add_keyword(payload: KeywordIn, user: User = Depends(get_current_user), db: Session = Depends(get_db)):
    marketplace = payload.marketplace.strip().lower()
    if marketplace not in {"all", "wb", "ozon"}:
        raise HTTPException(status_code=400, detail="marketplace РґРѕР»Р¶РµРЅ Р±С‹С‚СЊ all, wb РёР»Рё ozon")
    keyword = payload.keyword.strip()
    if not keyword:
        raise HTTPException(status_code=400, detail="РљР»СЋС‡РµРІРѕРµ СЃР»РѕРІРѕ РЅРµ РґРѕР»Р¶РЅРѕ Р±С‹С‚СЊ РїСѓСЃС‚С‹Рј")

    exists = db.scalar(
        select(UserKeyword).where(
            UserKeyword.user_id == user.id,
            UserKeyword.marketplace == marketplace,
            UserKeyword.keyword == keyword,
            _owned_by_actor_or_owner_filter(UserKeyword, user),
        )
    )
    if exists:
        return exists

    row = UserKeyword(user_id=user.id, marketplace=marketplace, keyword=keyword)
    _assign_owner_member(row, _resolve_owner_member_id(db, user))
    db.add(row)
    _audit(
        db,
        user,
        action="keyword_added",
        details=f"marketplace={marketplace};keyword={keyword}",
        module_code="seo_generation",
        entity_type="keyword",
    )
    db.commit()
    db.refresh(row)
    return row


@router.delete("/keywords/{keyword_id}", response_model=MessageOut)
def delete_keyword(keyword_id: int, user: User = Depends(get_current_user), db: Session = Depends(get_db)):
    row = db.scalar(
        select(UserKeyword).where(
            UserKeyword.id == keyword_id,
            UserKeyword.user_id == user.id,
            _owned_by_actor_or_owner_filter(UserKeyword, user),
        )
    )
    if not row:
        raise HTTPException(status_code=404, detail="РљР»СЋС‡ РЅРµ РЅР°Р№РґРµРЅ")
    db.delete(row)
    _audit(
        db,
        user,
        action="keyword_deleted",
        details=f"id={keyword_id}",
        module_code="seo_generation",
        entity_type="keyword",
        entity_id=str(keyword_id),
    )
    db.commit()
    return MessageOut(message="РљР»СЋС‡ СѓРґР°Р»РµРЅ")


@router.get("/wb/reviews", response_model=WbReviewsOut)
def wb_reviews(
    stars: int | None = None,
    date_from: date | None = None,
    date_to: date | None = None,
    fast: bool = False,
    user: User = Depends(get_current_user),
    db: Session = Depends(get_db),
):
    ensure_module_enabled(db, user, "wb_reviews_ai")
    wb_key = _get_active_marketplace_api_key(db, user.id, "wb")
    if not wb_key:
        raise HTTPException(status_code=400, detail="РЎРЅР°С‡Р°Р»Р° СЃРѕС…СЂР°РЅРёС‚Рµ API РєР»СЋС‡ РґР»СЏ wb")
    if stars is not None and (stars < 1 or stars > 5):
        raise HTTPException(status_code=400, detail="stars РґРѕР»Р¶РµРЅ Р±С‹С‚СЊ РѕС‚ 1 РґРѕ 5")
    left = date_from.isoformat() if date_from else ""
    right = date_to.isoformat() if date_to else ""
    cache_key = build_market_cache_key(
        {
            "kind": "reviews",
            "stars": int(stars or 0),
            "date_from": left,
            "date_to": right,
            "fast": int(bool(fast)),
            "key_rev": _secret_revision(wb_key),
        }
    )

    def _load_reviews_payload() -> dict[str, Any]:
        if fast:
            return fetch_wb_reviews_fast(
                wb_key,
                stars=stars,
                date_from=left or None,
                date_to=right or None,
            )
        return fetch_wb_reviews(
            wb_key,
            stars=stars,
            date_from=left or None,
            date_to=right or None,
            max_pages=8,
        )

    data, cache_meta = get_or_refresh_market_cache(
        db,
        user_id=int(user.id),
        module_code="wb_reviews_ai",
        marketplace="wb",
        cache_key=cache_key,
        ttl_sec=_market_cache_ttl("wb_reviews_ai", fast_mode=fast),
        fetcher=_load_reviews_payload,
        stale_if_error_sec=20 * 60,
    )
    new_rows = _filter_claimed_feedback_rows(
        db,
        user,
        module_code="wb_reviews_ai",
        marketplace="wb",
        item_type="review",
        rows=list(data.get("new") or []),
    )
    answered_rows = _filter_claimed_feedback_rows(
        db,
        user,
        module_code="wb_reviews_ai",
        marketplace="wb",
        item_type="review",
        rows=list(data.get("answered") or []),
    )
    if not new_rows and not answered_rows:
        ok, message = probe_wb_feedback_access(wb_key, feedback_kind="reviews")
        if not ok:
            raise HTTPException(status_code=400, detail=message)
    _audit(
        db,
        user,
        action="wb_reviews_read",
        details=f"new={len(new_rows)};answered={len(answered_rows)};source={cache_meta.get('source')};age={cache_meta.get('age_sec')}",
        module_code="wb_reviews_ai",
        entity_type="review",
    )
    db.commit()
    return WbReviewsOut(new=new_rows, answered=answered_rows)


@router.post("/wb/reviews/reply", response_model=WbReviewReplyOut)
def wb_reply_review(payload: WbReviewReplyIn, user: User = Depends(get_current_user), db: Session = Depends(get_db)):
    ensure_module_enabled(db, user, "wb_reviews_ai")
    wb_key = _get_active_marketplace_api_key(db, user.id, "wb")
    if not wb_key:
        raise HTTPException(status_code=400, detail="РЎРЅР°С‡Р°Р»Р° СЃРѕС…СЂР°РЅРёС‚Рµ API РєР»СЋС‡ РґР»СЏ wb")
    _claim_or_validate_work_item(
        db,
        user,
        module_code="wb_reviews_ai",
        marketplace="wb",
        item_type="review",
        item_external_id=str(payload.id or "").strip(),
    )
    ok, message = post_wb_review_reply(wb_key, payload.id, payload.text)
    detail_payload = {
        "review_id": str(payload.id or ""),
        "ok": bool(ok),
        "marketplace": "wb",
        "reply": str(payload.text or "")[:800],
    }
    _audit(
        db,
        user,
        action="wb_review_reply",
        details=json.dumps(detail_payload, ensure_ascii=False),
        module_code="wb_reviews_ai",
        entity_type="review",
        entity_id=str(payload.id or ""),
        status="ok" if ok else "error",
    )
    db.commit()
    if not ok:
        raise HTTPException(status_code=400, detail=message)
    return WbReviewReplyOut(ok=True, message=message)


@router.post("/wb/reviews/generate-reply", response_model=GenerateReviewReplyOut)
def wb_generate_reply(payload: GenerateReviewReplyIn, user: User = Depends(get_current_user), db: Session = Depends(get_db)):
    ensure_module_enabled(db, user, "wb_reviews_ai")
    settings_row = _get_or_create_ai_settings(db, user.id)
    runtime = _resolve_user_ai_runtime(db, user.id)
    knowledge_ctx = _build_user_knowledge_context(
        db,
        user.id,
        query_text=f"{payload.product_name} {payload.review_text} {payload.reviewer_name}",
    )
    prompt = _compose_ai_prompt(settings_row.prompt if settings_row and settings_row.prompt else "", knowledge_ctx, content_kind="review")
    ai_trace: dict[str, Any] = {}
    reply = generate_review_reply(
        review_text=payload.review_text,
        product_name=payload.product_name,
        stars=payload.stars,
        prompt=prompt,
        reviewer_name=payload.reviewer_name,
        marketplace="wb",
        content_kind="review",
        api_key=str(runtime.get("api_key") or ""),
        model=str(runtime.get("model") or ""),
        provider=str(runtime.get("provider") or ""),
        base_url=str(runtime.get("base_url") or ""),
        fallback_chain=list(runtime.get("fallback_chain") or []),
        trace=ai_trace,
    )
    trace_attempts = [x for x in (ai_trace.get("attempts") or []) if isinstance(x, dict)]
    details_payload = {
        "provider": ai_trace.get("used_provider") or runtime.get("provider") or "builtin",
        "model": ai_trace.get("used_model") or runtime.get("model") or settings.openai_model,
        "mode": ai_trace.get("used_mode") or runtime.get("mode") or "builtin",
        "service_id": ai_trace.get("used_service_id") or runtime.get("service_id"),
        "switched": bool(ai_trace.get("switched")),
        "attempts": len(trace_attempts),
        "attempt_statuses": [str(x.get("status_code") or x.get("error") or "ok") for x in trace_attempts[:5]],
        "marketplace": "wb",
        "content_kind": "review",
        "product": str(payload.product_name or "")[:180],
        "question_or_review": str(payload.review_text or "")[:320],
        "answer": str(reply or "")[:420],
    }
    _audit(
        db,
        user,
        action="wb_review_reply_generated",
        details=json.dumps(details_payload, ensure_ascii=False),
        module_code="wb_reviews_ai",
        entity_type="review",
    )
    db.commit()
    return GenerateReviewReplyOut(reply=reply)


@router.get("/ozon/reviews", response_model=WbReviewsOut)
def ozon_reviews(
    stars: int | None = None,
    date_from: date | None = None,
    date_to: date | None = None,
    fast: bool = False,
    user: User = Depends(get_current_user),
    db: Session = Depends(get_db),
):
    ensure_module_enabled(db, user, "wb_reviews_ai")
    ozon_key = _get_active_marketplace_api_key(db, user.id, "ozon")
    if not ozon_key:
        raise HTTPException(status_code=400, detail="РЎРЅР°С‡Р°Р»Р° СЃРѕС…СЂР°РЅРёС‚Рµ API РєР»СЋС‡ РґР»СЏ ozon")
    if stars is not None and (stars < 1 or stars > 5):
        raise HTTPException(status_code=400, detail="stars РґРѕР»Р¶РµРЅ Р±С‹С‚СЊ РѕС‚ 1 РґРѕ 5")
    left = date_from.isoformat() if date_from else ""
    right = date_to.isoformat() if date_to else ""
    cache_key = build_market_cache_key(
        {
            "kind": "reviews",
            "stars": int(stars or 0),
            "date_from": left,
            "date_to": right,
            "fast": int(bool(fast)),
            "key_rev": _secret_revision(ozon_key),
        }
    )

    def _load_reviews_payload() -> dict[str, Any]:
        return fetch_ozon_reviews(
            ozon_key,
            stars=stars,
            date_from=left or None,
            date_to=right or None,
            max_pages=1 if fast else 8,
            enrich_products=not fast,
        )

    data, cache_meta = get_or_refresh_market_cache(
        db,
        user_id=int(user.id),
        module_code="wb_reviews_ai",
        marketplace="ozon",
        cache_key=cache_key,
        ttl_sec=_market_cache_ttl("wb_reviews_ai", fast_mode=fast),
        fetcher=_load_reviews_payload,
        stale_if_error_sec=20 * 60,
    )
    new_rows = _filter_claimed_feedback_rows(
        db,
        user,
        module_code="wb_reviews_ai",
        marketplace="ozon",
        item_type="review",
        rows=list(data.get("new") or []),
    )
    answered_rows = _filter_claimed_feedback_rows(
        db,
        user,
        module_code="wb_reviews_ai",
        marketplace="ozon",
        item_type="review",
        rows=list(data.get("answered") or []),
    )
    if not new_rows and not answered_rows:
        ok, message = probe_ozon_feedback_access(ozon_key, feedback_kind="reviews")
        if not ok:
            raise HTTPException(status_code=400, detail=message)
    _audit(
        db,
        user,
        action="ozon_reviews_read",
        details=f"new={len(new_rows)};answered={len(answered_rows)};source={cache_meta.get('source')};age={cache_meta.get('age_sec')}",
        module_code="wb_reviews_ai",
        entity_type="review",
    )
    db.commit()
    return WbReviewsOut(new=new_rows, answered=answered_rows)


@router.post("/ozon/reviews/reply", response_model=WbReviewReplyOut)
def ozon_reply_review(payload: WbReviewReplyIn, user: User = Depends(get_current_user), db: Session = Depends(get_db)):
    ensure_module_enabled(db, user, "wb_reviews_ai")
    ozon_key = _get_active_marketplace_api_key(db, user.id, "ozon")
    if not ozon_key:
        raise HTTPException(status_code=400, detail="РЎРЅР°С‡Р°Р»Р° СЃРѕС…СЂР°РЅРёС‚Рµ API РєР»СЋС‡ РґР»СЏ ozon")
    _claim_or_validate_work_item(
        db,
        user,
        module_code="wb_reviews_ai",
        marketplace="ozon",
        item_type="review",
        item_external_id=str(payload.id or "").strip(),
    )
    ok, message = post_ozon_review_reply(ozon_key, payload.id, payload.text)
    message_short = re.sub(r"\s+", " ", str(message or "")).strip()[:260]
    detail_payload = {
        "review_id": str(payload.id or ""),
        "ok": bool(ok),
        "marketplace": "ozon",
        "reply": str(payload.text or "")[:800],
        "message": message_short,
    }
    _audit(
        db,
        user,
        action="ozon_review_reply",
        details=json.dumps(detail_payload, ensure_ascii=False),
        module_code="wb_reviews_ai",
        entity_type="review",
        entity_id=str(payload.id or ""),
        status="ok" if ok else "error",
    )
    db.commit()
    if not ok:
        raise HTTPException(status_code=400, detail=message)
    return WbReviewReplyOut(ok=True, message=message)


@router.post("/ozon/reviews/generate-reply", response_model=GenerateReviewReplyOut)
def ozon_generate_reply(payload: GenerateReviewReplyIn, user: User = Depends(get_current_user), db: Session = Depends(get_db)):
    ensure_module_enabled(db, user, "wb_reviews_ai")
    settings_row = _get_or_create_ai_settings(db, user.id)
    runtime = _resolve_user_ai_runtime(db, user.id)
    knowledge_ctx = _build_user_knowledge_context(
        db,
        user.id,
        query_text=f"{payload.product_name} {payload.review_text} {payload.reviewer_name}",
    )
    prompt = _compose_ai_prompt(settings_row.prompt if settings_row and settings_row.prompt else "", knowledge_ctx, content_kind="review")
    ai_trace: dict[str, Any] = {}
    reply = generate_review_reply(
        review_text=payload.review_text,
        product_name=payload.product_name,
        stars=payload.stars,
        prompt=prompt,
        reviewer_name=payload.reviewer_name,
        marketplace="ozon",
        content_kind="review",
        api_key=str(runtime.get("api_key") or ""),
        model=str(runtime.get("model") or ""),
        provider=str(runtime.get("provider") or ""),
        base_url=str(runtime.get("base_url") or ""),
        fallback_chain=list(runtime.get("fallback_chain") or []),
        trace=ai_trace,
    )
    trace_attempts = [x for x in (ai_trace.get("attempts") or []) if isinstance(x, dict)]
    details_payload = {
        "provider": ai_trace.get("used_provider") or runtime.get("provider") or "builtin",
        "model": ai_trace.get("used_model") or runtime.get("model") or settings.openai_model,
        "mode": ai_trace.get("used_mode") or runtime.get("mode") or "builtin",
        "service_id": ai_trace.get("used_service_id") or runtime.get("service_id"),
        "switched": bool(ai_trace.get("switched")),
        "attempts": len(trace_attempts),
        "attempt_statuses": [str(x.get("status_code") or x.get("error") or "ok") for x in trace_attempts[:5]],
        "marketplace": "ozon",
        "content_kind": "review",
        "product": str(payload.product_name or "")[:180],
        "question_or_review": str(payload.review_text or "")[:320],
        "answer": str(reply or "")[:420],
    }
    _audit(
        db,
        user,
        action="ozon_review_reply_generated",
        details=json.dumps(details_payload, ensure_ascii=False),
        module_code="wb_reviews_ai",
        entity_type="review",
    )
    db.commit()
    return GenerateReviewReplyOut(reply=reply)


@router.get("/wb/questions", response_model=WbReviewsOut)
def wb_questions(
    date_from: date | None = None,
    date_to: date | None = None,
    fast: bool = False,
    user: User = Depends(get_current_user),
    db: Session = Depends(get_db),
):
    ensure_module_enabled(db, user, "wb_questions_ai")
    wb_key = _get_active_marketplace_api_key(db, user.id, "wb")
    if not wb_key:
        raise HTTPException(status_code=400, detail="РЎРЅР°С‡Р°Р»Р° СЃРѕС…СЂР°РЅРёС‚Рµ API РєР»СЋС‡ РґР»СЏ wb")
    left = date_from.isoformat() if date_from else ""
    right = date_to.isoformat() if date_to else ""
    cache_key = build_market_cache_key(
        {
            "kind": "questions",
            "date_from": left,
            "date_to": right,
            "fast": int(bool(fast)),
            "key_rev": _secret_revision(wb_key),
        }
    )

    def _load_questions_payload() -> dict[str, Any]:
        if fast:
            return fetch_wb_questions_fast(
                wb_key,
                date_from=left or None,
                date_to=right or None,
            )
        return fetch_wb_questions(
            wb_key,
            date_from=left or None,
            date_to=right or None,
            max_pages=8,
        )

    data, cache_meta = get_or_refresh_market_cache(
        db,
        user_id=int(user.id),
        module_code="wb_questions_ai",
        marketplace="wb",
        cache_key=cache_key,
        ttl_sec=_market_cache_ttl("wb_questions_ai", fast_mode=fast),
        fetcher=_load_questions_payload,
        stale_if_error_sec=20 * 60,
    )
    new_rows = _filter_claimed_feedback_rows(
        db,
        user,
        module_code="wb_questions_ai",
        marketplace="wb",
        item_type="question",
        rows=list(data.get("new") or []),
    )
    answered_rows = _filter_claimed_feedback_rows(
        db,
        user,
        module_code="wb_questions_ai",
        marketplace="wb",
        item_type="question",
        rows=list(data.get("answered") or []),
    )
    if not new_rows and not answered_rows:
        ok, message = probe_wb_feedback_access(wb_key, feedback_kind="questions")
        if not ok:
            raise HTTPException(status_code=400, detail=message)
    _audit(
        db,
        user,
        action="wb_questions_read",
        details=f"new={len(new_rows)};answered={len(answered_rows)};source={cache_meta.get('source')};age={cache_meta.get('age_sec')}",
        module_code="wb_questions_ai",
        entity_type="question",
    )
    db.commit()
    return WbReviewsOut(new=new_rows, answered=answered_rows)


@router.post("/wb/questions/reply", response_model=WbReviewReplyOut)
def wb_reply_question(payload: WbReviewReplyIn, user: User = Depends(get_current_user), db: Session = Depends(get_db)):
    ensure_module_enabled(db, user, "wb_questions_ai")
    wb_key = _get_active_marketplace_api_key(db, user.id, "wb")
    if not wb_key:
        raise HTTPException(status_code=400, detail="РЎРЅР°С‡Р°Р»Р° СЃРѕС…СЂР°РЅРёС‚Рµ API РєР»СЋС‡ РґР»СЏ wb")
    _claim_or_validate_work_item(
        db,
        user,
        module_code="wb_questions_ai",
        marketplace="wb",
        item_type="question",
        item_external_id=str(payload.id or "").strip(),
    )
    state_value = str(payload.state or "").strip()
    try:
        ok, message = post_wb_question_reply(
            wb_key,
            payload.id,
            payload.text,
            state=state_value,
        )
    except TypeError as exc:
        # Backward-compatible fallback for older service signatures without `state`.
        if "state" not in str(exc):
            raise
        ok, message = post_wb_question_reply(
            wb_key,
            payload.id,
            payload.text,
        )
    message_short = re.sub(r"\s+", " ", str(message or "")).strip()[:260]
    detail_payload = {
        "question_id": str(payload.id or ""),
        "ok": bool(ok),
        "marketplace": "wb",
        "reply": str(payload.text or "")[:800],
        "state": str(payload.state or "").strip(),
        "message": message_short,
    }
    _audit(
        db,
        user,
        action="wb_question_reply",
        details=json.dumps(detail_payload, ensure_ascii=False),
        module_code="wb_questions_ai",
        entity_type="question",
        entity_id=str(payload.id or ""),
        status="ok" if ok else "error",
    )
    db.commit()
    if not ok:
        raise HTTPException(status_code=400, detail=message)
    return WbReviewReplyOut(ok=True, message=message)


@router.post("/wb/questions/generate-reply", response_model=GenerateReviewReplyOut)
def wb_generate_question_reply(payload: GenerateReviewReplyIn, user: User = Depends(get_current_user), db: Session = Depends(get_db)):
    ensure_module_enabled(db, user, "wb_questions_ai")
    settings_row = _get_or_create_question_ai_settings(db, user.id)
    runtime = _resolve_user_ai_runtime(db, user.id)
    knowledge_ctx = _build_user_knowledge_context(
        db,
        user.id,
        query_text=f"{payload.product_name} {payload.review_text} {payload.reviewer_name}",
    )
    prompt = _compose_ai_prompt(settings_row.prompt if settings_row and settings_row.prompt else "", knowledge_ctx, content_kind="question")
    ai_trace: dict[str, Any] = {}
    reply = generate_review_reply(
        review_text=payload.review_text,
        product_name=payload.product_name,
        stars=None,
        prompt=prompt,
        reviewer_name=payload.reviewer_name,
        marketplace="wb",
        content_kind="question",
        api_key=str(runtime.get("api_key") or ""),
        model=str(runtime.get("model") or ""),
        provider=str(runtime.get("provider") or ""),
        base_url=str(runtime.get("base_url") or ""),
        fallback_chain=list(runtime.get("fallback_chain") or []),
        trace=ai_trace,
    )
    trace_attempts = [x for x in (ai_trace.get("attempts") or []) if isinstance(x, dict)]
    details_payload = {
        "provider": ai_trace.get("used_provider") or runtime.get("provider") or "builtin",
        "model": ai_trace.get("used_model") or runtime.get("model") or settings.openai_model,
        "mode": ai_trace.get("used_mode") or runtime.get("mode") or "builtin",
        "service_id": ai_trace.get("used_service_id") or runtime.get("service_id"),
        "switched": bool(ai_trace.get("switched")),
        "attempts": len(trace_attempts),
        "attempt_statuses": [str(x.get("status_code") or x.get("error") or "ok") for x in trace_attempts[:5]],
        "marketplace": "wb",
        "content_kind": "question",
        "product": str(payload.product_name or "")[:180],
        "question_or_review": str(payload.review_text or "")[:320],
        "answer": str(reply or "")[:420],
    }
    _audit(
        db,
        user,
        action="wb_question_reply_generated",
        details=json.dumps(details_payload, ensure_ascii=False),
        module_code="wb_questions_ai",
        entity_type="question",
    )
    db.commit()
    return GenerateReviewReplyOut(reply=reply)


@router.get("/ozon/questions", response_model=WbReviewsOut)
def ozon_questions(
    date_from: date | None = None,
    date_to: date | None = None,
    fast: bool = False,
    user: User = Depends(get_current_user),
    db: Session = Depends(get_db),
):
    ensure_module_enabled(db, user, "wb_questions_ai")
    ozon_key = _get_active_marketplace_api_key(db, user.id, "ozon")
    if not ozon_key:
        raise HTTPException(status_code=400, detail="РЎРЅР°С‡Р°Р»Р° СЃРѕС…СЂР°РЅРёС‚Рµ API РєР»СЋС‡ РґР»СЏ ozon")
    left = date_from.isoformat() if date_from else ""
    right = date_to.isoformat() if date_to else ""
    cache_key = build_market_cache_key(
        {
            "kind": "questions",
            "date_from": left,
            "date_to": right,
            "fast": int(bool(fast)),
            "key_rev": _secret_revision(ozon_key),
        }
    )

    def _load_questions_payload() -> dict[str, Any]:
        return fetch_ozon_questions(
            ozon_key,
            date_from=left or None,
            date_to=right or None,
            max_pages=1 if fast else 8,
            enrich_products=not fast,
        )

    data, cache_meta = get_or_refresh_market_cache(
        db,
        user_id=int(user.id),
        module_code="wb_questions_ai",
        marketplace="ozon",
        cache_key=cache_key,
        ttl_sec=_market_cache_ttl("wb_questions_ai", fast_mode=fast),
        fetcher=_load_questions_payload,
        stale_if_error_sec=20 * 60,
    )
    new_rows = _filter_claimed_feedback_rows(
        db,
        user,
        module_code="wb_questions_ai",
        marketplace="ozon",
        item_type="question",
        rows=list(data.get("new") or []),
    )
    answered_rows = _filter_claimed_feedback_rows(
        db,
        user,
        module_code="wb_questions_ai",
        marketplace="ozon",
        item_type="question",
        rows=list(data.get("answered") or []),
    )
    if not new_rows and not answered_rows:
        ok, message = probe_ozon_feedback_access(ozon_key, feedback_kind="questions")
        if not ok:
            raise HTTPException(status_code=400, detail=message)
    _audit(
        db,
        user,
        action="ozon_questions_read",
        details=f"new={len(new_rows)};answered={len(answered_rows)};source={cache_meta.get('source')};age={cache_meta.get('age_sec')}",
        module_code="wb_questions_ai",
        entity_type="question",
    )
    db.commit()
    return WbReviewsOut(new=new_rows, answered=answered_rows)


@router.post("/ozon/questions/reply", response_model=WbReviewReplyOut)
def ozon_reply_question(payload: WbReviewReplyIn, user: User = Depends(get_current_user), db: Session = Depends(get_db)):
    ensure_module_enabled(db, user, "wb_questions_ai")
    ozon_key = _get_active_marketplace_api_key(db, user.id, "ozon")
    if not ozon_key:
        raise HTTPException(status_code=400, detail="РЎРЅР°С‡Р°Р»Р° СЃРѕС…СЂР°РЅРёС‚Рµ API РєР»СЋС‡ РґР»СЏ ozon")
    _claim_or_validate_work_item(
        db,
        user,
        module_code="wb_questions_ai",
        marketplace="ozon",
        item_type="question",
        item_external_id=str(payload.id or "").strip(),
    )
    ok, message = post_ozon_question_reply(
        ozon_key,
        payload.id,
        payload.text,
        sku=int(payload.sku or 0) if int(payload.sku or 0) > 0 else None,
    )
    message_short = re.sub(r"\s+", " ", str(message or "")).strip()[:260]
    detail_payload = {
        "question_id": str(payload.id or ""),
        "ok": bool(ok),
        "marketplace": "ozon",
        "reply": str(payload.text or "")[:800],
        "sku": int(payload.sku or 0) if int(payload.sku or 0) > 0 else None,
        "message": message_short,
    }
    _audit(
        db,
        user,
        action="ozon_question_reply",
        details=json.dumps(detail_payload, ensure_ascii=False),
        module_code="wb_questions_ai",
        entity_type="question",
        entity_id=str(payload.id or ""),
        status="ok" if ok else "error",
    )
    db.commit()
    if not ok:
        raise HTTPException(status_code=400, detail=message)
    return WbReviewReplyOut(ok=True, message=message)


@router.post("/ozon/questions/generate-reply", response_model=GenerateReviewReplyOut)
def ozon_generate_question_reply(payload: GenerateReviewReplyIn, user: User = Depends(get_current_user), db: Session = Depends(get_db)):
    ensure_module_enabled(db, user, "wb_questions_ai")
    settings_row = _get_or_create_question_ai_settings(db, user.id)
    runtime = _resolve_user_ai_runtime(db, user.id)
    knowledge_ctx = _build_user_knowledge_context(
        db,
        user.id,
        query_text=f"{payload.product_name} {payload.review_text} {payload.reviewer_name}",
    )
    prompt = _compose_ai_prompt(settings_row.prompt if settings_row and settings_row.prompt else "", knowledge_ctx, content_kind="question")
    ai_trace: dict[str, Any] = {}
    reply = generate_review_reply(
        review_text=payload.review_text,
        product_name=payload.product_name,
        stars=None,
        prompt=prompt,
        reviewer_name=payload.reviewer_name,
        marketplace="ozon",
        content_kind="question",
        api_key=str(runtime.get("api_key") or ""),
        model=str(runtime.get("model") or ""),
        provider=str(runtime.get("provider") or ""),
        base_url=str(runtime.get("base_url") or ""),
        fallback_chain=list(runtime.get("fallback_chain") or []),
        trace=ai_trace,
    )
    trace_attempts = [x for x in (ai_trace.get("attempts") or []) if isinstance(x, dict)]
    details_payload = {
        "provider": ai_trace.get("used_provider") or runtime.get("provider") or "builtin",
        "model": ai_trace.get("used_model") or runtime.get("model") or settings.openai_model,
        "mode": ai_trace.get("used_mode") or runtime.get("mode") or "builtin",
        "service_id": ai_trace.get("used_service_id") or runtime.get("service_id"),
        "switched": bool(ai_trace.get("switched")),
        "attempts": len(trace_attempts),
        "attempt_statuses": [str(x.get("status_code") or x.get("error") or "ok") for x in trace_attempts[:5]],
        "marketplace": "ozon",
        "content_kind": "question",
        "product": str(payload.product_name or "")[:180],
        "question_or_review": str(payload.review_text or "")[:320],
        "answer": str(reply or "")[:420],
    }
    _audit(
        db,
        user,
        action="ozon_question_reply_generated",
        details=json.dumps(details_payload, ensure_ascii=False),
        module_code="wb_questions_ai",
        entity_type="question",
    )
    db.commit()
    return GenerateReviewReplyOut(reply=reply)


@router.get("/wb/returns", response_model=ReturnsOut)
def wb_returns_list(
    status: str = "",
    date_from: date | None = None,
    date_to: date | None = None,
    user: User = Depends(get_current_user),
    db: Session = Depends(get_db),
):
    ensure_module_enabled(db, user, "returns")
    wb_key = _get_active_marketplace_api_key(db, user.id, "wb")
    if not wb_key:
        raise HTTPException(status_code=400, detail="РЎРЅР°С‡Р°Р»Р° СЃРѕС…СЂР°РЅРёС‚Рµ API РєР»СЋС‡ РґР»СЏ wb")
    left = date_from.isoformat() if date_from else ""
    right = date_to.isoformat() if date_to else ""
    cache_key = build_market_cache_key(
        {
            "kind": "returns",
            "parser_rev": 4,
            "status": str(status or "").strip().lower(),
            "date_from": left,
            "date_to": right,
            "key_rev": _secret_revision(wb_key),
        }
    )

    def _load_returns_payload() -> dict[str, Any]:
        return fetch_wb_returns(
            wb_key,
            status=status or None,
            date_from=left or None,
            date_to=right or None,
        )

    payload, cache_meta = get_or_refresh_market_cache(
        db,
        user_id=int(user.id),
        module_code="returns",
        marketplace="wb",
        cache_key=cache_key,
        ttl_sec=_market_cache_ttl("returns"),
        fetcher=_load_returns_payload,
        stale_if_error_sec=30 * 60,
    )
    normalized_rows = _normalize_returns_rows(_extract_returns_payload_rows(payload), "wb")
    rows = _filter_claimed_feedback_rows(
        db,
        user,
        module_code="returns",
        marketplace="wb",
        item_type="return",
        rows=normalized_rows,
    )
    warnings = [str(x) for x in (payload.get("warnings") or [])]
    _audit(
        db,
        user,
        action="wb_returns_read",
        details=f"rows={len(rows)};warnings={len(warnings)};source={cache_meta.get('source')};age={cache_meta.get('age_sec')}",
        module_code="returns",
        entity_type="return",
        status="ok" if not warnings else "partial",
    )
    db.commit()
    return ReturnsOut(rows=rows, warnings=warnings)


@router.get("/wb/returns/{return_id}", response_model=dict[str, Any])
def wb_returns_detail(return_id: str, user: User = Depends(get_current_user), db: Session = Depends(get_db)):
    ensure_module_enabled(db, user, "returns")
    wb_key = _get_active_marketplace_api_key(db, user.id, "wb")
    if not wb_key:
        raise HTTPException(status_code=400, detail="РЎРЅР°С‡Р°Р»Р° СЃРѕС…СЂР°РЅРёС‚Рµ API РєР»СЋС‡ РґР»СЏ wb")
    rid = str(return_id or "").strip()
    if not rid:
        raise HTTPException(status_code=400, detail="РќРµРєРѕСЂСЂРµРєС‚РЅС‹Р№ ID РІРѕР·РІСЂР°С‚Р°")
    if not _actor_is_owner(user):
        claim = db.scalar(
            select(WorkItemClaim).where(
                WorkItemClaim.user_id == user.id,
                WorkItemClaim.module_code == "returns",
                WorkItemClaim.marketplace == "wb",
                WorkItemClaim.item_type == "return",
                WorkItemClaim.item_external_id == rid,
            )
        )
        if claim and int(claim.owner_member_id or 0) != _actor_member_id(user):
            raise HTTPException(status_code=403, detail="Р—Р°РїРёСЃСЊ Р·Р°РєСЂРµРїР»РµРЅР° Р·Р° РґСЂСѓРіРёРј СЃРѕС‚СЂСѓРґРЅРёРєРѕРј")
    cache_key = build_market_cache_key(
        {
            "kind": "return_detail",
            "parser_rev": 4,
            "return_id": rid,
            "key_rev": _secret_revision(wb_key),
        }
    )
    row, cache_meta = get_or_refresh_market_cache(
        db,
        user_id=int(user.id),
        module_code="returns",
        marketplace="wb",
        cache_key=cache_key,
        ttl_sec=_market_cache_ttl("returns"),
        fetcher=lambda: fetch_wb_return_details(wb_key, rid),
        stale_if_error_sec=45 * 60,
    )
    if not row:
        raise HTTPException(status_code=404, detail="Р—Р°СЏРІРєР° РЅР° РІРѕР·РІСЂР°С‚ РЅРµ РЅР°Р№РґРµРЅР°")
    normalized = _normalize_returns_rows([row], "wb")
    if normalized:
        safe = dict(row)
        norm = normalized[0]
        for key in ("id", "status", "status_code", "status_note", "created_at", "updated_at", "date", "article", "vendor_code", "nm_id", "product", "quantity", "amount", "customer_comment", "seller_comment", "reason", "description", "photos", "marketplace"):
            if key in norm and norm.get(key) not in (None, "", [], {}):
                safe[key] = norm.get(key)
        safe["description"] = str(safe.get("reason") or safe.get("description") or "")
        safe["raw"] = row
        row = safe
    _audit(
        db,
        user,
        action="wb_return_detail_read",
        details=f"id={rid};source={cache_meta.get('source')}",
        module_code="returns",
        entity_type="return",
        entity_id=rid,
    )
    db.commit()
    return row


@router.patch("/wb/returns/action", response_model=ReturnActionOut)
def wb_returns_action(payload: ReturnActionIn, user: User = Depends(get_current_user), db: Session = Depends(get_db)):
    ensure_module_enabled(db, user, "returns")
    wb_key = _get_active_marketplace_api_key(db, user.id, "wb")
    if not wb_key:
        raise HTTPException(status_code=400, detail="РЎРЅР°С‡Р°Р»Р° СЃРѕС…СЂР°РЅРёС‚Рµ API РєР»СЋС‡ РґР»СЏ wb")
    rid = str(payload.id or "").strip()
    if not rid:
        raise HTTPException(status_code=400, detail="РќРµРєРѕСЂСЂРµРєС‚РЅС‹Р№ ID РІРѕР·РІСЂР°С‚Р°")
    _claim_or_validate_work_item(
        db,
        user,
        module_code="returns",
        marketplace="wb",
        item_type="return",
        item_external_id=rid,
    )
    ok, message, _raw = action_wb_return(wb_key, rid, payload.action, payload.comment)
    _audit(
        db,
        user,
        action="wb_return_action",
        details=f"id={rid};action={payload.action};ok={ok}",
        module_code="returns",
        entity_type="return",
        entity_id=rid,
        status="ok" if ok else "error",
    )
    db.commit()
    if not ok:
        raise HTTPException(status_code=400, detail=message)
    return ReturnActionOut(ok=True, message=message, id=rid, action=payload.action)


@router.get("/ozon/returns", response_model=ReturnsOut)
def ozon_returns_list(
    status: str = "",
    date_from: date | None = None,
    date_to: date | None = None,
    user: User = Depends(get_current_user),
    db: Session = Depends(get_db),
):
    ensure_module_enabled(db, user, "returns")
    ozon_key = _get_active_marketplace_api_key(db, user.id, "ozon")
    if not ozon_key:
        raise HTTPException(status_code=400, detail="РЎРЅР°С‡Р°Р»Р° СЃРѕС…СЂР°РЅРёС‚Рµ API РєР»СЋС‡ РґР»СЏ ozon")
    left = date_from.isoformat() if date_from else ""
    right = date_to.isoformat() if date_to else ""
    cache_key = build_market_cache_key(
        {
            "kind": "returns",
            "parser_rev": 4,
            "status": str(status or "").strip().lower(),
            "date_from": left,
            "date_to": right,
            "key_rev": _secret_revision(ozon_key),
        }
    )

    def _load_returns_payload() -> dict[str, Any]:
        return fetch_ozon_returns(
            ozon_key,
            status=status or None,
            date_from=left or None,
            date_to=right or None,
        )

    payload, cache_meta = get_or_refresh_market_cache(
        db,
        user_id=int(user.id),
        module_code="returns",
        marketplace="ozon",
        cache_key=cache_key,
        ttl_sec=_market_cache_ttl("returns"),
        fetcher=_load_returns_payload,
        stale_if_error_sec=30 * 60,
    )
    normalized_rows = _normalize_returns_rows(_extract_returns_payload_rows(payload), "ozon")
    rows = _filter_claimed_feedback_rows(
        db,
        user,
        module_code="returns",
        marketplace="ozon",
        item_type="return",
        rows=normalized_rows,
    )
    warnings = [str(x) for x in (payload.get("warnings") or [])]
    _audit(
        db,
        user,
        action="ozon_returns_read",
        details=f"rows={len(rows)};warnings={len(warnings)};source={cache_meta.get('source')};age={cache_meta.get('age_sec')}",
        module_code="returns",
        entity_type="return",
        status="ok" if not warnings else "partial",
    )
    db.commit()
    return ReturnsOut(rows=rows, warnings=warnings)


@router.get("/ozon/returns/{return_id}", response_model=dict[str, Any])
def ozon_returns_detail(return_id: str, user: User = Depends(get_current_user), db: Session = Depends(get_db)):
    ensure_module_enabled(db, user, "returns")
    ozon_key = _get_active_marketplace_api_key(db, user.id, "ozon")
    if not ozon_key:
        raise HTTPException(status_code=400, detail="РЎРЅР°С‡Р°Р»Р° СЃРѕС…СЂР°РЅРёС‚Рµ API РєР»СЋС‡ РґР»СЏ ozon")
    rid = str(return_id or "").strip()
    if not rid:
        raise HTTPException(status_code=400, detail="РќРµРєРѕСЂСЂРµРєС‚РЅС‹Р№ ID РІРѕР·РІСЂР°С‚Р°")
    if not _actor_is_owner(user):
        claim = db.scalar(
            select(WorkItemClaim).where(
                WorkItemClaim.user_id == user.id,
                WorkItemClaim.module_code == "returns",
                WorkItemClaim.marketplace == "ozon",
                WorkItemClaim.item_type == "return",
                WorkItemClaim.item_external_id == rid,
            )
        )
        if claim and int(claim.owner_member_id or 0) != _actor_member_id(user):
            raise HTTPException(status_code=403, detail="Р—Р°РїРёСЃСЊ Р·Р°РєСЂРµРїР»РµРЅР° Р·Р° РґСЂСѓРіРёРј СЃРѕС‚СЂСѓРґРЅРёРєРѕРј")
    cache_key = build_market_cache_key(
        {
            "kind": "return_detail",
            "parser_rev": 4,
            "return_id": rid,
            "key_rev": _secret_revision(ozon_key),
        }
    )
    row, cache_meta = get_or_refresh_market_cache(
        db,
        user_id=int(user.id),
        module_code="returns",
        marketplace="ozon",
        cache_key=cache_key,
        ttl_sec=_market_cache_ttl("returns"),
        fetcher=lambda: fetch_ozon_return_details(ozon_key, rid),
        stale_if_error_sec=45 * 60,
    )
    if not row:
        raise HTTPException(status_code=404, detail="Р—Р°СЏРІРєР° РЅР° РІРѕР·РІСЂР°С‚ РЅРµ РЅР°Р№РґРµРЅР°")
    normalized = _normalize_returns_rows([row], "ozon")
    if normalized:
        safe = dict(row)
        norm = normalized[0]
        for key in ("id", "status", "status_code", "status_note", "created_at", "updated_at", "date", "article", "vendor_code", "nm_id", "product", "quantity", "amount", "customer_comment", "seller_comment", "reason", "description", "photos", "marketplace"):
            if key in norm and norm.get(key) not in (None, "", [], {}):
                safe[key] = norm.get(key)
        safe["description"] = str(safe.get("reason") or safe.get("description") or "")
        safe["raw"] = row
        row = safe
    _audit(
        db,
        user,
        action="ozon_return_detail_read",
        details=f"id={rid};source={cache_meta.get('source')}",
        module_code="returns",
        entity_type="return",
        entity_id=rid,
    )
    db.commit()
    return row


@router.get("/wb/questions/ai-settings", response_model=ReviewAiSettingsOut)
def wb_questions_get_ai_settings(user: User = Depends(get_current_user), db: Session = Depends(get_db)):
    ensure_module_enabled(db, user, "wb_questions_ai")
    row = _get_or_create_question_ai_settings(db, user.id)
    db.commit()
    return ReviewAiSettingsOut(reply_mode=row.reply_mode, prompt=row.prompt)


@router.post("/wb/questions/ai-settings", response_model=ReviewAiSettingsOut)
def wb_questions_save_ai_settings(payload: ReviewAiSettingsIn, user: User = Depends(get_current_user), db: Session = Depends(get_db)):
    ensure_module_enabled(db, user, "wb_questions_ai")
    row = _get_or_create_question_ai_settings(db, user.id)
    mode = payload.reply_mode.strip().lower()
    if mode not in {"manual", "suggest", "auto"}:
        raise HTTPException(status_code=400, detail="reply_mode РґРѕР»Р¶РµРЅ Р±С‹С‚СЊ manual, suggest РёР»Рё auto")
    row.reply_mode = mode
    row.prompt = _sanitize_ai_prompt(payload.prompt)
    _audit(
        db,
        user,
        action="wb_questions_ai_settings_saved",
        details=f"reply_mode={mode}",
        module_code="wb_questions_ai",
        entity_type="ai_settings",
    )
    db.commit()
    return ReviewAiSettingsOut(reply_mode=row.reply_mode, prompt=row.prompt)


@router.get("/wb/reviews/ai-settings", response_model=ReviewAiSettingsOut)
def wb_get_ai_settings(user: User = Depends(get_current_user), db: Session = Depends(get_db)):
    ensure_module_enabled(db, user, "wb_reviews_ai")
    row = _get_or_create_ai_settings(db, user.id)
    db.commit()
    return ReviewAiSettingsOut(reply_mode=row.reply_mode, prompt=row.prompt)


@router.post("/wb/reviews/ai-settings", response_model=ReviewAiSettingsOut)
def wb_save_ai_settings(payload: ReviewAiSettingsIn, user: User = Depends(get_current_user), db: Session = Depends(get_db)):
    ensure_module_enabled(db, user, "wb_reviews_ai")
    row = _get_or_create_ai_settings(db, user.id)
    mode = payload.reply_mode.strip().lower()
    if mode not in {"manual", "suggest", "auto"}:
        raise HTTPException(status_code=400, detail="reply_mode РґРѕР»Р¶РµРЅ Р±С‹С‚СЊ manual, suggest РёР»Рё auto")
    row.reply_mode = mode
    row.prompt = _sanitize_ai_prompt(payload.prompt)
    _audit(
        db,
        user,
        action="wb_ai_settings_saved",
        details=f"reply_mode={mode}",
        module_code="wb_reviews_ai",
        entity_type="ai_settings",
    )
    db.commit()
    return ReviewAiSettingsOut(reply_mode=row.reply_mode, prompt=row.prompt)


@router.get("/wb/ads/campaigns", response_model=WbCampaignsOut)
def wb_ads_campaigns(
    fast: bool = True,
    user: User = Depends(get_current_user),
    db: Session = Depends(get_db),
):
    ensure_module_enabled(db, user, "wb_ads")
    wb_key = _get_active_marketplace_api_key(db, user.id, "wb")
    if not wb_key:
        raise HTTPException(status_code=400, detail="РЎРЅР°С‡Р°Р»Р° СЃРѕС…СЂР°РЅРёС‚Рµ API РєР»СЋС‡ РґР»СЏ wb")
    rows = get_wb_snapshot_rows(db, user.id)
    source = "snapshot"
    stale = is_wb_snapshot_stale(db, user.id)
    refresh_queued = False
    queue_depth_now = 0
    queue_available_now = False
    if not rows or stale:
        should_queue_refresh = not (bool(fast) and bool(rows))
        if should_queue_refresh:
            queue_depth_now = queue_depth()
            queue_available_now = queue_available()
            queue_result = enqueue_task(
                "sync_wb_snapshots",
                {"user_id": int(user.id)},
                dedupe_key=f"wb_snapshots:{int(user.id)}",
                dedupe_ttl_sec=120,
            )
            refresh_queued = bool(queue_result.get("queued"))
            queue_depth_now = queue_depth()
            force_sync_now = bool(not rows) or bool(stale and queue_depth_now > 220 and not fast)
            if force_sync_now:
                sync_wb_campaign_snapshots(db, user.id, wb_key)
                rows = get_wb_snapshot_rows(db, user.id)
                source = "sync-fallback" if not rows else "sync-refresh"
                stale = is_wb_snapshot_stale(db, user.id)
            elif not rows:
                source = "snapshot-empty"
            else:
                source = "snapshot+queue-refresh"
        else:
            source = "snapshot-stale-fast"
    def _collect_placeholder_campaign_ids(items: list[dict[str, Any]]) -> list[int]:
        out: list[int] = []
        for item in items:
            cid_local = _to_int_safe(_campaign_id_from_any(item))
            if cid_local <= 0:
                continue
            summary_local = _campaign_summary_from_base_row(item, cid_local)
            name_local = str(summary_local.get("name") or "").strip()
            if _campaign_name_is_placeholder(name_local, cid_local) or not _campaign_summary_has_context(summary_local, cid_local):
                out.append(cid_local)
        return out

    placeholder_ids = _collect_placeholder_campaign_ids(rows)
    if (not fast) and stale and rows and len(placeholder_ids) >= max(6, int(len(rows) * 0.35)):
        # Snapshot can stay stale while queue is busy. If many rows are placeholders,
        # run direct sync for this user to avoid blank campaign names on UI.
        try:
            sync_wb_campaign_snapshots(db, user.id, wb_key)
            rows = get_wb_snapshot_rows(db, user.id)
            source = "sync-on-placeholder"
            stale = is_wb_snapshot_stale(db, user.id)
        except Exception:
            pass
        placeholder_ids = _collect_placeholder_campaign_ids(rows)

    if (not fast) and placeholder_ids:
        preview_ids = sorted(set(placeholder_ids))[:600]
        placeholder_key = build_market_cache_key(
            {
                "kind": "wb_campaigns_placeholder_summaries",
                "ids": preview_ids,
                "key_rev": _secret_revision(wb_key),
            }
        )
        try:
            extra_summaries, _placeholder_meta = get_or_refresh_market_cache(
                db,
                user_id=int(user.id),
                module_code="wb_ads",
                marketplace="wb",
                cache_key=placeholder_key,
                ttl_sec=max(120, _market_cache_ttl("wb_ads")),
                fetcher=lambda: fetch_wb_campaign_summaries(
                    wb_key,
                    preview_ids,
                    fallback_limit=max(8, min(36, len(preview_ids) // 3 + 6)),
                ),
                stale_if_error_sec=60 * 60,
            )
        except Exception:
            extra_summaries = {}
        if isinstance(extra_summaries, dict) and extra_summaries:
            for row in rows:
                cid = _to_int_safe(_campaign_id_from_any(row))
                if cid <= 0:
                    continue
                summary = extra_summaries.get(str(cid))
                if not isinstance(summary, dict):
                    continue
                name = str(summary.get("name") or "").strip()
                status = str(summary.get("status") or "").strip()
                ctype = str(summary.get("type") or "").strip()
                budget = str(summary.get("budget") or "").strip()
                if name and not _campaign_name_is_placeholder(name, cid):
                    row["name"] = name
                if status and status not in {"-", "вЂ”"}:
                    row["status"] = status
                if ctype and ctype not in {"-", "вЂ”"}:
                    row["type"] = ctype
                if budget and budget not in {"-", "вЂ”"}:
                    row["budget"] = budget
    ids = sorted({_to_int_safe(_campaign_id_from_any(row)) for row in rows if _to_int_safe(_campaign_id_from_any(row)) > 0})
    hydrated_stats: dict[str, dict[str, Any]] = {}
    hydrated_summaries: dict[str, dict[str, Any]] = {}
    hydrate_ids = ids[:600]
    if (not fast) and hydrate_ids:
        summary_hydrate_key = build_market_cache_key(
            {
                "kind": "wb_campaigns_hydrate_summaries",
                "ids": hydrate_ids,
                "key_rev": _secret_revision(wb_key),
            }
        )
        stats_hydrate_key = build_market_cache_key(
            {
                "kind": "wb_campaigns_hydrate_stats",
                "ids": hydrate_ids,
                "date_from": "",
                "date_to": "",
                "key_rev": _secret_revision(wb_key),
            }
        )
        try:
            hydrated_summaries, _summary_meta = get_or_refresh_market_cache(
                db,
                user_id=int(user.id),
                module_code="wb_ads",
                marketplace="wb",
                cache_key=summary_hydrate_key,
                ttl_sec=max(120, _market_cache_ttl("wb_ads")),
                fetcher=lambda: fetch_wb_campaign_summaries(
                    wb_key,
                    hydrate_ids,
                    fallback_limit=max(12, min(64, len(hydrate_ids) // 3 + 8)),
                ),
                stale_if_error_sec=45 * 60,
                prefer_stale_sec=20 * 60,
            )
        except Exception:
            hydrated_summaries = {}
        try:
            hydrated_stats, _stats_meta = get_or_refresh_market_cache(
                db,
                user_id=int(user.id),
                module_code="wb_ads_analytics",
                marketplace="wb",
                cache_key=stats_hydrate_key,
                ttl_sec=max(120, _market_cache_ttl("wb_ads_analytics")),
                fetcher=lambda: fetch_wb_campaign_stats_bulk(wb_key, hydrate_ids, date_from=None, date_to=None),
                stale_if_error_sec=45 * 60,
                prefer_stale_sec=20 * 60,
            )
        except Exception:
            hydrated_stats = {}
        if isinstance(hydrated_summaries, dict) or isinstance(hydrated_stats, dict):
            merged_rows: list[dict[str, Any]] = []
            for row in rows:
                cid = _to_int_safe(_campaign_id_from_any(row))
                if cid <= 0:
                    merged_rows.append(row)
                    continue
                key = str(cid)
                summary = hydrated_summaries.get(key) if isinstance(hydrated_summaries, dict) else None
                stat = hydrated_stats.get(key) if isinstance(hydrated_stats, dict) else None
                merged = dict(row)
                if isinstance(summary, dict):
                    current_name = str(merged.get("name") or "").strip()
                    if summary.get("name") and (_campaign_name_is_placeholder(current_name, cid) or not current_name):
                        merged["name"] = str(summary.get("name") or "")
                    if summary.get("status") and str(merged.get("status") or "").strip() in {"", "-", "вЂ”"}:
                        merged["status"] = str(summary.get("status") or "")
                    if summary.get("type") and str(merged.get("type") or "").strip() in {"", "-", "вЂ”"}:
                        merged["type"] = str(summary.get("type") or "")
                    if summary.get("budget") and str(merged.get("budget") or merged.get("dailyBudget") or "").strip() in {"", "-", "вЂ”"}:
                        merged["budget"] = str(summary.get("budget") or "")
                if isinstance(stat, dict):
                    for metric_key in ("views", "clicks", "orders", "spent", "ctr", "cr", "cpc", "cpo", "add_to_cart"):
                        if metric_key in stat:
                            merged[metric_key] = _to_float_safe(stat.get(metric_key), 0.0)
                merged_rows.append(merged)
            rows = merged_rows
    _audit(
        db,
        user,
        action="wb_ads_campaigns_read",
        details=(
            f"count={len(rows)};ids={len(ids)};source={source};stale={int(stale)};"
            f"queue={int(refresh_queued)};queue_depth={queue_depth_now};"
            f"summary_hydrated={len(hydrated_summaries) if isinstance(hydrated_summaries, dict) else 0};"
            f"stats_hydrated={len(hydrated_stats) if isinstance(hydrated_stats, dict) else 0}"
        ),
        module_code="wb_ads",
        entity_type="campaign",
    )
    db.commit()
    return WbCampaignsOut(
        campaigns=rows,
        stats=hydrated_stats if isinstance(hydrated_stats, dict) else {},
        meta={
            "source": source,
            "fast_mode": bool(fast),
            "stale": stale,
            "count": len(rows),
            "refresh_queued": refresh_queued,
            "queue_available": queue_available_now,
            "queue_depth": queue_depth_now,
            "placeholder_count": len(placeholder_ids),
            "summary_hydrated": len(hydrated_summaries) if isinstance(hydrated_summaries, dict) else 0,
            "stats_hydrated": len(hydrated_stats) if isinstance(hydrated_stats, dict) else 0,
        },
    )


@router.post("/wb/ads/campaigns/sync", response_model=dict[str, Any])
def wb_ads_campaigns_sync(user: User = Depends(get_current_user), db: Session = Depends(get_db)):
    ensure_module_enabled(db, user, "wb_ads")
    wb_key = _get_active_marketplace_api_key(db, user.id, "wb")
    if not wb_key:
        raise HTTPException(status_code=400, detail="РЎРЅР°С‡Р°Р»Р° СЃРѕС…СЂР°РЅРёС‚Рµ API РєР»СЋС‡ РґР»СЏ wb")
    payload = enqueue_task(
        "sync_wb_snapshots",
        {"user_id": int(user.id)},
        dedupe_key=f"wb_snapshots:{int(user.id)}",
        dedupe_ttl_sec=120,
    )
    if not payload.get("queued") and not payload.get("ok"):
        payload = sync_wb_campaign_snapshots(db, user.id, wb_key)
    payload["queue_available"] = queue_available()
    payload["queue_depth"] = queue_depth()
    _audit(
        db,
        user,
        action="wb_ads_campaigns_sync",
        details=json.dumps(payload, ensure_ascii=False)[:2000],
        module_code="wb_ads",
        entity_type="campaign",
    )
    db.commit()
    return payload


@router.get("/ozon/ads/campaigns", response_model=WbCampaignsOut)
def ozon_ads_campaigns(user: User = Depends(get_current_user), db: Session = Depends(get_db)):
    ensure_module_enabled(db, user, "wb_ads")
    ozon_key = _get_active_marketplace_api_key(db, user.id, "ozon")
    if not ozon_key:
        raise HTTPException(status_code=400, detail="РЎРЅР°С‡Р°Р»Р° СЃРѕС…СЂР°РЅРёС‚Рµ API РєР»СЋС‡ РґР»СЏ ozon")
    cache_key = build_market_cache_key(
        {
            "kind": "ozon_campaigns",
            "key_rev": _secret_revision(ozon_key),
        }
    )
    payload, cache_meta = get_or_refresh_market_cache(
        db,
        user_id=int(user.id),
        module_code="wb_ads",
        marketplace="ozon",
        cache_key=cache_key,
        ttl_sec=_market_cache_ttl("wb_ads"),
        fetcher=lambda: fetch_ozon_ads_campaigns(ozon_key),
        stale_if_error_sec=30 * 60,
    )
    rows = list(payload.get("rows") or [])
    warnings = [str(x) for x in (payload.get("warnings") or [])]
    _audit(
        db,
        user,
        action="ozon_ads_campaigns_read",
        details=f"count={len(rows)};warnings={len(warnings)};source={cache_meta.get('source')};age={cache_meta.get('age_sec')}",
        module_code="wb_ads",
        entity_type="campaign",
        status="ok" if not warnings else "partial",
    )
    db.commit()
    return WbCampaignsOut(campaigns=rows, stats={}, meta={"warnings": warnings, "count": len(rows)})


@router.get("/ozon/ads/analytics", response_model=WbAdsAnalyticsOut)
def ozon_ads_analytics(
    date_from: date | None = None,
    date_to: date | None = None,
    campaign_id: int | None = None,
    user: User = Depends(get_current_user),
    db: Session = Depends(get_db),
):
    ensure_module_enabled(db, user, "wb_ads_analytics")
    ozon_key = _get_active_marketplace_api_key(db, user.id, "ozon")
    if not ozon_key:
        raise HTTPException(status_code=400, detail="РЎРЅР°С‡Р°Р»Р° СЃРѕС…СЂР°РЅРёС‚Рµ API РєР»СЋС‡ РґР»СЏ ozon")
    left = date_from.isoformat() if date_from else ""
    right = date_to.isoformat() if date_to else ""
    cache_key = build_market_cache_key(
        {
            "kind": "ozon_ads_analytics",
            "date_from": left,
            "date_to": right,
            "campaign_id": int(campaign_id or 0),
            "key_rev": _secret_revision(ozon_key),
        }
    )

    def _load_analytics_payload() -> dict[str, Any]:
        return fetch_ozon_ads_analytics(
            ozon_key,
            date_from=left or None,
            date_to=right or None,
            campaign_id=campaign_id,
        )

    payload, cache_meta = get_or_refresh_market_cache(
        db,
        user_id=int(user.id),
        module_code="wb_ads_analytics",
        marketplace="ozon",
        cache_key=cache_key,
        ttl_sec=_market_cache_ttl("wb_ads_analytics"),
        fetcher=_load_analytics_payload,
        stale_if_error_sec=30 * 60,
    )
    rows = list(payload.get("rows") or [])
    warnings = [str(x) for x in (payload.get("warnings") or [])]
    left = str(payload.get("date_from") or (date_from.isoformat() if date_from else date.today().isoformat()))
    right = str(payload.get("date_to") or (date_to.isoformat() if date_to else date.today().isoformat()))
    totals = dict(payload.get("totals") or {"views": 0.0, "clicks": 0.0, "orders": 0.0, "spent": 0.0, "ctr_avg": 0.0, "cr_avg": 0.0})
    if warnings:
        totals["warning_count"] = float(len(warnings))
    _audit(
        db,
        user,
        action="ozon_ads_analytics_read",
        details=f"date_from={left};date_to={right};rows={len(rows)};warnings={len(warnings)};source={cache_meta.get('source')};age={cache_meta.get('age_sec')}",
        module_code="wb_ads_analytics",
        entity_type="campaign",
        status="ok" if not warnings else "partial",
    )
    db.commit()
    return WbAdsAnalyticsOut(date_from=left, date_to=right, rows=rows, totals=totals)


@router.post("/wb/ads/campaigns/enrich", response_model=WbCampaignEnrichOut)
def wb_ads_campaigns_enrich(payload: CampaignIdsIn, user: User = Depends(get_current_user), db: Session = Depends(get_db)):
    ensure_module_enabled(db, user, "wb_ads")
    wb_key = _get_active_marketplace_api_key(db, user.id, "wb")
    if not wb_key:
        raise HTTPException(status_code=400, detail="РЎРЅР°С‡Р°Р»Р° СЃРѕС…СЂР°РЅРёС‚Рµ API РєР»СЋС‡ РґР»СЏ wb")

    ids = sorted({int(x) for x in payload.ids if int(x) > 0})[:600]
    if not ids:
        return WbCampaignEnrichOut(summaries={}, stats={})
    started_at = time.monotonic()
    fallback_deadline_sec = 10.0

    summaries: dict[str, dict[str, Any]] = {}
    stats: dict[str, dict[str, Any]] = {}
    warnings: list[str] = []
    error_flags: list[str] = []
    summaries_cache_meta: dict[str, Any] = {"source": "snapshot"}
    stats_cache_meta: dict[str, Any] = {"source": "snapshot"}

    id_set = {int(x) for x in ids}
    snapshot_by_id: dict[str, dict[str, Any]] = {}
    for row in get_wb_snapshot_rows(db, user.id):
        cid = _to_int_safe(_campaign_id_from_any(row))
        if cid > 0 and cid in id_set and str(cid) not in snapshot_by_id:
            snapshot_by_id[str(cid)] = row

    for cid in ids:
        key = str(cid)
        snap_row = snapshot_by_id.get(key)
        if not isinstance(snap_row, dict):
            continue
        snap_summary = _campaign_summary_from_base_row(snap_row, cid)
        if _campaign_summary_has_context(snap_summary, cid):
            summaries[key] = snap_summary
        snap_stat = _campaign_stat_from_base_row(snap_row)
        if snap_stat:
            stats[key] = snap_stat

    summary_fetch_ids = [
        cid
        for cid in ids
        if (
            not _campaign_summary_has_context(summaries.get(str(cid)), cid)
            or _campaign_name_is_placeholder(str((summaries.get(str(cid)) or {}).get("name") or ""), cid)
        )
    ]
    stats_fetch_ids = [cid for cid in ids if not _campaign_stat_has_context(stats.get(str(cid)))]

    if summary_fetch_ids:
        summaries_cache_key = build_market_cache_key(
            {
                "kind": "wb_campaign_summaries_bulk",
                "ids": summary_fetch_ids,
                "key_rev": _secret_revision(wb_key),
            }
        )
        try:
            fresh_summaries, summaries_cache_meta = get_or_refresh_market_cache(
                db,
                user_id=int(user.id),
                module_code="wb_ads",
                marketplace="wb",
                cache_key=summaries_cache_key,
                ttl_sec=max(120, _market_cache_ttl("wb_ads")),
                fetcher=lambda: fetch_wb_campaign_summaries(
                    wb_key,
                    summary_fetch_ids,
                    fallback_limit=max(16, min(180, len(summary_fetch_ids) // 2 + 12)),
                    detail_lookup_limit=max(10, min(90, len(summary_fetch_ids) // 3 + 8)),
                ),
                stale_if_error_sec=45 * 60,
                prefer_stale_sec=60 * 60,
            )
            if isinstance(fresh_summaries, dict):
                for key, value in fresh_summaries.items():
                    if isinstance(value, dict):
                        summaries[str(key)] = value
        except Exception:
            error_flags.append("summaries")
            warnings.append("summary_fetch_failed")

    if stats_fetch_ids:
        stats_cache_key = build_market_cache_key(
            {
                "kind": "wb_campaign_stats_bulk",
                "ids": stats_fetch_ids,
                "date_from": "",
                "date_to": "",
                "key_rev": _secret_revision(wb_key),
            }
        )
        try:
            fresh_stats, stats_cache_meta = get_or_refresh_market_cache(
                db,
                user_id=int(user.id),
                module_code="wb_ads_analytics",
                marketplace="wb",
                cache_key=stats_cache_key,
                ttl_sec=max(120, _market_cache_ttl("wb_ads_analytics")),
                fetcher=lambda: fetch_wb_campaign_stats_bulk(wb_key, stats_fetch_ids, date_from=None, date_to=None),
                stale_if_error_sec=45 * 60,
                prefer_stale_sec=60 * 60,
            )
            if isinstance(fresh_stats, dict):
                for key, value in fresh_stats.items():
                    if isinstance(value, dict):
                        stats[str(key)] = value
        except Exception:
            error_flags.append("stats")
            warnings.append("stats_fetch_failed")

    unresolved_stats_ids = [cid for cid in stats_fetch_ids if not _campaign_stat_has_context(stats.get(str(cid)))]
    if unresolved_stats_ids:
        single_retry_limit = 0 if len(ids) > 48 else (4 if len(ids) > 20 else 8)
        for cid in unresolved_stats_ids[:single_retry_limit]:
            if (time.monotonic() - started_at) >= fallback_deadline_sec:
                warnings.append("stats_retry_timeout")
                break
            try:
                one_map = fetch_wb_campaign_stats_bulk(wb_key, [int(cid)], date_from=None, date_to=None)
            except Exception:
                one_map = {}
            one = one_map.get(str(cid)) if isinstance(one_map, dict) else None
            if isinstance(one, dict) and one:
                stats[str(cid)] = one

    unresolved_summary_ids = [
        cid
        for cid in summary_fetch_ids
        if (
            not _campaign_summary_has_context(summaries.get(str(cid)), cid)
            or _campaign_name_is_placeholder(str((summaries.get(str(cid)) or {}).get("name") or ""), cid)
        )
    ]
    if unresolved_summary_ids:
        retry_limit = 10 if len(ids) > 240 else (16 if len(ids) > 96 else (10 if len(ids) > 24 else 14))
        for cid in unresolved_summary_ids[:retry_limit]:
            if (time.monotonic() - started_at) >= fallback_deadline_sec:
                warnings.append("summary_retry_timeout")
                break
            try:
                detail_payload = fetch_wb_campaign_details(wb_key, campaign_id=int(cid))
            except Exception:
                detail_payload = {}
            detail_summary = detail_payload.get("summary") if isinstance(detail_payload, dict) else None
            if isinstance(detail_summary, dict) and detail_summary:
                merged = _merge_campaign_row({}, detail_summary, {})
                summaries[str(cid)] = {
                    "campaign_id": int(detail_summary.get("campaign_id") or cid),
                    "name": str(merged.get("name") or detail_summary.get("name") or f"РљР°РјРїР°РЅРёСЏ {cid}"),
                    "status": str(merged.get("status") or detail_summary.get("status") or "-"),
                    "type": str(merged.get("type") or detail_summary.get("type") or "-"),
                    "budget": str(merged.get("budget") or detail_summary.get("budget") or "-"),
                }
            detail_stats = detail_payload.get("stats") if isinstance(detail_payload, dict) else None
            if isinstance(detail_stats, dict) and detail_stats and not _campaign_stat_has_context(stats.get(str(cid))):
                stats[str(cid)] = detail_stats

    missing_ids: list[int] = []
    missing_summary_ids: list[int] = []
    missing_stats_ids: list[int] = []
    hard_missing_ids: list[int] = []
    partial_summary_ids: list[int] = []
    partial_stats_ids: list[int] = []
    for cid in ids:
        key = str(cid)
        summary_ok = _campaign_summary_has_context(summaries.get(key), cid)
        stats_ok = _campaign_stat_has_context(stats.get(key))
        if not summary_ok:
            missing_summary_ids.append(cid)
        if not stats_ok:
            missing_stats_ids.append(cid)
        if not summary_ok and not stats_ok:
            hard_missing_ids.append(cid)
            missing_ids.append(cid)
        elif not summary_ok:
            partial_summary_ids.append(cid)
        elif not stats_ok:
            partial_stats_ids.append(cid)
    summary_count = len([x for x in ids if _campaign_summary_has_context(summaries.get(str(x)), x)])
    stats_count = len([x for x in ids if _campaign_stat_has_context(stats.get(str(x)))])
    temporary_unavailable = False
    if ids and not error_flags and summary_count <= 0 and stats_count <= 0:
        temporary_unavailable = True
        missing_ids = []
        hard_missing_ids = []
        missing_summary_ids = list(ids)
        missing_stats_ids = list(ids)
        partial_summary_ids = list(ids)
        partial_stats_ids = list(ids)
        warnings.append("temporary_unavailable")
    resolved_count = max(0, len(ids) - len(hard_missing_ids))
    if hard_missing_ids:
        warnings.append("partial_data")
    if partial_summary_ids:
        warnings.append("summary_partial")
    if partial_stats_ids:
        warnings.append("stats_partial")

    meta = {
        "requested_count": len(ids),
        "summary_count": summary_count,
        "stats_count": stats_count,
        "resolved_count": resolved_count,
        "missing_count": len(hard_missing_ids),
        "missing_ids": hard_missing_ids[:160],
        "hard_missing_ids": hard_missing_ids[:160],
        "partial_summary_count": len(partial_summary_ids),
        "partial_summary_ids": partial_summary_ids[:160],
        "partial_stats_count": len(partial_stats_ids),
        "partial_stats_ids": partial_stats_ids[:160],
        "missing_summary_ids": missing_summary_ids[:160],
        "missing_stats_ids": missing_stats_ids[:160],
        "temporary_unavailable": temporary_unavailable,
        "summary_source": str(summaries_cache_meta.get("source") or "snapshot"),
        "summary_age_sec": int(summaries_cache_meta.get("age_sec") or 0),
        "stats_source": str(stats_cache_meta.get("source") or "snapshot"),
        "stats_age_sec": int(stats_cache_meta.get("age_sec") or 0),
        "warnings": warnings,
        "errors": error_flags,
    }
    _audit(
        db,
        user,
        action="wb_ads_campaigns_enrich",
        details=(
            f"ids={len(ids)};resolved={resolved_count};missing={len(hard_missing_ids)};"
            f"summaries={meta.get('summary_count')};stats={meta.get('stats_count')};"
            f"errors={','.join(error_flags) if error_flags else '-'};warnings={','.join(warnings) if warnings else '-'}"
        ),
        module_code="wb_ads",
        entity_type="campaign",
    )
    db.commit()
    return WbCampaignEnrichOut(summaries=summaries, stats=stats, meta=meta)


@router.post("/wb/ads/rates", response_model=WbCampaignRatesOut)
def wb_ads_rates(payload: WbCampaignRatesIn, user: User = Depends(get_current_user), db: Session = Depends(get_db)):
    ensure_module_enabled(db, user, "wb_ads")
    wb_key = _get_active_marketplace_api_key(db, user.id, "wb")
    if not wb_key:
        raise HTTPException(status_code=400, detail="РЎРЅР°С‡Р°Р»Р° СЃРѕС…СЂР°РЅРёС‚Рµ API РєР»СЋС‡ РґР»СЏ wb")
    cache_key = build_market_cache_key(
        {
            "kind": "wb_campaign_rates",
            "campaign_id": int(payload.campaign_id or 0),
            "campaign_type": str(payload.campaign_type or "").strip().lower(),
            "key_rev": _secret_revision(wb_key),
        }
    )
    data, cache_meta = get_or_refresh_market_cache(
        db,
        user_id=int(user.id),
        module_code="wb_ads",
        marketplace="wb",
        cache_key=cache_key,
        ttl_sec=_market_cache_ttl("wb_ads"),
        fetcher=lambda: fetch_wb_campaign_rates(wb_key, payload.campaign_id, payload.campaign_type),
        stale_if_error_sec=30 * 60,
    )
    if data is None:
        raise HTTPException(status_code=400, detail="РќРµ СѓРґР°Р»РѕСЃСЊ РїРѕР»СѓС‡РёС‚СЊ СЃС‚Р°РІРєРё РїРѕ РєР°РјРїР°РЅРёРё")
    _audit(
        db,
        user,
        action="wb_ads_rates_read",
        details=f"campaign_id={payload.campaign_id};type={payload.campaign_type};source={cache_meta.get('source')}",
        module_code="wb_ads",
        entity_type="campaign",
        entity_id=str(payload.campaign_id),
    )
    db.commit()
    return WbCampaignRatesOut(campaign_id=payload.campaign_id, campaign_type=payload.campaign_type, data=data)


@router.get("/wb/ads/campaign-details", response_model=WbCampaignDetailOut)
def wb_ads_campaign_details(campaign_id: int, user: User = Depends(get_current_user), db: Session = Depends(get_db)):
    ensure_module_enabled(db, user, "wb_ads")
    wb_key = _get_active_marketplace_api_key(db, user.id, "wb")
    if not wb_key:
        raise HTTPException(status_code=400, detail="РЎРЅР°С‡Р°Р»Р° СЃРѕС…СЂР°РЅРёС‚Рµ API РєР»СЋС‡ РґР»СЏ wb")
    if campaign_id <= 0:
        raise HTTPException(status_code=400, detail="campaign_id РґРѕР»Р¶РµРЅ Р±С‹С‚СЊ > 0")

    snapshot_payload: dict[str, Any] = {}
    for row in get_wb_snapshot_rows(db, user.id):
        if _to_int_safe(_campaign_id_from_any(row)) != int(campaign_id):
            continue
        snapshot_summary = _campaign_summary_from_base_row(row, campaign_id)
        snapshot_stats = _campaign_stat_from_base_row(row)
        snapshot_payload = {
            "summary": snapshot_summary,
            "stats": snapshot_stats,
            "products": [],
            "rates": {},
            "raw": {"snapshot": row},
        }
        break

    cache_key = build_market_cache_key(
        {
            "kind": "wb_campaign_details",
            "campaign_id": int(campaign_id),
            "key_rev": _secret_revision(wb_key),
        }
    )
    try:
        data, cache_meta = get_or_refresh_market_cache(
            db,
            user_id=int(user.id),
            module_code="wb_ads",
            marketplace="wb",
            cache_key=cache_key,
            ttl_sec=_market_cache_ttl("wb_ads"),
            fetcher=lambda: fetch_wb_campaign_details(wb_key, campaign_id=campaign_id),
            stale_if_error_sec=45 * 60,
            prefer_stale_sec=30 * 60,
        )
    except Exception as exc:
        if not snapshot_payload:
            raise
        data = snapshot_payload
        cache_meta = {
            "source": "snapshot-fallback",
            "stale": True,
            "age_sec": 0,
            "ttl_sec": 0,
            "error": str(exc or "")[:240],
        }

    if snapshot_payload and isinstance(data, dict):
        merged_data = dict(data)
        live_summary = merged_data.get("summary") if isinstance(merged_data.get("summary"), dict) else {}
        live_stats = merged_data.get("stats") if isinstance(merged_data.get("stats"), dict) else {}
        if not _campaign_summary_has_context(live_summary, campaign_id):
            merged_data["summary"] = snapshot_payload.get("summary")
        if not _campaign_stat_has_context(live_stats) and snapshot_payload.get("stats"):
            merged_data["stats"] = snapshot_payload.get("stats")
        if not merged_data.get("raw") and snapshot_payload.get("raw"):
            merged_data["raw"] = snapshot_payload.get("raw")
        data = merged_data

    _audit(
        db,
        user,
        action="wb_ads_campaign_details_read",
        details=f"campaign_id={campaign_id};source={cache_meta.get('source')}",
        module_code="wb_ads",
        entity_type="campaign",
        entity_id=str(campaign_id),
    )
    db.commit()
    return WbCampaignDetailOut(campaign_id=campaign_id, data=data)


@router.get("/wb/ads/balance", response_model=WbAdsBalanceOut)
def wb_ads_balance(user: User = Depends(get_current_user), db: Session = Depends(get_db)):
    ensure_module_enabled(db, user, "wb_ads")
    wb_key = _get_active_marketplace_api_key(db, user.id, "wb")
    if not wb_key:
        raise HTTPException(status_code=400, detail="РЎРЅР°С‡Р°Р»Р° СЃРѕС…СЂР°РЅРёС‚Рµ API РєР»СЋС‡ РґР»СЏ wb")
    cache_key = build_market_cache_key(
        {
            "kind": "wb_ads_balance",
            "key_rev": _secret_revision(wb_key),
        }
    )
    data, cache_meta = get_or_refresh_market_cache(
        db,
        user_id=int(user.id),
        module_code="wb_ads",
        marketplace="wb",
        cache_key=cache_key,
        ttl_sec=max(120, _market_cache_ttl("wb_ads")),
        fetcher=lambda: fetch_wb_ads_balance(wb_key),
        stale_if_error_sec=30 * 60,
    )
    if data is None:
        raise HTTPException(status_code=400, detail="РќРµ СѓРґР°Р»РѕСЃСЊ РїРѕР»СѓС‡РёС‚СЊ Р±Р°Р»Р°РЅСЃ WB Ads")
    _audit(
        db,
        user,
        action="wb_ads_balance_read",
        details=f"ok=1;source={cache_meta.get('source')}",
        module_code="wb_ads",
        entity_type="balance",
    )
    db.commit()
    return WbAdsBalanceOut(data=data)


@router.post("/wb/ads/action", response_model=WbAdsActionOut)
def wb_ads_action(payload: WbAdsActionIn, user: User = Depends(get_current_user), db: Session = Depends(get_db)):
    ensure_module_enabled(db, user, "wb_ads")
    wb_key = _get_active_marketplace_api_key(db, user.id, "wb")
    if not wb_key:
        raise HTTPException(status_code=400, detail="РЎРЅР°С‡Р°Р»Р° СЃРѕС…СЂР°РЅРёС‚Рµ API РєР»СЋС‡ РґР»СЏ wb")
    ok, message, raw = update_wb_campaign_state(wb_key, campaign_id=payload.campaign_id, action=payload.action)
    _audit(
        db,
        user,
        action="wb_ads_action",
        details=f"campaign_id={payload.campaign_id};action={payload.action};ok={ok};raw={json.dumps(raw, ensure_ascii=False)[:600]}",
        module_code="wb_ads",
        entity_type="campaign",
        entity_id=str(payload.campaign_id),
        status="ok" if ok else "error",
    )
    db.commit()
    if not ok:
        raise HTTPException(status_code=400, detail=message)
    return WbAdsActionOut(campaign_id=payload.campaign_id, action=payload.action, ok=ok, message=message)


@router.get("/wb/ads/bidder/rules", response_model=WbBidderRulesOut)
def wb_ads_bidder_rules(user: User = Depends(get_current_user), db: Session = Depends(get_db)):
    ensure_module_enabled(db, user, "wb_ads")
    rows = [serialize_bidder_rule(x) for x in list_bidder_rules(db, user.id)]
    active_count = len([x for x in rows if bool(x.get("is_active"))])
    _audit(
        db,
        user,
        action="wb_ads_bidder_rules_read",
        details=f"rows={len(rows)};active={active_count}",
        module_code="wb_ads",
        entity_type="bidder_rule",
    )
    db.commit()
    return WbBidderRulesOut(rows=[WbBidderRuleOut(**x) for x in rows], meta={"count": len(rows), "active": active_count})


@router.post("/wb/ads/bidder/rules", response_model=WbBidderRuleOut)
def wb_ads_bidder_rules_create(
    payload: WbBidderRuleIn,
    user: User = Depends(get_current_user),
    db: Session = Depends(get_db),
):
    ensure_module_enabled(db, user, "wb_ads")
    try:
        normalized = normalize_rule_payload(payload.model_dump())
    except ValueError as exc:
        raise HTTPException(status_code=400, detail=str(exc))
    row = WbAdsBidderRule(user_id=int(user.id), **normalized)
    db.add(row)
    db.flush()
    serialized = serialize_bidder_rule(row)
    _audit(
        db,
        user,
        action="wb_ads_bidder_rule_created",
        details=f"rule_id={row.id};campaign_id={row.campaign_id};target_kind={row.target_kind};nm_id={row.nm_id};active={int(row.is_active)}",
        module_code="wb_ads",
        entity_type="bidder_rule",
        entity_id=str(row.id),
    )
    try:
        db.commit()
    except Exception as exc:
        db.rollback()
        low = str(exc or "").lower()
        if "uq_wb_ads_bidder_rule_target" in low or "unique" in low:
            raise HTTPException(status_code=400, detail="РўР°РєРѕРµ РїСЂР°РІРёР»Рѕ СѓР¶Рµ СЃСѓС‰РµСЃС‚РІСѓРµС‚")
        raise
    return WbBidderRuleOut(**serialized)


@router.patch("/wb/ads/bidder/rules/{rule_id}", response_model=WbBidderRuleOut)
def wb_ads_bidder_rules_update(
    rule_id: int,
    payload: WbBidderRuleUpdateIn,
    user: User = Depends(get_current_user),
    db: Session = Depends(get_db),
):
    ensure_module_enabled(db, user, "wb_ads")
    row = db.scalar(
        select(WbAdsBidderRule).where(
            WbAdsBidderRule.user_id == int(user.id),
            WbAdsBidderRule.id == int(rule_id),
        )
    )
    if not row:
        raise HTTPException(status_code=404, detail="РџСЂР°РІРёР»Рѕ РЅРµ РЅР°Р№РґРµРЅРѕ")
    raw_patch = payload.model_dump(exclude_unset=True)
    if not raw_patch:
        return WbBidderRuleOut(**serialize_bidder_rule(row))
    current = {
        "campaign_id": row.campaign_id,
        "target_kind": row.target_kind,
        "nm_id": row.nm_id,
        "target_value": row.target_value,
        "placement": row.placement,
        "strategy": row.strategy,
        "desired_bid": row.desired_bid,
        "min_bid": row.min_bid,
        "max_bid": row.max_bid,
        "step_bid": row.step_bid,
        "target_pos_from": row.target_pos_from,
        "target_pos_to": row.target_pos_to,
        "min_clicks": row.min_clicks,
        "is_active": row.is_active,
        "cooldown_sec": row.cooldown_sec,
        "notes": row.notes,
    }
    try:
        normalized = normalize_rule_payload(raw_patch, partial=True, current=current)
    except ValueError as exc:
        raise HTTPException(status_code=400, detail=str(exc))
    apply_rule_payload(row, normalized)
    serialized = serialize_bidder_rule(row)
    _audit(
        db,
        user,
        action="wb_ads_bidder_rule_updated",
        details=f"rule_id={row.id};fields={','.join(sorted(normalized.keys()))[:300]}",
        module_code="wb_ads",
        entity_type="bidder_rule",
        entity_id=str(row.id),
    )
    try:
        db.commit()
    except Exception as exc:
        db.rollback()
        low = str(exc or "").lower()
        if "uq_wb_ads_bidder_rule_target" in low or "unique" in low:
            raise HTTPException(status_code=400, detail="РўР°РєРѕРµ РїСЂР°РІРёР»Рѕ СѓР¶Рµ СЃСѓС‰РµСЃС‚РІСѓРµС‚")
        raise
    return WbBidderRuleOut(**serialized)


@router.delete("/wb/ads/bidder/rules/{rule_id}", response_model=MessageOut)
def wb_ads_bidder_rules_delete(
    rule_id: int,
    user: User = Depends(get_current_user),
    db: Session = Depends(get_db),
):
    ensure_module_enabled(db, user, "wb_ads")
    row = db.scalar(
        select(WbAdsBidderRule).where(
            WbAdsBidderRule.user_id == int(user.id),
            WbAdsBidderRule.id == int(rule_id),
        )
    )
    if not row:
        raise HTTPException(status_code=404, detail="РџСЂР°РІРёР»Рѕ РЅРµ РЅР°Р№РґРµРЅРѕ")
    db.execute(
        delete(WbAdsBidderRun).where(
            WbAdsBidderRun.user_id == int(user.id),
            WbAdsBidderRun.rule_id == int(row.id),
        )
    )
    db.delete(row)
    _audit(
        db,
        user,
        action="wb_ads_bidder_rule_deleted",
        details=f"rule_id={int(rule_id)}",
        module_code="wb_ads",
        entity_type="bidder_rule",
        entity_id=str(rule_id),
    )
    db.commit()
    return MessageOut(message="РџСЂР°РІРёР»Рѕ СѓРґР°Р»РµРЅРѕ")


@router.get("/wb/ads/bidder/runs", response_model=WbBidderRunsOut)
def wb_ads_bidder_runs(
    limit: int = 120,
    user: User = Depends(get_current_user),
    db: Session = Depends(get_db),
):
    ensure_module_enabled(db, user, "wb_ads")
    rows = list_bidder_runs(db, user.id, limit=limit)
    serialized = [serialize_bidder_run(x) for x in rows]
    db.commit()
    return WbBidderRunsOut(
        rows=[WbBidderRunRowOut(**x) for x in serialized],
        meta={
            "count": len(serialized),
            "limit": max(1, min(int(limit or 1), 400)),
        },
    )


@router.post("/wb/ads/bidder/run", response_model=WbBidderRunOut)
def wb_ads_bidder_run(
    payload: WbBidderRunIn,
    user: User = Depends(get_current_user),
    db: Session = Depends(get_db),
):
    ensure_module_enabled(db, user, "wb_ads")
    wb_key = _get_active_marketplace_api_key(db, user.id, "wb")
    if not wb_key:
        raise HTTPException(status_code=400, detail="РЎРЅР°С‡Р°Р»Р° СЃРѕС…СЂР°РЅРёС‚Рµ API РєР»СЋС‡ РґР»СЏ wb")

    ids = sorted({int(x) for x in payload.rule_ids if int(x) > 0})[:200]
    if not payload.force and queue_available():
        queued = enqueue_task(
            "wb_bidder_run",
            {"user_id": int(user.id), "rule_ids": ids, "force": bool(payload.force)},
            dedupe_key=f"wb_bidder_run:{int(user.id)}:{','.join(str(x) for x in ids) or 'all'}",
            dedupe_ttl_sec=90,
        )
        if queued.get("queued"):
            db.commit()
            return WbBidderRunOut(
                ok=True,
                message="Р—Р°РїСѓСЃРє Р±РёРґРґРµСЂР° РїРѕСЃС‚Р°РІР»РµРЅ РІ РѕС‡РµСЂРµРґСЊ",
                results=[],
                meta={
                    "queued": True,
                    "queue_depth": queue_depth(),
                    "queue_available": queue_available(),
                    "rule_ids": ids,
                },
            )

    result = run_bidder_rules(
        db,
        user_id=int(user.id),
        wb_api_key=wb_key,
        rule_ids=ids,
        force=bool(payload.force),
    )
    _audit(
        db,
        user,
        action="wb_ads_bidder_run",
        details=(
            f"rules={len(ids) if ids else 'all'};force={int(payload.force)};"
            f"executed={int((result.get('meta') or {}).get('executed') or 0)};"
            f"changed={int((result.get('meta') or {}).get('changed') or 0)};"
            f"errors={int((result.get('meta') or {}).get('errors') or 0)}"
        ),
        module_code="wb_ads",
        entity_type="bidder_rule",
        status="ok" if bool(result.get("ok")) else "partial",
    )
    db.commit()
    return WbBidderRunOut(
        ok=bool(result.get("ok")),
        message=str(result.get("message") or ""),
        results=list(result.get("results") or []),
        meta=dict(result.get("meta") or {}),
    )


@router.get("/wb/ads/analytics", response_model=WbAdsAnalyticsOut)
def wb_ads_analytics(
    date_from: date | None = None,
    date_to: date | None = None,
    campaign_id: int | None = None,
    offset: int = 0,
    limit: int = 80,
    user: User = Depends(get_current_user),
    db: Session = Depends(get_db),
):
    ensure_module_enabled(db, user, "wb_ads_analytics")
    wb_key = _get_active_marketplace_api_key(db, user.id, "wb")
    if not wb_key:
        raise HTTPException(status_code=400, detail="РЎРЅР°С‡Р°Р»Р° СЃРѕС…СЂР°РЅРёС‚Рµ API РєР»СЋС‡ РґР»СЏ wb")
    base_key = build_market_cache_key(
        {
            "kind": "wb_campaigns_base",
            "key_rev": _secret_revision(wb_key),
        }
    )
    base_error = ""
    try:
        rows, base_cache_meta = get_or_refresh_market_cache(
            db,
            user_id=int(user.id),
            module_code="wb_ads",
            marketplace="wb",
            cache_key=base_key,
            ttl_sec=_market_cache_ttl("wb_ads"),
            fetcher=lambda: fetch_wb_campaigns(wb_key, enrich=False),
            stale_if_error_sec=30 * 60,
        )
    except Exception as exc:
        rows = []
        base_cache_meta = {"source": "error", "age_sec": -1}
        base_error = str(exc or "")[:220]
    rows = list(rows or [])
    base_summary_map: dict[str, dict[str, Any]] = {}
    for row in rows:
        cid_text = _campaign_id_from_any(row)
        cid = _to_int_safe(cid_text)
        if cid <= 0:
            continue
        base_summary_map[str(cid)] = _campaign_summary_from_base_row(row, cid)
    all_ids = sorted({_to_int_safe(x) for x in base_summary_map.keys() if _to_int_safe(x) > 0})
    if campaign_id is not None and campaign_id > 0:
        ids = [campaign_id]
    else:
        safe_offset = max(0, int(offset or 0))
        safe_limit = max(1, min(int(limit or 1), 160))
        ids = all_ids[safe_offset:safe_offset + safe_limit]
    left = date_from.isoformat() if date_from else (date.today() - timedelta(days=6)).isoformat()
    right = date_to.isoformat() if date_to else date.today().isoformat()
    if not ids:
        _audit(
            db,
            user,
            action="wb_ads_analytics_read",
            details=f"date_from={left};date_to={right};campaigns=0;note=no_ids;base_source={base_cache_meta.get('source')}",
            module_code="wb_ads_analytics",
            entity_type="campaign",
        )
        db.commit()
        return WbAdsAnalyticsOut(
            date_from=left,
            date_to=right,
            rows=[],
            totals={"views": 0.0, "clicks": 0.0, "orders": 0.0, "spent": 0.0, "ctr_avg": 0.0, "cr_avg": 0.0},
            meta={
                "requested_count": 0,
                "summary_count": 0,
                "stats_count": 0,
                "temporary_unavailable": False,
                "warnings": ["no_ids"],
                "base_source": str(base_cache_meta.get("source") or ""),
                "base_error": base_error,
            },
        )

    summary_key = build_market_cache_key(
        {
            "kind": "wb_campaign_summaries_analytics",
            "ids": ids,
            "key_rev": _secret_revision(wb_key),
        }
    )
    summary_map: dict[str, dict[str, Any]] = {}
    summary_cache_meta: dict[str, Any] = {"source": "skip", "age_sec": -1}
    summary_error = ""
    try:
        summary_map, summary_cache_meta = get_or_refresh_market_cache(
            db,
            user_id=int(user.id),
            module_code="wb_ads",
            marketplace="wb",
            cache_key=summary_key,
            ttl_sec=max(120, _market_cache_ttl("wb_ads")),
            fetcher=lambda: fetch_wb_campaign_summaries(
                wb_key,
                ids,
                fallback_limit=max(8, min(28, len(ids) // 4 + 8)),
            ),
            stale_if_error_sec=30 * 60,
            prefer_stale_sec=20 * 60,
        )
    except Exception as exc:
        summary_map = {}
        summary_cache_meta = {"source": "error", "age_sec": -1}
        summary_error = str(exc or "")[:220]
    for cid in ids:
        key = str(cid)
        existing = dict(base_summary_map.get(key) or {
            "campaign_id": cid,
            "name": f"РљР°РјРїР°РЅРёСЏ {cid}",
            "status": "-",
            "type": "-",
            "budget": "-",
        })
        summary = summary_map.get(key) if isinstance(summary_map, dict) else None
        if isinstance(summary, dict):
            next_name = str(summary.get("name") or "").strip()
            if next_name and _campaign_name_is_placeholder(str(existing.get("name") or "").strip(), cid):
                existing["name"] = next_name
            next_status = str(summary.get("status") or "").strip()
            if next_status and str(existing.get("status") or "").strip() in {"", "-", "вЂ”"}:
                existing["status"] = next_status
            next_type = str(summary.get("type") or "").strip()
            if next_type and str(existing.get("type") or "").strip() in {"", "-", "вЂ”"}:
                existing["type"] = next_type
            next_budget = str(summary.get("budget") or "").strip()
            if next_budget and str(existing.get("budget") or "").strip() in {"", "-", "вЂ”"}:
                existing["budget"] = next_budget
        base_summary_map[key] = existing

    stats_key = build_market_cache_key(
        {
            "kind": "wb_campaign_stats",
            "ids": ids,
            "date_from": date_from.isoformat() if date_from else "",
            "date_to": date_to.isoformat() if date_to else "",
            "key_rev": _secret_revision(wb_key),
        }
    )
    stats_error = ""
    try:
        stats, stats_cache_meta = get_or_refresh_market_cache(
            db,
            user_id=int(user.id),
            module_code="wb_ads_analytics",
            marketplace="wb",
            cache_key=stats_key,
            ttl_sec=_market_cache_ttl("wb_ads_analytics"),
            fetcher=lambda: fetch_wb_campaign_stats_bulk(
                wb_key,
                ids,
                date_from=date_from.isoformat() if date_from else None,
                date_to=date_to.isoformat() if date_to else None,
            ),
            stale_if_error_sec=30 * 60,
            prefer_stale_sec=20 * 60,
        )
    except Exception as exc:
        stats = {}
        stats_cache_meta = {"source": "error", "age_sec": -1}
        stats_error = str(exc or "")[:220]

    unresolved_stats_ids = [
        cid
        for cid in ids
        if not _campaign_stat_has_context(stats.get(str(cid)))
    ]
    if unresolved_stats_ids:
        single_retry_limit = 12 if len(ids) > 220 else (18 if len(ids) > 80 else 26)
        for cid in unresolved_stats_ids[:single_retry_limit]:
            try:
                one_map = fetch_wb_campaign_stats_bulk(
                    wb_key,
                    [int(cid)],
                    date_from=date_from.isoformat() if date_from else None,
                    date_to=date_to.isoformat() if date_to else None,
                )
            except Exception:
                one_map = {}
            one = one_map.get(str(cid)) if isinstance(one_map, dict) else None
            if isinstance(one, dict) and one:
                stats[str(cid)] = one

    out_rows: list[dict[str, Any]] = []
    partial_summary_ids: list[int] = []
    partial_stats_ids: list[int] = []
    summary_count = 0
    stats_count = 0
    for cid in ids:
        key = str(cid)
        summary = base_summary_map.get(key, {"campaign_id": cid, "name": f"РљР°РјРїР°РЅРёСЏ {cid}", "status": "-", "type": "-", "budget": "-"})
        stat = stats.get(key, {}) if isinstance(stats, dict) else {}
        summary_ok = _campaign_summary_has_context(summary, cid)
        stat_ok = _campaign_stat_has_context(stat)
        if summary_ok:
            summary_count += 1
        else:
            partial_summary_ids.append(cid)
        if stat_ok:
            stats_count += 1
        else:
            partial_stats_ids.append(cid)
        out_rows.append(
            {
                "campaign_id": cid,
                "name": summary.get("name") or f"РљР°РјРїР°РЅРёСЏ {cid}",
                "status": summary.get("status") or "-",
                "type": summary.get("type") or "-",
                "budget": summary.get("budget") or "-",
                "views": _to_float_safe(stat.get("views"), 0.0) if stat.get("views") not in (None, "") else None,
                "clicks": _to_float_safe(stat.get("clicks"), 0.0) if stat.get("clicks") not in (None, "") else None,
                "orders": _to_float_safe(stat.get("orders"), 0.0) if stat.get("orders") not in (None, "") else None,
                "spent": _to_float_safe(stat.get("spent"), 0.0) if stat.get("spent") not in (None, "") else None,
                "ctr": _to_float_safe(stat.get("ctr"), 0.0) if stat.get("ctr") not in (None, "") else None,
                "cr": _to_float_safe(stat.get("cr"), 0.0) if stat.get("cr") not in (None, "") else None,
                "cpc": _to_float_safe(stat.get("cpc"), 0.0) if stat.get("cpc") not in (None, "") else None,
                "cpo": _to_float_safe(stat.get("cpo"), 0.0) if stat.get("cpo") not in (None, "") else None,
                "stat_has_context": stat_ok,
                "summary_has_context": summary_ok,
            }
        )
    out_rows.sort(key=lambda x: _to_float_safe(x.get("spent"), 0.0), reverse=True)
    totals = {
        "views": float(round(sum(_to_float_safe(x.get("views"), 0.0) for x in out_rows if x.get("views") not in (None, "")), 3)),
        "clicks": float(round(sum(_to_float_safe(x.get("clicks"), 0.0) for x in out_rows if x.get("clicks") not in (None, "")), 3)),
        "orders": float(round(sum(_to_float_safe(x.get("orders"), 0.0) for x in out_rows if x.get("orders") not in (None, "")), 3)),
        "spent": float(round(sum(_to_float_safe(x.get("spent"), 0.0) for x in out_rows if x.get("spent") not in (None, "")), 3)),
        "ctr_avg": float(round((sum(_to_float_safe(x.get("ctr"), 0.0) for x in out_rows if x.get("ctr") not in (None, "")) / stats_count) if stats_count else 0.0, 4)),
        "cr_avg": float(round((sum(_to_float_safe(x.get("cr"), 0.0) for x in out_rows if x.get("cr") not in (None, "")) / stats_count) if stats_count else 0.0, 4)),
    }
    warnings: list[str] = []
    if base_error:
        warnings.append("base_fetch_failed")
    if summary_error:
        warnings.append("summary_fetch_failed")
    if stats_error:
        warnings.append("stats_fetch_failed")
    temporary_unavailable = bool(ids) and summary_count <= 0 and stats_count <= 0 and not warnings
    if temporary_unavailable:
        warnings.append("temporary_unavailable")
    if partial_summary_ids:
        warnings.append("summary_partial")
    if partial_stats_ids:
        warnings.append("stats_partial")
    meta = {
        "requested_count": len(ids),
        "summary_count": summary_count,
        "stats_count": stats_count,
        "partial_summary_count": len(partial_summary_ids),
        "partial_stats_count": len(partial_stats_ids),
        "partial_summary_ids": partial_summary_ids[:160],
        "partial_stats_ids": partial_stats_ids[:160],
        "temporary_unavailable": temporary_unavailable,
        "warnings": warnings,
        "base_source": str(base_cache_meta.get("source") or ""),
        "base_age_sec": int(base_cache_meta.get("age_sec") or 0),
        "summary_source": str(summary_cache_meta.get("source") or ""),
        "summary_age_sec": int(summary_cache_meta.get("age_sec") or 0),
        "stats_source": str(stats_cache_meta.get("source") or ""),
        "stats_age_sec": int(stats_cache_meta.get("age_sec") or 0),
        "base_error": base_error,
        "summary_error": summary_error,
        "stats_error": stats_error,
    }

    _audit(
        db,
        user,
        action="wb_ads_analytics_read",
        details=(
            f"date_from={left};date_to={right};campaigns={len(out_rows)};base_source={base_cache_meta.get('source')};"
            f"summary_source={summary_cache_meta.get('source')};stats_source={stats_cache_meta.get('source')};"
            f"summary_ok={summary_count};stats_ok={stats_count};warnings={len(warnings)}"
        ),
        module_code="wb_ads_analytics",
        entity_type="campaign",
        status="ok" if not warnings else "partial",
    )
    db.commit()
    return WbAdsAnalyticsOut(date_from=left, date_to=right, rows=out_rows, totals=totals, meta=meta)

@router.get("/wb/ads/recommendations", response_model=WbAdsRecommendationsOut)
def wb_ads_recommendations(
    date_from: date | None = None,
    date_to: date | None = None,
    min_spent: float = 200.0,
    campaign_id: int | None = None,
    offset: int = 0,
    limit: int = 80,
    user: User = Depends(get_current_user),
    db: Session = Depends(get_db),
):
    ensure_module_enabled(db, user, "wb_ads_recommendations")
    wb_key = _get_active_marketplace_api_key(db, user.id, "wb")
    if not wb_key:
        raise HTTPException(status_code=400, detail="РЎРЅР°С‡Р°Р»Р° СЃРѕС…СЂР°РЅРёС‚Рµ API РєР»СЋС‡ РґР»СЏ wb")

    base_error = ""
    base_key = build_market_cache_key(
        {
            "kind": "wb_campaigns_base",
            "key_rev": _secret_revision(wb_key),
        }
    )
    try:
        base_rows, base_cache_meta = get_or_refresh_market_cache(
            db,
            user_id=int(user.id),
            module_code="wb_ads",
            marketplace="wb",
            cache_key=base_key,
            ttl_sec=_market_cache_ttl("wb_ads"),
            fetcher=lambda: fetch_wb_campaigns(wb_key, enrich=False),
            stale_if_error_sec=30 * 60,
        )
    except Exception as exc:
        base_rows = []
        base_cache_meta = {"source": "error", "age_sec": -1}
        base_error = str(exc or "")
    base_rows = list(base_rows or [])
    base_summary_map: dict[str, dict[str, Any]] = {}
    for row in base_rows:
        cid_text = _campaign_id_from_any(row)
        cid = _to_int_safe(cid_text)
        if cid <= 0:
            continue
        base_summary_map[str(cid)] = _campaign_summary_from_base_row(row, cid)
    all_ids = sorted({_to_int_safe(x) for x in base_summary_map.keys() if _to_int_safe(x) > 0})
    if campaign_id is not None and campaign_id > 0:
        ids = [campaign_id]
        total_available = 1
        slice_offset = 0
        slice_limit = 1
    else:
        safe_offset = max(0, int(offset or 0))
        safe_limit = max(1, min(int(limit or 1), 120))
        ids = all_ids[safe_offset:safe_offset + safe_limit]
        total_available = len(all_ids)
        slice_offset = safe_offset
        slice_limit = safe_limit
    if not ids:
        left = date_from.isoformat() if date_from else (date.today() - timedelta(days=6)).isoformat()
        right = date_to.isoformat() if date_to else date.today().isoformat()
        _audit(
            db,
            user,
            action="wb_ads_recommendations_read",
            details=f"date_from={left};date_to={right};rows=0;min_spent={max(0.0, float(min_spent or 0.0))};note=no_ids",
            module_code="wb_ads_recommendations",
            entity_type="campaign",
        )
        db.commit()
        empty_meta: dict[str, Any] = {
            "min_spent": max(0.0, float(min_spent or 0.0)),
            "campaigns_scanned": 0,
            "total_campaigns": 0,
            "offset": 0,
            "limit": 0,
            "has_more": False,
            "next_offset": None,
        }
        if base_error:
            empty_meta["note"] = f"API campaigns unavailable: {base_error[:220]}"
        return WbAdsRecommendationsOut(
            date_from=left,
            date_to=right,
            rows=[],
            meta=empty_meta,
        )
    stats_error = ""
    stats_key = build_market_cache_key(
        {
            "kind": "wb_campaign_stats",
            "ids": ids,
            "date_from": date_from.isoformat() if date_from else "",
            "date_to": date_to.isoformat() if date_to else "",
            "key_rev": _secret_revision(wb_key),
        }
    )
    try:
        stats, stats_cache_meta = get_or_refresh_market_cache(
            db,
            user_id=int(user.id),
            module_code="wb_ads_recommendations",
            marketplace="wb",
            cache_key=stats_key,
            ttl_sec=_market_cache_ttl("wb_ads_recommendations"),
            fetcher=lambda: fetch_wb_campaign_stats_bulk(
                wb_key,
                ids,
                date_from=date_from.isoformat() if date_from else None,
                date_to=date_to.isoformat() if date_to else None,
            ),
            stale_if_error_sec=30 * 60,
        )
    except Exception as exc:
        stats = {}
        stats_cache_meta = {"source": "error", "age_sec": -1}
        stats_error = str(exc or "")

    safe_min_spent = max(0.0, float(min_spent or 0.0))
    recommendations: list[dict[str, Any]] = []
    fallback_candidates: list[dict[str, Any]] = []

    def _fallback_recommendation(row: dict[str, Any]) -> tuple[str, str, str, str]:
        spent_v = float(row.get("spent") or 0.0)
        views_v = float(row.get("views") or 0.0)
        clicks_v = float(row.get("clicks") or 0.0)
        orders_v = float(row.get("orders") or 0.0)
        ctr_v = float(row.get("ctr") or 0.0)
        cpo_v = float(row.get("cpo") or 0.0)
        is_running_v = bool(row.get("is_running"))

        if spent_v < safe_min_spent and (views_v <= 0 and clicks_v <= 0 and orders_v <= 0):
            return ("РЎРѕР±СЂР°С‚СЊ РґР°РЅРЅС‹Рµ", "low", "monitor", "Р—Р° РїРµСЂРёРѕРґ РїРѕС‡С‚Рё РЅРµС‚ С‚СЂР°С„РёРєР°/СЂР°СЃС…РѕРґР°, СЂР°РЅРѕ РїСЂРёРЅРёРјР°С‚СЊ Р¶РµСЃС‚РєРёРµ СЂРµСЂРµРЅРёСЏ.")
        if spent_v < safe_min_spent and orders_v <= 0:
            return ("РќР°Р±Р»СЋРґР°С‚СЊ Рё РЅР°РєРѕРїРёС‚СЊ СЃС‚Р°С‚РёСЃС‚РёРєСѓ", "low", "monitor", "Р Р°СЃС…РѕРґ РЅРёР¶Рµ РїРѕСЂРѕРіР° РґР»СЏ РЅР°РґРµР¶РЅРѕР№ РѕС†РµРЅРєРё СЌС„С„РµРєС‚РёРІРЅРѕСЃС‚Рё.")
        if views_v > 0 and clicks_v <= 0:
            return ("РџСЂРѕРІРµСЂРёС‚СЊ СЃС‚Р°РІРєРё Рё РєСЂРµР°С‚РёРІ", "medium", "increase_bids", "Р•СЃС‚СЊ РїРѕРєР°Р·С‹, РЅРѕ РЅРµС‚ РєР»РёРєРѕРІ.")
        if clicks_v > 0 and orders_v <= 0:
            return ("РџСЂРѕРІРµСЂРёС‚СЊ РєР°СЂС‚РѕС‡РєСѓ Рё СЃРµРјР°РЅС‚РёРєСѓ", "medium", "optimize_listing", "Р•СЃС‚СЊ РєР»РёРєРё, РЅРѕ РЅРµС‚ Р·Р°РєР°Р·РѕРІ.")
        if orders_v > 0 and cpo_v > 0 and cpo_v <= 1200:
            return ("РџРѕРґРґРµСЂР¶РёРІР°С‚СЊ Рё РјР°СЃС€С‚Р°Р±РёСЂРѕРІР°С‚СЊ", "low", "scale", "РљР°РјРїР°РЅРёСЏ РїСЂРёРЅРѕСЃРёС‚ Р·Р°РєР°Р·С‹ СЃ РїСЂРёРµРјР»РµРјС‹Рј CPO.")
        if orders_v > 0 and cpo_v > 1200:
            return ("РћРїС‚РёРјРёР·РёСЂРѕРІР°С‚СЊ СЂР°СЃС…РѕРґС‹", "medium", "decrease_bids", "Р—Р°РєР°Р·С‹ РµСЃС‚СЊ, РЅРѕ CPO РІС‹СЂРµ С†РµР»РµРІРѕРіРѕ.")
        if (not is_running_v) and (views_v > 0 or clicks_v > 0):
            return ("РџСЂРѕРІРµСЂРёС‚СЊ Рё РІРѕР·РѕР±РЅРѕРІРёС‚СЊ", "low", "start", "РљР°РјРїР°РЅРёСЏ СЃРµР№С‡Р°СЃ РЅРµ Р°РєС‚РёРІРЅР°, РЅРѕ РїРѕ РЅРµР№ Р±С‹Р»Р° Р°РєС‚РёРІРЅРѕСЃС‚СЊ.")
        if ctr_v < 0.3 and views_v > 0:
            return ("РћР±РЅРѕРІРёС‚СЊ РєСЂРµР°С‚РёРІ Рё Р·Р°РіРѕР»РѕРІРєРё", "medium", "refresh", "CTR РЅРёР·РєРёР№ РґР»СЏ С‚РµРєСѓС‰РµРіРѕ РѕР±СЉРµРјР° РїРѕРєР°Р·РѕРІ.")
        return ("Р СѓС‡РЅРѕР№ Р°СѓРґРёС‚", "low", "audit", "РќРµРґРѕСЃС‚Р°С‚РѕС‡РЅРѕ РґР°РЅРЅС‹С… РґР»СЏ Р°РІС‚РѕРјР°С‚РёС‡РµСЃРєРѕРіРѕ СЃС†РµРЅР°СЂРёСЏ.")
    for cid in ids:
        key = str(cid)
        summary = base_summary_map.get(key, {})
        stat = stats.get(key, {})
        views = _to_float_safe(stat.get("views"), 0.0)
        clicks = _to_float_safe(stat.get("clicks"), 0.0)
        orders = _to_float_safe(stat.get("orders"), 0.0)
        spent = _to_float_safe(stat.get("spent"), 0.0)
        ctr = _to_float_safe(stat.get("ctr"), 0.0)
        cpc = _to_float_safe(stat.get("cpc"), 0.0)
        cpo = _to_float_safe(stat.get("cpo"), 0.0)
        status_text = str(summary.get("status") or "-")
        type_text = str(summary.get("type") or "-")
        status_label, is_running = _wb_status_label(status_text)
        type_label = _wb_type_label(type_text)

        recommendation = ""
        priority = "low"
        reason = ""
        action = ""
        base_payload = {
            "campaign_id": cid,
            "name": summary.get("name") or f"РљР°РјРїР°РЅРёСЏ {cid}",
            "status": status_label,
            "type": type_label,
            "is_running": is_running,
            "views": round(views, 3),
            "clicks": round(clicks, 3),
            "orders": round(orders, 3),
            "spent": round(spent, 3),
            "ctr": round(ctr, 4),
            "cpc": round(cpc, 4),
            "cpo": round(cpo, 4),
        }
        if spent >= safe_min_spent and orders <= 0:
            recommendation = "РџР°СѓР·Р° Рё РґРѕСЂР°Р±РѕС‚РєР°"
            priority = "high"
            action = "pause"
            reason = "Р•СЃС‚СЊ СЂР°СЃС…РѕРґ Р·Р° РїРµСЂРёРѕРґ, РЅРѕ РЅРµС‚ Р·Р°РєР°Р·РѕРІ."
        elif spent >= safe_min_spent and clicks >= 20 and ctr < 0.5:
            recommendation = "РџРµСЂРµР·Р°РїСѓСЃРє РєСЂРµР°С‚РёРІР°/СЃРµРјР°РЅС‚РёРєРё"
            priority = "high"
            action = "refresh"
            reason = "РќРёР·РєРёР№ CTR РїСЂРё РґРѕСЃС‚Р°С‚РѕС‡РЅРѕРј РѕР±СЉРµРјРµ РєР»РёРєРѕРІ."
        elif orders >= 3 and cpo > 0 and cpo <= 700 and ctr >= 1.0:
            recommendation = "РњР°СЃСЂС‚Р°Р±РёСЂРѕРІР°С‚СЊ"
            priority = "medium"
            action = "scale"
            reason = "РЎС‚Р°Р±РёР»СЊРЅС‹Рµ Р·Р°РєР°Р·С‹ СЃ РїСЂРёРµРјР»РµРјС‹Рј CPO."
        elif orders > 0 and cpo >= 1800:
            recommendation = "РЎРЅРёР·РёС‚СЊ СЃС‚Р°РІРєРё"
            priority = "medium"
            action = "decrease_bids"
            reason = "Р’С‹СЃРѕРєРёР№ CPO, РєР°РјРїР°РЅРёСЏ РЅРµСЌС„С„РµРєС‚РёРІРЅР°."
        elif (not is_running) and orders >= 2 and cpo > 0 and cpo <= 900:
            recommendation = "Р’РѕР·РѕР±РЅРѕРІРёС‚СЊ РїРѕРєР°Р·С‹"
            priority = "low"
            action = "start"
            reason = "РљР°РјРїР°РЅРёСЏ РЅР° РїР°СѓР·Рµ/Р·Р°РІРµСЂС€РµРЅР°, РЅРѕ РјРµС‚СЂРёРєРё Р±С‹Р»Рё СЂР°Р±РѕС‡РёРµ."

        if not recommendation:
            fallback_candidates.append(base_payload)
            continue
        recommendations.append(
            {
                **base_payload,
                "priority": priority,
                "recommendation": recommendation,
                "action": action,
                "reason": reason,
            }
        )

    fallback_mode = False
    if not recommendations and fallback_candidates:
        fallback_mode = True
        ranked_fallback = sorted(
            fallback_candidates,
            key=lambda row: (
                -float(row.get("spent") or 0),
                -float(row.get("clicks") or 0),
                -float(row.get("views") or 0),
            ),
        )[: min(40, len(fallback_candidates))]
        for row in ranked_fallback:
            recommendation, priority, action, reason = _fallback_recommendation(row)
            recommendations.append(
                {
                    **row,
                    "priority": priority,
                    "recommendation": recommendation,
                    "action": action,
                    "reason": reason,
                }
            )

    # Safety fallback: keep table informative even if metrics API returned sparse/empty data.
    if not recommendations and ids:
        fallback_mode = True
        for cid in ids[: min(30, len(ids))]:
            summary = base_summary_map.get(str(cid), {})
            row = {
                "campaign_id": cid,
                "name": summary.get("name") or f"РљР°РјРїР°РЅРёСЏ {cid}",
                "status": _wb_status_label(str(summary.get("status") or "-"))[0],
                "type": _wb_type_label(str(summary.get("type") or "-")),
                "is_running": _wb_status_label(str(summary.get("status") or "-"))[1],
                "views": 0.0,
                "clicks": 0.0,
                "orders": 0.0,
                "spent": 0.0,
                "ctr": 0.0,
                "cpc": 0.0,
                "cpo": 0.0,
                "priority": "low",
                "recommendation": "РЎРѕР±СЂР°С‚СЊ РґР°РЅРЅС‹Рµ",
                "action": "monitor",
                "reason": "РќРµРґРѕСЃС‚Р°С‚РѕС‡РЅРѕ РґР°РЅРЅС‹С… Р·Р° РїРµСЂРёРѕРґ: РїСЂРѕРІРµСЂСЊС‚Рµ СЃС‚Р°РІРєРё Рё РґР°Р№С‚Рµ РєР°РјРїР°РЅРёРё РЅР°Р±СЂР°С‚СЊ СЃС‚Р°С‚РёСЃС‚РёРєСѓ.",
            }
            recommendations.append(row)

    priority_weight = {"high": 3, "medium": 2, "low": 1}
    recommendations.sort(
        key=lambda row: (
            -priority_weight.get(str(row.get("priority") or "").lower(), 0),
            -float(row.get("spent") or 0),
        )
    )

    left = date_from.isoformat() if date_from else (date.today() - timedelta(days=6)).isoformat()
    right = date_to.isoformat() if date_to else date.today().isoformat()
    _audit(
        db,
        user,
        action="wb_ads_recommendations_read",
        details=f"date_from={left};date_to={right};rows={len(recommendations)};min_spent={safe_min_spent};base_source={base_cache_meta.get('source')};stats_source={stats_cache_meta.get('source')}",
        module_code="wb_ads_recommendations",
        entity_type="campaign",
    )
    db.commit()
    return WbAdsRecommendationsOut(
        date_from=left,
        date_to=right,
        rows=recommendations,
        meta={
            "min_spent": safe_min_spent,
            "campaigns_scanned": len(ids),
            "total_campaigns": total_available,
            "offset": slice_offset,
            "limit": slice_limit,
            "has_more": (slice_offset + slice_limit) < total_available,
            "next_offset": (slice_offset + slice_limit) if (slice_offset + slice_limit) < total_available else None,
            "fallback_mode": fallback_mode,
            "note": (f"partial stats fallback: {stats_error[:220]}") if stats_error else "",
        },
    )


@router.get("/ai/docs", response_model=list[KnowledgeDocOut])
def list_ai_docs(user: User = Depends(get_current_user), db: Session = Depends(get_db)):
    rows = db.scalars(
        select(UserKnowledgeDoc)
        .where(
            UserKnowledgeDoc.user_id == user.id,
            _owned_by_actor_or_owner_filter(UserKnowledgeDoc, user),
        )
        .order_by(UserKnowledgeDoc.id.desc())
    ).all()
    return [
        KnowledgeDocOut(
            id=row.id,
            filename=row.filename,
            content_type=row.content_type,
            size_chars=len(row.content_text or ""),
            created_at=row.created_at.isoformat() if row.created_at else "",
        )
        for row in rows
    ]


@router.post("/ai/docs/upload", response_model=KnowledgeDocOut)
async def upload_ai_doc(
    file: UploadFile = File(...),
    user: User = Depends(get_current_user),
    db: Session = Depends(get_db),
):
    raw = await file.read()
    text = _extract_text_from_upload(file.filename or "document", file.content_type or "", raw)
    if len(text) < 30:
        raise HTTPException(status_code=400, detail="Р”РѕРєСѓРјРµРЅС‚ СЃР»РёСЂРєРѕРј РєРѕСЂРѕС‚РєРёР№ РёР»Рё РЅРµ СѓРґР°Р»РѕСЃСЊ РёР·РІР»РµС‡СЊ С‚РµРєСЃС‚")

    row = UserKnowledgeDoc(
        user_id=user.id,
        filename=(file.filename or "document")[:255],
        content_type=(file.content_type or "application/octet-stream")[:120],
        content_text=text[:160000],
    )
    _assign_owner_member(row, _resolve_owner_member_id(db, user))
    db.add(row)
    _audit(
        db,
        user,
        action="ai_doc_uploaded",
        details=f"filename={row.filename};size={len(row.content_text)}",
        module_code="wb_questions_ai",
        entity_type="knowledge_doc",
    )
    db.commit()
    db.refresh(row)
    return KnowledgeDocOut(
        id=row.id,
        filename=row.filename,
        content_type=row.content_type,
        size_chars=len(row.content_text or ""),
        created_at=row.created_at.isoformat() if row.created_at else "",
    )


@router.delete("/ai/docs/{doc_id}", response_model=MessageOut)
def delete_ai_doc(doc_id: int, user: User = Depends(get_current_user), db: Session = Depends(get_db)):
    row = db.scalar(
        select(UserKnowledgeDoc).where(
            UserKnowledgeDoc.id == doc_id,
            UserKnowledgeDoc.user_id == user.id,
            _owned_by_actor_or_owner_filter(UserKnowledgeDoc, user),
        )
    )
    if not row:
        raise HTTPException(status_code=404, detail="Р”РѕРєСѓРјРµРЅС‚ РЅРµ РЅР°Р№РґРµРЅ")
    fname = row.filename
    db.delete(row)
    _audit(
        db,
        user,
        action="ai_doc_deleted",
        details=f"id={doc_id};filename={fname}",
        module_code="wb_questions_ai",
        entity_type="knowledge_doc",
        entity_id=str(doc_id),
    )
    db.commit()
    return MessageOut(message="Р”РѕРєСѓРјРµРЅС‚ СѓРґР°Р»РµРЅ")


@router.get("/help/docs", response_model=list[HelpDocOut])
def get_help_docs(
    module_code: str = "",
    lang: str = "ru",
    user: User = Depends(get_current_user),
    db: Session = Depends(get_db),
):
    ensure_module_enabled(db, user, "help_center")
    docs = HELP_DOCS_EN if (lang or "").strip().lower() == "en" else HELP_DOCS_RU
    normalized_module = str(module_code or "").strip()
    if normalized_module == "ads_bidder":
        normalized_module = "wb_ads_bidder"
    items = []
    for code, payload in docs.items():
        if code == "dashboard":
            continue
        if normalized_module and code != normalized_module:
            continue
        items.append(HelpDocOut(module_code=code, title=payload["title"], content=payload["content"]))
        if code == "wb_ads_bidder":
            items.append(HelpDocOut(module_code="ads_bidder", title=payload["title"], content=payload["content"]))
    return items


@router.get("/help/releases", response_model=list[HelpReleaseOut])
def get_help_releases(
    lang: str = "ru",
    user: User = Depends(get_current_user),
    db: Session = Depends(get_db),
):
    ensure_module_enabled(db, user, "help_center")
    is_en = (lang or "").strip().lower() == "en"
    out: list[HelpReleaseOut] = []
    for row in HELP_RELEASES:
        item = dict(row)
        if is_en:
            # Lightweight EN localization for release cards without splitting data model.
            version_key = str(item.get("version") or "").strip()
            item["summary"] = {
                "0.4.9": "APK 1.5.15: APK chat now uses a cleaner Telegram-like header/composer flow, mojibake strings are fixed, and module headers no longer jump between background sections.",
                "0.4.8": "APK 1.5.13: one-button Google Calendar connect now works per user account, WB Ads enrich is deeper, and the APK restores Bidder plus cleaner module headers.",
                "0.4.7": "APK 1.5.12: reissued with a new stable release key, so users must uninstall the old app before installing this build.",
                "0.4.6": "APK 1.5.11: calendar sync is stable, WB Ads now shows real names and metrics, and the mobile chat flow is fixed for the APK.",
                "0.4.5": "APK 1.5.10: returns card cleaned up, employee rights tightened, and bidder/help are pinned in the mobile navigation.",
                "0.4.4": "APK 1.5.9: quick navigation to Bidder/Help added, product photo editor improved, and task notifications strengthened.",
                "0.4.3": "APK 1.5.8: Android update installation fixed, product details are faster, and WB Ads enrich now uses faster caching.",
            }.get(version_key, str(item.get("summary") or ""))
            item["notes"] = {
                "0.4.9": "APK 1.5.15 installs over 1.5.12+ builds signed by the release key. If the device still has a pre-1.5.12 build, uninstall once and then install APK 1.5.15.",
                "0.4.8": "APK 1.5.13 installs over 1.5.12 and newer builds signed by the new release key. If the device still has an older pre-1.5.12 build, uninstall it once and then install APK 1.5.13.",
                "0.4.7": "APK 1.5.12 cannot install over the old app because the signing certificate changed. Uninstall the current SEO WIBE app first, then install APK 1.5.12 and sign in again.",
                "0.4.6": "APK 1.5.11 installs over previous versions. After the update, open Social -> Calendar, Ads -> WB Ads Analytics, and any mobile chat to verify sync, real metrics, and the fixed back/reactions flow.",
                "0.4.5": "APK 1.5.10 installs over the previous version. After the update, open Ads -> Bidder, Help, and Reviews/Returns to verify the refreshed interface.",
                "0.4.4": "APK 1.5.9 installs over previous versions. After update, open Ads -> Bidder and Products -> Edit to verify the new elements.",
                "0.4.3": "APK 1.5.8 installs over previous versions. After the update, check product details, WB Ads enrich, and the Android updater flow.",
            }.get(version_key, str(item.get("notes") or ""))
        out.append(HelpReleaseOut(**item))
    return out


@router.get("/mobile/apk/latest", response_model=MobileApkLatestOut)
def get_mobile_apk_latest(request: Request):
    row = next((x for x in HELP_RELEASES if bool(x.get("current"))), None)
    if row is None and HELP_RELEASES:
        row = HELP_RELEASES[0]
    if row is None:
        raise HTTPException(status_code=404, detail="APK release data not found")
    raw_url = str(row.get("android_download_url") or "").strip() or "/static/downloads/seo-wibe-mobile-latest.apk"
    if raw_url.startswith("/"):
        base = _social_public_base_url(request)
        raw_url = f"{base}{raw_url}"
    return MobileApkLatestOut(
        version=str(row.get("version") or ""),
        version_code=max(0, int(row.get("android_version_code") or 0)),
        released_at=str(row.get("released_at") or ""),
        summary=str(row.get("summary") or ""),
        android_download_url=raw_url,
        android_download_name=str(row.get("android_download_name") or "SEO WIBE Mobile Android (.apk)"),
        notes=str(row.get("notes") or ""),
    )


@router.post("/help/assistant", response_model=AiAssistantOut)
def help_assistant(payload: AiAssistantIn, user: User = Depends(get_current_user), db: Session = Depends(get_db)):
    ensure_module_enabled(db, user, "ai_assistant")
    question = " ".join((payload.question or "").split()).strip()
    if len(question) < 3:
        raise HTTPException(status_code=400, detail="Р’РІРµРґРёС‚Рµ Р±РѕР»РµРµ РїРѕРґСЂРѕР±РЅС‹Р№ РІРѕРїСЂРѕСЃ (РјРёРЅРёРјСѓРј 3 СЃРёРјРІРѕР»Р°).")
    module_code = str(payload.module_code or "").strip().lower()
    docs_map = HELP_DOCS_RU
    context_parts: list[str] = []
    if module_code and module_code in docs_map:
        context_parts.append(f"[{docs_map[module_code]['title']}] {docs_map[module_code]['content']}")
    if not context_parts:
        for code, row in docs_map.items():
            if code == "dashboard":
                continue
            context_parts.append(f"[{row['title']}] {row['content']}")
            if len(" ".join(context_parts)) > 12000:
                break
    user_knowledge = _build_user_knowledge_context(
        db,
        user.id,
        max_chars=10000,
        query_text=question,
    )
    if user_knowledge:
        context_parts.append(f"[РџРѕР»СЊР·РѕРІР°С‚РµР»СЊСЃРєР°СЏ Р±Р°Р·Р° Р·РЅР°РЅРёР№]\n{user_knowledge}")
    runtime = _resolve_user_ai_runtime(db, user.id)
    answer = generate_help_assistant_reply(
        question=question,
        context_text="\n\n".join(context_parts),
        prompt="",
        api_key=str(runtime.get("api_key") or ""),
        model=str(runtime.get("model") or ""),
        provider=str(runtime.get("provider") or ""),
        base_url=str(runtime.get("base_url") or ""),
    )
    detail_payload = {
        "module": module_code or "-",
        "provider": runtime.get("provider") or "builtin",
        "mode": runtime.get("mode") or "builtin",
        "knowledge": int(bool(user_knowledge)),
        "question": question[:600],
        "answer": str(answer or "")[:800],
    }
    _audit(
        db,
        user,
        action="help_assistant_asked",
        details=json.dumps(detail_payload, ensure_ascii=False),
        module_code="ai_assistant",
        entity_type="question",
        status="ok",
    )
    db.commit()
    return AiAssistantOut(
        answer=answer,
        provider=str(runtime.get("provider") or "builtin"),
        mode=str(runtime.get("mode") or "builtin"),
        service_name=str(runtime.get("service_name") or ""),
    )


def _resolve_credential(db: Session, user_id: int, marketplace: str) -> ApiCredential:
    cred = db.scalar(
        select(ApiCredential)
        .where(
            ApiCredential.user_id == user_id,
            ApiCredential.marketplace == marketplace,
            ApiCredential.active.is_(True),
        )
        .order_by(ApiCredential.id.desc())
    )
    if not cred:
        raise HTTPException(status_code=400, detail=f"РЎРЅР°С‡Р°Р»Р° СЃРѕС…СЂР°РЅРёС‚Рµ API РєР»СЋС‡ РґР»СЏ {marketplace}")
    return cred


def _resolve_product_marketplaces(value: str) -> list[str]:
    code = str(value or "").strip().lower()
    if code in {"wb", "ozon"}:
        return [code]
    if code == "all":
        return ["wb", "ozon"]
    raise HTTPException(status_code=400, detail="marketplace РґРѕР»Р¶РµРЅ Р±С‹С‚СЊ wb, ozon РёР»Рё all")


def _normalize_product_photo_url(value: str | None) -> str:
    raw = str(value or "").strip()
    if not raw:
        return ""
    if raw.startswith("//"):
        return f"https:{raw}"
    malformed_static = re.match(r"^https?://static/(.+)$", raw, flags=re.IGNORECASE)
    if malformed_static:
        return f"/static/{str(malformed_static.group(1) or '').lstrip('/')}"
    if raw.startswith("http://") or raw.startswith("https://"):
        return raw
    if raw.startswith("/static/"):
        return raw
    if raw.startswith("static/"):
        return f"/{raw}"
    if raw.startswith("/"):
        return raw
    return f"https://{raw.lstrip('/')}"


def _product_photo_identity_key(value: str | None) -> str:
    normalized = _normalize_product_photo_url(value)
    if not normalized:
        return ""
    try:
        parsed = urllib.parse.urlsplit(normalized)
    except Exception:
        return normalized.lower()
    path = str(parsed.path or "")
    path = re.sub(r"/(tm|small|preview|big|orig|x1|x2|c\d+x\d+)/", "/", path, flags=re.IGNORECASE)
    path = re.sub(r"/+", "/", path).rstrip("/")
    return f"{parsed.netloc.lower()}{path.lower()}"


def _product_photo_quality(value: str | None) -> int:
    low = str(value or "").lower()
    score = 0
    if "/orig/" in low or "/big/" in low:
        score += 300
    if re.search(r"/c\d+x\d+/", low):
        score += 220
    if "/x2/" in low:
        score += 180
    if "/x1/" in low:
        score += 160
    if "/tm/" in low or "/small/" in low or "/preview/" in low:
        score += 80
    if "?" not in low:
        score += 5
    return score


def _normalize_product_photo_list(values: Any) -> list[str]:
    source: list[Any]
    if isinstance(values, list):
        source = values
    elif values is None:
        source = []
    else:
        source = [values]
    order: list[str] = []
    chosen: dict[str, tuple[str, int]] = {}
    for item in source:
        normalized = _normalize_product_photo_url(str(item or ""))
        if not normalized:
            continue
        key = _product_photo_identity_key(normalized) or normalized.lower()
        score = _product_photo_quality(normalized)
        prev = chosen.get(key)
        if prev is None:
            chosen[key] = (normalized, score)
            order.append(key)
            continue
        if score > prev[1]:
            chosen[key] = (normalized, score)
    return [chosen[key][0] for key in order if key in chosen]


def _product_photos_from_row(row: Product) -> list[str]:
    raw = str(getattr(row, "photos_json", "") or "").strip()
    if not raw:
        return _normalize_product_photo_list([getattr(row, "photo_url", "")])
    try:
        parsed = json.loads(raw)
    except Exception:
        return _normalize_product_photo_list([getattr(row, "photo_url", "")])
    photos = _normalize_product_photo_list(parsed if isinstance(parsed, list) else [parsed])
    if not photos:
        return _normalize_product_photo_list([getattr(row, "photo_url", "")])
    return photos


@router.post("/products/import", response_model=list[ProductOut])
def import_products(payload: ImportProductsRequest, user: User = Depends(get_current_user), db: Session = Depends(get_db)):
    marketplaces = _resolve_product_marketplaces(payload.marketplace)
    import_all = bool(payload.import_all or not payload.articles)
    owner_member_id = _resolve_owner_member_id(db, user)
    upserted: list[Product] = []
    missing_keys: list[str] = []
    for marketplace in marketplaces:
        cred = db.scalar(
            select(ApiCredential)
            .where(
                ApiCredential.user_id == user.id,
                ApiCredential.marketplace == marketplace,
                ApiCredential.active.is_(True),
            )
            .order_by(ApiCredential.id.desc())
        )
        if not cred:
            missing_keys.append(marketplace)
            continue
        upserted.extend(
            upsert_products(
                db,
                user.id,
                marketplace,
                cred.api_key,
                payload.articles,
                import_all,
                owner_member_id=owner_member_id,
                actor_is_owner=_actor_is_owner(user),
            )
        )
    if not upserted and missing_keys:
        raise HTTPException(status_code=400, detail=f"РЎРЅР°С‡Р°Р»Р° СЃРѕС…СЂР°РЅРёС‚Рµ API РєР»СЋС‡Рё: {', '.join(missing_keys)}")
    _audit(
        db,
        user,
        action="products_imported",
        details=f"count={len(upserted)};marketplaces={','.join(marketplaces)};import_all={1 if import_all else 0};missing={','.join(missing_keys)}",
        module_code="products",
        entity_type="product",
    )
    db.commit()
    return upserted


@router.post("/products/reload", response_model=list[ProductOut])
def reload_products(payload: ProductReloadRequest, user: User = Depends(get_current_user), db: Session = Depends(get_db)):
    marketplaces = _resolve_product_marketplaces(payload.marketplace)
    import_all = True
    missing_keys: list[str] = []
    for marketplace in marketplaces:
        has_key = db.scalar(
            select(ApiCredential.id)
            .where(
                ApiCredential.user_id == user.id,
                ApiCredential.marketplace == marketplace,
                ApiCredential.active.is_(True),
            )
            .order_by(ApiCredential.id.desc())
        )
        if not has_key:
            missing_keys.append(marketplace)
    if len(missing_keys) == len(marketplaces):
        raise HTTPException(status_code=400, detail=f"РЎРЅР°С‡Р°Р»Р° СЃРѕС…СЂР°РЅРёС‚Рµ API РєР»СЋС‡Рё: {', '.join(missing_keys)}")

    owner_member_id = _resolve_owner_member_id(db, user)
    upserted: list[Product] = []
    for marketplace in marketplaces:
        cred = db.scalar(
            select(ApiCredential)
            .where(
                ApiCredential.user_id == user.id,
                ApiCredential.marketplace == marketplace,
                ApiCredential.active.is_(True),
            )
            .order_by(ApiCredential.id.desc())
        )
        if not cred:
            continue
        upserted.extend(
            upsert_products(
                db,
                user.id,
                marketplace,
                cred.api_key,
                payload.articles,
                import_all,
                owner_member_id=owner_member_id,
                actor_is_owner=_actor_is_owner(user),
            )
        )
    _audit(
        db,
        user,
        action="products_reloaded",
        details=f"count={len(upserted)};marketplaces={','.join(marketplaces)};import_all=1;mode=upsert;missing={','.join(missing_keys)}",
        module_code="products",
        entity_type="product",
    )
    db.commit()
    return upserted


@router.post("/products/refresh", response_model=list[ProductOut])
def refresh_products_alias(payload: ProductReloadRequest, user: User = Depends(get_current_user), db: Session = Depends(get_db)):
    # Backward compatibility alias for older frontend builds.
    return reload_products(payload, user, db)


@router.post("/products/reset", response_model=list[ProductOut])
def reset_products_alias(payload: ProductReloadRequest, user: User = Depends(get_current_user), db: Session = Depends(get_db)):
    return reload_products(payload, user, db)


@router.post("/products/reimport", response_model=list[ProductOut])
def reimport_products_alias(payload: ProductReloadRequest, user: User = Depends(get_current_user), db: Session = Depends(get_db)):
    return reload_products(payload, user, db)


@router.get("/products", response_model=ProductPageOut)
def list_products(
    marketplace: str = "all",
    category: str = "all",
    q: str = "",
    page: int = 1,
    page_size: int = 30,
    user: User = Depends(get_current_user),
    db: Session = Depends(get_db),
):
    placeholder_categories = {"", "-", "вЂ”", "n/a", "na", "none", "null", "РЅРµС‚", "unknown"}

    def _normalize_category_value(value: str) -> str:
        text = str(value or "").replace("\u00a0", " ")
        text = " ".join(text.split())
        return text.strip().lower()

    def _is_placeholder_category_value(value: str) -> bool:
        return _normalize_category_value(value) in placeholder_categories

    def _normalized_category_expr():
        raw = func.coalesce(Product.category_name, "")
        # Normalize NBSP + repeated spaces so UI category option matches DB value.
        return func.lower(
            func.trim(
                func.replace(
                    func.replace(
                        func.replace(raw, "\u00a0", " "),
                        "  ",
                        " ",
                    ),
                    "  ",
                    " ",
                )
            )
        )

    def _compact_category_expr():
        return func.replace(_normalized_category_expr(), " ", "")

    safe_market = str(marketplace or "all").strip().lower()
    if safe_market not in {"all", "wb", "ozon"}:
        safe_market = "all"
    safe_category = str(category or "all").strip()
    if not safe_category:
        safe_category = "all"
    safe_category_key = _normalize_category_value(safe_category)
    safe_category_compact = safe_category_key.replace(" ", "")
    safe_q = str(q or "").strip().lower()[:200]
    safe_page = max(1, int(page or 1))
    safe_page_size = int(page_size or 30)
    if safe_page_size not in PRODUCT_PAGE_SIZE_OPTIONS:
        safe_page_size = 30
    normalized_category_name = _normalized_category_expr()
    compact_category_name = _compact_category_expr()
    normalized_category_raw = func.lower(func.trim(func.coalesce(Product.category_name, "")))

    query = select(Product).where(
        Product.user_id == user.id,
        _owned_by_actor_or_owner_filter(Product, user),
    )
    if safe_market != "all":
        query = query.where(Product.marketplace == safe_market)
    if safe_market != "all" and safe_category_key != "all":
        category_match = normalized_category_name == safe_category_key
        if safe_category_compact and safe_category_compact != safe_category_key:
            category_match = or_(category_match, compact_category_name == safe_category_compact)
        query = query.where(category_match)
    if safe_q:
        pattern = f"%{safe_q}%"
        query = query.where(
            or_(
                func.lower(func.coalesce(Product.article, "")).like(pattern),
                func.lower(func.coalesce(Product.name, "")).like(pattern),
                func.lower(func.coalesce(Product.barcode, "")).like(pattern),
                normalized_category_name.like(pattern),
                func.lower(func.coalesce(Product.marketplace, "")).like(pattern),
            )
        )

    if safe_market == "ozon":
        missing_ozon_rows = db.scalars(
            select(Product)
            .where(
                Product.user_id == user.id,
                _owned_by_actor_or_owner_filter(Product, user),
                Product.marketplace == "ozon",
                or_(
                    Product.category_name.is_(None),
                    normalized_category_raw.in_(list(placeholder_categories)),
                ),
            )
            .order_by(Product.id.desc())
            .limit(1200)
        ).all()
        if missing_ozon_rows:
            # 1) Fast local backfill: reuse known category from same article/external_id in DB.
            local_refs: dict[tuple[str, str], str] = {}
            known_rows = db.scalars(
                select(Product)
                .where(
                    Product.user_id == user.id,
                    _owned_by_actor_or_owner_filter(Product, user),
                    Product.marketplace == "ozon",
                    Product.category_name.is_not(None),
                    ~normalized_category_raw.in_(list(placeholder_categories)),
                )
                .order_by(Product.id.desc())
                .limit(4000)
            ).all()
            for item in known_rows:
                article_key = str(item.article or "").strip().lower()
                external_key = str(item.external_id or "").strip()
                category_text = str(item.category_name or "").strip()
                if not category_text:
                    continue
                if article_key:
                    local_refs[(article_key, "")] = category_text
                if external_key:
                    local_refs[("", external_key)] = category_text
                if article_key or external_key:
                    local_refs[(article_key, external_key)] = category_text

            local_changed = False
            for row in missing_ozon_rows:
                article_key = str(row.article or "").strip().lower()
                external_key = str(row.external_id or "").strip()
                inferred = (
                    local_refs.get((article_key, external_key))
                    or local_refs.get((article_key, ""))
                    or local_refs.get(("", external_key))
                    or ""
                )
                if inferred:
                    if str(row.category_name or "").strip() != inferred:
                        row.category_name = inferred[:255]
                        local_changed = True

            ozon_key = _get_active_marketplace_api_key(db, user.id, "ozon")
            if ozon_key:
                refs = [
                    {
                        "article": str(row.article or ""),
                        "external_id": str(row.external_id or ""),
                    }
                    for row in missing_ozon_rows
                ]
                mapped = enrich_ozon_category_names(ozon_key, refs)
                if mapped:
                    changed = False
                    for row in missing_ozon_rows:
                        current_category = str(row.category_name or "").strip()
                        if current_category and not _is_placeholder_category_value(current_category):
                            continue
                        key = (str(row.article or "").strip().lower(), str(row.external_id or "").strip())
                        next_category = str(mapped.get(key) or "").strip()
                        if not next_category:
                            continue
                        if current_category == next_category:
                            continue
                        row.category_name = next_category[:255]
                        changed = True
                    if changed or local_changed:
                        db.flush()
            elif local_changed:
                db.flush()

    categories: list[str] = []
    if safe_market != "all":
        categories_query = (
            select(func.trim(Product.category_name))
            .where(
                Product.user_id == user.id,
                _owned_by_actor_or_owner_filter(Product, user),
                Product.marketplace == safe_market,
                ~normalized_category_raw.in_(list(placeholder_categories)),
            )
            .order_by(func.lower(func.trim(Product.category_name)).asc())
        )
        category_rows = db.scalars(categories_query).all()
        by_normalized: dict[str, str] = {}
        for raw in category_rows:
            label = " ".join(str(raw or "").replace("\u00a0", " ").split()).strip()
            if not label or _is_placeholder_category_value(label):
                continue
            normalized = _normalize_category_value(label)
            if normalized not in by_normalized:
                by_normalized[normalized] = label
        categories = sorted(by_normalized.values(), key=lambda s: s.lower())

    total = int(db.scalar(select(func.count()).select_from(query.subquery())) or 0)
    total_pages = max(1, math.ceil(total / safe_page_size)) if total else 0
    if total_pages and safe_page > total_pages:
        safe_page = total_pages
    offset = max(0, (safe_page - 1) * safe_page_size)
    rows = db.scalars(
        query.order_by(Product.id.desc()).offset(offset).limit(safe_page_size)
    ).all()
    for row in rows:
        stored_photos = _product_photos_from_row(row)
        normalized_main = _normalize_product_photo_url(getattr(row, "photo_url", ""))
        if normalized_main != str(row.photo_url or ""):
            row.photo_url = normalized_main
        if stored_photos and not row.photo_url:
            row.photo_url = stored_photos[0]
        if not row.photo_url:
            row.photo_url = f"https://placehold.co/120x120/e8eefc/1b2a52?text={row.marketplace.upper()}%20{row.id}"
    return ProductPageOut(
        rows=rows,
        categories=categories[:1000],
        total=total,
        page=safe_page,
        page_size=safe_page_size,
        total_pages=total_pages,
    )


@router.get("/products/{product_id}/details", response_model=ProductDetailOut)
def product_details(
    product_id: int,
    refresh: bool = False,
    user: User = Depends(get_current_user),
    db: Session = Depends(get_db),
):
    product = db.scalar(
        select(Product).where(
            Product.id == product_id,
            Product.user_id == user.id,
            _owned_by_actor_or_owner_filter(Product, user),
        )
    )
    if not product:
        raise HTTPException(status_code=404, detail="РўРѕРІР°СЂ РЅРµ РЅР°Р№РґРµРЅ")
    _hydrate_external_id_if_needed(db, user.id, product)
    details_payload: dict[str, Any] = {"photos": [], "attributes": {}, "raw": {}}
    details_cache_meta: dict[str, Any] = {}
    warnings: list[str] = []
    credential = _get_active_marketplace_api_key(db, user.id, product.marketplace)
    if credential:
        detail_cache_key = build_market_cache_key(
            {
                "kind": "product_details",
                "parser_rev": 2,
                "product_id": int(product.id),
                "marketplace": str(product.marketplace or "").strip().lower(),
                "article": str(product.article or "").strip(),
                "external_id": str(product.external_id or "").strip(),
                "key_rev": _secret_revision(credential),
            }
        )
        fetcher = lambda: fetch_marketplace_product_details(
            marketplace=product.marketplace,
            api_key=credential,
            article=product.article,
            external_id=product.external_id or "",
        )
        try:
            if bool(refresh):
                details_payload = fetcher()
                details_cache_meta = {"source": "api-live-forced", "stale": False, "age_sec": 0}
            else:
                details_payload, details_cache_meta = get_or_refresh_market_cache(
                    db,
                    user_id=int(user.id),
                    module_code="products_details",
                    marketplace=str(product.marketplace or "").strip().lower(),
                    cache_key=detail_cache_key,
                    ttl_sec=max(120, _market_cache_ttl("products")),
                    fetcher=fetcher,
                    stale_if_error_sec=60 * 60,
                    prefer_stale_sec=20 * 60,
                )
        except Exception as exc:
            warnings.append(f"Р”РµС‚Р°Р»Рё РёР· API РІСЂРµРјРµРЅРЅРѕ РЅРµРґРѕСЃС‚СѓРїРЅС‹: {str(exc or '')[:220]}")
            details_payload = {"photos": [], "attributes": {}, "raw": {}}
            details_cache_meta = {"source": "error", "stale": True, "age_sec": -1}
        if details_cache_meta.get("stale"):
            warnings.append("РџРѕРєР°Р·Р°РЅС‹ РєСЌСЂРёСЂРѕРІР°РЅРЅС‹Рµ РґРµС‚Р°Р»Рё РєР°СЂС‚РѕС‡РєРё, API РѕР±РЅРѕРІР»СЏРµС‚СЃСЏ РІ С„РѕРЅРµ.")
    else:
        warnings.append(f"API РєР»СЋС‡ {product.marketplace.upper()} РЅРµ РїРѕРґРєР»СЋС‡РµРЅ.")
    live_photos = _normalize_product_photo_list(details_payload.get("photos") if isinstance(details_payload, dict) else [])
    stored_photos = _product_photos_from_row(product)
    resolved_photos = _normalize_product_photo_list(live_photos + stored_photos + [product.photo_url])
    if resolved_photos:
        product.photo_url = resolved_photos[0][:500]
        next_json = json.dumps(resolved_photos, ensure_ascii=False)
        if str(product.photos_json or "") != next_json:
            product.photos_json = next_json
            db.flush()
    raw_payload = details_payload.get("raw") if isinstance(details_payload, dict) else {}
    if not isinstance(raw_payload, dict):
        raw_payload = {"value": str(raw_payload)[:5000]}
    attributes_payload = details_payload.get("attributes") if isinstance(details_payload, dict) else {}
    if not isinstance(attributes_payload, dict):
        attributes_payload = {}

    def _pick_first_price(*values: Any) -> float:
        for value in values:
            amount = float(round(_to_money(value), 2))
            if amount > 0:
                return amount
        return 0.0

    def _value_by_path(node: Any, path: str) -> Any:
        current = node
        for part in str(path or "").split("."):
            key = str(part).strip()
            if not key:
                return None
            if not isinstance(current, dict):
                return None
            current = current.get(key)
        return current

    def _collect_price_values(node: Any, keys: set[str], out: list[Any], depth: int = 0) -> None:
        if node is None or depth > 5:
            return
        if isinstance(node, dict):
            for raw_key, value in node.items():
                key = str(raw_key or "").strip().lower()
                if key in keys:
                    out.append(value)
                if isinstance(value, (dict, list)):
                    _collect_price_values(value, keys, out, depth + 1)
            return
        if isinstance(node, list):
            for item in node[:120]:
                _collect_price_values(item, keys, out, depth + 1)

    price_discount_values: list[Any] = []
    price_base_values: list[Any] = []
    price_min_values: list[Any] = []
    price_marketing_values: list[Any] = []
    _collect_price_values(
        raw_payload,
        {
            "price",
            "sale_price",
            "discount_price",
            "discounted_price",
            "price_with_discount",
            "pricewithdiscount",
            "final_price",
            "current_price",
        },
        price_discount_values,
    )
    _collect_price_values(
        raw_payload,
        {"old_price", "oldprice", "list_price", "base_price", "price_without_discount", "before_discount_price", "original_price"},
        price_base_values,
    )
    _collect_price_values(
        raw_payload,
        {"min_price", "minimum_price", "minprice", "min_ozon_price", "price_min", "auto_action_min_price"},
        price_min_values,
    )
    _collect_price_values(
        raw_payload,
        {"marketing_price", "promo_price", "promotion_price", "action_price", "campaign_price", "recommended_price", "premium_price", "special_price"},
        price_marketing_values,
    )
    for path in (
        "price_info.price",
        "price_info.price.price",
        "price_info.price.discount_price",
        "price_info.price.sale_price",
        "price_info.price.final_price",
        "price_info.price.current_price",
        "price_info.old_price",
        "price_info.price.old_price",
        "price_info.price.list_price",
        "price_info.price.base_price",
        "price_info.price.price_without_discount",
        "price_info.min_price",
        "price_info.price.min_price",
        "price_info.price.min_ozon_price",
        "price_info.price.auto_action_min_price",
        "price_info.marketing_price",
        "price_info.promo_price",
        "price_info.price.marketing_price",
        "price_info.price.promo_price",
        "price_info.price.recommended_price",
        "price_info.price.premium_price",
        "result.price",
        "result.old_price",
        "result.min_price",
        "result.marketing_price",
        "merged_source.price",
        "merged_source.old_price",
        "merged_source.min_price",
        "merged_source.marketing_price",
    ):
        value = _value_by_path(raw_payload, path)
        if value is None:
            continue
        path_low = str(path).lower()
        if any(token in path_low for token in ("old_price", "oldprice", "price_without_discount", "list_price", "base_price", "original_price")):
            price_base_values.append(value)
            continue
        if any(token in path_low for token in ("min_price", "minimum_price", "min_ozon_price", "price_min", "auto_action_min_price")):
            price_min_values.append(value)
            continue
        if any(token in path_low for token in ("marketing_price", "promo_price", "promotion_price", "campaign_price", "recommended_price", "premium_price")):
            price_marketing_values.append(value)
            continue
        price_discount_values.append(value)

    next_category = ""
    if not str(product.category_name or "").strip():
        next_category = str(attributes_payload.get("category_name") or "").strip()
        if not next_category:
            next_category = str(raw_payload.get("category_name") or "").strip()
        if not next_category:
            category_obj = raw_payload.get("category")
            if isinstance(category_obj, dict):
                next_category = str(
                    category_obj.get("category_name")
                    or category_obj.get("name")
                    or category_obj.get("title")
                    or ""
                ).strip()
        if next_category:
            product.category_name = next_category[:255]
            db.flush()
    if float(product.price_discount or 0.0) <= 0:
        product.price_discount = _pick_first_price(
            attributes_payload.get("price"),
            attributes_payload.get("price_with_discount"),
            attributes_payload.get("discount_price"),
            attributes_payload.get("sale_price"),
            attributes_payload.get("final_price"),
            raw_payload.get("price"),
            raw_payload.get("sale_price"),
            raw_payload.get("price_with_discount"),
            raw_payload.get("discount_price"),
            raw_payload.get("final_price"),
            *price_discount_values,
            (raw_payload.get("result") or {}).get("price") if isinstance(raw_payload.get("result"), dict) else None,
        )
    if float(product.price_base or 0.0) <= 0:
        product.price_base = _pick_first_price(
            attributes_payload.get("old_price"),
            attributes_payload.get("list_price"),
            attributes_payload.get("base_price"),
            attributes_payload.get("price_without_discount"),
            attributes_payload.get("original_price"),
            raw_payload.get("old_price"),
            raw_payload.get("price_without_discount"),
            raw_payload.get("list_price"),
            raw_payload.get("base_price"),
            *price_base_values,
        )
    if float(product.price_min or 0.0) <= 0:
        product.price_min = _pick_first_price(
            attributes_payload.get("min_price"),
            attributes_payload.get("minimum_price"),
            attributes_payload.get("min_ozon_price"),
            attributes_payload.get("auto_action_min_price"),
            raw_payload.get("min_price"),
            raw_payload.get("minimum_price"),
            *price_min_values,
        )
    if float(product.price_marketing or 0.0) <= 0:
        product.price_marketing = _pick_first_price(
            attributes_payload.get("marketing_price"),
            attributes_payload.get("campaign_price"),
            attributes_payload.get("promo_price"),
            attributes_payload.get("recommended_price"),
            attributes_payload.get("premium_price"),
            raw_payload.get("marketing_price"),
            raw_payload.get("promo_price"),
            raw_payload.get("promotion_price"),
            *price_marketing_values,
        )
    _audit(
        db,
        user,
        action="product_details_read",
        details=(
            f"product_id={product.id};marketplace={product.marketplace};refresh={1 if refresh else 0};"
            f"source={str(details_cache_meta.get('source') or '-')};age={int(details_cache_meta.get('age_sec') or 0)}"
        ),
        module_code="products",
        entity_type="product",
        entity_id=str(product.id),
    )
    db.commit()
    return ProductDetailOut(
        product=product,
        photos=resolved_photos,
        attributes={str(k): str(v) for k, v in attributes_payload.items()},
        raw=raw_payload,
        warnings=warnings,
    )


@router.patch("/products/{product_id}", response_model=ProductOut)
def update_product(product_id: int, payload: ProductUpdateIn, user: User = Depends(get_current_user), db: Session = Depends(get_db)):
    product = db.scalar(
        select(Product).where(
            Product.id == product_id,
            Product.user_id == user.id,
            _owned_by_actor_or_owner_filter(Product, user),
        )
    )
    if not product:
        raise HTTPException(status_code=404, detail="РўРѕРІР°СЂ РЅРµ РЅР°Р№РґРµРЅ")
    next_name = str(payload.name or "").strip()
    next_barcode = str(payload.barcode or "").strip()
    next_category = str(payload.category_name or "").strip()
    next_description = str(payload.current_description or "").strip()
    prev_description = str(product.current_description or "")
    prev_photos_order = _product_photos_from_row(product)
    next_photo = str(payload.photo_url or "").strip()
    next_keywords = str(payload.target_keywords or "").strip()
    next_purchase_price = max(0.0, round(_to_money(payload.purchase_price), 2)) if payload.purchase_price is not None else None
    next_price_base = max(0.0, round(_to_money(payload.price_base), 2)) if payload.price_base is not None else None
    next_price_discount = max(0.0, round(_to_money(payload.price_discount), 2)) if payload.price_discount is not None else None
    next_price_min = max(0.0, round(_to_money(payload.price_min), 2)) if payload.price_min is not None else None
    next_price_marketing = max(0.0, round(_to_money(payload.price_marketing), 2)) if payload.price_marketing is not None else None
    next_photos_order = (
        _normalize_product_photo_list(payload.photos_order or [])
        if payload.photos_order is not None
        else None
    )
    next_description_trimmed = next_description[:16000]
    description_changed = payload.current_description is not None and next_description_trimmed != prev_description
    photos_order_changed = (
        payload.photos_order is not None
        and (next_photos_order or []) != _normalize_product_photo_list(prev_photos_order)
    )
    if next_name:
        product.name = next_name[:255]
    if payload.barcode is not None:
        product.barcode = next_barcode[:64]
    if payload.category_name is not None:
        product.category_name = next_category[:255]
    if payload.current_description is not None:
        product.current_description = next_description_trimmed
    if next_purchase_price is not None:
        product.purchase_price = next_purchase_price
    if next_price_base is not None:
        product.price_base = next_price_base
    if next_price_discount is not None:
        product.price_discount = next_price_discount
    if next_price_min is not None:
        product.price_min = next_price_min
    if next_price_marketing is not None:
        product.price_marketing = next_price_marketing
    if payload.photo_url is not None:
        normalized_photo = _normalize_product_photo_url(next_photo)
        product.photo_url = normalized_photo[:500]
        if next_photos_order is None:
            photo_list = _normalize_product_photo_list([next_photo])
            product.photos_json = json.dumps(photo_list, ensure_ascii=False) if photo_list else "[]"
    if next_photos_order is not None:
        product.photos_json = json.dumps(next_photos_order, ensure_ascii=False) if next_photos_order else "[]"
        if next_photos_order:
            product.photo_url = str(next_photos_order[0])[:500]
        elif payload.photo_url is None:
            product.photo_url = ""
    if payload.target_keywords is not None:
        product.target_keywords = next_keywords[:5000]

    details = [
        f"product_id={product.id}",
        f"marketplace={product.marketplace}",
        f"name={'1' if bool(next_name) else '0'}",
        f"barcode={'1' if bool(payload.barcode is not None) else '0'}",
        f"category={'1' if bool(payload.category_name is not None) else '0'}",
        f"description={'1' if bool(payload.current_description is not None) else '0'}",
        f"photo={'1' if bool(payload.photo_url is not None) else '0'}",
        f"photos_order={'1' if payload.photos_order is not None else '0'}",
        f"keywords={'1' if bool(payload.target_keywords is not None) else '0'}",
        f"purchase_price={'1' if payload.purchase_price is not None else '0'}",
        f"price_base={'1' if payload.price_base is not None else '0'}",
        f"price_discount={'1' if payload.price_discount is not None else '0'}",
        f"price_min={'1' if payload.price_min is not None else '0'}",
        f"price_marketing={'1' if payload.price_marketing is not None else '0'}",
    ]
    api_key = _get_active_marketplace_api_key(db, user.id, product.marketplace)
    if product.marketplace == "wb" and (payload.current_description is not None or payload.photos_order is not None):
        _hydrate_external_id_if_needed(db, user.id, product)
        details.append(f"wb_external_id={'set' if bool(product.external_id) else 'missing'}")
    if payload.current_description is not None:
        if not description_changed:
            details.append("remote_update=skipped_unchanged")
        elif api_key:
            ok = update_product_description(
                product.marketplace,
                api_key,
                product.article,
                product.current_description,
                external_id=product.external_id,
            )
            details.append(f"remote_update={'ok' if ok else 'failed'}")
        else:
            details.append("remote_update=skipped_no_key")
    if payload.photos_order is not None:
        if not photos_order_changed:
            details.append("remote_photo_order=skipped_unchanged")
        elif api_key and next_photos_order:
            photos_ok, photos_status = update_product_photos_order(
                product.marketplace,
                api_key,
                product.article,
                product.external_id,
                next_photos_order,
            )
            details.append(f"remote_photo_order={'ok' if photos_ok else photos_status}")
        elif not api_key:
            details.append("remote_photo_order=skipped_no_key")
        else:
            details.append("remote_photo_order=empty")
    _audit(
        db,
        user,
        action="product_updated",
        details=";".join(details),
        module_code="products",
        entity_type="product",
        entity_id=str(product.id),
    )
    db.commit()
    db.refresh(product)
    return product


@router.post("/products/{product_id}/photos/upload", response_model=AvatarUploadOut)
def product_photo_upload(
    product_id: int,
    file: UploadFile = File(...),
    user: User = Depends(get_current_user),
    db: Session = Depends(get_db),
):
    product = db.scalar(
        select(Product).where(
            Product.id == product_id,
            Product.user_id == user.id,
            _owned_by_actor_or_owner_filter(Product, user),
        )
    )
    if not product:
        raise HTTPException(status_code=404, detail="РўРѕРІР°СЂ РЅРµ РЅР°Р№РґРµРЅ")
    if not file or not file.filename:
        raise HTTPException(status_code=400, detail="Р¤Р°Р№Р» РЅРµ РІС‹Р±СЂР°РЅ")

    content_type = str(file.content_type or "").lower()
    if not content_type.startswith("image/"):
        raise HTTPException(status_code=400, detail="Р¤Р°Р№Р» РґРѕР»Р¶РµРЅ Р±С‹С‚СЊ РёР·РѕР±СЂР°Р¶РµРЅРёРµРј")

    ext = os.path.splitext(str(file.filename or ""))[1].lower()
    if ext not in {".png", ".jpg", ".jpeg", ".webp", ".gif", ".svg"}:
        ext = ".png"
    raw = file.file.read()
    if not raw:
        raise HTTPException(status_code=400, detail="Р¤Р°Р№Р» РїСѓСЃС‚РѕР№")
    if len(raw) > 8 * 1024 * 1024:
        raise HTTPException(status_code=400, detail="Р¤Р°Р№Р» СЃР»РёСЂРєРѕРј Р±РѕР»СЊСЂРѕР№ (РґРѕ 8 РњР‘)")

    static_root = Path(__file__).resolve().parent.parent / "static"
    target_dir = static_root / "uploads" / "products" / f"user-{int(user.id)}"
    target_dir.mkdir(parents=True, exist_ok=True)
    filename = f"product-{int(product.id)}-{secrets.token_hex(8)}{ext}"
    path = target_dir / filename
    path.write_bytes(raw)
    url = f"/static/uploads/products/user-{int(user.id)}/{filename}"

    existing_photos = _product_photos_from_row(product)
    next_photos = _normalize_product_photo_list([url] + existing_photos)
    product.photos_json = json.dumps(next_photos, ensure_ascii=False) if next_photos else "[]"
    product.photo_url = next_photos[0] if next_photos else url

    _audit(
        db,
        user,
        action="product_photo_uploaded",
        details=f"product_id={product.id};size={len(raw)};name={filename}",
        module_code="products",
        entity_type="product",
        entity_id=str(product.id),
    )
    db.commit()
    return AvatarUploadOut(url=url)


@router.delete("/products/{product_id}/local", response_model=MessageOut)
def delete_product_local(product_id: int, user: User = Depends(get_current_user), db: Session = Depends(get_db)):
    product = db.scalar(
        select(Product).where(
            Product.id == product_id,
            Product.user_id == user.id,
            _owned_by_actor_or_owner_filter(Product, user),
        )
    )
    if not product:
        raise HTTPException(status_code=404, detail="РўРѕРІР°СЂ РЅРµ РЅР°Р№РґРµРЅ")
    marketplace = str(product.marketplace or "").strip().lower()
    article = str(product.article or "").strip()
    name = str(product.name or "").strip()
    db.delete(product)
    _audit(
        db,
        user,
        action="product_local_deleted",
        details=f"product_id={product_id};marketplace={marketplace};article={article};name={name[:255]};local_only=1",
        module_code="products",
        entity_type="product",
        entity_id=str(product_id),
    )
    db.commit()
    return MessageOut(message="РўРѕРІР°СЂ СѓРґР°Р»РµРЅ РёР· Р»РѕРєР°Р»СЊРЅРѕР№ Р±Р°Р·С‹")


@router.post("/products/local/delete", response_model=MessageOut)
def delete_products_local_bulk(payload: ProductBulkDeleteIn, user: User = Depends(get_current_user), db: Session = Depends(get_db)):
    ids = sorted({int(x) for x in (payload.product_ids or []) if int(x) > 0})
    if not ids:
        raise HTTPException(status_code=400, detail="Р’С‹Р±РµСЂРёС‚Рµ С‚РѕРІР°СЂС‹ РґР»СЏ СѓРґР°Р»РµРЅРёСЏ")
    rows = db.scalars(
        select(Product).where(
            Product.user_id == user.id,
            Product.id.in_(ids),
            _owned_by_actor_or_owner_filter(Product, user),
        )
    ).all()
    if not rows:
        raise HTTPException(status_code=404, detail="РўРѕРІР°СЂС‹ РЅРµ РЅР°Р№РґРµРЅС‹")
    deleted_ids: list[int] = []
    for row in rows:
        deleted_ids.append(int(row.id))
        db.delete(row)
    deleted_ids.sort()
    deleted_count = len(deleted_ids)
    requested_count = len(ids)
    missing_count = max(0, requested_count - deleted_count)
    _audit(
        db,
        user,
        action="product_local_bulk_deleted",
        details=(
            f"requested={requested_count};deleted={deleted_count};missing={missing_count};"
            f"ids={','.join(str(x) for x in deleted_ids[:120])}"
        )[:2000],
        module_code="products",
        entity_type="product",
        entity_id=f"bulk:{deleted_count}",
    )
    db.commit()
    return MessageOut(
        message=(
            f"РЈРґР°Р»РµРЅРѕ РёР· Р»РѕРєР°Р»СЊРЅРѕР№ Р±Р°Р·С‹: {deleted_count}"
            if missing_count <= 0
            else f"РЈРґР°Р»РµРЅРѕ РёР· Р»РѕРєР°Р»СЊРЅРѕР№ Р±Р°Р·С‹: {deleted_count} (РЅРµ РЅР°Р№РґРµРЅРѕ: {missing_count})"
        )
    )


@router.get("/products/{product_id}/keyword-suggestions", response_model=list[str])
def product_keyword_suggestions(product_id: int, user: User = Depends(get_current_user), db: Session = Depends(get_db)):
    product = db.scalar(
        select(Product).where(
            Product.id == product_id,
            Product.user_id == user.id,
            _owned_by_actor_or_owner_filter(Product, user),
        )
    )
    if not product:
        raise HTTPException(status_code=404, detail="РўРѕРІР°СЂ РЅРµ РЅР°Р№РґРµРЅ")

    _hydrate_external_id_if_needed(db, user.id, product)
    competitors = find_competitors(
        product.marketplace,
        product.name,
        product.current_description,
        exclude_external_id=product.external_id or "",
    )
    discovered = discover_keywords(
        product.name,
        product.current_description,
        competitors,
        get_user_keywords(db, user.id, product.marketplace, None if _actor_is_owner(user) else int(product.owner_member_id or 0)),
        [],
    )
    primary = _preferred_keyword_from_name(product.name)
    ranked = []
    if primary:
        ranked.append(primary)
    ranked.extend(discovered)
    dedup: list[str] = []
    seen: set[str] = set()
    for kw in ranked:
        k = kw.strip()
        if not k:
            continue
        lk = k.lower()
        if lk in seen:
            continue
        seen.add(lk)
        dedup.append(k)
        if len(dedup) >= 10:
            break
    return dedup


@router.post("/seo/positions/check", response_model=list[PositionCheckOut])
def check_current_positions(payload: PositionCheckRequest, user: User = Depends(get_current_user), db: Session = Depends(get_db)):
    ensure_module_enabled(db, user, "rank_tracking")

    product_ids = payload.product_ids
    if payload.apply_to_all:
        product_ids = db.scalars(
            select(Product.id).where(
                Product.user_id == user.id,
                _owned_by_actor_or_owner_filter(Product, user),
            )
        ).all()
    if not product_ids:
        raise HTTPException(status_code=400, detail="Р’С‹Р±РµСЂРёС‚Рµ С‚РѕРІР°СЂС‹ РґР»СЏ РїСЂРѕРІРµСЂРєРё РїРѕР·РёС†РёР№")

    result: list[PositionCheckOut] = []
    for product_id in product_ids:
        product = db.scalar(
            select(Product).where(
                Product.id == product_id,
                Product.user_id == user.id,
                _owned_by_actor_or_owner_filter(Product, user),
            )
        )
        if not product:
            continue
        _hydrate_external_id_if_needed(db, user.id, product)
        marketplace_api_key = _get_active_marketplace_api_key(db, user.id, product.marketplace)

        explicit_keywords = [x.strip() for x in payload.keywords if x.strip()]
        # Preserve order and remove duplicates to keep deterministic "primary keyword".
        used_keywords = list(dict.fromkeys(explicit_keywords))
        explicit_mode = bool(used_keywords)
        if not used_keywords:
            competitors = find_competitors(
                product.marketplace,
                product.name,
                product.current_description,
                exclude_external_id=product.external_id or "",
            )
            used_keywords = discover_keywords(
                product.name,
                product.current_description,
                competitors,
                get_user_keywords(db, user.id, product.marketplace, None if _actor_is_owner(user) else int(product.owner_member_id or 0)),
                [],
            )[:5]

        keyword_positions = evaluate_positions_for_keywords(
            product.marketplace,
            product.article,
            used_keywords,
            external_id=product.external_id,
            product_name=product.name,
            wb_api_key=marketplace_api_key if product.marketplace == "wb" else "",
        )
        # When user explicitly requested keywords, every keyword should produce a visible position.
        # If parser cannot find exact card for a keyword, mark it as 501 (outside top-500).
        if explicit_mode:
            if not keyword_positions:
                keyword_positions = {kw: 501 for kw in used_keywords}
            else:
                for kw in used_keywords:
                    keyword_positions.setdefault(kw, 501)
        if not keyword_positions:
            fallback_pos = _safe_known_position(product.last_position)
            if fallback_pos == 0 and used_keywords:
                fallback_pos = 501
            product.last_position = fallback_pos
            linked_jobs = db.scalars(
                select(SeoJob).where(
                    SeoJob.user_id == user.id,
                    SeoJob.product_id == product.id,
                    SeoJob.status.in_(["generated", "in_progress", "applied", "top_reached"]),
                    _owned_by_actor_or_owner_filter(SeoJob, user),
                )
            ).all()
            for job in linked_jobs:
                job.current_position = fallback_pos

            result.append(
                PositionCheckOut(
                    product_id=product.id,
                    article=product.article,
                    barcode=product.barcode,
                    name=product.name,
                    used_keywords=used_keywords,
                    best_position=fallback_pos,
                    avg_position=fallback_pos,
                    keyword_positions={kw: fallback_pos for kw in used_keywords},
                )
            )
            continue
        if explicit_mode and used_keywords:
            primary_kw = used_keywords[0]
            best_position = int(keyword_positions.get(primary_kw, 501))
        else:
            best_position = min(keyword_positions.values())
        avg_position = int(round(sum(keyword_positions.values()) / len(keyword_positions)))
        product.last_position = best_position
        linked_jobs = db.scalars(
            select(SeoJob).where(
                SeoJob.user_id == user.id,
                SeoJob.product_id == product.id,
                SeoJob.status.in_(["generated", "in_progress", "applied", "top_reached"]),
                _owned_by_actor_or_owner_filter(SeoJob, user),
            )
        ).all()
        for job in linked_jobs:
            job.current_position = best_position
            if best_position <= job.target_position:
                job.status = "top_reached"
            elif job.status == "top_reached":
                job.status = "in_progress"
            job.next_check_at = schedule_next_check(best_position, job.target_position)
        db.add(
            PositionSnapshot(
                user_id=user.id,
                product_id=product.id,
                source="manual_check",
                position=best_position,
                keywords=", ".join(used_keywords),
            )
        )

        result.append(
            PositionCheckOut(
                product_id=product.id,
                article=product.article,
                barcode=product.barcode,
                name=product.name,
                used_keywords=used_keywords,
                best_position=best_position,
                avg_position=avg_position,
                keyword_positions=keyword_positions,
            )
        )

    _audit(
        db,
        user,
        action="positions_checked",
        details=f"count={len(result)}",
        module_code="rank_tracking",
        entity_type="product",
    )
    db.commit()
    return result


@router.post("/seo/generate", response_model=list[SeoJobOut])
def generate_seo(payload: SeoGenerateRequest, user: User = Depends(get_current_user), db: Session = Depends(get_db)):
    ensure_module_enabled(db, user, "seo_generation")

    product_ids = payload.product_ids
    if payload.apply_to_all:
        product_ids = db.scalars(
            select(Product.id).where(
                Product.user_id == user.id,
                _owned_by_actor_or_owner_filter(Product, user),
            )
        ).all()

    if not product_ids:
        raise HTTPException(status_code=400, detail="РќРµС‚ РІС‹Р±СЂР°РЅРЅС‹С… С‚РѕРІР°СЂРѕРІ РґР»СЏ SEO")

    jobs: list[SeoJob] = []
    for product_id in product_ids:
        product = db.scalar(
            select(Product).where(
                Product.id == product_id,
                Product.user_id == user.id,
                _owned_by_actor_or_owner_filter(Product, user),
            )
        )
        if not product:
            continue
        _hydrate_external_id_if_needed(db, user.id, product)
        marketplace_api_key = _get_active_marketplace_api_key(db, user.id, product.marketplace)

        competitors = find_competitors(
            product.marketplace,
            product.name,
            product.current_description,
            exclude_external_id=product.external_id or "",
        )
        keywords = discover_keywords(
            product.name,
            product.current_description,
            competitors,
            get_user_keywords(db, user.id, product.marketplace, None if _actor_is_owner(user) else int(product.owner_member_id or 0)),
            payload.extra_keywords,
        )
        generated = build_seo_description(product.name, product.current_description, keywords, competitors)
        current_position = evaluate_position(
            product.marketplace,
            product.article,
            keywords,
            external_id=product.external_id,
            product_name=product.name,
            wb_api_key=marketplace_api_key if product.marketplace == "wb" else "",
        )
        if current_position is None:
            current_position = _safe_known_position(product.last_position)
            if current_position == 0 and keywords:
                current_position = 501
        product.target_keywords = ", ".join(keywords)
        product.last_position = current_position
        db.add(
            PositionSnapshot(
                user_id=user.id,
                product_id=product.id,
                source="generate",
                position=current_position,
                keywords=", ".join(keywords),
            )
        )

        job = SeoJob(
            user_id=user.id,
            owner_member_id=int(product.owner_member_id or _resolve_owner_member_id(db, user) or 0) or None,
            product_id=product.id,
            status="generated",
            generated_description=generated,
            keywords_snapshot=", ".join(keywords),
            competitor_snapshot=json.dumps(
                [
                    {
                        "name": c.name,
                        "position": c.position,
                        "keywords": c.keywords[:6],
                        "url": c.url,
                    }
                    for c in competitors[:5]
                ],
                ensure_ascii=False,
            ),
            target_position=payload.target_position,
            current_position=current_position,
            next_check_at=schedule_next_check(current_position, payload.target_position),
        )
        db.add(job)
        jobs.append(job)

    _audit(
        db,
        user,
        action="seo_generated",
        details=f"count={len(jobs)};apply_to_all={payload.apply_to_all}",
        module_code="seo_generation",
        entity_type="seo_job",
    )
    db.commit()
    return [build_seo_job_out(db, x) for x in jobs]


@router.get("/seo/jobs", response_model=list[SeoJobOut])
def list_seo_jobs(user: User = Depends(get_current_user), db: Session = Depends(get_db)):
    jobs = db.scalars(
        select(SeoJob).where(
            SeoJob.user_id == user.id,
            _owned_by_actor_or_owner_filter(SeoJob, user),
        ).order_by(SeoJob.id.desc())
    ).all()
    return [build_seo_job_out(db, x) for x in jobs]


@router.post("/seo/jobs/delete", response_model=MessageOut)
def delete_seo_jobs(payload: SeoDeleteRequest, user: User = Depends(get_current_user), db: Session = Depends(get_db)):
    if payload.delete_all:
        jobs = db.scalars(
            select(SeoJob).where(
                SeoJob.user_id == user.id,
                _owned_by_actor_or_owner_filter(SeoJob, user),
            )
        ).all()
    else:
        if not payload.job_ids:
            raise HTTPException(status_code=400, detail="РЈРєР°Р¶РёС‚Рµ Р·Р°РґР°С‡Рё РґР»СЏ СѓРґР°Р»РµРЅРёСЏ")
        jobs = db.scalars(
            select(SeoJob).where(
                SeoJob.user_id == user.id,
                SeoJob.id.in_(payload.job_ids),
                _owned_by_actor_or_owner_filter(SeoJob, user),
            )
        ).all()

    deleted = 0
    for job in jobs:
        db.delete(job)
        deleted += 1

    _audit(
        db,
        user,
        action="seo_deleted",
        details=f"count={deleted};all={payload.delete_all}",
        module_code="seo_generation",
        entity_type="seo_job",
    )
    db.commit()
    return MessageOut(message=f"РЈРґР°Р»РµРЅРѕ SEO Р·Р°РґР°С‡: {deleted}")


@router.post("/seo/delete", response_model=MessageOut)
def delete_seo_jobs_alias(payload: SeoDeleteRequest, user: User = Depends(get_current_user), db: Session = Depends(get_db)):
    # Backward compatibility alias for older frontend builds.
    return delete_seo_jobs(payload, user, db)


@router.post("/seo/clear", response_model=MessageOut)
def clear_seo_jobs_alias(payload: SeoDeleteRequest, user: User = Depends(get_current_user), db: Session = Depends(get_db)):
    return delete_seo_jobs(payload, user, db)


@router.post("/seo/jobs/clear", response_model=MessageOut)
def clear_seo_jobs_alias_v2(payload: SeoDeleteRequest, user: User = Depends(get_current_user), db: Session = Depends(get_db)):
    return delete_seo_jobs(payload, user, db)


@router.post("/seo/apply", response_model=list[SeoJobOut])
def apply_seo(payload: SeoApplyRequest, user: User = Depends(get_current_user), db: Session = Depends(get_db)):
    ensure_module_enabled(db, user, "auto_apply")

    if not payload.job_ids:
        raise HTTPException(status_code=400, detail="Р’С‹Р±РµСЂРёС‚Рµ С…РѕС‚СЏ Р±С‹ РѕРґРЅСѓ SEO-Р·Р°РґР°С‡Сѓ")

    jobs = db.scalars(
        select(SeoJob).where(
            SeoJob.user_id == user.id,
            SeoJob.id.in_(payload.job_ids),
            _owned_by_actor_or_owner_filter(SeoJob, user),
        )
    ).all()
    if not jobs:
        raise HTTPException(status_code=404, detail="Р—Р°РґР°С‡Рё РЅРµ РЅР°Р№РґРµРЅС‹")

    updated_jobs: list[SeoJob] = []
    for job in jobs:
        _enforce_record_owner_access(job, user)
        product = db.get(Product, job.product_id)
        if not product:
            continue
        _enforce_record_owner_access(product, user)
        _hydrate_external_id_if_needed(db, user.id, product)

        cred = _resolve_credential(db, user.id, product.marketplace)
        ok = update_product_description(
            product.marketplace,
            cred.api_key,
            product.article,
            job.generated_description,
            external_id=product.external_id or "",
        )
        if not ok:
            raise HTTPException(status_code=500, detail="РћС€РёР±РєР° РїСЂРёРјРµРЅРµРЅРёСЏ РёР·РјРµРЅРµРЅРёР№ РІ РјР°СЂРєРµС‚РїР»РµР№СЃ")

        product.current_description = job.generated_description
        job.status = "applied"
        job.next_check_at = schedule_next_check(job.current_position, job.target_position)
        updated_jobs.append(job)

    _audit(
        db,
        user,
        action="seo_applied",
        details=f"count={len(updated_jobs)}",
        module_code="auto_apply",
        entity_type="seo_job",
    )
    db.commit()
    return [build_seo_job_out(db, x) for x in updated_jobs]


@router.post("/seo/recheck", response_model=list[SeoJobOut])
def recheck_seo(payload: SeoRecheckRequest, user: User = Depends(get_current_user), db: Session = Depends(get_db)):
    ensure_module_enabled(db, user, "rank_tracking")

    if payload.recheck_all_due:
        jobs = db.scalars(
            select(SeoJob).where(
                SeoJob.user_id == user.id,
                SeoJob.status.in_(["applied", "in_progress", "generated"]),
                SeoJob.next_check_at.is_not(None),
                SeoJob.next_check_at <= datetime.utcnow(),
                _owned_by_actor_or_owner_filter(SeoJob, user),
            )
        ).all()
    else:
        jobs = db.scalars(
            select(SeoJob).where(
                SeoJob.user_id == user.id,
                SeoJob.id.in_(payload.job_ids),
                _owned_by_actor_or_owner_filter(SeoJob, user),
            )
        ).all()

    result: list[SeoJob] = []
    for job in jobs:
        _enforce_record_owner_access(job, user)
        product = db.get(Product, job.product_id)
        if not product:
            continue
        _enforce_record_owner_access(product, user)
        marketplace_api_key = _get_active_marketplace_api_key(db, user.id, product.marketplace)

        keywords = [k.strip() for k in job.keywords_snapshot.split(",") if k.strip()]
        current_position = evaluate_position(
            product.marketplace,
            product.article,
            keywords,
            external_id=product.external_id,
            product_name=product.name,
            wb_api_key=marketplace_api_key if product.marketplace == "wb" else "",
        )
        if current_position is None:
            current_position = _safe_known_position(job.current_position)
            if current_position == 0:
                current_position = _safe_known_position(product.last_position)
            if current_position == 0 and keywords:
                current_position = 501
        job.current_position = current_position
        product.last_position = current_position
        db.add(
            PositionSnapshot(
                user_id=user.id,
                product_id=product.id,
                source="recheck",
                position=current_position,
                keywords=", ".join(keywords),
            )
        )
        job.attempt_count += 1

        if current_position <= job.target_position:
            job.status = "top_reached"
        else:
            job.status = "in_progress"

        job.next_check_at = schedule_next_check(current_position, job.target_position)
        result.append(job)

    _audit(
        db,
        user,
        action="seo_rechecked",
        details=f"count={len(result)}",
        module_code="rank_tracking",
        entity_type="seo_job",
    )
    db.commit()
    return [build_seo_job_out(db, x) for x in result]


@router.get("/dashboard", response_model=DashboardOut)
def dashboard(user: User = Depends(get_current_user), db: Session = Depends(get_db)):
    actor_id = int(_actor_member_id(user) or 0)
    actor_marker = f"member:{actor_id}" if actor_id > 0 and not _actor_is_owner(user) else "owner"
    cache_key = build_market_cache_key(
        {
            "kind": "dashboard_aggregates",
            "user_id": int(user.id),
            "actor": actor_marker,
        }
    )

    def _load_dashboard_payload() -> dict[str, Any]:
        total_products = db.scalar(
            select(func.count()).select_from(Product).where(
                Product.user_id == user.id,
                _owned_by_actor_or_owner_filter(Product, user),
            )
        ) or 0
        total_jobs = db.scalar(
            select(func.count()).select_from(SeoJob).where(
                SeoJob.user_id == user.id,
                _owned_by_actor_or_owner_filter(SeoJob, user),
            )
        ) or 0
        applied_jobs = db.scalar(
            select(func.count()).select_from(SeoJob).where(
                SeoJob.user_id == user.id,
                SeoJob.status == "applied",
                _owned_by_actor_or_owner_filter(SeoJob, user),
            )
        ) or 0
        in_progress_jobs = db.scalar(
            select(func.count()).select_from(SeoJob).where(
                SeoJob.user_id == user.id,
                SeoJob.status.in_(["in_progress", "generated"]),
                _owned_by_actor_or_owner_filter(SeoJob, user),
            )
        ) or 0
        top5_products = db.scalar(
            select(func.count()).select_from(Product).where(
                Product.user_id == user.id,
                Product.last_position.is_not(None),
                Product.last_position <= 5,
                _owned_by_actor_or_owner_filter(Product, user),
            )
        ) or 0
        return {
            "total_products": int(total_products or 0),
            "total_jobs": int(total_jobs or 0),
            "applied_jobs": int(applied_jobs or 0),
            "in_progress_jobs": int(in_progress_jobs or 0),
            "top5_products": int(top5_products or 0),
        }

    payload, _cache_meta = get_or_refresh_market_cache(
        db,
        user_id=int(user.id),
        module_code="dashboard",
        marketplace="all",
        cache_key=cache_key,
        ttl_sec=max(30, min(120, _market_cache_ttl("dashboard"))),
        fetcher=_load_dashboard_payload,
        stale_if_error_sec=10 * 60,
    )
    safe_payload = payload if isinstance(payload, dict) else {}
    return DashboardOut(
        total_products=int(safe_payload.get("total_products") or 0),
        total_jobs=int(safe_payload.get("total_jobs") or 0),
        applied_jobs=int(safe_payload.get("applied_jobs") or 0),
        in_progress_jobs=int(safe_payload.get("in_progress_jobs") or 0),
        top5_products=int(safe_payload.get("top5_products") or 0),
    )


@router.get("/seo/trend", response_model=TrendOut)
def seo_trend(
    days: int = 21,
    product_id: int | None = None,
    user: User = Depends(get_current_user),
    db: Session = Depends(get_db),
):
    clamped_days = max(3, min(days, 90))
    since = datetime.utcnow() - timedelta(days=clamped_days - 1)
    query = select(PositionSnapshot).where(
        PositionSnapshot.user_id == user.id,
        PositionSnapshot.created_at >= since,
    )
    if not _actor_is_owner(user):
        visible_product_ids = db.scalars(
            select(Product.id).where(
                Product.user_id == user.id,
                _owned_by_actor_or_owner_filter(Product, user),
            )
        ).all()
        if not visible_product_ids:
            return TrendOut(points=[])
        query = query.where(PositionSnapshot.product_id.in_(visible_product_ids))
    if product_id is not None:
        query = query.where(PositionSnapshot.product_id == product_id)

    snapshots = db.scalars(query.order_by(PositionSnapshot.created_at.asc())).all()
    buckets: dict[str, list[int]] = {}
    for snap in snapshots:
        day_key = snap.created_at.strftime("%Y-%m-%d")
        buckets.setdefault(day_key, []).append(int(snap.position))

    points: list[TrendPointOut] = []
    for offset in range(clamped_days):
        day = (since + timedelta(days=offset)).strftime("%Y-%m-%d")
        vals = buckets.get(day, [])
        checks = len(vals)
        avg = round(sum(vals) / checks, 2) if checks else 0.0
        top5 = sum(1 for x in vals if x <= 5)
        points.append(TrendPointOut(date=day, checks=checks, avg_position=avg, top5_hits=top5))

    return TrendOut(points=points)


@router.get("/sales/stats", response_model=SalesStatsOut)
def sales_stats(
    marketplace: str = "all",
    date_from: date | None = None,
    date_to: date | None = None,
    granularity: str = "auto",
    tz: str = "UTC",
    force_refresh: bool = False,
    user: User = Depends(get_current_user),
    db: Session = Depends(get_db),
):
    ensure_module_enabled(db, user, "sales_stats")
    selected_market = (marketplace or "all").strip().lower()
    if selected_market not in {"all", "wb", "ozon"}:
        raise HTTPException(status_code=400, detail="marketplace РґРѕР»Р¶РµРЅ Р±С‹С‚СЊ all, wb РёР»Рё ozon")

    right = date_to or date.today()
    left = date_from or right
    if left > right:
        left, right = right, left
    if (right - left).days > 365:
        left = right - timedelta(days=365)

    gran = str(granularity or "auto").strip().lower()
    if gran not in {"auto", "hour", "day"}:
        raise HTTPException(status_code=400, detail="granularity РґРѕР»Р¶РµРЅ Р±С‹С‚СЊ auto, hour РёР»Рё day")
    tz_name = str(tz or "UTC").strip() or "UTC"
    try:
        ZoneInfo(tz_name)
    except Exception:
        tz_name = "UTC"

    wb_key = _get_active_marketplace_api_key(db, user.id, "wb")
    ozon_key = _get_active_marketplace_api_key(db, user.id, "ozon")
    key_rev = _secret_revision(wb_key, ozon_key)
    live_window = bool(left <= datetime.now(ZoneInfo(tz_name)).date() <= right)
    live_market = selected_market in {"all", "wb"}
    use_live_mode = bool(live_window and live_market)
    sales_ttl_sec = max(90, min(_market_cache_ttl("sales_stats"), 180)) if use_live_mode else _market_cache_ttl("sales_stats")
    sales_prefer_stale_sec = 3 * 60 if use_live_mode and not force_refresh else 0
    sales_cache_key = build_market_cache_key(
        {
            "marketplace": selected_market,
            "date_from": left.isoformat(),
            "date_to": right.isoformat(),
            "granularity": gran,
            "tz": tz_name,
            "key_rev": key_rev,
        }
    )

    def _load_sales_payload(
        period_from: date,
        period_to: date,
        *,
        prefer_live: bool = False,
        force_fresh_wb: bool = False,
    ) -> dict[str, Any]:
        return build_sales_report(
            marketplace=selected_market,
            date_from=period_from,
            date_to=period_to,
            wb_api_key=wb_key,
            ozon_api_key=ozon_key,
            granularity=gran,
            timezone=tz_name,
            prefer_live=prefer_live,
            force_fresh_wb=force_fresh_wb,
        )

    try:
        payload, sales_cache_meta = get_or_refresh_market_cache(
            db,
            user_id=int(user.id),
            module_code="sales_stats",
            marketplace=selected_market,
            cache_key=sales_cache_key,
            ttl_sec=sales_ttl_sec,
            fetcher=lambda: _load_sales_payload(
                left,
                right,
                prefer_live=use_live_mode,
                force_fresh_wb=bool(force_refresh),
            ),
            stale_if_error_sec=20 * 60,
            prefer_stale_sec=sales_prefer_stale_sec,
            force_refresh=bool(force_refresh),
        )
    except Exception as exc:
        sales_cache_meta = {"source": "error", "age_sec": -1}
        payload = {
            "rows": [],
            "chart": [],
            "totals": {
                "orders": 0,
                "units": 0,
                "buyouts": 0,
                "revenue": 0.0,
                "returns": 0,
                "ad_spend": 0.0,
                "penalties": 0.0,
                "days": 0,
                "gross_profit": 0.0,
            },
            "warnings": [f"Sales stats load failed: {str(exc or '')[:220]}"],
            "granularity": "day",
            "timezone": tz_name,
        }

    rows = payload.get("rows") if isinstance(payload, dict) else []
    chart = payload.get("chart") if isinstance(payload, dict) else []
    totals = payload.get("totals") if isinstance(payload, dict) else {}
    warnings = payload.get("warnings") if isinstance(payload, dict) else []
    rows = rows if isinstance(rows, list) else []
    chart = chart if isinstance(chart, list) else []
    totals = totals if isinstance(totals, dict) else {}
    warnings = warnings if isinstance(warnings, list) else []
    if (not _sales_payload_has_data(payload)) and _warnings_indicate_upstream_failure(warnings):
        fallback_payload, fallback_meta = _market_cache_latest_payload(
            db,
            user_id=int(user.id),
            module_code="sales_stats",
            marketplace=selected_market,
            max_age_sec=7 * 24 * 60 * 60,
            exclude_cache_keys={str(sales_cache_meta.get("cache_key") or "").strip()},
            scan_limit=140,
        )
        if _sales_payload_has_data(fallback_payload):
            payload = fallback_payload or {}
            sales_cache_meta = fallback_meta or {"source": "db-latest-module-fallback", "age_sec": -1}
            rows = payload.get("rows") if isinstance(payload, dict) else []
            chart = payload.get("chart") if isinstance(payload, dict) else []
            totals = payload.get("totals") if isinstance(payload, dict) else {}
            warnings = payload.get("warnings") if isinstance(payload, dict) else []
            rows = rows if isinstance(rows, list) else []
            chart = chart if isinstance(chart, list) else []
            totals = totals if isinstance(totals, dict) else {}
            warnings = warnings if isinstance(warnings, list) else []
            warnings.append("РџРѕРєР°Р·Р°РЅС‹ РїРѕСЃР»РµРґРЅРёРµ СЃС‚Р°Р±РёР»СЊРЅС‹Рµ РґР°РЅРЅС‹Рµ: С‚РµРєСѓС‰РёР№ Р·Р°РїСЂРѕСЃ Рє API РІРµСЂРЅСѓР» РїСѓСЃС‚РѕР№ РѕС‚РІРµС‚.")
    warm_result = _enqueue_sales_cache_warmup(
        int(user.id),
        marketplace=selected_market,
        date_from=left,
        date_to=right,
        granularity=gran,
        tz_name=tz_name,
    )
    comparison: dict[str, Any] = {}
    comparison_rows: list[dict[str, Any]] = []
    comparison_chart: list[dict[str, Any]] = []
    period_days = max(1, (right - left).days + 1)
    comparison_limit_days = 31 if force_refresh else 45
    if period_days > comparison_limit_days:
        warnings.append(
            f"РЎСЂР°РІРЅРµРЅРёРµ СЃ РїСЂРµРґС‹РґСѓС‰РёРј РїРµСЂРёРѕРґРѕРј РѕС‚РєР»СЋС‡РµРЅРѕ РґР»СЏ РґРёР°РїР°Р·РѕРЅР° Р±РѕР»РµРµ {comparison_limit_days} РґРЅРµР№ (СѓСЃРєРѕСЂРµРЅРЅС‹Р№ СЂРµР¶РёРј)."
        )
        prev_cache_meta = {"source": "skipped-long-period"}
    else:
        try:
            prev_to = left - timedelta(days=1)
            prev_from = prev_to - timedelta(days=period_days - 1)
            prev_cache_key = build_market_cache_key(
                {
                    "marketplace": selected_market,
                    "date_from": prev_from.isoformat(),
                    "date_to": prev_to.isoformat(),
                    "granularity": gran,
                    "tz": tz_name,
                    "key_rev": key_rev,
                    "comparison": 1,
                }
            )
            prev_payload, prev_cache_meta = get_or_refresh_market_cache(
                db,
                user_id=int(user.id),
                module_code="sales_stats",
                marketplace=selected_market,
                cache_key=prev_cache_key,
                ttl_sec=_market_cache_ttl("sales_stats"),
                fetcher=lambda: _load_sales_payload(prev_from, prev_to, prefer_live=False, force_fresh_wb=False),
                stale_if_error_sec=45 * 60,
                prefer_stale_sec=12 * 60 * 60,
            )
            prev_totals = prev_payload.get("totals") if isinstance(prev_payload, dict) else {}
            prev_totals = prev_totals if isinstance(prev_totals, dict) else {}
            prev_rows = prev_payload.get("rows") if isinstance(prev_payload, dict) else []
            prev_chart = prev_payload.get("chart") if isinstance(prev_payload, dict) else []
            comparison_rows = prev_rows if isinstance(prev_rows, list) else []
            comparison_chart = prev_chart if isinstance(prev_chart, list) else []

            def _cmp(metric: str) -> dict[str, Any]:
                cur = float(totals.get(metric) or 0.0)
                prev = float(prev_totals.get(metric) or 0.0)
                if abs(prev) < 1e-9:
                    delta_pct = 100.0 if abs(cur) > 1e-9 else 0.0
                else:
                    delta_pct = ((cur - prev) / abs(prev)) * 100.0
                return {
                    "current": cur,
                    "previous": prev,
                    "delta": cur - prev,
                    "delta_pct": round(delta_pct, 2),
                }

            comparison = {
                "period_days": period_days,
                "current_from": left.isoformat(),
                "current_to": right.isoformat(),
                "previous_from": prev_from.isoformat(),
                "previous_to": prev_to.isoformat(),
                "metrics": {
                    "orders": _cmp("orders"),
                    "units": _cmp("units"),
                    "buyouts": _cmp("buyouts"),
                    "order_amount": _cmp("order_amount"),
                    "buyout_amount": _cmp("buyout_amount"),
                    "revenue": _cmp("revenue"),
                    "returns": _cmp("returns"),
                    "ad_spend": _cmp("ad_spend"),
                    "penalties": _cmp("penalties"),
                    "income": _cmp("income"),
                    "expense": _cmp("expense"),
                    "net": _cmp("net"),
                    "commission": _cmp("commission"),
                    "logistics": _cmp("logistics"),
                    "storage": _cmp("storage"),
                    "deductions": _cmp("deductions"),
                    "acceptance": _cmp("acceptance"),
                    "other_expense": _cmp("other_expense"),
                    "gross_profit": _cmp("gross_profit"),
                },
            }
            _enqueue_sales_cache_warmup(
                int(user.id),
                marketplace=selected_market,
                date_from=prev_from,
                date_to=prev_to,
                granularity=gran,
                tz_name=tz_name,
            )
        except Exception as exc:
            comparison = {}
            warnings.append(f"РЎСЂР°РІРЅРµРЅРёРµ СЃ РїСЂРµРґС‹РґСѓС‰РёРј РїРµСЂРёРѕРґРѕРј РЅРµРґРѕСЃС‚СѓРїРЅРѕ: {str(exc or '')[:140]}")
            prev_cache_meta = {"source": "error"}

    report_granularity = str(payload.get("granularity") or ("hour" if gran == "hour" else "day"))
    report_tz = str(payload.get("timezone") or tz_name)
    cache_stats = get_market_cache_stats(db, user_id=int(user.id), module_code="sales_stats")
    _audit(
        db,
        user,
        action="sales_stats_read",
        details=(
            f"market={selected_market};from={left.isoformat()};to={right.isoformat()};rows={len(rows)};"
            f"granularity={report_granularity};tz={report_tz};comparison={1 if comparison else 0};"
            f"force={int(bool(force_refresh))};"
            f"source={sales_cache_meta.get('source')};prev_source={prev_cache_meta.get('source')};"
            f"cache_entries={cache_stats.get('entries')};cache_hits={cache_stats.get('hits')};"
            f"cache_refreshes={cache_stats.get('refreshes')};warm_queued={int(bool(warm_result.get('queued')))}"
        ),
        module_code="sales_stats",
        entity_type="sales",
    )
    db.commit()
    return SalesStatsOut(
        marketplace=selected_market,
        date_from=left.isoformat(),
        date_to=right.isoformat(),
        granularity=report_granularity,
        timezone=report_tz,
        rows=rows,
        chart=chart,
        comparison_rows=comparison_rows,
        comparison_chart=comparison_chart,
        totals=totals,
        comparison=comparison,
        warnings=[str(x) for x in warnings],
    )


@router.get("/accounting/settings", response_model=AccountingSettingsOut)
def accounting_settings_get(user: User = Depends(get_current_user), db: Session = Depends(get_db)):
    ensure_module_enabled(db, user, "accounting")
    row = _get_or_create_accounting_settings(db, user)
    db.commit()
    return _accounting_settings_to_out(row)


@router.put("/accounting/settings", response_model=AccountingSettingsOut)
def accounting_settings_update(payload: AccountingSettingsIn, user: User = Depends(get_current_user), db: Session = Depends(get_db)):
    ensure_module_enabled(db, user, "accounting")
    _require_owner_actor(user)
    row = _get_or_create_accounting_settings(db, user)
    row.vat_rate = max(0.0, min(100.0, _to_money(payload.vat_rate)))
    row.tax_rate = max(0.0, min(100.0, _to_money(payload.tax_rate)))
    row.additional_rate = max(0.0, min(100.0, _to_money(payload.additional_rate)))
    row.fixed_cost_per_month = max(0.0, _to_money(payload.fixed_cost_per_month))
    _audit(
        db,
        user,
        action="accounting_settings_updated",
        details=(
            f"vat={row.vat_rate};tax={row.tax_rate};additional={row.additional_rate};"
            f"fixed={row.fixed_cost_per_month}"
        ),
        module_code="accounting",
        entity_type="accounting_settings",
        entity_id=str(row.id),
    )
    db.commit()
    db.refresh(row)
    return _accounting_settings_to_out(row)


@router.get("/accounting/expenses", response_model=AccountingExpenseListOut)
def accounting_expenses_get(user: User = Depends(get_current_user), db: Session = Depends(get_db)):
    ensure_module_enabled(db, user, "accounting")
    rows = db.scalars(
        select(AccountingExpense)
        .where(AccountingExpense.user_id == user.id)
        .order_by(AccountingExpense.id.desc())
    ).all()
    db.commit()
    return AccountingExpenseListOut(rows=[_accounting_expense_to_out(x) for x in rows])


@router.post("/accounting/expenses", response_model=AccountingExpenseOut)
def accounting_expense_create(payload: AccountingExpenseIn, user: User = Depends(get_current_user), db: Session = Depends(get_db)):
    ensure_module_enabled(db, user, "accounting")
    _require_owner_actor(user)
    category = " ".join(str(payload.category or "").split()).strip()
    if not category:
        raise HTTPException(status_code=400, detail="РЈРєР°Р¶РёС‚Рµ РєР°С‚РµРіРѕСЂРёСЋ СЂР°СЃС…РѕРґР°")
    amount = max(0.0, _to_money(payload.amount))
    start_date = _parse_datetime_or_none(payload.start_date)
    end_date = _parse_datetime_or_none(payload.end_date)
    if start_date and end_date and start_date > end_date:
        start_date, end_date = end_date, start_date
    row = AccountingExpense(
        user_id=user.id,
        marketplace=_normalize_accounting_marketplace(payload.marketplace),
        category=category[:120],
        amount=amount,
        recurrence=_normalize_expense_recurrence(payload.recurrence),
        start_date=start_date,
        end_date=end_date,
        note=str(payload.note or "")[:1000],
        is_active=bool(payload.is_active),
    )
    db.add(row)
    db.flush()
    _audit(
        db,
        user,
        action="accounting_expense_created",
        details=(
            f"id={row.id};marketplace={row.marketplace};category={row.category};"
            f"amount={row.amount};recurrence={row.recurrence};active={1 if row.is_active else 0}"
        )[:2000],
        module_code="accounting",
        entity_type="accounting_expense",
        entity_id=str(row.id),
    )
    db.commit()
    db.refresh(row)
    return _accounting_expense_to_out(row)


@router.put("/accounting/expenses/{expense_id}", response_model=AccountingExpenseOut)
def accounting_expense_update(
    expense_id: int,
    payload: AccountingExpenseIn,
    user: User = Depends(get_current_user),
    db: Session = Depends(get_db),
):
    ensure_module_enabled(db, user, "accounting")
    _require_owner_actor(user)
    row = db.scalar(select(AccountingExpense).where(AccountingExpense.id == expense_id, AccountingExpense.user_id == user.id))
    if not row:
        raise HTTPException(status_code=404, detail="Р Р°СЃС…РѕРґ РЅРµ РЅР°Р№РґРµРЅ")
    category = " ".join(str(payload.category or "").split()).strip()
    if not category:
        raise HTTPException(status_code=400, detail="РЈРєР°Р¶РёС‚Рµ РєР°С‚РµРіРѕСЂРёСЋ СЂР°СЃС…РѕРґР°")
    row.marketplace = _normalize_accounting_marketplace(payload.marketplace)
    row.category = category[:120]
    row.amount = max(0.0, _to_money(payload.amount))
    row.recurrence = _normalize_expense_recurrence(payload.recurrence)
    row.start_date = _parse_datetime_or_none(payload.start_date)
    row.end_date = _parse_datetime_or_none(payload.end_date)
    if row.start_date and row.end_date and row.start_date > row.end_date:
        row.start_date, row.end_date = row.end_date, row.start_date
    row.note = str(payload.note or "")[:1000]
    row.is_active = bool(payload.is_active)
    _audit(
        db,
        user,
        action="accounting_expense_updated",
        details=(
            f"id={row.id};marketplace={row.marketplace};category={row.category};"
            f"amount={row.amount};recurrence={row.recurrence};active={1 if row.is_active else 0}"
        )[:2000],
        module_code="accounting",
        entity_type="accounting_expense",
        entity_id=str(row.id),
    )
    db.commit()
    db.refresh(row)
    return _accounting_expense_to_out(row)


@router.delete("/accounting/expenses/{expense_id}", response_model=MessageOut)
def accounting_expense_delete(expense_id: int, user: User = Depends(get_current_user), db: Session = Depends(get_db)):
    ensure_module_enabled(db, user, "accounting")
    _require_owner_actor(user)
    row = db.scalar(select(AccountingExpense).where(AccountingExpense.id == expense_id, AccountingExpense.user_id == user.id))
    if not row:
        raise HTTPException(status_code=404, detail="Р Р°СЃС…РѕРґ РЅРµ РЅР°Р№РґРµРЅ")
    db.delete(row)
    _audit(
        db,
        user,
        action="accounting_expense_deleted",
        details=f"id={expense_id};category={str(row.category or '')[:120]}",
        module_code="accounting",
        entity_type="accounting_expense",
        entity_id=str(expense_id),
    )
    db.commit()
    return MessageOut(message="Р Р°СЃС…РѕРґ СѓРґР°Р»РµРЅ")


@router.get("/accounting/data", response_model=AccountingDataOut)
def accounting_data(
    marketplace: str = "all",
    date_from: date | None = None,
    date_to: date | None = None,
    granularity: str = "auto",
    tz: str = "UTC",
    q: str = "",
    sort_by: str = "net_profit_desc",
    fast: bool = True,
    user: User = Depends(get_current_user),
    db: Session = Depends(get_db),
):
    ensure_module_enabled(db, user, "accounting")
    selected_market = _normalize_accounting_marketplace(marketplace)
    right = date_to or date.today()
    left = date_from or right
    if left > right:
        left, right = right, left
    if (right - left).days > 365:
        left = right - timedelta(days=365)
    gran = str(granularity or "auto").strip().lower()
    if gran not in {"auto", "hour", "day"}:
        gran = "auto"
    tz_name = str(tz or "UTC").strip() or "UTC"
    try:
        ZoneInfo(tz_name)
    except Exception:
        tz_name = "UTC"

    wb_key = _get_active_marketplace_api_key(db, user.id, "wb")
    ozon_key = _get_active_marketplace_api_key(db, user.id, "ozon")
    settings_row = _get_or_create_accounting_settings(db, user)
    settings_payload = {
        "vat_rate": float(round(settings_row.vat_rate or 0.0, 2)),
        "tax_rate": float(round(settings_row.tax_rate or 0.0, 2)),
        "additional_rate": float(round(settings_row.additional_rate or 0.0, 2)),
        "fixed_cost_per_month": float(round(settings_row.fixed_cost_per_month or 0.0, 2)),
    }
    expense_rows = db.scalars(
        select(AccountingExpense)
        .where(AccountingExpense.user_id == user.id)
        .order_by(AccountingExpense.id.asc())
    ).all()
    products_payload = _collect_product_cost_payload(db, user, selected_market)
    expenses_payload = _collect_accounting_expense_payload(expense_rows)
    expense_sig_raw = "|".join(
        [
            ";".join(
                [
                    str(int(row.id)),
                    _normalize_accounting_marketplace(row.marketplace),
                    f"{float(round(row.amount or 0.0, 2)):.2f}",
                    row.period_from.isoformat(),
                    row.period_to.isoformat(),
                    "1" if bool(row.is_active) else "0",
                    row.updated_at.isoformat() if row.updated_at else "",
                ]
            )
            for row in expense_rows
        ]
    )
    products_sig_raw = "|".join(
        sorted(
            [
                ";".join(
                    [
                        str(x.get("marketplace") or ""),
                        str(x.get("external_id") or ""),
                        str(x.get("article") or ""),
                        f"{float(round(_to_float_safe(x.get('purchase_price'), 0.0), 2)):.2f}",
                    ]
                )
                for x in products_payload
            ]
        )
    )
    cache_key = build_market_cache_key(
        {
            "kind": "accounting_data",
            "market": selected_market,
            "date_from": left.isoformat(),
            "date_to": right.isoformat(),
            "granularity": gran,
            "tz": tz_name,
            "q": str(q or "").strip().lower(),
            "sort_by": str(sort_by or "net_profit_desc").strip().lower(),
            "wb_key_rev": _secret_revision(wb_key),
            "ozon_key_rev": _secret_revision(ozon_key),
            "settings": settings_payload,
            "expense_sig": hashlib.sha1(expense_sig_raw.encode("utf-8")).hexdigest(),
            "products_sig": hashlib.sha1(products_sig_raw.encode("utf-8")).hexdigest(),
        }
    )
    previous_same_key_payload: dict[str, Any] | None = None
    prev_same_key_row = db.scalar(
        select(MarketplaceApiCache).where(
            MarketplaceApiCache.user_id == int(user.id),
            MarketplaceApiCache.module_code == "accounting",
            MarketplaceApiCache.marketplace == selected_market[:30],
            MarketplaceApiCache.cache_key == cache_key,
        )
    )
    if prev_same_key_row and str(prev_same_key_row.payload_json or "").strip():
        try:
            parsed_prev = json.loads(str(prev_same_key_row.payload_json or ""))
            if isinstance(parsed_prev, dict):
                previous_same_key_payload = parsed_prev
        except Exception:
            previous_same_key_payload = None
    if bool(fast):
        quick_data: dict[str, Any] | None = None
        quick_meta: dict[str, Any] = {}
        if _accounting_payload_has_data(previous_same_key_payload):
            quick_data = dict(previous_same_key_payload or {})
            quick_meta = {"source": "db-same-key-fastpath", "age_sec": -1, "cache_key": cache_key}
        else:
            probe_data, probe_meta = _market_cache_latest_payload(
                db,
                user_id=int(user.id),
                module_code="accounting",
                marketplace=selected_market,
                max_age_sec=14 * 24 * 60 * 60,
                exclude_cache_keys={cache_key},
                scan_limit=220,
            )
            if _accounting_payload_has_data(probe_data):
                quick_data = dict(probe_data or {})
                quick_meta = probe_meta or {"source": "db-latest-module-fastpath", "age_sec": -1}
        if _accounting_payload_has_data(quick_data):
            quick_warnings = [
                _decode_mojibake_text(str(x or "")).strip()
                for x in list((quick_data or {}).get("warnings") or [])
                if str(x or "").strip()
            ]
            quick_warnings.append("Cached data is shown. Full refresh is running in background.")
            quick_data["warnings"] = list(dict.fromkeys([x for x in quick_warnings if x]))
            rows = list((quick_data or {}).get("analysis_rows") or [])
            if len(rows) > 2000:
                rows = rows[:2000]
                quick_data.setdefault("warnings", []).append("Showing first 2000 analysis rows to keep UI responsive.")

            _audit(
                db,
                user,
                action="accounting_read",
                details=(
                    f"market={selected_market};from={left.isoformat()};to={right.isoformat()};"
                    f"rows={len(rows)};granularity={gran};tz={tz_name};sort={sort_by};q_len={len(str(q or ''))};"
                    f"source={quick_meta.get('source')}"
                ),
                module_code="accounting",
                entity_type="accounting_report",
                status="partial",
            )
            db.commit()
            return AccountingDataOut(
                marketplace=selected_market,
                date_from=left.isoformat(),
                date_to=right.isoformat(),
                overview=(quick_data or {}).get("overview") or {},
                chart=(quick_data or {}).get("chart") or [],
                analysis_rows=rows,
                warnings=[
                    _decode_mojibake_text(str(x or "")).strip()
                    for x in ((quick_data or {}).get("warnings") or [])
                    if _decode_mojibake_text(str(x or "")).strip()
                ],
            )

    data, accounting_cache_meta = get_or_refresh_market_cache(
        db,
        user_id=int(user.id),
        module_code="accounting",
        marketplace=selected_market,
        cache_key=cache_key,
        ttl_sec=max(180, _market_cache_ttl("accounting")),
        fetcher=lambda: build_accounting_payload(
            marketplace=selected_market,
            date_from=left,
            date_to=right,
            wb_api_key=wb_key,
            ozon_api_key=ozon_key,
            products=products_payload,
            expenses=expenses_payload,
            settings=settings_payload,
            granularity=gran,
            tz_name=tz_name,
            search=str(q or ""),
            sort_by=str(sort_by or "net_profit_desc"),
        ),
        stale_if_error_sec=6 * 60 * 60,
        prefer_stale_sec=2 * 60 * 60,
    )
    warnings_now = [str(x or "") for x in (data.get("warnings") if isinstance(data, dict) else [])]
    has_upstream_warning = _warnings_indicate_upstream_failure(warnings_now)
    finance_fallback_needed = _accounting_payload_needs_finance_fallback(data, warnings_now)
    if has_upstream_warning and (not _accounting_payload_has_data(data) or finance_fallback_needed):
        fallback_data: dict[str, Any] | None = None
        fallback_meta: dict[str, Any] = {}
        if _accounting_payload_has_data(previous_same_key_payload):
            fallback_data = previous_same_key_payload
            fallback_meta = {"source": "db-same-key-fallback", "age_sec": -1, "cache_key": cache_key}
        else:
            excluded = {str(cache_key or "").strip()}
            for _ in range(10):
                probe_data, probe_meta = _market_cache_latest_payload(
                    db,
                    user_id=int(user.id),
                    module_code="accounting",
                    marketplace=selected_market,
                    max_age_sec=14 * 24 * 60 * 60,
                    exclude_cache_keys=excluded,
                    scan_limit=220,
                )
                if not isinstance(probe_data, dict):
                    break
                probe_key = str((probe_meta or {}).get("cache_key") or "").strip()
                if probe_key:
                    excluded.add(probe_key)
                if _accounting_payload_has_data(probe_data):
                    fallback_data = probe_data
                    fallback_meta = probe_meta or {"source": "db-latest-module-fallback", "age_sec": -1}
                    break
        if _accounting_payload_has_data(fallback_data):
            data = fallback_data or {}
            accounting_cache_meta = fallback_meta or {"source": "db-latest-module-fallback", "age_sec": -1}
            if finance_fallback_needed:
                data.setdefault("warnings", []).append(
                    "РџРѕРєР°Р·Р°РЅС‹ РїРѕСЃР»РµРґРЅРёРµ СЃС‚Р°Р±РёР»СЊРЅС‹Рµ РґР°РЅРЅС‹Рµ: С‚РµРєСѓС‰РёР№ РѕС‚РІРµС‚ API С‡Р°СЃС‚РёС‡РЅС‹Р№ (С„РёРЅР°РЅСЃРѕРІС‹Рµ РјРµС‚СЂРёРєРё РІСЂРµРјРµРЅРЅРѕ РЅРµРґРѕСЃС‚СѓРїРЅС‹)."
                )
            else:
                data.setdefault("warnings", []).append(
                    "РџРѕРєР°Р·Р°РЅС‹ РїРѕСЃР»РµРґРЅРёРµ СЃС‚Р°Р±РёР»СЊРЅС‹Рµ РґР°РЅРЅС‹Рµ: С‚РµРєСѓС‰РёР№ Р·Р°РїСЂРѕСЃ Рє API РІРµСЂРЅСѓР» РїСѓСЃС‚РѕР№ РѕС‚РІРµС‚."
                )
    if accounting_cache_meta.get("source") == "db-stale-fallback":
        data.setdefault("warnings", []).append(
            "РџРѕРєР°Р·Р°РЅС‹ РєСЌСЂРёСЂРѕРІР°РЅРЅС‹Рµ РґР°РЅРЅС‹Рµ РёР·-Р·Р° РІСЂРµРјРµРЅРЅРѕР№ РЅРµРґРѕСЃС‚СѓРїРЅРѕСЃС‚Рё API РјР°СЂРєРµС‚РїР»РµР№СЃР°."
        )
    rows = list(data.get("analysis_rows") or [])
    if len(rows) > 2000:
        rows = rows[:2000]
        data.setdefault("warnings", []).append("РџРѕРєР°Р·Р°РЅС‹ РїРµСЂРІС‹Рµ 2000 СЃС‚СЂРѕРє Р°РЅР°Р»РёР·Р° РґР»СЏ СѓСЃРєРѕСЂРµРЅРёСЏ РёРЅС‚РµСЂС„РµР№СЃР°.")

    _audit(
        db,
        user,
        action="accounting_read",
        details=(
            f"market={selected_market};from={left.isoformat()};to={right.isoformat()};"
            f"rows={len(rows)};granularity={gran};tz={tz_name};sort={sort_by};q_len={len(str(q or ''))}"
        ),
        module_code="accounting",
        entity_type="accounting_report",
    )
    db.commit()
    return AccountingDataOut(
        marketplace=selected_market,
        date_from=left.isoformat(),
        date_to=right.isoformat(),
        overview=data.get("overview") or {},
        chart=data.get("chart") or [],
        analysis_rows=rows,
        warnings=[_decode_mojibake_text(str(x)).strip() for x in (data.get("warnings") or []) if _decode_mojibake_text(str(x)).strip()],
    )


@router.get("/accounting/monthly-summary", response_model=AccountingMonthlySummaryOut)
def accounting_monthly_summary(
    months: int = 12,
    tz: str = "Europe/Moscow",
    fast: bool = True,
    user: User = Depends(get_current_user),
    db: Session = Depends(get_db),
):
    ensure_module_enabled(db, user, "accounting")
    month_count = max(1, min(12, int(months or 12)))
    tz_name = str(tz or "Europe/Moscow").strip() or "Europe/Moscow"
    try:
        ZoneInfo(tz_name)
    except Exception:
        tz_name = "UTC"

    wb_key = _get_active_marketplace_api_key(db, user.id, "wb")
    ozon_key = _get_active_marketplace_api_key(db, user.id, "ozon")
    settings_row = _get_or_create_accounting_settings(db, user)
    settings_payload = {
        "vat_rate": float(round(settings_row.vat_rate or 0.0, 2)),
        "tax_rate": float(round(settings_row.tax_rate or 0.0, 2)),
        "additional_rate": float(round(settings_row.additional_rate or 0.0, 2)),
        "fixed_cost_per_month": float(round(settings_row.fixed_cost_per_month or 0.0, 2)),
    }

    expense_rows = db.scalars(
        select(AccountingExpense)
        .where(AccountingExpense.user_id == user.id)
        .order_by(AccountingExpense.id.asc())
    ).all()
    products_payload = _collect_product_cost_payload(db, user, "all")
    expenses_payload = _collect_accounting_expense_payload(expense_rows)

    expense_sig_raw = "|".join(
        [
            ";".join(
                [
                    str(int(row.id)),
                    _normalize_accounting_marketplace(row.marketplace),
                    f"{float(round(row.amount or 0.0, 2)):.2f}",
                    row.period_from.isoformat(),
                    row.period_to.isoformat(),
                    "1" if bool(row.is_active) else "0",
                    row.updated_at.isoformat() if row.updated_at else "",
                ]
            )
            for row in expense_rows
        ]
    )
    products_sig_raw = "|".join(
        sorted(
            [
                ";".join(
                    [
                        str(x.get("marketplace") or ""),
                        str(x.get("external_id") or ""),
                        str(x.get("article") or ""),
                        f"{float(round(_to_float_safe(x.get('purchase_price'), 0.0), 2)):.2f}",
                    ]
                )
                for x in products_payload
            ]
        )
    )

    cache_key = build_market_cache_key(
        {
            "kind": "accounting_monthly",
            "months": month_count,
            "tz": tz_name,
            "wb_key_rev": _secret_revision(wb_key),
            "ozon_key_rev": _secret_revision(ozon_key),
            "settings": settings_payload,
            "expense_sig": hashlib.sha1(expense_sig_raw.encode("utf-8")).hexdigest(),
            "products_sig": hashlib.sha1(products_sig_raw.encode("utf-8")).hexdigest(),
        }
    )

    previous_same_key_payload: dict[str, Any] | None = None
    prev_same_key_row = db.scalar(
        select(MarketplaceApiCache).where(
            MarketplaceApiCache.user_id == int(user.id),
            MarketplaceApiCache.module_code == "accounting_monthly",
            MarketplaceApiCache.marketplace == "all",
            MarketplaceApiCache.cache_key == cache_key,
        )
    )
    if prev_same_key_row and str(prev_same_key_row.payload_json or "").strip():
        try:
            parsed_prev = json.loads(str(prev_same_key_row.payload_json or ""))
            if isinstance(parsed_prev, dict):
                previous_same_key_payload = parsed_prev
        except Exception:
            previous_same_key_payload = None

    if bool(fast):
        quick_data: dict[str, Any] | None = None
        quick_meta: dict[str, Any] = {}
        if _accounting_monthly_payload_has_data(previous_same_key_payload):
            quick_data = dict(previous_same_key_payload or {})
            quick_meta = {"source": "db-same-key-fastpath", "age_sec": -1, "cache_key": cache_key}
        else:
            probe_data, probe_meta = _market_cache_latest_payload(
                db,
                user_id=int(user.id),
                module_code="accounting_monthly",
                marketplace="all",
                max_age_sec=14 * 24 * 60 * 60,
                exclude_cache_keys={cache_key},
                scan_limit=180,
            )
            if _accounting_monthly_payload_has_data(probe_data):
                quick_data = dict(probe_data or {})
                quick_meta = probe_meta or {"source": "db-latest-module-fastpath", "age_sec": -1}

        if _accounting_monthly_payload_has_data(quick_data):
            meta = dict((quick_data or {}).get("meta") or {})
            quick_warnings = [
                _decode_mojibake_text(str(x or "")).strip()
                for x in list(meta.get("warnings") or [])
                if _decode_mojibake_text(str(x or "")).strip()
            ]
            quick_warnings.append("Cached monthly summary is shown. Live refresh is running in background.")
            months_rows = list((quick_data or {}).get("months") or [])
            meta_payload = {
                "source": str(quick_meta.get("source") or meta.get("source") or "cache"),
                "partial": True,
                "warnings": list(dict.fromkeys([x for x in quick_warnings if x])),
                "generated_at": str(meta.get("generated_at") or datetime.utcnow().replace(microsecond=0).isoformat() + "Z"),
            }
            _audit(
                db,
                user,
                action="accounting_monthly_read",
                details=(
                    f"months={month_count};tz={tz_name};rows={len(months_rows)};"
                    f"source={meta_payload.get('source')};fast=1"
                ),
                module_code="accounting",
                entity_type="accounting_monthly_report",
                status="partial",
            )
            db.commit()
            return AccountingMonthlySummaryOut(
                months=months_rows,
                meta=AccountingMonthlyMetaOut(**meta_payload),
            )

    data, monthly_cache_meta = get_or_refresh_market_cache(
        db,
        user_id=int(user.id),
        module_code="accounting_monthly",
        marketplace="all",
        cache_key=cache_key,
        ttl_sec=max(180, _market_cache_ttl("accounting_monthly", fast_mode=bool(fast))),
        fetcher=lambda: build_accounting_monthly_summary(
            months=month_count,
            tz_name=tz_name,
            wb_api_key=wb_key,
            ozon_api_key=ozon_key,
            products=products_payload,
            expenses=expenses_payload,
            settings=settings_payload,
        ),
        stale_if_error_sec=6 * 60 * 60,
        prefer_stale_sec=2 * 60 * 60,
    )

    warnings_now = []
    if isinstance(data, dict):
        meta_now = data.get("meta") if isinstance(data.get("meta"), dict) else {}
        warnings_now = [str(x or "") for x in (meta_now.get("warnings") or [])]

    if _warnings_indicate_upstream_failure(warnings_now) and not _accounting_monthly_payload_has_data(data):
        fallback_data: dict[str, Any] | None = None
        fallback_meta: dict[str, Any] = {}
        if _accounting_monthly_payload_has_data(previous_same_key_payload):
            fallback_data = previous_same_key_payload
            fallback_meta = {"source": "db-same-key-fallback", "age_sec": -1, "cache_key": cache_key}
        else:
            probe_data, probe_meta = _market_cache_latest_payload(
                db,
                user_id=int(user.id),
                module_code="accounting_monthly",
                marketplace="all",
                max_age_sec=14 * 24 * 60 * 60,
                exclude_cache_keys={cache_key},
                scan_limit=180,
            )
            if _accounting_monthly_payload_has_data(probe_data):
                fallback_data = probe_data
                fallback_meta = probe_meta or {"source": "db-latest-module-fallback", "age_sec": -1}
        if _accounting_monthly_payload_has_data(fallback_data):
            data = fallback_data or {}
            monthly_cache_meta = fallback_meta or {"source": "db-latest-module-fallback", "age_sec": -1}
            existing_meta = data.get("meta") if isinstance(data.get("meta"), dict) else {}
            merged_warnings = [str(x or "") for x in list(existing_meta.get("warnings") or []) if str(x or "").strip()]
            merged_warnings.append("Showing last stable monthly summary due temporary upstream API limits.")
            data["meta"] = {
                **existing_meta,
                "warnings": list(dict.fromkeys(merged_warnings)),
                "partial": True,
            }

    months_rows = list((data or {}).get("months") or []) if isinstance(data, dict) else []
    meta_payload_raw = (data or {}).get("meta") if isinstance((data or {}).get("meta"), dict) else {}
    meta_payload = {
        "source": str(monthly_cache_meta.get("source") or meta_payload_raw.get("source") or "live"),
        "partial": bool(meta_payload_raw.get("partial") or False),
        "warnings": [
            _decode_mojibake_text(str(x or "")).strip()
            for x in list(meta_payload_raw.get("warnings") or [])
            if _decode_mojibake_text(str(x or "")).strip()
        ],
        "generated_at": str(meta_payload_raw.get("generated_at") or datetime.utcnow().replace(microsecond=0).isoformat() + "Z"),
    }

    _audit(
        db,
        user,
        action="accounting_monthly_read",
        details=(
            f"months={month_count};tz={tz_name};rows={len(months_rows)};"
            f"source={meta_payload.get('source')};fast={1 if bool(fast) else 0}"
        ),
        module_code="accounting",
        entity_type="accounting_monthly_report",
    )
    db.commit()
    return AccountingMonthlySummaryOut(
        months=months_rows,
        meta=AccountingMonthlyMetaOut(**meta_payload),
    )

@router.get("/accounting/purchase-prices/template")
def accounting_purchase_price_template(
    marketplace: str = "all",
    user: User = Depends(get_current_user),
    db: Session = Depends(get_db),
):
    ensure_module_enabled(db, user, "accounting")
    selected_market = _normalize_accounting_marketplace(marketplace)
    rows_src = _collect_product_cost_payload(db, user, selected_market)
    rows: list[list[Any]] = [
        ["marketplace", "article", "external_id", "name", "purchase_price"],
    ]
    for row in rows_src:
        rows.append(
            [
                row.get("marketplace") or "",
                row.get("article") or "",
                row.get("external_id") or "",
                row.get("name") or "",
                row.get("purchase_price") or 0.0,
            ]
        )
    raw, mime = _workbook_bytes_from_rows(rows)
    ext = "xlsx" if "sheet" in mime else "csv"
    filename = f"purchase_price_template_{selected_market}_{datetime.utcnow().strftime('%Y%m%d_%H%M%S')}.{ext}"
    headers = {"Content-Disposition": f'attachment; filename="{filename}"'}
    db.commit()
    return StreamingResponse(io.BytesIO(raw), media_type=mime, headers=headers)


@router.get("/accounting/purchase-prices/export")
def accounting_purchase_price_export(
    marketplace: str = "all",
    user: User = Depends(get_current_user),
    db: Session = Depends(get_db),
):
    ensure_module_enabled(db, user, "accounting")
    selected_market = _normalize_accounting_marketplace(marketplace)
    rows_src = _collect_product_cost_payload(db, user, selected_market)
    rows: list[list[Any]] = [
        ["marketplace", "article", "external_id", "name", "purchase_price"],
    ]
    for row in rows_src:
        rows.append(
            [
                row.get("marketplace") or "",
                row.get("article") or "",
                row.get("external_id") or "",
                row.get("name") or "",
                float(round(_to_money(row.get("purchase_price") or 0.0), 2)),
            ]
        )
    raw, mime = _workbook_bytes_from_rows(rows)
    ext = "xlsx" if "sheet" in mime else "csv"
    filename = f"purchase_prices_export_{selected_market}_{datetime.utcnow().strftime('%Y%m%d_%H%M%S')}.{ext}"
    headers = {"Content-Disposition": f'attachment; filename="{filename}"'}
    db.commit()
    return StreamingResponse(io.BytesIO(raw), media_type=mime, headers=headers)


@router.post("/accounting/purchase-prices/import", response_model=AccountingPurchasePriceImportOut)
def accounting_purchase_price_import(
    file: UploadFile = File(...),
    user: User = Depends(get_current_user),
    db: Session = Depends(get_db),
):
    ensure_module_enabled(db, user, "accounting")
    _require_owner_actor(user)
    rows = _parse_purchase_price_upload(file)
    if not rows:
        raise HTTPException(status_code=400, detail="Р¤Р°Р№Р» РїСѓСЃС‚ РёР»Рё РЅРµ СЂР°СЃРїРѕР·РЅР°РЅ. РСЃРїРѕР»СЊР·СѓР№С‚Рµ СЂР°Р±Р»РѕРЅ РёР· РјРѕРґСѓР»СЏ Р‘СѓС…РіР°Р»С‚РµСЂРёСЏ.")
    updated = 0
    skipped = 0
    unmatched: list[str] = []
    errors: list[str] = []
    for item in rows:
        row_no = int(_to_int(item.get("row_no") or 0))
        market = _normalize_accounting_marketplace(item.get("marketplace") or "all")
        article = str(item.get("article") or "").strip()
        external_id = str(item.get("external_id") or "").strip()
        price_raw = str(item.get("price_raw") or "").strip()
        if not article and not external_id:
            errors.append(f"РЎС‚СЂРѕРєР° {row_no}: РЅРµ СѓРєР°Р·Р°РЅ article/external_id.")
            continue
        if not price_raw:
            errors.append(f"РЎС‚СЂРѕРєР° {row_no}: РЅРµ СѓРєР°Р·Р°РЅР° Р·Р°РєСѓРїРѕС‡РЅР°СЏ С†РµРЅР°.")
            continue
        if not re.fullmatch(r"-?\d+(?:[.,]\d+)?", price_raw):
            errors.append(f"РЎС‚СЂРѕРєР° {row_no}: РЅРµРєРѕСЂСЂРµРєС‚РЅР°СЏ С†РµРЅР° '{price_raw}'.")
            continue
        price = max(0.0, _to_money(price_raw))
        query = select(Product).where(
            Product.user_id == user.id,
            _owned_by_actor_or_owner_filter(Product, user),
        )
        if market in {"wb", "ozon"}:
            query = query.where(Product.marketplace == market)
        match_filters: list[Any] = []
        if article:
            match_filters.append(func.lower(func.coalesce(Product.article, "")) == article.lower())
        if external_id:
            normalized = external_id.lower()
            match_filters.append(func.lower(func.coalesce(Product.external_id, "")) == normalized)
            match_filters.append(func.lower(func.coalesce(Product.barcode, "")) == normalized)
        if not match_filters:
            errors.append(f"РЎС‚СЂРѕРєР° {row_no}: РЅРµ СѓРґР°Р»РѕСЃСЊ СЃРѕР±СЂР°С‚СЊ СѓСЃР»РѕРІРёСЏ СЃРѕРїРѕСЃС‚Р°РІР»РµРЅРёСЏ.")
            continue
        matched = db.scalars(query.where(or_(*match_filters))).all()
        if not matched:
            unmatched.append(f"{market}:{article or external_id}")
            continue
        changed = 0
        for product in matched:
            old_value = float(round(product.purchase_price or 0.0, 2))
            if abs(old_value - price) < 0.0001:
                continue
            product.purchase_price = price
            changed += 1
        if changed > 0:
            updated += changed
        else:
            skipped += len(matched)

    _audit(
        db,
        user,
        action="accounting_purchase_price_imported",
        details=(
            f"updated={updated};skipped={skipped};unmatched={len(unmatched)};errors={len(errors)};"
            f"filename={str(file.filename or '')[:120]}"
        ),
        module_code="accounting",
        entity_type="product_purchase_price",
    )
    db.commit()
    return AccountingPurchasePriceImportOut(
        updated=updated,
        skipped=skipped,
        unmatched=unmatched[:200],
        errors=errors[:200],
    )


@router.get("/profile", response_model=UserProfileOut)
def profile_state(user: User = Depends(get_current_user), db: Session = Depends(get_db)):
    ensure_module_enabled(db, user, "user_profile")
    profile = _get_or_create_user_profile(db, user.id)
    account = _get_or_create_billing_account(db, user.id)
    payload = _build_user_profile_payload(db, user, profile, account)
    if not _actor_is_owner(user):
        actor_email = _actor_email(user)
        payload.team_members = [x for x in payload.team_members if str(x.email or "").strip().lower() == actor_email]
        actor_member = next((x for x in payload.team_members if str(x.email or "").strip().lower() == actor_email), None)
        if actor_member:
            payload.full_name = str(actor_member.full_name or actor_member.nickname or payload.full_name or "")
            payload.phone = str(actor_member.phone or payload.phone or "")
            payload.avatar_url = str(actor_member.avatar_url or payload.avatar_url or "")
    db.commit()
    return payload


@router.put("/profile", response_model=UserProfileOut)
def profile_update(payload: UserProfileUpdateIn, user: User = Depends(get_current_user), db: Session = Depends(get_db)):
    ensure_module_enabled(db, user, "user_profile")
    profile = _get_or_create_user_profile(db, user.id)
    if _actor_is_owner(user):
        profile.full_name = payload.full_name.strip()[:255]
        profile.company_name = payload.company_name.strip()[:255]
        profile.city = payload.city.strip()[:120]
        profile.legal_name = payload.legal_name.strip()[:255]
        profile.legal_address = payload.legal_address.strip()[:255]
        profile.tax_id = payload.tax_id.strip()[:40]
        profile.tax_rate = max(0.0, min(float(payload.tax_rate or 0.0), 100.0))
        profile.phone = payload.phone.strip()[:40]
        profile.position_title = payload.position_title.strip()[:120]
        profile.team_size = max(1, min(int(payload.team_size or 1), 100000))
        profile.company_structure = payload.company_structure.strip()[:12000]
        owner_member = _ensure_owner_team_member(db, user)
        owner_member.avatar_url = payload.avatar_url.strip()[:500]
        _audit(
            db,
            user,
            action="profile_updated",
            details=f"company={profile.company_name};city={profile.city};team={profile.team_size}",
            module_code="user_profile",
            entity_type="profile",
        )
    else:
        member_id = _actor_member_id(user)
        row = db.get(TeamMember, member_id)
        if not row or row.user_id != user.id or bool(row.is_owner):
            raise HTTPException(status_code=403, detail="РќРµРґРѕСЃС‚Р°С‚РѕС‡РЅРѕ РїСЂР°РІ")
        row.full_name = payload.full_name.strip()[:255]
        row.phone = payload.phone.strip()[:40]
        row.avatar_url = payload.avatar_url.strip()[:500]
        _audit(
            db,
            user,
            action="profile_member_updated",
            details=f"member_id={row.id}",
            module_code="user_profile",
            entity_type="team_member",
            entity_id=str(row.id),
        )
    account = _get_or_create_billing_account(db, user.id)
    db.commit()
    return _build_user_profile_payload(db, user, profile, account)


def _save_avatar_upload(file: UploadFile, *, user_id: int, prefix: str) -> str:
    if not file or not file.filename:
        raise HTTPException(status_code=400, detail="Р¤Р°Р№Р» РЅРµ РІС‹Р±СЂР°РЅ")
    content_type = str(file.content_type or "").lower()
    if not content_type.startswith("image/"):
        raise HTTPException(status_code=400, detail="Р¤Р°Р№Р» РґРѕР»Р¶РµРЅ Р±С‹С‚СЊ РёР·РѕР±СЂР°Р¶РµРЅРёРµРј")
    ext = os.path.splitext(file.filename)[1].lower()
    if ext not in {".png", ".jpg", ".jpeg", ".webp", ".gif", ".svg"}:
        ext = ".png"
    raw = file.file.read()
    if not raw:
        raise HTTPException(status_code=400, detail="Р¤Р°Р№Р» РїСѓСЃС‚РѕР№")
    if len(raw) > 4 * 1024 * 1024:
        raise HTTPException(status_code=400, detail="Р¤Р°Р№Р» СЃР»РёСЂРєРѕРј Р±РѕР»СЊСЂРѕР№ (РґРѕ 4 РњР‘)")
    static_root = Path(__file__).resolve().parent.parent / "static"
    target_dir = static_root / "uploads" / "avatars"
    target_dir.mkdir(parents=True, exist_ok=True)
    filename = f"{prefix}-{user_id}-{secrets.token_hex(6)}{ext}"
    path = target_dir / filename
    path.write_bytes(raw)
    return f"/static/uploads/avatars/{filename}"


@router.post("/profile/avatar/upload", response_model=AvatarUploadOut)
def profile_avatar_upload(
    file: UploadFile = File(...),
    user: User = Depends(get_current_user),
    db: Session = Depends(get_db),
):
    ensure_module_enabled(db, user, "user_profile")
    url = _save_avatar_upload(file, user_id=user.id, prefix="avatar")
    if _actor_is_owner(user):
        owner_member = _ensure_owner_team_member(db, user)
        owner_member.avatar_url = url
    else:
        member_id = _actor_member_id(user)
        row = db.get(TeamMember, member_id)
        if not row or row.user_id != user.id:
            raise HTTPException(status_code=404, detail="РЎРѕС‚СЂСѓРґРЅРёРє РЅРµ РЅР°Р№РґРµРЅ")
        row.avatar_url = url
    _audit(
        db,
        user,
        action="profile_avatar_uploaded",
        details=f"url={url}",
        module_code="user_profile",
        entity_type="profile",
    )
    db.commit()
    return AvatarUploadOut(url=url)


@router.post("/profile/team/{member_id}/avatar/upload", response_model=AvatarUploadOut)
def profile_team_avatar_upload(
    member_id: int,
    file: UploadFile = File(...),
    user: User = Depends(get_current_user),
    db: Session = Depends(get_db),
):
    ensure_module_enabled(db, user, "user_profile")
    row = db.get(TeamMember, member_id)
    if not row or row.user_id != user.id:
        raise HTTPException(status_code=404, detail="РЎРѕС‚СЂСѓРґРЅРёРє РЅРµ РЅР°Р№РґРµРЅ")
    if not _actor_is_owner(user):
        actor_member_id = _actor_member_id(user)
        if int(actor_member_id or 0) != int(member_id or 0):
            raise HTTPException(status_code=403, detail="РќРµРґРѕСЃС‚Р°С‚РѕС‡РЅРѕ РїСЂР°РІ")
    url = _save_avatar_upload(file, user_id=user.id, prefix=f"member{member_id}")
    row.avatar_url = url
    _audit(
        db,
        user,
        action="profile_team_avatar_uploaded",
        details=f"member_id={row.id};url={url}",
        module_code="user_profile",
        entity_type="team_member",
        entity_id=str(row.id),
    )
    db.commit()
    return AvatarUploadOut(url=url)


@router.get("/profile/ai", response_model=AiProfileOut)
def profile_ai_state(user: User = Depends(get_current_user), db: Session = Depends(get_db)):
    ensure_module_enabled(db, user, "user_profile")
    payload = _build_ai_profile_payload(db, user.id)
    db.commit()
    return payload


@router.post("/profile/ai/select", response_model=AiProfileOut)
def profile_ai_select(payload: AiSelectionIn, user: User = Depends(get_current_user), db: Session = Depends(get_db)):
    ensure_module_enabled(db, user, "user_profile")
    _save_user_ai_selection(
        db,
        user.id,
        use_global_default=bool(payload.use_global_default),
        mode=payload.mode,
        service_id=payload.service_id,
    )
    _audit(
        db,
        user,
        action="profile_ai_selected",
        details=f"use_global_default={bool(payload.use_global_default)};mode={payload.mode};service_id={payload.service_id}",
        module_code="user_profile",
        entity_type="ai_selection",
    )
    db.commit()
    return _build_ai_profile_payload(db, user.id)


@router.post("/profile/ai/services", response_model=AiServiceOut)
def profile_ai_service_add(payload: AiServiceIn, user: User = Depends(get_current_user), db: Session = Depends(get_db)):
    ensure_module_enabled(db, user, "user_profile")
    row = _upsert_ai_service(
        db,
        user_id=user.id,
        payload=payload,
    )
    _audit(
        db,
        user,
        action="profile_ai_service_added",
        details=f"service_id={row.id};provider={row.provider}",
        module_code="user_profile",
        entity_type="ai_service",
        entity_id=str(row.id),
    )
    db.commit()
    return _ai_service_to_out(row, scope="user")


@router.put("/profile/ai/services/{service_id}", response_model=AiServiceOut)
def profile_ai_service_update(service_id: int, payload: AiServiceIn, user: User = Depends(get_current_user), db: Session = Depends(get_db)):
    ensure_module_enabled(db, user, "user_profile")
    row = db.get(AiServiceAccount, service_id)
    if not row or row.user_id != user.id:
        raise HTTPException(status_code=404, detail="AI СЃРµСЂРІРёСЃ РЅРµ РЅР°Р№РґРµРЅ")
    _update_ai_service_row(row, payload)
    _audit(
        db,
        user,
        action="profile_ai_service_updated",
        details=f"service_id={row.id};provider={row.provider}",
        module_code="user_profile",
        entity_type="ai_service",
        entity_id=str(row.id),
    )
    db.commit()
    return _ai_service_to_out(row, scope="user")


@router.delete("/profile/ai/services/{service_id}", response_model=MessageOut)
def profile_ai_service_delete(service_id: int, user: User = Depends(get_current_user), db: Session = Depends(get_db)):
    ensure_module_enabled(db, user, "user_profile")
    row = db.get(AiServiceAccount, service_id)
    if not row or row.user_id != user.id:
        raise HTTPException(status_code=404, detail="AI СЃРµСЂРІРёСЃ РЅРµ РЅР°Р№РґРµРЅ")
    _reset_ai_selection_if_deleted_service(db, user.id, service_id)
    db.delete(row)
    _audit(
        db,
        user,
        action="profile_ai_service_deleted",
        details=f"service_id={service_id}",
        module_code="user_profile",
        entity_type="ai_service",
        entity_id=str(service_id),
    )
    db.commit()
    return MessageOut(message="AI СЃРµСЂРІРёСЃ СѓРґР°Р»РµРЅ")


@router.post("/profile/ai/services/reorder", response_model=list[AiServiceOut])
def profile_ai_service_reorder(payload: AiServiceReorderIn, user: User = Depends(get_current_user), db: Session = Depends(get_db)):
    ensure_module_enabled(db, user, "user_profile")
    ordered = _reorder_ai_services(db, user_id=user.id, service_ids=list(payload.service_ids or []))
    _audit(
        db,
        user,
        action="profile_ai_service_reordered",
        details=json.dumps({"service_ids": [int(x) for x in payload.service_ids or []]}, ensure_ascii=False),
        module_code="user_profile",
        entity_type="ai_service",
    )
    db.commit()
    return [_ai_service_to_out(row, scope="user") for row in ordered]


@router.get("/profile/team", response_model=list[TeamMemberOut])
def profile_team_list(user: User = Depends(get_current_user), db: Session = Depends(get_db)):
    ensure_module_enabled(db, user, "user_profile")
    _require_owner_actor(user)
    rows = _list_team_members(db, user.id)
    db.commit()
    return rows


@router.post("/profile/team", response_model=TeamMemberOut)
def profile_team_add(payload: TeamMemberIn, user: User = Depends(get_current_user), db: Session = Depends(get_db)):
    ensure_module_enabled(db, user, "user_profile")
    _require_owner_actor(user)
    email = payload.email.strip().lower()
    _ensure_team_email_is_available(db, email, user_id=user.id)
    exists = db.scalar(select(TeamMember).where(TeamMember.user_id == user.id, TeamMember.email == email))
    if exists:
        raise HTTPException(status_code=400, detail="РЎРѕС‚СЂСѓРґРЅРёРє СЃ С‚Р°РєРёРј email СѓР¶Рµ РґРѕР±Р°РІР»РµРЅ")
    password = _validate_team_member_password(payload.password, required=True)
    row = TeamMember(
        user_id=user.id,
        email=email[:255],
        phone=payload.phone.strip()[:40],
        full_name=payload.full_name.strip()[:255],
        city=payload.city.strip()[:120],
        position_title=payload.position_title.strip()[:120],
        nickname=payload.nickname.strip()[:120],
        avatar_url=payload.avatar_url.strip()[:500],
        hashed_password=get_password_hash(password),
        access_scope=json.dumps(_safe_team_scope(payload.access_scope), ensure_ascii=False),
        is_owner=False,
        is_active=True,
    )
    db.add(row)
    _audit(
        db,
        user,
        action="profile_team_member_added",
        details=f"email={row.email}",
        module_code="user_profile",
        entity_type="team_member",
        entity_id=str(row.id),
    )
    db.commit()
    return _team_member_to_out(row)


@router.put("/profile/team/{member_id}", response_model=TeamMemberOut)
def profile_team_update(member_id: int, payload: TeamMemberIn, user: User = Depends(get_current_user), db: Session = Depends(get_db)):
    ensure_module_enabled(db, user, "user_profile")
    row = db.get(TeamMember, member_id)
    if not row or row.user_id != user.id:
        raise HTTPException(status_code=404, detail="РЎРѕС‚СЂСѓРґРЅРёРє РЅРµ РЅР°Р№РґРµРЅ")
    if not _actor_is_owner(user):
        actor_member_id = _actor_member_id(user)
        if int(actor_member_id or 0) != int(member_id or 0):
            raise HTTPException(status_code=403, detail="РќРµРґРѕСЃС‚Р°С‚РѕС‡РЅРѕ РїСЂР°РІ")
        row.phone = payload.phone.strip()[:40]
        row.full_name = payload.full_name.strip()[:255]
        row.city = payload.city.strip()[:120]
        row.position_title = payload.position_title.strip()[:120]
        row.nickname = payload.nickname.strip()[:120]
        row.avatar_url = payload.avatar_url.strip()[:500]
        _audit(
            db,
            user,
            action="profile_team_member_self_updated",
            details=f"member_id={row.id}",
            module_code="user_profile",
            entity_type="team_member",
            entity_id=str(row.id),
        )
        db.commit()
        return _team_member_to_out(row)
    if row.is_owner:
        row.phone = payload.phone.strip()[:40]
        row.full_name = payload.full_name.strip()[:255]
        row.city = payload.city.strip()[:120]
        row.position_title = payload.position_title.strip()[:120]
        row.nickname = (payload.nickname.strip() or "owner")[:120]
        row.avatar_url = payload.avatar_url.strip()[:500]
        row.access_scope = json.dumps(["*"], ensure_ascii=False)
    else:
        email = payload.email.strip().lower()
        _ensure_team_email_is_available(db, email, user_id=user.id, exclude_member_id=row.id)
        duplicate = db.scalar(select(TeamMember).where(TeamMember.user_id == user.id, TeamMember.email == email, TeamMember.id != row.id))
        if duplicate:
            raise HTTPException(status_code=400, detail="РЎРѕС‚СЂСѓРґРЅРёРє СЃ С‚Р°РєРёРј email СѓР¶Рµ РґРѕР±Р°РІР»РµРЅ")
        row.email = email[:255]
        row.phone = payload.phone.strip()[:40]
        row.full_name = payload.full_name.strip()[:255]
        row.city = payload.city.strip()[:120]
        row.position_title = payload.position_title.strip()[:120]
        row.nickname = payload.nickname.strip()[:120]
        row.avatar_url = payload.avatar_url.strip()[:500]
        row.access_scope = json.dumps(_safe_team_scope(payload.access_scope), ensure_ascii=False)
        password = _validate_team_member_password(payload.password, required=False)
        if password:
            row.hashed_password = get_password_hash(password)
    _audit(
        db,
        user,
        action="profile_team_member_updated",
        details=f"member_id={row.id}",
        module_code="user_profile",
        entity_type="team_member",
        entity_id=str(row.id),
    )
    db.commit()
    return _team_member_to_out(row)


@router.delete("/profile/team/{member_id}", response_model=MessageOut)
def profile_team_delete(member_id: int, user: User = Depends(get_current_user), db: Session = Depends(get_db)):
    ensure_module_enabled(db, user, "user_profile")
    _require_owner_actor(user)
    row = db.get(TeamMember, member_id)
    if not row or row.user_id != user.id:
        raise HTTPException(status_code=404, detail="РЎРѕС‚СЂСѓРґРЅРёРє РЅРµ РЅР°Р№РґРµРЅ")
    if row.is_owner:
        raise HTTPException(status_code=400, detail="РќРµР»СЊР·СЏ СѓРґР°Р»РёС‚СЊ РІР»Р°РґРµР»СЊС†Р° РєР°Р±РёРЅРµС‚Р°")
    db.delete(row)
    _audit(
        db,
        user,
        action="profile_team_member_deleted",
        details=f"member_id={member_id}",
        module_code="user_profile",
        entity_type="team_member",
        entity_id=str(member_id),
    )
    db.commit()
    return MessageOut(message="РЎРѕС‚СЂСѓРґРЅРёРє СѓРґР°Р»РµРЅ")


@router.post("/profile/password", response_model=MessageOut)
def profile_change_password(
    payload: UserProfilePasswordIn,
    token: str = Depends(oauth2_scheme),
    user: User = Depends(get_current_user),
    db: Session = Depends(get_db),
):
    ensure_module_enabled(db, user, "user_profile")
    if len(payload.new_password or "") < 8:
        raise HTTPException(status_code=400, detail="РќРѕРІС‹Р№ РїР°СЂРѕР»СЊ РґРѕР»Р¶РµРЅ Р±С‹С‚СЊ РјРёРЅРёРјСѓРј 8 СЃРёРјРІРѕР»РѕРІ")

    subject_raw = str(decode_access_token(token) or "").strip()
    subject = subject_raw.lower()
    if subject.startswith("u:"):
        user_id = _to_int_safe(subject.split(":", 1)[1])
        if user_id and int(user_id) == int(user.id):
            if not verify_password(payload.current_password, user.hashed_password):
                raise HTTPException(status_code=400, detail="РўРµРєСѓС‰РёР№ РїР°СЂРѕР»СЊ СѓРєР°Р·Р°РЅ РЅРµРІРµСЂРЅРѕ")
            user.hashed_password = get_password_hash(payload.new_password)
            owner_member = db.scalar(select(TeamMember).where(TeamMember.user_id == user.id, TeamMember.is_owner.is_(True)))
            if owner_member:
                owner_member.hashed_password = user.hashed_password
            _audit(
                db,
                user,
                action="profile_password_changed",
                details="kind=owner",
                module_code="user_profile",
                entity_type="password",
            )
            db.commit()
            return MessageOut(message="РџР°СЂРѕР»СЊ РѕР±РЅРѕРІР»РµРЅ")
    if subject.startswith("m:"):
        member_id = _to_int_safe(subject.split(":", 1)[1])
        if member_id:
            member = db.get(TeamMember, member_id)
            if member and int(member.user_id) == int(user.id) and member.is_active and not member.is_owner:
                if not member.hashed_password or not verify_password(payload.current_password, member.hashed_password):
                    raise HTTPException(status_code=400, detail="РўРµРєСѓС‰РёР№ РїР°СЂРѕР»СЊ СѓРєР°Р·Р°РЅ РЅРµРІРµСЂРЅРѕ")
                member.hashed_password = get_password_hash(payload.new_password)
                _audit(
                    db,
                    user,
                    action="profile_password_changed",
                    details=f"kind=team_member;member_id={member.id}",
                    module_code="user_profile",
                    entity_type="password",
                    entity_id=str(member.id),
                )
                db.commit()
                return MessageOut(message="РџР°СЂРѕР»СЊ СЃРѕС‚СЂСѓРґРЅРёРєР° РѕР±РЅРѕРІР»РµРЅ")

    if subject and subject == user.email:
        if not verify_password(payload.current_password, user.hashed_password):
            raise HTTPException(status_code=400, detail="РўРµРєСѓС‰РёР№ РїР°СЂРѕР»СЊ СѓРєР°Р·Р°РЅ РЅРµРІРµСЂРЅРѕ")
        user.hashed_password = get_password_hash(payload.new_password)
        owner_member = db.scalar(select(TeamMember).where(TeamMember.user_id == user.id, TeamMember.is_owner.is_(True)))
        if owner_member:
            owner_member.hashed_password = user.hashed_password
        _audit(
            db,
            user,
            action="profile_password_changed",
            details="kind=owner",
            module_code="user_profile",
            entity_type="password",
        )
        db.commit()
        return MessageOut(message="РџР°СЂРѕР»СЊ РѕР±РЅРѕРІР»РµРЅ")

    member = db.scalar(
        select(TeamMember).where(
            TeamMember.user_id == user.id,
            TeamMember.email == subject,
            TeamMember.is_active.is_(True),
            TeamMember.is_owner.is_(False),
        )
    )
    if not member or not member.hashed_password or not verify_password(payload.current_password, member.hashed_password):
        raise HTTPException(status_code=400, detail="РўРµРєСѓС‰РёР№ РїР°СЂРѕР»СЊ СѓРєР°Р·Р°РЅ РЅРµРІРµСЂРЅРѕ")
    member.hashed_password = get_password_hash(payload.new_password)
    _audit(
        db,
        user,
        action="profile_password_changed",
        details=f"kind=team_member;member_id={member.id}",
        module_code="user_profile",
        entity_type="password",
        entity_id=str(member.id),
    )
    db.commit()
    return MessageOut(message="РџР°СЂРѕР»СЊ СЃРѕС‚СЂСѓРґРЅРёРєР° РѕР±РЅРѕРІР»РµРЅ")


@router.post("/profile/plan", response_model=UserProfileOut)
def profile_change_plan(payload: BillingPlanChangeIn, user: User = Depends(get_current_user), db: Session = Depends(get_db)):
    ensure_module_enabled(db, user, "user_profile")
    _require_owner_actor(user)
    plan_code = (payload.plan_code or "").strip().lower()
    if plan_code not in BILLING_PLANS:
        raise HTTPException(status_code=400, detail=f"РќРµРёР·РІРµСЃС‚РЅС‹Р№ РїР»Р°РЅ: {plan_code}")

    account = _get_or_create_billing_account(db, user.id)
    account.plan_code = plan_code
    account.status = "active"
    account.monthly_price = int(BILLING_PLANS[plan_code]["price"])
    if not account.renew_at or account.renew_at < datetime.utcnow():
        account.renew_at = datetime.utcnow() + timedelta(days=30)
    db.add(account)
    db.add(
        BillingEvent(
            user_id=user.id,
            event_type="plan_changed",
            amount=account.monthly_price,
            note=f"РџР»Р°РЅ РёР·РјРµРЅРµРЅ РЅР° {plan_code}",
        )
    )
    _audit(
        db,
        user,
        action="profile_plan_changed",
        details=f"plan={plan_code}",
        module_code="billing",
        entity_type="plan",
    )
    profile = _get_or_create_user_profile(db, user.id)
    db.commit()
    return _build_user_profile_payload(db, user, profile, account)


@router.post("/profile/renew", response_model=UserProfileOut)
def profile_renew_plan(user: User = Depends(get_current_user), db: Session = Depends(get_db)):
    ensure_module_enabled(db, user, "user_profile")
    _require_owner_actor(user)
    account = _get_or_create_billing_account(db, user.id)
    base = account.renew_at if account.renew_at and account.renew_at > datetime.utcnow() else datetime.utcnow()
    account.renew_at = base + timedelta(days=30)
    account.status = "active"
    db.add(account)
    db.add(
        BillingEvent(
            user_id=user.id,
            event_type="renew",
            amount=account.monthly_price,
            note=f"РџСЂРѕРґР»РµРЅРёРµ РґРѕ {account.renew_at.isoformat()}",
        )
    )
    _audit(
        db,
        user,
        action="profile_plan_renewed",
        details=f"renew_at={account.renew_at.isoformat()}",
        module_code="billing",
        entity_type="plan",
    )
    profile = _get_or_create_user_profile(db, user.id)
    db.commit()
    return _build_user_profile_payload(db, user, profile, account)


@router.get("/billing", response_model=BillingOut)
def billing_state(user: User = Depends(get_current_user), db: Session = Depends(get_db)):
    ensure_module_enabled(db, user, "billing")
    account = _get_or_create_billing_account(db, user.id)
    payload = _build_billing_payload(db, user.id, account)
    db.commit()
    return payload


@router.post("/billing/plan", response_model=BillingOut)
def billing_change_plan(payload: BillingPlanChangeIn, user: User = Depends(get_current_user), db: Session = Depends(get_db)):
    ensure_module_enabled(db, user, "billing")
    plan_code = (payload.plan_code or "").strip().lower()
    if plan_code not in BILLING_PLANS:
        raise HTTPException(status_code=400, detail=f"РќРµРёР·РІРµСЃС‚РЅС‹Р№ РїР»Р°РЅ: {plan_code}")

    account = _get_or_create_billing_account(db, user.id)
    account.plan_code = plan_code
    account.status = "active"
    account.monthly_price = int(BILLING_PLANS[plan_code]["price"])
    if not account.renew_at or account.renew_at < datetime.utcnow():
        account.renew_at = datetime.utcnow() + timedelta(days=30)
    db.add(account)
    db.add(
        BillingEvent(
            user_id=user.id,
            event_type="plan_changed",
            amount=account.monthly_price,
            note=f"РџР»Р°РЅ РёР·РјРµРЅРµРЅ РЅР° {plan_code}",
        )
    )
    _audit(
        db,
        user,
        action="billing_plan_changed",
        details=f"plan={plan_code}",
        module_code="billing",
        entity_type="plan",
    )
    db.commit()
    return _build_billing_payload(db, user.id, account)


@router.post("/billing/renew", response_model=BillingOut)
def billing_renew(user: User = Depends(get_current_user), db: Session = Depends(get_db)):
    ensure_module_enabled(db, user, "billing")
    account = _get_or_create_billing_account(db, user.id)
    base = account.renew_at if account.renew_at and account.renew_at > datetime.utcnow() else datetime.utcnow()
    account.renew_at = base + timedelta(days=30)
    account.status = "active"
    db.add(account)
    db.add(
        BillingEvent(
            user_id=user.id,
            event_type="renew",
            amount=account.monthly_price,
            note=f"РџСЂРѕРґР»РµРЅРёРµ РґРѕ {account.renew_at.isoformat()}",
        )
    )
    _audit(
        db,
        user,
        action="billing_renewed",
        details=f"renew_at={account.renew_at.isoformat()}",
        module_code="billing",
        entity_type="plan",
    )
    db.commit()
    return _build_billing_payload(db, user.id, account)


@router.get("/admin/users", response_model=list[UserOut])
def admin_users(_: User = Depends(get_admin_user), db: Session = Depends(get_db)):
    return db.scalars(select(User).order_by(User.id.desc())).all()


@router.get("/admin/users/{user_id}/profile", response_model=AdminUserProfileOut)
def admin_user_profile(user_id: int, _: User = Depends(get_admin_user), db: Session = Depends(get_db)):
    target = db.get(User, user_id)
    if not target:
        raise HTTPException(status_code=404, detail="РџРѕР»СЊР·РѕРІР°С‚РµР»СЊ РЅРµ РЅР°Р№РґРµРЅ")
    db.commit()
    return _build_admin_user_profile_payload(db, target)


@router.put("/admin/users/{user_id}/profile", response_model=AdminUserProfileOut)
def admin_user_profile_update(
    user_id: int,
    payload: UserProfileUpdateIn,
    me: User = Depends(get_admin_user),
    db: Session = Depends(get_db),
):
    target = db.get(User, user_id)
    if not target:
        raise HTTPException(status_code=404, detail="РџРѕР»СЊР·РѕРІР°С‚РµР»СЊ РЅРµ РЅР°Р№РґРµРЅ")
    profile = _get_or_create_user_profile(db, target.id)

    profile.full_name = payload.full_name.strip()[:255]
    profile.company_name = payload.company_name.strip()[:255]
    profile.city = payload.city.strip()[:120]
    profile.legal_name = payload.legal_name.strip()[:255]
    profile.legal_address = payload.legal_address.strip()[:255]
    profile.tax_id = payload.tax_id.strip()[:40]
    profile.tax_rate = max(0.0, min(float(payload.tax_rate or 0.0), 100.0))
    profile.phone = payload.phone.strip()[:40]
    profile.position_title = payload.position_title.strip()[:120]
    profile.team_size = max(1, min(int(payload.team_size or 1), 100000))
    profile.company_structure = payload.company_structure.strip()[:12000]
    profile.avatar_url = payload.avatar_url.strip()[:500]

    _audit(
        db,
        me,
        action="admin_user_profile_updated",
        details=f"user_id={target.id};company={profile.company_name};city={profile.city};team={profile.team_size}",
        module_code="admin",
        entity_type="user",
        entity_id=str(target.id),
    )
    db.commit()
    return _build_admin_user_profile_payload(db, target)


@router.post("/admin/users/{user_id}/plan", response_model=AdminUserProfileOut)
def admin_user_change_plan(
    user_id: int,
    payload: BillingPlanChangeIn,
    me: User = Depends(get_admin_user),
    db: Session = Depends(get_db),
):
    target = db.get(User, user_id)
    if not target:
        raise HTTPException(status_code=404, detail="РџРѕР»СЊР·РѕРІР°С‚РµР»СЊ РЅРµ РЅР°Р№РґРµРЅ")
    plan_code = (payload.plan_code or "").strip().lower()
    if plan_code not in BILLING_PLANS:
        raise HTTPException(status_code=400, detail=f"РќРµРёР·РІРµСЃС‚РЅС‹Р№ РїР»Р°РЅ: {plan_code}")

    account = _get_or_create_billing_account(db, target.id)
    account.plan_code = plan_code
    account.status = "active"
    account.monthly_price = int(BILLING_PLANS[plan_code]["price"])
    if not account.renew_at or account.renew_at < datetime.utcnow():
        account.renew_at = datetime.utcnow() + timedelta(days=30)
    db.add(account)
    db.add(
        BillingEvent(
            user_id=target.id,
            event_type="admin_plan_changed",
            amount=account.monthly_price,
            note=f"РђРґРјРёРЅРёСЃС‚СЂР°С‚РѕСЂ РёР·РјРµРЅРёР» РїР»Р°РЅ РЅР° {plan_code}",
        )
    )
    _audit(
        db,
        me,
        action="admin_user_plan_changed",
        details=f"user_id={target.id};plan={plan_code}",
        module_code="admin",
        entity_type="user",
        entity_id=str(target.id),
    )
    db.commit()
    return _build_admin_user_profile_payload(db, target)


@router.post("/admin/users/{user_id}/team", response_model=TeamMemberOut)
def admin_team_add(
    user_id: int,
    payload: TeamMemberIn,
    me: User = Depends(get_admin_user),
    db: Session = Depends(get_db),
):
    target = db.get(User, user_id)
    if not target:
        raise HTTPException(status_code=404, detail="РџРѕР»СЊР·РѕРІР°С‚РµР»СЊ РЅРµ РЅР°Р№РґРµРЅ")
    email = payload.email.strip().lower()
    _ensure_team_email_is_available(db, email, user_id=user_id)
    exists = db.scalar(select(TeamMember).where(TeamMember.user_id == user_id, TeamMember.email == email))
    if exists:
        raise HTTPException(status_code=400, detail="РЎРѕС‚СЂСѓРґРЅРёРє СЃ С‚Р°РєРёРј email СѓР¶Рµ РґРѕР±Р°РІР»РµРЅ")
    password = _validate_team_member_password(payload.password, required=True)
    row = TeamMember(
        user_id=user_id,
        email=email[:255],
        phone=payload.phone.strip()[:40],
        full_name=payload.full_name.strip()[:255],
        city=payload.city.strip()[:120],
        position_title=payload.position_title.strip()[:120],
        nickname=payload.nickname.strip()[:120],
        avatar_url=payload.avatar_url.strip()[:500],
        hashed_password=get_password_hash(password),
        access_scope=json.dumps(_safe_team_scope(payload.access_scope), ensure_ascii=False),
        is_owner=False,
        is_active=True,
    )
    db.add(row)
    _audit(
        db,
        me,
        action="admin_team_member_added",
        details=f"user_id={user_id};email={row.email}",
        module_code="admin",
        entity_type="team_member",
        entity_id=str(row.id),
    )
    db.commit()
    return _team_member_to_out(row)


@router.put("/admin/users/{user_id}/team/{member_id}", response_model=TeamMemberOut)
def admin_team_update(
    user_id: int,
    member_id: int,
    payload: TeamMemberIn,
    me: User = Depends(get_admin_user),
    db: Session = Depends(get_db),
):
    target = db.get(User, user_id)
    if not target:
        raise HTTPException(status_code=404, detail="РџРѕР»СЊР·РѕРІР°С‚РµР»СЊ РЅРµ РЅР°Р№РґРµРЅ")
    row = db.get(TeamMember, member_id)
    if not row or row.user_id != user_id:
        raise HTTPException(status_code=404, detail="РЎРѕС‚СЂСѓРґРЅРёРє РЅРµ РЅР°Р№РґРµРЅ")
    if row.is_owner:
        row.phone = payload.phone.strip()[:40]
        row.full_name = payload.full_name.strip()[:255]
        row.city = payload.city.strip()[:120]
        row.position_title = payload.position_title.strip()[:120]
        row.nickname = (payload.nickname.strip() or "owner")[:120]
        row.avatar_url = payload.avatar_url.strip()[:500]
        row.access_scope = json.dumps(["*"], ensure_ascii=False)
    else:
        email = payload.email.strip().lower()
        _ensure_team_email_is_available(db, email, user_id=user_id, exclude_member_id=row.id)
        duplicate = db.scalar(select(TeamMember).where(TeamMember.user_id == user_id, TeamMember.email == email, TeamMember.id != row.id))
        if duplicate:
            raise HTTPException(status_code=400, detail="РЎРѕС‚СЂСѓРґРЅРёРє СЃ С‚Р°РєРёРј email СѓР¶Рµ РґРѕР±Р°РІР»РµРЅ")
        row.email = email[:255]
        row.phone = payload.phone.strip()[:40]
        row.full_name = payload.full_name.strip()[:255]
        row.city = payload.city.strip()[:120]
        row.position_title = payload.position_title.strip()[:120]
        row.nickname = payload.nickname.strip()[:120]
        row.avatar_url = payload.avatar_url.strip()[:500]
        row.access_scope = json.dumps(_safe_team_scope(payload.access_scope), ensure_ascii=False)
        password = _validate_team_member_password(payload.password, required=False)
        if password:
            row.hashed_password = get_password_hash(password)
    _audit(
        db,
        me,
        action="admin_team_member_updated",
        details=f"user_id={user_id};member_id={member_id}",
        module_code="admin",
        entity_type="team_member",
        entity_id=str(member_id),
    )
    db.commit()
    return _team_member_to_out(row)


@router.delete("/admin/users/{user_id}/team/{member_id}", response_model=MessageOut)
def admin_team_delete(user_id: int, member_id: int, me: User = Depends(get_admin_user), db: Session = Depends(get_db)):
    row = db.get(TeamMember, member_id)
    if not row or row.user_id != user_id:
        raise HTTPException(status_code=404, detail="РЎРѕС‚СЂСѓРґРЅРёРє РЅРµ РЅР°Р№РґРµРЅ")
    if row.is_owner:
        raise HTTPException(status_code=400, detail="РќРµР»СЊР·СЏ СѓРґР°Р»РёС‚СЊ РІР»Р°РґРµР»СЊС†Р° РєР°Р±РёРЅРµС‚Р°")
    db.delete(row)
    _audit(
        db,
        me,
        action="admin_team_member_deleted",
        details=f"user_id={user_id};member_id={member_id}",
        module_code="admin",
        entity_type="team_member",
        entity_id=str(member_id),
    )
    db.commit()
    return MessageOut(message="РЎРѕС‚СЂСѓРґРЅРёРє СѓРґР°Р»РµРЅ")


@router.get("/admin/stats", response_model=AdminStatsOut)
def admin_stats(_: User = Depends(get_admin_user), db: Session = Depends(get_db)):
    week_ago = datetime.utcnow() - timedelta(days=7)
    day_ago = datetime.utcnow() - timedelta(hours=24)
    total_users = db.scalar(select(func.count()).select_from(User)) or 0
    new_users_7d = db.scalar(select(func.count()).select_from(User).where(User.created_at >= week_ago)) or 0
    total_products = db.scalar(select(func.count()).select_from(Product)) or 0
    total_jobs = db.scalar(select(func.count()).select_from(SeoJob)) or 0
    active_jobs = db.scalar(select(func.count()).select_from(SeoJob).where(SeoJob.status.in_(["generated", "in_progress"]))) or 0
    total_team_members = db.scalar(select(func.count()).select_from(TeamMember)) or 0
    employees_total = db.scalar(
        select(func.count()).select_from(TeamMember).where(TeamMember.is_owner.is_(False))
    ) or 0
    active_users_24h = db.scalar(
        select(func.count(func.distinct(AuditLog.user_id))).where(
            AuditLog.created_at >= day_ago,
            AuditLog.user_id.is_not(None),
        )
    ) or 0
    audit_events_24h = db.scalar(
        select(func.count()).select_from(AuditLog).where(AuditLog.created_at >= day_ago)
    ) or 0
    return AdminStatsOut(
        total_users=total_users,
        new_users_7d=new_users_7d,
        total_products=total_products,
        total_jobs=total_jobs,
        active_jobs=active_jobs,
        total_team_members=total_team_members,
        employees_total=employees_total,
        active_users_24h=active_users_24h,
        audit_events_24h=audit_events_24h,
    )


def _proc_meminfo_kb() -> dict[str, int]:
    out: dict[str, int] = {}
    try:
        with open("/proc/meminfo", "r", encoding="utf-8") as fh:
            for line in fh:
                if ":" not in line:
                    continue
                key, raw_val = line.split(":", 1)
                token = raw_val.strip().split(" ")[0]
                if token.isdigit():
                    out[key.strip()] = int(token)
    except Exception:
        return {}
    return out


def _proc_cpu_times() -> tuple[int, int]:
    try:
        with open("/proc/stat", "r", encoding="utf-8") as fh:
            first = fh.readline().strip().split()
        if not first or first[0] != "cpu":
            return (0, 0)
        nums = [int(x) for x in first[1:] if str(x).isdigit()]
        if not nums:
            return (0, 0)
        total = int(sum(nums))
        idle = int(nums[3] + (nums[4] if len(nums) > 4 else 0))
        return (total, idle)
    except Exception:
        return (0, 0)


def _proc_net_bytes() -> tuple[int, int]:
    rx = 0
    tx = 0
    try:
        with open("/proc/net/dev", "r", encoding="utf-8") as fh:
            rows = fh.readlines()[2:]
        for line in rows:
            if ":" not in line:
                continue
            iface, data = line.split(":", 1)
            iface_name = iface.strip()
            if iface_name in {"lo"}:
                continue
            cols = [x for x in data.strip().split() if x]
            if len(cols) < 16:
                continue
            rx += int(cols[0] or 0)
            tx += int(cols[8] or 0)
    except Exception:
        return (0, 0)
    return (rx, tx)


def _collect_server_metrics() -> dict[str, Any]:
    global _SERVER_CPU_SNAPSHOT, _SERVER_NET_SNAPSHOT, _SERVER_NET_TS
    now = datetime.utcnow().timestamp()
    disk = os.statvfs("/")
    disk_total = int(disk.f_blocks * disk.f_frsize)
    disk_free = int(disk.f_bavail * disk.f_frsize)
    disk_used = max(0, disk_total - disk_free)
    disk_pct = round((disk_used / disk_total * 100.0), 2) if disk_total > 0 else 0.0

    mem = _proc_meminfo_kb()
    mem_total = int(mem.get("MemTotal", 0) * 1024)
    mem_available = int(mem.get("MemAvailable", mem.get("MemFree", 0)) * 1024)
    mem_used = max(0, mem_total - mem_available)
    mem_pct = round((mem_used / mem_total * 100.0), 2) if mem_total > 0 else 0.0
    swap_total = int(mem.get("SwapTotal", 0) * 1024)
    swap_free = int(mem.get("SwapFree", 0) * 1024)
    swap_used = max(0, swap_total - swap_free)

    cpu_total, cpu_idle = _proc_cpu_times()
    cpu_pct = 0.0
    if _SERVER_CPU_SNAPSHOT is not None:
        prev_total, prev_idle = _SERVER_CPU_SNAPSHOT
        delta_total = max(0, cpu_total - prev_total)
        delta_idle = max(0, cpu_idle - prev_idle)
        if delta_total > 0:
            cpu_pct = round((1.0 - (delta_idle / delta_total)) * 100.0, 2)
    else:
        try:
            load1 = float(os.getloadavg()[0])
            cpu_count = max(1, int(os.cpu_count() or 1))
            cpu_pct = round(min(100.0, (load1 / cpu_count) * 100.0), 2)
        except Exception:
            cpu_pct = 0.0
    _SERVER_CPU_SNAPSHOT = (cpu_total, cpu_idle)

    net_rx, net_tx = _proc_net_bytes()
    rx_rate = 0.0
    tx_rate = 0.0
    if _SERVER_NET_SNAPSHOT is not None and _SERVER_NET_TS > 0:
        prev_rx, prev_tx = _SERVER_NET_SNAPSHOT
        delta_sec = max(1e-6, now - _SERVER_NET_TS)
        rx_rate = max(0.0, (net_rx - prev_rx) / delta_sec)
        tx_rate = max(0.0, (net_tx - prev_tx) / delta_sec)
    _SERVER_NET_SNAPSHOT = (net_rx, net_tx)
    _SERVER_NET_TS = now

    uptime_seconds = 0
    try:
        with open("/proc/uptime", "r", encoding="utf-8") as fh:
            uptime_seconds = int(float(fh.read().split()[0]))
    except Exception:
        uptime_seconds = 0

    try:
        load_avg = os.getloadavg()
    except Exception:
        load_avg = (0.0, 0.0, 0.0)

    return {
        "timestamp": datetime.utcnow().isoformat(),
        "uptime_seconds": uptime_seconds,
        "cpu": {
            "usage_percent": cpu_pct,
            "cores": int(os.cpu_count() or 1),
            "load_avg_1m": round(float(load_avg[0] or 0.0), 3),
            "load_avg_5m": round(float(load_avg[1] or 0.0), 3),
            "load_avg_15m": round(float(load_avg[2] or 0.0), 3),
        },
        "memory": {
            "total_bytes": mem_total,
            "used_bytes": mem_used,
            "available_bytes": mem_available,
            "usage_percent": mem_pct,
            "swap_total_bytes": swap_total,
            "swap_used_bytes": swap_used,
            "swap_free_bytes": swap_free,
        },
        "disk": {
            "mount": "/",
            "total_bytes": disk_total,
            "used_bytes": disk_used,
            "free_bytes": disk_free,
            "usage_percent": disk_pct,
        },
        "network": {
            "rx_bytes_total": int(net_rx),
            "tx_bytes_total": int(net_tx),
            "rx_bytes_per_sec": round(float(rx_rate), 2),
            "tx_bytes_per_sec": round(float(tx_rate), 2),
        },
    }


@router.get("/admin/server/metrics", response_model=dict[str, Any])
def admin_server_metrics(_: User = Depends(get_admin_user), db: Session = Depends(get_db)):
    payload = _collect_server_metrics()
    cache_stats = get_market_cache_stats(db)
    payload["market_cache"] = {
        "entries": int(cache_stats.get("entries") or 0),
        "hits": int(cache_stats.get("hits") or 0),
        "refreshes": int(cache_stats.get("refreshes") or 0),
        "expired": int(cache_stats.get("expired") or 0),
        "api_calls_saved": int(cache_stats.get("api_calls_saved") or 0),
    }
    payload["performance"] = collect_perf_metrics(window_sec=15 * 60)
    payload["task_queue"] = {
        "enabled": queue_enabled(),
        "available": queue_available(),
        "depth": queue_depth(),
    }
    return payload


@router.post("/admin/users/password", response_model=MessageOut)
def admin_reset_password(payload: AdminPasswordResetIn, me: User = Depends(get_admin_user), db: Session = Depends(get_db)):
    user = db.get(User, payload.user_id)
    if not user:
        raise HTTPException(status_code=404, detail="РџРѕР»СЊР·РѕРІР°С‚РµР»СЊ РЅРµ РЅР°Р№РґРµРЅ")
    if len(payload.new_password) < 8:
        raise HTTPException(status_code=400, detail="РџР°СЂРѕР»СЊ РґРѕР»Р¶РµРЅ Р±С‹С‚СЊ РјРёРЅРёРјСѓРј 8 СЃРёРјРІРѕР»РѕРІ")

    user.hashed_password = get_password_hash(payload.new_password)
    owner_member = db.scalar(
        select(TeamMember)
        .where(
            TeamMember.user_id == user.id,
            TeamMember.is_owner.is_(True),
        )
        .order_by(TeamMember.id.asc())
    )
    if owner_member:
        owner_member.hashed_password = user.hashed_password
    _audit(
        db,
        me,
        action="admin_password_reset",
        details=f"user_id={user.id};password_updated=1",
        module_code="admin",
        entity_type="user",
        entity_id=str(user.id),
    )
    db.commit()
    return MessageOut(message="РџР°СЂРѕР»СЊ РїРѕР»СЊР·РѕРІР°С‚РµР»СЏ РѕР±РЅРѕРІР»РµРЅ")


@router.post("/admin/users/role", response_model=MessageOut)
def admin_set_role(payload: AdminRoleUpdateIn, me: User = Depends(get_admin_user), db: Session = Depends(get_db)):
    user = db.get(User, payload.user_id)
    if not user:
        raise HTTPException(status_code=404, detail="РџРѕР»СЊР·РѕРІР°С‚РµР»СЊ РЅРµ РЅР°Р№РґРµРЅ")
    role = payload.role.strip().lower()
    if role not in {"admin", "client"}:
        raise HTTPException(status_code=400, detail="role РґРѕР»Р¶РµРЅ Р±С‹С‚СЊ admin РёР»Рё client")
    if user.id == me.id and role != "admin":
        raise HTTPException(status_code=400, detail="РќРµР»СЊР·СЏ СЃРЅСЏС‚СЊ admin c С‚РµРєСѓС‰РµРіРѕ РїРѕР»СЊР·РѕРІР°С‚РµР»СЏ")
    user.role = role
    _audit(
        db,
        me,
        action="admin_role_updated",
        details=f"user_id={user.id};role={role}",
        module_code="admin",
        entity_type="user",
        entity_id=str(user.id),
    )
    db.commit()
    return MessageOut(message="Р РѕР»СЊ РїРѕР»СЊР·РѕРІР°С‚РµР»СЏ РѕР±РЅРѕРІР»РµРЅР°")


@router.delete("/admin/users/{user_id}", response_model=MessageOut)
def admin_delete_user(user_id: int, me: User = Depends(get_admin_user), db: Session = Depends(get_db)):
    user = db.get(User, user_id)
    if not user:
        raise HTTPException(status_code=404, detail="РџРѕР»СЊР·РѕРІР°С‚РµР»СЊ РЅРµ РЅР°Р№РґРµРЅ")
    if user.id == me.id:
        raise HTTPException(status_code=400, detail="РќРµР»СЊР·СЏ СѓРґР°Р»РёС‚СЊ С‚РµРєСѓС‰РµРіРѕ Р°РґРјРёРЅРёСЃС‚СЂР°С‚РѕСЂР°")
    db.delete(user)
    _audit(
        db,
        me,
        action="admin_user_deleted",
        details=f"user_id={user_id}",
        module_code="admin",
        entity_type="user",
        entity_id=str(user_id),
    )
    db.commit()
    return MessageOut(message="РџРѕР»СЊР·РѕРІР°С‚РµР»СЊ СѓРґР°Р»РµРЅ")


@router.get("/admin/modules", response_model=list[ModuleAccessOut])
def admin_modules(_: User = Depends(get_admin_user), db: Session = Depends(get_db)):
    rows = db.scalars(select(ModuleAccess)).all()
    return [ModuleAccessOut(user_id=r.user_id, module_code=r.module_code, enabled=r.enabled) for r in rows]


@router.get("/admin/modules/mobile", response_model=list[ModuleAccessOut])
def admin_mobile_modules(
    user_id: int | None = None,
    _: User = Depends(get_admin_user),
    db: Session = Depends(get_db),
):
    overrides = _get_mobile_module_overrides(db)
    out: list[ModuleAccessOut] = []
    for user_key, module_map in overrides.items():
        uid = _to_int_safe(user_key)
        if uid <= 0:
            continue
        if user_id and int(user_id) > 0 and uid != int(user_id):
            continue
        for module_code, enabled in module_map.items():
            out.append(ModuleAccessOut(user_id=uid, module_code=module_code, enabled=bool(enabled)))
    out.sort(key=lambda x: (int(x.user_id or 0), str(x.module_code or "")))
    return out


@router.get("/admin/ui/settings", response_model=UiSettingsOut)
def admin_get_ui_settings(_: User = Depends(get_admin_user), db: Session = Depends(get_db)):
    return _get_ui_settings(db)


@router.post("/admin/ui/settings", response_model=UiSettingsOut)
def admin_save_ui_settings(payload: UiSettingsIn, me: User = Depends(get_admin_user), db: Session = Depends(get_db)):
    next_payload = _sanitize_ui_settings_payload(
        {
            "theme_choice_enabled": bool(payload.theme_choice_enabled),
            "force_theme": bool(payload.force_theme),
            "default_theme": payload.default_theme,
            "allowed_themes": payload.allowed_themes,
        }
    )
    _set_system_setting(db, "ui_settings", json.dumps(next_payload, ensure_ascii=False))
    _audit(
        db,
        me,
        action="admin_ui_settings_updated",
        details=json.dumps(next_payload, ensure_ascii=False),
        module_code="admin",
        entity_type="ui_settings",
    )
    db.commit()
    return UiSettingsOut(**next_payload)


@router.post("/admin/modules", response_model=ModuleAccessOut)
def set_module_access(payload: ModuleAccessIn, me: User = Depends(get_admin_user), db: Session = Depends(get_db)):
    if payload.module_code not in DEFAULT_MODULES:
        raise HTTPException(status_code=400, detail=f"РќРµРёР·РІРµСЃС‚РЅС‹Р№ module_code: {payload.module_code}")
    row = db.scalar(select(ModuleAccess).where(ModuleAccess.user_id == payload.user_id, ModuleAccess.module_code == payload.module_code))
    if not row:
        row = ModuleAccess(user_id=payload.user_id, module_code=payload.module_code, enabled=payload.enabled)
        db.add(row)
    else:
        row.enabled = payload.enabled

    _audit(
        db,
        me,
        action="admin_module_updated",
        details=f"user_id={payload.user_id};module={payload.module_code};enabled={payload.enabled}",
        module_code="admin",
        entity_type="module_access",
        entity_id=f"{payload.user_id}:{payload.module_code}",
    )
    db.commit()
    return ModuleAccessOut(user_id=row.user_id, module_code=row.module_code, enabled=row.enabled)


@router.post("/admin/modules/mobile", response_model=ModuleAccessOut)
def set_mobile_module_access(payload: ModuleAccessIn, me: User = Depends(get_admin_user), db: Session = Depends(get_db)):
    code = str(payload.module_code or "").strip().lower()
    if code not in DEFAULT_MODULES:
        raise HTTPException(status_code=400, detail=f"РќРµРёР·РІРµСЃС‚РЅС‹Р№ module_code: {payload.module_code}")
    target_user = db.get(User, int(payload.user_id or 0))
    if not target_user:
        raise HTTPException(status_code=404, detail="РџРѕР»СЊР·РѕРІР°С‚РµР»СЊ РЅРµ РЅР°Р№РґРµРЅ")
    safe = _set_mobile_module_override(db, user_id=int(payload.user_id), module_code=code, enabled=bool(payload.enabled))
    _audit(
        db,
        me,
        action="admin_mobile_module_updated",
        details=f"user_id={payload.user_id};module={code};enabled={1 if payload.enabled else 0}",
        module_code="admin",
        entity_type="mobile_module_access",
        entity_id=f"{payload.user_id}:{code}",
    )
    db.commit()
    user_map = safe.get(str(int(payload.user_id)), {})
    current_enabled = bool(user_map.get(code, bool(payload.enabled)))
    return ModuleAccessOut(user_id=int(payload.user_id), module_code=code, enabled=current_enabled)


@router.get("/admin/credentials", response_model=list[ApiCredentialOut])
def admin_list_credentials(user_id: int, _: User = Depends(get_admin_user), db: Session = Depends(get_db)):
    creds = db.scalars(
        select(ApiCredential).where(ApiCredential.user_id == user_id).order_by(ApiCredential.id.desc())
    ).all()
    return [ApiCredentialOut(id=c.id, marketplace=c.marketplace, api_key_masked=mask_key(c.api_key), active=c.active) for c in creds]


@router.get("/admin/credentials/all", response_model=list[AdminCredentialRowOut])
def admin_list_all_credentials(_: User = Depends(get_admin_user), db: Session = Depends(get_db)):
    rows = db.execute(
        select(ApiCredential, User.email)
        .join(User, User.id == ApiCredential.user_id)
        .order_by(ApiCredential.id.desc())
    ).all()
    return [
        AdminCredentialRowOut(
            id=cred.id,
            user_id=cred.user_id,
            user_email=email,
            marketplace=cred.marketplace,
            api_key_masked=mask_key(cred.api_key),
            active=cred.active,
            created_at=cred.created_at,
        )
        for cred, email in rows
    ]


@router.post("/admin/credentials", response_model=ApiCredentialOut)
def admin_save_credential(payload: AdminCredentialIn, me: User = Depends(get_admin_user), db: Session = Depends(get_db)):
    marketplace = validate_marketplace(payload.marketplace)
    user = db.get(User, payload.user_id)
    if not user:
        raise HTTPException(status_code=404, detail="РџРѕР»СЊР·РѕРІР°С‚РµР»СЊ РЅРµ РЅР°Р№РґРµРЅ")

    creds = db.scalars(
        select(ApiCredential)
        .where(ApiCredential.user_id == payload.user_id, ApiCredential.marketplace == marketplace)
        .order_by(ApiCredential.id.desc())
    ).all()
    if creds:
        cred = creds[0]
        cred.api_key = payload.api_key
        cred.active = True
        for stale in creds[1:]:
            stale.active = False
    else:
        cred = ApiCredential(user_id=payload.user_id, marketplace=marketplace, api_key=payload.api_key, active=True)
        db.add(cred)

    _audit(
        db,
        me,
        action="admin_credential_saved",
        details=f"user_id={payload.user_id};marketplace={marketplace}",
        module_code="admin",
        entity_type="api_credential",
        entity_id=f"{payload.user_id}:{marketplace}",
    )
    db.commit()
    return ApiCredentialOut(id=cred.id, marketplace=cred.marketplace, api_key_masked=mask_key(cred.api_key), active=cred.active)


@router.delete("/admin/credentials/{credential_id}", response_model=MessageOut)
def admin_delete_credential(credential_id: int, _: User = Depends(get_admin_user), db: Session = Depends(get_db)):
    cred = db.get(ApiCredential, credential_id)
    if not cred:
        raise HTTPException(status_code=404, detail="РљР»СЋС‡ РЅРµ РЅР°Р№РґРµРЅ")
    db.delete(cred)
    db.commit()
    return MessageOut(message="РљР»СЋС‡ СѓРґР°Р»РµРЅ")


@router.get("/admin/ai/global", response_model=dict[str, Any])
def admin_ai_global_state(_: User = Depends(get_admin_user), db: Session = Depends(get_db)):
    default = _get_global_ai_default(db)
    services = db.scalars(
        select(AiServiceAccount)
        .where(AiServiceAccount.user_id.is_(None))
        .order_by(AiServiceAccount.priority.asc(), AiServiceAccount.id.desc())
    ).all()
    return {
        "global_default": {
            "use_global_default": False,
            "mode": str(default.get("mode") or "builtin"),
            "service_id": _to_int_safe(default.get("service_id")),
        },
        "global_services": [_ai_service_to_out(x, scope="global") for x in services],
    }


@router.post("/admin/ai/global/default", response_model=dict[str, Any])
def admin_ai_global_default_save(payload: AiSelectionIn, me: User = Depends(get_admin_user), db: Session = Depends(get_db)):
    mode = _normalize_ai_mode(payload.mode)
    if mode == "user":
        raise HTTPException(status_code=400, detail="Р“Р»РѕР±Р°Р»СЊРЅС‹Р№ default РЅРµ РјРѕР¶РµС‚ СЃСЃС‹Р»Р°С‚СЊСЃСЏ РЅР° user-СЃРµСЂРІРёСЃ")
    service_id = _validate_ai_service_binding(db, mode=mode, service_id=payload.service_id, user_id=None)
    _set_system_setting(db, "ai_global_default", json.dumps({"mode": mode, "service_id": service_id}, ensure_ascii=False))
    _audit(
        db,
        me,
        action="admin_ai_global_default_saved",
        details=f"mode={mode};service_id={service_id}",
        module_code="admin",
        entity_type="ai_selection",
    )
    db.commit()
    return {
        "global_default": {
            "use_global_default": False,
            "mode": mode,
            "service_id": service_id,
        }
    }


@router.post("/admin/ai/global/services", response_model=AiServiceOut)
def admin_ai_global_service_add(payload: AiServiceIn, me: User = Depends(get_admin_user), db: Session = Depends(get_db)):
    row = _upsert_ai_service(db, user_id=None, payload=payload)
    _audit(
        db,
        me,
        action="admin_ai_global_service_added",
        details=f"service_id={row.id};provider={row.provider}",
        module_code="admin",
        entity_type="ai_service",
        entity_id=str(row.id),
    )
    db.commit()
    return _ai_service_to_out(row, scope="global")


@router.put("/admin/ai/global/services/{service_id}", response_model=AiServiceOut)
def admin_ai_global_service_update(service_id: int, payload: AiServiceIn, me: User = Depends(get_admin_user), db: Session = Depends(get_db)):
    row = db.get(AiServiceAccount, service_id)
    if not row or row.user_id is not None:
        raise HTTPException(status_code=404, detail="AI СЃРµСЂРІРёСЃ РЅРµ РЅР°Р№РґРµРЅ")
    _update_ai_service_row(row, payload)
    _audit(
        db,
        me,
        action="admin_ai_global_service_updated",
        details=f"service_id={row.id};provider={row.provider}",
        module_code="admin",
        entity_type="ai_service",
        entity_id=str(row.id),
    )
    db.commit()
    return _ai_service_to_out(row, scope="global")


@router.delete("/admin/ai/global/services/{service_id}", response_model=MessageOut)
def admin_ai_global_service_delete(service_id: int, me: User = Depends(get_admin_user), db: Session = Depends(get_db)):
    row = db.get(AiServiceAccount, service_id)
    if not row or row.user_id is not None:
        raise HTTPException(status_code=404, detail="AI СЃРµСЂРІРёСЃ РЅРµ РЅР°Р№РґРµРЅ")
    global_default = _get_global_ai_default(db)
    if str(global_default.get("mode") or "") == "global" and _to_int_safe(global_default.get("service_id")) == service_id:
        _set_system_setting(db, "ai_global_default", json.dumps({"mode": "builtin", "service_id": None}, ensure_ascii=False))
    prefs = db.scalars(
        select(UserAiPreference).where(
            UserAiPreference.mode == "global",
            UserAiPreference.service_id == service_id,
            UserAiPreference.use_global_default.is_(False),
        )
    ).all()
    for pref in prefs:
        pref.use_global_default = True
        pref.mode = "builtin"
        pref.service_id = None
    db.delete(row)
    _audit(
        db,
        me,
        action="admin_ai_global_service_deleted",
        details=f"service_id={service_id}",
        module_code="admin",
        entity_type="ai_service",
        entity_id=str(service_id),
    )
    db.commit()
    return MessageOut(message="AI СЃРµСЂРІРёСЃ СѓРґР°Р»РµРЅ")


@router.post("/admin/ai/global/services/reorder", response_model=list[AiServiceOut])
def admin_ai_global_service_reorder(payload: AiServiceReorderIn, me: User = Depends(get_admin_user), db: Session = Depends(get_db)):
    ordered = _reorder_ai_services(db, user_id=None, service_ids=list(payload.service_ids or []))
    _audit(
        db,
        me,
        action="admin_ai_global_service_reordered",
        details=json.dumps({"service_ids": [int(x) for x in payload.service_ids or []]}, ensure_ascii=False),
        module_code="admin",
        entity_type="ai_service",
    )
    db.commit()
    return [_ai_service_to_out(row, scope="global") for row in ordered]


@router.get("/admin/users/{user_id}/ai", response_model=AiProfileOut)
def admin_user_ai_state(user_id: int, _: User = Depends(get_admin_user), db: Session = Depends(get_db)):
    target = db.get(User, user_id)
    if not target:
        raise HTTPException(status_code=404, detail="РџРѕР»СЊР·РѕРІР°С‚РµР»СЊ РЅРµ РЅР°Р№РґРµРЅ")
    return _build_ai_profile_payload(db, user_id)


@router.post("/admin/users/{user_id}/ai/select", response_model=AiProfileOut)
def admin_user_ai_select(
    user_id: int,
    payload: AiSelectionIn,
    me: User = Depends(get_admin_user),
    db: Session = Depends(get_db),
):
    target = db.get(User, user_id)
    if not target:
        raise HTTPException(status_code=404, detail="РџРѕР»СЊР·РѕРІР°С‚РµР»СЊ РЅРµ РЅР°Р№РґРµРЅ")
    _save_user_ai_selection(
        db,
        user_id,
        use_global_default=bool(payload.use_global_default),
        mode=payload.mode,
        service_id=payload.service_id,
    )
    _audit(
        db,
        me,
        action="admin_user_ai_selected",
        details=f"user_id={user_id};use_global_default={bool(payload.use_global_default)};mode={payload.mode};service_id={payload.service_id}",
        module_code="admin",
        entity_type="ai_selection",
        entity_id=str(user_id),
    )
    db.commit()
    return _build_ai_profile_payload(db, user_id)


@router.post("/admin/users/{user_id}/ai/services", response_model=AiServiceOut)
def admin_user_ai_service_add(
    user_id: int,
    payload: AiServiceIn,
    me: User = Depends(get_admin_user),
    db: Session = Depends(get_db),
):
    target = db.get(User, user_id)
    if not target:
        raise HTTPException(status_code=404, detail="РџРѕР»СЊР·РѕРІР°С‚РµР»СЊ РЅРµ РЅР°Р№РґРµРЅ")
    row = _upsert_ai_service(db, user_id=user_id, payload=payload)
    _audit(
        db,
        me,
        action="admin_user_ai_service_added",
        details=f"user_id={user_id};service_id={row.id};provider={row.provider}",
        module_code="admin",
        entity_type="ai_service",
        entity_id=f"{user_id}:{row.id}",
    )
    db.commit()
    return _ai_service_to_out(row, scope="user")


@router.put("/admin/users/{user_id}/ai/services/{service_id}", response_model=AiServiceOut)
def admin_user_ai_service_update(
    user_id: int,
    service_id: int,
    payload: AiServiceIn,
    me: User = Depends(get_admin_user),
    db: Session = Depends(get_db),
):
    target = db.get(User, user_id)
    if not target:
        raise HTTPException(status_code=404, detail="РџРѕР»СЊР·РѕРІР°С‚РµР»СЊ РЅРµ РЅР°Р№РґРµРЅ")
    row = db.get(AiServiceAccount, service_id)
    if not row or row.user_id != user_id:
        raise HTTPException(status_code=404, detail="AI СЃРµСЂРІРёСЃ РЅРµ РЅР°Р№РґРµРЅ")
    _update_ai_service_row(row, payload)
    _audit(
        db,
        me,
        action="admin_user_ai_service_updated",
        details=f"user_id={user_id};service_id={row.id};provider={row.provider}",
        module_code="admin",
        entity_type="ai_service",
        entity_id=f"{user_id}:{row.id}",
    )
    db.commit()
    return _ai_service_to_out(row, scope="user")


@router.delete("/admin/users/{user_id}/ai/services/{service_id}", response_model=MessageOut)
def admin_user_ai_service_delete(user_id: int, service_id: int, me: User = Depends(get_admin_user), db: Session = Depends(get_db)):
    target = db.get(User, user_id)
    if not target:
        raise HTTPException(status_code=404, detail="РџРѕР»СЊР·РѕРІР°С‚РµР»СЊ РЅРµ РЅР°Р№РґРµРЅ")
    row = db.get(AiServiceAccount, service_id)
    if not row or row.user_id != user_id:
        raise HTTPException(status_code=404, detail="AI СЃРµСЂРІРёСЃ РЅРµ РЅР°Р№РґРµРЅ")
    _reset_ai_selection_if_deleted_service(db, user_id, service_id)
    db.delete(row)
    _audit(
        db,
        me,
        action="admin_user_ai_service_deleted",
        details=f"user_id={user_id};service_id={service_id}",
        module_code="admin",
        entity_type="ai_service",
        entity_id=f"{user_id}:{service_id}",
    )
    db.commit()
    return MessageOut(message="AI СЃРµСЂРІРёСЃ СѓРґР°Р»РµРЅ")


@router.post("/admin/users/{user_id}/ai/services/reorder", response_model=list[AiServiceOut])
def admin_user_ai_service_reorder(
    user_id: int,
    payload: AiServiceReorderIn,
    me: User = Depends(get_admin_user),
    db: Session = Depends(get_db),
):
    target = db.get(User, user_id)
    if not target:
        raise HTTPException(status_code=404, detail="РџРѕР»СЊР·РѕРІР°С‚РµР»СЊ РЅРµ РЅР°Р№РґРµРЅ")
    ordered = _reorder_ai_services(db, user_id=user_id, service_ids=list(payload.service_ids or []))
    _audit(
        db,
        me,
        action="admin_user_ai_service_reordered",
        details=json.dumps({"user_id": user_id, "service_ids": [int(x) for x in payload.service_ids or []]}, ensure_ascii=False),
        module_code="admin",
        entity_type="ai_service",
        entity_id=str(user_id),
    )
    db.commit()
    return [_ai_service_to_out(row, scope="user") for row in ordered]


@router.get("/admin/audit", response_model=AuditLogPageOut)
def admin_audit(
    limit: int = 0,
    page_size: int = 100,
    page: int = 1,
    action: str = "",
    module_code: str = "",
    status: str = "",
    user_id: int | None = None,
    actor_email: str = "",
    actor_member_id: int | None = None,
    q: str = "",
    date_from: date | None = None,
    date_to: date | None = None,
    _: User = Depends(get_admin_user),
    db: Session = Depends(get_db),
):
    safe_limit = int(limit or 0)
    requested_page_size = safe_limit if safe_limit > 0 else int(page_size or 100)
    clamped_page_size = max(10, min(requested_page_size, 500))
    current_page = max(1, int(page or 1))
    query = select(AuditLog)
    safe_action = str(action or "").strip()
    safe_module = str(module_code or "").strip()
    safe_status = str(status or "").strip()
    safe_actor_email = str(actor_email or "").strip().lower()
    safe_q = str(q or "").strip().lower()
    if safe_action:
        query = query.where(AuditLog.action.ilike(f"%{safe_action}%"))
    if safe_module:
        query = query.where(AuditLog.module_code.ilike(f"%{safe_module}%"))
    if safe_status:
        query = query.where(AuditLog.status.ilike(f"%{safe_status}%"))
    if user_id is not None and int(user_id) > 0:
        query = query.where(AuditLog.user_id == int(user_id))
    if actor_member_id is not None and int(actor_member_id) > 0:
        query = query.where(AuditLog.actor_member_id == int(actor_member_id))
    if safe_actor_email:
        query = query.where(func.lower(AuditLog.actor_email).like(f"%{safe_actor_email}%"))
    if date_from is not None:
        query = query.where(AuditLog.created_at >= datetime.combine(date_from, datetime.min.time()))
    if date_to is not None:
        query = query.where(AuditLog.created_at <= datetime.combine(date_to, datetime.max.time()))
    if safe_q:
        pattern = f"%{safe_q}%"
        query = query.where(
            or_(
                func.lower(AuditLog.action).like(pattern),
                func.lower(AuditLog.module_code).like(pattern),
                func.lower(AuditLog.status).like(pattern),
                func.lower(AuditLog.details).like(pattern),
                func.lower(AuditLog.entity_type).like(pattern),
                func.lower(AuditLog.entity_id).like(pattern),
                func.lower(AuditLog.actor_email).like(pattern),
                func.lower(cast(AuditLog.user_id, String)).like(pattern),
                func.lower(cast(AuditLog.actor_member_id, String)).like(pattern),
            )
        )
    total = int(db.scalar(select(func.count()).select_from(query.subquery())) or 0)
    offset = max(0, (current_page - 1) * clamped_page_size)
    rows = db.scalars(
        query.order_by(AuditLog.id.desc()).offset(offset).limit(clamped_page_size)
    ).all()
    total_pages = max(1, math.ceil(total / clamped_page_size)) if total else 0
    if total_pages and current_page > total_pages:
        current_page = total_pages
        offset = max(0, (current_page - 1) * clamped_page_size)
        rows = db.scalars(
            query.order_by(AuditLog.id.desc()).offset(offset).limit(clamped_page_size)
        ).all()
    return AuditLogPageOut(
        rows=rows,
        total=total,
        page=current_page,
        page_size=clamped_page_size,
        total_pages=total_pages,
    )


SOCIAL_GAMES: dict[str, str] = {
    "snake": "Р—РјРµР№РєР°",
    "tetris": "РўРµС‚СЂРёСЃ",
    "2048": "2048",
    "checkers": "РЁР°С€РєРё",
    "chess": "РЁР°С…РјР°С‚С‹",
    "battleship": "РњРѕСЂСЃРєРѕР№ Р±РѕР№",
}

SOCIAL_SCORE_GAMES: set[str] = {"snake", "tetris", "2048"}

_SOCIAL_BOARD_PROFILE_PREFIX_GAMES: set[str] = {CHESS_GAME_CODE, BATTLESHIP_GAME_CODE}

_SOCIAL_MSG_REQUEST_CACHE: dict[str, tuple[int, float]] = {}
_SOCIAL_MSG_REQUEST_CACHE_TTL_SEC = 15 * 60


def _social_msg_cache_get(cache_key: str) -> int:
    now_ts = time.time()
    stale = [
        key
        for key, (_, ts) in _SOCIAL_MSG_REQUEST_CACHE.items()
        if (now_ts - float(ts or 0.0)) > _SOCIAL_MSG_REQUEST_CACHE_TTL_SEC
    ]
    for key in stale:
        _SOCIAL_MSG_REQUEST_CACHE.pop(key, None)
    data = _SOCIAL_MSG_REQUEST_CACHE.get(cache_key)
    return int(data[0] or 0) if data else 0


def _social_msg_cache_set(cache_key: str, message_id: int) -> None:
    _SOCIAL_MSG_REQUEST_CACHE[cache_key] = (int(message_id or 0), time.time())


def _social_actor_key(user: User) -> str:
    member_id = _actor_member_id(user)
    if not _actor_is_owner(user) and member_id > 0:
        return f"m:{member_id}"
    return f"u:{int(user.id)}"


def _social_canonical_actor_key(db: Session, actor_key: str) -> str:
    key = str(actor_key or "").strip().lower()
    if key.startswith("m:"):
        member_id = _to_int_safe(key.split(":", 1)[1])
        member = db.get(TeamMember, member_id) if member_id > 0 else None
        if member and bool(member.is_owner):
            return f"u:{int(member.user_id)}"
    return key


def _social_actor_alias_keys(db: Session, actor_key: str) -> list[str]:
    canonical = _social_canonical_actor_key(db, actor_key)
    out: set[str] = {canonical}
    if canonical.startswith("u:"):
        user_id = _to_int_safe(canonical.split(":", 1)[1])
        if user_id > 0:
            owner_member = db.scalar(
                select(TeamMember.id).where(
                    TeamMember.user_id == user_id,
                    TeamMember.is_owner.is_(True),
                    TeamMember.is_active.is_(True),
                ).order_by(TeamMember.id.asc())
            )
            if owner_member:
                out.add(f"m:{int(owner_member)}")
    return sorted(x for x in out if x)


def _social_actor_identity(db: Session, user: User) -> tuple[str, str, int | None]:
    actor_key = _social_actor_key(user)
    if actor_key.startswith("m:"):
        member_id = _to_int_safe(actor_key.split(":", 1)[1])
        member = db.get(TeamMember, member_id) if member_id else None
        nick = str((member.nickname if member and member.nickname else member.full_name if member else "") or "").strip()
        if not nick:
            nick = (member.email if member else _actor_email(user) or user.email).split("@")[0]
        return actor_key, nick[:120], member_id
    owner_member = db.scalar(
        select(TeamMember).where(
            TeamMember.user_id == user.id,
            TeamMember.is_owner.is_(True),
        ).order_by(TeamMember.id.asc())
    )
    nick = str((owner_member.nickname if owner_member and owner_member.nickname else owner_member.full_name if owner_member else "") or "").strip()
    if not nick:
        profile = db.scalar(select(UserProfile).where(UserProfile.user_id == user.id))
        nick = str((profile.full_name if profile and profile.full_name else "") or "").strip()
    if not nick:
        nick = str(user.email or "").split("@")[0]
    return actor_key, nick[:120], (owner_member.id if owner_member else None)


def _social_identity_by_key(db: Session, actor_key: str) -> tuple[int, int | None, str]:
    key = str(actor_key or "").strip().lower()
    if key.startswith("m:"):
        member_id = _to_int_safe(key.split(":", 1)[1])
        member = db.get(TeamMember, member_id) if member_id else None
        if not member or not member.is_active:
            raise HTTPException(status_code=404, detail="РЈС‡Р°СЃС‚РЅРёРє РЅРµ РЅР°Р№РґРµРЅ")
        nick = str((member.nickname or member.full_name or member.email).strip() or f"member-{member.id}")
        return int(member.user_id), int(member.id), nick[:120]
    if key.startswith("u:"):
        user_id = _to_int_safe(key.split(":", 1)[1])
        owner = db.get(User, user_id) if user_id else None
        if not owner:
            raise HTTPException(status_code=404, detail="РџРѕР»СЊР·РѕРІР°С‚РµР»СЊ РЅРµ РЅР°Р№РґРµРЅ")
        owner_member = db.scalar(
            select(TeamMember).where(
                TeamMember.user_id == owner.id,
                TeamMember.is_owner.is_(True),
            ).order_by(TeamMember.id.asc())
        )
        nick = ""
        if owner_member:
            nick = str((owner_member.nickname or owner_member.full_name or "").strip())
        if not nick:
            nick = str(owner.email or "").split("@")[0]
        return int(owner.id), None, nick[:120]
    raise HTTPException(status_code=400, detail="РќРµРєРѕСЂСЂРµРєС‚РЅС‹Р№ actor_key")


def _social_current_nick_by_key(db: Session, actor_key: str) -> str:
    key = str(actor_key or "").strip().lower()
    if not key:
        return ""
    if key.startswith("m:"):
        member_id = _to_int_safe(key.split(":", 1)[1])
        member = db.get(TeamMember, member_id) if member_id else None
        if member and member.is_active:
            return str((member.nickname or member.full_name or member.email).strip())
        return ""
    if key.startswith("u:"):
        user_id = _to_int_safe(key.split(":", 1)[1])
        owner = db.get(User, user_id) if user_id else None
        if not owner:
            return ""
        owner_member = db.scalar(
            select(TeamMember).where(
                TeamMember.user_id == owner.id,
                TeamMember.is_owner.is_(True),
            ).order_by(TeamMember.id.asc())
        )
        nick = ""
        if owner_member:
            nick = str((owner_member.nickname or owner_member.full_name or "").strip())
        if not nick:
            profile = db.scalar(select(UserProfile).where(UserProfile.user_id == owner.id))
            nick = str((profile.full_name if profile and profile.full_name else "") or "").strip()
        if not nick:
            nick = str(owner.email or "").split("@")[0]
        return nick
    return ""


def _social_current_avatar_by_key(db: Session, actor_key: str) -> str:
    key = str(actor_key or "").strip().lower()
    if not key:
        return ""
    if key.startswith("m:"):
        member_id = _to_int_safe(key.split(":", 1)[1])
        member = db.get(TeamMember, member_id) if member_id else None
        if member and member.is_active:
            return str(member.avatar_url or "").strip()
        return ""
    if key.startswith("u:"):
        user_id = _to_int_safe(key.split(":", 1)[1])
        owner = db.get(User, user_id) if user_id else None
        if not owner:
            return ""
        owner_member = db.scalar(
            select(TeamMember).where(
                TeamMember.user_id == owner.id,
                TeamMember.is_owner.is_(True),
            ).order_by(TeamMember.id.asc())
        )
        if owner_member and owner_member.avatar_url:
            return str(owner_member.avatar_url or "").strip()
        profile = db.scalar(select(UserProfile).where(UserProfile.user_id == owner.id))
        return str(profile.avatar_url or "").strip() if profile else ""
    return ""


def _social_ensure_thread_member(
    db: Session,
    *,
    thread_id: int,
    actor_key: str,
    user_id: int,
    member_id: int | None,
    actor_nick: str,
) -> SocialChatThreadMember:
    row = db.scalar(
        select(SocialChatThreadMember).where(
            SocialChatThreadMember.thread_id == thread_id,
            SocialChatThreadMember.actor_key == actor_key,
        )
    )
    if row:
        row.user_id = user_id
        row.member_id = member_id
        row.actor_nick = actor_nick[:120]
        row.updated_at = datetime.utcnow()
        return row
    row = SocialChatThreadMember(
        thread_id=thread_id,
        user_id=user_id,
        member_id=member_id,
        actor_key=actor_key,
        actor_nick=actor_nick[:120],
        last_read_message_id=0,
    )
    db.add(row)
    db.flush()
    return row


def _social_ensure_global_thread(db: Session) -> SocialChatThread:
    thread = db.scalar(
        select(SocialChatThread).where(
            SocialChatThread.kind == "global",
            SocialChatThread.owner_user_id.is_(None),
        ).order_by(SocialChatThread.id.asc())
    )
    if thread:
        return thread
    thread = SocialChatThread(
        kind="global",
        owner_user_id=None,
        title="Р“Р»РѕР±Р°Р»СЊРЅС‹Р№ С‡Р°С‚",
    )
    db.add(thread)
    db.flush()
    return thread


def _social_ensure_company_thread(db: Session, user_id: int) -> SocialChatThread:
    thread = db.scalar(
        select(SocialChatThread).where(
            SocialChatThread.kind == "company",
            SocialChatThread.owner_user_id == user_id,
        ).order_by(SocialChatThread.id.asc())
    )
    if thread:
        return thread
    thread = SocialChatThread(
        kind="company",
        owner_user_id=user_id,
        title="Р§Р°С‚ РєРѕРјРїР°РЅРёРё",
    )
    db.add(thread)
    db.flush()
    return thread


def _social_thread_last_message(db: Session, thread_id: int) -> SocialChatMessage | None:
    return db.scalar(
        select(SocialChatMessage)
        .where(SocialChatMessage.thread_id == thread_id)
        .order_by(SocialChatMessage.id.desc())
    )


def _social_last_activity_map(db: Session, actor_keys: list[str]) -> dict[str, datetime]:
    keys = [str(x or "").strip().lower() for x in actor_keys if str(x or "").strip()]
    if not keys:
        return {}
    out: dict[str, datetime] = {}
    key_to_member_id: dict[str, int] = {}
    owner_key_to_user_id: dict[str, int] = {}
    owner_user_ids: set[int] = set()

    for key in keys:
        if key.startswith("m:"):
            member_id = _to_int_safe(key.split(":", 1)[1])
            if member_id > 0:
                key_to_member_id[key] = int(member_id)
            continue
        if not key.startswith("u:"):
            continue
        user_id = _to_int_safe(key.split(":", 1)[1])
        if user_id <= 0:
            continue
        owner_key_to_user_id[key] = int(user_id)
        owner_user_ids.add(int(user_id))

    owner_member_by_user: dict[int, int] = {}
    owner_email_by_user: dict[int, str] = {}
    if owner_user_ids:
        owner_rows = db.execute(
            select(TeamMember.user_id, TeamMember.id, TeamMember.email)
            .where(
                TeamMember.user_id.in_(list(owner_user_ids)),
                TeamMember.is_owner.is_(True),
                TeamMember.is_active.is_(True),
            )
            .order_by(TeamMember.user_id.asc(), TeamMember.id.asc())
        ).all()
        for raw_user_id, raw_member_id, raw_email in owner_rows:
            user_id = _to_int_safe(raw_user_id)
            member_id = _to_int_safe(raw_member_id)
            if user_id <= 0 or member_id <= 0 or user_id in owner_member_by_user:
                continue
            owner_member_by_user[user_id] = member_id
            owner_email_by_user[user_id] = str(raw_email or "").strip().lower()
        missing_owner_users = [uid for uid in owner_user_ids if uid not in owner_email_by_user]
        if missing_owner_users:
            user_rows = db.execute(select(User.id, User.email).where(User.id.in_(missing_owner_users))).all()
            for raw_user_id, raw_email in user_rows:
                user_id = _to_int_safe(raw_user_id)
                if user_id > 0 and user_id not in owner_email_by_user:
                    owner_email_by_user[user_id] = str(raw_email or "").strip().lower()

    for key, user_id in owner_key_to_user_id.items():
        member_id = owner_member_by_user.get(int(user_id))
        if member_id:
            key_to_member_id[key] = int(member_id)

    member_ids = sorted({int(v) for v in key_to_member_id.values() if int(v) > 0})
    if member_ids:
        member_rows = db.execute(
            select(
                AuditLog.actor_member_id,
                func.max(AuditLog.created_at).label("last_at"),
            )
            .where(AuditLog.actor_member_id.in_(member_ids))
            .group_by(AuditLog.actor_member_id)
        ).all()
        by_member: dict[int, datetime] = {}
        for raw_member_id, last_at in member_rows:
            member_id = _to_int_safe(raw_member_id)
            if member_id > 0 and isinstance(last_at, datetime):
                by_member[member_id] = last_at
        for key, member_id in key_to_member_id.items():
            dt = by_member.get(int(member_id))
            if isinstance(dt, datetime):
                out[key] = dt

    if owner_user_ids:
        user_ids = sorted({int(v) for v in owner_user_ids if int(v) > 0})
        user_rows = db.execute(
            select(
                AuditLog.user_id,
                func.max(AuditLog.created_at).label("last_at"),
            )
            .where(
                AuditLog.user_id.in_(user_ids),
                AuditLog.user_id.is_not(None),
                AuditLog.actor_is_owner.is_(True),
            )
            .group_by(AuditLog.user_id)
        ).all()
        by_user: dict[int, datetime] = {}
        for raw_user_id, last_at in user_rows:
            user_id = _to_int_safe(raw_user_id)
            if user_id > 0 and isinstance(last_at, datetime):
                by_user[user_id] = last_at
        for key, user_id in owner_key_to_user_id.items():
            dt = by_user.get(int(user_id))
            if not isinstance(dt, datetime):
                continue
            prev = out.get(key)
            if not prev or dt > prev:
                out[key] = dt

    email_to_keys: dict[str, list[str]] = {}
    for key in keys:
        if key in out:
            continue
        # For team members we intentionally avoid email fallback:
        # different users can share one email and that corrupts "last seen".
        if not key.startswith("u:"):
            continue
        owner_user_id = int(owner_key_to_user_id.get(key) or 0)
        email = owner_email_by_user.get(owner_user_id, "")
        safe_email = str(email or "").strip().lower()
        if safe_email:
            email_to_keys.setdefault(safe_email, []).append(key)

    if not email_to_keys:
        return out
    by_email_rows = db.execute(
        select(
            func.lower(AuditLog.actor_email).label("actor_email"),
            func.max(AuditLog.created_at).label("last_at"),
        )
        .where(func.lower(AuditLog.actor_email).in_(list(email_to_keys.keys())))
        .group_by(func.lower(AuditLog.actor_email))
    ).all()
    for actor_email, last_at in by_email_rows:
        email = str(actor_email or "").strip().lower()
        if not email or not isinstance(last_at, datetime):
            continue
        for key in email_to_keys.get(email, []):
            prev = out.get(key)
            if not prev or last_at > prev:
                out[key] = last_at
    return out


_SOCIAL_ONLINE_WINDOW_SEC = 120


def _social_is_online(last_seen: datetime | None) -> bool:
    if not isinstance(last_seen, datetime):
        return False
    safe_seen = last_seen
    if safe_seen.tzinfo is not None:
        safe_seen = safe_seen.astimezone(timezone.utc).replace(tzinfo=None)
    delta = (datetime.utcnow() - safe_seen).total_seconds()
    return 0 <= float(delta or 0.0) <= float(_SOCIAL_ONLINE_WINDOW_SEC)


def _social_public_profile_by_key(db: Session, actor_key: str) -> dict[str, Any]:
    key = _social_canonical_actor_key(db, actor_key)
    payload: dict[str, Any] = {
        "actor_key": key,
        "nick": _social_current_nick_by_key(db, key) or key,
        "full_name": "",
        "email": "",
        "company_name": "",
        "city": "",
        "position_title": "",
        "avatar_url": _social_current_avatar_by_key(db, key) or "",
        "is_owner": False,
        "user_id": 0,
        "member_id": 0,
    }
    if key.startswith("m:"):
        member_id = _to_int_safe(key.split(":", 1)[1])
        member = db.get(TeamMember, member_id) if member_id else None
        if not member:
            return payload
        profile = db.scalar(select(UserProfile).where(UserProfile.user_id == int(member.user_id)))
        payload.update(
            {
                "full_name": str(member.full_name or "").strip(),
                "email": str(member.email or "").strip().lower(),
                "company_name": str(profile.company_name or "").strip() if profile else "",
                "city": str(member.city or "").strip() or (str(profile.city or "").strip() if profile else ""),
                "position_title": str(member.position_title or "").strip() or (str(profile.position_title or "").strip() if profile else ""),
                "is_owner": bool(member.is_owner),
                "user_id": int(member.user_id or 0),
                "member_id": int(member.id or 0),
            }
        )
        return payload
    if key.startswith("u:"):
        user_id = _to_int_safe(key.split(":", 1)[1])
        owner = db.get(User, user_id) if user_id else None
        owner_member = db.scalar(
            select(TeamMember).where(
                TeamMember.user_id == int(user_id or 0),
                TeamMember.is_owner.is_(True),
            ).order_by(TeamMember.id.asc())
        ) if user_id else None
        profile = db.scalar(select(UserProfile).where(UserProfile.user_id == int(user_id or 0))) if user_id else None
        payload.update(
            {
                "full_name": str((owner_member.full_name if owner_member else "") or (profile.full_name if profile else "") or "").strip(),
                "email": str((owner.email if owner else "") or "").strip().lower(),
                "company_name": str(profile.company_name or "").strip() if profile else "",
                "city": str(profile.city or "").strip() if profile else "",
                "position_title": str(profile.position_title or "").strip() if profile else "",
                "is_owner": True,
                "user_id": int(user_id or 0),
                "member_id": int(owner_member.id or 0) if owner_member else 0,
            }
        )
    return payload


def _to_utc_iso(dt: datetime | None) -> str:
    if not isinstance(dt, datetime):
        return ""
    safe = dt
    if safe.tzinfo is None:
        safe = safe.replace(tzinfo=timezone.utc)
    else:
        safe = safe.astimezone(timezone.utc)
    return safe.isoformat().replace("+00:00", "Z")


def _social_parse_attachments(raw: str | None) -> list[dict[str, Any]]:
    try:
        data = json.loads(str(raw or "[]"))
    except Exception:
        return []
    if not isinstance(data, list):
        return []
    out: list[dict[str, Any]] = []
    for item in data:
        if not isinstance(item, dict):
            continue
        url = str(item.get("url") or "").strip()[:500]
        filename = str(item.get("filename") or "").strip()[:255]
        if not url:
            continue
        out.append(
            {
                "url": url,
                "filename": filename or "file",
                "content_type": str(item.get("content_type") or "").strip()[:120],
                "size_bytes": int(max(0, _to_int_safe(item.get("size_bytes")))),
            }
        )
    return out


def _social_parse_reactions(raw: str | None) -> dict[str, list[dict[str, str]]]:
    try:
        data = json.loads(str(raw or "{}"))
    except Exception:
        return {}
    if not isinstance(data, dict):
        return {}
    out: dict[str, list[dict[str, str]]] = {}
    for emoji_raw, actors_raw in data.items():
        emoji = str(emoji_raw or "").strip()
        if not emoji:
            continue
        actor_rows: list[dict[str, str]] = []
        if isinstance(actors_raw, list):
            for item in actors_raw:
                if isinstance(item, dict):
                    key = str(item.get("actor_key") or "").strip()[:60]
                    nick = str(item.get("nick") or "").strip()[:120]
                else:
                    key = str(item or "").strip()[:60]
                    nick = ""
                if not key:
                    continue
                actor_rows.append({"actor_key": key, "nick": nick})
        if actor_rows:
            out[emoji] = actor_rows
    return out


def _social_message_delivery_meta(db: Session, row: SocialChatMessage) -> tuple[str, int, int]:
    cache: dict[str, str] = {}

    def _canon(raw_key: str) -> str:
        key = str(raw_key or "").strip().lower()
        if not key:
            return ""
        if key in cache:
            return cache[key]
        safe = _social_canonical_actor_key(db, key)
        cache[key] = safe
        return safe

    sender_key = _canon(str(row.sender_key or ""))
    members = db.scalars(
        select(SocialChatThreadMember).where(SocialChatThreadMember.thread_id == int(row.thread_id))
    ).all()
    recipient_reads: dict[str, int] = {}
    for member in members:
        raw_key = str(member.actor_key or "").strip().lower()
        canonical_key = _canon(raw_key)
        if not canonical_key or canonical_key == sender_key:
            continue
        last_read = int(member.last_read_message_id or 0)
        prev = recipient_reads.get(canonical_key, 0)
        if last_read > prev:
            recipient_reads[canonical_key] = last_read
    total = len(recipient_reads)
    read_by = sum(1 for last_read in recipient_reads.values() if int(last_read) >= int(row.id or 0))
    status = "read" if total > 0 and read_by >= total else "sent"
    return status, int(read_by), int(total)


def _social_message_to_out(db: Session, actor_key: str, row: SocialChatMessage) -> SocialChatMessageOut:
    sender_key = str(row.sender_key or "")
    sender_nick = _social_current_nick_by_key(db, sender_key) or str(row.sender_nick or "")
    safe_sender_key = _social_canonical_actor_key(db, sender_key.strip().lower())
    safe_actor_key = _social_canonical_actor_key(db, str(actor_key or "").strip().lower())
    is_mine = bool(safe_sender_key and safe_actor_key and safe_sender_key == safe_actor_key)
    delivery_status: str | None = None
    delivery_read_by = 0
    delivery_total = 0
    if is_mine:
        delivery_status, delivery_read_by, delivery_total = _social_message_delivery_meta(db, row)
    reply_payload: dict[str, Any] | None = None
    reply_id = int(row.reply_to_message_id or 0)
    if reply_id > 0:
        reply_row = db.get(SocialChatMessage, reply_id)
        if reply_row and int(reply_row.thread_id or 0) == int(row.thread_id or 0):
            reply_sender_key = str(reply_row.sender_key or "")
            reply_payload = {
                "id": int(reply_row.id),
                "sender_key": reply_sender_key,
                "sender_nick": _social_current_nick_by_key(db, reply_sender_key) or str(reply_row.sender_nick or ""),
                "text": str(reply_row.text or "")[:500],
            }
    parsed_reactions = _social_parse_reactions(getattr(row, "reactions_json", "") or "{}")
    reactions: list[dict[str, Any]] = []
    for emoji, actors in parsed_reactions.items():
        actor_keys = [str(item.get("actor_key") or "") for item in actors if isinstance(item, dict)]
        actor_nicks = [str(item.get("nick") or "").strip() for item in actors if isinstance(item, dict)]
        reactions.append(
            {
                "emoji": emoji,
                "count": len(actor_keys),
                "my": actor_key in actor_keys,
                "actors": [x for x in actor_nicks if x][:8],
            }
        )
    return SocialChatMessageOut(
        id=int(row.id),
        thread_id=int(row.thread_id),
        sender_key=sender_key,
        sender_nick=sender_nick,
        sender_avatar=_social_current_avatar_by_key(db, sender_key) or None,
        text=str(row.text or ""),
        created_at=_to_utc_iso(row.created_at),
        reply_to=reply_payload,
        attachments=_social_parse_attachments(getattr(row, "attachments_json", "") or "[]"),
        reactions=reactions,
        is_mine=is_mine,
        delivery_status=delivery_status,
        delivery_read_by=int(delivery_read_by),
        delivery_total=int(delivery_total),
    )


def _social_company_allowed_actor_keys(db: Session, user_id: int) -> set[str]:
    allowed = {f"u:{int(user_id)}"}
    rows = db.scalars(
        select(TeamMember).where(
            TeamMember.user_id == int(user_id),
            TeamMember.is_active.is_(True),
        )
    ).all()
    for row in rows:
        if bool(row.is_owner):
            allowed.add(f"u:{int(user_id)}")
        else:
            allowed.add(f"m:{int(row.id)}")
    return allowed


def _social_clean_group_member_keys(
    db: Session,
    *,
    user_id: int,
    actor_key: str,
    member_keys: list[str] | None,
) -> list[str]:
    allowed = _social_company_allowed_actor_keys(db, user_id)
    source = member_keys or []
    seen: set[str] = set()
    out: list[str] = []
    for raw in source:
        key = str(raw or "").strip().lower()
        if not key or key in seen or key not in allowed:
            continue
        seen.add(key)
        out.append(key)
    me = str(actor_key or "").strip().lower()
    if me and me not in seen:
        out.insert(0, me)
        seen.add(me)
    owner_key = f"u:{int(user_id)}"
    if owner_key not in seen:
        out.insert(0, owner_key)
    return out


def _social_chat_storage_dir() -> Path:
    static_root = Path(__file__).resolve().parent.parent / "static"
    target_dir = static_root / "uploads" / "social_chat"
    target_dir.mkdir(parents=True, exist_ok=True)
    return target_dir


def _social_chat_clean_filename(raw: str) -> str:
    base = os.path.basename(str(raw or "").strip()) or "file"
    safe = re.sub(r"[^\w.\-]+", "_", base, flags=re.UNICODE).strip("._")
    return (safe or "file")[:180]


def _social_chat_guess_ext(filename: str, content_type: str) -> str:
    ext = str(Path(filename or "").suffix or "").strip().lower()
    if ext and len(ext) <= 10:
        return ext
    guessed = mimetypes.guess_extension(str(content_type or "").split(";", 1)[0].strip().lower()) or ""
    if guessed == ".jpe":
        guessed = ".jpg"
    if guessed and len(guessed) <= 10:
        return guessed
    return ".bin"


def _social_thread_to_out(db: Session, actor_key: str, row: SocialChatThread, member_row: SocialChatThreadMember) -> SocialChatThreadOut:
    last = _social_thread_last_message(db, row.id)
    last_payload: dict[str, Any] = {}
    if last:
        last_avatar = _social_current_avatar_by_key(db, str(last.sender_key or ""))
        last_payload = {
            "id": int(last.id),
            "sender_key": str(last.sender_key or ""),
            "sender_nick": str(last.sender_nick or ""),
            "sender_avatar": str(last_avatar or ""),
            "text": str(last.text or ""),
            "created_at": _to_utc_iso(last.created_at),
        }
    unread = db.scalar(
        select(func.count())
        .select_from(SocialChatMessage)
        .where(
            SocialChatMessage.thread_id == row.id,
            SocialChatMessage.id > int(member_row.last_read_message_id or 0),
            SocialChatMessage.sender_key != actor_key,
        )
    ) or 0
    participants_rows = db.scalars(
        select(SocialChatThreadMember)
        .where(SocialChatThreadMember.thread_id == row.id)
        .order_by(SocialChatThreadMember.id.asc())
    ).all()
    participant_keys = [str(x.actor_key or "").strip().lower() for x in participants_rows if str(x.actor_key or "").strip()]
    last_activity = _social_last_activity_map(db, participant_keys)
    participants = []
    for x in participants_rows:
        key = str(x.actor_key or "")
        current_nick = _social_current_nick_by_key(db, key) or str(x.actor_nick or "")
        avatar_url = _social_current_avatar_by_key(db, key)
        if current_nick and current_nick != str(x.actor_nick or ""):
            x.actor_nick = current_nick[:120]
        last_seen = last_activity.get(key.strip().lower())
        is_online = _social_is_online(last_seen)
        participants.append(
            {
                "actor_key": key,
                "nick": current_nick,
                "avatar_url": str(avatar_url or ""),
                "last_seen_at": _to_utc_iso(last_seen),
                "is_online": bool(is_online),
                "is_me": key == actor_key,
            }
        )
    title = str(row.title or "")
    if row.kind == "direct":
        other = next((p for p in participants if not p["is_me"] and p["nick"]), None)
        if other:
            title = other["nick"]
    return SocialChatThreadOut(
        id=row.id,
        kind=row.kind or "",
        title=title or row.title or "",
        avatar_url=str(row.avatar_url or "") or None,
        last_message=last_payload,
        unread=int(unread),
        participants=participants,
    )


_CBR_DAILY_URL = "https://www.cbr.ru/scripts/XML_daily.asp"
_CBR_CODES = {"USD", "EUR", "CNY", "BYN", "TRY", "GBP", "UAH"}
_CBR_CACHE_TTL = 60 * 60
_CBR_CACHE: dict[str, Any] = {"stamp": 0.0, "payload": None}
_NOTIFICATION_SOUND_SETTINGS_KEY = "notification_sound_settings"
_DEFAULT_NOTIFICATION_SOUND_SETTINGS: dict[str, Any] = {
    "desktop_enabled": True,
    "chat_enabled": True,
    "task_enabled": True,
    "calendar_enabled": True,
    "default_sound_url": "",
    "chat_sound_url": "",
    "task_sound_url": "",
    "calendar_sound_url": "",
}


def _parse_cbr_number(raw: str) -> float:
    safe = str(raw or "").strip().replace(",", ".")
    try:
        return float(safe)
    except Exception:
        return 0.0


def _fetch_cbr_rates() -> dict[str, Any]:
    with urllib.request.urlopen(_CBR_DAILY_URL, timeout=8) as resp:
        data = resp.read()
    root = ET.fromstring(data)
    date_raw = str(root.attrib.get("Date", "")).strip()
    date_iso = ""
    if date_raw:
        try:
            date_iso = datetime.strptime(date_raw, "%d.%m.%Y").date().isoformat()
        except Exception:
            date_iso = date_raw
    rates: dict[str, float] = {"RUB": 1.0}
    for node in root.findall("Valute"):
        code = str(node.findtext("CharCode", "") or "").strip().upper()
        if not code or code not in _CBR_CODES:
            continue
        nominal = _parse_cbr_number(node.findtext("Nominal", "1"))
        value = _parse_cbr_number(node.findtext("Value", "0"))
        if nominal <= 0 or value <= 0:
            continue
        rates[code] = value / nominal
    return {
        "base": "RUB",
        "date": date_iso or date_raw,
        "updated_at": datetime.utcnow().isoformat(),
        "source": "cbr",
        "stale": False,
        "rates": rates,
    }


def _get_cbr_rates() -> dict[str, Any]:
    now = time.time()
    cached = _CBR_CACHE.get("payload")
    stamp = float(_CBR_CACHE.get("stamp") or 0.0)
    if cached and (now - stamp) < _CBR_CACHE_TTL:
        return cached
    try:
        payload = _fetch_cbr_rates()
        _CBR_CACHE["payload"] = payload
        _CBR_CACHE["stamp"] = now
        return payload
    except Exception:
        if cached:
            fallback = dict(cached)
            fallback["stale"] = True
            return fallback
        raise


def _sanitize_notification_sound_settings(raw: dict[str, Any] | None) -> dict[str, Any]:
    source = raw or {}
    return {
        "desktop_enabled": bool(source.get("desktop_enabled", _DEFAULT_NOTIFICATION_SOUND_SETTINGS["desktop_enabled"])),
        "chat_enabled": bool(source.get("chat_enabled", _DEFAULT_NOTIFICATION_SOUND_SETTINGS["chat_enabled"])),
        "task_enabled": bool(source.get("task_enabled", _DEFAULT_NOTIFICATION_SOUND_SETTINGS["task_enabled"])),
        "calendar_enabled": bool(source.get("calendar_enabled", _DEFAULT_NOTIFICATION_SOUND_SETTINGS["calendar_enabled"])),
        "default_sound_url": str(source.get("default_sound_url") or "").strip()[:500],
        "chat_sound_url": str(source.get("chat_sound_url") or "").strip()[:500],
        "task_sound_url": str(source.get("task_sound_url") or "").strip()[:500],
        "calendar_sound_url": str(source.get("calendar_sound_url") or "").strip()[:500],
    }


def _get_notification_sound_settings(db: Session) -> NotificationSoundSettingsOut:
    raw = _get_system_setting(db, _NOTIFICATION_SOUND_SETTINGS_KEY)
    if not raw:
        safe_default = _sanitize_notification_sound_settings(_DEFAULT_NOTIFICATION_SOUND_SETTINGS)
        _set_system_setting(db, _NOTIFICATION_SOUND_SETTINGS_KEY, json.dumps(safe_default, ensure_ascii=False))
        db.commit()
        return NotificationSoundSettingsOut(**safe_default)
    try:
        parsed = json.loads(raw)
    except Exception:
        parsed = {}
    safe = _sanitize_notification_sound_settings(parsed if isinstance(parsed, dict) else {})
    if json.dumps(safe, ensure_ascii=False) != json.dumps(parsed if isinstance(parsed, dict) else {}, ensure_ascii=False):
        _set_system_setting(db, _NOTIFICATION_SOUND_SETTINGS_KEY, json.dumps(safe, ensure_ascii=False))
        db.commit()
    return NotificationSoundSettingsOut(**safe)


def _social_emit_due_reminders(db: Session, *, user_id: int, actor_key: str, actor_nick: str) -> None:
    safe_user_id = int(user_id or 0)
    safe_actor_key = str(actor_key or "").strip().lower()
    if safe_user_id <= 0 or not safe_actor_key:
        return
    now_ts = time.monotonic()
    throttle_key = f"{safe_user_id}:{safe_actor_key}"
    last_run = float(_SOCIAL_REMINDER_LAST_RUN.get(throttle_key) or 0.0)
    if (now_ts - last_run) < float(_SOCIAL_REMINDER_THROTTLE_SEC):
        return
    _SOCIAL_REMINDER_LAST_RUN[throttle_key] = now_ts
    if len(_SOCIAL_REMINDER_LAST_RUN) > 2500:
        for key, stamp in list(_SOCIAL_REMINDER_LAST_RUN.items()):
            if (now_ts - float(stamp or 0.0)) > (4 * _SOCIAL_REMINDER_THROTTLE_SEC):
                _SOCIAL_REMINDER_LAST_RUN.pop(key, None)

    now_local = datetime.now(_social_calendar_tzinfo()).replace(tzinfo=None, microsecond=0)

    # Task reminders near deadline (30m window) in calendar timezone.
    task_rows = db.scalars(
        select(SocialTask).where(
            SocialTask.user_id == safe_user_id,
            SocialTask.assignee_key == safe_actor_key,
            SocialTask.status != "done",
            SocialTask.due_date.is_not(None),
            SocialTask.due_date <= (now_local + timedelta(minutes=30)),
            SocialTask.due_date >= (now_local - timedelta(minutes=30)),
        )
    ).all()
    for row in task_rows:
        due_local = _social_localize_dt(row.due_date) or row.due_date
        if not isinstance(due_local, datetime):
            continue
        due_slot = due_local.strftime("%Y%m%d%H%M")
        due_text = due_local.strftime("%d.%m.%Y %H:%M")
        _social_push_notification(
            db,
            user_id=safe_user_id,
            recipient_key=safe_actor_key,
            kind="task_reminder",
            dedupe_key=f"task_due:{int(row.id)}:{due_slot}:{safe_actor_key}",
            title="Task reminder",
            body=f"{str(row.title or '')[:140]} - deadline {due_text}",
            payload={
                "task_id": int(row.id),
                "kind": "task",
                "assignee": actor_nick,
                "i18n_key": "task_reminder_3h",
                "i18n_params": {
                    "task_title": str(row.title or "")[:180],
                    "assignee_nick": str(actor_nick or ""),
                    "due_text": due_text,
                },
            },
        )

    # Calendar reminders: default slots -7d/-3d/-1d/-3h/-1h/0m with per-actor overrides.
    event_rows = db.scalars(
        select(SocialCalendarEvent).where(
            SocialCalendarEvent.user_id == safe_user_id,
            SocialCalendarEvent.start_at >= (now_local - timedelta(minutes=30)),
            SocialCalendarEvent.start_at <= (now_local + timedelta(days=8)),
            or_(SocialCalendarEvent.is_public.is_(True), SocialCalendarEvent.actor_key == safe_actor_key),
        )
    ).all()
    for row in event_rows:
        event_start = _social_localize_dt(row.start_at) or row.start_at
        if not isinstance(event_start, datetime):
            continue
        if not bool(row.is_public) and str(row.actor_key or "").strip().lower() != safe_actor_key:
            continue
        reminder_settings = _social_calendar_reminder_settings_get(
            db,
            user_id=safe_user_id,
            event_id=int(row.id or 0),
            actor_key=safe_actor_key,
        )
        if not bool(reminder_settings.get("enabled", True)):
            continue
        offsets = reminder_settings.get("offsets_min") if isinstance(reminder_settings.get("offsets_min"), list) else []
        safe_offsets = _social_calendar_reminder_normalize_offsets(offsets)
        title = str(row.title or "Event")[:180]
        start_text = event_start.strftime("%d.%m.%Y %H:%M")
        for offset_min in safe_offsets:
            reminder_at = event_start + timedelta(minutes=int(offset_min or 0))
            if now_local < reminder_at or now_local > (reminder_at + timedelta(minutes=30)):
                continue
            slot_key = reminder_at.strftime("%Y%m%d%H%M")
            _social_push_notification(
                db,
                user_id=safe_user_id,
                recipient_key=safe_actor_key,
                kind="calendar_reminder",
                dedupe_key=f"calendar_due:{int(row.id)}:{int(offset_min)}:{slot_key}:{safe_actor_key}",
                title="Calendar reminder",
                body=f"{title} - {start_text}",
                payload={
                    "event_id": int(row.id),
                    "kind": "calendar",
                    "i18n_key": "calendar_event_reminder",
                    "i18n_params": {
                        "event_title": title,
                        "start_text": start_text,
                        "offset_min": int(offset_min or 0),
                    },
                },
            )

    _social_emit_due_announcements(db, user_id=safe_user_id, actor_key=safe_actor_key)

def _social_announcement_to_out(row: SocialAnnouncement) -> SocialAnnouncementOut:
    target_user_ids = _social_announcement_target_user_ids(row)
    target_user_id = int(row.user_id) if row.user_id is not None else (target_user_ids[0] if len(target_user_ids) == 1 else None)
    return SocialAnnouncementOut(
        id=int(row.id),
        title=str(row.title or ""),
        body=str(row.body or ""),
        starts_at=_to_utc_iso(row.starts_at),
        ends_at=_to_utc_iso(row.ends_at),
        is_active=bool(row.is_active),
        user_id=target_user_id,
        user_ids=target_user_ids,
        created_by_user_id=int(row.created_by_user_id) if row.created_by_user_id is not None else None,
        created_at=_to_utc_iso(row.created_at),
        updated_at=_to_utc_iso(row.updated_at),
    )


def _social_announcement_to_public_out(row: SocialAnnouncement) -> SocialAnnouncementPublicOut:
    return SocialAnnouncementPublicOut(
        id=int(row.id),
        title=str(row.title or ""),
        body=str(row.body or ""),
        starts_at=_to_utc_iso(row.starts_at),
        ends_at=_to_utc_iso(row.ends_at),
    )


def _social_announcement_target_user_ids(row: SocialAnnouncement) -> list[int]:
    result: list[int] = []
    seen: set[int] = set()
    raw = str(getattr(row, "target_user_ids_json", "") or "").strip()
    parsed: Any = []
    if raw:
        try:
            parsed = json.loads(raw)
        except Exception:
            parsed = []
    source = parsed if isinstance(parsed, list) else []
    for value in source:
        uid = int(value or 0)
        if uid <= 0 or uid in seen:
            continue
        seen.add(uid)
        result.append(uid)
    legacy_uid = int(getattr(row, "user_id", 0) or 0)
    if legacy_uid > 0 and legacy_uid not in seen:
        result.append(legacy_uid)
    return result


def _social_announcement_is_for_user(row: SocialAnnouncement, user_id: int) -> bool:
    targets = _social_announcement_target_user_ids(row)
    if not targets:
        return True
    return int(user_id or 0) in set(targets)


def _social_resolve_announcement_targets(payload: SocialAnnouncementIn) -> tuple[int | None, list[int]]:
    resolved: list[int] = []
    seen: set[int] = set()
    single_user_id = int(payload.user_id or 0)
    if single_user_id > 0:
        resolved.append(single_user_id)
        seen.add(single_user_id)
    for raw in payload.user_ids or []:
        uid = int(raw or 0)
        if uid <= 0 or uid in seen:
            continue
        seen.add(uid)
        resolved.append(uid)
        if len(resolved) >= 1000:
            break
    if len(resolved) == 1:
        return resolved[0], resolved
    return None, resolved


def _social_emit_due_announcements(db: Session, *, user_id: int, actor_key: str) -> None:
    safe_user_id = int(user_id or 0)
    safe_actor_key = str(actor_key or "").strip()
    if safe_user_id <= 0 or not safe_actor_key:
        return
    now = datetime.utcnow()
    rows = db.scalars(
        select(SocialAnnouncement).where(
            SocialAnnouncement.is_active.is_(True),
            SocialAnnouncement.starts_at <= now,
            or_(SocialAnnouncement.ends_at.is_(None), SocialAnnouncement.ends_at >= now),
        ).order_by(SocialAnnouncement.starts_at.asc(), SocialAnnouncement.id.asc())
    ).all()
    if not rows:
        return
    acked_ids = set(
        db.scalars(
            select(SocialAnnouncementAck.announcement_id).where(
                SocialAnnouncementAck.user_id == safe_user_id,
                SocialAnnouncementAck.actor_key == safe_actor_key,
            )
        ).all()
    )
    for row in rows:
        ann_id = int(row.id or 0)
        if ann_id <= 0 or ann_id in acked_ids:
            continue
        if not _social_announcement_is_for_user(row, safe_user_id):
            continue
        _social_push_notification(
            db,
            user_id=safe_user_id,
            recipient_key=safe_actor_key,
            kind="announcement",
            dedupe_key=f"announcement:{ann_id}:{safe_actor_key}",
            title=str(row.title or "РћР±СЉСЏРІР»РµРЅРёРµ")[:255],
            body=str(row.body or "")[:5000],
            payload={"announcement_id": ann_id, "kind": "announcement"},
        )


def _social_push_notification(
    db: Session,
    *,
    user_id: int,
    recipient_key: str,
    kind: str,
    dedupe_key: str,
    title: str,
    body: str,
    payload: dict[str, Any] | None = None,
) -> None:
    key = str(dedupe_key or "").strip()[:120]
    if not key:
        return
    safe_title = (_decode_mojibake_text(title or "") or str(title or "")).strip()
    safe_body = _decode_mojibake_text(body or "")
    existing = db.scalar(
        select(SocialNotification).where(
            SocialNotification.user_id == int(user_id),
            SocialNotification.recipient_key == recipient_key,
            SocialNotification.kind == kind,
            SocialNotification.dedupe_key == key,
        )
    )
    if existing:
        return
    db.add(
        SocialNotification(
            user_id=user_id,
            recipient_key=recipient_key[:60],
            kind=kind[:40],
            dedupe_key=key,
            title=safe_title[:255],
            body=safe_body[:5000],
            payload_json=json.dumps(payload or {}, ensure_ascii=False),
            is_read=False,
        )
    )



def _social_emit_game_turn_notification(
    db: Session,
    *,
    room: SocialCheckersRoom,
    game_code: str,
    mover_actor_key: str,
    mover_nick: str,
    move_marker: str,
    move_at_iso: str,
) -> None:
    if str(room.mode or "").strip().lower() != "human":
        return
    if str(room.status or "").strip().lower() != "active":
        return
    mover_key = str(mover_actor_key or "").strip().lower()
    host_key = str(room.host_actor_key or "").strip().lower()
    guest_key = str(room.guest_actor_key or "").strip().lower()
    recipient_key = ""
    recipient_user_id = 0
    if mover_key and mover_key == host_key:
        recipient_key = guest_key
        recipient_user_id = int(room.guest_user_id or 0)
    elif mover_key and mover_key == guest_key:
        recipient_key = host_key
        recipient_user_id = int(room.host_user_id or 0)
    if recipient_user_id <= 0 or not recipient_key:
        return
    marker = str(move_marker or "").strip()[:40] or str(int(datetime.utcnow().timestamp()))
    _social_push_notification(
        db,
        user_id=int(recipient_user_id),
        recipient_key=str(recipient_key),
        kind="game_turn",
        dedupe_key=f"game_turn:{str(game_code or '').strip().lower()}:{int(room.id or 0)}:{marker}:{recipient_key}"[:120],
        title="Opponent moved",
        body=f"{str(mover_nick or 'Opponent')[:120]} made a move.",
        payload={
            "game_code": str(game_code or "").strip().lower(),
            "room_id": int(room.id or 0),
            "opponent_nick": str(mover_nick or "")[:120],
            "move_at": str(move_at_iso or "")[:40],
            "i18n_key": "game_turn",
            "i18n_params": {
                "opponent_nick": str(mover_nick or "")[:120],
                "game_code": str(game_code or "").strip().lower(),
            },
        },
    )

def _social_parse_dt(raw: str | None) -> datetime | None:
    text_raw = str(raw or "").strip()
    if not text_raw:
        return None
    try:
        normalized = text_raw.replace("Z", "+00:00")
        return _social_localize_dt(datetime.fromisoformat(normalized))
    except Exception:
        return None


def _social_ical_unfold_lines(raw_text: str) -> list[str]:
    text_value = str(raw_text or "").replace("\r\n", "\n").replace("\r", "\n")
    lines = text_value.split("\n")
    out: list[str] = []
    for line in lines:
        if (line.startswith(" ") or line.startswith("\t")) and out:
            out[-1] += line[1:]
            continue
        out.append(line)
    return out


def _social_ical_unescape(value: str) -> str:
    text_value = str(value or "")
    text_value = text_value.replace("\\n", "\n").replace("\\N", "\n")
    text_value = text_value.replace("\\,", ",").replace("\\;", ";").replace("\\\\", "\\")
    return text_value.strip()


def _social_ical_parse_dt(raw: str, *, tzid: str = "", is_date_only: bool = False) -> datetime | None:
    text_value = str(raw or "").strip()
    if not text_value:
        return None
    if re.fullmatch(r"\d{8}", text_value):
        try:
            dt_local = datetime.strptime(text_value, "%Y%m%d")
            return dt_local
        except Exception:
            return None
    fmt_candidates = ["%Y%m%dT%H%M%SZ", "%Y%m%dT%H%M%S", "%Y%m%dT%H%M"]
    parsed: datetime | None = None
    used_utc_suffix = False
    for fmt in fmt_candidates:
        try:
            parsed = datetime.strptime(text_value, fmt)
            used_utc_suffix = fmt.endswith("Z")
            break
        except Exception:
            continue
    if parsed is None:
        return None
    if used_utc_suffix:
        return _social_localize_dt(parsed.replace(tzinfo=timezone.utc))
    if tzid:
        try:
            tzinfo = ZoneInfo(str(tzid))
            return _social_localize_dt(parsed.replace(tzinfo=tzinfo))
        except Exception:
            return parsed.replace(microsecond=0)
    if is_date_only:
        return parsed.replace(hour=0, minute=0, second=0, microsecond=0)
    return parsed.replace(microsecond=0)


def _social_parse_ical_events(raw_text: str) -> tuple[list[dict[str, Any]], list[str]]:
    lines = _social_ical_unfold_lines(raw_text)
    events: list[dict[str, Any]] = []
    warnings: list[str] = []
    current: dict[str, Any] | None = None
    for line in lines:
        text_line = str(line or "").strip()
        if not text_line:
            continue
        upper = text_line.upper()
        if upper == "BEGIN:VEVENT":
            current = {}
            continue
        if upper == "END:VEVENT":
            if not isinstance(current, dict):
                current = None
                continue
            status = str(current.get("STATUS") or "").strip().upper()
            if status == "CANCELLED":
                current = None
                continue
            uid = str(current.get("UID") or "").strip()
            title = str(current.get("SUMMARY") or "").strip() or "Google event"
            details = _social_ical_unescape(str(current.get("DESCRIPTION") or ""))
            dt_start_raw, dt_start_tz, dt_start_is_date = current.get("_DTSTART", ("", "", False))
            dt_end_raw, dt_end_tz, dt_end_is_date = current.get("_DTEND", ("", "", False))
            start_at = _social_ical_parse_dt(dt_start_raw, tzid=dt_start_tz, is_date_only=bool(dt_start_is_date))
            end_at = _social_ical_parse_dt(dt_end_raw, tzid=dt_end_tz, is_date_only=bool(dt_end_is_date))
            if not start_at:
                current = None
                continue
            if end_at and end_at <= start_at:
                end_at = start_at + timedelta(hours=1)
            events.append(
                {
                    "uid": uid or hashlib.sha1(f"{title}|{start_at.isoformat()}|{details}".encode("utf-8")).hexdigest()[:24],
                    "title": title[:255],
                    "details": details[:5000],
                    "start_at": start_at,
                    "end_at": end_at,
                }
            )
            current = None
            continue
        if current is None:
            continue
        if ":" not in text_line:
            continue
        key_part, value = text_line.split(":", 1)
        key_tokens = [str(x or "").strip() for x in key_part.split(";") if str(x or "").strip()]
        if not key_tokens:
            continue
        key = key_tokens[0].upper()
        params: dict[str, str] = {}
        for token in key_tokens[1:]:
            if "=" not in token:
                continue
            pkey, pval = token.split("=", 1)
            params[str(pkey or "").strip().upper()] = str(pval or "").strip()
        if key in {"UID", "SUMMARY", "DESCRIPTION", "STATUS"}:
            current[key] = value
            continue
        if key in {"DTSTART", "DTEND"}:
            tzid = str(params.get("TZID") or "").strip()
            is_date = str(params.get("VALUE") or "").strip().upper() == "DATE"
            current[f"_{key}"] = (value, tzid, is_date)
            continue
    if not events:
        warnings.append("Google Calendar РЅРµ РІРµСЂРЅСѓР» СЃРѕР±С‹С‚РёР№ РІ ICS.")
    return events, warnings


@router.get("/social/bootstrap", response_model=dict[str, Any])
def social_bootstrap(
    request: Request,
    user: User = Depends(get_current_user),
    db: Session = Depends(get_db),
):
    ensure_module_enabled(db, user, "social_hub")
    actor_key, actor_nick, _ = _social_actor_identity(db, user)
    company_thread = _social_ensure_company_thread(db, user.id)
    _social_sync_company_thread_members(
        db,
        thread=company_thread,
        user_id=int(user.id),
        actor_key=actor_key,
        actor_nick=actor_nick,
    )
    company_actors = db.scalars(
        select(TeamMember).where(
            TeamMember.user_id == user.id,
            TeamMember.is_active.is_(True),
        ).order_by(TeamMember.is_owner.desc(), TeamMember.id.asc())
    ).all()
    db.commit()
    _audit(
        db,
        user,
        action="social_bootstrap_loaded",
        details=f"actor={actor_key}",
        module_code="social_hub",
        entity_type="social",
        entity_id=actor_key,
        request=request,
    )
    db.commit()
    return {
        "actor": {
            "actor_key": actor_key,
            "nick": actor_nick,
            "is_owner": bool(_actor_is_owner(user)),
        },
        "games": [{"code": code, "title": title} for code, title in SOCIAL_GAMES.items()],
        "company_actors": [
            {
                "actor_key": f"u:{user.id}" if bool(row.is_owner) else f"m:{row.id}",
                "nick": str((row.nickname or row.full_name or row.email).strip() or f"member-{row.id}"),
                "is_owner": bool(row.is_owner),
            }
            for row in company_actors
        ],
    }


@router.post("/social/games/score", response_model=SocialGameScoreOut)
def social_save_game_score(
    payload: SocialGameScoreIn,
    request: Request,
    user: User = Depends(get_current_user),
    db: Session = Depends(get_db),
):
    ensure_module_enabled(db, user, "social_hub")
    game_code = str(payload.game_code or "").strip().lower()
    if game_code not in SOCIAL_SCORE_GAMES:
        raise HTTPException(status_code=400, detail="РќРµРёР·РІРµСЃС‚РЅР°СЏ РёРіСЂР°")
    actor_key, actor_nick, _ = _social_actor_identity(db, user)
    score = max(0, int(payload.score or 0))
    row = db.scalar(
        select(SocialGameScore).where(
            SocialGameScore.game_code == game_code,
            SocialGameScore.actor_key == actor_key,
        )
    )
    if not row:
        row = SocialGameScore(
            user_id=user.id,
            game_code=game_code,
            actor_key=actor_key,
            actor_nick=actor_nick,
            best_score=score,
            last_score=score,
            play_count=1,
        )
        db.add(row)
        db.flush()
    else:
        row.actor_nick = actor_nick[:120]
        row.last_score = score
        row.play_count = int(row.play_count or 0) + 1
        row.best_score = max(int(row.best_score or 0), score)
    _audit(
        db,
        user,
        action="social_game_score_saved",
        details=f"game={game_code};score={score};best={int(row.best_score or 0)}",
        module_code="social_hub",
        entity_type="game",
        entity_id=f"{game_code}:{actor_key}",
        request=request,
    )
    db.commit()
    return SocialGameScoreOut(
        game_code=game_code,
        actor_key=actor_key,
        actor_nick=str(row.actor_nick or ""),
        best_score=int(row.best_score or 0),
        last_score=int(row.last_score or 0),
        play_count=int(row.play_count or 0),
        updated_at=row.updated_at.isoformat() if row.updated_at else None,
    )


@router.get("/social/games/leaderboard", response_model=SocialLeaderboardOut)
def social_leaderboard(
    game_code: str,
    limit: int = 50,
    user: User = Depends(get_current_user),
    db: Session = Depends(get_db),
):
    ensure_module_enabled(db, user, "social_hub")
    safe_game = str(game_code or "").strip().lower()
    if safe_game not in SOCIAL_SCORE_GAMES:
        raise HTTPException(status_code=400, detail="РќРµРёР·РІРµСЃС‚РЅР°СЏ РёРіСЂР°")
    safe_limit = max(5, min(int(limit or 50), 200))
    rows = db.scalars(
        select(SocialGameScore)
        .where(SocialGameScore.game_code == safe_game)
        .order_by(SocialGameScore.best_score.desc(), SocialGameScore.updated_at.asc(), SocialGameScore.id.asc())
        .limit(safe_limit)
    ).all()
    actor_key, _, _ = _social_actor_identity(db, user)
    my_row = db.scalar(
        select(SocialGameScore).where(
            SocialGameScore.game_code == safe_game,
            SocialGameScore.actor_key == actor_key,
        )
    )
    my_best = int(my_row.best_score or 0) if my_row else 0
    my_rank = None
    if my_row:
        higher = db.scalar(
            select(func.count())
            .select_from(SocialGameScore)
            .where(
                SocialGameScore.game_code == safe_game,
                SocialGameScore.best_score > my_best,
            )
        ) or 0
        my_rank = int(higher) + 1
    top = []
    for idx, row in enumerate(rows, start=1):
        top.append(
            {
                "rank": idx,
                "nick": str(row.actor_nick or ""),
                "score": int(row.best_score or 0),
                "is_me": str(row.actor_key or "") == actor_key,
            }
        )
    return SocialLeaderboardOut(
        game_code=safe_game,
        top=top,
        my_rank=my_rank,
        my_best=my_best,
    )



def _social_checkers_profile_row(
    db: Session,
    *,
    actor_key: str,
    actor_nick: str = "",
    user_id: int | None = None,
    member_id: int | None = None,
) -> SocialCheckersProfile:
    canonical = _social_canonical_actor_key(db, actor_key)
    row = db.scalar(select(SocialCheckersProfile).where(SocialCheckersProfile.actor_key == canonical))
    safe_user_id = int(user_id or 0)
    safe_member_id = int(member_id or 0) or None
    safe_nick = str(actor_nick or "").strip()
    if safe_user_id <= 0 or not safe_nick:
        try:
            resolved_user_id, resolved_member_id, resolved_nick = _social_identity_by_key(db, canonical)
            if safe_user_id <= 0:
                safe_user_id = int(resolved_user_id or 0)
            if not safe_member_id and resolved_member_id:
                safe_member_id = int(resolved_member_id or 0)
            if not safe_nick:
                safe_nick = str(resolved_nick or "").strip()
        except HTTPException:
            pass
    safe_nick = safe_nick or _social_current_nick_by_key(db, canonical) or canonical
    if safe_user_id <= 0:
        raise HTTPException(status_code=400, detail="РќРµРєРѕСЂСЂРµРєС‚РЅС‹Р№ РёРіСЂРѕРє РґР»СЏ СЂРµР№С‚РёРЅРіР°")
    if row:
        row.user_id = safe_user_id
        row.member_id = safe_member_id
        row.actor_nick = safe_nick[:120]
        row.updated_at = datetime.utcnow()
        return row
    row = SocialCheckersProfile(
        user_id=safe_user_id,
        member_id=safe_member_id,
        actor_key=canonical[:60],
        actor_nick=safe_nick[:120],
        rating=1200,
        wins=0,
        losses=0,
        draws=0,
        play_count=0,
    )
    db.add(row)
    db.flush()
    return row


def _social_checkers_my_profile(db: Session, user: User) -> tuple[str, str, int | None, SocialCheckersProfile]:
    actor_key, actor_nick, member_id = _social_actor_identity(db, user)
    row = _social_checkers_profile_row(
        db,
        actor_key=actor_key,
        actor_nick=actor_nick,
        user_id=int(user.id),
        member_id=member_id,
    )
    return actor_key, actor_nick, member_id, row


def _social_checkers_profile_out(row: SocialCheckersProfile, *, actor_key: str = "") -> dict[str, Any]:
    return {
        "actor_key": actor_key or str(row.actor_key or ""),
        "nick": str(row.actor_nick or "") or actor_key,
        "rating": int(row.rating or 1200),
        "wins": int(row.wins or 0),
        "losses": int(row.losses or 0),
        "draws": int(row.draws or 0),
        "play_count": int(row.play_count or 0),
        "updated_at": _to_utc_iso(row.updated_at),
    }


def _social_checkers_public_player(db: Session, actor_key: str, fallback_nick: str = "", *, difficulty: str = "") -> dict[str, Any]:
    key = str(actor_key or "").strip().lower()
    if key.startswith("bot:"):
        bot = build_checkers_bot_identity(difficulty or key.split(":", 1)[1])
        return {
            "actor_key": str(bot["actor_key"]),
            "nick": str(bot["nick"]),
            "rating": int(bot["rating"]),
            "wins": 0,
            "losses": 0,
            "draws": 0,
            "play_count": 0,
            "is_bot": True,
            "avatar_url": "",
        }
    canonical = _social_canonical_actor_key(db, key)
    profile_row = db.scalar(select(SocialCheckersProfile).where(SocialCheckersProfile.actor_key == canonical))
    public_profile = _social_public_profile_by_key(db, canonical)
    nick = str(public_profile.get("nick") or fallback_nick or canonical).strip()
    return {
        "actor_key": canonical,
        "nick": nick[:120],
        "rating": int(profile_row.rating or 1200) if profile_row else 1200,
        "wins": int(profile_row.wins or 0) if profile_row else 0,
        "losses": int(profile_row.losses or 0) if profile_row else 0,
        "draws": int(profile_row.draws or 0) if profile_row else 0,
        "play_count": int(profile_row.play_count or 0) if profile_row else 0,
        "is_bot": False,
        "avatar_url": str(public_profile.get("avatar_url") or ""),
    }


def _social_checkers_room_side(db: Session, room: SocialCheckersRoom, actor_key: str) -> str:
    aliases = set(_social_actor_alias_keys(db, actor_key))
    host_key = _social_canonical_actor_key(db, str(room.host_actor_key or ""))
    guest_key = _social_canonical_actor_key(db, str(room.guest_actor_key or ""))
    if host_key and host_key in aliases:
        return "white"
    if guest_key and guest_key in aliases:
        return "black"
    return ""


def _social_checkers_room_visible(db: Session, room: SocialCheckersRoom, actor_key: str) -> bool:
    if bool(room.is_public) and str(room.status or "") == "waiting":
        return True
    return bool(_social_checkers_room_side(db, room, actor_key))


def _social_checkers_room_state(room: SocialCheckersRoom) -> dict[str, Any]:
    try:
        raw = json.loads(str(room.state_json or "{}"))
    except Exception:
        raw = {}
    return load_checkers_state(raw)


def _social_checkers_store_room_state(room: SocialCheckersRoom, state: dict[str, Any]) -> None:
    room.state_json = json.dumps(load_checkers_state(state), ensure_ascii=False)


def _social_checkers_note(room: SocialCheckersRoom, payload: dict[str, Any]) -> str:
    status = str(payload.get("status") or room.status or "")
    if status == "waiting":
        return "РљРѕРјРЅР°С‚Р° РІРёРґРЅР° РІ РѕР±С‰РµРј Р»РѕР±Р±Рё Рё Р¶РґРµС‚ РІС‚РѕСЂРѕРіРѕ РёРіСЂРѕРєР°."
    if status == "active" and str(room.mode or "") == "bot":
        return f"РЈСЂРѕРІРµРЅСЊ РР: {str(room.difficulty or 'medium')}"
    if status == "active":
        return "РРіСЂР° РёРґРµС‚ РІ СЂРµР°Р»СЊРЅРѕРј РІСЂРµРјРµРЅРё. РџРѕР·РёС†РёСЏ РѕР±РЅРѕРІР»СЏРµС‚СЃСЏ Р°РІС‚РѕРјР°С‚РёС‡РµСЃРєРё."
    if status == "finished":
        result = str(payload.get("result") or "")
        winner = str(payload.get("winner") or "")
        if result == "draw":
            return "РџР°СЂС‚РёСЏ Р·Р°РІРµСЂС€РёР»Р°СЃСЊ РЅРёС‡СЊРµР№."
        if winner == "white":
            return "Р‘РµР»С‹Рµ Р·Р°РІРµСЂС€РёР»Рё РїР°СЂС‚РёСЋ РїРѕР±РµРґРѕР№."
        if winner == "black":
            return "Р§РµСЂРЅС‹Рµ Р·Р°РІРµСЂС€РёР»Рё РїР°СЂС‚РёСЋ РїРѕР±РµРґРѕР№."
    if status == "cancelled":
        return "РљРѕРјРЅР°С‚Р° Р·Р°РєСЂС‹С‚Р° РґРѕ СЃС‚Р°СЂС‚Р° РёР»Рё Р·Р°РІРµСЂС€РµРЅР° РІСЂСѓС‡РЅСѓСЋ."
    return ""


def _social_checkers_room_payload(db: Session, room: SocialCheckersRoom, viewer_key: str) -> dict[str, Any]:
    state = _social_checkers_room_state(room)
    my_side = _social_checkers_room_side(db, room, viewer_key)
    can_join = bool(room.is_public) and str(room.mode or "") == "human" and str(room.status or "") == "waiting" and not my_side and not str(room.guest_actor_key or "")
    can_move = bool(my_side) and str(room.status or "") == "active" and str(state.get("turn") or "") == my_side and not str(state.get("winner") or "") and str(state.get("result") or "") != "draw"
    white_player = _social_checkers_public_player(db, str(room.host_actor_key or ""), str(room.host_nick or ""), difficulty=str(room.difficulty or ""))
    if str(room.mode or "") == "bot":
        bot_meta = build_checkers_bot_identity(room.difficulty)
        black_key = str(bot_meta.get("actor_key") or "bot:medium")
        black_nick = str(bot_meta.get("nick") or "SEO WIBE AI")
    else:
        black_key = str(room.guest_actor_key or "")
        black_nick = str(room.guest_nick or "")
    black_player = _social_checkers_public_player(db, black_key, black_nick, difficulty=str(room.difficulty or ""))
    legal_moves = get_checkers_legal_moves(state, my_side) if can_move else []
    payload = {
        "id": int(room.id or 0),
        "room_code": str(room.room_code or ""),
        "title": str(room.title or "").strip() or f"РљРѕРјРЅР°С‚Р° {room.room_code}",
        "mode": str(room.mode or "human"),
        "difficulty": str(room.difficulty or "medium"),
        "status": str(room.status or "waiting"),
        "is_public": bool(room.is_public),
        "turn": str(state.get("turn") or "white"),
        "winner": str(state.get("winner") or ""),
        "result": str(state.get("result") or ""),
        "board": state.get("board") or [],
        "last_move": state.get("last_move") or {},
        "history": state.get("history") or [],
        "players": {
            "white": white_player,
            "black": black_player,
        },
        "my_side": my_side,
        "my_turn": bool(can_move),
        "can_join": bool(can_join),
        "can_move": bool(can_move),
        "legal_moves": legal_moves,
        "created_at": _to_utc_iso(room.created_at),
        "updated_at": _to_utc_iso(room.updated_at),
        "last_move_at": _to_utc_iso(room.last_move_at),
        "finished_at": _to_utc_iso(room.finished_at),
    }
    payload["note"] = _social_checkers_note(room, payload)
    return payload


def _social_checkers_rank_rows(db: Session, viewer_key: str, limit: int = 100) -> tuple[list[dict[str, Any]], int | None, int]:
    safe_limit = max(5, min(int(limit or 100), 200))
    canonical = _social_canonical_actor_key(db, viewer_key)
    rows = db.scalars(
        select(SocialCheckersProfile)
        .order_by(
            SocialCheckersProfile.rating.desc(),
            SocialCheckersProfile.wins.desc(),
            SocialCheckersProfile.play_count.desc(),
            SocialCheckersProfile.updated_at.asc(),
            SocialCheckersProfile.id.asc(),
        )
        .limit(safe_limit)
    ).all()
    ordered_keys = db.scalars(
        select(SocialCheckersProfile.actor_key).order_by(
            SocialCheckersProfile.rating.desc(),
            SocialCheckersProfile.wins.desc(),
            SocialCheckersProfile.play_count.desc(),
            SocialCheckersProfile.updated_at.asc(),
            SocialCheckersProfile.id.asc(),
        )
    ).all()
    my_rank = None
    for idx, key in enumerate(ordered_keys, start=1):
        if str(key or "") == canonical:
            my_rank = idx
            break
    my_row = db.scalar(select(SocialCheckersProfile).where(SocialCheckersProfile.actor_key == canonical))
    my_rating = int(my_row.rating or 1200) if my_row else 1200
    data_rows: list[dict[str, Any]] = []
    for idx, row in enumerate(rows, start=1):
        data_rows.append(
            {
                "rank": idx,
                "actor_key": str(row.actor_key or ""),
                "nick": str(row.actor_nick or row.actor_key or ""),
                "rating": int(row.rating or 1200),
                "wins": int(row.wins or 0),
                "losses": int(row.losses or 0),
                "draws": int(row.draws or 0),
                "play_count": int(row.play_count or 0),
                "is_me": str(row.actor_key or "") == canonical,
            }
        )
    return data_rows, my_rank, my_rating


def _social_checkers_apply_stats(db: Session, room: SocialCheckersRoom, state: dict[str, Any]) -> None:
    if bool(room.stats_applied):
        return
    result = str(state.get("result") or "")
    winner = str(state.get("winner") or "")
    if str(room.status or "") != "finished":
        return
    if str(room.mode or "") == "bot":
        host_profile = _social_checkers_profile_row(
            db,
            actor_key=str(room.host_actor_key or ""),
            actor_nick=str(room.host_nick or ""),
            user_id=int(room.host_user_id or 0),
            member_id=int(room.host_member_id or 0) or None,
        )
        bot_meta = build_checkers_bot_identity(room.difficulty)
        if result == "draw":
            host_profile.draws = int(host_profile.draws or 0) + 1
            score_value = 0.5
        elif winner == "white":
            host_profile.wins = int(host_profile.wins or 0) + 1
            score_value = 1.0
        else:
            host_profile.losses = int(host_profile.losses or 0) + 1
            score_value = 0.0
        host_profile.play_count = int(host_profile.play_count or 0) + 1
        host_profile.rating, _ = apply_checkers_elo(int(host_profile.rating or 1200), int(bot_meta.get("rating") or 1200), score_value)
        room.stats_applied = True
        return
    if not str(room.guest_actor_key or ""):
        return
    white_profile = _social_checkers_profile_row(
        db,
        actor_key=str(room.host_actor_key or ""),
        actor_nick=str(room.host_nick or ""),
        user_id=int(room.host_user_id or 0),
        member_id=int(room.host_member_id or 0) or None,
    )
    black_profile = _social_checkers_profile_row(
        db,
        actor_key=str(room.guest_actor_key or ""),
        actor_nick=str(room.guest_nick or ""),
        user_id=int(room.guest_user_id or 0),
        member_id=int(room.guest_member_id or 0) or None,
    )
    if result == "draw":
        white_score = 0.5
        black_score = 0.5
        white_profile.draws = int(white_profile.draws or 0) + 1
        black_profile.draws = int(black_profile.draws or 0) + 1
    elif winner == "white":
        white_score = 1.0
        black_score = 0.0
        white_profile.wins = int(white_profile.wins or 0) + 1
        black_profile.losses = int(black_profile.losses or 0) + 1
    else:
        white_score = 0.0
        black_score = 1.0
        white_profile.losses = int(white_profile.losses or 0) + 1
        black_profile.wins = int(black_profile.wins or 0) + 1
    white_profile.play_count = int(white_profile.play_count or 0) + 1
    black_profile.play_count = int(black_profile.play_count or 0) + 1
    next_white, _ = apply_checkers_elo(int(white_profile.rating or 1200), int(black_profile.rating or 1200), white_score)
    next_black, _ = apply_checkers_elo(int(black_profile.rating or 1200), int(white_profile.rating or 1200), black_score)
    white_profile.rating = next_white
    black_profile.rating = next_black
    room.stats_applied = True


@router.get("/social/games/checkers/overview", response_model=dict[str, Any])
def social_checkers_overview(
    user: User = Depends(get_current_user),
    db: Session = Depends(get_db),
):
    ensure_module_enabled(db, user, "social_hub")
    actor_key, actor_nick, member_id, profile_row = _social_checkers_my_profile(db, user)
    aliases = _social_actor_alias_keys(db, actor_key)
    public_rooms = db.scalars(
        select(SocialCheckersRoom)
        .where(
            SocialCheckersRoom.game_code == CHECKERS_GAME_CODE,
            SocialCheckersRoom.mode == "human",
            SocialCheckersRoom.is_public.is_(True),
            SocialCheckersRoom.status == "waiting",
        )
        .order_by(SocialCheckersRoom.updated_at.desc(), SocialCheckersRoom.id.desc())
        .limit(24)
    ).all()
    my_rooms = db.scalars(
        select(SocialCheckersRoom)
        .where(
            SocialCheckersRoom.game_code == CHECKERS_GAME_CODE,
            or_(
                SocialCheckersRoom.host_actor_key.in_(aliases),
                SocialCheckersRoom.guest_actor_key.in_(aliases),
            ),
            SocialCheckersRoom.status.in_(["waiting", "active", "finished", "cancelled"]),
        )
        .order_by(SocialCheckersRoom.updated_at.desc(), SocialCheckersRoom.id.desc())
        .limit(12)
    ).all()
    leaderboard_rows, my_rank, my_rating = _social_checkers_rank_rows(db, actor_key, limit=20)
    db.commit()
    return {
        "profile": {
            **_social_checkers_profile_out(profile_row, actor_key=actor_key),
            "nick": actor_nick,
            "member_id": int(member_id or 0),
        },
        "leaderboard": {
            "rows": leaderboard_rows,
            "my_rank": my_rank,
            "my_rating": my_rating,
        },
        "rooms": {
            "public": [_social_checkers_room_payload(db, row, actor_key) for row in public_rooms],
            "mine": [_social_checkers_room_payload(db, row, actor_key) for row in my_rooms],
        },
        "difficulties": get_checkers_difficulties(),
    }


@router.get("/social/games/checkers/leaderboard", response_model=dict[str, Any])
def social_checkers_leaderboard(
    limit: int = 100,
    user: User = Depends(get_current_user),
    db: Session = Depends(get_db),
):
    ensure_module_enabled(db, user, "social_hub")
    actor_key, _, _, profile_row = _social_checkers_my_profile(db, user)
    rows, my_rank, my_rating = _social_checkers_rank_rows(db, actor_key, limit=limit)
    db.commit()
    return {
        "rows": rows,
        "my_rank": my_rank,
        "my_rating": my_rating,
        "my_profile": _social_checkers_profile_out(profile_row, actor_key=actor_key),
    }


@router.post("/social/games/checkers/rooms", response_model=dict[str, Any])
def social_checkers_create_room(
    payload: dict[str, Any],
    request: Request,
    user: User = Depends(get_current_user),
    db: Session = Depends(get_db),
):
    ensure_module_enabled(db, user, "social_hub")
    actor_key, actor_nick, member_id, _ = _social_checkers_my_profile(db, user)
    mode = str((payload or {}).get("mode") or "human").strip().lower()
    if mode not in {"human", "bot"}:
        raise HTTPException(status_code=400, detail="РќРµРєРѕСЂСЂРµРєС‚РЅС‹Р№ СЂРµР¶РёРј РєРѕРјРЅР°С‚С‹")
    difficulty = str((payload or {}).get("difficulty") or "medium").strip().lower()
    bot_meta = build_checkers_bot_identity(difficulty)
    safe_title = str((payload or {}).get("title") or "").strip()[:120]
    title = safe_title or (f"РљРѕРјРЅР°С‚Р° {actor_nick}" if mode == "human" else f"РњР°С‚С‡ СЃ {bot_meta['title']}")
    is_public = bool((payload or {}).get("is_public", True)) if mode == "human" else False
    existing_codes = set(db.scalars(select(SocialCheckersRoom.room_code)).all())
    room = SocialCheckersRoom(
        game_code=CHECKERS_GAME_CODE,
        room_code=create_checkers_room_code(existing_codes),
        title=title,
        owner_user_id=int(user.id),
        host_user_id=int(user.id),
        host_member_id=int(member_id or 0) or None,
        host_actor_key=_social_canonical_actor_key(db, actor_key),
        host_nick=actor_nick[:120],
        guest_user_id=None,
        guest_member_id=None,
        guest_actor_key=str(bot_meta["actor_key"]) if mode == "bot" else "",
        guest_nick=str(bot_meta["nick"])[:120] if mode == "bot" else "",
        mode=mode,
        difficulty=str(bot_meta["difficulty"]),
        is_public=bool(is_public),
        status="active" if mode == "bot" else "waiting",
        state_json=json.dumps(create_checkers_state(), ensure_ascii=False),
        stats_applied=False,
        last_move_at=None,
        finished_at=None,
    )
    db.add(room)
    db.flush()
    _audit(
        db,
        user,
        action="social_checkers_room_created",
        details=json.dumps({"room_id": int(room.id), "mode": mode, "title": title}, ensure_ascii=False),
        module_code="social_hub",
        entity_type="checkers_room",
        entity_id=str(room.id),
        request=request,
    )
    db.commit()
    db.refresh(room)
    return _social_checkers_room_payload(db, room, actor_key)


@router.get("/social/games/checkers/rooms/{room_id}", response_model=dict[str, Any])
def social_checkers_get_room(
    room_id: int,
    user: User = Depends(get_current_user),
    db: Session = Depends(get_db),
):
    ensure_module_enabled(db, user, "social_hub")
    actor_key, _, _, _ = _social_checkers_my_profile(db, user)
    room = db.get(SocialCheckersRoom, int(room_id or 0))
    if not room or str(room.game_code or "") != CHECKERS_GAME_CODE:
        raise HTTPException(status_code=404, detail="РљРѕРјРЅР°С‚Р° РЅРµ РЅР°Р№РґРµРЅР°")
    if not _social_checkers_room_visible(db, room, actor_key):
        raise HTTPException(status_code=403, detail="РќРµС‚ РґРѕСЃС‚СѓРїР° Рє РєРѕРјРЅР°С‚Рµ")
    db.commit()
    return _social_checkers_room_payload(db, room, actor_key)

@router.post("/social/games/checkers/rooms/{room_id}/join", response_model=dict[str, Any])
def social_checkers_join_room(
    room_id: int,
    request: Request,
    user: User = Depends(get_current_user),
    db: Session = Depends(get_db),
):
    ensure_module_enabled(db, user, "social_hub")
    actor_key, actor_nick, member_id, _ = _social_checkers_my_profile(db, user)
    room = db.get(SocialCheckersRoom, int(room_id or 0))
    if not room or str(room.game_code or "") != CHECKERS_GAME_CODE:
        raise HTTPException(status_code=404, detail="РљРѕРјРЅР°С‚Р° РЅРµ РЅР°Р№РґРµРЅР°")
    if str(room.mode or "") != "human":
        raise HTTPException(status_code=400, detail="Рљ Р±РѕС‚Сѓ РїРѕРґРєР»СЋС‡РµРЅРёРµ РЅРµ С‚СЂРµР±СѓРµС‚СЃСЏ")
    if str(room.status or "") != "waiting":
        raise HTTPException(status_code=409, detail="РљРѕРјРЅР°С‚Р° СѓР¶Рµ Р·Р°РЅСЏС‚Р°")
    if not bool(room.is_public):
        raise HTTPException(status_code=403, detail="РљРѕРјРЅР°С‚Р° Р·Р°РєСЂС‹С‚Р° РґР»СЏ РїРѕРґРєР»СЋС‡РµРЅРёСЏ")
    if _social_checkers_room_side(db, room, actor_key) == "white":
        raise HTTPException(status_code=400, detail="РќРµР»СЊР·СЏ РїРѕРґРєР»СЋС‡РёС‚СЊСЃСЏ Рє СЃРІРѕРµР№ РєРѕРјРЅР°С‚Рµ")
    room.guest_user_id = int(user.id)
    room.guest_member_id = int(member_id or 0) or None
    room.guest_actor_key = _social_canonical_actor_key(db, actor_key)
    room.guest_nick = actor_nick[:120]
    room.status = "active"
    room.updated_at = datetime.utcnow()
    _audit(
        db,
        user,
        action="social_checkers_room_joined",
        details=json.dumps({"room_id": int(room.id), "room_code": str(room.room_code or "")}, ensure_ascii=False),
        module_code="social_hub",
        entity_type="checkers_room",
        entity_id=str(room.id),
        request=request,
    )
    db.commit()
    db.refresh(room)
    return _social_checkers_room_payload(db, room, actor_key)


@router.post("/social/games/checkers/rooms/{room_id}/move", response_model=dict[str, Any])
def social_checkers_make_move(
    room_id: int,
    payload: dict[str, Any],
    request: Request,
    user: User = Depends(get_current_user),
    db: Session = Depends(get_db),
):
    ensure_module_enabled(db, user, "social_hub")
    actor_key, _, _, _ = _social_checkers_my_profile(db, user)
    room = db.get(SocialCheckersRoom, int(room_id or 0))
    if not room or str(room.game_code or "") != CHECKERS_GAME_CODE:
        raise HTTPException(status_code=404, detail="РљРѕРјРЅР°С‚Р° РЅРµ РЅР°Р№РґРµРЅР°")
    my_side = _social_checkers_room_side(db, room, actor_key)
    if not my_side:
        raise HTTPException(status_code=403, detail="РќРµС‚ РґРѕСЃС‚СѓРїР° Рє РїР°СЂС‚РёРё")
    if str(room.status or "") != "active":
        raise HTTPException(status_code=409, detail="РљРѕРјРЅР°С‚Р° СЃРµР№С‡Р°СЃ РЅРµР°РєС‚РёРІРЅР°")
    state = _social_checkers_room_state(room)
    if str(state.get("turn") or "") != my_side:
        raise HTTPException(status_code=409, detail="РЎРµР№С‡Р°СЃ С…РѕРґ РґСЂСѓРіРѕРіРѕ РёРіСЂРѕРєР°")
    try:
        next_state = apply_checkers_move(state, (payload or {}).get("path"))
    except ValueError as exc:
        raise HTTPException(status_code=400, detail=str(exc))
    if str(room.mode or "") == "bot" and not str(next_state.get("winner") or "") and str(next_state.get("result") or "") != "draw" and str(next_state.get("turn") or "") == "black":
        bot_move = pick_checkers_bot_move(next_state, room.difficulty)
        if bot_move and isinstance(bot_move, dict):
            next_state = apply_checkers_move(next_state, bot_move.get("path"))
    move_stamp = _to_utc_iso(datetime.utcnow())
    if next_state.get("last_move"):
        next_state["last_move"]["at"] = move_stamp
    if next_state.get("history"):
        next_state["history"][-1]["at"] = move_stamp
    _social_checkers_store_room_state(room, next_state)
    room.last_move_at = datetime.utcnow()
    if str(next_state.get("winner") or "") or str(next_state.get("result") or "") == "draw":
        room.status = "finished"
        room.finished_at = datetime.utcnow()
        _social_checkers_apply_stats(db, room, next_state)
    move_marker = str(len(next_state.get("history") or []))
    _social_emit_game_turn_notification(
        db,
        room=room,
        game_code=CHECKERS_GAME_CODE,
        mover_actor_key=actor_key,
        mover_nick=_social_current_nick_by_key(db, actor_key) or actor_key,
        move_marker=move_marker,
        move_at_iso=move_stamp,
    )
    _audit(
        db,
        user,
        action="social_checkers_move",
        details=json.dumps({"room_id": int(room.id), "side": my_side, "path": next_state.get("last_move", {}).get("path", [])}, ensure_ascii=False),
        module_code="social_hub",
        entity_type="checkers_room",
        entity_id=str(room.id),
        request=request,
    )
    db.commit()
    db.refresh(room)
    return _social_checkers_room_payload(db, room, actor_key)


@router.post("/social/games/checkers/rooms/{room_id}/leave", response_model=dict[str, Any])
def social_checkers_leave_room(
    room_id: int,
    request: Request,
    user: User = Depends(get_current_user),
    db: Session = Depends(get_db),
):
    ensure_module_enabled(db, user, "social_hub")
    actor_key, _, _, _ = _social_checkers_my_profile(db, user)
    room = db.get(SocialCheckersRoom, int(room_id or 0))
    if not room or str(room.game_code or "") != CHECKERS_GAME_CODE:
        raise HTTPException(status_code=404, detail="РљРѕРјРЅР°С‚Р° РЅРµ РЅР°Р№РґРµРЅР°")
    my_side = _social_checkers_room_side(db, room, actor_key)
    if not my_side:
        raise HTTPException(status_code=403, detail="РќРµС‚ РґРѕСЃС‚СѓРїР° Рє РїР°СЂС‚РёРё")
    state = _social_checkers_room_state(room)
    if str(room.status or "") == "waiting":
        room.status = "cancelled"
        state["result"] = "cancelled"
        state["winner"] = ""
    elif str(room.status or "") == "active":
        winner = "black" if my_side == "white" else "white"
        state["winner"] = winner
        state["result"] = "resigned"
        state["last_move"] = {"side": my_side, "path": [], "capture_count": 0, "promoted": False, "at": _to_utc_iso(datetime.utcnow())}
        room.status = "finished"
        room.finished_at = datetime.utcnow()
        _social_checkers_apply_stats(db, room, state)
    _social_checkers_store_room_state(room, state)
    room.last_move_at = datetime.utcnow()
    _audit(
        db,
        user,
        action="social_checkers_room_left",
        details=json.dumps({"room_id": int(room.id), "side": my_side, "status": str(room.status or "")}, ensure_ascii=False),
        module_code="social_hub",
        entity_type="checkers_room",
        entity_id=str(room.id),
        request=request,
    )
    db.commit()
    db.refresh(room)
    return _social_checkers_room_payload(db, room, actor_key)



def _social_board_profile_storage_key(db: Session, actor_key: str, game_code: str) -> tuple[str, str]:
    canonical = _social_canonical_actor_key(db, actor_key)
    safe_game = str(game_code or CHECKERS_GAME_CODE).strip().lower() or CHECKERS_GAME_CODE
    if safe_game in _SOCIAL_BOARD_PROFILE_PREFIX_GAMES:
        return canonical, f"{safe_game}:{canonical}"[:60]
    return canonical, canonical


def _social_board_profile_row(
    db: Session,
    *,
    actor_key: str,
    actor_nick: str = "",
    user_id: int | None = None,
    member_id: int | None = None,
    game_code: str = CHECKERS_GAME_CODE,
) -> SocialCheckersProfile:
    canonical, storage_key = _social_board_profile_storage_key(db, actor_key, game_code)
    row = db.scalar(select(SocialCheckersProfile).where(SocialCheckersProfile.actor_key == storage_key))
    safe_user_id = int(user_id or 0)
    safe_member_id = int(member_id or 0) or None
    safe_nick = str(actor_nick or "").strip()
    if safe_user_id <= 0 or not safe_nick:
        try:
            resolved_user_id, resolved_member_id, resolved_nick = _social_identity_by_key(db, canonical)
            if safe_user_id <= 0:
                safe_user_id = int(resolved_user_id or 0)
            if not safe_member_id and resolved_member_id:
                safe_member_id = int(resolved_member_id or 0)
            if not safe_nick:
                safe_nick = str(resolved_nick or "").strip()
        except HTTPException:
            pass
    safe_nick = safe_nick or _social_current_nick_by_key(db, canonical) or canonical
    if safe_user_id <= 0:
        raise HTTPException(status_code=400, detail="РќРµРєРѕСЂСЂРµРєС‚РЅС‹Р№ РёРіСЂРѕРє РґР»СЏ СЂРµР№С‚РёРЅРіР°")
    if row:
        row.user_id = safe_user_id
        row.member_id = safe_member_id
        row.actor_nick = safe_nick[:120]
        row.updated_at = datetime.utcnow()
        return row
    row = SocialCheckersProfile(
        user_id=safe_user_id,
        member_id=safe_member_id,
        actor_key=storage_key,
        actor_nick=safe_nick[:120],
        rating=1200,
        wins=0,
        losses=0,
        draws=0,
        play_count=0,
    )
    db.add(row)
    db.flush()
    return row


def _social_board_my_profile(db: Session, user: User, game_code: str) -> tuple[str, str, int | None, SocialCheckersProfile]:
    actor_key, actor_nick, member_id = _social_actor_identity(db, user)
    row = _social_board_profile_row(
        db,
        actor_key=actor_key,
        actor_nick=actor_nick,
        user_id=int(user.id),
        member_id=member_id,
        game_code=game_code,
    )
    return actor_key, actor_nick, member_id, row


def _social_board_public_player(
    db: Session,
    actor_key: str,
    fallback_nick: str,
    *,
    game_code: str,
    difficulty: str,
    bot_builder,
) -> dict[str, Any]:
    key = str(actor_key or "").strip().lower()
    if key.startswith("bot:"):
        bot = bot_builder(difficulty or key.rsplit(":", 1)[-1])
        return {
            "actor_key": str(bot.get("actor_key") or "bot"),
            "nick": str(bot.get("nick") or "SEO WIBE AI"),
            "rating": int(bot.get("rating") or 1200),
            "wins": 0,
            "losses": 0,
            "draws": 0,
            "play_count": 0,
            "is_bot": True,
            "avatar_url": "",
        }
    canonical, storage_key = _social_board_profile_storage_key(db, key, game_code)
    profile_row = db.scalar(select(SocialCheckersProfile).where(SocialCheckersProfile.actor_key == storage_key))
    public_profile = _social_public_profile_by_key(db, canonical)
    nick = str(public_profile.get("nick") or fallback_nick or canonical).strip()
    return {
        "actor_key": canonical,
        "nick": nick[:120],
        "rating": int(profile_row.rating or 1200) if profile_row else 1200,
        "wins": int(profile_row.wins or 0) if profile_row else 0,
        "losses": int(profile_row.losses or 0) if profile_row else 0,
        "draws": int(profile_row.draws or 0) if profile_row else 0,
        "play_count": int(profile_row.play_count or 0) if profile_row else 0,
        "is_bot": False,
        "avatar_url": str(public_profile.get("avatar_url") or ""),
    }


def _social_board_rank_rows(db: Session, viewer_key: str, *, game_code: str, limit: int = 100) -> tuple[list[dict[str, Any]], int | None, int]:
    safe_limit = max(5, min(int(limit or 100), 200))
    canonical, profile_key = _social_board_profile_storage_key(db, viewer_key, game_code)
    query = select(SocialCheckersProfile)
    key_query = select(SocialCheckersProfile.actor_key)
    if str(game_code) in _SOCIAL_BOARD_PROFILE_PREFIX_GAMES:
        prefix = f"{str(game_code)}:%"
        query = query.where(SocialCheckersProfile.actor_key.like(prefix))
        key_query = key_query.where(SocialCheckersProfile.actor_key.like(prefix))
    else:
        for pref_game in sorted(_SOCIAL_BOARD_PROFILE_PREFIX_GAMES):
            query = query.where(~SocialCheckersProfile.actor_key.like(f"{pref_game}:%"))
            key_query = key_query.where(~SocialCheckersProfile.actor_key.like(f"{pref_game}:%"))
    query = query.order_by(
        SocialCheckersProfile.rating.desc(),
        SocialCheckersProfile.wins.desc(),
        SocialCheckersProfile.play_count.desc(),
        SocialCheckersProfile.updated_at.asc(),
        SocialCheckersProfile.id.asc(),
    )
    rows = db.scalars(query.limit(safe_limit)).all()
    ordered_keys = db.scalars(key_query.order_by(
        SocialCheckersProfile.rating.desc(),
        SocialCheckersProfile.wins.desc(),
        SocialCheckersProfile.play_count.desc(),
        SocialCheckersProfile.updated_at.asc(),
        SocialCheckersProfile.id.asc(),
    )).all()
    my_rank = None
    for idx, key in enumerate(ordered_keys, start=1):
        if str(key or "") == profile_key:
            my_rank = idx
            break
    my_row = db.scalar(select(SocialCheckersProfile).where(SocialCheckersProfile.actor_key == profile_key))
    my_rating = int(my_row.rating or 1200) if my_row else 1200
    data_rows: list[dict[str, Any]] = []
    for idx, row in enumerate(rows, start=1):
        actor_storage = str(row.actor_key or "")
        actor_public = actor_storage
        if str(game_code) in _SOCIAL_BOARD_PROFILE_PREFIX_GAMES and actor_storage.startswith(f"{str(game_code)}:"):
            actor_public = actor_storage[len(str(game_code)) + 1 :]
        data_rows.append(
            {
                "rank": idx,
                "actor_key": actor_public,
                "nick": str(row.actor_nick or actor_public),
                "rating": int(row.rating or 1200),
                "wins": int(row.wins or 0),
                "losses": int(row.losses or 0),
                "draws": int(row.draws or 0),
                "play_count": int(row.play_count or 0),
                "is_me": actor_storage == profile_key,
            }
        )
    return data_rows, my_rank, my_rating


def _social_board_apply_stats(
    db: Session,
    room: SocialCheckersRoom,
    state: dict[str, Any],
    *,
    game_code: str,
    bot_builder,
) -> None:
    if bool(room.stats_applied):
        return
    result = str(state.get("result") or "")
    winner = str(state.get("winner") or "")
    if str(room.status or "") != "finished":
        return
    if str(room.mode or "") == "bot":
        host_profile = _social_board_profile_row(
            db,
            actor_key=str(room.host_actor_key or ""),
            actor_nick=str(room.host_nick or ""),
            user_id=int(room.host_user_id or 0),
            member_id=int(room.host_member_id or 0) or None,
            game_code=game_code,
        )
        bot_meta = bot_builder(room.difficulty)
        if result == "draw":
            host_profile.draws = int(host_profile.draws or 0) + 1
            score_value = 0.5
        elif winner == "white":
            host_profile.wins = int(host_profile.wins or 0) + 1
            score_value = 1.0
        else:
            host_profile.losses = int(host_profile.losses or 0) + 1
            score_value = 0.0
        host_profile.play_count = int(host_profile.play_count or 0) + 1
        host_profile.rating, _ = apply_checkers_elo(int(host_profile.rating or 1200), int(bot_meta.get("rating") or 1200), score_value)
        room.stats_applied = True
        return
    if not str(room.guest_actor_key or ""):
        return
    white_profile = _social_board_profile_row(
        db,
        actor_key=str(room.host_actor_key or ""),
        actor_nick=str(room.host_nick or ""),
        user_id=int(room.host_user_id or 0),
        member_id=int(room.host_member_id or 0) or None,
        game_code=game_code,
    )
    black_profile = _social_board_profile_row(
        db,
        actor_key=str(room.guest_actor_key or ""),
        actor_nick=str(room.guest_nick or ""),
        user_id=int(room.guest_user_id or 0),
        member_id=int(room.guest_member_id or 0) or None,
        game_code=game_code,
    )
    if result == "draw":
        white_score = 0.5
        black_score = 0.5
        white_profile.draws = int(white_profile.draws or 0) + 1
        black_profile.draws = int(black_profile.draws or 0) + 1
    elif winner == "white":
        white_score = 1.0
        black_score = 0.0
        white_profile.wins = int(white_profile.wins or 0) + 1
        black_profile.losses = int(black_profile.losses or 0) + 1
    else:
        white_score = 0.0
        black_score = 1.0
        white_profile.losses = int(white_profile.losses or 0) + 1
        black_profile.wins = int(black_profile.wins or 0) + 1
    white_profile.play_count = int(white_profile.play_count or 0) + 1
    black_profile.play_count = int(black_profile.play_count or 0) + 1
    next_white, _ = apply_checkers_elo(int(white_profile.rating or 1200), int(black_profile.rating or 1200), white_score)
    next_black, _ = apply_checkers_elo(int(black_profile.rating or 1200), int(white_profile.rating or 1200), black_score)
    white_profile.rating = next_white
    black_profile.rating = next_black
    room.stats_applied = True


def _social_chess_room_state(room: SocialCheckersRoom) -> dict[str, Any]:
    try:
        raw = json.loads(str(room.state_json or "{}"))
    except Exception:
        raw = {}
    return load_chess_state(raw)


def _social_chess_store_room_state(room: SocialCheckersRoom, state: dict[str, Any]) -> None:
    room.state_json = json.dumps(load_chess_state(state), ensure_ascii=False)


def _social_chess_note(room: SocialCheckersRoom, payload: dict[str, Any]) -> str:
    status = str(payload.get("status") or room.status or "")
    if status == "waiting":
        return "РљРѕРјРЅР°С‚Р° РІРёРґРЅР° РІ РѕР±С‰РµРј Р»РѕР±Р±Рё Рё Р¶РґРµС‚ РІС‚РѕСЂРѕРіРѕ РёРіСЂРѕРєР°."
    if status == "active" and str(room.mode or "") == "bot":
        return f"РЈСЂРѕРІРµРЅСЊ РР: {str(room.difficulty or 'medium')}"
    if status == "active":
        return "РРіСЂР° РёРґРµС‚ РІ СЂРµР°Р»СЊРЅРѕРј РІСЂРµРјРµРЅРё."
    if status == "finished":
        result = str(payload.get("result") or "")
        winner = str(payload.get("winner") or "")
        if result == "draw":
            return "РџР°СЂС‚РёСЏ Р·Р°РІРµСЂС€РёР»Р°СЃСЊ РЅРёС‡СЊРµР№."
        if winner == "white":
            return "Р‘РµР»С‹Рµ Р·Р°РІРµСЂС€РёР»Рё РїР°СЂС‚РёСЋ РїРѕР±РµРґРѕР№."
        if winner == "black":
            return "Р§РµСЂРЅС‹Рµ Р·Р°РІРµСЂС€РёР»Рё РїР°СЂС‚РёСЋ РїРѕР±РµРґРѕР№."
    if status == "cancelled":
        return "РљРѕРјРЅР°С‚Р° Р·Р°РєСЂС‹С‚Р° РґРѕ СЃС‚Р°СЂС‚Р° РёР»Рё Р·Р°РІРµСЂС€РµРЅР° РІСЂСѓС‡РЅСѓСЋ."
    return ""


def _social_chess_room_payload(db: Session, room: SocialCheckersRoom, viewer_key: str) -> dict[str, Any]:
    state = _social_chess_room_state(room)
    my_side = _social_checkers_room_side(db, room, viewer_key)
    can_join = bool(room.is_public) and str(room.mode or "") == "human" and str(room.status or "") == "waiting" and not my_side and not str(room.guest_actor_key or "")
    can_move = bool(my_side) and str(room.status or "") == "active" and str(state.get("turn") or "") == my_side and not str(state.get("winner") or "") and str(state.get("result") or "") != "draw"
    white_player = _social_board_public_player(
        db,
        str(room.host_actor_key or ""),
        str(room.host_nick or ""),
        game_code=CHESS_GAME_CODE,
        difficulty=str(room.difficulty or ""),
        bot_builder=build_chess_bot_identity,
    )
    if str(room.mode or "") == "bot":
        bot_meta = build_chess_bot_identity(room.difficulty)
        black_key = str(bot_meta.get("actor_key") or "bot:chess:medium")
        black_nick = str(bot_meta.get("nick") or "SEO WIBE AI")
    else:
        black_key = str(room.guest_actor_key or "")
        black_nick = str(room.guest_nick or "")
    black_player = _social_board_public_player(
        db,
        black_key,
        black_nick,
        game_code=CHESS_GAME_CODE,
        difficulty=str(room.difficulty or ""),
        bot_builder=build_chess_bot_identity,
    )
    legal_moves = get_chess_legal_moves(state, my_side) if can_move else []
    payload = {
        "id": int(room.id or 0),
        "room_code": str(room.room_code or ""),
        "title": str(room.title or "").strip() or f"РљРѕРјРЅР°С‚Р° {room.room_code}",
        "mode": str(room.mode or "human"),
        "difficulty": str(room.difficulty or "medium"),
        "status": str(room.status or "waiting"),
        "is_public": bool(room.is_public),
        "turn": str(state.get("turn") or "white"),
        "winner": str(state.get("winner") or ""),
        "result": str(state.get("result") or ""),
        "board": state.get("board") or [],
        "last_move": state.get("last_move") or {},
        "history": state.get("history") or [],
        "players": {
            "white": white_player,
            "black": black_player,
        },
        "my_side": my_side,
        "my_turn": bool(can_move),
        "can_join": bool(can_join),
        "can_move": bool(can_move),
        "legal_moves": legal_moves,
        "created_at": _to_utc_iso(room.created_at),
        "updated_at": _to_utc_iso(room.updated_at),
        "last_move_at": _to_utc_iso(room.last_move_at),
        "finished_at": _to_utc_iso(room.finished_at),
    }
    payload["note"] = _social_chess_note(room, payload)
    return payload


@router.get("/social/games/chess/overview", response_model=dict[str, Any])
def social_chess_overview(
    user: User = Depends(get_current_user),
    db: Session = Depends(get_db),
):
    ensure_module_enabled(db, user, "social_hub")
    actor_key, actor_nick, member_id, profile_row = _social_board_my_profile(db, user, CHESS_GAME_CODE)
    aliases = _social_actor_alias_keys(db, actor_key)
    public_rooms = db.scalars(
        select(SocialCheckersRoom)
        .where(
            SocialCheckersRoom.game_code == CHESS_GAME_CODE,
            SocialCheckersRoom.mode == "human",
            SocialCheckersRoom.is_public.is_(True),
            SocialCheckersRoom.status == "waiting",
        )
        .order_by(SocialCheckersRoom.updated_at.desc(), SocialCheckersRoom.id.desc())
        .limit(24)
    ).all()
    my_rooms = db.scalars(
        select(SocialCheckersRoom)
        .where(
            SocialCheckersRoom.game_code == CHESS_GAME_CODE,
            or_(
                SocialCheckersRoom.host_actor_key.in_(aliases),
                SocialCheckersRoom.guest_actor_key.in_(aliases),
            ),
            SocialCheckersRoom.status.in_(["waiting", "active", "finished", "cancelled"]),
        )
        .order_by(SocialCheckersRoom.updated_at.desc(), SocialCheckersRoom.id.desc())
        .limit(12)
    ).all()
    leaderboard_rows, my_rank, my_rating = _social_board_rank_rows(db, actor_key, game_code=CHESS_GAME_CODE, limit=20)
    db.commit()
    return {
        "profile": {
            **_social_checkers_profile_out(profile_row, actor_key=actor_key),
            "nick": actor_nick,
            "member_id": int(member_id or 0),
        },
        "leaderboard": {
            "rows": leaderboard_rows,
            "my_rank": my_rank,
            "my_rating": my_rating,
        },
        "rooms": {
            "public": [_social_chess_room_payload(db, row, actor_key) for row in public_rooms],
            "mine": [_social_chess_room_payload(db, row, actor_key) for row in my_rooms],
        },
        "difficulties": get_chess_difficulties(),
    }


@router.get("/social/games/chess/leaderboard", response_model=dict[str, Any])
def social_chess_leaderboard(
    limit: int = 100,
    user: User = Depends(get_current_user),
    db: Session = Depends(get_db),
):
    ensure_module_enabled(db, user, "social_hub")
    actor_key, _, _, profile_row = _social_board_my_profile(db, user, CHESS_GAME_CODE)
    rows, my_rank, my_rating = _social_board_rank_rows(db, actor_key, game_code=CHESS_GAME_CODE, limit=limit)
    db.commit()
    return {
        "rows": rows,
        "my_rank": my_rank,
        "my_rating": my_rating,
        "my_profile": _social_checkers_profile_out(profile_row, actor_key=actor_key),
    }


@router.post("/social/games/chess/rooms", response_model=dict[str, Any])
def social_chess_create_room(
    payload: dict[str, Any],
    request: Request,
    user: User = Depends(get_current_user),
    db: Session = Depends(get_db),
):
    ensure_module_enabled(db, user, "social_hub")
    actor_key, actor_nick, member_id, _ = _social_board_my_profile(db, user, CHESS_GAME_CODE)
    mode = str((payload or {}).get("mode") or "human").strip().lower()
    if mode not in {"human", "bot"}:
        raise HTTPException(status_code=400, detail="РќРµРєРѕСЂСЂРµРєС‚РЅС‹Р№ СЂРµР¶РёРј РєРѕРјРЅР°С‚С‹")
    difficulty = str((payload or {}).get("difficulty") or "medium").strip().lower()
    bot_meta = build_chess_bot_identity(difficulty)
    safe_title = str((payload or {}).get("title") or "").strip()[:120]
    title = safe_title or (f"РљРѕРјРЅР°С‚Р° {actor_nick}" if mode == "human" else f"РњР°С‚С‡ СЃ {bot_meta['title']}")
    is_public = bool((payload or {}).get("is_public", True)) if mode == "human" else False
    existing_codes = set(db.scalars(select(SocialCheckersRoom.room_code)).all())
    room = SocialCheckersRoom(
        game_code=CHESS_GAME_CODE,
        room_code=create_checkers_room_code(existing_codes),
        title=title,
        owner_user_id=int(user.id),
        host_user_id=int(user.id),
        host_member_id=int(member_id or 0) or None,
        host_actor_key=_social_canonical_actor_key(db, actor_key),
        host_nick=actor_nick[:120],
        guest_user_id=None,
        guest_member_id=None,
        guest_actor_key=str(bot_meta["actor_key"]) if mode == "bot" else "",
        guest_nick=str(bot_meta["nick"])[:120] if mode == "bot" else "",
        mode=mode,
        difficulty=str(bot_meta["difficulty"]),
        is_public=bool(is_public),
        status="active" if mode == "bot" else "waiting",
        state_json=json.dumps(create_chess_state(), ensure_ascii=False),
        stats_applied=False,
        last_move_at=None,
        finished_at=None,
    )
    db.add(room)
    db.flush()
    _audit(
        db,
        user,
        action="social_chess_room_created",
        details=json.dumps({"room_id": int(room.id), "mode": mode, "title": title}, ensure_ascii=False),
        module_code="social_hub",
        entity_type="chess_room",
        entity_id=str(room.id),
        request=request,
    )
    db.commit()
    db.refresh(room)
    return _social_chess_room_payload(db, room, actor_key)


@router.get("/social/games/chess/rooms/{room_id}", response_model=dict[str, Any])
def social_chess_get_room(
    room_id: int,
    user: User = Depends(get_current_user),
    db: Session = Depends(get_db),
):
    ensure_module_enabled(db, user, "social_hub")
    actor_key, _, _, _ = _social_board_my_profile(db, user, CHESS_GAME_CODE)
    room = db.get(SocialCheckersRoom, int(room_id or 0))
    if not room or str(room.game_code or "") != CHESS_GAME_CODE:
        raise HTTPException(status_code=404, detail="РљРѕРјРЅР°С‚Р° РЅРµ РЅР°Р№РґРµРЅР°")
    if not _social_checkers_room_visible(db, room, actor_key):
        raise HTTPException(status_code=403, detail="РќРµС‚ РґРѕСЃС‚СѓРїР° Рє РєРѕРјРЅР°С‚Рµ")
    db.commit()
    return _social_chess_room_payload(db, room, actor_key)


@router.post("/social/games/chess/rooms/{room_id}/join", response_model=dict[str, Any])
def social_chess_join_room(
    room_id: int,
    request: Request,
    user: User = Depends(get_current_user),
    db: Session = Depends(get_db),
):
    ensure_module_enabled(db, user, "social_hub")
    actor_key, actor_nick, member_id, _ = _social_board_my_profile(db, user, CHESS_GAME_CODE)
    room = db.get(SocialCheckersRoom, int(room_id or 0))
    if not room or str(room.game_code or "") != CHESS_GAME_CODE:
        raise HTTPException(status_code=404, detail="РљРѕРјРЅР°С‚Р° РЅРµ РЅР°Р№РґРµРЅР°")
    if str(room.mode or "") != "human":
        raise HTTPException(status_code=400, detail="Рљ Р±РѕС‚Сѓ РїРѕРґРєР»СЋС‡РµРЅРёРµ РЅРµ С‚СЂРµР±СѓРµС‚СЃСЏ")
    if str(room.status or "") != "waiting":
        raise HTTPException(status_code=409, detail="РљРѕРјРЅР°С‚Р° СѓР¶Рµ Р·Р°РЅСЏС‚Р°")
    if not bool(room.is_public):
        raise HTTPException(status_code=403, detail="РљРѕРјРЅР°С‚Р° Р·Р°РєСЂС‹С‚Р° РґР»СЏ РїРѕРґРєР»СЋС‡РµРЅРёСЏ")
    if _social_checkers_room_side(db, room, actor_key) == "white":
        raise HTTPException(status_code=400, detail="РќРµР»СЊР·СЏ РїРѕРґРєР»СЋС‡РёС‚СЊСЃСЏ Рє СЃРІРѕРµР№ РєРѕРјРЅР°С‚Рµ")
    room.guest_user_id = int(user.id)
    room.guest_member_id = int(member_id or 0) or None
    room.guest_actor_key = _social_canonical_actor_key(db, actor_key)
    room.guest_nick = actor_nick[:120]
    room.status = "active"
    room.updated_at = datetime.utcnow()
    _audit(
        db,
        user,
        action="social_chess_room_joined",
        details=json.dumps({"room_id": int(room.id), "room_code": str(room.room_code or "")}, ensure_ascii=False),
        module_code="social_hub",
        entity_type="chess_room",
        entity_id=str(room.id),
        request=request,
    )
    db.commit()
    db.refresh(room)
    return _social_chess_room_payload(db, room, actor_key)


@router.post("/social/games/chess/rooms/{room_id}/move", response_model=dict[str, Any])
def social_chess_make_move(
    room_id: int,
    payload: dict[str, Any],
    request: Request,
    user: User = Depends(get_current_user),
    db: Session = Depends(get_db),
):
    ensure_module_enabled(db, user, "social_hub")
    actor_key, _, _, _ = _social_board_my_profile(db, user, CHESS_GAME_CODE)
    room = db.get(SocialCheckersRoom, int(room_id or 0))
    if not room or str(room.game_code or "") != CHESS_GAME_CODE:
        raise HTTPException(status_code=404, detail="РљРѕРјРЅР°С‚Р° РЅРµ РЅР°Р№РґРµРЅР°")
    my_side = _social_checkers_room_side(db, room, actor_key)
    if not my_side:
        raise HTTPException(status_code=403, detail="РќРµС‚ РґРѕСЃС‚СѓРїР° Рє РїР°СЂС‚РёРё")
    if str(room.status or "") != "active":
        raise HTTPException(status_code=409, detail="РљРѕРјРЅР°С‚Р° СЃРµР№С‡Р°СЃ РЅРµР°РєС‚РёРІРЅР°")
    state = _social_chess_room_state(room)
    if str(state.get("turn") or "") != my_side:
        raise HTTPException(status_code=409, detail="РЎРµР№С‡Р°СЃ С…РѕРґ РґСЂСѓРіРѕРіРѕ РёРіСЂРѕРєР°")
    try:
        next_state = apply_chess_move(state, payload or {})
    except ValueError as exc:
        raise HTTPException(status_code=400, detail=str(exc))
    if (
        str(room.mode or "") == "bot"
        and not str(next_state.get("winner") or "")
        and str(next_state.get("result") or "") != "draw"
        and str(next_state.get("turn") or "") == "black"
    ):
        bot_move = pick_chess_bot_move(next_state, room.difficulty)
        if bot_move and isinstance(bot_move, dict):
            next_state = apply_chess_move(next_state, bot_move)
    move_stamp = _to_utc_iso(datetime.utcnow())
    if next_state.get("last_move"):
        next_state["last_move"]["at"] = move_stamp
    if next_state.get("history"):
        next_state["history"][-1]["at"] = move_stamp
    _social_chess_store_room_state(room, next_state)
    room.last_move_at = datetime.utcnow()
    if str(next_state.get("winner") or "") or str(next_state.get("result") or "") == "draw":
        room.status = "finished"
        room.finished_at = datetime.utcnow()
        _social_board_apply_stats(
            db,
            room,
            next_state,
            game_code=CHESS_GAME_CODE,
            bot_builder=build_chess_bot_identity,
        )
    move_marker = str(len(next_state.get("history") or []))
    _social_emit_game_turn_notification(
        db,
        room=room,
        game_code=CHESS_GAME_CODE,
        mover_actor_key=actor_key,
        mover_nick=_social_current_nick_by_key(db, actor_key) or actor_key,
        move_marker=move_marker,
        move_at_iso=move_stamp,
    )
    _audit(
        db,
        user,
        action="social_chess_move",
        details=json.dumps({"room_id": int(room.id), "side": my_side, "move": next_state.get("last_move", {})}, ensure_ascii=False),
        module_code="social_hub",
        entity_type="chess_room",
        entity_id=str(room.id),
        request=request,
    )
    db.commit()
    db.refresh(room)
    return _social_chess_room_payload(db, room, actor_key)


@router.post("/social/games/chess/rooms/{room_id}/leave", response_model=dict[str, Any])
def social_chess_leave_room(
    room_id: int,
    request: Request,
    user: User = Depends(get_current_user),
    db: Session = Depends(get_db),
):
    ensure_module_enabled(db, user, "social_hub")
    actor_key, _, _, _ = _social_board_my_profile(db, user, CHESS_GAME_CODE)
    room = db.get(SocialCheckersRoom, int(room_id or 0))
    if not room or str(room.game_code or "") != CHESS_GAME_CODE:
        raise HTTPException(status_code=404, detail="РљРѕРјРЅР°С‚Р° РЅРµ РЅР°Р№РґРµРЅР°")
    my_side = _social_checkers_room_side(db, room, actor_key)
    if not my_side:
        raise HTTPException(status_code=403, detail="РќРµС‚ РґРѕСЃС‚СѓРїР° Рє РїР°СЂС‚РёРё")
    state = _social_chess_room_state(room)
    if str(room.status or "") == "waiting":
        room.status = "cancelled"
        state["result"] = "cancelled"
        state["winner"] = ""
    elif str(room.status or "") == "active":
        winner = "black" if my_side == "white" else "white"
        state["winner"] = winner
        state["result"] = "resigned"
        state["last_move"] = {
            "side": my_side,
            "from": None,
            "to": None,
            "piece": "",
            "capture": False,
            "promotion": "",
            "at": _to_utc_iso(datetime.utcnow()),
        }
        room.status = "finished"
        room.finished_at = datetime.utcnow()
        _social_board_apply_stats(
            db,
            room,
            state,
            game_code=CHESS_GAME_CODE,
            bot_builder=build_chess_bot_identity,
        )
    _social_chess_store_room_state(room, state)
    room.last_move_at = datetime.utcnow()
    _audit(
        db,
        user,
        action="social_chess_room_left",
        details=json.dumps({"room_id": int(room.id), "side": my_side, "status": str(room.status or "")}, ensure_ascii=False),
        module_code="social_hub",
        entity_type="chess_room",
        entity_id=str(room.id),
        request=request,
    )
    db.commit()
    db.refresh(room)
    return _social_chess_room_payload(db, room, actor_key)




def _social_battleship_room_state(room: SocialCheckersRoom) -> dict[str, Any]:
    try:
        raw = json.loads(str(room.state_json or "{}"))
    except Exception:
        raw = {}
    return load_battleship_state(raw)


def _social_battleship_store_room_state(room: SocialCheckersRoom, state: dict[str, Any]) -> None:
    room.state_json = json.dumps(load_battleship_state(state), ensure_ascii=False)


def _social_battleship_note(room: SocialCheckersRoom, payload: dict[str, Any]) -> str:
    status = str(payload.get("status") or room.status or "")
    if status == "waiting":
        return "РљРѕРјРЅР°С‚Р° РІРёРґРЅР° РІ РѕР±С‰РµРј Р»РѕР±Р±Рё Рё Р¶РґРµС‚ РІС‚РѕСЂРѕРіРѕ РёРіСЂРѕРєР°."
    if status == "active" and str(room.mode or "") == "bot":
        return f"РЈСЂРѕРІРµРЅСЊ РР: {str(room.difficulty or 'medium')}"
    if status == "active":
        return "РњР°С‚С‡ Р°РєС‚РёРІРµРЅ. Р”РµР»Р°Р№С‚Рµ РІС‹СЃС‚СЂРµР»С‹ РїРѕ РїРѕР»СЋ СЃРѕРїРµСЂРЅРёРєР°."
    if status == "finished":
        result = str(payload.get("result") or "")
        winner = str(payload.get("winner") or "")
        if result == "draw":
            return "РњР°С‚С‡ Р·Р°РІРµСЂС€РёР»СЃСЏ РЅРёС‡СЊРµР№."
        if winner == "white":
            return "Р‘РµР»С‹Р№ С„Р»РѕС‚ РїРѕР±РµРґРёР»."
        if winner == "black":
            return "Р§РµСЂРЅС‹Р№ С„Р»РѕС‚ РїРѕР±РµРґРёР»."
    if status == "cancelled":
        return "РљРѕРјРЅР°С‚Р° Р·Р°РєСЂС‹С‚Р° РґРѕ СЃС‚Р°СЂС‚Р° РёР»Рё Р·Р°РІРµСЂС€РµРЅР° РІСЂСѓС‡РЅСѓСЋ."
    return ""


def _social_battleship_room_payload(db: Session, room: SocialCheckersRoom, viewer_key: str) -> dict[str, Any]:
    state = _social_battleship_room_state(room)
    my_side = _social_checkers_room_side(db, room, viewer_key)
    can_join = bool(room.is_public) and str(room.mode or "") == "human" and str(room.status or "") == "waiting" and not my_side and not str(room.guest_actor_key or "")
    can_move = bool(my_side) and str(room.status or "") == "active" and str(state.get("turn") or "") == my_side and not str(state.get("winner") or "") and str(state.get("result") or "") != "draw"
    white_player = _social_board_public_player(
        db,
        str(room.host_actor_key or ""),
        str(room.host_nick or ""),
        game_code=BATTLESHIP_GAME_CODE,
        difficulty=str(room.difficulty or ""),
        bot_builder=build_battleship_bot_identity,
    )
    if str(room.mode or "") == "bot":
        bot_meta = build_battleship_bot_identity(room.difficulty)
        black_key = str(bot_meta.get("actor_key") or "bot:battleship:medium")
        black_nick = str(bot_meta.get("nick") or "SEO WIBE AI")
    else:
        black_key = str(room.guest_actor_key or "")
        black_nick = str(room.guest_nick or "")
    black_player = _social_board_public_player(
        db,
        black_key,
        black_nick,
        game_code=BATTLESHIP_GAME_CODE,
        difficulty=str(room.difficulty or ""),
        bot_builder=build_battleship_bot_identity,
    )

    players_state = state.get("players") if isinstance(state.get("players"), dict) else {}
    white_state = players_state.get("white") if isinstance(players_state.get("white"), dict) else {}
    black_state = players_state.get("black") if isinstance(players_state.get("black"), dict) else {}
    white_board = white_state.get("board") if isinstance(white_state.get("board"), list) else []
    black_board = black_state.get("board") if isinstance(black_state.get("board"), list) else []

    if my_side == "white":
        own_board = white_board
        enemy_board = mask_enemy_board(black_board)
        available = get_battleship_available_shots(state, "white") if can_move else []
    elif my_side == "black":
        own_board = black_board
        enemy_board = mask_enemy_board(white_board)
        available = get_battleship_available_shots(state, "black") if can_move else []
    else:
        own_board = mask_enemy_board(white_board)
        enemy_board = mask_enemy_board(black_board)
        available = []

    payload = {
        "id": int(room.id or 0),
        "room_code": str(room.room_code or ""),
        "title": str(room.title or "").strip() or f"РљРѕРјРЅР°С‚Р° {room.room_code}",
        "mode": str(room.mode or "human"),
        "difficulty": str(room.difficulty or "medium"),
        "status": str(room.status or "waiting"),
        "is_public": bool(room.is_public),
        "turn": str(state.get("turn") or "white"),
        "winner": str(state.get("winner") or ""),
        "result": str(state.get("result") or ""),
        "last_move": state.get("last_move") or {},
        "history": state.get("history") or [],
        "players": {
            "white": white_player,
            "black": black_player,
        },
        "my_side": my_side,
        "my_turn": bool(can_move),
        "can_join": bool(can_join),
        "can_move": bool(can_move),
        "own_board": own_board,
        "enemy_board": enemy_board,
        "available_shots": available,
        "created_at": _to_utc_iso(room.created_at),
        "updated_at": _to_utc_iso(room.updated_at),
        "last_move_at": _to_utc_iso(room.last_move_at),
        "finished_at": _to_utc_iso(room.finished_at),
    }
    payload["note"] = _social_battleship_note(room, payload)
    return payload


@router.get("/social/games/battleship/overview", response_model=dict[str, Any])
def social_battleship_overview(
    user: User = Depends(get_current_user),
    db: Session = Depends(get_db),
):
    ensure_module_enabled(db, user, "social_hub")
    actor_key, actor_nick, member_id, profile_row = _social_board_my_profile(db, user, BATTLESHIP_GAME_CODE)
    aliases = _social_actor_alias_keys(db, actor_key)
    public_rooms = db.scalars(
        select(SocialCheckersRoom)
        .where(
            SocialCheckersRoom.game_code == BATTLESHIP_GAME_CODE,
            SocialCheckersRoom.mode == "human",
            SocialCheckersRoom.is_public.is_(True),
            SocialCheckersRoom.status == "waiting",
        )
        .order_by(SocialCheckersRoom.updated_at.desc(), SocialCheckersRoom.id.desc())
        .limit(24)
    ).all()
    my_rooms = db.scalars(
        select(SocialCheckersRoom)
        .where(
            SocialCheckersRoom.game_code == BATTLESHIP_GAME_CODE,
            or_(
                SocialCheckersRoom.host_actor_key.in_(aliases),
                SocialCheckersRoom.guest_actor_key.in_(aliases),
            ),
            SocialCheckersRoom.status.in_(["waiting", "active", "finished", "cancelled"]),
        )
        .order_by(SocialCheckersRoom.updated_at.desc(), SocialCheckersRoom.id.desc())
        .limit(12)
    ).all()
    leaderboard_rows, my_rank, my_rating = _social_board_rank_rows(db, actor_key, game_code=BATTLESHIP_GAME_CODE, limit=20)
    db.commit()
    return {
        "profile": {
            **_social_checkers_profile_out(profile_row, actor_key=actor_key),
            "nick": actor_nick,
            "member_id": int(member_id or 0),
        },
        "leaderboard": {
            "rows": leaderboard_rows,
            "my_rank": my_rank,
            "my_rating": my_rating,
        },
        "rooms": {
            "public": [_social_battleship_room_payload(db, row, actor_key) for row in public_rooms],
            "mine": [_social_battleship_room_payload(db, row, actor_key) for row in my_rooms],
        },
        "difficulties": get_battleship_difficulties(),
    }


@router.get("/social/games/battleship/leaderboard", response_model=dict[str, Any])
def social_battleship_leaderboard(
    limit: int = 100,
    user: User = Depends(get_current_user),
    db: Session = Depends(get_db),
):
    ensure_module_enabled(db, user, "social_hub")
    actor_key, _, _, profile_row = _social_board_my_profile(db, user, BATTLESHIP_GAME_CODE)
    rows, my_rank, my_rating = _social_board_rank_rows(db, actor_key, game_code=BATTLESHIP_GAME_CODE, limit=limit)
    db.commit()
    return {
        "rows": rows,
        "my_rank": my_rank,
        "my_rating": my_rating,
        "my_profile": _social_checkers_profile_out(profile_row, actor_key=actor_key),
    }


@router.post("/social/games/battleship/rooms", response_model=dict[str, Any])
def social_battleship_create_room(
    payload: dict[str, Any],
    request: Request,
    user: User = Depends(get_current_user),
    db: Session = Depends(get_db),
):
    ensure_module_enabled(db, user, "social_hub")
    actor_key, actor_nick, member_id, _ = _social_board_my_profile(db, user, BATTLESHIP_GAME_CODE)
    mode = str((payload or {}).get("mode") or "human").strip().lower()
    if mode not in {"human", "bot"}:
        raise HTTPException(status_code=400, detail="РќРµРєРѕСЂСЂРµРєС‚РЅС‹Р№ СЂРµР¶РёРј РєРѕРјРЅР°С‚С‹")
    difficulty = str((payload or {}).get("difficulty") or "medium").strip().lower()
    bot_meta = build_battleship_bot_identity(difficulty)
    safe_title = str((payload or {}).get("title") or "").strip()[:120]
    title = safe_title or (f"РљРѕРјРЅР°С‚Р° {actor_nick}" if mode == "human" else f"РњР°С‚С‡ СЃ {bot_meta['title']}")
    is_public = bool((payload or {}).get("is_public", True)) if mode == "human" else False
    existing_codes = set(db.scalars(select(SocialCheckersRoom.room_code)).all())
    state = create_battleship_state(include_black=(mode == "bot"))
    room = SocialCheckersRoom(
        game_code=BATTLESHIP_GAME_CODE,
        room_code=create_checkers_room_code(existing_codes),
        title=title,
        owner_user_id=int(user.id),
        host_user_id=int(user.id),
        host_member_id=int(member_id or 0) or None,
        host_actor_key=_social_canonical_actor_key(db, actor_key),
        host_nick=actor_nick[:120],
        guest_user_id=None,
        guest_member_id=None,
        guest_actor_key=str(bot_meta["actor_key"]) if mode == "bot" else "",
        guest_nick=str(bot_meta["nick"])[:120] if mode == "bot" else "",
        mode=mode,
        difficulty=str(bot_meta["difficulty"]),
        is_public=bool(is_public),
        status="active" if mode == "bot" else "waiting",
        state_json=json.dumps(load_battleship_state(state), ensure_ascii=False),
        stats_applied=False,
        last_move_at=None,
        finished_at=None,
    )
    db.add(room)
    db.flush()
    _audit(
        db,
        user,
        action="social_battleship_room_created",
        details=json.dumps({"room_id": int(room.id), "mode": mode, "title": title}, ensure_ascii=False),
        module_code="social_hub",
        entity_type="battleship_room",
        entity_id=str(room.id),
        request=request,
    )
    db.commit()
    db.refresh(room)
    return _social_battleship_room_payload(db, room, actor_key)


@router.get("/social/games/battleship/rooms/{room_id}", response_model=dict[str, Any])
def social_battleship_get_room(
    room_id: int,
    user: User = Depends(get_current_user),
    db: Session = Depends(get_db),
):
    ensure_module_enabled(db, user, "social_hub")
    actor_key, _, _, _ = _social_board_my_profile(db, user, BATTLESHIP_GAME_CODE)
    room = db.get(SocialCheckersRoom, int(room_id or 0))
    if not room or str(room.game_code or "") != BATTLESHIP_GAME_CODE:
        raise HTTPException(status_code=404, detail="РљРѕРјРЅР°С‚Р° РЅРµ РЅР°Р№РґРµРЅР°")
    if not _social_checkers_room_visible(db, room, actor_key):
        raise HTTPException(status_code=403, detail="РќРµС‚ РґРѕСЃС‚СѓРїР° Рє РєРѕРјРЅР°С‚Рµ")
    db.commit()
    return _social_battleship_room_payload(db, room, actor_key)


@router.post("/social/games/battleship/rooms/{room_id}/join", response_model=dict[str, Any])
def social_battleship_join_room(
    room_id: int,
    request: Request,
    user: User = Depends(get_current_user),
    db: Session = Depends(get_db),
):
    ensure_module_enabled(db, user, "social_hub")
    actor_key, actor_nick, member_id, _ = _social_board_my_profile(db, user, BATTLESHIP_GAME_CODE)
    room = db.get(SocialCheckersRoom, int(room_id or 0))
    if not room or str(room.game_code or "") != BATTLESHIP_GAME_CODE:
        raise HTTPException(status_code=404, detail="РљРѕРјРЅР°С‚Р° РЅРµ РЅР°Р№РґРµРЅР°")
    if str(room.mode or "") != "human":
        raise HTTPException(status_code=400, detail="Рљ Р±РѕС‚Сѓ РїРѕРґРєР»СЋС‡РµРЅРёРµ РЅРµ С‚СЂРµР±СѓРµС‚СЃСЏ")
    if str(room.status or "") != "waiting":
        raise HTTPException(status_code=409, detail="РљРѕРјРЅР°С‚Р° СѓР¶Рµ Р·Р°РЅСЏС‚Р°")
    if not bool(room.is_public):
        raise HTTPException(status_code=403, detail="РљРѕРјРЅР°С‚Р° Р·Р°РєСЂС‹С‚Р° РґР»СЏ РїРѕРґРєР»СЋС‡РµРЅРёСЏ")
    if _social_checkers_room_side(db, room, actor_key) == "white":
        raise HTTPException(status_code=400, detail="РќРµР»СЊР·СЏ РїРѕРґРєР»СЋС‡РёС‚СЊСЃСЏ Рє СЃРІРѕРµР№ РєРѕРјРЅР°С‚Рµ")
    room.guest_user_id = int(user.id)
    room.guest_member_id = int(member_id or 0) or None
    room.guest_actor_key = _social_canonical_actor_key(db, actor_key)
    room.guest_nick = actor_nick[:120]
    state = _social_battleship_room_state(room)
    state = assign_battleship_side(state, "black")
    _social_battleship_store_room_state(room, state)
    room.status = "active"
    room.updated_at = datetime.utcnow()
    _audit(
        db,
        user,
        action="social_battleship_room_joined",
        details=json.dumps({"room_id": int(room.id), "room_code": str(room.room_code or "")}, ensure_ascii=False),
        module_code="social_hub",
        entity_type="battleship_room",
        entity_id=str(room.id),
        request=request,
    )
    db.commit()
    db.refresh(room)
    return _social_battleship_room_payload(db, room, actor_key)


@router.post("/social/games/battleship/rooms/{room_id}/move", response_model=dict[str, Any])
def social_battleship_make_move(
    room_id: int,
    payload: dict[str, Any],
    request: Request,
    user: User = Depends(get_current_user),
    db: Session = Depends(get_db),
):
    ensure_module_enabled(db, user, "social_hub")
    actor_key, _, _, _ = _social_board_my_profile(db, user, BATTLESHIP_GAME_CODE)
    room = db.get(SocialCheckersRoom, int(room_id or 0))
    if not room or str(room.game_code or "") != BATTLESHIP_GAME_CODE:
        raise HTTPException(status_code=404, detail="РљРѕРјРЅР°С‚Р° РЅРµ РЅР°Р№РґРµРЅР°")
    my_side = _social_checkers_room_side(db, room, actor_key)
    if not my_side:
        raise HTTPException(status_code=403, detail="РќРµС‚ РґРѕСЃС‚СѓРїР° Рє РїР°СЂС‚РёРё")
    if str(room.status or "") != "active":
        raise HTTPException(status_code=409, detail="РљРѕРјРЅР°С‚Р° СЃРµР№С‡Р°СЃ РЅРµР°РєС‚РёРІРЅР°")
    state = _social_battleship_room_state(room)
    if str(state.get("turn") or "") != my_side:
        raise HTTPException(status_code=409, detail="РЎРµР№С‡Р°СЃ С…РѕРґ РґСЂСѓРіРѕРіРѕ РёРіСЂРѕРєР°")
    try:
        row = int((payload or {}).get("row"))
        col = int((payload or {}).get("col"))
    except Exception:
        raise HTTPException(status_code=400, detail="РќРµРєРѕСЂСЂРµРєС‚РЅР°СЏ С†РµР»СЊ")
    try:
        next_state = apply_battleship_shot(state, my_side, row, col)
    except ValueError as exc:
        raise HTTPException(status_code=400, detail=str(exc))

    if (
        str(room.mode or "") == "bot"
        and not str(next_state.get("winner") or "")
        and str(next_state.get("result") or "") != "draw"
        and str(next_state.get("turn") or "") == "black"
    ):
        bot_move = pick_battleship_bot_move(next_state, room.difficulty)
        if bot_move and isinstance(bot_move, dict):
            next_state = apply_battleship_shot(
                next_state,
                "black",
                int(bot_move.get("row") or 0),
                int(bot_move.get("col") or 0),
            )

    move_stamp = _to_utc_iso(datetime.utcnow())
    if next_state.get("last_move"):
        next_state["last_move"]["at"] = move_stamp
    if next_state.get("history"):
        next_state["history"][-1]["at"] = move_stamp
    _social_battleship_store_room_state(room, next_state)
    room.last_move_at = datetime.utcnow()
    if str(next_state.get("winner") or "") or str(next_state.get("result") or "") == "draw":
        room.status = "finished"
        room.finished_at = datetime.utcnow()
        _social_board_apply_stats(
            db,
            room,
            next_state,
            game_code=BATTLESHIP_GAME_CODE,
            bot_builder=build_battleship_bot_identity,
        )
    move_marker = str(len(next_state.get("history") or []))
    _social_emit_game_turn_notification(
        db,
        room=room,
        game_code=BATTLESHIP_GAME_CODE,
        mover_actor_key=actor_key,
        mover_nick=_social_current_nick_by_key(db, actor_key) or actor_key,
        move_marker=move_marker,
        move_at_iso=move_stamp,
    )
    _audit(
        db,
        user,
        action="social_battleship_move",
        details=json.dumps({"room_id": int(room.id), "side": my_side, "target": [row, col]}, ensure_ascii=False),
        module_code="social_hub",
        entity_type="battleship_room",
        entity_id=str(room.id),
        request=request,
    )
    db.commit()
    db.refresh(room)
    return _social_battleship_room_payload(db, room, actor_key)


@router.post("/social/games/battleship/rooms/{room_id}/leave", response_model=dict[str, Any])
def social_battleship_leave_room(
    room_id: int,
    request: Request,
    user: User = Depends(get_current_user),
    db: Session = Depends(get_db),
):
    ensure_module_enabled(db, user, "social_hub")
    actor_key, _, _, _ = _social_board_my_profile(db, user, BATTLESHIP_GAME_CODE)
    room = db.get(SocialCheckersRoom, int(room_id or 0))
    if not room or str(room.game_code or "") != BATTLESHIP_GAME_CODE:
        raise HTTPException(status_code=404, detail="РљРѕРјРЅР°С‚Р° РЅРµ РЅР°Р№РґРµРЅР°")
    my_side = _social_checkers_room_side(db, room, actor_key)
    if not my_side:
        raise HTTPException(status_code=403, detail="РќРµС‚ РґРѕСЃС‚СѓРїР° Рє РїР°СЂС‚РёРё")
    state = _social_battleship_room_state(room)
    if str(room.status or "") == "waiting":
        room.status = "cancelled"
        state["result"] = "cancelled"
        state["winner"] = ""
    elif str(room.status or "") == "active":
        winner = "black" if my_side == "white" else "white"
        state["winner"] = winner
        state["result"] = "resigned"
        state["last_move"] = {
            "side": my_side,
            "target": None,
            "hit": False,
            "sunk": False,
            "at": _to_utc_iso(datetime.utcnow()),
        }
        room.status = "finished"
        room.finished_at = datetime.utcnow()
        _social_board_apply_stats(
            db,
            room,
            state,
            game_code=BATTLESHIP_GAME_CODE,
            bot_builder=build_battleship_bot_identity,
        )
    _social_battleship_store_room_state(room, state)
    room.last_move_at = datetime.utcnow()
    _audit(
        db,
        user,
        action="social_battleship_room_left",
        details=json.dumps({"room_id": int(room.id), "side": my_side, "status": str(room.status or "")}, ensure_ascii=False),
        module_code="social_hub",
        entity_type="battleship_room",
        entity_id=str(room.id),
        request=request,
    )
    db.commit()
    db.refresh(room)
    return _social_battleship_room_payload(db, room, actor_key)


@router.get("/social/chat/threads", response_model=list[SocialChatThreadOut])
def social_chat_threads(
    user: User = Depends(get_current_user),
    db: Session = Depends(get_db),
):
    ensure_module_enabled(db, user, "social_hub")
    actor_key, actor_nick, member_id = _social_actor_identity(db, user)
    company_thread = _social_ensure_company_thread(db, user.id)
    _social_sync_company_thread_members(
        db,
        thread=company_thread,
        user_id=int(user.id),
        actor_key=actor_key,
        actor_nick=actor_nick,
    )
    actor_aliases = _social_actor_alias_keys(db, actor_key)
    mine_rows = db.scalars(
        select(SocialChatThreadMember)
        .where(SocialChatThreadMember.actor_key.in_(actor_aliases))
        .order_by(SocialChatThreadMember.updated_at.desc(), SocialChatThreadMember.id.desc())
    ).all()
    # Migrate legacy owner alias memberships (m:<owner_member_id>) into canonical u:<user_id>.
    by_thread: dict[int, SocialChatThreadMember] = {}
    for row in mine_rows:
        tid = int(row.thread_id or 0)
        if tid <= 0:
            continue
        prev = by_thread.get(tid)
        is_exact = str(row.actor_key or "").strip().lower() == actor_key
        if prev is None or is_exact:
            by_thread[tid] = row
    mine: list[SocialChatThreadMember] = []
    for tid, row in by_thread.items():
        if str(row.actor_key or "").strip().lower() == actor_key:
            mine.append(row)
            continue
        migrated = _social_ensure_thread_member(
            db,
            thread_id=int(tid),
            actor_key=actor_key,
            user_id=int(user.id),
            member_id=member_id,
            actor_nick=actor_nick,
        )
        if int(migrated.last_read_message_id or 0) < int(row.last_read_message_id or 0):
            migrated.last_read_message_id = int(row.last_read_message_id or 0)
        db.delete(row)
        mine.append(migrated)
    out: list[SocialChatThreadOut] = []
    for member_row in mine:
        thread = db.get(SocialChatThread, member_row.thread_id)
        if not thread:
            continue
        if str(thread.kind or "") == "global":
            continue
        if thread.kind in {"company", "group"} and int(thread.owner_user_id or 0) != int(user.id):
            continue
        out.append(_social_thread_to_out(db, actor_key, thread, member_row))
    out.sort(key=lambda x: (x.unread, x.last_message.get("id", 0)), reverse=True)
    db.commit()
    return out


@router.get("/social/chat/participant/profile", response_model=dict[str, Any])
def social_chat_participant_profile(
    thread_id: int,
    actor_key: str,
    user: User = Depends(get_current_user),
    db: Session = Depends(get_db),
):
    ensure_module_enabled(db, user, "social_hub")
    safe_thread_id = int(thread_id or 0)
    safe_actor_key = _social_canonical_actor_key(db, str(actor_key or "").strip().lower())
    if safe_thread_id <= 0 or not safe_actor_key:
        raise HTTPException(status_code=400, detail="РќРµРєРѕСЂСЂРµРєС‚РЅС‹Рµ РїР°СЂР°РјРµС‚СЂС‹")
    viewer_key, _, _ = _social_actor_identity(db, user)
    viewer_aliases = _social_actor_alias_keys(db, viewer_key)
    has_access = db.scalar(
        select(SocialChatThreadMember.id).where(
            SocialChatThreadMember.thread_id == safe_thread_id,
            SocialChatThreadMember.actor_key.in_(viewer_aliases),
        )
    )
    if not has_access:
        raise HTTPException(status_code=403, detail="РќРµС‚ РґРѕСЃС‚СѓРїР° Рє С‡Р°С‚Сѓ")
    target_aliases = _social_actor_alias_keys(db, safe_actor_key)
    target_row = db.scalar(
        select(SocialChatThreadMember).where(
            SocialChatThreadMember.thread_id == safe_thread_id,
            SocialChatThreadMember.actor_key.in_(target_aliases),
        )
    )
    if not target_row:
        raise HTTPException(status_code=404, detail="РЈС‡Р°СЃС‚РЅРёРє РЅРµ РЅР°Р№РґРµРЅ")

    target_key = _social_canonical_actor_key(db, str(target_row.actor_key or "").strip().lower())
    profile_payload = _social_public_profile_by_key(db, target_key)
    activity_map = _social_last_activity_map(db, [target_key])
    last_seen = activity_map.get(target_key)
    profile_payload["last_seen_at"] = _to_utc_iso(last_seen)
    profile_payload["is_online"] = bool(_social_is_online(last_seen))
    profile_payload["can_edit"] = False
    profile_payload["is_me"] = target_key == viewer_key
    return profile_payload


@router.put("/social/chat/threads/{thread_id}/avatar", response_model=SocialChatThreadOut)
def social_chat_thread_avatar(
    thread_id: int,
    payload: SocialChatThreadAvatarIn,
    user: User = Depends(get_current_user),
    db: Session = Depends(get_db),
):
    ensure_module_enabled(db, user, "social_hub")
    actor_key, actor_nick, member_id = _social_actor_identity(db, user)
    thread = db.get(SocialChatThread, int(thread_id or 0))
    if not thread:
        raise HTTPException(status_code=404, detail="Р§Р°С‚ РЅРµ РЅР°Р№РґРµРЅ")
    if thread.kind == "direct":
        raise HTTPException(status_code=400, detail="РђРІР°С‚Р°СЂ РґРѕСЃС‚СѓРїРµРЅ С‚РѕР»СЊРєРѕ РґР»СЏ РіСЂСѓРїРїРѕРІС‹С… С‡Р°С‚РѕРІ")
    if thread.kind == "global" and user.role != "admin":
        raise HTTPException(status_code=403, detail="РќРµРґРѕСЃС‚Р°С‚РѕС‡РЅРѕ РїСЂР°РІ")
    if thread.kind == "company" and int(thread.owner_user_id or 0) != int(user.id) and user.role != "admin":
        raise HTTPException(status_code=403, detail="РќРµРґРѕСЃС‚Р°С‚РѕС‡РЅРѕ РїСЂР°РІ")
    if thread.kind == "group":
        if int(thread.owner_user_id or 0) != int(user.id):
            raise HTTPException(status_code=403, detail="РќРµРґРѕСЃС‚Р°С‚РѕС‡РЅРѕ РїСЂР°РІ")
        member_row = db.scalar(
            select(SocialChatThreadMember).where(
                SocialChatThreadMember.thread_id == int(thread.id),
                SocialChatThreadMember.actor_key == actor_key,
            )
        )
        if not member_row:
            raise HTTPException(status_code=403, detail="РќРµС‚ РґРѕСЃС‚СѓРїР° Рє РіСЂСѓРїРїРµ")
    elif thread.kind == "company":
        member_row = _social_sync_company_thread_members(
            db,
            thread=thread,
            user_id=int(user.id),
            actor_key=actor_key,
            actor_nick=actor_nick,
        )
    else:
        member_row = _social_ensure_thread_member(
            db,
            thread_id=thread.id,
            actor_key=actor_key,
            user_id=user.id,
            member_id=member_id,
            actor_nick=actor_nick,
        )
    safe_url = str(payload.avatar_url or "").strip()[:500]
    thread.avatar_url = safe_url
    thread.updated_at = datetime.utcnow()
    _audit(
        db,
        user,
        action="social_thread_avatar_updated",
        details=json.dumps({"thread_id": int(thread.id), "avatar_url": safe_url}, ensure_ascii=False),
        module_code="social_hub",
        entity_type="social_thread",
        entity_id=str(thread.id),
    )
    db.commit()
    return _social_thread_to_out(db, actor_key, thread, member_row)


def _social_chat_upload_thread_avatar(
    *,
    thread_id: int,
    file: UploadFile,
    user: User,
    db: Session,
) -> SocialChatThreadOut:
    ensure_module_enabled(db, user, "social_hub")
    actor_key, actor_nick, _ = _social_actor_identity(db, user)
    thread = db.get(SocialChatThread, int(thread_id or 0))
    if not thread:
        raise HTTPException(status_code=404, detail="Р§Р°С‚ РЅРµ РЅР°Р№РґРµРЅ")
    kind = str(thread.kind or "")
    if kind not in {"group", "company"}:
        raise HTTPException(status_code=400, detail="Р—Р°РіСЂСѓР·РєР° Р°РІР°С‚Р°СЂР° РґРѕСЃС‚СѓРїРЅР° С‚РѕР»СЊРєРѕ РґР»СЏ С‡Р°С‚Р° РєРѕРјРїР°РЅРёРё Рё РіСЂСѓРїРї")

    if kind == "group":
        if int(thread.owner_user_id or 0) != int(user.id):
            raise HTTPException(status_code=403, detail="РќРµРґРѕСЃС‚Р°С‚РѕС‡РЅРѕ РїСЂР°РІ")
        member_row = db.scalar(
            select(SocialChatThreadMember).where(
                SocialChatThreadMember.thread_id == int(thread.id),
                SocialChatThreadMember.actor_key == actor_key,
            )
        )
        if not member_row:
            raise HTTPException(status_code=403, detail="РќРµС‚ РґРѕСЃС‚СѓРїР° Рє РіСЂСѓРїРїРµ")
        prefix = f"group{int(thread.id)}"
        audit_action = "social_group_avatar_uploaded"
    else:
        if int(thread.owner_user_id or 0) != int(user.id) and user.role != "admin":
            raise HTTPException(status_code=403, detail="РќРµРґРѕСЃС‚Р°С‚РѕС‡РЅРѕ РїСЂР°РІ")
        member_row = _social_sync_company_thread_members(
            db,
            thread=thread,
            user_id=int(user.id),
            actor_key=actor_key,
            actor_nick=actor_nick,
        )
        prefix = f"company{int(thread.id)}"
        audit_action = "social_company_chat_avatar_uploaded"

    url = _save_avatar_upload(file, user_id=user.id, prefix=prefix)
    thread.avatar_url = str(url or "")
    thread.updated_at = datetime.utcnow()
    _audit(
        db,
        user,
        action=audit_action,
        details=json.dumps({"thread_id": int(thread.id), "avatar_url": str(url or "")}, ensure_ascii=False),
        module_code="social_hub",
        entity_type="social_thread",
        entity_id=str(thread.id),
    )
    db.commit()
    return _social_thread_to_out(db, actor_key, thread, member_row)


@router.post("/social/chat/threads/{thread_id}/avatar/upload", response_model=SocialChatThreadOut)
def social_chat_thread_avatar_upload(
    thread_id: int,
    file: UploadFile = File(...),
    user: User = Depends(get_current_user),
    db: Session = Depends(get_db),
):
    return _social_chat_upload_thread_avatar(
        thread_id=thread_id,
        file=file,
        user=user,
        db=db,
    )


@router.post("/social/chat/groups/{thread_id}/avatar/upload", response_model=SocialChatThreadOut)
def social_chat_group_avatar_upload(
    thread_id: int,
    file: UploadFile = File(...),
    user: User = Depends(get_current_user),
    db: Session = Depends(get_db),
):
    return _social_chat_upload_thread_avatar(
        thread_id=thread_id,
        file=file,
        user=user,
        db=db,
    )


@router.put("/social/chat/company/{thread_id}", response_model=SocialChatThreadOut)
def social_update_company_chat(
    thread_id: int,
    payload: SocialChatCompanyUpdateIn,
    request: Request,
    user: User = Depends(get_current_user),
    db: Session = Depends(get_db),
):
    ensure_module_enabled(db, user, "social_hub")
    actor_key, actor_nick, _ = _social_actor_identity(db, user)
    thread = db.get(SocialChatThread, int(thread_id or 0))
    if not thread or str(thread.kind or "") != "company":
        raise HTTPException(status_code=404, detail="Р§Р°С‚ РєРѕРјРїР°РЅРёРё РЅРµ РЅР°Р№РґРµРЅ")
    if int(thread.owner_user_id or 0) != int(user.id) and user.role != "admin":
        raise HTTPException(status_code=403, detail="РќРµРґРѕСЃС‚Р°С‚РѕС‡РЅРѕ РїСЂР°РІ")
    member_row = _social_sync_company_thread_members(
        db,
        thread=thread,
        user_id=int(user.id),
        actor_key=actor_key,
        actor_nick=actor_nick,
    )

    if payload.title is not None:
        title = str(payload.title or "").strip()[:255]
        if len(title) < 2:
            raise HTTPException(status_code=400, detail="РЈРєР°Р¶РёС‚Рµ РЅР°Р·РІР°РЅРёРµ С‡Р°С‚Р°")
        thread.title = title
    if payload.avatar_url is not None:
        thread.avatar_url = str(payload.avatar_url or "").strip()[:500]

    thread.updated_at = datetime.utcnow()
    _audit(
        db,
        user,
        action="social_company_chat_updated",
        details=json.dumps(
            {
                "thread_id": int(thread.id),
                "title": str(thread.title or ""),
                "avatar_url": str(thread.avatar_url or ""),
            },
            ensure_ascii=False,
        ),
        module_code="social_hub",
        entity_type="social_thread",
        entity_id=str(thread.id),
        request=request,
    )
    db.commit()
    return _social_thread_to_out(db, actor_key, thread, member_row)


@router.get("/social/currency/rates", response_model=SocialCurrencyRatesOut)
def social_currency_rates(
    user: User = Depends(get_current_user),
    db: Session = Depends(get_db),
):
    ensure_module_enabled(db, user, "social_hub")
    payload = _get_cbr_rates()
    return SocialCurrencyRatesOut(**payload)


@router.post("/social/chat/direct", response_model=SocialChatThreadOut)
def social_start_direct_chat(
    payload: SocialChatDirectStartIn,
    user: User = Depends(get_current_user),
    db: Session = Depends(get_db),
):
    ensure_module_enabled(db, user, "social_hub")
    actor_key, actor_nick, actor_member_id = _social_actor_identity(db, user)
    peer_key = _social_canonical_actor_key(db, str(payload.actor_key or "").strip().lower())
    if not peer_key or peer_key == actor_key:
        raise HTTPException(status_code=400, detail="Р’С‹Р±РµСЂРёС‚Рµ СЃРѕР±РµСЃРµРґРЅРёРєР°")
    peer_user_id, peer_member_id, peer_nick = _social_identity_by_key(db, peer_key)
    actor_aliases = _social_actor_alias_keys(db, actor_key)
    peer_aliases = _social_actor_alias_keys(db, peer_key)
    peer_thread_ids = select(SocialChatThreadMember.thread_id).where(SocialChatThreadMember.actor_key.in_(peer_aliases))
    actor_thread_ids = select(SocialChatThreadMember.thread_id).where(SocialChatThreadMember.actor_key.in_(actor_aliases))
    thread = db.scalar(
        select(SocialChatThread).where(
            SocialChatThread.kind == "direct",
            SocialChatThread.id.in_(actor_thread_ids),
            SocialChatThread.id.in_(peer_thread_ids),
        )
        .order_by(SocialChatThread.id.desc())
    )
    if not thread:
        thread = db.scalar(
            select(SocialChatThread)
            .join(SocialChatThreadMember, SocialChatThreadMember.thread_id == SocialChatThread.id)
            .where(
                SocialChatThread.kind == "direct",
                SocialChatThreadMember.actor_key == actor_key,
                SocialChatThread.id.in_(peer_thread_ids),
            )
            .order_by(SocialChatThread.id.desc())
        )
    if not thread:
        thread = db.scalar(
            select(SocialChatThread)
            .join(SocialChatThreadMember, SocialChatThreadMember.thread_id == SocialChatThread.id)
            .where(
                SocialChatThread.kind == "direct",
                SocialChatThreadMember.actor_key == peer_key,
                SocialChatThread.id.in_(actor_thread_ids),
            )
            .order_by(SocialChatThread.id.desc())
        )
    if not thread:
        thread = SocialChatThread(
            kind="direct",
            owner_user_id=None,
            title=f"Р›РёС‡РЅС‹Р№ С‡Р°С‚: {actor_nick} в†” {peer_nick}",
        )
        db.add(thread)
        db.flush()
    me_member = _social_ensure_thread_member(
        db,
        thread_id=thread.id,
        actor_key=actor_key,
        user_id=user.id,
        member_id=actor_member_id,
        actor_nick=actor_nick,
    )
    peer_member = _social_ensure_thread_member(
        db,
        thread_id=thread.id,
        actor_key=peer_key,
        user_id=peer_user_id,
        member_id=peer_member_id,
        actor_nick=peer_nick,
    )
    for alias in peer_aliases:
        if alias != peer_key:
            row = db.scalar(
                select(SocialChatThreadMember).where(
                    SocialChatThreadMember.thread_id == int(thread.id),
                    SocialChatThreadMember.actor_key == alias,
                )
            )
            if row:
                if int(peer_member.last_read_message_id or 0) < int(row.last_read_message_id or 0):
                    peer_member.last_read_message_id = int(row.last_read_message_id or 0)
                db.delete(row)
    for alias in actor_aliases:
        if alias != actor_key:
            row = db.scalar(
                select(SocialChatThreadMember).where(
                    SocialChatThreadMember.thread_id == int(thread.id),
                    SocialChatThreadMember.actor_key == alias,
                )
            )
            if row:
                if int(me_member.last_read_message_id or 0) < int(row.last_read_message_id or 0):
                    me_member.last_read_message_id = int(row.last_read_message_id or 0)
                db.delete(row)
    db.commit()
    return _social_thread_to_out(db, actor_key, thread, me_member)


def _social_sync_group_members(
    db: Session,
    *,
    thread: SocialChatThread,
    user_id: int,
    actor_key: str,
    actor_nick: str,
    member_keys: list[str],
) -> SocialChatThreadMember:
    me_row: SocialChatThreadMember | None = None
    keep_keys = set(member_keys)
    existing_rows = db.scalars(
        select(SocialChatThreadMember).where(SocialChatThreadMember.thread_id == int(thread.id))
    ).all()
    for key in member_keys:
        uid, mid, nick = _social_identity_by_key(db, key)
        if int(uid) != int(user_id):
            continue
        row = _social_ensure_thread_member(
            db,
            thread_id=int(thread.id),
            actor_key=key,
            user_id=int(uid),
            member_id=int(mid) if mid else None,
            actor_nick=(actor_nick if key == actor_key else nick),
        )
        if key == actor_key:
            me_row = row
    for row in existing_rows:
        key = str(row.actor_key or "").strip().lower()
        if key not in keep_keys:
            db.delete(row)
    if not me_row:
        me_row = _social_ensure_thread_member(
            db,
            thread_id=int(thread.id),
            actor_key=actor_key,
            user_id=int(user_id),
            member_id=_to_int_safe(actor_key.split(":", 1)[1]) if actor_key.startswith("m:") else None,
            actor_nick=actor_nick,
        )
    thread.updated_at = datetime.utcnow()
    return me_row


def _social_company_member_keys(
    db: Session,
    *,
    user_id: int,
    actor_key: str,
) -> list[str]:
    allowed = sorted(_social_company_allowed_actor_keys(db, int(user_id)))
    return _social_clean_group_member_keys(
        db,
        user_id=int(user_id),
        actor_key=actor_key,
        member_keys=allowed,
    )


def _social_sync_company_thread_members(
    db: Session,
    *,
    thread: SocialChatThread,
    user_id: int,
    actor_key: str,
    actor_nick: str,
) -> SocialChatThreadMember:
    member_keys = _social_company_member_keys(
        db,
        user_id=int(user_id),
        actor_key=actor_key,
    )
    return _social_sync_group_members(
        db,
        thread=thread,
        user_id=int(user_id),
        actor_key=actor_key,
        actor_nick=actor_nick,
        member_keys=member_keys,
    )


@router.post("/social/chat/groups", response_model=SocialChatThreadOut)
def social_create_group_chat(
    payload: SocialChatGroupIn,
    request: Request,
    user: User = Depends(get_current_user),
    db: Session = Depends(get_db),
):
    ensure_module_enabled(db, user, "social_hub")
    actor_key, actor_nick, _ = _social_actor_identity(db, user)
    title = str(payload.title or "").strip()[:255]
    if len(title) < 2:
        raise HTTPException(status_code=400, detail="РЈРєР°Р¶РёС‚Рµ РЅР°Р·РІР°РЅРёРµ РіСЂСѓРїРїС‹")
    member_keys = _social_clean_group_member_keys(
        db,
        user_id=int(user.id),
        actor_key=actor_key,
        member_keys=list(payload.member_keys or []),
    )
    if len(member_keys) < 2:
        raise HTTPException(status_code=400, detail="Р”РѕР±Р°РІСЊС‚Рµ РјРёРЅРёРјСѓРј РґРІСѓС… СѓС‡Р°СЃС‚РЅРёРєРѕРІ")
    thread = SocialChatThread(
        kind="group",
        owner_user_id=int(user.id),
        title=title,
        avatar_url=str(payload.avatar_url or "").strip()[:500],
    )
    db.add(thread)
    db.flush()
    me_row = _social_sync_group_members(
        db,
        thread=thread,
        user_id=int(user.id),
        actor_key=actor_key,
        actor_nick=actor_nick,
        member_keys=member_keys,
    )
    _audit(
        db,
        user,
        action="social_group_chat_created",
        details=json.dumps(
            {"thread_id": int(thread.id), "title": title, "members": len(member_keys)},
            ensure_ascii=False,
        ),
        module_code="social_hub",
        entity_type="social_thread",
        entity_id=str(thread.id),
        request=request,
    )
    db.commit()
    return _social_thread_to_out(db, actor_key, thread, me_row)


@router.put("/social/chat/groups/{thread_id}", response_model=SocialChatThreadOut)
def social_update_group_chat(
    thread_id: int,
    payload: SocialChatGroupUpdateIn,
    request: Request,
    user: User = Depends(get_current_user),
    db: Session = Depends(get_db),
):
    ensure_module_enabled(db, user, "social_hub")
    actor_key, actor_nick, _ = _social_actor_identity(db, user)
    thread = db.get(SocialChatThread, int(thread_id or 0))
    if not thread or str(thread.kind or "") != "group" or int(thread.owner_user_id or 0) != int(user.id):
        raise HTTPException(status_code=404, detail="Р“СЂСѓРїРїР° РЅРµ РЅР°Р№РґРµРЅР°")
    me_member = db.scalar(
        select(SocialChatThreadMember).where(
            SocialChatThreadMember.thread_id == int(thread.id),
            SocialChatThreadMember.actor_key == actor_key,
        )
    )
    if not me_member:
        raise HTTPException(status_code=403, detail="РќРµС‚ РґРѕСЃС‚СѓРїР° Рє РіСЂСѓРїРїРµ")
    title = str(payload.title or "").strip()[:255] if payload.title is not None else str(thread.title or "")
    if len(title) < 2:
        raise HTTPException(status_code=400, detail="РЈРєР°Р¶РёС‚Рµ РЅР°Р·РІР°РЅРёРµ РіСЂСѓРїРїС‹")
    thread.title = title
    if payload.avatar_url is not None:
        thread.avatar_url = str(payload.avatar_url or "").strip()[:500]
    if payload.member_keys is not None:
        member_keys = _social_clean_group_member_keys(
            db,
            user_id=int(user.id),
            actor_key=actor_key,
            member_keys=list(payload.member_keys or []),
        )
        if len(member_keys) < 2:
            raise HTTPException(status_code=400, detail="Р’ РіСЂСѓРїРїРµ РґРѕР»Р¶РЅРѕ Р±С‹С‚СЊ РјРёРЅРёРјСѓРј РґРІР° СѓС‡Р°СЃС‚РЅРёРєР°")
        me_member = _social_sync_group_members(
            db,
            thread=thread,
            user_id=int(user.id),
            actor_key=actor_key,
            actor_nick=actor_nick,
            member_keys=member_keys,
        )
    thread.updated_at = datetime.utcnow()
    _audit(
        db,
        user,
        action="social_group_chat_updated",
        details=json.dumps(
            {
                "thread_id": int(thread.id),
                "title": str(thread.title or ""),
                "members": int(
                    db.scalar(
                        select(func.count())
                        .select_from(SocialChatThreadMember)
                        .where(SocialChatThreadMember.thread_id == int(thread.id))
                    )
                    or 0
                ),
            },
            ensure_ascii=False,
        ),
        module_code="social_hub",
        entity_type="social_thread",
        entity_id=str(thread.id),
        request=request,
    )
    db.commit()
    return _social_thread_to_out(db, actor_key, thread, me_member)


@router.delete("/social/chat/groups/{thread_id}", response_model=MessageOut)
def social_delete_group_chat(
    thread_id: int,
    request: Request,
    user: User = Depends(get_current_user),
    db: Session = Depends(get_db),
):
    ensure_module_enabled(db, user, "social_hub")
    actor_key, _, _ = _social_actor_identity(db, user)
    thread = db.get(SocialChatThread, int(thread_id or 0))
    if not thread or str(thread.kind or "") != "group" or int(thread.owner_user_id or 0) != int(user.id):
        raise HTTPException(status_code=404, detail="Р“СЂСѓРїРїР° РЅРµ РЅР°Р№РґРµРЅР°")
    me_member = db.scalar(
        select(SocialChatThreadMember).where(
            SocialChatThreadMember.thread_id == int(thread.id),
            SocialChatThreadMember.actor_key == actor_key,
        )
    )
    if not me_member:
        raise HTTPException(status_code=403, detail="РќРµС‚ РґРѕСЃС‚СѓРїР° Рє РіСЂСѓРїРїРµ")

    deleted_messages = db.execute(
        delete(SocialChatMessage).where(SocialChatMessage.thread_id == int(thread.id))
    ).rowcount or 0
    deleted_members = db.execute(
        delete(SocialChatThreadMember).where(SocialChatThreadMember.thread_id == int(thread.id))
    ).rowcount or 0

    db.delete(thread)
    _audit(
        db,
        user,
        action="social_group_chat_deleted",
        details=json.dumps(
            {
                "thread_id": int(thread.id),
                "title": str(thread.title or ""),
                "deleted_messages": int(deleted_messages),
                "deleted_members": int(deleted_members),
            },
            ensure_ascii=False,
        ),
        module_code="social_hub",
        entity_type="social_thread",
        entity_id=str(thread.id),
        request=request,
    )
    db.commit()
    return MessageOut(message="Р“СЂСѓРїРїР° СѓРґР°Р»РµРЅР°")


@router.get("/social/chat/search", response_model=dict[str, Any])
def social_chat_search(
    q: str = "",
    limit_threads: int = 250,
    user: User = Depends(get_current_user),
    db: Session = Depends(get_db),
):
    ensure_module_enabled(db, user, "social_hub")
    search = str(q or "").strip().lower()
    if len(search) < 2:
        return {"query": search, "thread_ids": [], "limited": False}

    actor_key, _, _ = _social_actor_identity(db, user)
    actor_aliases = _social_actor_alias_keys(db, actor_key)
    allowed_rows = db.scalars(
        select(SocialChatThread.id)
        .join(SocialChatThreadMember, SocialChatThreadMember.thread_id == SocialChatThread.id)
        .where(
            SocialChatThreadMember.actor_key.in_(actor_aliases),
            SocialChatThread.kind != "global",
            or_(
                ~SocialChatThread.kind.in_(["company", "group"]),
                SocialChatThread.owner_user_id == int(user.id),
            ),
        )
        .order_by(SocialChatThread.updated_at.desc(), SocialChatThread.id.desc())
        .limit(4000)
    ).all()

    allowed_thread_ids: list[int] = []
    seen_thread_ids: set[int] = set()
    for raw_tid in allowed_rows:
        tid = int(raw_tid or 0)
        if tid <= 0 or tid in seen_thread_ids:
            continue
        seen_thread_ids.add(tid)
        allowed_thread_ids.append(tid)

    if not allowed_thread_ids:
        return {"query": search, "thread_ids": [], "limited": False}

    safe_limit = max(20, min(int(limit_threads or 250), 600))
    search_pattern = f"%{search}%"
    match_rows = db.scalars(
        select(SocialChatMessage.thread_id)
        .where(
            SocialChatMessage.thread_id.in_(allowed_thread_ids),
            func.lower(cast(SocialChatMessage.text, String)).like(search_pattern),
        )
        .order_by(SocialChatMessage.id.desc())
        .limit(12000)
    ).all()

    matched_thread_ids: list[int] = []
    seen_matched_ids: set[int] = set()
    for raw_tid in match_rows:
        tid = int(raw_tid or 0)
        if tid <= 0 or tid in seen_matched_ids:
            continue
        seen_matched_ids.add(tid)
        matched_thread_ids.append(tid)
        if len(matched_thread_ids) >= safe_limit:
            break

    return {
        "query": search,
        "thread_ids": matched_thread_ids,
        "limited": len(matched_thread_ids) >= safe_limit,
    }

@router.get("/social/chat/actors", response_model=list[dict[str, Any]])
def social_chat_actor_directory(
    q: str = "",
    user: User = Depends(get_current_user),
    db: Session = Depends(get_db),
):
    ensure_module_enabled(db, user, "social_hub")
    search = str(q or "").strip().lower()
    allow_cross_company = "@" in search and len(search) >= 3
    owners = db.scalars(select(User).order_by(User.id.asc()).limit(5000)).all()
    members = db.scalars(
        select(TeamMember).where(TeamMember.is_active.is_(True)).order_by(TeamMember.id.asc()).limit(15000)
    ).all()
    out: list[dict[str, Any]] = []
    seen: set[str] = set()
    owner_has_member: set[int] = set()
    for row in members:
        if bool(row.is_owner):
            owner_has_member.add(int(row.user_id))
        if row.user_id != user.id and not allow_cross_company:
            continue
        base_key = f"u:{int(row.user_id)}" if bool(row.is_owner) else f"m:{int(row.id)}"
        key = _social_canonical_actor_key(db, base_key)
        nick = str((row.nickname or row.full_name or row.email).strip() or f"member-{row.id}")
        company = str((row.user.email if row.user else "") or "").strip().lower()
        if row.user_id != user.id:
            hay = f"{row.email or ''}".lower()
            if search and search not in hay:
                continue
        else:
            hay = f"{nick} {row.email or ''} {company}".lower()
            if search and search not in hay:
                continue
        if key in seen:
            continue
        seen.add(key)
        out.append({
            "actor_key": key,
            "nick": nick[:120],
            "company": company,
            "avatar_url": str(row.avatar_url or ""),
            "is_owner": bool(row.is_owner),
        })
    for row in owners:
        if int(row.id) in owner_has_member:
            continue
        if row.id != user.id and not allow_cross_company:
            continue
        key = f"u:{int(row.id)}"
        nick = str((row.email or "").split("@")[0] or f"user-{row.id}")
        if row.id != user.id:
            hay = f"{row.email or ''}".lower()
            if search and search not in hay:
                continue
        else:
            hay = f"{nick} {row.email or ''}".lower()
            if search and search not in hay:
                continue
        if key in seen:
            continue
        seen.add(key)
        out.append({
            "actor_key": key,
            "nick": nick[:120],
            "company": str(row.email or "").strip().lower(),
            "avatar_url": "",
            "is_owner": True,
        })
    return out[:500]


@router.get("/social/chat/messages/{thread_id}", response_model=list[SocialChatMessageOut])
def social_chat_messages(
    thread_id: int,
    before_id: int = 0,
    limit: int = 80,
    user: User = Depends(get_current_user),
    db: Session = Depends(get_db),
):
    ensure_module_enabled(db, user, "social_hub")
    actor_key, actor_nick, _ = _social_actor_identity(db, user)
    actor_aliases = _social_actor_alias_keys(db, actor_key)
    thread = db.get(SocialChatThread, int(thread_id or 0))
    if not thread:
        raise HTTPException(status_code=404, detail="Р§Р°С‚ РЅРµ РЅР°Р№РґРµРЅ")
    if str(thread.kind or "") == "global":
        raise HTTPException(status_code=410, detail="Р“Р»РѕР±Р°Р»СЊРЅС‹Р№ С‡Р°С‚ РѕС‚РєР»СЋС‡РµРЅ")
    if str(thread.kind or "") in {"company", "group"} and int(thread.owner_user_id or 0) != int(user.id):
        raise HTTPException(status_code=403, detail="РќРµС‚ РґРѕСЃС‚СѓРїР° Рє С‡Р°С‚Сѓ")
    if str(thread.kind or "") == "company":
        member_row = _social_sync_company_thread_members(
            db,
            thread=thread,
            user_id=int(user.id),
            actor_key=actor_key,
            actor_nick=actor_nick,
        )
    else:
        member_row = db.scalar(
            select(SocialChatThreadMember).where(
                SocialChatThreadMember.thread_id == int(thread.id),
                SocialChatThreadMember.actor_key.in_(actor_aliases),
            ).order_by(SocialChatThreadMember.id.asc())
        )
        if not member_row:
            raise HTTPException(status_code=403, detail="РќРµС‚ РґРѕСЃС‚СѓРїР° Рє С‡Р°С‚Сѓ")
    safe_limit = max(20, min(int(limit or 80), 200))
    query = select(SocialChatMessage).where(SocialChatMessage.thread_id == thread_id)
    if int(before_id or 0) > 0:
        query = query.where(SocialChatMessage.id < int(before_id))
    rows = db.scalars(query.order_by(SocialChatMessage.id.desc()).limit(safe_limit)).all()
    rows = list(reversed(rows))
    if rows:
        member_row.last_read_message_id = max(int(member_row.last_read_message_id or 0), int(rows[-1].id))
        member_row.updated_at = datetime.utcnow()
    db.commit()
    return [_social_message_to_out(db, actor_key, row) for row in rows]


@router.post("/social/chat/messages/{thread_id}", response_model=SocialChatMessageOut)
def social_chat_send_message(
    thread_id: int,
    payload: SocialChatMessageIn,
    request: Request,
    user: User = Depends(get_current_user),
    db: Session = Depends(get_db),
):
    ensure_module_enabled(db, user, "social_hub")
    actor_key, actor_nick, _ = _social_actor_identity(db, user)
    actor_aliases = _social_actor_alias_keys(db, actor_key)
    thread = db.get(SocialChatThread, int(thread_id or 0))
    if not thread:
        raise HTTPException(status_code=404, detail="Р§Р°С‚ РЅРµ РЅР°Р№РґРµРЅ")
    if str(thread.kind or "") == "global":
        raise HTTPException(status_code=410, detail="Р“Р»РѕР±Р°Р»СЊРЅС‹Р№ С‡Р°С‚ РѕС‚РєР»СЋС‡РµРЅ")
    if str(thread.kind or "") in {"company", "group"} and int(thread.owner_user_id or 0) != int(user.id):
        raise HTTPException(status_code=403, detail="РќРµС‚ РґРѕСЃС‚СѓРїР° Рє С‡Р°С‚Сѓ")
    if str(thread.kind or "") == "company":
        member_row = _social_sync_company_thread_members(
            db,
            thread=thread,
            user_id=int(user.id),
            actor_key=actor_key,
            actor_nick=actor_nick,
        )
    else:
        member_row = db.scalar(
            select(SocialChatThreadMember).where(
                SocialChatThreadMember.thread_id == thread_id,
                SocialChatThreadMember.actor_key.in_(actor_aliases),
            ).order_by(SocialChatThreadMember.id.asc())
        )
        if not member_row:
            raise HTTPException(status_code=403, detail="РќРµС‚ РґРѕСЃС‚СѓРїР° Рє С‡Р°С‚Сѓ")
    text_msg = str(payload.text or "").strip()
    if not text_msg:
        raise HTTPException(status_code=400, detail="РЎРѕРѕР±С‰РµРЅРёРµ РїСѓСЃС‚РѕРµ")
    request_id = str(request.headers.get("x-request-id") or "").strip()[:120]
    cache_key = ""
    if request_id:
        cache_key = f"{int(user.id)}:{int(thread_id)}:{request_id}"
        cached_id = _social_msg_cache_get(cache_key)
        if cached_id > 0:
            cached_msg = db.get(SocialChatMessage, int(cached_id))
            if cached_msg and int(cached_msg.sender_user_id or 0) == int(user.id):
                return _social_message_to_out(db, actor_key, cached_msg)
    reply_to_id = int(payload.reply_to_message_id or 0) if payload.reply_to_message_id else 0
    if reply_to_id > 0:
        reply_row = db.get(SocialChatMessage, reply_to_id)
        if not reply_row or int(reply_row.thread_id or 0) != int(thread_id):
            raise HTTPException(status_code=400, detail="РЎРѕРѕР±С‰РµРЅРёРµ РґР»СЏ РѕС‚РІРµС‚Р° РЅРµ РЅР°Р№РґРµРЅРѕ")
    message = SocialChatMessage(
        thread_id=thread_id,
        sender_user_id=user.id,
        sender_member_id=_actor_member_id(user) if not _actor_is_owner(user) else None,
        sender_key=actor_key,
        sender_nick=actor_nick[:120],
        text=text_msg[:5000],
        reply_to_message_id=reply_to_id or None,
        attachments_json="[]",
        reactions_json="{}",
    )
    db.add(message)
    db.flush()
    thread.updated_at = datetime.utcnow()
    member_row.last_read_message_id = int(message.id)
    member_row.updated_at = datetime.utcnow()
    recipients = db.scalars(
        select(SocialChatThreadMember).where(
            SocialChatThreadMember.thread_id == thread_id,
            SocialChatThreadMember.actor_key.notin_(actor_aliases),
        )
    ).all()
    thread_title = str(thread.title if thread else "Р§Р°С‚").strip() or "Р§Р°С‚"
    for rcpt in recipients:
        _social_push_notification(
            db,
            user_id=int(rcpt.user_id or 0),
            recipient_key=str(rcpt.actor_key or ""),
            kind="chat_message",
            dedupe_key=f"chat:{message.id}:{rcpt.actor_key}",
            title=f"РќРѕРІРѕРµ СЃРѕРѕР±С‰РµРЅРёРµ: {thread_title}",
            body=f"{actor_nick}: {text_msg[:180]}",
            payload={
                "thread_id": thread_id,
                "message_id": int(message.id),
                "sender_key": actor_key,
                "sender_nick": actor_nick,
            },
        )
    detail_payload = {
        "thread_id": int(thread_id),
        "message_id": int(message.id),
        "thread_title": thread_title[:120],
        "preview": text_msg[:500],
        "recipients": len(recipients),
    }
    _audit(
        db,
        user,
        action="social_chat_message_sent",
        details=json.dumps(detail_payload, ensure_ascii=False),
        module_code="social_hub",
        entity_type="chat_message",
        entity_id=str(message.id),
        request=request,
    )
    db.commit()
    if cache_key:
        _social_msg_cache_set(cache_key, int(message.id))
    return _social_message_to_out(db, actor_key, message)


@router.post("/social/chat/messages/{thread_id}/files", response_model=SocialChatMessageOut)
def social_chat_send_file(
    thread_id: int,
    request: Request,
    file: UploadFile = File(...),
    text: str = Form(""),
    reply_to_message_id: int | None = Form(None),
    user: User = Depends(get_current_user),
    db: Session = Depends(get_db),
):
    ensure_module_enabled(db, user, "social_hub")
    actor_key, actor_nick, _ = _social_actor_identity(db, user)
    actor_aliases = _social_actor_alias_keys(db, actor_key)
    thread = db.get(SocialChatThread, int(thread_id or 0))
    if not thread:
        raise HTTPException(status_code=404, detail="Р§Р°С‚ РЅРµ РЅР°Р№РґРµРЅ")
    if str(thread.kind or "") == "global":
        raise HTTPException(status_code=410, detail="Р“Р»РѕР±Р°Р»СЊРЅС‹Р№ С‡Р°С‚ РѕС‚РєР»СЋС‡РµРЅ")
    if str(thread.kind or "") in {"company", "group"} and int(thread.owner_user_id or 0) != int(user.id):
        raise HTTPException(status_code=403, detail="РќРµС‚ РґРѕСЃС‚СѓРїР° Рє С‡Р°С‚Сѓ")
    if str(thread.kind or "") == "company":
        member_row = _social_sync_company_thread_members(
            db,
            thread=thread,
            user_id=int(user.id),
            actor_key=actor_key,
            actor_nick=actor_nick,
        )
    else:
        member_row = db.scalar(
            select(SocialChatThreadMember).where(
                SocialChatThreadMember.thread_id == int(thread_id),
                SocialChatThreadMember.actor_key.in_(actor_aliases),
            ).order_by(SocialChatThreadMember.id.asc())
        )
        if not member_row:
            raise HTTPException(status_code=403, detail="РќРµС‚ РґРѕСЃС‚СѓРїР° Рє С‡Р°С‚Сѓ")
    if not file or not file.filename:
        raise HTTPException(status_code=400, detail="Р¤Р°Р№Р» РЅРµ РІС‹Р±СЂР°РЅ")
    raw = file.file.read()
    if not raw:
        raise HTTPException(status_code=400, detail="Р¤Р°Р№Р» РїСѓСЃС‚РѕР№")
    max_size = 25 * 1024 * 1024
    if len(raw) > max_size:
        limit_mb = max_size // (1024 * 1024)
        raise HTTPException(status_code=400, detail=f"File is too large (max {limit_mb} MB)")
    original_name = _social_chat_clean_filename(file.filename or "file")
    ext = _social_chat_guess_ext(original_name, str(file.content_type or ""))
    reply_id = int(reply_to_message_id or 0) if reply_to_message_id else 0
    if reply_id > 0:
        reply_row = db.get(SocialChatMessage, reply_id)
        if not reply_row or int(reply_row.thread_id or 0) != int(thread_id):
            raise HTTPException(status_code=400, detail="РЎРѕРѕР±С‰РµРЅРёРµ РґР»СЏ РѕС‚РІРµС‚Р° РЅРµ РЅР°Р№РґРµРЅРѕ")
    safe_text = str(text or "").strip()[:5000]
    request_id = str(request.headers.get("x-request-id") or "").strip()[:120]
    file_hash = hashlib.sha256(raw).hexdigest()[:24]
    payload_hash = hashlib.sha1(
        f"{safe_text}\n{original_name.lower()}".encode("utf-8", "ignore")
    ).hexdigest()[:20]
    dedupe_keys: list[str] = []
    if request_id:
        dedupe_keys.append(f"{int(user.id)}:{int(thread_id)}:file:req:{request_id}")
    dedupe_keys.append(
        f"{int(user.id)}:{int(thread_id)}:file:sig:{int(reply_id)}:{len(raw)}:{file_hash}:{payload_hash}"
    )
    for cache_key in dedupe_keys:
        cached_id = _social_msg_cache_get(cache_key)
        if cached_id <= 0:
            continue
        cached_msg = db.get(SocialChatMessage, int(cached_id))
        if cached_msg and int(cached_msg.sender_user_id or 0) == int(user.id):
            return _social_message_to_out(db, actor_key, cached_msg)
    storage_name = f"chat-{int(thread_id)}-{secrets.token_hex(8)}{ext}"
    path = _social_chat_storage_dir() / storage_name
    path.write_bytes(raw)
    url = f"/static/uploads/social_chat/{storage_name}"
    attachment = {
        "url": url,
        "filename": original_name[:255],
        "content_type": str(file.content_type or "application/octet-stream")[:120],
        "size_bytes": len(raw),
    }
    message = SocialChatMessage(
        thread_id=int(thread_id),
        sender_user_id=int(user.id),
        sender_member_id=_actor_member_id(user) if not _actor_is_owner(user) else None,
        sender_key=actor_key,
        sender_nick=actor_nick[:120],
        text=safe_text,
        reply_to_message_id=reply_id or None,
        attachments_json=json.dumps([attachment], ensure_ascii=False),
        reactions_json="{}",
    )
    db.add(message)
    db.flush()
    thread.updated_at = datetime.utcnow()
    member_row.last_read_message_id = int(message.id)
    member_row.updated_at = datetime.utcnow()
    recipients = db.scalars(
        select(SocialChatThreadMember).where(
            SocialChatThreadMember.thread_id == int(thread_id),
            SocialChatThreadMember.actor_key.notin_(actor_aliases),
        )
    ).all()
    thread_title = str(thread.title or "Р§Р°С‚").strip() or "Р§Р°С‚"
    preview = safe_text or f"рџ“Ћ {original_name}"
    for rcpt in recipients:
        _social_push_notification(
            db,
            user_id=int(rcpt.user_id or 0),
            recipient_key=str(rcpt.actor_key or ""),
            kind="chat_message",
            dedupe_key=f"chat:{message.id}:{rcpt.actor_key}",
            title=f"РќРѕРІРѕРµ СЃРѕРѕР±С‰РµРЅРёРµ: {thread_title}",
            body=f"{actor_nick}: {preview[:180]}",
            payload={
                "thread_id": int(thread_id),
                "message_id": int(message.id),
                "sender_key": actor_key,
                "sender_nick": actor_nick,
            },
        )
    _audit(
        db,
        user,
        action="social_chat_file_sent",
        details=json.dumps(
            {
                "thread_id": int(thread_id),
                "message_id": int(message.id),
                "file_name": original_name[:180],
                "size": len(raw),
            },
            ensure_ascii=False,
        ),
        module_code="social_hub",
        entity_type="chat_message",
        entity_id=str(message.id),
        request=request,
    )
    db.commit()
    for cache_key in dedupe_keys:
        _social_msg_cache_set(cache_key, int(message.id))
    return _social_message_to_out(db, actor_key, message)


@router.post("/social/chat/messages/{thread_id}/{message_id}/reactions", response_model=SocialChatMessageOut)
def social_chat_toggle_reaction(
    thread_id: int,
    message_id: int,
    payload: SocialChatReactionIn,
    request: Request,
    user: User = Depends(get_current_user),
    db: Session = Depends(get_db),
):
    ensure_module_enabled(db, user, "social_hub")
    actor_key, actor_nick, _ = _social_actor_identity(db, user)
    actor_aliases = _social_actor_alias_keys(db, actor_key)
    member_row = db.scalar(
        select(SocialChatThreadMember).where(
            SocialChatThreadMember.thread_id == int(thread_id),
            SocialChatThreadMember.actor_key.in_(actor_aliases),
        ).order_by(SocialChatThreadMember.id.asc())
    )
    if not member_row:
        raise HTTPException(status_code=403, detail="РќРµС‚ РґРѕСЃС‚СѓРїР° Рє С‡Р°С‚Сѓ")
    message = db.get(SocialChatMessage, int(message_id or 0))
    if not message or int(message.thread_id or 0) != int(thread_id):
        raise HTTPException(status_code=404, detail="РЎРѕРѕР±С‰РµРЅРёРµ РЅРµ РЅР°Р№РґРµРЅРѕ")
    emoji = str(payload.emoji or "").strip()
    if not emoji or len(emoji) > 16:
        raise HTTPException(status_code=400, detail="РќРµРєРѕСЂСЂРµРєС‚РЅР°СЏ СЂРµР°РєС†РёСЏ")
    reactions = _social_parse_reactions(getattr(message, "reactions_json", "") or "{}")
    rows = reactions.get(emoji, [])
    exists_idx = next(
        (idx for idx, item in enumerate(rows) if str(item.get("actor_key") or "") == actor_key),
        -1,
    )
    if exists_idx >= 0:
        rows.pop(exists_idx)
    else:
        rows.append({"actor_key": actor_key, "nick": actor_nick[:120]})
    rows = rows[:100]
    if rows:
        reactions[emoji] = rows
    else:
        reactions.pop(emoji, None)
    message.reactions_json = json.dumps(reactions, ensure_ascii=False)
    if str(message.sender_key or "") != actor_key:
        _social_push_notification(
            db,
            user_id=int(message.sender_user_id or user.id),
            recipient_key=str(message.sender_key or ""),
            kind="chat_reaction",
            dedupe_key=f"chat_reaction:{int(message.id)}:{actor_key}:{emoji}:{1 if exists_idx < 0 else 0}",
            title="Р РµР°РєС†РёСЏ РЅР° СЃРѕРѕР±С‰РµРЅРёРµ",
            body=f"{actor_nick}: {emoji}",
            payload={"thread_id": int(thread_id), "message_id": int(message.id), "emoji": emoji},
        )
    _audit(
        db,
        user,
        action="social_chat_reaction_toggled",
        details=json.dumps(
            {
                "thread_id": int(thread_id),
                "message_id": int(message.id),
                "emoji": emoji,
                "active": 1 if exists_idx < 0 else 0,
            },
            ensure_ascii=False,
        ),
        module_code="social_hub",
        entity_type="chat_message",
        entity_id=str(message.id),
        request=request,
    )
    db.commit()
    return _social_message_to_out(db, actor_key, message)


@router.post("/social/chat/read/{thread_id}", response_model=MessageOut)
def social_mark_chat_read(
    thread_id: int,
    user: User = Depends(get_current_user),
    db: Session = Depends(get_db),
):
    ensure_module_enabled(db, user, "social_hub")
    actor_key, actor_nick, _ = _social_actor_identity(db, user)
    actor_aliases = _social_actor_alias_keys(db, actor_key)
    thread = db.get(SocialChatThread, int(thread_id or 0))
    if not thread:
        raise HTTPException(status_code=404, detail="Р§Р°С‚ РЅРµ РЅР°Р№РґРµРЅ")
    if str(thread.kind or "") == "global":
        raise HTTPException(status_code=410, detail="Р“Р»РѕР±Р°Р»СЊРЅС‹Р№ С‡Р°С‚ РѕС‚РєР»СЋС‡РµРЅ")
    if str(thread.kind or "") in {"company", "group"} and int(thread.owner_user_id or 0) != int(user.id):
        raise HTTPException(status_code=403, detail="РќРµС‚ РґРѕСЃС‚СѓРїР° Рє С‡Р°С‚Сѓ")
    if str(thread.kind or "") == "company":
        member_row = _social_sync_company_thread_members(
            db,
            thread=thread,
            user_id=int(user.id),
            actor_key=actor_key,
            actor_nick=actor_nick,
        )
    else:
        member_row = db.scalar(
            select(SocialChatThreadMember).where(
                SocialChatThreadMember.thread_id == int(thread.id),
                SocialChatThreadMember.actor_key.in_(actor_aliases),
            ).order_by(SocialChatThreadMember.id.asc())
        )
        if not member_row:
            raise HTTPException(status_code=403, detail="РќРµС‚ РґРѕСЃС‚СѓРїР° Рє С‡Р°С‚Сѓ")
    last_id = db.scalar(
        select(func.max(SocialChatMessage.id)).where(SocialChatMessage.thread_id == thread_id)
    ) or 0
    member_row.last_read_message_id = int(last_id)
    member_row.updated_at = datetime.utcnow()
    # Mark related chat notifications as read for this thread.
    notif_rows = db.scalars(
        select(SocialNotification).where(
            SocialNotification.user_id == int(user.id),
            SocialNotification.recipient_key.in_(actor_aliases),
            SocialNotification.kind == "chat_message",
            SocialNotification.is_read.is_(False),
        )
    ).all()
    for notif in notif_rows:
        try:
            payload = json.loads(str(notif.payload_json or "{}"))
        except Exception:
            payload = {}
        if int(payload.get("thread_id") or 0) == int(thread_id):
            notif.is_read = True
    db.commit()
    return MessageOut(message="РћРє")


@router.get("/social/tasks/actors", response_model=list[dict[str, Any]])
def social_task_actors(
    user: User = Depends(get_current_user),
    db: Session = Depends(get_db),
):
    ensure_module_enabled(db, user, "social_hub")
    rows = db.scalars(
        select(TeamMember).where(
            TeamMember.user_id == user.id,
            TeamMember.is_active.is_(True),
        ).order_by(TeamMember.is_owner.desc(), TeamMember.id.asc())
    ).all()
    out = []
    for row in rows:
        out.append(
            {
                "actor_key": f"u:{user.id}" if bool(row.is_owner) else f"m:{row.id}",
                "nick": str((row.nickname or row.full_name or row.email).strip() or f"member-{row.id}"),
                "is_owner": bool(row.is_owner),
            }
        )
    return out


@router.get("/social/tasks/projects", response_model=list[SocialTaskProjectOut])
def social_task_projects(
    user: User = Depends(get_current_user),
    db: Session = Depends(get_db),
):
    ensure_module_enabled(db, user, "social_hub")
    rows = db.scalars(
        select(SocialTaskProject)
        .where(SocialTaskProject.user_id == user.id)
        .order_by(SocialTaskProject.updated_at.desc(), SocialTaskProject.id.desc())
    ).all()
    return [
        SocialTaskProjectOut(
            id=int(row.id),
            title=str(row.title or ""),
            description=str(row.description or ""),
            created_by_key=str(row.created_by_key or ""),
            created_by_nick=str(row.created_by_nick or ""),
            created_at=row.created_at.isoformat() if row.created_at else "",
        )
        for row in rows
    ]


@router.post("/social/tasks/projects", response_model=SocialTaskProjectOut)
def social_create_project(
    payload: SocialTaskProjectIn,
    request: Request,
    user: User = Depends(get_current_user),
    db: Session = Depends(get_db),
):
    ensure_module_enabled(db, user, "social_hub")
    actor_key, actor_nick, _ = _social_actor_identity(db, user)
    title = str(payload.title or "").strip()
    if not title:
        raise HTTPException(status_code=400, detail="Enter project title")
    row = SocialTaskProject(
        user_id=user.id,
        title=title[:255],
        description=str(payload.description or "")[:5000],
        created_by_key=actor_key,
        created_by_nick=actor_nick[:120],
    )
    db.add(row)
    db.flush()

    member_rows = db.scalars(
        select(TeamMember)
        .where(
            TeamMember.user_id == user.id,
            TeamMember.is_active.is_(True),
        )
        .order_by(TeamMember.is_owner.desc(), TeamMember.id.asc())
    ).all()
    seen_keys: set[str] = set()
    for member in member_rows:
        key = f"u:{int(user.id)}" if bool(member.is_owner) else f"m:{int(member.id)}"
        canon_key = _social_canonical_actor_key(db, key)
        if not canon_key or canon_key in seen_keys:
            continue
        seen_keys.add(canon_key)
        db.add(
            SocialTaskProjectMember(
                project_id=int(row.id),
                actor_key=canon_key,
                added_by_key=actor_key[:60],
            )
        )

    creator_key = _social_canonical_actor_key(db, actor_key)
    if creator_key and creator_key not in seen_keys:
        db.add(
            SocialTaskProjectMember(
                project_id=int(row.id),
                actor_key=creator_key,
                added_by_key=actor_key[:60],
            )
        )

    _audit(
        db,
        user,
        action="social_project_created",
        details=f"project_id={row.id}",
        module_code="social_hub",
        entity_type="task_project",
        entity_id=str(row.id),
        request=request,
    )
    db.commit()
    return SocialTaskProjectOut(
        id=int(row.id),
        title=str(row.title or ""),
        description=str(row.description or ""),
        created_by_key=str(row.created_by_key or ""),
        created_by_nick=str(row.created_by_nick or ""),
        created_at=row.created_at.isoformat() if row.created_at else "",
    )


def _social_task_company_actor_rows(db: Session, user: User) -> list[dict[str, Any]]:
    rows = db.scalars(
        select(TeamMember)
        .where(
            TeamMember.user_id == user.id,
            TeamMember.is_active.is_(True),
        )
        .order_by(TeamMember.is_owner.desc(), TeamMember.id.asc())
    ).all()
    out: list[dict[str, Any]] = []
    for row in rows:
        actor_key = f"u:{int(user.id)}" if bool(row.is_owner) else f"m:{int(row.id)}"
        nick = str((row.nickname or row.full_name or row.email).strip() or f"member-{row.id}")
        out.append(
            {
                "actor_key": _social_canonical_actor_key(db, actor_key),
                "nick": nick,
                "avatar_url": str(row.avatar_url or "").strip(),
                "is_owner": bool(row.is_owner),
            }
        )
    uniq: dict[str, dict[str, Any]] = {}
    for row in out:
        key = str(row.get("actor_key") or "").strip().lower()
        if key and key not in uniq:
            uniq[key] = row
    return list(uniq.values())


def _social_task_project_member_keys(db: Session, project_id: int) -> set[str]:
    rows = db.scalars(
        select(SocialTaskProjectMember.actor_key).where(
            SocialTaskProjectMember.project_id == int(project_id)
        )
    ).all()
    out: set[str] = set()
    for raw in rows:
        key = _social_canonical_actor_key(db, str(raw or "").strip().lower())
        if key:
            out.add(key)
    return out


def _social_task_actor_alias_set(db: Session, actor_key: str) -> set[str]:
    return set(_social_actor_alias_keys(db, actor_key))


@router.get("/social/tasks/projects/{project_id}/members", response_model=list[SocialTaskProjectMemberOut])
def social_task_project_members_list(
    project_id: int,
    user: User = Depends(get_current_user),
    db: Session = Depends(get_db),
):
    ensure_module_enabled(db, user, "social_hub")
    project = db.get(SocialTaskProject, int(project_id))
    if not project or int(project.user_id) != int(user.id):
        raise HTTPException(status_code=404, detail="Project not found")
    actor_rows = _social_task_company_actor_rows(db, user)
    member_keys = _social_task_project_member_keys(db, int(project.id))
    out: list[SocialTaskProjectMemberOut] = []
    for row in actor_rows:
        actor_key = str(row.get("actor_key") or "")
        out.append(
            SocialTaskProjectMemberOut(
                actor_key=actor_key,
                nick=str(row.get("nick") or ""),
                avatar_url=str(row.get("avatar_url") or "") or _social_current_avatar_by_key(db, actor_key) or "",
                is_owner=bool(row.get("is_owner")),
                in_project=bool(actor_key and actor_key in member_keys),
            )
        )
    return out


@router.put("/social/tasks/projects/{project_id}/members", response_model=list[SocialTaskProjectMemberOut])
def social_task_project_members_update(
    project_id: int,
    payload: SocialTaskProjectMembersUpdateIn,
    request: Request,
    user: User = Depends(get_current_user),
    db: Session = Depends(get_db),
):
    ensure_module_enabled(db, user, "social_hub")
    if not _actor_is_owner(user):
        raise HTTPException(status_code=403, detail="Only owner can manage project members")

    project = db.get(SocialTaskProject, int(project_id))
    if not project or int(project.user_id) != int(user.id):
        raise HTTPException(status_code=404, detail="Project not found")

    actor_rows = _social_task_company_actor_rows(db, user)
    actor_map = {str(row.get("actor_key") or ""): row for row in actor_rows}

    requested_keys: set[str] = set()
    for raw_key in list(payload.actor_keys or []):
        key = _social_canonical_actor_key(db, str(raw_key or "").strip().lower())
        if key and key in actor_map:
            requested_keys.add(key)

    owner_key = f"u:{int(user.id)}"
    requested_keys.add(owner_key)
    creator_key = _social_canonical_actor_key(db, str(project.created_by_key or "").strip().lower())
    if creator_key and creator_key in actor_map:
        requested_keys.add(creator_key)

    db.execute(delete(SocialTaskProjectMember).where(SocialTaskProjectMember.project_id == int(project.id)))
    for key in sorted(requested_keys):
        db.add(
            SocialTaskProjectMember(
                project_id=int(project.id),
                actor_key=key,
                added_by_key=owner_key,
            )
        )

    fallback_key = creator_key if creator_key in requested_keys else owner_key
    fallback_nick = _social_current_nick_by_key(db, fallback_key) or str(project.created_by_nick or "") or "Owner"
    project_tasks = db.scalars(
        select(SocialTask).where(
            SocialTask.user_id == user.id,
            SocialTask.project_id == int(project.id),
            SocialTask.task_kind == "company",
        )
    ).all()
    for task in project_tasks:
        assignee_key = _social_canonical_actor_key(db, str(task.assignee_key or "").strip().lower())
        if assignee_key and assignee_key in requested_keys:
            continue
        task.assignee_key = fallback_key
        task.assignee_nick = fallback_nick[:120]

    _audit(
        db,
        user,
        action="social_project_members_updated",
        details=f"project_id={project.id};members={len(requested_keys)}",
        module_code="social_hub",
        entity_type="task_project",
        entity_id=str(project.id),
        request=request,
    )
    db.commit()
    return social_task_project_members_list(project_id=int(project.id), user=user, db=db)


def _social_task_bucket(task: SocialTask, now_local: datetime) -> str:
    status = str(task.status or "todo").strip().lower()
    if status == "done":
        return "done"
    due = _social_localize_dt(task.due_date) if isinstance(task.due_date, datetime) else None
    if due is None:
        return "upcoming"
    if due < now_local:
        return "overdue"
    if due.date() == now_local.date():
        return "today"
    if due.date() == (now_local.date() + timedelta(days=1)):
        return "tomorrow"
    return "upcoming"


def _social_task_apply_bucket(task: SocialTask, bucket: str, now_local: datetime) -> None:
    safe_bucket = str(bucket or "").strip().lower()
    due = _social_localize_dt(task.due_date) if isinstance(task.due_date, datetime) else None
    base_hour = int(due.hour if due else 18)
    base_minute = int(due.minute if due else 0)

    if safe_bucket == "done":
        task.status = "done"
        task.closed_at = datetime.utcnow()
        task.completed_at = datetime.utcnow()
        return

    if str(task.status or "").strip().lower() == "done":
        task.status = "todo"
    task.closed_at = None
    task.completed_at = None

    if safe_bucket == "today":
        target_date = now_local.date()
    elif safe_bucket == "tomorrow":
        target_date = now_local.date() + timedelta(days=1)
    elif safe_bucket == "overdue":
        target_date = now_local.date() - timedelta(days=1)
    else:
        if due and due.date() >= (now_local.date() + timedelta(days=2)):
            target_date = due.date()
        else:
            target_date = now_local.date() + timedelta(days=2)

    task.due_date = datetime(target_date.year, target_date.month, target_date.day, base_hour, base_minute, 0)


def _social_task_can_view(db: Session, task: SocialTask, *, actor_aliases: set[str], is_owner: bool) -> bool:
    if is_owner and str(task.task_kind or "company") != "personal":
        return True

    task_kind = str(task.task_kind or "company").strip().lower()
    creator_key = _social_canonical_actor_key(db, str(task.creator_key or "").strip().lower())
    assignee_key = _social_canonical_actor_key(db, str(task.assignee_key or "").strip().lower())

    if task_kind == "personal":
        return bool(creator_key and creator_key in actor_aliases)

    if task.project_id:
        member_keys = _social_task_project_member_keys(db, int(task.project_id))
        if member_keys:
            return bool(actor_aliases & member_keys)

    return bool((creator_key and creator_key in actor_aliases) or (assignee_key and assignee_key in actor_aliases) or is_owner)

def _social_task_to_out(db: Session, task: SocialTask, *, actor_key: str, is_owner: bool) -> SocialTaskOut:
    actor_aliases = _social_task_actor_alias_set(db, actor_key)
    project = db.get(SocialTaskProject, task.project_id) if task.project_id else None
    comments = db.scalars(
        select(SocialTaskComment)
        .where(SocialTaskComment.task_id == task.id)
        .order_by(SocialTaskComment.id.asc())
    ).all()
    now_local = datetime.now(_social_calendar_tzinfo()).replace(tzinfo=None, microsecond=0)
    bucket = _social_task_bucket(task, now_local)

    creator_key = _social_canonical_actor_key(db, str(task.creator_key or "").strip().lower())
    assignee_key = _social_canonical_actor_key(db, str(task.assignee_key or "").strip().lower())
    personal_only = str(task.task_kind or "company").strip().lower() == "personal"

    can_delete = bool(is_owner or (creator_key and creator_key in actor_aliases))
    can_complete = bool(is_owner or (assignee_key and assignee_key in actor_aliases))
    can_edit = bool(can_delete or can_complete)
    if personal_only and not (creator_key and creator_key in actor_aliases):
        can_delete = False
        can_complete = False
        can_edit = False

    return SocialTaskOut(
        id=int(task.id),
        project_id=int(task.project_id) if task.project_id else None,
        project_title=str(project.title if project else ""),
        title=str(task.title or ""),
        description=str(task.description or ""),
        status=str(task.status or "todo"),
        priority=str(task.priority or "normal"),
        due_date=task.due_date.isoformat() if task.due_date else None,
        assignee_key=str(task.assignee_key or ""),
        assignee_nick=str(task.assignee_nick or ""),
        assignee_avatar_url=_social_current_avatar_by_key(db, str(task.assignee_key or "")) or "",
        creator_key=str(task.creator_key or ""),
        creator_nick=str(task.creator_nick or ""),
        task_kind=str(task.task_kind or "company"),
        bucket=bucket,
        can_edit=can_edit,
        can_delete=can_delete,
        can_complete=can_complete,
        sort_order=int(task.sort_order or 0),
        completed_at=task.completed_at.isoformat() if task.completed_at else None,
        comments=[
            {
                "id": int(c.id),
                "author_key": str(c.author_key or ""),
                "author_nick": str(c.author_nick or ""),
                "text": str(c.text or ""),
                "created_at": c.created_at.isoformat() if c.created_at else "",
            }
            for c in comments
        ],
        created_at=task.created_at.isoformat() if task.created_at else "",
        updated_at=task.updated_at.isoformat() if task.updated_at else "",
    )


@router.get("/social/tasks", response_model=list[SocialTaskOut])
def social_tasks_list(
    project_id: int = 0,
    status: str = "",
    task_kind: str = "all",
    include_done: bool = False,
    user: User = Depends(get_current_user),
    db: Session = Depends(get_db),
):
    ensure_module_enabled(db, user, "social_hub")
    actor_key, actor_nick, _ = _social_actor_identity(db, user)
    actor_aliases = _social_task_actor_alias_set(db, actor_key)
    is_owner = bool(_actor_is_owner(user))

    query = select(SocialTask).where(SocialTask.user_id == user.id)
    if int(project_id or 0) > 0:
        query = query.where(SocialTask.project_id == int(project_id))

    safe_kind = str(task_kind or "all").strip().lower()
    if safe_kind in {"company", "personal"}:
        query = query.where(SocialTask.task_kind == safe_kind)

    safe_status = str(status or "").strip().lower()
    if safe_status in {"todo", "in_progress", "done"}:
        query = query.where(SocialTask.status == safe_status)
    elif not bool(include_done):
        query = query.where(SocialTask.status != "done")

    rows = db.scalars(
        query.order_by(
            SocialTask.sort_order.asc(),
            SocialTask.due_date.asc(),
            SocialTask.updated_at.desc(),
            SocialTask.id.desc(),
        ).limit(2000)
    ).all()

    now_local = datetime.now(_social_calendar_tzinfo()).replace(tzinfo=None, microsecond=0)
    visible_rows: list[SocialTask] = []
    notif_created = False

    for task in rows:
        if not _social_task_can_view(db, task, actor_aliases=actor_aliases, is_owner=is_owner):
            continue
        visible_rows.append(task)

        status_value = str(task.status or "todo").strip().lower()
        if status_value == "done" or not isinstance(task.due_date, datetime):
            continue

        due_local = _social_localize_dt(task.due_date) or task.due_date
        assignee_key = _social_canonical_actor_key(db, str(task.assignee_key or "").strip().lower())
        creator_key = _social_canonical_actor_key(db, str(task.creator_key or "").strip().lower())

        if assignee_key:
            if due_local < now_local:
                _social_push_notification(
                    db,
                    user_id=user.id,
                    recipient_key=str(task.assignee_key or ""),
                    kind="task_overdue",
                    dedupe_key=f"task_overdue:{task.id}:{now_local.date().isoformat()}:{assignee_key}",
                    title="Р—Р°РґР°С‡Р° РїСЂРѕСЃСЂРѕС‡РµРЅР°",
                    body=f"{str(task.title or '')[:120]} вЂ” СЃСЂРѕРє Р·Р°РґР°С‡Рё РёСЃС‚РµРє.",
                    payload={
                        "task_id": int(task.id),
                        "i18n_key": "task_overdue",
                        "i18n_params": {
                            "task_title": str(task.title or "")[:180],
                            "assignee_nick": str(task.assignee_nick or ""),
                        },
                    },
                )
                notif_created = True
            else:
                delta_sec = (due_local - now_local).total_seconds()
                if 0 < delta_sec <= (3 * 3600):
                    slot = due_local.strftime("%Y%m%d%H")
                    due_text = due_local.strftime("%d.%m.%Y %H:%M")
                    _social_push_notification(
                        db,
                        user_id=user.id,
                        recipient_key=str(task.assignee_key or ""),
                        kind="task_reminder_3h",
                        dedupe_key=f"task_reminder_3h:{task.id}:{slot}:{assignee_key}",
                        title="РЎСЂРѕРє Р·Р°РґР°С‡Рё СЃРєРѕСЂРѕ",
                        body=f"{str(task.title or '')[:120]} вЂ” РґРѕ РґРµРґР»Р°Р№РЅР° РјРµРЅСЊС€Рµ 3 С‡Р°СЃРѕРІ.",
                        payload={
                            "task_id": int(task.id),
                            "i18n_key": "task_reminder_3h",
                            "i18n_params": {
                                "task_title": str(task.title or "")[:180],
                                "assignee_nick": str(task.assignee_nick or ""),
                                "due_text": due_text,
                            },
                        },
                    )
                    notif_created = True

        if due_local < now_local and creator_key and creator_key not in {assignee_key, ""}:
            _social_push_notification(
                db,
                user_id=user.id,
                recipient_key=str(task.creator_key or ""),
                kind="task_overdue",
                dedupe_key=f"task_overdue:{task.id}:{now_local.date().isoformat()}:{creator_key}",
                title="Р—Р°РґР°С‡Р° РїСЂРѕСЃСЂРѕС‡РµРЅР°",
                body=f"{str(task.title or '')[:120]} вЂ” РёСЃРїРѕР»РЅРёС‚РµР»СЊ РїСЂРѕРїСѓСЃС‚РёР» СЃСЂРѕРє.",
                payload={
                    "task_id": int(task.id),
                    "i18n_key": "task_overdue",
                    "i18n_params": {
                        "task_title": str(task.title or "")[:180],
                        "assignee_nick": str(task.assignee_nick or ""),
                    },
                },
            )
            notif_created = True

    bucket_order = {"overdue": 0, "today": 1, "tomorrow": 2, "upcoming": 3, "done": 4}
    visible_rows.sort(
        key=lambda row: (
            bucket_order.get(_social_task_bucket(row, now_local), 9),
            int(getattr(row, "sort_order", 0) or 0),
            getattr(row, "due_date", None) or datetime.max,
            -int(getattr(row, "id", 0) or 0),
        )
    )

    if notif_created:
        db.commit()

    return [_social_task_to_out(db, row, actor_key=actor_key, is_owner=is_owner) for row in visible_rows]


@router.post("/social/tasks", response_model=SocialTaskOut)
def social_create_task(
    payload: SocialTaskIn,
    request: Request,
    user: User = Depends(get_current_user),
    db: Session = Depends(get_db),
):
    ensure_module_enabled(db, user, "social_hub")
    actor_key, actor_nick, _ = _social_actor_identity(db, user)
    actor_aliases = _social_task_actor_alias_set(db, actor_key)

    title = str(payload.title or "").strip()
    if not title:
        raise HTTPException(status_code=400, detail="Task title is required")

    safe_kind = str(payload.task_kind or "company").strip().lower()
    if safe_kind not in {"company", "personal"}:
        safe_kind = "company"

    due_dt = _social_parse_dt(payload.due_date)
    project_id = _to_int_safe(payload.project_id) or None

    if safe_kind == "personal":
        project_id = None
        assignee_key = actor_key
        assignee_nick = actor_nick
    else:
        assignee_key = str(payload.assignee_key or actor_key).strip().lower() or actor_key
        assignee_user_id, _, assignee_nick = _social_identity_by_key(db, assignee_key)
        if int(assignee_user_id) != int(user.id):
            raise HTTPException(status_code=400, detail="Assignee must be from your company")

    if project_id:
        project = db.get(SocialTaskProject, project_id)
        if not project or int(project.user_id) != int(user.id):
            raise HTTPException(status_code=404, detail="Project not found")
        member_keys = _social_task_project_member_keys(db, int(project_id))
        if member_keys and not bool(actor_aliases & member_keys) and not _actor_is_owner(user):
            raise HTTPException(status_code=403, detail="No access to this project")
        if safe_kind == "company":
            assignee_aliases = _social_task_actor_alias_set(db, assignee_key)
            if member_keys and not bool(member_keys & assignee_aliases):
                raise HTTPException(status_code=400, detail="Assignee is not a project member")

    max_sort_query = select(func.max(SocialTask.sort_order)).where(
        SocialTask.user_id == user.id,
        SocialTask.task_kind == safe_kind,
    )
    if project_id:
        max_sort_query = max_sort_query.where(SocialTask.project_id == int(project_id))
    else:
        max_sort_query = max_sort_query.where(SocialTask.project_id.is_(None))
    max_sort = int(db.scalar(max_sort_query) or 0)

    task = SocialTask(
        user_id=user.id,
        project_id=project_id,
        title=title[:255],
        description=str(payload.description or "")[:5000],
        status="todo",
        priority=str(payload.priority or "normal")[:20],
        task_kind=safe_kind,
        sort_order=max_sort + 10,
        due_date=due_dt,
        assignee_key=assignee_key,
        assignee_nick=assignee_nick[:120],
        creator_key=actor_key,
        creator_nick=actor_nick[:120],
    )
    db.add(task)
    db.flush()

    notif_stamp = int(datetime.utcnow().timestamp())
    notify_recipients = {str(assignee_key or "").strip(), str(actor_key or "").strip()}
    notify_recipients.discard("")
    for recipient in sorted(notify_recipients):
        _social_push_notification(
            db,
            user_id=user.id,
            recipient_key=recipient,
            kind="task_assigned",
            dedupe_key=f"task_assigned:{task.id}:{notif_stamp}:{recipient}",
            title="РќРѕРІР°СЏ Р·Р°РґР°С‡Р°",
            body=f"{actor_nick}: {title[:180]}",
            payload={
                "task_id": int(task.id),
                "i18n_key": "task_assigned",
                "i18n_params": {
                    "task_title": str(title or "")[:180],
                    "actor_nick": str(actor_nick or ""),
                    "assignee_nick": str(task.assignee_nick or ""),
                },
            },
        )

    _audit(
        db,
        user,
        action="social_task_created",
        details=f"task_id={task.id};assignee={assignee_key};kind={safe_kind}",
        module_code="social_hub",
        entity_type="task",
        entity_id=str(task.id),
        request=request,
    )
    db.commit()
    return _social_task_to_out(db, task, actor_key=actor_key, is_owner=bool(_actor_is_owner(user)))


@router.put("/social/tasks/{task_id}", response_model=SocialTaskOut)
def social_update_task(
    task_id: int,
    payload: SocialTaskUpdateIn,
    request: Request,
    user: User = Depends(get_current_user),
    db: Session = Depends(get_db),
):
    ensure_module_enabled(db, user, "social_hub")
    actor_key, actor_nick, _ = _social_actor_identity(db, user)
    actor_aliases = _social_task_actor_alias_set(db, actor_key)
    is_owner = bool(_actor_is_owner(user))

    task = db.get(SocialTask, task_id)
    if not task or int(task.user_id) != int(user.id):
        raise HTTPException(status_code=404, detail="Task not found")

    if not _social_task_can_view(db, task, actor_aliases=actor_aliases, is_owner=is_owner):
        raise HTTPException(status_code=403, detail="No access to this task")

    creator_key = _social_canonical_actor_key(db, str(task.creator_key or "").strip().lower())
    assignee_key_old = _social_canonical_actor_key(db, str(task.assignee_key or "").strip().lower())
    can_edit = bool(is_owner or (creator_key and creator_key in actor_aliases) or (assignee_key_old and assignee_key_old in actor_aliases))
    if not can_edit:
        raise HTTPException(status_code=403, detail="No rights to update this task")

    old_assignee = str(task.assignee_key or "")
    old_status = str(task.status or "todo")

    if payload.task_kind is not None:
        safe_kind = str(payload.task_kind or "company").strip().lower()
        if safe_kind in {"company", "personal"}:
            task.task_kind = safe_kind
            if safe_kind == "personal":
                task.project_id = None
                task.assignee_key = actor_key
                task.assignee_nick = actor_nick[:120]

    if payload.project_id is not None and str(task.task_kind or "company") != "personal":
        pid = _to_int_safe(payload.project_id) or None
        if pid:
            project = db.get(SocialTaskProject, pid)
            if not project or int(project.user_id) != int(user.id):
                raise HTTPException(status_code=404, detail="Project not found")
            member_keys = _social_task_project_member_keys(db, int(pid))
            if member_keys and not bool(actor_aliases & member_keys) and not is_owner:
                raise HTTPException(status_code=403, detail="No access to this project")
        task.project_id = pid

    if payload.title is not None:
        title = str(payload.title or "").strip()
        if not title:
            raise HTTPException(status_code=400, detail="Task title cannot be empty")
        task.title = title[:255]

    if payload.description is not None:
        task.description = str(payload.description or "")[:5000]

    if payload.priority is not None:
        task.priority = str(payload.priority or "normal")[:20]

    if payload.due_date is not None:
        task.due_date = _social_parse_dt(payload.due_date)

    if payload.assignee_key is not None and str(task.task_kind or "company") != "personal":
        assignee_key = str(payload.assignee_key or "").strip().lower()
        if assignee_key:
            assignee_user_id, _, assignee_nick = _social_identity_by_key(db, assignee_key)
            if int(assignee_user_id) != int(user.id):
                raise HTTPException(status_code=400, detail="Assignee must be from your company")
            if task.project_id:
                member_keys = _social_task_project_member_keys(db, int(task.project_id))
                assignee_aliases = _social_task_actor_alias_set(db, assignee_key)
                if member_keys and not bool(member_keys & assignee_aliases):
                    raise HTTPException(status_code=400, detail="Assignee is not a project member")
            task.assignee_key = assignee_key
            task.assignee_nick = assignee_nick[:120]

    if payload.status is not None:
        safe_status = str(payload.status or "").strip().lower()
        if safe_status not in {"todo", "in_progress", "done"}:
            raise HTTPException(status_code=400, detail="Invalid task status")
        if safe_status == "done" and not is_owner:
            assignee_key = _social_canonical_actor_key(db, str(task.assignee_key or "").strip().lower())
            if not assignee_key or assignee_key not in actor_aliases:
                raise HTTPException(status_code=403, detail="Employee can close only own tasks")
        task.status = safe_status
        if safe_status == "done":
            task.closed_at = datetime.utcnow()
            task.completed_at = datetime.utcnow()
        else:
            task.closed_at = None
            task.completed_at = None

    if str(task.assignee_key or "") != old_assignee:
        reassigned_stamp = int(datetime.utcnow().timestamp())
        reassign_recipients = {str(task.assignee_key or "").strip(), str(task.creator_key or "").strip()}
        reassign_recipients.discard("")
        for recipient in sorted(reassign_recipients):
            _social_push_notification(
                db,
                user_id=user.id,
                recipient_key=recipient,
                kind="task_assigned",
                dedupe_key=f"task_assigned:{task.id}:{reassigned_stamp}:{recipient}",
                title="Р—Р°РґР°С‡Р° РЅР°Р·РЅР°С‡РµРЅР°",
                body=f"{actor_nick}: {str(task.title or '')[:180]}",
                payload={
                    "task_id": int(task.id),
                    "i18n_key": "task_assigned",
                    "i18n_params": {
                        "task_title": str(task.title or "")[:180],
                        "actor_nick": str(actor_nick or ""),
                        "assignee_nick": str(task.assignee_nick or ""),
                    },
                },
            )

    if old_status != "done" and str(task.status or "") == "done":
        done_stamp = int(datetime.utcnow().timestamp())
        done_recipients = {str(task.creator_key or "").strip(), str(task.assignee_key or "").strip()}
        done_recipients.discard("")
        for recipient in sorted(done_recipients):
            _social_push_notification(
                db,
                user_id=user.id,
                recipient_key=recipient,
                kind="task_done",
                dedupe_key=f"task_done:{task.id}:{done_stamp}:{recipient}",
                title="Р—Р°РґР°С‡Р° РІС‹РїРѕР»РЅРµРЅР°",
                body=f"{actor_nick}: {str(task.title or '')[:140]}",
                payload={
                    "task_id": int(task.id),
                    "i18n_key": "task_done",
                    "i18n_params": {
                        "task_title": str(task.title or "")[:180],
                        "actor_nick": str(actor_nick or ""),
                        "assignee_nick": str(task.assignee_nick or ""),
                    },
                },
            )

    _audit(
        db,
        user,
        action="social_task_updated",
        details=f"task_id={task.id};status={task.status};assignee={task.assignee_key};kind={task.task_kind}",
        module_code="social_hub",
        entity_type="task",
        entity_id=str(task.id),
        request=request,
    )
    db.commit()
    return _social_task_to_out(db, task, actor_key=actor_key, is_owner=is_owner)


@router.delete("/social/tasks/{task_id}", response_model=MessageOut)
def social_delete_task(
    task_id: int,
    request: Request,
    user: User = Depends(get_current_user),
    db: Session = Depends(get_db),
):
    ensure_module_enabled(db, user, "social_hub")
    actor_key, _, _ = _social_actor_identity(db, user)
    actor_aliases = _social_task_actor_alias_set(db, actor_key)
    is_owner = bool(_actor_is_owner(user))

    task = db.get(SocialTask, task_id)
    if not task or int(task.user_id) != int(user.id):
        raise HTTPException(status_code=404, detail="Task not found")

    creator_key = _social_canonical_actor_key(db, str(task.creator_key or "").strip().lower())
    if not is_owner and not (creator_key and creator_key in actor_aliases):
        raise HTTPException(status_code=403, detail="Only owner or task creator can delete task")

    db.execute(delete(SocialTaskComment).where(SocialTaskComment.task_id == int(task.id)))
    db.delete(task)
    _audit(
        db,
        user,
        action="social_task_deleted",
        details=f"task_id={task_id}",
        module_code="social_hub",
        entity_type="task",
        entity_id=str(task_id),
        request=request,
    )
    db.commit()
    return MessageOut(message="ok")


@router.post("/social/tasks/reorder", response_model=list[SocialTaskOut])
def social_reorder_tasks(
    payload: SocialTaskReorderIn,
    request: Request,
    user: User = Depends(get_current_user),
    db: Session = Depends(get_db),
):
    ensure_module_enabled(db, user, "social_hub")
    actor_key, _, _ = _social_actor_identity(db, user)
    actor_aliases = _social_task_actor_alias_set(db, actor_key)
    is_owner = bool(_actor_is_owner(user))

    items = list(payload.items or [])
    if not items:
        return social_tasks_list(user=user, db=db)

    now_local = datetime.now(_social_calendar_tzinfo()).replace(tzinfo=None, microsecond=0)
    bucket_base = {
        "overdue": 100000,
        "today": 200000,
        "tomorrow": 300000,
        "upcoming": 400000,
        "done": 900000,
    }

    touched_ids: list[int] = []
    for item in items:
        task_id = int(getattr(item, "task_id", 0) or 0)
        if task_id <= 0:
            continue
        task = db.get(SocialTask, task_id)
        if not task or int(task.user_id) != int(user.id):
            continue
        if not _social_task_can_view(db, task, actor_aliases=actor_aliases, is_owner=is_owner):
            continue

        creator_key = _social_canonical_actor_key(db, str(task.creator_key or "").strip().lower())
        assignee_key = _social_canonical_actor_key(db, str(task.assignee_key or "").strip().lower())
        can_edit = bool(is_owner or (creator_key and creator_key in actor_aliases) or (assignee_key and assignee_key in actor_aliases))
        if not can_edit:
            continue

        bucket = str(getattr(item, "bucket", "") or "").strip().lower()
        if bucket:
            _social_task_apply_bucket(task, bucket, now_local)

        to_index = getattr(item, "to_index", None)
        base = int(bucket_base.get(bucket or _social_task_bucket(task, now_local), 500000))
        if to_index is None:
            task.sort_order = int(base + (len(touched_ids) + 1) * 10)
        else:
            task.sort_order = int(base + max(0, int(to_index)) * 10 + 10)

        touched_ids.append(int(task.id))

    _audit(
        db,
        user,
        action="social_task_reordered",
        details=f"count={len(touched_ids)}",
        module_code="social_hub",
        entity_type="task",
        entity_id=",".join(str(x) for x in touched_ids[:20]),
        request=request,
    )
    db.commit()
    return social_tasks_list(include_done=True, task_kind="all", user=user, db=db)


@router.post("/social/tasks/{task_id}/comments", response_model=SocialTaskOut)
def social_add_task_comment(
    task_id: int,
    payload: SocialTaskCommentIn,
    request: Request,
    user: User = Depends(get_current_user),
    db: Session = Depends(get_db),
):
    ensure_module_enabled(db, user, "social_hub")
    actor_key, actor_nick, _ = _social_actor_identity(db, user)
    actor_aliases = _social_task_actor_alias_set(db, actor_key)

    task = db.get(SocialTask, task_id)
    if not task or int(task.user_id) != int(user.id):
        raise HTTPException(status_code=404, detail="Task not found")
    if not _social_task_can_view(db, task, actor_aliases=actor_aliases, is_owner=bool(_actor_is_owner(user))):
        raise HTTPException(status_code=403, detail="No access to this task")

    text_comment = str(payload.text or "").strip()
    if not text_comment:
        raise HTTPException(status_code=400, detail="Comment is empty")

    db.add(
        SocialTaskComment(
            task_id=task.id,
            author_key=actor_key,
            author_nick=actor_nick[:120],
            text=text_comment[:5000],
        )
    )

    recipients = {str(task.creator_key or ""), str(task.assignee_key or "")}
    recipients.discard(actor_key)
    for recipient in recipients:
        if not recipient:
            continue
        _social_push_notification(
            db,
            user_id=user.id,
            recipient_key=recipient,
            kind="task_comment",
            dedupe_key=f"task_comment:{task.id}:{int(datetime.utcnow().timestamp())}:{recipient}",
            title="New task comment",
            body=f"{actor_nick}: {text_comment[:180]}",
            payload={"task_id": int(task.id)},
        )

    _audit(
        db,
        user,
        action="social_task_comment_added",
        details=f"task_id={task.id}",
        module_code="social_hub",
        entity_type="task",
        entity_id=str(task.id),
        request=request,
    )
    db.commit()
    return _social_task_to_out(db, task, actor_key=actor_key, is_owner=bool(_actor_is_owner(user)))

def _social_calendar_tzinfo() -> ZoneInfo:
    raw = str(os.getenv("SEO_WIBE_CALENDAR_TZ") or os.getenv("TZ") or SOCIAL_CALENDAR_DEFAULT_TZ).strip()
    if raw:
        try:
            return ZoneInfo(raw)
        except Exception:
            pass
    return ZoneInfo(SOCIAL_CALENDAR_DEFAULT_TZ)


def _social_localize_dt(value: datetime | None) -> datetime | None:
    if not isinstance(value, datetime):
        return None
    if value.tzinfo is None:
        return value.replace(microsecond=0)
    return value.astimezone(_social_calendar_tzinfo()).replace(tzinfo=None, microsecond=0)


def _social_forwarded_first(value: Any) -> str:
    return str(value or "").split(",", 1)[0].strip()


def _social_public_base_url(request: Request) -> str:
    from_env = str(os.getenv("SEO_WIBE_PUBLIC_BASE_URL") or "").strip().rstrip("/")
    if from_env:
        return from_env

    def _host_is_local(raw_host: str) -> bool:
        host = str(raw_host or "").split(":", 1)[0].strip().lower()
        if not host:
            return True
        if host in {"localhost", "127.0.0.1", "0.0.0.0", "::1", "5.129.207.106"}:
            return True
        return bool(re.fullmatch(r"\d+\.\d+\.\d+\.\d+", host))

    forwarded_host = _social_forwarded_first(request.headers.get("x-forwarded-host") or request.headers.get("host"))
    if forwarded_host and not _host_is_local(forwarded_host):
        forwarded_proto = _social_forwarded_first(
            request.headers.get("x-forwarded-proto") or request.headers.get("x-forwarded-scheme") or request.url.scheme
        ) or "https"
        forwarded_prefix = _social_forwarded_first(request.headers.get("x-forwarded-prefix") or "").strip()
        prefix = f"/{forwarded_prefix.lstrip('/')}" if forwarded_prefix and forwarded_prefix != "/" else ""
        return f"{forwarded_proto}://{forwarded_host}{prefix}".rstrip("/")

    request_host = str(request.url.hostname or "").strip().lower()
    if _host_is_local(request_host):
        return "https://seowibe.ru"
    return str(request.base_url).rstrip("/")


def _social_google_oauth_config(request: Request) -> tuple[str, str, str]:
    client_id = str(os.getenv("SEO_WIBE_GOOGLE_CLIENT_ID") or "").strip()
    client_secret = str(os.getenv("SEO_WIBE_GOOGLE_CLIENT_SECRET") or "").strip()
    redirect_uri = str(os.getenv("SEO_WIBE_GOOGLE_REDIRECT_URI") or "").strip()
    if not redirect_uri:
        redirect_uri = f"{_social_public_base_url(request)}/api/social/calendar/google-oauth/callback"
    return client_id, client_secret, redirect_uri


def _social_actor_store_context(db: Session, user: User) -> dict[str, Any]:
    actor_key, _, _ = _social_actor_identity(db, user)
    store_key = _social_canonical_actor_key(db, actor_key) or actor_key or f"u:{int(user.id)}"
    fallback_keys = [
        key
        for key in _social_actor_alias_keys(db, actor_key)
        if key and key != store_key
    ]
    legacy_key = str(int(user.id))
    if legacy_key and legacy_key != store_key and legacy_key not in fallback_keys:
        fallback_keys.append(legacy_key)
    return {
        "actor_key": actor_key,
        "store_key": store_key,
        "fallback_keys": fallback_keys,
        "actor_email": _actor_email(user) or str(user.email or "").strip().lower(),
    }


def _social_store_lookup_keys(store_key: str | int, fallback_keys: list[str] | None = None) -> list[str]:
    out: list[str] = []
    for raw in [store_key, *(fallback_keys or [])]:
        key = str(raw or "").strip()
        if not key or key in out:
            continue
        out.append(key)
    return out


def _social_google_oauth_state_for_actor(*, user_id: int, store_key: str, return_target: str = "web") -> str:
    payload = {
        "u": int(user_id),
        "k": str(store_key or "").strip()[:120],
        "rt": "apk" if str(return_target or "").strip().lower() == "apk" else "web",
        "ts": int(time.time()),
    }
    encoded = base64.urlsafe_b64encode(
        json.dumps(payload, ensure_ascii=True, separators=(",", ":")).encode("utf-8")
    ).decode("ascii").rstrip("=")
    return create_access_token(f"gcal_oauth_v2:{encoded}", expires_minutes=20)


def _social_google_oauth_state_from_token(state: str) -> dict[str, Any]:
    subject = str(decode_access_token(str(state or "").strip()) or "").strip()
    if subject.startswith("gcal_oauth_v2:"):
        encoded = subject.split(":", 1)[1]
        padded = encoded + ("=" * ((4 - len(encoded) % 4) % 4))
        try:
            payload = json.loads(base64.urlsafe_b64decode(padded.encode("ascii")).decode("utf-8"))
        except Exception:
            return {"user_id": 0, "store_key": "", "return_target": "web"}
        return {
            "user_id": max(0, int(payload.get("u") or 0)),
            "store_key": str(payload.get("k") or "").strip(),
            "return_target": "apk" if str(payload.get("rt") or "").strip().lower() == "apk" else "web",
        }
    if subject.startswith("gcal_oauth:"):
        chunks = subject.split(":")
        if len(chunks) >= 3:
            try:
                user_id = int(chunks[1] or 0)
            except Exception:
                user_id = 0
            if user_id > 0:
                return {"user_id": user_id, "store_key": f"u:{user_id}", "return_target": "web"}
    return {"user_id": 0, "store_key": "", "return_target": "web"}


def _social_google_tokens_load(db: Session) -> dict[str, dict[str, Any]]:
    raw = _get_system_setting(db, SOCIAL_GOOGLE_OAUTH_TOKENS_KEY)
    if not raw:
        return {}
    try:
        parsed = json.loads(raw)
    except Exception:
        return {}
    if not isinstance(parsed, dict):
        return {}
    out: dict[str, dict[str, Any]] = {}
    for key, value in parsed.items():
        store_key = str(key or "").strip()
        if not store_key or not isinstance(value, dict):
            continue
        out[store_key] = dict(value)
    return out


def _social_google_tokens_get(db: Session, store_key: str | int, *, fallback_keys: list[str] | None = None) -> dict[str, Any]:
    all_tokens = _social_google_tokens_load(db)
    for key in _social_store_lookup_keys(store_key, fallback_keys):
        payload = all_tokens.get(key)
        if isinstance(payload, dict):
            return dict(payload)
    return {}


def _social_google_tokens_save(db: Session, store_key: str | int, payload: dict[str, Any]) -> dict[str, Any]:
    key = str(store_key or "").strip()
    if not key:
        raise HTTPException(status_code=400, detail="Google token store key is empty")
    all_tokens = _social_google_tokens_load(db)
    safe_payload = {
        "access_token": str(payload.get("access_token") or "").strip(),
        "refresh_token": str(payload.get("refresh_token") or "").strip(),
        "token_type": str(payload.get("token_type") or "Bearer").strip() or "Bearer",
        "scope": str(payload.get("scope") or "").strip(),
        "expires_at": int(payload.get("expires_at") or 0),
        "account_email": str(payload.get("account_email") or "").strip().lower(),
        "updated_at": datetime.utcnow().isoformat(),
    }
    all_tokens[key] = safe_payload
    _set_system_setting(db, SOCIAL_GOOGLE_OAUTH_TOKENS_KEY, json.dumps(all_tokens, ensure_ascii=False))
    return safe_payload


def _social_calendar_sync_status_load(db: Session) -> dict[str, dict[str, Any]]:
    raw = _get_system_setting(db, SOCIAL_CALENDAR_SYNC_STATUS_KEY)
    if not raw:
        return {}
    try:
        parsed = json.loads(raw)
    except Exception:
        return {}
    if not isinstance(parsed, dict):
        return {}
    out: dict[str, dict[str, Any]] = {}
    for key, value in parsed.items():
        store_key = str(key or "").strip()
        if not store_key or not isinstance(value, dict):
            continue
        out[store_key] = dict(value)
    return out


def _social_calendar_sync_status_get(db: Session, store_key: str | int, *, fallback_keys: list[str] | None = None) -> dict[str, Any]:
    all_status = _social_calendar_sync_status_load(db)
    for key in _social_store_lookup_keys(store_key, fallback_keys):
        payload = all_status.get(key)
        if isinstance(payload, dict):
            return dict(payload)
    return {}


def _social_calendar_sync_status_save(db: Session, store_key: str | int, payload: dict[str, Any]) -> dict[str, Any]:
    key = str(store_key or "").strip()
    if not key:
        raise HTTPException(status_code=400, detail="Calendar sync store key is empty")
    all_status = _social_calendar_sync_status_load(db)
    warnings = payload.get("warnings") if isinstance(payload.get("warnings"), list) else []
    safe_payload = {
        "last_sync_at": str(payload.get("last_sync_at") or datetime.utcnow().isoformat()).strip(),
        "last_sync_source": str(payload.get("last_sync_source") or "").strip(),
        "last_sync_state": str(payload.get("last_sync_state") or "idle").strip(),
        "last_sync_ok": bool(payload.get("last_sync_ok")),
        "imported": max(0, int(payload.get("imported") or 0)),
        "updated": max(0, int(payload.get("updated") or 0)),
        "deleted": max(0, int(payload.get("deleted") or 0)),
        "skipped": max(0, int(payload.get("skipped") or 0)),
        "warning_count": len([str(x or "").strip() for x in warnings if str(x or "").strip()]),
        "warnings": [str(x or "").strip()[:160] for x in warnings if str(x or "").strip()][:12],
        "last_sync_error": str(payload.get("last_sync_error") or "").strip()[:260],
    }
    all_status[key] = safe_payload
    _set_system_setting(db, SOCIAL_CALENDAR_SYNC_STATUS_KEY, json.dumps(all_status, ensure_ascii=False))
    return safe_payload


def _social_calendar_sync_mark(
    db: Session,
    *,
    store_key: str,
    source: str,
    state: str,
    imported: int = 0,
    updated: int = 0,
    deleted: int = 0,
    skipped: int = 0,
    warnings: list[str] | None = None,
    error: str = "",
) -> dict[str, Any]:
    safe_warnings = [str(x or "").strip() for x in (warnings or []) if str(x or "").strip()]
    return _social_calendar_sync_status_save(
        db,
        store_key,
        {
            "last_sync_at": datetime.utcnow().isoformat(),
            "last_sync_source": str(source or "").strip(),
            "last_sync_state": str(state or "idle").strip(),
            "last_sync_ok": str(state or "").strip() in {"ok", "partial", "empty"},
            "imported": imported,
            "updated": updated,
            "deleted": deleted,
            "skipped": skipped,
            "warnings": safe_warnings,
            "last_sync_error": error,
        },
    )



def _social_calendar_reminder_store_key(user_id: int) -> str:
    uid = max(0, int(user_id or 0))
    return f"{SOCIAL_CALENDAR_REMINDER_SETTINGS_PREFIX}{uid}"[:80]


def _social_calendar_reminder_normalize_offsets(offsets: list[Any] | None) -> list[int]:
    allowed = set(int(x) for x in SOCIAL_CALENDAR_REMINDER_ALLOWED_OFFSETS_MIN)
    source = offsets if isinstance(offsets, list) else []
    out: list[int] = []
    seen: set[int] = set()
    for raw in source:
        value = int(raw or 0)
        if value not in allowed or value in seen:
            continue
        seen.add(value)
        out.append(value)
    out.sort()
    if out:
        return out
    return list(SOCIAL_CALENDAR_REMINDER_DEFAULT_OFFSETS_MIN)


def _social_calendar_reminders_load_user(db: Session, user_id: int) -> dict[str, dict[str, Any]]:
    raw = _get_system_setting(db, _social_calendar_reminder_store_key(user_id))
    if not raw:
        return {}
    try:
        parsed = json.loads(raw)
    except Exception:
        return {}
    if not isinstance(parsed, dict):
        return {}
    out: dict[str, dict[str, Any]] = {}
    for key, value in parsed.items():
        item_key = str(key or "").strip()
        if not item_key or not isinstance(value, dict):
            continue
        out[item_key] = {
            "enabled": bool(value.get("enabled", True)),
            "offsets_min": _social_calendar_reminder_normalize_offsets(value.get("offsets_min") if isinstance(value.get("offsets_min"), list) else []),
            "updated_at": str(value.get("updated_at") or "").strip(),
        }
    return out


def _social_calendar_reminders_save_user(db: Session, user_id: int, payload: dict[str, dict[str, Any]]) -> None:
    safe: dict[str, dict[str, Any]] = {}
    for key, value in (payload or {}).items():
        item_key = str(key or "").strip()
        if not item_key or not isinstance(value, dict):
            continue
        safe[item_key] = {
            "enabled": bool(value.get("enabled", True)),
            "offsets_min": _social_calendar_reminder_normalize_offsets(value.get("offsets_min") if isinstance(value.get("offsets_min"), list) else []),
            "updated_at": str(value.get("updated_at") or datetime.utcnow().isoformat()).strip(),
        }
    _set_system_setting(
        db,
        _social_calendar_reminder_store_key(user_id),
        json.dumps(safe, ensure_ascii=False),
    )


def _social_calendar_event_reminder_key(event_id: int, actor_key: str) -> str:
    return f"e:{max(0, int(event_id or 0))}:a:{str(actor_key or '').strip().lower()}"


def _social_calendar_reminder_settings_get(
    db: Session,
    *,
    user_id: int,
    event_id: int,
    actor_key: str,
) -> dict[str, Any]:
    all_rows = _social_calendar_reminders_load_user(db, int(user_id))
    key = _social_calendar_event_reminder_key(event_id, actor_key)
    row = all_rows.get(key) if isinstance(all_rows, dict) else None
    offsets = _social_calendar_reminder_normalize_offsets(row.get("offsets_min") if isinstance(row, dict) else None)
    enabled = bool(row.get("enabled", True)) if isinstance(row, dict) else True
    return {
        "event_id": int(event_id or 0),
        "enabled": enabled,
        "offsets_min": offsets,
        "available_offsets_min": list(SOCIAL_CALENDAR_REMINDER_ALLOWED_OFFSETS_MIN),
    }


def _social_calendar_reminder_settings_put(
    db: Session,
    *,
    user_id: int,
    event_id: int,
    actor_key: str,
    enabled: bool,
    offsets_min: list[int] | None,
) -> dict[str, Any]:
    all_rows = _social_calendar_reminders_load_user(db, int(user_id))
    key = _social_calendar_event_reminder_key(event_id, actor_key)
    all_rows[key] = {
        "enabled": bool(enabled),
        "offsets_min": _social_calendar_reminder_normalize_offsets(offsets_min),
        "updated_at": datetime.utcnow().isoformat(),
    }
    _social_calendar_reminders_save_user(db, int(user_id), all_rows)
    return _social_calendar_reminder_settings_get(
        db,
        user_id=int(user_id),
        event_id=int(event_id),
        actor_key=str(actor_key or ""),
    )


def _social_calendar_reminder_settings_remove_event(db: Session, *, user_id: int, event_id: int) -> None:
    all_rows = _social_calendar_reminders_load_user(db, int(user_id))
    if not all_rows:
        return
    prefix = f"e:{max(0, int(event_id or 0))}:a:"
    changed = False
    for key in list(all_rows.keys()):
        if str(key).startswith(prefix):
            all_rows.pop(key, None)
            changed = True
    if changed:
        _social_calendar_reminders_save_user(db, int(user_id), all_rows)

def _social_google_extract_email(value: Any) -> str:
    text = str(value or "").strip().lower()
    if not text or "@" not in text:
        return ""
    match = re.search(r"[a-z0-9._%+\-]+@[a-z0-9.\-]+\.[a-z]{2,}", text)
    return match.group(0) if match else ""


def _social_google_fetch_account_email(access_token: str) -> str:
    token = str(access_token or "").strip()
    if not token:
        return ""
    req = urllib.request.Request(
        "https://www.googleapis.com/calendar/v3/users/me/calendarList?minAccessRole=reader&maxResults=50",
        headers={
            "Authorization": f"Bearer {token}",
            "Accept": "application/json",
            "User-Agent": "SEO-WIBE/1.0",
        },
        method="GET",
    )
    try:
        with urllib.request.urlopen(req, timeout=25) as response:
            payload = json.loads(response.read().decode("utf-8", errors="ignore"))
    except Exception:
        return ""
    items = payload.get("items") if isinstance(payload, dict) else []
    if not isinstance(items, list):
        return ""
    for primary_only in (True, False):
        for item in items:
            if not isinstance(item, dict):
                continue
            if primary_only and not bool(item.get("primary")):
                continue
            for candidate in (
                item.get("id"),
                item.get("summaryOverride"),
                item.get("summary"),
                item.get("description"),
            ):
                email = _social_google_extract_email(candidate)
                if email:
                    return email
    return ""


def _social_google_refresh_access_token(
    db: Session,
    *,
    store_key: str,
    fallback_keys: list[str] | None = None,
    refresh_token: str,
    client_id: str,
    client_secret: str,
) -> dict[str, Any]:
    form = urllib.parse.urlencode(
        {
            "client_id": client_id,
            "client_secret": client_secret,
            "refresh_token": refresh_token,
            "grant_type": "refresh_token",
        }
    ).encode("utf-8")
    req = urllib.request.Request(
        "https://oauth2.googleapis.com/token",
        data=form,
        headers={"Content-Type": "application/x-www-form-urlencoded"},
        method="POST",
    )
    try:
        with urllib.request.urlopen(req, timeout=25) as response:
            data = json.loads(response.read().decode("utf-8", errors="ignore"))
    except Exception as exc:
        raise HTTPException(status_code=400, detail=f"Google refresh token error: {str(exc or '')[:220]}")
    access_token = str(data.get("access_token") or "").strip()
    if not access_token:
        raise HTTPException(status_code=400, detail="Google OAuth did not return access_token while refreshing")
    expires_in = int(data.get("expires_in") or 3600)
    next_payload = _social_google_tokens_get(db, store_key, fallback_keys=fallback_keys)
    next_payload["access_token"] = access_token
    next_payload["token_type"] = str(data.get("token_type") or next_payload.get("token_type") or "Bearer")
    next_payload["scope"] = str(data.get("scope") or next_payload.get("scope") or "")
    next_payload["expires_at"] = int(time.time()) + max(60, expires_in)
    if not str(next_payload.get("refresh_token") or "").strip():
        next_payload["refresh_token"] = refresh_token
    if not str(next_payload.get("account_email") or "").strip():
        next_payload["account_email"] = _social_google_fetch_account_email(access_token)
    return _social_google_tokens_save(db, store_key, next_payload)


def _social_google_event_dt(value: dict[str, Any] | None, *, is_end: bool = False) -> datetime | None:
    node = value if isinstance(value, dict) else {}
    date_time = str(node.get("dateTime") or "").strip()
    if date_time:
        try:
            parsed = datetime.fromisoformat(date_time.replace("Z", "+00:00"))
            return _social_localize_dt(parsed)
        except Exception:
            return None
    date_only = str(node.get("date") or "").strip()
    if not date_only:
        return None
    try:
        base = datetime.strptime(date_only, "%Y-%m-%d")
    except Exception:
        return None
    if is_end:
        return base + timedelta(days=1)
    return base


def _social_google_fetch_primary_events(
    db: Session,
    *,
    store_key: str,
    fallback_keys: list[str] | None,
    client_id: str,
    client_secret: str,
) -> tuple[list[dict[str, Any]], list[str]]:
    tokens = _social_google_tokens_get(db, store_key, fallback_keys=fallback_keys)
    access_token = str(tokens.get("access_token") or "").strip()
    refresh_token = str(tokens.get("refresh_token") or "").strip()
    expires_at = int(tokens.get("expires_at") or 0)
    if (not access_token) or (expires_at > 0 and expires_at <= int(time.time()) + 45):
        if not refresh_token:
            raise HTTPException(status_code=400, detail="Google Calendar РЅРµ РїРѕРґРєР»СЋС‡РµРЅ. РќР°Р¶РјРёС‚Рµ В«РџРѕРґРєР»СЋС‡РёС‚СЊ GoogleВ»." )
        refreshed = _social_google_refresh_access_token(
            db,
            store_key=store_key,
            fallback_keys=fallback_keys,
            refresh_token=refresh_token,
            client_id=client_id,
            client_secret=client_secret,
        )
        access_token = str(refreshed.get("access_token") or "").strip()
        if not access_token:
            raise HTTPException(status_code=400, detail="РќРµ СѓРґР°Р»РѕСЃСЊ РѕР±РЅРѕРІРёС‚СЊ Google access_token")
        tokens = refreshed

    if not str(tokens.get("account_email") or "").strip():
        account_email = _social_google_fetch_account_email(access_token)
        if account_email:
            tokens["account_email"] = account_email
            _social_google_tokens_save(db, store_key, tokens)

    time_min = (datetime.utcnow() - timedelta(days=180)).replace(microsecond=0).isoformat() + "Z"
    time_max = (datetime.utcnow() + timedelta(days=365)).replace(microsecond=0).isoformat() + "Z"
    out: list[dict[str, Any]] = []
    warnings: list[str] = []
    page_token = ""
    page_count = 0
    while page_count < 6:
        params_map = {
            "singleEvents": "true",
            "orderBy": "startTime",
            "maxResults": "2500",
            "timeMin": time_min,
            "timeMax": time_max,
        }
        if page_token:
            params_map["pageToken"] = page_token
        params = urllib.parse.urlencode(params_map)
        req = urllib.request.Request(
            f"https://www.googleapis.com/calendar/v3/calendars/primary/events?{params}",
            headers={
                "Authorization": f"Bearer {access_token}",
                "Accept": "application/json",
                "User-Agent": "SEO-WIBE/1.0",
            },
            method="GET",
        )
        try:
            with urllib.request.urlopen(req, timeout=25) as response:
                payload = json.loads(response.read().decode("utf-8", errors="ignore"))
        except Exception as exc:
            raise HTTPException(status_code=400, detail=f"РќРµ СѓРґР°Р»РѕСЃСЊ Р·Р°РіСЂСѓР·РёС‚СЊ СЃРѕР±С‹С‚РёСЏ Google Calendar: {str(exc or '')[:220]}")
        rows = payload.get("items") if isinstance(payload, dict) else []
        if not isinstance(rows, list):
            rows = []
        for row in rows:
            if not isinstance(row, dict):
                continue
            uid = str(row.get("id") or "").strip()
            if not uid:
                continue
            start_at = _social_google_event_dt(row.get("start"), is_end=False)
            end_at = _social_google_event_dt(row.get("end"), is_end=True)
            if not start_at:
                continue
            summary = str(row.get("summary") or "").strip() or "Google event"
            description = str(row.get("description") or "").strip()
            location = str(row.get("location") or "").strip()
            details_parts = [description]
            if location:
                details_parts.append(f"Location: {location}")
            details_text = "\n".join([x for x in details_parts if x]).strip()
            out.append(
                {
                    "uid": uid,
                    "title": summary[:255],
                    "details": details_text[:5000],
                    "start_at": start_at,
                    "end_at": end_at,
                }
            )
        page_token = str(payload.get("nextPageToken") or "").strip() if isinstance(payload, dict) else ""
        if not page_token:
            break
        page_count += 1
    if page_token:
        warnings.append("google_events_truncated")
    return out, warnings


def _social_calendar_redirect_url(public_base: str, *, mobile_app: bool, connected: bool = False, error: str = "") -> str:
    params = {
        "tab": "social",
        "social_subtab": "calendar",
    }
    if mobile_app:
        params["mobile_app"] = "1"
    if connected:
        params["google_oauth_connected"] = "1"
        params["auto_google_sync"] = "1"
    if error:
        params["google_oauth_error"] = str(error or "").strip()[:180]
    path = "/mobile" if mobile_app else "/"
    return f"{public_base}{path}?{urllib.parse.urlencode(params)}"


def _social_calendar_apk_deep_link(*, connected: bool = False, error: str = "") -> str:
    params = {
        "mobile_app": "1",
        "tab": "social",
        "social_subtab": "calendar",
    }
    if connected:
        params["google_oauth_connected"] = "1"
        params["auto_google_sync"] = "1"
    if error:
        params["google_oauth_error"] = str(error or "").strip()[:180]
    return f"seowibe://open?{urllib.parse.urlencode(params)}"


def _social_google_oauth_finish_response(public_base: str, *, return_target: str, connected: bool = False, error: str = "") -> Response:
    mobile_app = str(return_target or "").strip().lower() == "apk"
    redirect_url = _social_calendar_redirect_url(public_base, mobile_app=mobile_app, connected=connected, error=error)
    if not mobile_app:
        return RedirectResponse(url=redirect_url, status_code=302)
    deep_link = _social_calendar_apk_deep_link(connected=connected, error=error)
    title = "SEO WIBE"
    message = "Возвращаем вас в приложение..." if connected else "Возвращаем вас в SEO WIBE..."
    html = f"""<!doctype html>
<html lang=\"ru\">
<head>
  <meta charset=\"utf-8\">
  <meta name=\"viewport\" content=\"width=device-width, initial-scale=1\">
  <title>{title}</title>
</head>
<body style=\"font-family:system-ui,-apple-system,Segoe UI,sans-serif;padding:32px;color:#10233d;background:#f5f8fc;\">
  <h1 style=\"margin:0 0 12px;font-size:22px;\">{title}</h1>
  <p style=\"margin:0 0 18px;font-size:15px;\">{message}</p>
  <p style=\"margin:0;font-size:14px;\"><a href={json.dumps(deep_link)} style=\"color:#0f7ad7;\">Открыть приложение</a></p>
  <script>
    (function () {{
      var deepLink = {json.dumps(deep_link)};
      var webFallback = {json.dumps(redirect_url)};
      try {{ window.location.replace(deepLink); }} catch (_) {{}}
      window.setTimeout(function () {{
        try {{ window.location.replace(webFallback); }} catch (_) {{}}
      }}, 1400);
    }})();
  </script>
</body>
</html>"""
    return HTMLResponse(content=html, status_code=200)

@router.get("/social/calendar/events", response_model=list[SocialCalendarEventOut])
def social_calendar_events(
    date_from: str = "",
    date_to: str = "",
    user: User = Depends(get_current_user),
    db: Session = Depends(get_db),
):
    ensure_module_enabled(db, user, "social_hub")
    actor_key, _, _ = _social_actor_identity(db, user)
    left = _social_parse_dt(date_from) or (datetime.utcnow() - timedelta(days=180))
    right = _social_parse_dt(date_to) or (datetime.utcnow() + timedelta(days=365))
    rows = db.scalars(
        select(SocialCalendarEvent)
        .where(
            SocialCalendarEvent.user_id == user.id,
            SocialCalendarEvent.start_at >= left,
            SocialCalendarEvent.start_at <= right,
            or_(SocialCalendarEvent.is_public.is_(True), SocialCalendarEvent.actor_key == actor_key),
        )
        .order_by(SocialCalendarEvent.start_at.asc(), SocialCalendarEvent.id.asc())
    ).all()
    return [
        SocialCalendarEventOut(
            id=int(row.id),
            title=str(row.title or ""),
            details=str(row.details or ""),
            start_at=row.start_at.isoformat() if row.start_at else "",
            end_at=row.end_at.isoformat() if row.end_at else None,
            created_at=row.created_at.isoformat() if row.created_at else "",
            is_public=bool(row.is_public),
        )
        for row in rows
    ]


@router.get("/social/calendar/events/{event_id}/reminders", response_model=SocialCalendarReminderSettingsOut)
def social_calendar_event_reminders_get(
    event_id: int,
    user: User = Depends(get_current_user),
    db: Session = Depends(get_db),
):
    ensure_module_enabled(db, user, "social_hub")
    actor_key, _, _ = _social_actor_identity(db, user)
    row = db.get(SocialCalendarEvent, int(event_id or 0))
    if not row or int(row.user_id or 0) != int(user.id):
        raise HTTPException(status_code=404, detail="РЎРѕР±С‹С‚РёРµ РЅРµ РЅР°Р№РґРµРЅРѕ")
    creator_key = str(row.actor_key or "").strip().lower()
    current_actor_key = str(actor_key or "").strip().lower()
    if not bool(row.is_public) and creator_key != current_actor_key:
        raise HTTPException(status_code=403, detail="РќРµС‚ РґРѕСЃС‚СѓРїР° Рє СЃРѕР±С‹С‚РёСЋ")
    payload = _social_calendar_reminder_settings_get(
        db,
        user_id=int(user.id),
        event_id=int(row.id),
        actor_key=current_actor_key,
    )
    return SocialCalendarReminderSettingsOut(**payload)


@router.put("/social/calendar/events/{event_id}/reminders", response_model=SocialCalendarReminderSettingsOut)
def social_calendar_event_reminders_put(
    event_id: int,
    payload: SocialCalendarReminderSettingsIn,
    user: User = Depends(get_current_user),
    db: Session = Depends(get_db),
):
    ensure_module_enabled(db, user, "social_hub")
    actor_key, _, _ = _social_actor_identity(db, user)
    row = db.get(SocialCalendarEvent, int(event_id or 0))
    if not row or int(row.user_id or 0) != int(user.id):
        raise HTTPException(status_code=404, detail="РЎРѕР±С‹С‚РёРµ РЅРµ РЅР°Р№РґРµРЅРѕ")
    creator_key = str(row.actor_key or "").strip().lower()
    current_actor_key = str(actor_key or "").strip().lower()
    if not bool(row.is_public) and creator_key != current_actor_key:
        raise HTTPException(status_code=403, detail="РќРµС‚ РґРѕСЃС‚СѓРїР° Рє СЃРѕР±С‹С‚РёСЋ")
    saved = _social_calendar_reminder_settings_put(
        db,
        user_id=int(user.id),
        event_id=int(row.id),
        actor_key=current_actor_key,
        enabled=bool(payload.enabled),
        offsets_min=[int(x) for x in (payload.offsets_min or [])],
    )
    db.commit()
    return SocialCalendarReminderSettingsOut(**saved)


@router.get("/social/calendar/google-oauth/status")
def social_calendar_google_oauth_status(
    request: Request,
    user: User = Depends(get_current_user),
    db: Session = Depends(get_db),
):
    ensure_module_enabled(db, user, "social_hub")
    actor_ctx = _social_actor_store_context(db, user)
    tokens = _social_google_tokens_get(db, actor_ctx["store_key"], fallback_keys=actor_ctx["fallback_keys"])
    access_token = str(tokens.get("access_token") or "").strip()
    refresh_token = str(tokens.get("refresh_token") or "").strip()
    expires_at = int(tokens.get("expires_at") or 0)
    client_id, client_secret, redirect_uri = _social_google_oauth_config(request)
    sync_status = _social_calendar_sync_status_get(
        db,
        actor_ctx["store_key"],
        fallback_keys=actor_ctx["fallback_keys"],
    )
    setup_parts: list[str] = []
    if not client_id:
        setup_parts.append("Р”РѕР±Р°РІСЊС‚Рµ SEO_WIBE_GOOGLE_CLIENT_ID")
    if not client_secret:
        setup_parts.append("Р”РѕР±Р°РІСЊС‚Рµ SEO_WIBE_GOOGLE_CLIENT_SECRET")
    if not (client_id and client_secret):
        setup_parts.append(f"Redirect URI РґР»СЏ Google OAuth: {redirect_uri}")
    if not str(os.getenv("SEO_WIBE_PUBLIC_BASE_URL") or "").strip():
        setup_parts.append("Р РµРєРѕРјРµРЅРґСѓРµС‚СЃСЏ SEO_WIBE_PUBLIC_BASE_URL=https://seowibe.ru")
    return {
        "connected": bool(access_token or refresh_token),
        "oauth_configured": bool(client_id and client_secret),
        "required_redirect_uri": redirect_uri,
        "public_base_url": _social_public_base_url(request),
        "redirect_uri": redirect_uri,
        "expires_at": expires_at,
        "scope": str(tokens.get("scope") or "").strip(),
        "has_refresh_token": bool(refresh_token),
        "account_email": str(tokens.get("account_email") or "").strip().lower(),
        "actor_email": str(actor_ctx.get("actor_email") or "").strip().lower(),
        "setup_hint": "; ".join(setup_parts)[:320],
        "last_sync_at": str(sync_status.get("last_sync_at") or "").strip(),
        "last_sync_source": str(sync_status.get("last_sync_source") or "").strip(),
        "last_sync_state": str(sync_status.get("last_sync_state") or "idle").strip(),
        "last_sync_ok": bool(sync_status.get("last_sync_ok")),
        "last_sync_error": str(sync_status.get("last_sync_error") or "").strip(),
        "last_sync_summary": {
            "imported": max(0, int(sync_status.get("imported") or 0)),
            "updated": max(0, int(sync_status.get("updated") or 0)),
            "deleted": max(0, int(sync_status.get("deleted") or 0)),
            "skipped": max(0, int(sync_status.get("skipped") or 0)),
            "warning_count": max(0, int(sync_status.get("warning_count") or 0)),
            "warnings": [str(x or "").strip() for x in (sync_status.get("warnings") or []) if str(x or "").strip()][:12],
        },
    }


@router.get("/social/calendar/google-oauth/start")
def social_calendar_google_oauth_start(
    request: Request,
    return_target: str = "",
    user: User = Depends(get_current_user),
    db: Session = Depends(get_db),
):
    ensure_module_enabled(db, user, "social_hub")
    client_id, _, redirect_uri = _social_google_oauth_config(request)
    if not client_id:
        raise HTTPException(
            status_code=400,
            detail="Google OAuth РЅРµ РЅР°СЃС‚СЂРѕРµРЅ: Р·Р°РґР°Р№С‚Рµ SEO_WIBE_GOOGLE_CLIENT_ID Рё SEO_WIBE_GOOGLE_CLIENT_SECRET РЅР° СЃРµСЂРІРµСЂРµ.",
        )
    actor_ctx = _social_actor_store_context(db, user)
    safe_target = str(return_target or "").strip().lower()
    if safe_target not in {"apk", "web"}:
        referer = str(request.headers.get("referer") or "")
        safe_target = "apk" if ("mobile_app=1" in referer or "/mobile" in referer) else "web"
    state = _social_google_oauth_state_for_actor(
        user_id=int(user.id),
        store_key=str(actor_ctx["store_key"]),
        return_target=safe_target,
    )
    params = urllib.parse.urlencode(
        {
            "client_id": client_id,
            "redirect_uri": redirect_uri,
            "response_type": "code",
            "scope": SOCIAL_GOOGLE_OAUTH_SCOPE,
            "access_type": "offline",
            "prompt": "consent",
            "include_granted_scopes": "true",
            "state": state,
        }
    )
    url = f"https://accounts.google.com/o/oauth2/v2/auth?{params}"
    _audit(
        db,
        user,
        action="social_calendar_google_oauth_started",
        details=f"redirect_uri={redirect_uri};target={safe_target};store_key={actor_ctx['store_key']}",
        module_code="social_hub",
        entity_type="calendar_oauth",
        status="ok",
        request=request,
    )
    db.commit()
    return {"ok": True, "url": url}


@router.get("/social/calendar/google-oauth/callback")
def social_calendar_google_oauth_callback(
    request: Request,
    code: str = "",
    state: str = "",
    error: str = "",
    db: Session = Depends(get_db),
):
    public_base = _social_public_base_url(request)
    state_payload = _social_google_oauth_state_from_token(state)
    return_target = str(state_payload.get("return_target") or "web").strip().lower()
    if error:
        return _social_google_oauth_finish_response(public_base, return_target=return_target, error=str(error)[:180])

    user_id = max(0, int(state_payload.get("user_id") or 0))
    store_key = str(state_payload.get("store_key") or "").strip() or f"u:{user_id}"
    if user_id <= 0:
        return _social_google_oauth_finish_response(public_base, return_target=return_target, error="invalid_state")
    user = db.get(User, user_id)
    if not user:
        return _social_google_oauth_finish_response(public_base, return_target=return_target, error="user_not_found")
    client_id, client_secret, redirect_uri = _social_google_oauth_config(request)
    if not client_id or not client_secret:
        return _social_google_oauth_finish_response(public_base, return_target=return_target, error="oauth_not_configured")
    form = urllib.parse.urlencode(
        {
            "code": str(code or "").strip(),
            "client_id": client_id,
            "client_secret": client_secret,
            "redirect_uri": redirect_uri,
            "grant_type": "authorization_code",
        }
    ).encode("utf-8")
    token_req = urllib.request.Request(
        "https://oauth2.googleapis.com/token",
        data=form,
        headers={"Content-Type": "application/x-www-form-urlencoded"},
        method="POST",
    )
    try:
        with urllib.request.urlopen(token_req, timeout=25) as response:
            token_data = json.loads(response.read().decode("utf-8", errors="ignore"))
    except Exception:
        return _social_google_oauth_finish_response(public_base, return_target=return_target, error="token_exchange_failed")
    access_token = str(token_data.get("access_token") or "").strip()
    refresh_token = str(token_data.get("refresh_token") or "").strip()
    if not access_token:
        return _social_google_oauth_finish_response(public_base, return_target=return_target, error="missing_access_token")
    expires_in = int(token_data.get("expires_in") or 3600)
    existing = _social_google_tokens_get(db, store_key, fallback_keys=[str(int(user_id))])
    payload = {
        "access_token": access_token,
        "refresh_token": refresh_token or str(existing.get("refresh_token") or "").strip(),
        "token_type": str(token_data.get("token_type") or "Bearer"),
        "scope": str(token_data.get("scope") or SOCIAL_GOOGLE_OAUTH_SCOPE),
        "expires_at": int(time.time()) + max(60, expires_in),
        "account_email": _social_google_fetch_account_email(access_token) or str(existing.get("account_email") or "").strip().lower(),
    }
    _social_google_tokens_save(db, store_key, payload)
    _audit(
        db,
        user,
        action="social_calendar_google_oauth_connected",
        details=(
            f"scope={payload.get('scope')};has_refresh={'1' if payload.get('refresh_token') else '0'};"
            f"store_key={store_key};account={payload.get('account_email') or '-'}"
        ),
        module_code="social_hub",
        entity_type="calendar_oauth",
        status="ok",
        request=request,
    )
    db.commit()
    return _social_google_oauth_finish_response(public_base, return_target=return_target, connected=True)


@router.post("/social/calendar/google-sync", response_model=SocialCalendarGoogleSyncOut)
def social_calendar_google_sync(
    payload: SocialCalendarGoogleSyncIn,
    request: Request,
    user: User = Depends(get_current_user),
    db: Session = Depends(get_db),
):
    ensure_module_enabled(db, user, "social_hub")
    actor_ctx = _social_actor_store_context(db, user)
    actor_key = str(actor_ctx["actor_key"])
    actor_store_key = str(actor_ctx["store_key"])
    actor_fallback_keys = list(actor_ctx.get("fallback_keys") or [])
    url_raw = str(payload.ical_url or "").strip()
    source_kind = "ics_url" if url_raw else "google_oauth"
    warnings: list[str] = []
    parsed_events: list[dict[str, Any]] = []
    source_hash = ""
    imported = 0
    updated = 0
    deleted = 0
    skipped = 0
    try:
        if url_raw:
            url_value = url_raw
            if url_value.lower().startswith("webcal://"):
                url_value = "https://" + url_value[9:]
            parsed_url = urllib.parse.urlparse(url_value)
            if parsed_url.scheme not in {"http", "https"}:
                raise HTTPException(status_code=400, detail="РџРѕРґРґРµСЂР¶РёРІР°СЋС‚СЃСЏ С‚РѕР»СЊРєРѕ СЃСЃС‹Р»РєРё http/https (РёР»Рё webcal)")
            try:
                req = urllib.request.Request(url_value, headers={"User-Agent": "SEO-WIBE/1.0"})
                with urllib.request.urlopen(req, timeout=25) as response:
                    raw_bytes = response.read(2 * 1024 * 1024)
            except urllib.error.HTTPError as exc:
                raise HTTPException(
                    status_code=400,
                    detail=(
                        f"ICS URL РІРµСЂРЅСѓР» HTTP {int(exc.code or 0)}. РџСЂРѕРІРµСЂСЊС‚Рµ, С‡С‚Рѕ РєР°Р»РµРЅРґР°СЂСЊ РґРѕСЃС‚СѓРїРµРЅ РїРѕ РїСѓР±Р»РёС‡РЅРѕР№ РёР»Рё secret ICS СЃСЃС‹Р»РєРµ."
                    ),
                )
            except urllib.error.URLError as exc:
                reason = str(getattr(exc, "reason", exc) or "").strip()
                raise HTTPException(status_code=400, detail=f"РќРµ СѓРґР°Р»РѕСЃСЊ РїРѕРґРєР»СЋС‡РёС‚СЊСЃСЏ Рє ICS URL: {reason[:220]}")
            except Exception as exc:
                raise HTTPException(status_code=400, detail=f"РќРµ СѓРґР°Р»РѕСЃСЊ Р·Р°РіСЂСѓР·РёС‚СЊ ICS: {str(exc or '')[:220]}")

            text_payload = ""
            for encoding in ("utf-8", "utf-8-sig", "cp1251", "latin-1"):
                try:
                    text_payload = raw_bytes.decode(encoding)
                    break
                except Exception:
                    continue
            if not text_payload:
                raise HTTPException(status_code=400, detail="РќРµ СѓРґР°Р»РѕСЃСЊ РґРµРєРѕРґРёСЂРѕРІР°С‚СЊ ICS С„Р°Р№Р»")
            parsed_events, parse_warnings = _social_parse_ical_events(text_payload)
            warnings.extend(parse_warnings)
            source_hash = hashlib.sha1(url_value.encode("utf-8")).hexdigest()[:16]
        else:
            client_id, client_secret, _ = _social_google_oauth_config(request)
            if not client_id or not client_secret:
                raise HTTPException(
                    status_code=400,
                    detail="Google OAuth РЅРµ РЅР°СЃС‚СЂРѕРµРЅ РЅР° СЃРµСЂРІРµСЂРµ. Р”РѕР±Р°РІСЊС‚Рµ SEO_WIBE_GOOGLE_CLIENT_ID Рё SEO_WIBE_GOOGLE_CLIENT_SECRET.",
                )
            parsed_events, oauth_warnings = _social_google_fetch_primary_events(
                db,
                store_key=actor_store_key,
                fallback_keys=actor_fallback_keys,
                client_id=client_id,
                client_secret=client_secret,
            )
            warnings.extend(oauth_warnings)
            source_hash = hashlib.sha1(f"oauth-primary:{actor_store_key}".encode("utf-8")).hexdigest()[:16]

        if not parsed_events:
            _social_calendar_sync_mark(
                db,
                store_key=actor_store_key,
                source=source_kind,
                state="empty",
                warnings=warnings,
            )
            db.commit()
            return SocialCalendarGoogleSyncOut(ok=True, imported=0, updated=0, deleted=0, skipped=0, warnings=warnings)

        marker_prefix = f"[[gcal_sync source={source_hash} "
        marker_pattern = re.compile(r"\[\[gcal_sync source=([a-f0-9]{16}) uid=([^\]\s]+)\]\]")

        existing_rows = db.scalars(
            select(SocialCalendarEvent).where(
                SocialCalendarEvent.user_id == int(user.id),
                SocialCalendarEvent.actor_key == actor_key,
                SocialCalendarEvent.details.like(f"%{marker_prefix}%"),
            )
        ).all()
        existing_by_uid: dict[str, SocialCalendarEvent] = {}
        for row in existing_rows:
            details_text = str(row.details or "")
            match = marker_pattern.search(details_text)
            if not match:
                continue
            if match.group(1) != source_hash:
                continue
            uid_text = urllib.parse.unquote(str(match.group(2) or "").strip())
            if uid_text:
                existing_by_uid[uid_text] = row

        seen_uids: set[str] = set()
        safe_is_public = bool(payload.is_public)
        for event in parsed_events[:1800]:
            uid = str(event.get("uid") or "").strip()
            if not uid:
                skipped += 1
                continue
            if uid in seen_uids:
                continue
            seen_uids.add(uid)
            marker_uid = urllib.parse.quote(uid, safe="")[:220]
            marker = f"[[gcal_sync source={source_hash} uid={marker_uid}]]"
            details = str(event.get("details") or "").strip()
            stored_details = f"{details}\n\n{marker}" if details else marker
            start_at = event.get("start_at") if isinstance(event.get("start_at"), datetime) else None
            end_at = event.get("end_at") if isinstance(event.get("end_at"), datetime) else None
            if not start_at:
                skipped += 1
                continue
            row = existing_by_uid.get(uid)
            if row is None:
                db.add(
                    SocialCalendarEvent(
                        user_id=int(user.id),
                        actor_key=actor_key,
                        is_public=safe_is_public,
                        title=str(event.get("title") or "Google event")[:255],
                        details=stored_details[:5000],
                        start_at=start_at,
                        end_at=end_at,
                    )
                )
                imported += 1
                continue
            changed = False
            next_title = str(event.get("title") or row.title or "Google event")[:255]
            if str(row.title or "") != next_title:
                row.title = next_title
                changed = True
            if str(row.details or "") != stored_details[:5000]:
                row.details = stored_details[:5000]
                changed = True
            if bool(row.is_public) != safe_is_public:
                row.is_public = safe_is_public
                changed = True
            if row.start_at != start_at:
                row.start_at = start_at
                changed = True
            if row.end_at != end_at:
                row.end_at = end_at
                changed = True
            if changed:
                updated += 1
            else:
                skipped += 1

        if bool(payload.replace_source_events):
            for uid, row in existing_by_uid.items():
                if uid in seen_uids:
                    continue
                db.delete(row)
                deleted += 1

        state = "partial" if warnings else "ok"
        _audit(
            db,
            user,
            action="social_calendar_google_sync",
            details=(
                f"source={source_hash};store_key={actor_store_key};imported={imported};updated={updated};"
                f"deleted={deleted};skipped={skipped};events={len(parsed_events)}"
            ),
            module_code="social_hub",
            entity_type="calendar_sync",
            status="ok" if state == "ok" else "partial",
            request=request,
        )
        _social_calendar_sync_mark(
            db,
            store_key=actor_store_key,
            source=source_kind,
            state=state,
            imported=imported,
            updated=updated,
            deleted=deleted,
            skipped=skipped,
            warnings=warnings,
        )
        db.commit()
        return SocialCalendarGoogleSyncOut(
            ok=True,
            imported=imported,
            updated=updated,
            deleted=deleted,
            skipped=skipped,
            warnings=warnings,
        )
    except HTTPException as exc:
        db.rollback()
        _social_calendar_sync_mark(
            db,
            store_key=actor_store_key,
            source=source_kind,
            state="error",
            warnings=warnings,
            error=str(exc.detail or "").strip(),
        )
        db.commit()
        raise
    except Exception as exc:
        db.rollback()
        detail = f"РќРµ СѓРґР°Р»РѕСЃСЊ СЃРёРЅС…СЂРѕРЅРёР·РёСЂРѕРІР°С‚СЊ РєР°Р»РµРЅРґР°СЂСЊ: {str(exc or '')[:220]}"
        _social_calendar_sync_mark(
            db,
            store_key=actor_store_key,
            source=source_kind,
            state="error",
            warnings=warnings,
            error=detail,
        )
        db.commit()
        raise HTTPException(status_code=400, detail=detail)


@router.post("/social/calendar/events", response_model=SocialCalendarEventOut)
def social_calendar_create_event(
    payload: SocialCalendarEventIn,
    user: User = Depends(get_current_user),
    db: Session = Depends(get_db),
):
    ensure_module_enabled(db, user, "social_hub")
    actor_key, _, _ = _social_actor_identity(db, user)
    start_at = _social_parse_dt(payload.start_at)
    if not start_at:
        raise HTTPException(status_code=400, detail="РќРµРєРѕСЂСЂРµРєС‚РЅР°СЏ РґР°С‚Р° РЅР°С‡Р°Р»Р°")
    end_at = _social_parse_dt(payload.end_at)
    row = SocialCalendarEvent(
        user_id=user.id,
        actor_key=actor_key,
        is_public=bool(payload.is_public),
        title=str(payload.title or "").strip()[:255],
        details=str(payload.details or "")[:5000],
        start_at=start_at,
        end_at=end_at,
    )
    db.add(row)
    db.flush()
    _audit(
        db,
        user,
        action="social_calendar_event_created",
        details=json.dumps(
            {
                "event_id": int(row.id or 0),
                "title": str(row.title or "")[:200],
                "start_at": row.start_at.isoformat() if row.start_at else "",
                "is_public": bool(row.is_public),
            },
            ensure_ascii=False,
        ),
        module_code="social_hub",
        entity_type="calendar_event",
        entity_id=str(row.id or ""),
    )
    db.commit()
    return SocialCalendarEventOut(
        id=int(row.id),
        title=str(row.title or ""),
        details=str(row.details or ""),
        start_at=row.start_at.isoformat() if row.start_at else "",
        end_at=row.end_at.isoformat() if row.end_at else None,
        created_at=row.created_at.isoformat() if row.created_at else "",
        is_public=bool(row.is_public),
    )


@router.put("/social/calendar/events/{event_id}", response_model=SocialCalendarEventOut)
def social_calendar_update_event(
    event_id: int,
    payload: SocialCalendarEventIn,
    user: User = Depends(get_current_user),
    db: Session = Depends(get_db),
):
    ensure_module_enabled(db, user, "social_hub")
    actor_key, _, _ = _social_actor_identity(db, user)
    row = db.get(SocialCalendarEvent, event_id)
    if not row or int(row.user_id) != int(user.id):
        raise HTTPException(status_code=404, detail="РЎРѕР±С‹С‚РёРµ РЅРµ РЅР°Р№РґРµРЅРѕ")
    if str(row.actor_key or "") != actor_key and not _actor_is_owner(user):
        raise HTTPException(status_code=403, detail="РќРµС‚ РґРѕСЃС‚СѓРїР° Рє СЃРѕР±С‹С‚РёСЋ")
    start_at = _social_parse_dt(payload.start_at)
    if not start_at:
        raise HTTPException(status_code=400, detail="РќРµРєРѕСЂСЂРµРєС‚РЅР°СЏ РґР°С‚Р° РЅР°С‡Р°Р»Р°")
    row.title = str(payload.title or "").strip()[:255]
    row.details = str(payload.details or "")[:5000]
    row.start_at = start_at
    row.end_at = _social_parse_dt(payload.end_at)
    row.is_public = bool(payload.is_public)
    _audit(
        db,
        user,
        action="social_calendar_event_updated",
        details=json.dumps(
            {
                "event_id": int(row.id or 0),
                "title": str(row.title or "")[:200],
                "start_at": row.start_at.isoformat() if row.start_at else "",
                "is_public": bool(row.is_public),
            },
            ensure_ascii=False,
        ),
        module_code="social_hub",
        entity_type="calendar_event",
        entity_id=str(row.id or ""),
    )
    db.commit()
    return SocialCalendarEventOut(
        id=int(row.id),
        title=str(row.title or ""),
        details=str(row.details or ""),
        start_at=row.start_at.isoformat() if row.start_at else "",
        end_at=row.end_at.isoformat() if row.end_at else None,
        created_at=row.created_at.isoformat() if row.created_at else "",
        is_public=bool(row.is_public),
    )


@router.delete("/social/calendar/events/{event_id}", response_model=MessageOut)
def social_calendar_delete_event(
    event_id: int,
    user: User = Depends(get_current_user),
    db: Session = Depends(get_db),
):
    ensure_module_enabled(db, user, "social_hub")
    actor_key, _, _ = _social_actor_identity(db, user)
    row = db.get(SocialCalendarEvent, event_id)
    if not row or int(row.user_id) != int(user.id):
        raise HTTPException(status_code=404, detail="Event not found")
    if str(row.actor_key or "") != actor_key and not _actor_is_owner(user):
        raise HTTPException(status_code=403, detail="No access to event")
    _audit(
        db,
        user,
        action="social_calendar_event_deleted",
        details=json.dumps(
            {
                "event_id": int(row.id or 0),
                "title": str(row.title or "")[:200],
                "start_at": row.start_at.isoformat() if row.start_at else "",
                "is_public": bool(row.is_public),
            },
            ensure_ascii=False,
        ),
        module_code="social_hub",
        entity_type="calendar_event",
        entity_id=str(row.id or ""),
    )
    _social_calendar_reminder_settings_remove_event(db, user_id=int(user.id), event_id=int(row.id or 0))
    db.delete(row)
    db.commit()
    return MessageOut(message="Event deleted")

def _social_note_storage_dir() -> Path:
    static_root = Path(__file__).resolve().parent.parent / "static"
    target_dir = static_root / "uploads" / "social_notes"
    target_dir.mkdir(parents=True, exist_ok=True)
    return target_dir


def _social_note_clean_filename(raw: str) -> str:
    base = os.path.basename(str(raw or "").strip()) or "file"
    safe = re.sub(r"[^\w.\-]+", "_", base, flags=re.UNICODE).strip("._")
    if not safe:
        return "file"
    return safe[:180]


def _social_note_guess_ext(filename: str, content_type: str) -> str:
    ext = str(Path(filename or "").suffix or "").strip().lower()
    if ext and len(ext) <= 10:
        return ext
    guessed = mimetypes.guess_extension(str(content_type or "").split(";", 1)[0].strip().lower()) or ""
    if guessed == ".jpe":
        guessed = ".jpg"
    if guessed and len(guessed) <= 10:
        return guessed
    return ".bin"


_MOJIBAKE_TEXT_RE = re.compile(r"(?:\u0420[\u0400-\u04FF]|\u0421[\u0400-\u04FF]|\u0440[\u0450-\u045f]|\u0441[\u0400-\u040f\u0450-\u045f]|\u0432[\u0400-\u040f]|\u00d0.|\u00d1.)")
_KNOWN_MOJIBAKE_REPLACEMENTS: dict[str, str] = {
    "\u0420\u045c\u0420\u0455\u0420\u0406\u0420\u0455\u0420\xb5 \u0421\u0403\u0420\u0455\u0420\u0455\u0420\xb1\u0421\u2030\u0420\xb5\u0420\u0405\u0420\u0451\u0420\xb5": "\u041d\u043e\u0432\u043e\u0435 \u0441\u043e\u043e\u0431\u0449\u0435\u043d\u0438\u0435",
    "\u0420\xa0\u0421\u045a\u0420\xa0\u0421\u2022\u0420\xa0\u0420\u2020\u0420\xa0\u0421\u2022\u0420\xa0\u0412\xb5 \u0420\u040e\u0420\u0453\u0420\xa0\u0421\u2022\u0420\xa0\u0421\u2022\u0420\xa0\u0412\xb1\u0420\u040e\u0432\u0402\xb0\u0420\xa0\u0412\xb5\u0420\xa0\u0420\u2026\u0420\xa0\u0421\u2018\u0420\xa0\u0412\xb5": "РќРѕРІРѕРµ СЃРѕРѕР±С‰РµРЅРёРµ",
    "\u0421\u0402\u0421\u045f\u0432\u0402\u045a\u0420\u2039": "рџ“Ћ",
    "\u0440\u045f\u201c\u040b": "рџ“Ћ",
}

def _apply_known_mojibake_replacements(text: str) -> str:
    fixed = str(text or "")
    for bad, good in _KNOWN_MOJIBAKE_REPLACEMENTS.items():
        fixed = fixed.replace(bad, good)
    return fixed


def _mojibake_score(text: str) -> int:
    value = str(text or "")
    rare = sum(
        1
        for ch in value
        if (("\u0400" <= ch <= "\u040f") or ("\u0450" <= ch <= "\u045f")) and ch not in {"\u0401", "\u0451"}
    )
    return len(_MOJIBAKE_TEXT_RE.findall(value)) + value.count("\ufffd") * 3 + rare


def _cyrillic_score(text: str) -> int:
    value = str(text or "")
    return sum(1 for ch in value if ("\u0410" <= ch <= "\u044f") or ch in {"\u0401", "\u0451"})


def _decode_mojibake_text(value: Any) -> str:
    raw = _apply_known_mojibake_replacements(str(value or ""))
    if not raw:
        return ""
    if _mojibake_score(raw) <= 0:
        return raw

    def _decode_once(text: str) -> list[str]:
        out: list[str] = []
        for src_encoding in ("cp1251", "latin1", "cp1252"):
            for mode in ("strict", "ignore"):
                try:
                    candidate = text.encode(src_encoding, errors=mode).decode("utf-8", errors=mode)
                except Exception:
                    continue
                if not candidate or candidate == text:
                    continue
                if len(candidate) < max(4, int(len(text) * 0.6)):
                    continue
                out.append(candidate)
        return out

    candidates = [raw]
    seen = {raw}
    frontier = [raw]
    for _ in range(3):
        next_frontier: list[str] = []
        for item in frontier:
            for candidate in _decode_once(item):
                if candidate in seen:
                    continue
                seen.add(candidate)
                candidates.append(candidate)
                next_frontier.append(candidate)
        if not next_frontier:
            break
        frontier = next_frontier

    def _signal_score(text: str) -> int:
        emoji = sum(1 for ch in text if 0x1F300 <= ord(ch) <= 0x1FAFF)
        return _cyrillic_score(text) + emoji * 2

    best = raw
    best_score = _mojibake_score(raw)
    best_signal = _signal_score(raw)
    for candidate in candidates:
        score = _mojibake_score(candidate)
        signal = _signal_score(candidate)
        if score < best_score or (score == best_score and signal > best_signal + 1):
            best = candidate
            best_score = score
            best_signal = signal
    return _apply_known_mojibake_replacements(best)


def _social_note_delete_disk_file(url: str) -> None:
    safe = str(url or "").strip()
    prefix = "/static/uploads/social_notes/"
    if not safe.startswith(prefix):
        return
    filename = os.path.basename(safe)
    if not filename:
        return
    path = _social_note_storage_dir() / filename
    try:
        if path.exists():
            path.unlink()
    except Exception:
        pass


def _social_note_file_to_out(row: SocialNoteFile) -> SocialNoteFileOut:
    return SocialNoteFileOut(
        id=int(row.id),
        filename=str(row.filename or ""),
        url=str(row.url or ""),
        content_type=str(row.content_type or ""),
        size_bytes=int(row.size_bytes or 0),
        created_at=row.created_at.isoformat() if row.created_at else "",
    )


def _social_note_to_out(db: Session, row: SocialNote) -> SocialNoteOut:
    file_rows = db.scalars(
        select(SocialNoteFile)
        .where(SocialNoteFile.note_id == row.id)
        .order_by(SocialNoteFile.created_at.desc(), SocialNoteFile.id.desc())
    ).all()
    return SocialNoteOut(
        id=int(row.id),
        title=_decode_mojibake_text(str(row.title or "")),
        content=_decode_mojibake_text(str(row.content or "")),
        updated_at=row.updated_at.isoformat() if row.updated_at else "",
        files=[_social_note_file_to_out(x) for x in file_rows],
    )


@router.get("/social/notes", response_model=list[SocialNoteOut])
def social_notes_list(
    user: User = Depends(get_current_user),
    db: Session = Depends(get_db),
):
    ensure_module_enabled(db, user, "social_hub")
    actor_key, _, _ = _social_actor_identity(db, user)
    rows = db.scalars(
        select(SocialNote)
        .where(
            SocialNote.user_id == user.id,
            SocialNote.actor_key == actor_key,
        )
        .order_by(SocialNote.updated_at.desc(), SocialNote.id.desc())
    ).all()
    return [_social_note_to_out(db, row) for row in rows]


@router.post("/social/notes", response_model=SocialNoteOut)
def social_create_note(
    payload: SocialNoteIn,
    user: User = Depends(get_current_user),
    db: Session = Depends(get_db),
):
    ensure_module_enabled(db, user, "social_hub")
    actor_key, _, _ = _social_actor_identity(db, user)
    row = SocialNote(
        user_id=user.id,
        actor_key=actor_key,
        title=_decode_mojibake_text(payload.title or "").strip()[:255] or "РќРѕРІР°СЏ Р·Р°РјРµС‚РєР°",
        content=_decode_mojibake_text(payload.content or "")[:20000],
    )
    db.add(row)
    db.flush()
    _audit(
        db,
        user,
        action="social_note_created",
        details=json.dumps({"note_id": int(row.id or 0), "title": str(row.title or "")[:200]}, ensure_ascii=False),
        module_code="social_hub",
        entity_type="note",
        entity_id=str(row.id or ""),
    )
    db.commit()
    return _social_note_to_out(db, row)


@router.put("/social/notes/{note_id}", response_model=SocialNoteOut)
def social_update_note(
    note_id: int,
    payload: SocialNoteIn,
    user: User = Depends(get_current_user),
    db: Session = Depends(get_db),
):
    ensure_module_enabled(db, user, "social_hub")
    actor_key, _, _ = _social_actor_identity(db, user)
    row = db.get(SocialNote, note_id)
    if not row or int(row.user_id) != int(user.id) or str(row.actor_key or "") != actor_key:
        raise HTTPException(status_code=404, detail="Р—Р°РјРµС‚РєР° РЅРµ РЅР°Р№РґРµРЅР°")
    row.title = _decode_mojibake_text(payload.title or "").strip()[:255] or "Р‘РµР· РЅР°Р·РІР°РЅРёСЏ"
    row.content = _decode_mojibake_text(payload.content or "")[:20000]
    _audit(
        db,
        user,
        action="social_note_updated",
        details=json.dumps({"note_id": int(row.id or 0), "title": str(row.title or "")[:200]}, ensure_ascii=False),
        module_code="social_hub",
        entity_type="note",
        entity_id=str(row.id or ""),
    )
    db.commit()
    return _social_note_to_out(db, row)


@router.post("/social/notes/{note_id}/files", response_model=SocialNoteFileOut)
def social_note_upload_file(
    note_id: int,
    file: UploadFile = File(...),
    user: User = Depends(get_current_user),
    db: Session = Depends(get_db),
):
    ensure_module_enabled(db, user, "social_hub")
    actor_key, _, _ = _social_actor_identity(db, user)
    row = db.get(SocialNote, note_id)
    if not row or int(row.user_id) != int(user.id) or str(row.actor_key or "") != actor_key:
        raise HTTPException(status_code=404, detail="Р—Р°РјРµС‚РєР° РЅРµ РЅР°Р№РґРµРЅР°")
    if not file or not file.filename:
        raise HTTPException(status_code=400, detail="Р¤Р°Р№Р» РЅРµ РІС‹Р±СЂР°РЅ")
    raw = file.file.read()
    if not raw:
        raise HTTPException(status_code=400, detail="Р¤Р°Р№Р» РїСѓСЃС‚РѕР№")
    max_size = 12 * 1024 * 1024
    if len(raw) > max_size:
        raise HTTPException(status_code=400, detail="Р¤Р°Р№Р» СЃР»РёСЂРєРѕРј Р±РѕР»СЊСЂРѕР№ (РґРѕ 12 РњР‘)")
    original_name = _social_note_clean_filename(file.filename or "file")
    ext = _social_note_guess_ext(original_name, str(file.content_type or ""))
    storage_name = f"note-{int(row.id)}-{secrets.token_hex(8)}{ext}"
    path = _social_note_storage_dir() / storage_name
    path.write_bytes(raw)
    url = f"/static/uploads/social_notes/{storage_name}"
    file_row = SocialNoteFile(
        note_id=int(row.id),
        user_id=user.id,
        actor_key=actor_key,
        filename=original_name[:255],
        url=url[:500],
        content_type=str(file.content_type or "application/octet-stream")[:120],
        size_bytes=len(raw),
    )
    db.add(file_row)
    row.updated_at = datetime.utcnow()
    _audit(
        db,
        user,
        action="social_note_file_uploaded",
        details=json.dumps(
            {
                "note_id": int(row.id or 0),
                "file_name": str(file_row.filename or "")[:200],
                "size": int(file_row.size_bytes or 0),
            },
            ensure_ascii=False,
        ),
        module_code="social_hub",
        entity_type="note_file",
        entity_id=str(row.id or ""),
    )
    db.commit()
    return _social_note_file_to_out(file_row)


@router.delete("/social/notes/{note_id}/files/{file_id}", response_model=MessageOut)
def social_note_delete_file(
    note_id: int,
    file_id: int,
    user: User = Depends(get_current_user),
    db: Session = Depends(get_db),
):
    ensure_module_enabled(db, user, "social_hub")
    actor_key, _, _ = _social_actor_identity(db, user)
    row = db.get(SocialNote, note_id)
    if not row or int(row.user_id) != int(user.id) or str(row.actor_key or "") != actor_key:
        raise HTTPException(status_code=404, detail="Р—Р°РјРµС‚РєР° РЅРµ РЅР°Р№РґРµРЅР°")
    file_row = db.scalar(
        select(SocialNoteFile).where(
            SocialNoteFile.id == file_id,
            SocialNoteFile.note_id == int(row.id),
            SocialNoteFile.user_id == user.id,
            SocialNoteFile.actor_key == actor_key,
        )
    )
    if not file_row:
        raise HTTPException(status_code=404, detail="Р¤Р°Р№Р» РЅРµ РЅР°Р№РґРµРЅ")
    _social_note_delete_disk_file(str(file_row.url or ""))
    db.delete(file_row)
    row.updated_at = datetime.utcnow()
    _audit(
        db,
        user,
        action="social_note_file_deleted",
        details=json.dumps(
            {"note_id": int(row.id or 0), "file_id": int(file_id or 0)},
            ensure_ascii=False,
        ),
        module_code="social_hub",
        entity_type="note_file",
        entity_id=str(file_id or ""),
    )
    db.commit()
    return MessageOut(message="Р¤Р°Р№Р» СѓРґР°Р»РµРЅ")


@router.delete("/social/notes/{note_id}", response_model=MessageOut)
def social_delete_note(
    note_id: int,
    user: User = Depends(get_current_user),
    db: Session = Depends(get_db),
):
    ensure_module_enabled(db, user, "social_hub")
    actor_key, _, _ = _social_actor_identity(db, user)
    row = db.get(SocialNote, note_id)
    if not row or int(row.user_id) != int(user.id) or str(row.actor_key or "") != actor_key:
        raise HTTPException(status_code=404, detail="Р—Р°РјРµС‚РєР° РЅРµ РЅР°Р№РґРµРЅР°")
    files = db.scalars(
        select(SocialNoteFile).where(
            SocialNoteFile.note_id == int(row.id),
            SocialNoteFile.user_id == user.id,
            SocialNoteFile.actor_key == actor_key,
        )
    ).all()
    for file_row in files:
        _social_note_delete_disk_file(str(file_row.url or ""))
        db.delete(file_row)
    _audit(
        db,
        user,
        action="social_note_deleted",
        details=json.dumps({"note_id": int(row.id or 0), "title": str(row.title or "")[:200]}, ensure_ascii=False),
        module_code="social_hub",
        entity_type="note",
        entity_id=str(row.id or ""),
    )
    db.delete(row)
    db.commit()
    return MessageOut(message="Р—Р°РјРµС‚РєР° СѓРґР°Р»РµРЅР°")


@router.get("/social/notifications", response_model=dict[str, Any])
def social_notifications(
    since_id: int = 0,
    limit: int = 40,
    user: User = Depends(get_current_user),
    db: Session = Depends(get_db),
):
    ensure_module_enabled(db, user, "social_hub")
    actor_key, actor_nick, _ = _social_actor_identity(db, user)
    actor_aliases = _social_actor_alias_keys(db, actor_key)
    _social_emit_due_reminders(db, user_id=int(user.id), actor_key=actor_key, actor_nick=actor_nick)
    safe_limit = max(10, min(int(limit or 40), 200))
    query = select(SocialNotification).where(
        SocialNotification.user_id == int(user.id),
        SocialNotification.recipient_key.in_(actor_aliases),
    )
    if int(since_id or 0) > 0:
        query = query.where(SocialNotification.id > int(since_id))
    rows = db.scalars(query.order_by(SocialNotification.id.desc()).limit(safe_limit)).all()
    unread = db.scalar(
        select(func.count())
        .select_from(SocialNotification)
        .where(
            SocialNotification.user_id == int(user.id),
            SocialNotification.recipient_key.in_(actor_aliases),
            SocialNotification.is_read.is_(False),
        )
    ) or 0
    data_rows = []
    for row in reversed(rows):
        payload: dict[str, Any]
        try:
            payload = json.loads(str(row.payload_json or "{}"))
        except Exception:
            payload = {}
        data_rows.append(
            SocialNotificationOut(
                id=int(row.id),
                kind=str(row.kind or ""),
                title=_decode_mojibake_text(str(row.title or "")),
                body=_decode_mojibake_text(str(row.body or "")),
                payload=payload,
                is_read=bool(getattr(row, "is_read", False)),
                created_at=_to_utc_iso(row.created_at),
            ).model_dump()
        )
    return {"unread": int(unread), "rows": data_rows, "meta": {"generated_at_utc": _to_utc_iso(datetime.utcnow())}}


@router.post("/social/notifications/read-all", response_model=MessageOut)
def social_notifications_read_all(
    user: User = Depends(get_current_user),
    db: Session = Depends(get_db),
):
    ensure_module_enabled(db, user, "social_hub")
    actor_key, _, _ = _social_actor_identity(db, user)
    actor_aliases = _social_actor_alias_keys(db, actor_key)
    db.execute(
        text(
            """
            UPDATE social_notifications
            SET is_read = 1
            WHERE user_id = :user_id
              AND recipient_key IN (:k1, :k2)
              AND is_read = 0
            """
        ),
        {
            "user_id": int(user.id),
            "k1": actor_aliases[0] if actor_aliases else actor_key,
            "k2": actor_aliases[1] if len(actor_aliases) > 1 else actor_aliases[0] if actor_aliases else actor_key,
        },
    )
    db.commit()
    return MessageOut(message="РћРє")


@router.get("/social/announcements/pending", response_model=dict[str, Any])
def social_pending_announcements(
    limit: int = 5,
    user: User = Depends(get_current_user),
    db: Session = Depends(get_db),
):
    ensure_module_enabled(db, user, "social_hub")
    actor_key, actor_nick, _ = _social_actor_identity(db, user)
    _social_emit_due_reminders(db, user_id=int(user.id), actor_key=actor_key, actor_nick=actor_nick)
    safe_limit = max(1, min(int(limit or 5), 20))
    now = datetime.utcnow()
    acked_ids = set(
        db.scalars(
            select(SocialAnnouncementAck.announcement_id).where(
                SocialAnnouncementAck.user_id == int(user.id),
                SocialAnnouncementAck.actor_key == actor_key,
            )
        ).all()
    )
    rows = db.scalars(
        select(SocialAnnouncement).where(
            SocialAnnouncement.is_active.is_(True),
            SocialAnnouncement.starts_at <= now,
            or_(SocialAnnouncement.ends_at.is_(None), SocialAnnouncement.ends_at >= now),
        ).order_by(SocialAnnouncement.starts_at.asc(), SocialAnnouncement.id.asc())
    ).all()
    pending = [
        _social_announcement_to_public_out(x).model_dump()
        for x in rows
        if int(x.id or 0) not in acked_ids and _social_announcement_is_for_user(x, int(user.id))
    ]
    return {"rows": pending[:safe_limit]}


@router.post("/social/announcements/{announcement_id}/ack", response_model=MessageOut)
def social_ack_announcement(
    announcement_id: int,
    user: User = Depends(get_current_user),
    db: Session = Depends(get_db),
):
    ensure_module_enabled(db, user, "social_hub")
    ann_id = int(announcement_id or 0)
    if ann_id <= 0:
        raise HTTPException(status_code=400, detail="РќРµРєРѕСЂСЂРµРєС‚РЅС‹Р№ ID РѕР±СЉСЏРІР»РµРЅРёСЏ")
    actor_key, _, _ = _social_actor_identity(db, user)
    row = db.get(SocialAnnouncement, ann_id)
    if not row or not row.is_active:
        raise HTTPException(status_code=404, detail="РћР±СЉСЏРІР»РµРЅРёРµ РЅРµ РЅР°Р№РґРµРЅРѕ")
    if not _social_announcement_is_for_user(row, int(user.id)):
        raise HTTPException(status_code=403, detail="РќРµС‚ РґРѕСЃС‚СѓРїР° Рє РѕР±СЉСЏРІР»РµРЅРёСЋ")
    exists = db.scalar(
        select(SocialAnnouncementAck).where(
            SocialAnnouncementAck.announcement_id == ann_id,
            SocialAnnouncementAck.user_id == int(user.id),
            SocialAnnouncementAck.actor_key == actor_key,
        )
    )
    if not exists:
        db.add(
            SocialAnnouncementAck(
                announcement_id=ann_id,
                user_id=int(user.id),
                actor_key=actor_key[:60],
            )
        )
    db.execute(
        text(
            """
            UPDATE social_notifications
            SET is_read = 1
            WHERE user_id = :user_id
              AND recipient_key = :recipient_key
              AND kind = 'announcement'
              AND dedupe_key = :dedupe_key
            """
        ),
        {
            "user_id": int(user.id),
            "recipient_key": actor_key,
            "dedupe_key": f"announcement:{ann_id}:{actor_key}",
        },
    )
    db.commit()
    return MessageOut(message="РћРє")


def _notification_sound_storage_dir() -> Path:
    static_root = Path(__file__).resolve().parent.parent / "static"
    target_dir = static_root / "uploads" / "notification_sounds"
    target_dir.mkdir(parents=True, exist_ok=True)
    return target_dir


def _notification_sound_clean_filename(raw: str) -> str:
    base = os.path.basename(str(raw or "").strip()) or "sound"
    safe = re.sub(r"[^\w.\-]+", "_", base, flags=re.UNICODE).strip("._")
    return (safe or "sound")[:180]


def _notification_sound_guess_ext(filename: str, content_type: str) -> str:
    ext = str(Path(filename or "").suffix or "").strip().lower()
    allowed = {".mp3", ".wav", ".ogg", ".m4a", ".aac", ".webm", ".mp4"}
    if ext in allowed:
        return ext
    guessed = mimetypes.guess_extension(str(content_type or "").split(";", 1)[0].strip().lower()) or ""
    if guessed == ".jpe":
        guessed = ".jpg"
    if guessed in allowed:
        return guessed
    return ".mp3"


@router.get("/social/notification-settings", response_model=NotificationSoundSettingsOut)
def social_notification_settings(
    user: User = Depends(get_current_user),
    db: Session = Depends(get_db),
):
    ensure_module_enabled(db, user, "social_hub")
    return _get_notification_sound_settings(db)


@router.get("/admin/notification-settings", response_model=NotificationSoundSettingsOut)
def admin_get_notification_settings(
    _: User = Depends(get_admin_user),
    db: Session = Depends(get_db),
):
    return _get_notification_sound_settings(db)


@router.get("/admin/announcements", response_model=list[SocialAnnouncementOut])
def admin_announcements(
    active_only: bool = False,
    _: User = Depends(get_admin_user),
    db: Session = Depends(get_db),
):
    query = select(SocialAnnouncement)
    if active_only:
        query = query.where(SocialAnnouncement.is_active.is_(True))
    rows = db.scalars(query.order_by(SocialAnnouncement.starts_at.desc(), SocialAnnouncement.id.desc()).limit(300)).all()
    return [_social_announcement_to_out(x) for x in rows]


@router.post("/admin/announcements", response_model=SocialAnnouncementOut)
def admin_create_announcement(
    payload: SocialAnnouncementIn,
    me: User = Depends(get_admin_user),
    db: Session = Depends(get_db),
):
    title = str(payload.title or "").strip()
    if not title:
        raise HTTPException(status_code=400, detail="РЈРєР°Р¶РёС‚Рµ Р·Р°РіРѕР»РѕРІРѕРє РѕР±СЉСЏРІР»РµРЅРёСЏ")
    starts_at = _social_parse_dt(payload.starts_at)
    if not starts_at:
        raise HTTPException(status_code=400, detail="РќРµРєРѕСЂСЂРµРєС‚РЅР°СЏ РґР°С‚Р°/РІСЂРµРјСЏ РїСѓР±Р»РёРєР°С†РёРё")
    ends_at = _social_parse_dt(payload.ends_at) if payload.ends_at else None
    if ends_at and ends_at < starts_at:
        raise HTTPException(status_code=400, detail="Р”Р°С‚Р° Р·Р°РІРµСЂСЂРµРЅРёСЏ РјРµРЅСЊС€Рµ РґР°С‚С‹ РїСѓР±Р»РёРєР°С†РёРё")
    target_user_id, target_user_ids = _social_resolve_announcement_targets(payload)
    row = SocialAnnouncement(
        user_id=target_user_id,
        target_user_ids_json=json.dumps(target_user_ids, ensure_ascii=False),
        title=title[:255],
        body=str(payload.body or "")[:5000],
        starts_at=starts_at,
        ends_at=ends_at,
        is_active=bool(payload.is_active),
        created_by_user_id=int(me.id),
    )
    db.add(row)
    db.flush()
    _audit(
        db,
        me,
        action="admin_announcement_created",
        details=json.dumps(
            {
                "announcement_id": int(row.id or 0),
                "starts_at": _to_utc_iso(starts_at),
                "ends_at": _to_utc_iso(ends_at),
                "user_id": target_user_id,
                "user_ids": target_user_ids,
            },
            ensure_ascii=False,
        ),
        module_code="admin",
        entity_type="announcement",
        entity_id=str(row.id or ""),
    )
    db.commit()
    return _social_announcement_to_out(row)


@router.put("/admin/announcements/{announcement_id}", response_model=SocialAnnouncementOut)
def admin_update_announcement(
    announcement_id: int,
    payload: SocialAnnouncementIn,
    me: User = Depends(get_admin_user),
    db: Session = Depends(get_db),
):
    row = db.get(SocialAnnouncement, int(announcement_id or 0))
    if not row:
        raise HTTPException(status_code=404, detail="РћР±СЉСЏРІР»РµРЅРёРµ РЅРµ РЅР°Р№РґРµРЅРѕ")
    title = str(payload.title or "").strip()
    if not title:
        raise HTTPException(status_code=400, detail="РЈРєР°Р¶РёС‚Рµ Р·Р°РіРѕР»РѕРІРѕРє РѕР±СЉСЏРІР»РµРЅРёСЏ")
    starts_at = _social_parse_dt(payload.starts_at)
    if not starts_at:
        raise HTTPException(status_code=400, detail="РќРµРєРѕСЂСЂРµРєС‚РЅР°СЏ РґР°С‚Р°/РІСЂРµРјСЏ РїСѓР±Р»РёРєР°С†РёРё")
    ends_at = _social_parse_dt(payload.ends_at) if payload.ends_at else None
    if ends_at and ends_at < starts_at:
        raise HTTPException(status_code=400, detail="Р”Р°С‚Р° Р·Р°РІРµСЂСЂРµРЅРёСЏ РјРµРЅСЊС€Рµ РґР°С‚С‹ РїСѓР±Р»РёРєР°С†РёРё")
    target_user_id, target_user_ids = _social_resolve_announcement_targets(payload)
    row.title = title[:255]
    row.body = str(payload.body or "")[:5000]
    row.starts_at = starts_at
    row.ends_at = ends_at
    row.user_id = target_user_id
    row.target_user_ids_json = json.dumps(target_user_ids, ensure_ascii=False)
    row.is_active = bool(payload.is_active)
    row.updated_at = datetime.utcnow()
    _audit(
        db,
        me,
        action="admin_announcement_updated",
        details=json.dumps(
            {
                "announcement_id": int(row.id or 0),
                "starts_at": _to_utc_iso(starts_at),
                "ends_at": _to_utc_iso(ends_at),
                "user_id": target_user_id,
                "user_ids": target_user_ids,
                "is_active": bool(row.is_active),
            },
            ensure_ascii=False,
        ),
        module_code="admin",
        entity_type="announcement",
        entity_id=str(row.id or ""),
    )
    db.commit()
    return _social_announcement_to_out(row)


@router.delete("/admin/announcements/{announcement_id}", response_model=MessageOut)
def admin_delete_announcement(
    announcement_id: int,
    me: User = Depends(get_admin_user),
    db: Session = Depends(get_db),
):
    row = db.get(SocialAnnouncement, int(announcement_id or 0))
    if not row:
        raise HTTPException(status_code=404, detail="РћР±СЉСЏРІР»РµРЅРёРµ РЅРµ РЅР°Р№РґРµРЅРѕ")
    row.is_active = False
    row.updated_at = datetime.utcnow()
    _audit(
        db,
        me,
        action="admin_announcement_deleted",
        details=json.dumps({"announcement_id": int(row.id or 0)}, ensure_ascii=False),
        module_code="admin",
        entity_type="announcement",
        entity_id=str(row.id or ""),
    )
    db.commit()
    return MessageOut(message="РћРє")


@router.post("/admin/notification-settings", response_model=NotificationSoundSettingsOut)
def admin_save_notification_settings(
    payload: NotificationSoundSettingsIn,
    me: User = Depends(get_admin_user),
    db: Session = Depends(get_db),
):
    safe = _sanitize_notification_sound_settings(payload.model_dump())
    _set_system_setting(db, _NOTIFICATION_SOUND_SETTINGS_KEY, json.dumps(safe, ensure_ascii=False))
    _audit(
        db,
        me,
        action="admin_notification_settings_updated",
        details=json.dumps(
            {
                "desktop_enabled": bool(safe.get("desktop_enabled")),
                "chat_enabled": bool(safe.get("chat_enabled")),
                "task_enabled": bool(safe.get("task_enabled")),
                "calendar_enabled": bool(safe.get("calendar_enabled")),
            },
            ensure_ascii=False,
        ),
        module_code="admin",
        entity_type="notification_settings",
        entity_id="global",
    )
    db.commit()
    return NotificationSoundSettingsOut(**safe)


@router.post("/admin/notification-settings/upload", response_model=AvatarUploadOut)
def admin_upload_notification_sound(
    group: str = Form(...),
    file: UploadFile = File(...),
    me: User = Depends(get_admin_user),
    db: Session = Depends(get_db),
):
    key = str(group or "").strip().lower()
    map_key = {
        "default": "default_sound_url",
        "chat": "chat_sound_url",
        "task": "task_sound_url",
        "calendar": "calendar_sound_url",
    }.get(key, "")
    if not map_key:
        raise HTTPException(status_code=400, detail="РќРµРєРѕСЂСЂРµРєС‚РЅР°СЏ РіСЂСѓРїРїР° Р·РІСѓРєР°")
    if not file or not file.filename:
        raise HTTPException(status_code=400, detail="Р¤Р°Р№Р» РЅРµ РІС‹Р±СЂР°РЅ")
    raw = file.file.read()
    if not raw:
        raise HTTPException(status_code=400, detail="Р¤Р°Р№Р» РїСѓСЃС‚РѕР№")
    max_size = 4 * 1024 * 1024
    if len(raw) > max_size:
        raise HTTPException(status_code=400, detail="Р¤Р°Р№Р» СЃР»РёСЂРєРѕРј Р±РѕР»СЊСЂРѕР№ (РґРѕ 4 РњР‘)")
    original = _notification_sound_clean_filename(file.filename or "sound")
    ext = _notification_sound_guess_ext(original, str(file.content_type or ""))
    if ext not in {".mp3", ".wav", ".ogg", ".m4a", ".aac", ".webm", ".mp4"}:
        raise HTTPException(status_code=400, detail="РќРµРїРѕРґРґРµСЂР¶РёРІР°РµРјС‹Р№ С„РѕСЂРјР°С‚ Р°СѓРґРёРѕ")
    storage_name = f"notif-{key}-{secrets.token_hex(6)}{ext}"
    path = _notification_sound_storage_dir() / storage_name
    path.write_bytes(raw)
    url = f"/static/uploads/notification_sounds/{storage_name}"
    current = _get_notification_sound_settings(db).model_dump()
    current[map_key] = url
    safe = _sanitize_notification_sound_settings(current)
    _set_system_setting(db, _NOTIFICATION_SOUND_SETTINGS_KEY, json.dumps(safe, ensure_ascii=False))
    _audit(
        db,
        me,
        action="admin_notification_sound_uploaded",
        details=json.dumps(
            {
                "group": key,
                "url": url,
                "size": len(raw),
            },
            ensure_ascii=False,
        ),
        module_code="admin",
        entity_type="notification_sound",
        entity_id=key,
    )
    db.commit()
    return AvatarUploadOut(url=url)


def upsert_products(
    db: Session,
    user_id: int,
    marketplace: str,
    api_key: str,
    articles: list[str],
    import_all: bool,
    *,
    owner_member_id: int,
    actor_is_owner: bool,
) -> list[Product]:
    try:
        data = fetch_products_from_marketplace(marketplace, api_key, articles, import_all)
    except Exception as exc:
        safe_error = re.sub(r"\s+", " ", str(exc or "")).strip()[:220]
        safe_marketplace = str(marketplace or "").upper() or "MP"
        raise HTTPException(
            status_code=502,
            detail=f"\u041d\u0435 \u0443\u0434\u0430\u043b\u043e\u0441\u044c \u0437\u0430\u0433\u0440\u0443\u0437\u0438\u0442\u044c \u0442\u043e\u0432\u0430\u0440\u044b \u0438\u0437 {safe_marketplace}: {safe_error or '\u0432\u043d\u0435\u0448\u043d\u0438\u0439 API \u0432\u0435\u0440\u043d\u0443\u043b \u043e\u0448\u0438\u0431\u043a\u0443'}",
        ) from exc
    if not isinstance(data, list):
        data = []
    upserted: list[Product] = []
    for item in data:
        item_photos = _normalize_product_photo_list(getattr(item, "photos", []))
        item_main_photo = _normalize_product_photo_url(getattr(item, "photo_url", ""))
        if item_main_photo:
            item_photos = _normalize_product_photo_list([item_main_photo] + item_photos)
        placeholder_photo = f"https://placehold.co/120x120/e8eefc/1b2a52?text={marketplace.upper()}"
        base_query = select(Product).where(
            Product.user_id == user_id,
            Product.marketplace == marketplace,
            Product.article == item.article,
        )
        if actor_is_owner:
            product = db.scalar(base_query.order_by(Product.id.desc()))
        else:
            product = db.scalar(
                base_query.where(Product.owner_member_id == int(owner_member_id or 0)).order_by(Product.id.desc())
            )
        if not product:
            product = Product(
                user_id=user_id,
                owner_member_id=int(owner_member_id or 0) or None,
                marketplace=marketplace,
                article=item.article,
                external_id=item.external_id,
                barcode=item.barcode,
                photo_url=(item_photos[0] if item_photos else placeholder_photo),
                photos_json=json.dumps(item_photos or [placeholder_photo], ensure_ascii=False),
                name=item.name,
                category_name=item.category_name or "",
                current_description=item.description,
                target_keywords="",
            )
            db.add(product)
        else:
            product.name = item.name
            product.external_id = item.external_id or product.external_id
            product.barcode = item.barcode
            stored_photos = _product_photos_from_row(product)
            merged_photos = _normalize_product_photo_list(item_photos + stored_photos + [product.photo_url, item_main_photo])
            if merged_photos:
                product.photo_url = merged_photos[0]
                product.photos_json = json.dumps(merged_photos, ensure_ascii=False)
            elif not str(product.photo_url or "").strip():
                product.photo_url = placeholder_photo
                if not str(product.photos_json or "").strip():
                    product.photos_json = json.dumps([placeholder_photo], ensure_ascii=False)
            elif not str(product.photos_json or "").strip():
                product.photos_json = json.dumps(_normalize_product_photo_list([product.photo_url]), ensure_ascii=False)
            product.current_description = item.description
            if item.category_name:
                product.category_name = item.category_name[:255]
            if not product.owner_member_id:
                product.owner_member_id = int(owner_member_id or 0) or None
        upserted.append(product)
    return upserted


def build_seo_job_out(db: Session, job: SeoJob) -> SeoJobOut:
    product = db.get(Product, job.product_id)
    article = product.article if product else "-"
    name = product.name if product else "РЈРґР°Р»РµРЅРЅС‹Р№ С‚РѕРІР°СЂ"
    barcode = product.barcode if product else ""
    competitor_items: list[dict] = []
    competitor_snapshot = job.competitor_snapshot
    if competitor_snapshot:
        try:
            parsed = json.loads(competitor_snapshot)
            if isinstance(parsed, list):
                competitor_items = [x for x in parsed if isinstance(x, dict)]
                lines = []
                for idx, c in enumerate(competitor_items, start=1):
                    cname = c.get("name") or f"РљРѕРЅРєСѓСЂРµРЅС‚ {idx}"
                    cpos = c.get("position") or idx
                    ckw = ", ".join(c.get("keywords") or [])
                    lines.append(f"#{cpos}: {cname}" + (f" | РєР»СЋС‡Рё: {ckw}" if ckw else ""))
                competitor_snapshot = "\n".join(lines) if lines else competitor_snapshot
        except Exception:
            pass

    generated = _sanitize_generated_description(job.generated_description)

    return SeoJobOut(
        id=job.id,
        product_id=job.product_id,
        product_article=article,
        product_name=name,
        product_barcode=barcode,
        status=job.status,
        generated_description=generated,
        keywords_snapshot=job.keywords_snapshot,
        competitor_snapshot=competitor_snapshot,
        competitor_items=competitor_items,
        target_position=job.target_position,
        current_position=job.current_position,
        next_check_at=job.next_check_at,
    )


def get_user_keywords(db: Session, user_id: int, marketplace: str, owner_member_id: int | None = None) -> list[str]:
    query = select(UserKeyword.keyword).where(UserKeyword.user_id == user_id, UserKeyword.marketplace.in_(["all", marketplace]))
    if owner_member_id is not None and int(owner_member_id or 0) > 0:
        query = query.where(UserKeyword.owner_member_id == int(owner_member_id))
    rows = db.scalars(
        query
    ).all()
    return list(dict.fromkeys([x.strip() for x in rows if x and x.strip()]))


def _get_active_marketplace_api_key(db: Session, user_id: int, marketplace: str) -> str:
    cred = db.scalar(
        select(ApiCredential)
        .where(
            ApiCredential.user_id == user_id,
            ApiCredential.marketplace == marketplace,
            ApiCredential.active.is_(True),
        )
        .order_by(ApiCredential.id.desc())
    )
    return cred.api_key if cred and cred.api_key else ""


def _get_or_create_ai_settings(db: Session, user_id: int) -> UserAiSettings:
    row = db.scalar(select(UserAiSettings).where(UserAiSettings.user_id == user_id))
    if row:
        return row
    row = UserAiSettings(user_id=user_id, reply_mode="manual", prompt="")
    db.add(row)
    db.flush()
    return row


def _get_or_create_question_ai_settings(db: Session, user_id: int) -> UserQuestionAiSettings:
    row = db.scalar(select(UserQuestionAiSettings).where(UserQuestionAiSettings.user_id == user_id))
    if row:
        return row
    row = UserQuestionAiSettings(user_id=user_id, reply_mode="manual", prompt="")
    db.add(row)
    db.flush()
    return row


def _campaign_id_from_any(row: dict[str, Any]) -> str:
    if not isinstance(row, dict):
        return ""
    for key in ("advertId", "advert_id", "campaignId", "campaign_id", "id", "adId"):
        value = row.get(key)
        text = str(value).strip() if value is not None else ""
        if text and text != "0":
            return text
    return ""


def _to_int_safe(value: Any) -> int:
    try:
        return int(str(value).strip())
    except Exception:
        return 0


def _to_float_safe(value: Any, default: float = 0.0) -> float:
    try:
        num = float(value)
    except Exception:
        try:
            num = float(str(value).strip().replace(",", "."))
        except Exception:
            return float(default)
    if not math.isfinite(num):
        return float(default)
    return float(num)


def _wb_status_label(raw_value: str) -> tuple[str, bool]:
    raw = (raw_value or "").strip()
    code = _to_int_safe(raw)
    labels = {
        -1: "СѓРґР°Р»РµРЅР°",
        1: "С‡РµСЂРЅРѕРІРёРє",
        2: "РјРѕРґРµСЂР°С†РёСЏ",
        3: "РѕС‚РєР»РѕРЅРµРЅР°",
        4: "РіРѕС‚РѕРІР° Рє Р·Р°РїСѓСЃРєСѓ",
        5: "Р·Р°РїР»Р°РЅРёСЂРѕРІР°РЅР°",
        6: "РёРґСѓС‚ РїРѕРєР°Р·С‹",
        7: "Р·Р°РІРµСЂСЂРµРЅР°",
        8: "РѕС‚РјРµРЅРµРЅР°",
        9: "Р°РєС‚РёРІРЅР°",
        10: "РїР°СѓР·Р° РїРѕ РґРЅРµРІРЅРѕРјСѓ Р»РёРјРёС‚Сѓ",
        11: "РїР°СѓР·Р°",
    }
    if code:
        label = labels.get(code, f"СЃС‚Р°С‚СѓСЃ {code}")
        return f"{code} ({label})", code in {6, 9}
    lower = raw.lower()
    if not lower or lower == "-":
        return "-", False
    is_running = ("active" in lower) or ("Р°РєС‚РёРІ" in lower)
    return raw, is_running


def _wb_type_label(raw_value: str) -> str:
    raw = (raw_value or "").strip()
    code = _to_int_safe(raw)
    labels = {
        4: "search",
        5: "catalog",
        6: "cards",
        7: "recommendation",
        8: "auto-cpm",
        9: "search + catalog",
    }
    if code:
        label = labels.get(code, f"type-{code}")
        return f"{code} ({label})"
    return raw or "-"


def _merge_campaign_row(row: dict[str, Any], summary: dict[str, Any], stat: dict[str, Any]) -> dict[str, Any]:
    next_row = dict(row or {})
    campaign_id = _to_int_safe(_campaign_id_from_any(next_row) or (summary or {}).get("campaign_id"))
    if summary:
        if not _campaign_id_from_any(next_row) and summary.get("campaign_id"):
            next_row["advertId"] = summary.get("campaign_id")
        current_name = str(next_row.get("name") or "").strip()
        if summary.get("name") and (not current_name or _campaign_name_is_placeholder(current_name, campaign_id)):
            next_row["name"] = summary.get("name")
        if summary.get("status") and str(next_row.get("status") or next_row.get("state") or "").strip() in {"", "-", "вЂ”"}:
            next_row["status"] = summary.get("status")
        if summary.get("type") and str(next_row.get("type") or next_row.get("campaignType") or "").strip() in {"", "-", "вЂ”"}:
            next_row["type"] = summary.get("type")
        if summary.get("budget") and str(next_row.get("dailyBudget") or next_row.get("budget") or "").strip() in {"", "-", "вЂ”"}:
            next_row["budget"] = summary.get("budget")
    if stat:
        if "stat_has_context" in stat:
            next_row["stat_has_context"] = bool(stat.get("stat_has_context"))
        if "stat_nonzero" in stat:
            next_row["stat_nonzero"] = bool(stat.get("stat_nonzero"))
        for key in ("views", "clicks", "orders", "spent", "ctr", "cr", "cpc", "cpo", "add_to_cart"):
            if key not in stat:
                continue
            raw_value = stat.get(key)
            if raw_value in (None, ""):
                next_row[key] = None
                continue
            next_row[key] = _to_float_safe(raw_value, 0.0)
    return next_row


def _campaign_summary_from_base_row(row: dict[str, Any], campaign_id: int) -> dict[str, Any]:
    if not isinstance(row, dict):
        return {
            "campaign_id": campaign_id,
            "name": f"РљР°РјРїР°РЅРёСЏ {campaign_id}",
            "status": "-",
            "type": "-",
            "budget": "-",
        }
    settings_row = row.get("settings") if isinstance(row.get("settings"), dict) else {}
    finance_row = row.get("finance") if isinstance(row.get("finance"), dict) else {}
    name = (
        str(row.get("name") or row.get("campaignName") or row.get("campaign_name") or settings_row.get("name") or settings_row.get("title") or "").strip()
        or f"РљР°РјРїР°РЅРёСЏ {campaign_id}"
    )
    status = str(row.get("status") or row.get("state") or row.get("statusName") or settings_row.get("status") or "").strip() or "-"
    ctype = str(row.get("type") or row.get("campaignType") or row.get("adType") or row.get("typeName") or settings_row.get("type") or "").strip() or "-"
    budget = (
        str(row.get("dailyBudget") or row.get("budget") or row.get("sum") or row.get("balance") or finance_row.get("budget") or "").strip()
        or "-"
    )
    return {
        "campaign_id": campaign_id,
        "name": name,
        "status": status,
        "type": ctype,
        "budget": budget,
    }


def _campaign_stat_from_base_row(row: dict[str, Any]) -> dict[str, Any]:
    if not isinstance(row, dict):
        return {}
    metric_keys = ("views", "clicks", "orders", "spent", "ctr", "cr", "cpc", "cpo", "add_to_cart")
    rounded_four = {"ctr", "cr", "cpc", "cpo"}
    stat: dict[str, Any] = {}
    metric_hits = 0
    metric_nonzero_hits = 0
    for key in metric_keys:
        if key not in row:
            continue
        raw_value = row.get(key)
        if raw_value in (None, "", "-", "вЂ”"):
            continue
        value = _to_float_safe(raw_value, float("nan"))
        if not math.isfinite(value):
            continue
        metric_hits += 1
        if abs(value) > 1e-9:
            metric_nonzero_hits += 1
        precision = 4 if key in rounded_four else 3
        stat[key] = float(round(value, precision))
    if not stat:
        if bool(row.get("stat_has_context")):
            return {
                "stat_has_context": True,
                "stat_nonzero": bool(row.get("stat_nonzero")),
            }
        return {}
    stat["metric_hits"] = metric_hits
    stat["metric_nonzero_hits"] = metric_nonzero_hits
    stat["stat_has_context"] = True
    stat["stat_nonzero"] = bool(metric_nonzero_hits > 0 or row.get("stat_nonzero"))
    return stat


def _campaign_name_is_placeholder(name: str, campaign_id: int) -> bool:
    text = str(name or "").strip().lower()
    if not text:
        return True
    compact = re.sub(r"\s+", " ", text).strip()
    if re.fullmatch(r"\d{4,}", compact):
        return True

    ru_campaign_root = "\u043a\u0430\u043c\u043f\u0430\u043d"
    ru_ad_root = "\u0440\u0435\u043a\u043b\u0430\u043c"
    generic_re = rf"(campaign|camp|advert|advertising|ads?|{ru_ad_root}[\w\u0400-\u04ff]*|{ru_campaign_root}[\w\u0400-\u04ff]*)"
    if re.fullmatch(rf"{generic_re}\s*[#в„–:\-]?\s*\d+", compact):
        return True
    if re.fullmatch(rf"{generic_re}\s+{generic_re}\s*[#в„–:\-]?\s*\d+", compact):
        return True

    cid = int(campaign_id or 0)
    if cid > 0:
        cid_text = str(cid)
        if compact == cid_text:
            return True
        if re.search(rf"\b{re.escape(cid_text)}\b", compact):
            reduced = re.sub(rf"\b{re.escape(cid_text)}\b", " ", compact)
            reduced = re.sub(r"[#в„–:;,_\-.()\[\]/]+", " ", reduced)
            reduced = re.sub(r"\s+", " ", reduced).strip()
            if not reduced:
                return True
            words = re.findall(r"[a-z\u0400-\u04ff]+", reduced)
            if words:
                generic_words = {
                    "campaign",
                    "camp",
                    "advert",
                    "advertising",
                    "ad",
                    "ads",
                    "\u043a\u0430\u043c\u043f\u0430\u043d\u0438\u044f",
                    "\u043a\u0430\u043c\u043f\u0430\u043d\u0438\u0438",
                    "\u0440\u0435\u043a\u043b\u0430\u043c\u0430",
                    "\u0440\u0435\u043a\u043b\u0430\u043c\u043d\u0430\u044f",
                }
                if all(word in generic_words or word.startswith(ru_campaign_root) or word.startswith(ru_ad_root) for word in words):
                    return True
    return False


def _campaign_summary_has_context(summary: dict[str, Any] | None, campaign_id: int) -> bool:
    if not isinstance(summary, dict):
        return False
    cid = int(campaign_id or 0)
    name = str(summary.get("name") or "").strip()
    if name and not _campaign_name_is_placeholder(name, cid):
        return True
    status = str(summary.get("status") or "").strip()
    ctype = str(summary.get("type") or "").strip()
    budget = str(summary.get("budget") or "").strip()
    if status and status not in {"-", "0"}:
        return True
    if ctype and ctype not in {"-", "0"}:
        return True
    if budget and budget not in {"-", "0", "0.0", "0,0"}:
        return True
    return False


def _campaign_stat_has_context(stat: dict[str, Any] | None) -> bool:
    if not isinstance(stat, dict):
        return False
    explicit = stat.get("stat_has_context")
    if isinstance(explicit, bool):
        return explicit
    metric_keys = ("views", "clicks", "orders", "spent", "ctr", "cr", "cpc", "cpo")
    for key in metric_keys:
        if key not in stat:
            continue
        raw_value = stat.get(key)
        if raw_value in (None, ""):
            continue
        value = _to_float_safe(raw_value, float("nan"))
        if math.isfinite(value) and abs(value) > 0.000001:
            return True
    return False


def _knowledge_tokens(text: str) -> list[str]:
    words = re.findall(r"[a-zA-ZР°-СЏРђ-РЇ0-9_]{3,}", str(text or "").lower())
    stop = {
        "РґР»СЏ", "СЌС‚Рѕ", "С‡С‚Рѕ", "РєР°Рє", "РёР»Рё", "РµСЃР»Рё", "РїСЂРё", "Р±РµР·", "РµРіРѕ", "РµС‰Рµ", "РµС‰С‘",
        "with", "this", "that", "from", "into", "about", "your", "have", "will", "would",
    }
    out: list[str] = []
    seen: set[str] = set()
    for word in words:
        if word in stop or word in seen:
            continue
        seen.add(word)
        out.append(word)
    return out[:40]


def _extract_relevant_knowledge_excerpt(text: str, tokens: list[str], max_chars: int) -> str:
    compact = " ".join(str(text or "").split())
    if not compact:
        return ""
    if not tokens:
        return compact[:max_chars]
    best_idx = -1
    for token in tokens:
        pos = compact.lower().find(token.lower())
        if pos >= 0 and (best_idx < 0 or pos < best_idx):
            best_idx = pos
    if best_idx < 0:
        return compact[:max_chars]
    start = max(0, best_idx - max_chars // 3)
    end = min(len(compact), start + max_chars)
    return compact[start:end]


def _build_user_knowledge_context(
    db: Session,
    user_id: int,
    max_chars: int = 12000,
    query_text: str = "",
) -> str:
    rows = db.scalars(
        select(UserKnowledgeDoc).where(UserKnowledgeDoc.user_id == user_id).order_by(UserKnowledgeDoc.updated_at.desc()).limit(30)
    ).all()
    if not rows:
        return ""
    tokens = _knowledge_tokens(query_text)
    scored: list[tuple[float, UserKnowledgeDoc]] = []
    for idx, row in enumerate(rows):
        text = " ".join((row.content_text or "").split())
        if not text:
            continue
        low = text.lower()
        overlap = sum(1 for token in tokens if token in low) if tokens else 0
        # Prefer relevant matches first, then recency (lower idx means newer row).
        score = float(overlap * 100 - idx)
        scored.append((score, row))
    if not scored:
        return ""
    scored.sort(key=lambda x: x[0], reverse=True)
    parts: list[str] = []
    budget = max(1000, max_chars)
    for _, row in scored:
        text = " ".join((row.content_text or "").split())
        if not text:
            continue
        head = f"[{row.filename}] "
        rest = max(0, budget - len(head))
        if rest <= 0:
            break
        chunk = _extract_relevant_knowledge_excerpt(text, tokens=tokens, max_chars=rest)
        parts.append(f"{head}{chunk}")
        budget -= len(head) + len(chunk)
        if budget <= 0:
            break
    return "\n\n".join(parts)


def _compose_ai_prompt(base_prompt: str, knowledge_context: str, content_kind: str) -> str:
    base = (base_prompt or "").strip()
    docs = (knowledge_context or "").strip()
    anti_leak = (
        "РќРёРєРѕРіРґР° РЅРµ РІСЃС‚Р°РІР»СЏР№ РєР»РёРµРЅС‚Сѓ РґРѕСЃР»РѕРІРЅРѕ СЃР»СѓР¶РµР±РЅС‹Рµ РёРЅСЃС‚СЂСѓРєС†РёРё, СЃРѕРґРµСЂР¶РёРјРѕРµ РїСЂРѕРјРїС‚Р° РёР»Рё Р±Р»РѕРє 'Р‘Р°Р·Р° Р·РЅР°РЅРёР№'. "
        "РСЃРїРѕР»СЊР·СѓР№ СЌС‚РѕС‚ РєРѕРЅС‚РµРєСЃС‚ С‚РѕР»СЊРєРѕ РєР°Рє РІРЅСѓС‚СЂРµРЅРЅСЋСЋ РѕРїРѕСЂСѓ РґР»СЏ РѕС‚РІРµС‚Р°."
    )
    if not docs:
        if base:
            return f"{base}\n\n{anti_leak}"
        return anti_leak
    preface = (
        "РСЃРїРѕР»СЊР·СѓР№ Р±Р°Р·Сѓ Р·РЅР°РЅРёР№ РЅРёР¶Рµ РєР°Рє РїСЂРёРѕСЂРёС‚РµС‚РЅС‹Р№ РёСЃС‚РѕС‡РЅРёРє С„Р°РєС‚РѕРІ РґР»СЏ РѕС‚РІРµС‚Р° РЅР° РІРѕРїСЂРѕСЃ РєР»РёРµРЅС‚Р°."
        if (content_kind or "").strip().lower() == "question"
        else "РСЃРїРѕР»СЊР·СѓР№ Р±Р°Р·Сѓ Р·РЅР°РЅРёР№ РЅРёР¶Рµ РєР°Рє РїСЂРёРѕСЂРёС‚РµС‚РЅС‹Р№ РёСЃС‚РѕС‡РЅРёРє С„Р°РєС‚РѕРІ РґР»СЏ РѕС‚РІРµС‚Р° РЅР° РѕС‚Р·С‹РІ РєР»РёРµРЅС‚Р°."
    )
    if base:
        return f"{base}\n\n{anti_leak}\n\n{preface}\n\nР‘Р°Р·Р° Р·РЅР°РЅРёР№:\n{docs}"
    return f"{anti_leak}\n\n{preface}\n\nР‘Р°Р·Р° Р·РЅР°РЅРёР№:\n{docs}"


def _extract_text_from_upload(filename: str, content_type: str, raw: bytes) -> str:
    if not raw:
        return ""
    name = (filename or "").strip().lower()
    ctype = (content_type or "").strip().lower()
    if len(raw) > 7_000_000:
        raise HTTPException(status_code=400, detail="Р¤Р°Р№Р» СЃР»РёСЂРєРѕРј Р±РѕР»СЊСЂРѕР№ (РјР°РєСЃРёРјСѓРј 7MB)")

    is_pdf = name.endswith(".pdf") or "pdf" in ctype
    if is_pdf:
        text = _extract_pdf_text(raw)
        if not text:
            raise HTTPException(status_code=400, detail="РќРµ СѓРґР°Р»РѕСЃСЊ РёР·РІР»РµС‡СЊ С‚РµРєСЃС‚ РёР· PDF. РџСЂРѕРІРµСЂСЊС‚Рµ, С‡С‚Рѕ PDF РЅРµ СЃРєР°РЅ-РёР·РѕР±СЂР°Р¶РµРЅРёРµ.")
        return " ".join(text.split())

    allowed = (".txt", ".md", ".csv", ".tsv", ".log", ".json", ".xml", ".yml", ".yaml")
    if not any(name.endswith(ext) for ext in allowed) and "text/" not in ctype:
        raise HTTPException(status_code=400, detail="РџРѕРґРґРµСЂР¶РёРІР°СЋС‚СЃСЏ txt/md/csv/tsv/log/json/xml/yml/yaml/pdf")
    return " ".join(_decode_bytes(raw).split())


def _decode_bytes(raw: bytes) -> str:
    for enc in ("utf-8", "cp1251", "latin-1"):
        try:
            return raw.decode(enc)
        except Exception:
            continue
    return raw.decode("utf-8", errors="ignore")


def _extract_pdf_text(raw: bytes) -> str:
    bio = io.BytesIO(raw)
    # Optional dependency support to keep deploy lightweight.
    for module_name, class_name in (("pypdf", "PdfReader"), ("PyPDF2", "PdfReader")):
        try:
            module = __import__(module_name, fromlist=[class_name])
            reader_cls = getattr(module, class_name)
            reader = reader_cls(bio)
            parts: list[str] = []
            for page in list(reader.pages)[:120]:
                try:
                    parts.append(page.extract_text() or "")
                except Exception:
                    continue
            return "\n".join(parts)
        except Exception:
            bio.seek(0)
            continue
    raise HTTPException(
        status_code=400,
        detail="PDF-РїР°СЂСЃРёРЅРі РЅРµРґРѕСЃС‚СѓРїРµРЅ: СѓСЃС‚Р°РЅРѕРІРёС‚Рµ РїР°РєРµС‚ pypdf Рё РїРµСЂРµР·Р°РїСѓСЃС‚РёС‚Рµ СЃРµСЂРІРёСЃ.",
    )


def _get_or_create_user_profile(db: Session, user_id: int) -> UserProfile:
    row = db.scalar(select(UserProfile).where(UserProfile.user_id == user_id))
    if row:
        return row
    row = UserProfile(
        user_id=user_id,
        full_name="",
        company_name="",
        city="",
        legal_name="",
        legal_address="",
        tax_id="",
        tax_rate=0.0,
        phone="",
        position_title="",
        team_size=1,
        company_structure="",
        avatar_url="",
    )
    db.add(row)
    db.flush()
    return row


def _safe_team_scope(values: list[str] | None) -> list[str]:
    allowed = {
        "products",
        "seo_generation",
        "sales_stats",
        "accounting",
        "wb_reviews_ai",
        "wb_questions_ai",
        "returns",
        "wb_ads",
        "wb_ads_analytics",
        "wb_ads_recommendations",
        "user_profile",
        "help_center",
        "ai_assistant",
        "social_hub",
    }
    out: list[str] = []
    seen: set[str] = set()
    for item in values or []:
        code = str(item or "").strip().lower()
        if not code or code not in allowed or code in seen:
            continue
        seen.add(code)
        out.append(code)
    if not out:
        return ["products", "sales_stats", "accounting", "wb_reviews_ai", "wb_questions_ai", "returns", "wb_ads", "wb_ads_analytics", "wb_ads_recommendations", "user_profile", "help_center", "social_hub"]
    return out


def _ensure_team_email_is_available(
    db: Session,
    email: str,
    *,
    user_id: int,
    exclude_member_id: int | None = None,
) -> None:
    candidate = (email or "").strip().lower()
    if not candidate:
        raise HTTPException(status_code=400, detail="РЈРєР°Р¶РёС‚Рµ email СЃРѕС‚СЂСѓРґРЅРёРєР°")

    owner_user = db.scalar(select(User).where(User.email == candidate))
    if owner_user and owner_user.id != user_id:
        raise HTTPException(status_code=400, detail="Р­С‚РѕС‚ email СѓР¶Рµ РёСЃРїРѕР»СЊР·СѓРµС‚СЃСЏ РІ РґСЂСѓРіРѕРј РєР°Р±РёРЅРµС‚Рµ")

    duplicate = db.scalar(select(TeamMember).where(TeamMember.email == candidate))
    if duplicate and duplicate.user_id != user_id:
        raise HTTPException(status_code=400, detail="Р­С‚РѕС‚ email СѓР¶Рµ РёСЃРїРѕР»СЊР·СѓРµС‚СЃСЏ СЃРѕС‚СЂСѓРґРЅРёРєРѕРј РґСЂСѓРіРѕРіРѕ РєР°Р±РёРЅРµС‚Р°")
    if duplicate and exclude_member_id is not None and duplicate.id != exclude_member_id:
        raise HTTPException(status_code=400, detail="РЎРѕС‚СЂСѓРґРЅРёРє СЃ С‚Р°РєРёРј email СѓР¶Рµ РґРѕР±Р°РІР»РµРЅ")


def _validate_team_member_password(raw_password: str, *, required: bool) -> str:
    password = str(raw_password or "")
    if not password:
        if required:
            raise HTTPException(status_code=400, detail="РЈРєР°Р¶РёС‚Рµ РїР°СЂРѕР»СЊ СЃРѕС‚СЂСѓРґРЅРёРєР° (РјРёРЅРёРјСѓРј 8 СЃРёРјРІРѕР»РѕРІ)")
        return ""
    if len(password) < 8:
        raise HTTPException(status_code=400, detail="РџР°СЂРѕР»СЊ СЃРѕС‚СЂСѓРґРЅРёРєР° РґРѕР»Р¶РµРЅ Р±С‹С‚СЊ РјРёРЅРёРјСѓРј 8 СЃРёРјРІРѕР»РѕРІ")
    return password


def _team_scope_from_row(row: TeamMember) -> list[str]:
    raw = str(row.access_scope or "").strip()
    if not raw:
        return ["*"] if row.is_owner else ["products", "sales_stats", "accounting", "wb_reviews_ai", "wb_questions_ai", "returns", "social_hub"]
    try:
        parsed = json.loads(raw)
    except Exception:
        parsed = []
    if row.is_owner:
        return ["*"]
    if not isinstance(parsed, list):
        return ["products", "sales_stats", "accounting", "wb_reviews_ai", "wb_questions_ai", "returns", "wb_ads", "wb_ads_analytics", "wb_ads_recommendations", "user_profile", "help_center", "social_hub"]
    cleaned = _safe_team_scope([str(x) for x in parsed])
    if cleaned and all(x in {"products", "sales_stats", "accounting"} for x in cleaned):
        return ["products", "sales_stats", "accounting", "wb_reviews_ai", "wb_questions_ai", "returns", "wb_ads", "wb_ads_analytics", "wb_ads_recommendations", "user_profile", "help_center", "social_hub"]
    return cleaned


def _team_member_to_out(row: TeamMember) -> TeamMemberOut:
    return TeamMemberOut(
        id=row.id,
        email=row.email,
        has_password=bool(str(row.hashed_password or "").strip()),
        phone=row.phone or "",
        full_name=row.full_name or "",
        city=row.city or "",
        position_title=row.position_title or "",
        nickname=row.nickname or "",
        avatar_url=row.avatar_url or "",
        access_scope=_team_scope_from_row(row),
        is_owner=bool(row.is_owner),
        is_active=bool(row.is_active),
        created_at=row.created_at.isoformat() if row.created_at else None,
    )


def _ensure_owner_team_member(db: Session, user: User) -> TeamMember:
    owner = db.scalar(select(TeamMember).where(TeamMember.user_id == user.id, TeamMember.is_owner.is_(True)).order_by(TeamMember.id.asc()))
    if owner:
        owner.email = user.email
        if not owner.hashed_password:
            owner.hashed_password = user.hashed_password
        owner.access_scope = json.dumps(["*"], ensure_ascii=False)
        owner.is_active = True
        return owner
    owner = TeamMember(
        user_id=user.id,
        email=user.email,
        full_name="",
        nickname="owner",
        avatar_url="",
        hashed_password=user.hashed_password,
        access_scope=json.dumps(["*"], ensure_ascii=False),
        is_owner=True,
        is_active=True,
    )
    db.add(owner)
    db.flush()
    return owner


def _list_team_members(db: Session, user_id: int) -> list[TeamMemberOut]:
    rows = db.scalars(select(TeamMember).where(TeamMember.user_id == user_id).order_by(TeamMember.is_owner.desc(), TeamMember.id.asc())).all()
    return [_team_member_to_out(row) for row in rows]


def _build_admin_user_profile_payload(db: Session, target: User) -> AdminUserProfileOut:
    profile = _get_or_create_user_profile(db, target.id)
    _ensure_owner_team_member(db, target)
    account = _get_or_create_billing_account(db, target.id)
    credentials = db.scalars(select(ApiCredential).where(ApiCredential.user_id == target.id).order_by(ApiCredential.id.desc())).all()
    return AdminUserProfileOut(
        user_id=target.id,
        email=target.email,
        role=target.role,
        profile={
            "full_name": profile.full_name,
            "company_name": profile.company_name,
            "city": profile.city,
            "legal_name": profile.legal_name,
            "legal_address": profile.legal_address,
            "tax_id": profile.tax_id,
            "tax_rate": profile.tax_rate,
            "phone": profile.phone,
            "position_title": profile.position_title,
            "team_size": profile.team_size,
            "company_structure": profile.company_structure,
            "avatar_url": profile.avatar_url,
        },
        plan={
            "plan_code": account.plan_code,
            "status": account.status,
            "monthly_price": int(account.monthly_price or 0),
            "renew_at": account.renew_at.isoformat() if account.renew_at else None,
        },
        credentials=[
            ApiCredentialOut(
                id=c.id,
                marketplace=c.marketplace,
                api_key_masked=mask_key(c.api_key),
                active=bool(c.active),
            )
            for c in credentials
        ],
        team_members=_list_team_members(db, target.id),
    )


def _build_user_profile_payload(db: Session, user: User, profile: UserProfile, account: BillingAccount) -> UserProfileOut:
    _ensure_owner_team_member(db, user)
    creds = db.scalars(select(ApiCredential).where(ApiCredential.user_id == user.id).order_by(ApiCredential.id.desc())).all()
    owner_member_avatar = db.scalar(
        select(TeamMember.avatar_url).where(
            TeamMember.user_id == user.id,
            TeamMember.is_owner.is_(True),
        ).order_by(TeamMember.id.asc())
    ) or ""
    effective_avatar_url = str(owner_member_avatar or "").strip() or str(profile.avatar_url or "").strip()
    plans = [
        {
            "code": code,
            "title": str(info.get("title") or code),
            "price": int(info.get("price") or 0),
            "limits": dict(info.get("limits") or {}),
        }
        for code, info in BILLING_PLANS.items()
    ]
    return UserProfileOut(
        email=user.email,
        full_name=profile.full_name,
        company_name=profile.company_name,
        city=profile.city,
        legal_name=profile.legal_name,
        legal_address=profile.legal_address,
        tax_id=profile.tax_id,
        tax_rate=float(profile.tax_rate or 0.0),
        phone=profile.phone,
        position_title=profile.position_title,
        team_size=int(profile.team_size or 1),
        company_structure=profile.company_structure,
        avatar_url=effective_avatar_url,
        plan_code=account.plan_code,
        plan_status=account.status,
        monthly_price=int(account.monthly_price or 0),
        renew_at=account.renew_at.isoformat() if account.renew_at else None,
        available_plans=plans,
        credentials=[
            ApiCredentialOut(
                id=c.id,
                marketplace=c.marketplace,
                api_key_masked=mask_key(c.api_key),
                active=bool(c.active),
            )
            for c in creds
        ],
        team_members=_list_team_members(db, user.id),
    )


def _get_or_create_billing_account(db: Session, user_id: int) -> BillingAccount:
    row = db.scalar(select(BillingAccount).where(BillingAccount.user_id == user_id))
    if row:
        return row
    plan_code = "starter"
    info = BILLING_PLANS[plan_code]
    row = BillingAccount(
        user_id=user_id,
        plan_code=plan_code,
        status="active",
        monthly_price=int(info["price"]),
        renew_at=datetime.utcnow() + timedelta(days=30),
        auto_renew=True,
    )
    db.add(row)
    db.flush()
    db.add(BillingEvent(user_id=user_id, event_type="created", amount=int(info["price"]), note="РЎРѕР·РґР°РЅ billing Р°РєРєР°СѓРЅС‚"))
    return row


def _build_billing_payload(db: Session, user_id: int, account: BillingAccount) -> BillingOut:
    plan_code = (account.plan_code or "starter").lower()
    if plan_code not in BILLING_PLANS:
        plan_code = "starter"
    limits = dict(BILLING_PLANS[plan_code]["limits"])

    now = datetime.utcnow()
    month_start = datetime(now.year, now.month, 1)
    usage_products = db.scalar(select(func.count()).select_from(Product).where(Product.user_id == user_id)) or 0
    usage_jobs = db.scalar(select(func.count()).select_from(SeoJob).where(SeoJob.user_id == user_id, SeoJob.created_at >= month_start)) or 0
    usage_replies = db.scalar(
        select(func.count()).select_from(AuditLog).where(
            AuditLog.user_id == user_id,
            AuditLog.created_at >= month_start,
            AuditLog.action.in_(["wb_review_reply", "ozon_review_reply", "wb_question_reply", "ozon_question_reply"]),
        )
    ) or 0
    usage = {
        "products": int(usage_products),
        "seo_jobs_month": int(usage_jobs),
        "ai_replies_month": int(usage_replies),
    }

    mods = db.scalars(select(ModuleAccess).where(ModuleAccess.user_id == user_id).order_by(ModuleAccess.module_code.asc())).all()
    modules = [{"module_code": x.module_code, "enabled": bool(x.enabled)} for x in mods]

    hist_rows = db.scalars(
        select(BillingEvent).where(BillingEvent.user_id == user_id).order_by(BillingEvent.id.desc()).limit(100)
    ).all()
    history = [
        {
            "id": row.id,
            "event_type": row.event_type,
            "amount": row.amount,
            "note": row.note,
            "created_at": row.created_at.isoformat() if row.created_at else "",
        }
        for row in hist_rows
    ]

    plans = [
        {
            "code": code,
            "title": str(info.get("title") or code),
            "price": int(info.get("price") or 0),
            "limits": dict(info.get("limits") or {}),
        }
        for code, info in BILLING_PLANS.items()
    ]
    renew_at = account.renew_at.isoformat() if account.renew_at else None
    return BillingOut(
        plan_code=plan_code,
        status=account.status,
        monthly_price=int(account.monthly_price or 0),
        renew_at=renew_at,
        auto_renew=bool(account.auto_renew),
        limits=limits,
        usage=usage,
        available_plans=plans,
        modules=modules,
        history=history,
    )


def _normalize_accounting_marketplace(value: str) -> str:
    code = str(value or "all").strip().lower()
    if code not in {"all", "wb", "ozon"}:
        return "all"
    return code


def _normalize_expense_recurrence(value: str) -> str:
    code = str(value or "monthly").strip().lower()
    if code in {"one_time", "single"}:
        return "once"
    if code not in {"once", "daily", "weekly", "monthly", "quarterly", "yearly"}:
        return "monthly"
    return code


def _parse_datetime_or_none(value: str | None) -> datetime | None:
    text = str(value or "").strip()
    if not text:
        return None
    if text.endswith("Z"):
        text = text[:-1] + "+00:00"
    try:
        dt = datetime.fromisoformat(text)
        if dt.tzinfo is not None:
            return dt.astimezone(timezone.utc).replace(tzinfo=None)
        return dt
    except Exception:
        pass
    for fmt in ("%Y-%m-%d", "%d.%m.%Y", "%Y/%m/%d"):
        try:
            return datetime.strptime(text[:10], fmt)
        except Exception:
            continue
    return None


def _to_iso_or_none(value: datetime | None) -> str | None:
    if not value:
        return None
    try:
        return value.replace(microsecond=0).isoformat()
    except Exception:
        return None


def _to_money(value: Any) -> float:
    if value is None:
        return 0.0
    if isinstance(value, (int, float)):
        num = float(value)
        if not math.isfinite(num):
            return 0.0
        return float(round(num, 2))
    if isinstance(value, dict):
        for key in (
            "price",
            "amount",
            "value",
            "old_price",
            "base_price",
            "list_price",
            "min_price",
            "marketing_price",
            "promo_price",
            "discount_price",
            "sale_price",
            "price_with_discount",
        ):
            if key in value:
                amount = _to_money(value.get(key))
                if amount > 0:
                    return amount
        for nested in value.values():
            amount = _to_money(nested)
            if amount > 0:
                return amount
        return 0.0
    if isinstance(value, list):
        for item in value[:60]:
            amount = _to_money(item)
            if amount > 0:
                return amount
        return 0.0

    text = str(value or "").replace("\u00a0", " ").strip()
    if not text:
        return 0.0
    compact = re.sub(r"[^0-9,.\-]", "", text)
    if not compact:
        return 0.0
    if "," in compact and "." in compact:
        if compact.rfind(",") > compact.rfind("."):
            compact = compact.replace(".", "").replace(",", ".")
        else:
            compact = compact.replace(",", "")
    else:
        compact = compact.replace(",", ".")
    try:
        num = float(compact)
    except Exception:
        return 0.0
    if not math.isfinite(num):
        return 0.0
    return float(round(num, 2))


def _get_or_create_accounting_settings(db: Session, user: User) -> AccountingSettings:
    row = db.scalar(select(AccountingSettings).where(AccountingSettings.user_id == user.id))
    if row:
        return row
    profile = _get_or_create_user_profile(db, user.id)
    inferred_tax = _to_money(profile.tax_rate or 0.0)
    row = AccountingSettings(
        user_id=user.id,
        vat_rate=0.0,
        tax_rate=inferred_tax,
        additional_rate=0.0,
        fixed_cost_per_month=0.0,
    )
    db.add(row)
    db.flush()
    return row


def _accounting_settings_to_out(row: AccountingSettings) -> AccountingSettingsOut:
    return AccountingSettingsOut(
        vat_rate=float(round(row.vat_rate or 0.0, 2)),
        tax_rate=float(round(row.tax_rate or 0.0, 2)),
        additional_rate=float(round(row.additional_rate or 0.0, 2)),
        fixed_cost_per_month=float(round(row.fixed_cost_per_month or 0.0, 2)),
        updated_at=_to_iso_or_none(row.updated_at),
    )


def _accounting_expense_to_out(row: AccountingExpense) -> AccountingExpenseOut:
    return AccountingExpenseOut(
        id=int(row.id),
        marketplace=_normalize_accounting_marketplace(row.marketplace),
        category=str(row.category or ""),
        amount=float(round(row.amount or 0.0, 2)),
        recurrence=_normalize_expense_recurrence(row.recurrence),
        start_date=_to_iso_or_none(row.start_date),
        end_date=_to_iso_or_none(row.end_date),
        note=str(row.note or ""),
        is_active=bool(row.is_active),
        created_at=_to_iso_or_none(row.created_at),
        updated_at=_to_iso_or_none(row.updated_at),
    )


def _collect_accounting_expense_payload(rows: list[AccountingExpense]) -> list[dict[str, Any]]:
    out: list[dict[str, Any]] = []
    for row in rows:
        out.append(
            {
                "marketplace": _normalize_accounting_marketplace(row.marketplace),
                "category": str(row.category or ""),
                "amount": float(round(row.amount or 0.0, 2)),
                "recurrence": _normalize_expense_recurrence(row.recurrence),
                "start_date": _to_iso_or_none(row.start_date),
                "end_date": _to_iso_or_none(row.end_date),
                "note": str(row.note or ""),
                "is_active": bool(row.is_active),
            }
        )
    return out


def _collect_product_cost_payload(db: Session, user: User, marketplace: str) -> list[dict[str, Any]]:
    query = select(Product).where(
        Product.user_id == user.id,
        _owned_by_actor_or_owner_filter(Product, user),
    )
    market = _normalize_accounting_marketplace(marketplace)
    if market in {"wb", "ozon"}:
        query = query.where(Product.marketplace == market)
    rows = db.scalars(query).all()
    out: list[dict[str, Any]] = []
    for row in rows:
        out.append(
            {
                "marketplace": str(row.marketplace or "").strip().lower(),
                "article": str(row.article or ""),
                "external_id": str(row.external_id or ""),
                "barcode": str(row.barcode or ""),
                "name": str(row.name or ""),
                "purchase_price": float(round(row.purchase_price or 0.0, 2)),
            }
        )
    return out


def _workbook_bytes_from_rows(rows: list[list[Any]]) -> tuple[bytes, str]:
    if Workbook is not None:
        wb = Workbook()
        ws = wb.active
        ws.title = "purchase_prices"
        for row in rows:
            ws.append(list(row))
        bio = io.BytesIO()
        wb.save(bio)
        return bio.getvalue(), "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"

    text_rows: list[str] = []
    for row in rows:
        cells = [str(cell if cell is not None else "") for cell in row]
        escaped = ['"' + x.replace('"', '""') + '"' for x in cells]
        text_rows.append(",".join(escaped))
    raw = ("\n".join(text_rows)).encode("utf-8")
    return raw, "text/csv; charset=utf-8"


def _normalize_header(value: Any) -> str:
    text = str(value or "").strip().lower()
    text = text.replace("С‘", "Рµ")
    text = re.sub(r"[^a-zР°-СЏ0-9]+", "_", text)
    text = re.sub(r"_+", "_", text).strip("_")
    return text


def _parse_purchase_price_upload(upload: UploadFile) -> list[dict[str, Any]]:
    raw = upload.file.read() if upload and upload.file else b""
    if not raw:
        return []
    filename = str(upload.filename or "").lower()
    if (filename.endswith(".xlsx") or filename.endswith(".xlsm")) and load_workbook is not None:
        wb = load_workbook(io.BytesIO(raw), read_only=True, data_only=True)
        ws = wb.active
        rows_iter = ws.iter_rows(values_only=True)
        try:
            header = [str(x or "") for x in next(rows_iter)]
        except StopIteration:
            return []
        idx = { _normalize_header(name): pos for pos, name in enumerate(header) }
        return _extract_purchase_rows_from_iter(rows_iter, idx)

    text = raw.decode("utf-8", errors="ignore")
    lines = [line for line in text.splitlines() if line.strip()]
    if not lines:
        return []
    header_cells = [cell.strip().strip('"') for cell in lines[0].split(",")]
    idx = { _normalize_header(name): pos for pos, name in enumerate(header_cells) }
    data_rows = [line.split(",") for line in lines[1:]]
    return _extract_purchase_rows_from_iter(data_rows, idx)


def _extract_purchase_rows_from_iter(rows_iter: Any, idx: dict[str, int]) -> list[dict[str, Any]]:
    mp_idx = _first_index(idx, {"marketplace", "mp", "market", "РјР°СЂРєРµС‚РїР»РµРёСЃ", "РјР°СЂРєРµС‚РїР»РµР№СЃ"})
    article_idx = _first_index(idx, {"article", "Р°СЂС‚РёРєСѓР»", "sku", "offer_id", "vendor_code", "vendorcode"})
    external_idx = _first_index(idx, {"external_id", "externalid", "nm_id", "nmid", "barcode", "СЂС‚СЂРёС…РєРѕРґ"})
    price_idx = _first_index(idx, {"purchase_price", "purchaseprice", "Р·Р°РєСѓРїРѕС‡РЅР°СЏ_С†РµРЅР°", "Р·Р°РєСѓРїРѕС‡РЅР°СЏ_СЃС‚РѕРёРјРѕСЃС‚СЊ", "С†РµРЅР°_Р·Р°РєСѓРїРєРё", "cost", "cost_price"})
    if price_idx < 0:
        return []

    out: list[dict[str, Any]] = []
    row_no = 1
    for raw_row in rows_iter:
        row_no += 1
        row = list(raw_row) if isinstance(raw_row, (list, tuple)) else []
        def _cell(pos: int) -> str:
            if pos < 0 or pos >= len(row):
                return ""
            return str(row[pos] if row[pos] is not None else "").strip()

        marketplace = _normalize_accounting_marketplace(_cell(mp_idx) or "all")
        article = _cell(article_idx)
        external_id = _cell(external_idx)
        price_raw = _cell(price_idx)
        if not price_raw and not article and not external_id:
            continue
        out.append(
            {
                "row_no": row_no,
                "marketplace": marketplace,
                "article": article,
                "external_id": external_id,
                "price_raw": price_raw,
            }
        )
    return out


def _first_index(idx: dict[str, int], names: set[str]) -> int:
    for name in names:
        if name in idx:
            return int(idx[name])
    return -1


def _sanitize_ai_prompt(text: str) -> str:
    compact = " ".join((text or "").split())
    if len(compact) > 6000:
        return compact[:6000]
    return compact


def _hydrate_external_id_if_needed(db: Session, user_id: int, product: Product) -> None:
    if product.marketplace != "wb":
        return
    if product.external_id:
        return
    cred = db.scalar(
        select(ApiCredential)
        .where(
            ApiCredential.user_id == user_id,
            ApiCredential.marketplace == "wb",
            ApiCredential.active.is_(True),
        )
        .order_by(ApiCredential.id.desc())
    )
    if not cred:
        return
    resolved = resolve_wb_external_id(cred.api_key, product.article, product.name)
    if resolved:
        product.external_id = resolved


def _safe_known_position(value: int | None) -> int:
    if value is None:
        return 0
    if value <= 0:
        return 0
    if value > 500:
        return 501
    return int(value)


def _get_system_setting(db: Session, key: str) -> str:
    row = db.scalar(select(SystemSetting).where(SystemSetting.key == key))
    if not row:
        return ""
    return str(row.value or "")


def _set_system_setting(db: Session, key: str, value: str) -> None:
    row = db.scalar(select(SystemSetting).where(SystemSetting.key == key))
    if not row:
        row = SystemSetting(key=key, value=value)
        db.add(row)
        db.flush()
        return
    row.value = value


def _sanitize_mobile_module_overrides(raw: dict[str, Any] | None) -> dict[str, dict[str, bool]]:
    payload = raw if isinstance(raw, dict) else {}
    out: dict[str, dict[str, bool]] = {}
    allowed_codes = {str(code).strip().lower() for code in DEFAULT_MODULES}
    for user_key, row in payload.items():
        uid = _to_int_safe(user_key)
        if uid <= 0:
            continue
        mapping = row if isinstance(row, dict) else {}
        clean_row: dict[str, bool] = {}
        for module_code, enabled in mapping.items():
            code = str(module_code or "").strip().lower()
            if not code or code not in allowed_codes:
                continue
            clean_row[code] = bool(enabled)
        if clean_row:
            out[str(uid)] = clean_row
    return out


def _get_mobile_module_overrides(db: Session) -> dict[str, dict[str, bool]]:
    raw = _get_system_setting(db, MOBILE_MODULE_OVERRIDES_SETTING_KEY)
    if not raw:
        return {}
    try:
        parsed = json.loads(raw)
    except Exception:
        parsed = {}
    safe = _sanitize_mobile_module_overrides(parsed if isinstance(parsed, dict) else {})
    if safe != parsed:
        _set_system_setting(db, MOBILE_MODULE_OVERRIDES_SETTING_KEY, json.dumps(safe, ensure_ascii=False))
        db.flush()
    return safe


def _set_mobile_module_override(db: Session, *, user_id: int, module_code: str, enabled: bool) -> dict[str, dict[str, bool]]:
    code = str(module_code or "").strip().lower()
    if code not in DEFAULT_MODULES:
        return _get_mobile_module_overrides(db)
    safe = _get_mobile_module_overrides(db)
    user_key = str(int(user_id or 0))
    if not user_key.isdigit() or int(user_key) <= 0:
        return safe
    row = dict(safe.get(user_key) or {})
    row[code] = bool(enabled)
    safe[user_key] = row
    safe = _sanitize_mobile_module_overrides(safe)
    _set_system_setting(db, MOBILE_MODULE_OVERRIDES_SETTING_KEY, json.dumps(safe, ensure_ascii=False))
    db.flush()
    return safe


def _module_enabled_for_context(db: Session, user: User, module_code: str, default_enabled: bool) -> bool:
    base = bool(default_enabled)
    if not bool(getattr(user, "_is_mobile_app", False)):
        return base
    safe = _get_mobile_module_overrides(db)
    user_map = safe.get(str(int(user.id or 0)), {})
    if not isinstance(user_map, dict):
        return base
    code = str(module_code or "").strip().lower()
    if not code:
        return base
    if code in user_map:
        return bool(user_map.get(code))
    return base


def _sanitize_ui_settings_payload(raw: dict[str, Any] | None) -> dict[str, Any]:
    data = raw if isinstance(raw, dict) else {}
    enabled = bool(data.get("theme_choice_enabled", DEFAULT_UI_SETTINGS["theme_choice_enabled"]))
    force_theme = bool(data.get("force_theme", DEFAULT_UI_SETTINGS.get("force_theme", False)))
    default_theme = str(data.get("default_theme") or DEFAULT_UI_SETTINGS["default_theme"]).strip().lower()
    if default_theme not in AVAILABLE_THEMES:
        default_theme = str(DEFAULT_UI_SETTINGS["default_theme"])
    allowed_raw = data.get("allowed_themes")
    allowed_in = allowed_raw if isinstance(allowed_raw, list) else list(DEFAULT_UI_SETTINGS["allowed_themes"])
    allowed: list[str] = []
    seen: set[str] = set()
    for theme in allowed_in:
        code = str(theme or "").strip().lower()
        if not code or code not in AVAILABLE_THEMES or code in seen:
            continue
        seen.add(code)
        allowed.append(code)
    if not allowed:
        allowed = [str(DEFAULT_UI_SETTINGS["default_theme"])]
    if default_theme not in allowed:
        default_theme = allowed[0]
    if enabled:
        force_theme = False
    return {
        "theme_choice_enabled": enabled,
        "force_theme": force_theme,
        "default_theme": default_theme,
        "allowed_themes": allowed,
    }


def _get_ui_settings(db: Session) -> UiSettingsOut:
    raw = _get_system_setting(db, "ui_settings")
    if not raw:
        safe_default = _sanitize_ui_settings_payload(DEFAULT_UI_SETTINGS)
        _set_system_setting(db, "ui_settings", json.dumps(safe_default, ensure_ascii=False))
        db.commit()
        return UiSettingsOut(**safe_default)
    try:
        parsed = json.loads(raw)
    except Exception:
        parsed = {}
    safe = _sanitize_ui_settings_payload(parsed)
    if safe != parsed:
        _set_system_setting(db, "ui_settings", json.dumps(safe, ensure_ascii=False))
        db.commit()
    return UiSettingsOut(**safe)


def _preferred_keyword_from_name(name: str) -> str:
    low = name.lower()
    if "СѓС‚РµРїР»" in low and "С‚СЂСѓР±" in low:
        return "СѓС‚РµРїР»РёС‚РµР»СЊ РґР»СЏ С‚СЂСѓР±"
    if "РґС‹РјРѕС…РѕРґ" in low and "С‚СЂСѓР±" in low:
        return "С‚СЂСѓР±Р° РґС‹РјРѕС…РѕРґР°"
    if "РєРѕР»РµРЅРѕ" in low and "РґС‹РјРѕС…РѕРґ" in low:
        return "РєРѕР»РµРЅРѕ РґС‹РјРѕС…РѕРґР°"
    words = [w for w in low.replace("/", " ").replace("-", " ").split() if len(w) >= 4]
    if len(words) >= 2:
        return f"{words[0]} {words[1]}"
    if words:
        return words[0]
    return ""


def _sanitize_generated_description(text: str) -> str:
    raw = text or ""
    banned = (
        "Р¤РѕСЂРјСѓР»РёСЂРѕРІРєРё СЃРґРµР»Р°РЅС‹ РєРѕСЂРѕС‚РєРёРјРё Рё РїРѕРЅСЏС‚РЅС‹РјРё, С‡С‚РѕР±С‹ Р±С‹СЃС‚СЂРѕ РѕС†РµРЅРёС‚СЊ РЅР°Р·РЅР°С‡РµРЅРёРµ Рё СЃРѕРІРјРµСЃС‚РёРјРѕСЃС‚СЊ С‚РѕРІР°СЂР°.",
    )
    for phrase in banned:
        raw = raw.replace(phrase, "").strip()
    raw = " ".join(raw.split())
    return raw


def _actor_member_id(user: User) -> int:
    try:
        return int(getattr(user, "_actor_member_id", 0) or 0)
    except Exception:
        return 0


def _actor_is_owner(user: User) -> bool:
    return bool(getattr(user, "_actor_is_owner", True))


def _actor_email(user: User) -> str:
    return str(getattr(user, "_actor_email", "") or "").strip().lower()


def _require_owner_actor(user: User) -> None:
    if not _actor_is_owner(user):
        raise HTTPException(status_code=403, detail="РўРѕР»СЊРєРѕ РІР»Р°РґРµР»РµС† РєР°Р±РёРЅРµС‚Р° РјРѕР¶РµС‚ РІС‹РїРѕР»РЅСЏС‚СЊ СЌС‚Рѕ РґРµР№СЃС‚РІРёРµ")


def _actor_owner_member_id(user: User) -> int:
    return int(getattr(user, "_owner_member_id", 0) or 0)


def _owner_member_id_for_user(db: Session, user_id: int) -> int:
    row = db.scalar(
        select(TeamMember.id).where(
            TeamMember.user_id == user_id,
            TeamMember.is_owner.is_(True),
        ).order_by(TeamMember.id.asc())
    )
    return int(row or 0)


def _resolve_owner_member_id(db: Session, user: User) -> int:
    owner_member_id = _owner_member_id_for_user(db, user.id)
    if owner_member_id > 0:
        return owner_member_id
    return _actor_member_id(user)


def _owned_by_actor_or_owner_filter(model: Any, user: User) -> Any:
    if _actor_is_owner(user):
        return True
    actor_id = _actor_member_id(user)
    owner_id = _actor_owner_member_id(user)
    if actor_id <= 0 and owner_id <= 0:
        return False
    if actor_id > 0 and owner_id > 0 and actor_id != owner_id:
        return model.owner_member_id.in_([actor_id, owner_id]) | model.owner_member_id.is_(None)
    if owner_id > 0:
        return (model.owner_member_id == owner_id) | model.owner_member_id.is_(None)
    return model.owner_member_id == actor_id


def _enforce_record_owner_access(record: Any, user: User) -> None:
    if _actor_is_owner(user):
        return
    actor_id = _actor_member_id(user)
    owner_id = _actor_owner_member_id(user)
    if actor_id <= 0:
        raise HTTPException(status_code=403, detail="РќРµРґРѕСЃС‚Р°С‚РѕС‡РЅРѕ РїСЂР°РІ")
    owner_member_id = int(getattr(record, "owner_member_id", 0) or 0)
    if owner_member_id not in {actor_id, owner_id}:
        raise HTTPException(status_code=403, detail="РќРµРґРѕСЃС‚Р°С‚РѕС‡РЅРѕ РїСЂР°РІ РґР»СЏ РґРѕСЃС‚СѓРїР° Рє РґР°РЅРЅС‹Рј РґСЂСѓРіРѕРіРѕ СЃРѕС‚СЂСѓРґРЅРёРєР°")


def _assign_owner_member(record: Any, owner_member_id: int) -> None:
    if hasattr(record, "owner_member_id"):
        setattr(record, "owner_member_id", int(owner_member_id or 0) or None)


def _work_item_claim_key(module_code: str, marketplace: str, item_type: str, item_external_id: str) -> tuple[str, str, str, str]:
    return (
        str(module_code or "").strip().lower(),
        str(marketplace or "").strip().lower(),
        str(item_type or "").strip().lower(),
        str(item_external_id or "").strip(),
    )


def _claim_or_validate_work_item(
    db: Session,
    user: User,
    *,
    module_code: str,
    marketplace: str,
    item_type: str,
    item_external_id: str,
) -> None:
    if _actor_is_owner(user):
        return
    actor_member_id = _actor_member_id(user)
    if actor_member_id <= 0:
        raise HTTPException(status_code=403, detail="РќРµРґРѕСЃС‚Р°С‚РѕС‡РЅРѕ РїСЂР°РІ")
    _, _, _, item_id = _work_item_claim_key(module_code, marketplace, item_type, item_external_id)
    if not item_id:
        raise HTTPException(status_code=400, detail="РќРµ СѓРґР°Р»РѕСЃСЊ РѕРїСЂРµРґРµР»РёС‚СЊ ID Р·Р°РїРёСЃРё")
    row = db.scalar(
        select(WorkItemClaim).where(
            WorkItemClaim.user_id == user.id,
            WorkItemClaim.module_code == module_code,
            WorkItemClaim.marketplace == marketplace,
            WorkItemClaim.item_type == item_type,
            WorkItemClaim.item_external_id == item_id,
        )
    )
    if row:
        if int(row.owner_member_id or 0) != actor_member_id:
            raise HTTPException(status_code=403, detail="Р—Р°РїРёСЃСЊ СѓР¶Рµ Р·Р°РєСЂРµРїР»РµРЅР° Р·Р° РґСЂСѓРіРёРј СЃРѕС‚СЂСѓРґРЅРёРєРѕРј")
        if not row.claimed_at:
            row.claimed_at = datetime.utcnow()
        row.updated_at = datetime.utcnow()
        return
    row = WorkItemClaim(
        user_id=user.id,
        module_code=module_code,
        marketplace=marketplace,
        item_type=item_type,
        item_external_id=item_id,
        owner_member_id=actor_member_id,
        claimed_at=datetime.utcnow(),
        created_at=datetime.utcnow(),
        updated_at=datetime.utcnow(),
    )
    db.add(row)


def _feedback_synthetic_id(row: dict[str, Any], marketplace: str) -> str:
    raw = str(
        row.get("id")
        or row.get("feedbackId")
        or row.get("feedback_id")
        or row.get("reviewId")
        or row.get("review_id")
        or row.get("questionId")
        or row.get("question_id")
        or ""
    ).strip()
    if raw:
        return raw
    parts = [
        str(row.get("created_at") or row.get("date") or ""),
        str(row.get("product") or ""),
        str(row.get("article") or ""),
        str(row.get("barcode") or ""),
        str(row.get("user") or ""),
        str(row.get("text") or ""),
        str(row.get("answer") or ""),
    ]
    base = "|".join(parts).strip().lower()
    if not base:
        return ""
    digest = hashlib.sha1(base.encode("utf-8")).hexdigest()[:12]
    return f"{marketplace}-fb-{digest}"


def _filter_claimed_feedback_rows(
    db: Session,
    user: User,
    *,
    module_code: str,
    marketplace: str,
    item_type: str,
    rows: list[dict[str, Any]],
) -> list[dict[str, Any]]:
    safe_rows = [x for x in (rows or []) if isinstance(x, dict)]
    if _actor_is_owner(user):
        for row in safe_rows:
            row["id"] = _feedback_synthetic_id(row, marketplace)
        return safe_rows
    actor_member_id = _actor_member_id(user)
    if actor_member_id <= 0:
        return []
    ids: list[str] = []
    for row in safe_rows:
        row_id = _feedback_synthetic_id(row, marketplace)
        row["id"] = row_id
        if row_id:
            ids.append(row_id)
    if not ids:
        return safe_rows
    claims = db.scalars(
        select(WorkItemClaim).where(
            WorkItemClaim.user_id == user.id,
            WorkItemClaim.module_code == module_code,
            WorkItemClaim.marketplace == marketplace,
            WorkItemClaim.item_type == item_type,
            WorkItemClaim.item_external_id.in_(ids),
        )
    ).all()
    claimed_by: dict[str, int] = {
        str(x.item_external_id): int(x.owner_member_id or 0)
        for x in claims
    }
    out: list[dict[str, Any]] = []
    for row in safe_rows:
        row_id = str(row.get("id") or "")
        owner = claimed_by.get(row_id)
        if owner and owner != actor_member_id:
            continue
        out.append(row)
    return out


def _sqlite_db_main_path() -> str:
    raw_url = str(settings.database_url or "").strip()
    if not raw_url.lower().startswith("sqlite"):
        return ""
    if raw_url.startswith("sqlite:////"):
        return "/" + raw_url.split("sqlite:////", 1)[1]
    if raw_url.startswith("sqlite:///"):
        return os.path.abspath(raw_url.split("sqlite:///", 1)[1])
    return ""


def _estimate_audit_storage_bytes(db: Session) -> int:
    if not str(settings.database_url or "").lower().startswith("sqlite"):
        # For non-sqlite we currently skip byte-based retention.
        return 0
    page_size = int(db.scalar(text("PRAGMA page_size")) or 0)
    page_count = int(db.scalar(text("PRAGMA page_count")) or 0)
    free_count = int(db.scalar(text("PRAGMA freelist_count")) or 0)
    used_bytes = max(0, page_count - free_count) * max(0, page_size)
    main_path = _sqlite_db_main_path()
    wal_bytes = 0
    if main_path:
        wal_path = f"{main_path}-wal"
        try:
            wal_bytes = os.path.getsize(wal_path) if os.path.exists(wal_path) else 0
        except Exception:
            wal_bytes = 0
    return int(used_bytes + wal_bytes)


def _prune_audit_storage_if_needed(db: Session) -> None:
    global _AUDIT_PRUNE_LAST_CHECK_AT
    now_ts = datetime.utcnow().timestamp()
    if now_ts - _AUDIT_PRUNE_LAST_CHECK_AT < 180:
        return
    _AUDIT_PRUNE_LAST_CHECK_AT = now_ts

    try:
        used_bytes = _estimate_audit_storage_bytes(db)
    except Exception:
        return
    if used_bytes <= AUDIT_STORAGE_MAX_BYTES:
        return

    cutoff = datetime.utcnow() - timedelta(days=31)
    rounds = 0
    while used_bytes > AUDIT_STORAGE_TARGET_BYTES and rounds < 120:
        ids = db.scalars(
            select(AuditLog.id)
            .where(AuditLog.created_at < cutoff)
            .order_by(AuditLog.id.asc())
            .limit(AUDIT_PRUNE_BATCH)
        ).all()
        if not ids:
            ids = db.scalars(
                select(AuditLog.id)
                .order_by(AuditLog.id.asc())
                .limit(AUDIT_PRUNE_BATCH)
            ).all()
            if not ids:
                break
        db.execute(delete(AuditLog).where(AuditLog.id.in_(ids)))
        db.flush()
        rounds += 1
        try:
            used_bytes = _estimate_audit_storage_bytes(db)
        except Exception:
            break
    # Keep WAL bounded after large prune cycles.
    if rounds > 0:
        try:
            db.execute(text("PRAGMA wal_checkpoint(TRUNCATE)"))
        except Exception:
            pass


def _audit(
    db: Session,
    user: User | None,
    *,
    action: str,
    details: str = "",
    module_code: str = "",
    entity_type: str = "",
    entity_id: str = "",
    status: str = "ok",
    request: Request | None = None,
) -> None:
    # High-frequency read events and UI pings create write contention on SQLite.
    # Keep mutating/security events, but skip noisy telemetry-grade records.
    action_key = str(action or "").strip().lower()
    if (
        not action_key
        or action_key.startswith("ui_")
        or action_key.endswith("_read")
        or action_key in {"auth_session_check", "wb_ads_campaigns_enrich"}
    ):
        return

    actor_email = _actor_email(user) if user else ""
    actor_member_id = _actor_member_id(user) if user else None
    actor_is_owner = _actor_is_owner(user) if user else True
    client_ip = ""
    user_agent = ""
    if request is not None:
        client_ip = str(getattr(request.client, "host", "") or "")
        user_agent = str(request.headers.get("user-agent") or "")[:500]
    db.add(
        AuditLog(
            user_id=(user.id if user else None),
            action=str(action or "").strip()[:120],
            details=str(details or "")[:5000],
            actor_email=actor_email[:255],
            actor_member_id=actor_member_id,
            actor_is_owner=actor_is_owner,
            module_code=str(module_code or "")[:80],
            entity_type=str(entity_type or "")[:80],
            entity_id=str(entity_id or "")[:120],
            status=str(status or "ok")[:24],
            ip=client_ip[:80],
            user_agent=user_agent[:500],
        )
    )
    _prune_audit_storage_if_needed(db)


def _actor_scope(user: User) -> list[str]:
    raw = getattr(user, "_actor_member_scope", ["*"])
    if not isinstance(raw, list):
        return ["*"]
    out: list[str] = []
    seen: set[str] = set()
    for item in raw:
        code = str(item or "").strip().lower()
        if not code or code in seen:
            continue
        seen.add(code)
        out.append(code)
    return out or ["*"]


def _actor_can_use_module(user: User, module_code: str) -> bool:
    code = str(module_code or "").strip().lower()
    if not code:
        return True
    scope = _actor_scope(user)
    return "*" in scope or code in scope


def _normalize_ai_mode(value: str) -> str:
    mode = str(value or "").strip().lower()
    if mode not in {"builtin", "global", "user"}:
        return "builtin"
    return mode


def _normalize_ai_provider(value: str) -> str:
    code = str(value or "").strip().lower()
    if not code:
        return "openai"
    if code not in AI_PROVIDER_CODES:
        return "custom"
    return code


def _sanitize_ai_service_name(value: str) -> str:
    text = " ".join(str(value or "").split()).strip()
    if not text:
        return "AI service"
    return text[:120]


def _provider_default_model(provider: str) -> str:
    code = _normalize_ai_provider(provider)
    defaults = {
        "deepseek": "deepseek-chat",
    }
    return defaults.get(code, settings.openai_model or "gpt-4o-mini")


def _sanitize_ai_service_model(value: str, *, provider: str = "openai") -> str:
    normalized_provider = _normalize_ai_provider(provider)
    text = " ".join(str(value or "").split()).strip()
    default_model = _provider_default_model(normalized_provider)
    if not text:
        return default_model
    model = text[:120]
    low = model.lower()
    if normalized_provider == "deepseek":
        # DeepSeek endpoint rejects OpenAI model names; auto-fix common misconfiguration.
        if low.startswith("gpt-") or low in {"o1", "o1-mini", "o3", "o4-mini"}:
            return "deepseek-chat"
    return model


def _sanitize_ai_base_url(value: str) -> str:
    text = str(value or "").strip()
    if not text:
        return ""
    if text.startswith("http://") or text.startswith("https://"):
        return text[:500]
    return f"https://{text[:490]}"


def _get_global_ai_default(db: Session) -> dict[str, Any]:
    raw = _get_system_setting(db, "ai_global_default")
    if not raw:
        payload = {"mode": "builtin", "service_id": None}
        _set_system_setting(db, "ai_global_default", json.dumps(payload, ensure_ascii=False))
        return payload
    try:
        parsed = json.loads(raw)
    except Exception:
        parsed = {}
    mode = _normalize_ai_mode(str(parsed.get("mode") or "builtin"))
    service_id = _to_int_safe(parsed.get("service_id")) or None
    payload = {"mode": mode, "service_id": service_id}
    if payload != parsed:
        _set_system_setting(db, "ai_global_default", json.dumps(payload, ensure_ascii=False))
    return payload


def _get_or_create_user_ai_preference(db: Session, user_id: int) -> UserAiPreference:
    row = db.scalar(select(UserAiPreference).where(UserAiPreference.user_id == user_id))
    if row:
        return row
    row = UserAiPreference(
        user_id=user_id,
        use_global_default=True,
        mode="builtin",
        service_id=None,
    )
    db.add(row)
    db.flush()
    return row


def _validate_ai_service_binding(db: Session, *, mode: str, service_id: int | None, user_id: int | None) -> int | None:
    if mode == "builtin":
        return None
    sid = _to_int_safe(service_id)
    if sid <= 0:
        raise HTTPException(status_code=400, detail="Р’С‹Р±РµСЂРёС‚Рµ AI СЃРµСЂРІРёСЃ РґР»СЏ РІС‹Р±СЂР°РЅРЅРѕРіРѕ СЂРµР¶РёРјР°")
    row = db.get(AiServiceAccount, sid)
    if not row:
        raise HTTPException(status_code=404, detail="AI СЃРµСЂРІРёСЃ РЅРµ РЅР°Р№РґРµРЅ")
    if mode == "global" and row.user_id is not None:
        raise HTTPException(status_code=400, detail="Р”Р»СЏ СЂРµР¶РёРјР° global РЅСѓР¶РµРЅ РіР»РѕР±Р°Р»СЊРЅС‹Р№ AI СЃРµСЂРІРёСЃ")
    if mode == "user" and (user_id is None or row.user_id != user_id):
        raise HTTPException(status_code=400, detail="Р”Р»СЏ СЂРµР¶РёРјР° user РЅСѓР¶РµРЅ AI СЃРµСЂРІРёСЃ С‚РµРєСѓС‰РµРіРѕ РїРѕР»СЊР·РѕРІР°С‚РµР»СЏ")
    return sid


def _save_user_ai_selection(
    db: Session,
    user_id: int,
    *,
    use_global_default: bool,
    mode: str,
    service_id: int | None,
) -> UserAiPreference:
    row = _get_or_create_user_ai_preference(db, user_id)
    row.use_global_default = bool(use_global_default)
    if row.use_global_default:
        row.mode = "builtin"
        row.service_id = None
        return row
    safe_mode = _normalize_ai_mode(mode)
    safe_id = _validate_ai_service_binding(db, mode=safe_mode, service_id=service_id, user_id=user_id)
    row.mode = safe_mode
    row.service_id = safe_id
    return row


def _ai_service_to_out(row: AiServiceAccount, *, scope: str) -> AiServiceOut:
    return AiServiceOut(
        id=row.id,
        scope=scope,
        user_id=row.user_id,
        name=row.name or "",
        provider=row.provider or "openai",
        model=row.model or "",
        base_url=row.base_url or "",
        api_key_masked=mask_key(row.api_key or ""),
        priority=int(row.priority or 1000),
        is_active=bool(row.is_active),
        created_at=row.created_at.isoformat() if row.created_at else None,
    )


def _update_ai_service_row(row: AiServiceAccount, payload: AiServiceIn) -> None:
    row.name = _sanitize_ai_service_name(payload.name)
    provider = _normalize_ai_provider(payload.provider)
    row.provider = provider
    row.api_key = str(payload.api_key or "").strip()[:255]
    row.model = _sanitize_ai_service_model(payload.model, provider=provider)
    row.base_url = _sanitize_ai_base_url(payload.base_url)
    row.is_active = True


def _upsert_ai_service(db: Session, *, user_id: int | None, payload: AiServiceIn) -> AiServiceAccount:
    max_priority = db.scalar(
        select(func.max(AiServiceAccount.priority)).where(AiServiceAccount.user_id == user_id)
    )
    next_priority = int(max_priority or 0) + 1
    row = AiServiceAccount(
        user_id=user_id,
        name="",
        provider="openai",
        api_key="",
        model=settings.openai_model,
        base_url="",
        is_active=True,
        priority=next_priority,
    )
    _update_ai_service_row(row, payload)
    if not row.api_key:
        raise HTTPException(status_code=400, detail="API РєР»СЋС‡ AI СЃРµСЂРІРёСЃР° РѕР±СЏР·Р°С‚РµР»РµРЅ")
    db.add(row)
    db.flush()
    return row


def _reset_ai_selection_if_deleted_service(db: Session, user_id: int, service_id: int) -> None:
    pref = db.scalar(select(UserAiPreference).where(UserAiPreference.user_id == user_id))
    if pref and not pref.use_global_default and _to_int_safe(pref.service_id) == service_id:
        pref.use_global_default = True
        pref.mode = "builtin"
        pref.service_id = None


def _reorder_ai_services(db: Session, *, user_id: int | None, service_ids: list[int]) -> list[AiServiceAccount]:
    safe_ids = [int(x) for x in service_ids if int(x or 0) > 0]
    rows = db.scalars(
        select(AiServiceAccount).where(AiServiceAccount.user_id == user_id)
    ).all()
    if not rows:
        return []
    row_by_id = {int(row.id): row for row in rows}
    requested = [row_by_id[sid] for sid in safe_ids if sid in row_by_id]
    remaining = [row for row in rows if int(row.id) not in set(safe_ids)]
    remaining.sort(key=lambda r: (int(r.priority or 1000), int(r.id or 0)))
    ordered = requested + remaining
    for idx, row in enumerate(ordered, start=1):
        row.priority = idx
    return ordered


def _resolve_user_ai_runtime(db: Session, user_id: int) -> dict[str, Any]:
    global_default = _get_global_ai_default(db)
    pref = db.scalar(select(UserAiPreference).where(UserAiPreference.user_id == user_id))
    use_global_default = True
    mode = "builtin"
    service_id = None
    if pref:
        use_global_default = bool(pref.use_global_default)
        if not use_global_default:
            mode = _normalize_ai_mode(pref.mode)
            service_id = _to_int_safe(pref.service_id) or None
    if use_global_default:
        mode = _normalize_ai_mode(str(global_default.get("mode") or "builtin"))
        service_id = _to_int_safe(global_default.get("service_id")) or None

    global_rows = db.scalars(
        select(AiServiceAccount)
        .where(
            AiServiceAccount.user_id.is_(None),
            AiServiceAccount.is_active.is_(True),
        )
        .order_by(AiServiceAccount.priority.asc(), AiServiceAccount.id.desc())
    ).all()
    user_rows = db.scalars(
        select(AiServiceAccount)
        .where(
            AiServiceAccount.user_id == user_id,
            AiServiceAccount.is_active.is_(True),
        )
        .order_by(AiServiceAccount.priority.asc(), AiServiceAccount.id.desc())
    ).all()

    def _row_to_runtime(row: AiServiceAccount, runtime_mode: str) -> dict[str, Any]:
        resolved_provider = row.provider or "openai"
        return {
            "mode": runtime_mode,
            "service_id": row.id,
            "service_name": row.name or f"AI #{row.id}",
            "provider": resolved_provider,
            "api_key": row.api_key or "",
            "model": _sanitize_ai_service_model(row.model or "", provider=resolved_provider),
            "base_url": row.base_url or "",
            "source": "service",
        }

    def _builtin_runtime() -> dict[str, Any]:
        return {
            "mode": "builtin",
            "service_id": None,
            "service_name": "Built-in OpenAI",
            "provider": "openai",
            "api_key": settings.openai_api_key or "",
            "model": _sanitize_ai_service_model(settings.openai_model or "", provider="openai"),
            "base_url": "",
            "source": "builtin",
        }

    ai_chain: list[dict[str, Any]] = []
    selected_row: AiServiceAccount | None = None
    if mode == "global" and service_id:
        selected_row = next((x for x in global_rows if int(x.id or 0) == int(service_id)), None)
    elif mode == "user" and service_id:
        selected_row = next((x for x in user_rows if int(x.id or 0) == int(service_id)), None)
    if mode == "user" and not selected_row and user_rows:
        selected_row = user_rows[0]
    if mode == "global" and not selected_row and global_rows:
        selected_row = global_rows[0]

    if mode == "user" and selected_row:
        ai_chain.append(_row_to_runtime(selected_row, "user"))
        for row in user_rows:
            if selected_row and int(row.id or 0) == int(selected_row.id or 0):
                continue
            ai_chain.append(_row_to_runtime(row, "user"))
        for row in global_rows:
            ai_chain.append(_row_to_runtime(row, "global"))
    elif mode == "global" and selected_row:
        ai_chain.append(_row_to_runtime(selected_row, "global"))
        for row in global_rows:
            if selected_row and int(row.id or 0) == int(selected_row.id or 0):
                continue
            ai_chain.append(_row_to_runtime(row, "global"))
    else:
        # Built-in as primary (if configured), then global fallbacks.
        builtin = _builtin_runtime()
        if str(builtin.get("api_key") or "").strip():
            ai_chain.append(builtin)
        for row in global_rows:
            ai_chain.append(_row_to_runtime(row, "global"))

    # If selected mode has no valid service, fallback to available chain.
    if not ai_chain:
        builtin = _builtin_runtime()
        if str(builtin.get("api_key") or "").strip():
            ai_chain.append(builtin)

    # Ensure builtin is always the last fallback when configured.
    builtin = _builtin_runtime()
    builtin_token = str(builtin.get("api_key") or "").strip()
    if builtin_token:
        has_builtin = any(str(x.get("mode") or "") == "builtin" for x in ai_chain)
        if not has_builtin:
            ai_chain.append(builtin)

    if not ai_chain:
        # Keep shape stable even when no credentials are configured.
        empty_builtin = _builtin_runtime()
        empty_builtin["api_key"] = ""
        ai_chain = [empty_builtin]

    primary = dict(ai_chain[0])
    primary["fallback_chain"] = [dict(x) for x in ai_chain[1:]]
    primary["ai_chain"] = [dict(x) for x in ai_chain]
    if str(primary.get("mode") or "") == "global" and mode == "builtin":
        primary["source"] = "auto_global_fallback"
    return primary


def _build_ai_profile_payload(db: Session, user_id: int) -> AiProfileOut:
    pref = _get_or_create_user_ai_preference(db, user_id)
    global_default = _get_global_ai_default(db)
    user_services = db.scalars(
        select(AiServiceAccount)
        .where(AiServiceAccount.user_id == user_id)
        .order_by(AiServiceAccount.priority.asc(), AiServiceAccount.id.desc())
    ).all()
    global_services = db.scalars(
        select(AiServiceAccount)
        .where(AiServiceAccount.user_id.is_(None))
        .order_by(AiServiceAccount.priority.asc(), AiServiceAccount.id.desc())
    ).all()
    runtime = _resolve_user_ai_runtime(db, user_id)
    runtime_chain = [x for x in (runtime.get("ai_chain") or []) if isinstance(x, dict)]

    def _runtime_to_effective(row: dict[str, Any]) -> AiEffectiveOut:
        return AiEffectiveOut(
            mode=str(row.get("mode") or "builtin"),
            service_id=_to_int_safe(row.get("service_id")) or None,
            service_name=str(row.get("service_name") or ""),
            provider=str(row.get("provider") or "builtin"),
            model=str(row.get("model") or ""),
            source=str(row.get("source") or "service"),
        )

    return AiProfileOut(
        selection=AiSelectionOut(
            use_global_default=bool(pref.use_global_default),
            mode=str(pref.mode or "builtin"),
            service_id=_to_int_safe(pref.service_id) or None,
        ),
        global_default=AiSelectionOut(
            use_global_default=False,
            mode=str(global_default.get("mode") or "builtin"),
            service_id=_to_int_safe(global_default.get("service_id")) or None,
        ),
        effective=AiEffectiveOut(
            mode=str(runtime.get("mode") or "builtin"),
            service_id=_to_int_safe(runtime.get("service_id")) or None,
            service_name=str(runtime.get("service_name") or ""),
            provider=str(runtime.get("provider") or "builtin"),
            model=str(runtime.get("model") or ""),
            source=str(runtime.get("source") or "builtin"),
        ),
        effective_chain=[_runtime_to_effective(x) for x in runtime_chain],
        user_services=[_ai_service_to_out(x, scope="user") for x in user_services],
        global_services=[_ai_service_to_out(x, scope="global") for x in global_services],
    )


def ensure_module_enabled(db: Session, user: User, module_code: str):
    code = str(module_code or "").strip().lower()
    row = db.scalar(
        select(ModuleAccess).where(
            ModuleAccess.user_id == user.id,
            ModuleAccess.module_code == code,
        )
    )
    if not row:
        raise HTTPException(status_code=403, detail=f"РњРѕРґСѓР»СЊ '{module_code}' РѕС‚РєР»СЋС‡РµРЅ РґР»СЏ РІР°СЂРµРіРѕ С‚Р°СЂРёС„Р°")
    if not _module_enabled_for_context(db, user, code, bool(row.enabled)):
        raise HTTPException(status_code=403, detail=f"РњРѕРґСѓР»СЊ '{module_code}' РѕС‚РєР»СЋС‡РµРЅ РґР»СЏ РІР°СЂРµРіРѕ С‚Р°СЂРёС„Р°")
    if not _actor_can_use_module(user, code):
        raise HTTPException(status_code=403, detail=f"Р”РѕСЃС‚СѓРї Рє РјРѕРґСѓР»СЋ '{module_code}' РѕРіСЂР°РЅРёС‡РµРЅ РґР»СЏ РІР°СЂРµРіРѕ СЃРѕС‚СЂСѓРґРЅРёРєР°")


def validate_marketplace(value: str) -> str:
    marketplace = value.strip().lower()
    if marketplace not in {"wb", "ozon"}:
        raise HTTPException(status_code=400, detail="marketplace РґРѕР»Р¶РµРЅ Р±С‹С‚СЊ wb РёР»Рё ozon")
    return marketplace


def mask_key(api_key: str) -> str:
    if len(api_key) <= 8:
        return "*" * len(api_key)
    return f"{api_key[:4]}...{api_key[-4:]}"





























